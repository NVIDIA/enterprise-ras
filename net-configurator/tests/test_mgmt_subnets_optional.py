# SPDX-FileCopyrightText: Copyright (c) 2026 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: MIT
"""Task 8a/8b: mgmt_subnets is now an OPTIONAL Settings key whose
presence is a hard migration error.

A workbook without `mgmt_subnets` must no longer fail validation with
"Missing required key: 'mgmt_subnets'" (Task 8a) — this unblocked the later
task that removed `mgmt_subnets` from the default Excels. As of Task 8b,
mgmt_subnets is fully retired: a workbook that still HAS the key now fails
with a migration error pointing at the OOB VLAN row / Nodes 'OOB VLAN'
column, regardless of whether the value is a syntactically valid CIDR.
mgmt_subnets stays in OPTIONAL_SETTINGS_KEYS (not REQUIRED, not removed
from the known-keys union) purely so its presence trips this specific
migration error instead of a generic "unknown key" warning.
"""
import sys
from pathlib import Path

import openpyxl

SCRIPTS_DIR = Path(__file__).parent.parent / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from validate_excel import (
    validate_settings,
    ValidationResult,
    REQUIRED_SETTINGS_KEYS,
    OPTIONAL_SETTINGS_KEYS,
)


def _settings(pairs):
    """Build a Settings sheet from (key, value) pairs."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Settings"
    ws.cell(row=1, column=1, value="Setting")
    ws.cell(row=1, column=2, value="Value")
    for r, (k, v) in enumerate(pairs, start=2):
        ws.cell(row=r, column=1, value=k)
        ws.cell(row=r, column=2, value=v)
    return ws


def _run(pairs):
    result = ValidationResult()
    validate_settings(_settings(pairs), result)
    return result


def _missing_mgmt_subnets_errors(result):
    return [
        e for e in result.errors
        if "mgmt_subnets" in e and "Missing required key" in e
    ]


# ─── Membership: mgmt_subnets moved from REQUIRED to OPTIONAL ────────────

def test_mgmt_subnets_not_in_required_keys():
    assert "mgmt_subnets" not in REQUIRED_SETTINGS_KEYS


def test_mgmt_subnets_in_optional_keys():
    assert "mgmt_subnets" in OPTIONAL_SETTINGS_KEYS


# ─── Behavioral: absent mgmt_subnets no longer errors ────────────────────

_OTHER_REQUIRED_KEYS = [
    ("architecture", "2-4-3-200"),
    ("bgp_asn", 65000),
    ("loopback_base", "172.16.176"),
]


def test_workbook_without_mgmt_subnets_has_no_missing_key_error():
    result = _run(_OTHER_REQUIRED_KEYS)
    assert not _missing_mgmt_subnets_errors(result), result.errors


def test_workbook_without_mgmt_subnets_does_not_flag_unknown_key():
    # mgmt_subnets simply isn't a key in this sheet at all, so there's
    # nothing to flag as unknown — this just documents the absence case
    # doesn't produce any mgmt_subnets-related noise.
    result = _run(_OTHER_REQUIRED_KEYS)
    assert not any("mgmt_subnets" in e for e in result.errors), result.errors


# ─── Behavioral: presence of mgmt_subnets is a hard migration error (8b) ──

def _migration_errors(result):
    return [
        e for e in result.errors
        if "mgmt_subnets is no longer supported" in e or "OOB VLAN" in e
    ]


def test_workbook_with_valid_cidr_mgmt_subnets_now_errors():
    # Presence alone is now an error — even a
    # syntactically valid CIDR value no longer passes.
    pairs = _OTHER_REQUIRED_KEYS + [("mgmt_subnets", "192.168.200.0/24")]
    result = _run(pairs)
    assert _migration_errors(result), result.errors


def test_workbook_with_invalid_mgmt_subnets_cidr_also_errors():
    # Present-and-invalid also trips the same migration error — there's no
    # more CIDR-shape-specific validation to fall back on; any presence at
    # all is the failure condition.
    pairs = _OTHER_REQUIRED_KEYS + [("mgmt_subnets", "not-a-cidr")]
    result = _run(pairs)
    assert _migration_errors(result), result.errors


def test_workbook_with_mgmt_subnets_present_is_not_flagged_as_unknown_key():
    # Sanity: mgmt_subnets being OPTIONAL (not REQUIRED-and-removed) means
    # it stays in the known-keys union and is never flagged as unknown —
    # it fails with OUR specific migration message, not a generic warning.
    result = _run(_OTHER_REQUIRED_KEYS + [("mgmt_subnets", "192.168.200.0/24")])
    assert not any("Unknown key 'mgmt_subnets'" in e for e in result.warnings), result.warnings


def test_workbook_without_mgmt_subnets_has_no_migration_error():
    # Absence is the correct/expected state post-migration: no error at all.
    result = _run(_OTHER_REQUIRED_KEYS)
    assert not _migration_errors(result), result.errors
