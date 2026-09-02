# SPDX-FileCopyrightText: Copyright (c) 2026 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: MIT
"""
ERA-45: name/role consistency + plane symmetry. Both are WARNINGS — the Excel
Function stays authoritative (the parser trusts it); these only flag a likely
mislabel (root-caused on an 8SU partner submission, where gs-plane spines carried a
gl-plane Function and were left unprovisioned).
"""
import sys
from pathlib import Path

import openpyxl

SCRIPTS_DIR = Path(__file__).parent.parent / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from validate_excel import (ValidationResult, validate_nodes,  # noqa: E402
                            validate_plane_symmetry)

HEADERS = ['Function', 'Name', 'Type', 'MAC Address for ZTP', 'Mgmt IP Address',
           'Prefix', 'Gateway', 'ZTP', 'Enabled', 'Notes', 'OOB VLAN']


def _nodes_ws(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Nodes"
    for c, h in enumerate(HEADERS, 1):
        ws.cell(1, c, h)
    for r, row in enumerate(rows, 2):
        for c, v in enumerate(row, 1):
            ws.cell(r, c, v)
    return ws


def _row(function, name, ip='192.168.200.10'):
    return [function, name, '', '', ip, 24, '192.168.200.1', '', 'Yes', '', '']


class TestNameRoleWarn:
    """Part 1 — warn when the Name prefix implies a different role than Function."""

    def test_mislabeled_gpu_spine_warns(self):
        # gs-plane1-* (spine) name given a gl-plane1 (leaf) Function — the NLA case.
        ws = _nodes_ws([_row('gl-plane1', 'gs-plane1-01')])
        r = ValidationResult()
        validate_nodes(ws, r, settings={})
        assert any('gs-plane1-01' in w and 'implies role gs-plane1' in w
                   for w in r.warnings), r.warnings

    def test_matching_name_function_no_warn(self):
        ws = _nodes_ws([_row('gs-plane1', 'gs-plane1-01')])
        r = ValidationResult()
        validate_nodes(ws, r, settings={})
        assert not any('implies role' in w for w in r.warnings)

    def test_cl_csl_pair_is_exempt(self):
        # cl-* name resolves to 'csl' via the hostname fallback, but Function
        # 'cl' (dedicated leaf) is authoritative — this ambiguity must not warn.
        ws = _nodes_ws([_row('cl', 'cl-01')])
        r = ValidationResult()
        validate_nodes(ws, r, settings={})
        assert not any('implies role' in w for w in r.warnings)

    def test_server_name_without_role_no_warn(self):
        ws = _nodes_ws([_row('gpu', 'su-01-node-01')])
        r = ValidationResult()
        validate_nodes(ws, r, settings={})
        assert not any('implies role' in w for w in r.warnings)


def _n(function, name, enabled=True):
    return {'function': function, 'name': name, 'enabled': enabled,
            'is_air_documentary': False}


class TestPlaneSymmetry:
    """Part 2 — warn on asymmetric per-role switch counts across GPU planes."""

    def test_asymmetric_planes_warn(self):
        nodes = ([_n('gs-plane1', f'gs-plane1-0{i}') for i in range(1, 4)]
                 + [_n('gs-plane2', 'gs-plane2-01')])  # 3 vs 1
        r = ValidationResult()
        validate_plane_symmetry(nodes, r)
        assert any('Plane asymmetry' in w and 'gs-plane1' in w
                   for w in r.warnings), r.warnings

    def test_symmetric_planes_no_warn(self):
        nodes = [_n('gs-plane1', 'gs-plane1-01'), _n('gs-plane2', 'gs-plane2-01'),
                 _n('gl-plane1', 'gl-plane1-01'), _n('gl-plane2', 'gl-plane2-01')]
        r = ValidationResult()
        validate_plane_symmetry(nodes, r)
        assert not any('Plane asymmetry' in w for w in r.warnings)

    def test_single_plane_no_warn(self):
        nodes = [_n('gs-plane1', 'gs-plane1-01'), _n('gs-plane1', 'gs-plane1-02')]
        r = ValidationResult()
        validate_plane_symmetry(nodes, r)
        assert not any('Plane asymmetry' in w for w in r.warnings)

    def test_disabled_switch_not_counted(self):
        # a disabled plane2 switch shouldn't create a phantom asymmetry
        nodes = [_n('gs-plane1', 'gs-plane1-01'), _n('gs-plane2', 'gs-plane2-01'),
                 _n('gs-plane2', 'gs-plane2-02', enabled=False)]
        r = ValidationResult()
        validate_plane_symmetry(nodes, r)
        assert not any('Plane asymmetry' in w for w in r.warnings)
