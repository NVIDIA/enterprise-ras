# SPDX-FileCopyrightText: Copyright (c) 2026 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: MIT
"""ERA-58: `link auto-negotiate` must be expressible in the Port Profiles sheet.

Production disables auto-negotiation on data sub-ports — 35-96 stanzas per
switch, ~28% of the entire residual on one capture, and on one GPU leaf
`link speed` + `link auto-negotiate` together account for 192 of 300 misses.
We had no Port Profiles column, no Settings field, and no template emitted it.

The defect is that it was UNREPRESENTABLE, not that it was unset. The OEM
reference configs carry NO `link auto-negotiate` at all, so defaulting a value
would diverge from the endorsed design; the shipped models therefore leave the
column blank and nothing is emitted. These tests pin both halves: silent by
default, and actually emitted when an operator fills the column in.
"""
import re
import shutil
import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest
import yaml

NC = Path(__file__).resolve().parent.parent
ARCHS = ["2-4-3-200", "2-4-5-800", "2-8-5-200",
         "2-8-9-400", "2-8-9-400-SP", "2-8-9-800"]
_COPY_IGNORE = shutil.ignore_patterns(
    "output", "archive", ".git", ".venv*", "__pycache__", "*.pyc",
    ".pytest_cache", ".era-secrets", "legacy",
)


def _profile_header(xlsx):
    ws = openpyxl.load_workbook(xlsx)["VLANs & Profiles"]
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(r, 1).value or "").strip() == "Profile":
            return r, [ws.cell(r, c).value for c in range(1, 12)]
    return None, []


@pytest.mark.parametrize("arch", ARCHS)
def test_workbook_exposes_the_auto_negotiate_column(arch):
    """The column must exist so the attribute can be expressed at all."""
    xlsx = NC / "input" / arch / "default" / f"{arch}.xlsx"
    if not xlsx.exists():
        pytest.skip(f"no default workbook for {arch}")
    _, header = _profile_header(xlsx)
    assert header and header[10] == "Auto-Negotiate", (
        f"{arch}: Port Profiles has no Auto-Negotiate column (header={header})")


@pytest.mark.parametrize("arch", ARCHS)
def test_column_is_appended_not_inserted(arch):
    """Columns 1-10 must keep their positions.

    excel_parser reads LACP Bypass / Speed / Breakout / Lanes at fixed indices
    7-10. Inserting a column instead of appending would shift them — the exact
    failure mode ERA-81 records for the Nodes sheet.
    """
    xlsx = NC / "input" / arch / "default" / f"{arch}.xlsx"
    if not xlsx.exists():
        pytest.skip(f"no default workbook for {arch}")
    _, header = _profile_header(xlsx)
    assert header[:10] == [
        "Profile", "Port Mode", "Native/Access VLAN", "Allowed VLANs",
        "Untagged VLAN", "VRF", "LACP Bypass", "Speed", "Breakout", "Lanes",
    ], f"{arch}: Port Profiles columns 1-10 moved: {header[:10]}"


@pytest.mark.parametrize("arch", ARCHS)
def test_shipped_workbooks_emit_no_auto_negotiate(arch):
    """Blank by default: the OEM reference carries no `link auto-negotiate`."""
    d = NC / "output" / arch / "default" / "configs"
    if not d.is_dir():
        pytest.skip(f"no generated configs for {arch}")
    offenders = [p.name for p in d.glob("*-config.sh")
                 if "auto-negotiate" in p.read_text()]
    assert not offenders, (
        f"{arch}: emitted auto-negotiate with a blank column: {offenders}")


def test_setting_the_column_emits_the_stanza(tmp_path):
    """End to end: fill the column in and the config must carry it.

    Without this the column would be decorative — present in the sheet, read by
    nothing, which is precisely the state ERA-55 found the Speed column in.
    """
    arch = "2-4-3-200"
    src = NC / "input" / arch / "default" / f"{arch}.xlsx"
    if not src.exists():
        pytest.skip(f"no default workbook for {arch}")
    dst = tmp_path / "nc"
    shutil.copytree(NC, dst, ignore=_COPY_IGNORE)
    xlsx = dst / "input" / arch / "default" / f"{arch}.xlsx"

    wb = openpyxl.load_workbook(xlsx)
    ws = wb["VLANs & Profiles"]
    hdr = None
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(r, 1).value or "").strip() == "Profile":
            hdr = r
            break
    assert hdr, "no Port Profiles header"
    touched = 0
    for r in range(hdr + 1, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if name in (None, ""):
            break
        if str(name).strip() == "CPU/In-Band Network":
            ws.cell(r, 11).value = "disabled"
            touched += 1
    assert touched, "no CPU/In-Band Network profile row to set"
    wb.save(xlsx)

    r = subprocess.run(
        [sys.executable, "scripts/excel_parser.py", "--arch", arch,
         "--site", "default", "--skip-validate"],
        cwd=dst, capture_output=True, text=True)
    assert r.returncode == 0, f"parser failed:\n{r.stderr[-2000:]}"

    gen = subprocess.run(
        ["ansible-playbook", "playbooks/generate-cli-configs.yml",
         "-i", f"output/{arch}/default/inventory/hosts",
         "-e", f"config_output_dir=../output/{arch}/default/configs"],
        cwd=dst, capture_output=True, text=True)
    assert gen.returncode == 0, f"generate failed:\n{gen.stderr[-2000:]}"

    cfg = (dst / "output" / arch / "default" / "configs" / "core-01-config.sh").read_text()
    m = re.search(r"nv set interface (\S+) link auto-negotiate disabled", cfg)
    assert m, "Auto-Negotiate was set in the sheet but no stanza was emitted"
    assert "s" in m.group(1), (
        f"auto-negotiate should name sub-ports, got {m.group(1)}")


# --- injection safety -------------------------------------------------------
# The Auto-Negotiate cell renders UNQUOTED into a root-executed switch config.
# SHELL_INJECTION_PRONE_KEYS guards the Settings sheet only, so Port Profiles
# cells need their own protection. Both layers are asserted: validate_excel
# refuses the workbook, and the parser refuses the value even when validation
# is skipped.

_INJECTION = "disabled; touch /tmp/era_pwned_autoneg"


def _workbook_with_autoneg(tmp_path, value):
    arch = "2-4-3-200"
    src = NC / "input" / arch / "default" / f"{arch}.xlsx"
    if not src.exists():
        pytest.skip(f"no default workbook for {arch}")
    dst = tmp_path / "nc"
    shutil.copytree(NC, dst, ignore=_COPY_IGNORE)
    xlsx = dst / "input" / arch / "default" / f"{arch}.xlsx"
    wb = openpyxl.load_workbook(xlsx)
    ws = wb["VLANs & Profiles"]
    hdr = next(r for r in range(1, ws.max_row + 1)
               if str(ws.cell(r, 1).value or "").strip() == "Profile")
    for r in range(hdr + 1, ws.max_row + 1):
        n = ws.cell(r, 1).value
        if n in (None, ""):
            break
        if str(n).strip() == "CPU/In-Band Network":
            ws.cell(r, 11).value = value
            break
    wb.save(xlsx)
    return dst, arch, xlsx


def test_validate_excel_rejects_shell_metacharacters_in_auto_negotiate(tmp_path):
    dst, arch, xlsx = _workbook_with_autoneg(tmp_path, _INJECTION)
    r = subprocess.run(
        [sys.executable, "scripts/validate_excel.py", str(xlsx)],
        cwd=dst, capture_output=True, text=True)
    out = r.stdout + r.stderr
    assert r.returncode != 0, "workbook with an injected Auto-Negotiate value validated cleanly"
    assert "command injection" in out, out[-800:]


def test_parser_refuses_non_whitelisted_auto_negotiate(tmp_path):
    """Defence in depth: --skip-validate must not be a way around it."""
    dst, arch, xlsx = _workbook_with_autoneg(tmp_path, _INJECTION)
    r = subprocess.run(
        [sys.executable, "scripts/excel_parser.py", "--arch", arch,
         "--site", "default", "--skip-validate"],
        cwd=dst, capture_output=True, text=True)
    assert r.returncode == 0
    core = yaml.safe_load(
        (dst / "output" / arch / "default" / "inventory" / "group_vars" / "core.yml").read_text())
    got = (core.get("network_roles") or {}).get("cpu", {}).get("auto_negotiate")
    assert got is None, f"injected value survived parsing as {got!r}"


def test_legitimate_auto_negotiate_values_still_accepted(tmp_path):
    """The whitelist must not be so tight it breaks the feature."""
    for value in ("disabled", "enabled", "on", "off"):
        dst, arch, xlsx = _workbook_with_autoneg(tmp_path / value, value)
        subprocess.run(
            [sys.executable, "scripts/excel_parser.py", "--arch", arch,
             "--site", "default", "--skip-validate"],
            cwd=dst, capture_output=True, text=True, check=True)
        core = yaml.safe_load(
            (dst / "output" / arch / "default" / "inventory" / "group_vars" / "core.yml").read_text())
        got = (core.get("network_roles") or {}).get("cpu", {}).get("auto_negotiate")
        assert got == value, f"{value!r} was rejected, got {got!r}"
