# SPDX-FileCopyrightText: Copyright (c) 2026 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: MIT
"""ERA-40: every advertised subnet is covered by ERA_PREFIXES, and no isolated one is.

`ERA_PREFIXES` is the prefix-list the EXIT VRF's outbound route-map
(`OUTBOUND_ERA_PREFIXES`) applies toward `cust-net-edge`. Membership in it is
what decides whether a subnet is reachable from the customer network.

The parser *derives* that list: rule 10/20 cover the loopback supernet
(`<loopback_base>.0/21` at max-prefix-len 24, plus `.0/24` at 32), and it then
walks the INBAND and OOB VLANs adding explicit rules for anything those two
rules miss. Nothing checked the derivation, and ERA-40 asked whether such a
check existed. It did not.

Measured across all six archs x both shipped sites at the time of writing:

    VRF        covered   NOT covered
    INBAND        24           0
    OOB           12           0
    GPU            0          54

That is the correct shape, and both halves of it are load-bearing:

* An INBAND or OOB subnet that falls out of coverage is **silently
  unreachable** from the customer network. Nothing in the config looks wrong;
  the prefix simply never advertises.
* A GPU subnet that falls *into* coverage **leaks the East/West compute fabric
  to the customer network**. That is an isolation failure, and it is the more
  serious direction — which is why this asserts it rather than only checking
  for gaps.

The derivation hardcodes `('INBAND', 'OOB')`. If a future arch advertises a
third VRF, these tests fail rather than letting it go quietly unadvertised.
"""
import ipaddress
import re
from pathlib import Path

import openpyxl
import pytest

NC = Path(__file__).resolve().parent.parent
_SHIPPED_SITES = ("default", "largescale")

# VRFs whose subnets MUST reach cust-net-edge.
_ADVERTISED_VRFS = frozenset({"INBAND", "OOB"})
# VRFs that must NEVER be advertised — the East/West compute fabric is
# deliberately isolated from the customer network.
_ISOLATED_VRFS = frozenset({"GPU"})

_ERA_RULE = re.compile(
    r"prefix-list ERA_PREFIXES rule \d+ match (\S+) max-prefix-len (\d+)")


def _era_prefixes(arch, site):
    """The ERA_PREFIXES rules as (network, max_prefix_len), or None."""
    cfgdir = NC / "output" / arch / site / "configs"
    for cfg in sorted(cfgdir.glob("*-config.sh")):
        txt = cfg.read_text()
        if "ERA_PREFIXES rule" not in txt:
            continue
        return [(ipaddress.ip_network(m.group(1), strict=False), int(m.group(2)))
                for m in _ERA_RULE.finditer(txt)]
    return None


def _covered(net, rules):
    """NVUE semantics: inside the rule's network AND no longer than max-len."""
    return any(net.subnet_of(n) and net.prefixlen <= mx for n, mx in rules)


def _vlan_subnets(xlsx):
    """-> [(vrf, subnet_str, ip_network)] from the VLANs & Profiles sheet."""
    ws = openpyxl.load_workbook(xlsx, data_only=True)["VLANs & Profiles"]
    hdr = {}
    for r in range(1, min(ws.max_row + 1, 6)):
        for c in range(1, ws.max_column + 1):
            v = str(ws.cell(r, c).value or "").strip().lower()
            if v in ("vlan id", "subnet", "vrf"):
                hdr.setdefault(v, (r, c))
        if "vlan id" in hdr:
            break
    if not {"vlan id", "subnet", "vrf"} <= set(hdr):
        return []
    out = []
    for r in range(hdr["vlan id"][0] + 1, ws.max_row + 1):
        sn = str(ws.cell(r, hdr["subnet"][1]).value or "").strip()
        if "/" not in sn:
            continue
        vrf = str(ws.cell(r, hdr["vrf"][1]).value or "").strip()
        try:
            out.append((vrf, sn, ipaddress.ip_network(sn, strict=False)))
        except ValueError:
            continue
    return out


def _cases():
    for xlsx in sorted(NC.glob("input/*/*/*.xlsx")):
        arch, site = xlsx.parts[-3], xlsx.parts[-2]
        if site not in _SHIPPED_SITES:
            continue
        yield pytest.param(arch, site, xlsx, id=f"{arch}/{site}")


CASES = list(_cases())


def test_cases_exist():
    """An empty glob would make every parametrised test vacuously pass.

    This file exists partly because the first version of this analysis reported
    a clean "0 uncovered" while reading a key that did not exist.
    """
    assert CASES, "no shipped workbooks found"


@pytest.mark.parametrize("arch,site,xlsx", CASES)
def test_advertised_vrf_subnets_are_covered(arch, site, xlsx):
    """A gap here is silent: the prefix simply never reaches cust-net-edge."""
    rules = _era_prefixes(arch, site)
    if rules is None:
        pytest.skip(f"{arch}/{site}: no config emits ERA_PREFIXES")
    assert rules, f"{arch}/{site}: ERA_PREFIXES is empty"

    subnets = _vlan_subnets(xlsx)
    assert subnets, f"{arch}/{site}: no VLAN subnets read from the workbook"

    checked = 0
    missing = []
    for vrf, sn, net in subnets:
        if vrf.upper() not in _ADVERTISED_VRFS:
            continue
        checked += 1
        if not _covered(net, rules):
            missing.append(f"{vrf} {sn}")
    assert checked, (
        f"{arch}/{site}: no {sorted(_ADVERTISED_VRFS)} subnets examined — the "
        f"VRF column may have moved, which would make this test vacuous")
    assert not missing, (
        f"{arch}/{site}: advertised subnets not covered by ERA_PREFIXES: "
        f"{missing} — these would be silently unreachable from the customer "
        f"network")


@pytest.mark.parametrize("arch,site,xlsx", CASES)
def test_isolated_vrf_subnets_are_not_advertised(arch, site, xlsx):
    """The serious direction: compute-fabric prefixes must not leak outbound."""
    rules = _era_prefixes(arch, site)
    if rules is None:
        pytest.skip(f"{arch}/{site}: no config emits ERA_PREFIXES")

    subnets = _vlan_subnets(xlsx)
    checked = 0
    leaked = []
    for vrf, sn, net in subnets:
        if vrf.upper() not in _ISOLATED_VRFS:
            continue
        checked += 1
        if _covered(net, rules):
            leaked.append(f"{vrf} {sn}")
    if not checked:
        pytest.skip(f"{arch}/{site}: no isolated-VRF subnets declared")
    assert not leaked, (
        f"{arch}/{site}: isolated-VRF subnets ARE advertised by ERA_PREFIXES: "
        f"{leaked} — this exposes the East/West compute fabric to the customer "
        f"network")


@pytest.mark.parametrize("arch,site,xlsx", CASES)
def test_no_unclassified_vrf_slips_through(arch, site, xlsx):
    """Every VRF carrying a subnet must have a stated advertise/isolate intent.

    Without this, adding a VRF to the workbook would leave it covered by
    neither test above — unadvertised by default and unnoticed, which is the
    exact failure mode ERA-40 was raised about.
    """
    subnets = _vlan_subnets(xlsx)
    if not subnets:
        pytest.skip(f"{arch}/{site}: no VLAN subnets")
    known = _ADVERTISED_VRFS | _ISOLATED_VRFS
    unknown = sorted({v for v, _, _ in subnets if v and v.upper() not in known})
    assert not unknown, (
        f"{arch}/{site}: VRF(s) {unknown} carry subnets but are classified "
        f"neither advertised nor isolated. Decide which, and add them to the "
        f"set in this file — do not leave it implicit")


# ---------------------------------------------------------------------------
# The validator half (validate_excel.validate_prefix_alignment).
#
# The tests above check the six shipped workbooks. An operator's own workbook
# never reaches them, so the leak direction is also guarded at ingest.
# ---------------------------------------------------------------------------

import shutil
import subprocess
import sys

_LOOPBACK_SUPERNET = "172.16.176.0/21"


def _workbook_copy(tmp_path, arch="2-8-9-800"):
    src = NC / "input" / arch / "default" / f"{arch}.xlsx"
    if not src.exists():
        pytest.skip(f"no shipped workbook for {arch}")
    dst = tmp_path / f"{arch}.xlsx"
    shutil.copy(src, dst)
    return dst


def _validate(xlsx):
    return subprocess.run(
        [sys.executable, "scripts/validate_excel.py", str(xlsx)],
        cwd=NC, capture_output=True, text=True)


def _set_vlan_subnet(xlsx, want_vrf, new_subnet):
    """Move the first VLAN in `want_vrf` onto `new_subnet`. -> the VLAN id."""
    wb = openpyxl.load_workbook(xlsx)
    ws = wb["VLANs & Profiles"]
    hdr = {}
    for r in range(1, min(ws.max_row + 1, 6)):
        for c in range(1, ws.max_column + 1):
            v = str(ws.cell(r, c).value or "").strip().lower()
            if v in ("vlan id", "subnet", "vrf"):
                hdr.setdefault(v, (r, c))
        if "vlan id" in hdr:
            break
    for r in range(hdr["vlan id"][0] + 1, ws.max_row + 1):
        vrf = str(ws.cell(r, hdr["vrf"][1]).value or "").strip().upper()
        if vrf == want_vrf:
            ws.cell(r, hdr["subnet"][1]).value = new_subnet
            vid = ws.cell(r, hdr["vlan id"][1]).value
            wb.save(xlsx)
            return vid
    pytest.skip(f"no {want_vrf} VLAN to move")


def test_shipped_workbook_passes_prefix_alignment(tmp_path):
    """False-positive guard — a guard that fires on correct input is worse than none."""
    xlsx = _workbook_copy(tmp_path)
    r = _validate(xlsx)
    combined = r.stdout + r.stderr
    assert "falls inside the advertised supernet" not in combined, (
        f"validator fired on an unmodified shipped workbook:\n{combined[-1500:]}")


def test_gpu_subnet_inside_the_advertised_supernet_is_rejected(tmp_path):
    """The leak the validator exists for.

    A GPU VLAN addressed inside <loopback_base>.0/21 at <=/24 is advertised to
    cust-net-edge by ERA_PREFIXES rule 10 — a rule that never names it. Nothing
    in the workbook or the rendered config looks wrong.
    """
    xlsx = _workbook_copy(tmp_path)
    vid = _set_vlan_subnet(xlsx, "GPU", "172.16.180.0/24")  # inside the /21

    net = ipaddress.ip_network("172.16.180.0/24")
    assert net.subnet_of(ipaddress.ip_network(_LOOPBACK_SUPERNET)), (
        "test fixture is wrong — the injected subnet must be inside the supernet")

    r = _validate(xlsx)
    combined = r.stdout + r.stderr
    assert "falls inside the advertised supernet" in combined, (
        f"GPU subnet inside the supernet was NOT flagged (VLAN {vid}):\n"
        f"{combined[-2000:]}")
    assert r.returncode != 0, "leak was reported as a warning, not an error"


def test_gpu_subnet_outside_the_supernet_is_accepted(tmp_path):
    """The shipped arrangement: GPU lives in 192.168.x, well clear of the /21."""
    xlsx = _workbook_copy(tmp_path)
    _set_vlan_subnet(xlsx, "GPU", "192.168.99.0/24")
    r = _validate(xlsx)
    assert "falls inside the advertised supernet" not in (r.stdout + r.stderr)


# ---------------------------------------------------------------------------
# The warning path.
#
# This branch shipped BROKEN: it called `result.warning(...)`, and
# ValidationResult exposes `warn()`. It crashed `make import` with an
# AttributeError for any workbook carrying a VRF outside the classified sets.
#
# The original tests did not catch it because they only ever exercised the
# ERROR branch. All twelve shipped workbooks classify every VRF they use, so
# the warning line was never executed by anything — a guard with an unreachable
# branch is a guard with untested code in it.
#
# The e2e fixtures DO carry a STORAGE VLAN subnet (the shipped workbooks no
# longer do, per ADR-0047), so the first thing to touch this branch was CI.
# ---------------------------------------------------------------------------

_E2E_FIXTURES = sorted((NC.parent / "release" / "e2e" / "fixtures").glob("*.xlsx"))


@pytest.mark.parametrize("fx", _E2E_FIXTURES, ids=lambda p: p.stem)
def test_e2e_fixtures_validate_without_crashing(fx):
    """`make import` must not die on the fixtures CI actually deploys."""
    r = _validate(fx)
    combined = r.stdout + r.stderr
    assert "Traceback" not in combined and "AttributeError" not in combined, (
        f"{fx.name}: validate_excel crashed:\n{combined[-1500:]}")


def test_storage_vrf_does_not_warn(tmp_path):
    """STORAGE is routed via its own VRF and uplink, not via ERA_PREFIXES.

    Absence from ERA_PREFIXES is CORRECT for it, so warning would be a false
    positive — and false positives on a guard are how people learn to ignore it.
    """
    xlsx = _workbook_copy(tmp_path)
    _set_vlan_subnet(xlsx, "GPU", "192.168.99.0/24")  # keep GPU valid
    wb = openpyxl.load_workbook(xlsx)
    ws = wb["VLANs & Profiles"]
    hdr = {}
    for r in range(1, min(ws.max_row + 1, 6)):
        for c in range(1, ws.max_column + 1):
            v = str(ws.cell(r, c).value or "").strip().lower()
            if v in ("vlan id", "subnet", "vrf"):
                hdr.setdefault(v, (r, c))
        if "vlan id" in hdr:
            break
    # Retag one GPU row as STORAGE with an off-supernet subnet.
    for r in range(hdr["vlan id"][0] + 1, ws.max_row + 1):
        if str(ws.cell(r, hdr["vrf"][1]).value or "").strip().upper() == "GPU":
            ws.cell(r, hdr["vrf"][1]).value = "STORAGE"
            ws.cell(r, hdr["subnet"][1]).value = "10.10.10.0/24"
            break
    wb.save(xlsx)

    r = _validate(xlsx)
    combined = r.stdout + r.stderr
    assert "Traceback" not in combined, f"crashed:\n{combined[-1200:]}"
    assert "classified neither advertised" not in combined, (
        "STORAGE warned — it is separately routed, so this is a false positive")


def test_a_genuinely_unknown_vrf_warns_and_does_not_crash(tmp_path):
    """The branch that was dead code. Must warn cleanly, not raise."""
    xlsx = _workbook_copy(tmp_path)
    wb = openpyxl.load_workbook(xlsx)
    ws = wb["VLANs & Profiles"]
    hdr = {}
    for r in range(1, min(ws.max_row + 1, 6)):
        for c in range(1, ws.max_column + 1):
            v = str(ws.cell(r, c).value or "").strip().lower()
            if v in ("vlan id", "subnet", "vrf"):
                hdr.setdefault(v, (r, c))
        if "vlan id" in hdr:
            break
    for r in range(hdr["vlan id"][0] + 1, ws.max_row + 1):
        if str(ws.cell(r, hdr["vrf"][1]).value or "").strip().upper() == "GPU":
            ws.cell(r, hdr["vrf"][1]).value = "TENANT_X"
            ws.cell(r, hdr["subnet"][1]).value = "10.20.30.0/24"
            break
    wb.save(xlsx)

    r = _validate(xlsx)
    combined = r.stdout + r.stderr
    assert "Traceback" not in combined and "AttributeError" not in combined, (
        f"the warning branch still raises:\n{combined[-1500:]}")
    assert "classified neither advertised" in combined, (
        "an unknown VRF produced no warning — the branch is unreachable again")
