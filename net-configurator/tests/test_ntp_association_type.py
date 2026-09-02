# SPDX-FileCopyrightText: Copyright (c) 2026 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: MIT
"""ERA-59: `association-type` is emitted by every role and settable per server.

Before this, `nv set system ntp server <s> association-type server` was emitted
by the SPINE template only — core, gl and oob-switch omitted it — so every
generated config except a dedicated cs-/gs- spine lacked the suffix. One role
template differing from three by accident is a defect regardless of which value
is right.

The two references disagree by vintage and neither settles it: the OEM
reference configs set no NTP at all, and the July-2026 production captures are
split *by fabric* (2852 spine yes / oob no; nw2432 both yes; 2894 oob mixed) —
the signature of a site setting, not an architectural one. So it is emitted
explicitly everywhere, defaulting to NVUE's own default `server`, and made
settable per server so a site can ask for `pool` or `peer`.

Security note: `ntp_servers` is a SHELL_INJECTION_PRONE_KEY rendered into a
root-executed config. The address and the type are kept in separate structures
so the address still renders through `| quote` and the type comes from a closed
whitelist — neither half can carry a spaced or injected value.
"""
import re
import shutil
import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest
import yaml

NC = Path(__file__).resolve().parent.parent
ARCH = "2-8-9-800"
ARCHS = ["2-4-3-200", "2-4-5-800", "2-8-5-200",
         "2-8-9-400", "2-8-9-400-SP", "2-8-9-800"]
NTP_RE = re.compile(r"nv set system ntp server (\S+)(.*)$")
_COPY_IGNORE = shutil.ignore_patterns(
    "output", "archive", ".git", ".venv*", "__pycache__", "*.pyc",
    ".pytest_cache", ".era-secrets", "legacy",
)


def test_every_role_template_emits_association_type():
    """All four role templates must agree — this is ERA-59's actual complaint."""
    missing = []
    for tpl in sorted(NC.glob("roles/*/templates/*_nvue_cli.j2")):
        txt = tpl.read_text()
        if "nv set system ntp server" not in txt:
            continue
        if "association-type" not in txt:
            missing.append(tpl.name)
    assert not missing, f"templates emit NTP without association-type: {missing}"


@pytest.mark.parametrize("arch", ARCHS)
def test_generated_configs_carry_association_type(arch):
    d = NC / "output" / arch / "default" / "configs"
    if not d.is_dir():
        pytest.skip(f"no generated configs for {arch}")
    for cfg in sorted(d.glob("*-config.sh")):
        for line in cfg.read_text().splitlines():
            m = NTP_RE.match(line.strip())
            if m:
                assert "association-type" in m.group(2), (
                    f"{arch}/{cfg.name}: {line.strip()!r} has no association-type")


@pytest.mark.parametrize("arch", ARCHS)
def test_default_association_type_is_server(arch):
    """NVUE's own default. Emitting it explicitly changes nothing functionally."""
    d = NC / "output" / arch / "default" / "configs"
    if not d.is_dir():
        pytest.skip(f"no generated configs for {arch}")
    types = set()
    for cfg in d.glob("*-config.sh"):
        for m in re.finditer(r"ntp server \S+ association-type (\S+)", cfg.read_text()):
            types.add(m.group(1))
    if types:
        assert types == {"server"}, f"{arch}: unexpected association types {types}"


@pytest.mark.parametrize("value,expected", [
    ("2.pool.ntp.org association-type pool", "pool"),
    ("3.pool.ntp.org peer", "peer"),
])
def test_per_server_association_type_is_honoured(value, expected, tmp_path):
    """A site must be able to ask for pool/peer, in either spelling.

    The verbose `association-type pool` form mirrors NVUE so the sheet documents
    itself; the bare `pool` shorthand is accepted too.
    """
    src = NC / "input" / ARCH / "default" / f"{ARCH}.xlsx"
    if not src.exists():
        pytest.skip(f"no default workbook for {ARCH}")
    dst = tmp_path / "nc"
    shutil.copytree(NC, dst, ignore=_COPY_IGNORE)
    xlsx = dst / "input" / ARCH / "default" / f"{ARCH}.xlsx"

    wb = openpyxl.load_workbook(xlsx)
    ws = wb["Settings"]
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(r, 1).value or "").strip() == "ntp_servers":
            ws.cell(r, 2).value = f"0.pool.ntp.org\n{value}"
            break
    else:
        pytest.fail("no ntp_servers row in Settings")
    wb.save(xlsx)

    r = subprocess.run(
        [sys.executable, "scripts/excel_parser.py", "--arch", ARCH,
         "--site", "default", "--skip-validate"],
        cwd=dst, capture_output=True, text=True)
    assert r.returncode == 0, f"parser failed:\n{r.stderr[-2000:]}"

    allvars = yaml.safe_load(
        (dst / "output" / ARCH / "default" / "inventory" / "group_vars" / "all" / "main.yml").read_text())
    addr = value.split()[0]
    assert allvars["ntp_servers"] == ["0.pool.ntp.org", addr], (
        f"addresses not split cleanly: {allvars['ntp_servers']}")
    assert allvars["ntp_association_types"][addr] == expected
    assert allvars["ntp_association_types"]["0.pool.ntp.org"] == "server"


def test_address_never_carries_the_type_through(tmp_path):
    """The address list must stay bare addresses.

    `ntp_servers` is a SHELL_INJECTION_PRONE_KEY rendered through `| quote`.
    If the type were left on the address, the whole string would be quoted as
    one argument and NVUE would get a hostname with spaces in it.
    """
    src = NC / "input" / ARCH / "default" / f"{ARCH}.xlsx"
    if not src.exists():
        pytest.skip(f"no default workbook for {ARCH}")
    dst = tmp_path / "nc"
    shutil.copytree(NC, dst, ignore=_COPY_IGNORE)
    xlsx = dst / "input" / ARCH / "default" / f"{ARCH}.xlsx"
    wb = openpyxl.load_workbook(xlsx)
    ws = wb["Settings"]
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(r, 1).value or "").strip() == "ntp_servers":
            ws.cell(r, 2).value = "0.pool.ntp.org association-type pool"
            break
    wb.save(xlsx)
    subprocess.run(
        [sys.executable, "scripts/excel_parser.py", "--arch", ARCH,
         "--site", "default", "--skip-validate"],
        cwd=dst, capture_output=True, text=True, check=True)
    allvars = yaml.safe_load(
        (dst / "output" / ARCH / "default" / "inventory" / "group_vars" / "all" / "main.yml").read_text())
    for s in allvars["ntp_servers"]:
        assert " " not in s, f"address {s!r} still carries its association type"


def test_unknown_association_type_falls_back_to_server(tmp_path):
    """A typo must not reach a root-executed config script."""
    src = NC / "input" / ARCH / "default" / f"{ARCH}.xlsx"
    if not src.exists():
        pytest.skip(f"no default workbook for {ARCH}")
    dst = tmp_path / "nc"
    shutil.copytree(NC, dst, ignore=_COPY_IGNORE)
    xlsx = dst / "input" / ARCH / "default" / f"{ARCH}.xlsx"
    wb = openpyxl.load_workbook(xlsx)
    ws = wb["Settings"]
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(r, 1).value or "").strip() == "ntp_servers":
            ws.cell(r, 2).value = "0.pool.ntp.org association-type poool"
            break
    wb.save(xlsx)
    r = subprocess.run(
        [sys.executable, "scripts/excel_parser.py", "--arch", ARCH,
         "--site", "default", "--skip-validate"],
        cwd=dst, capture_output=True, text=True)
    assert r.returncode == 0
    allvars = yaml.safe_load(
        (dst / "output" / ARCH / "default" / "inventory" / "group_vars" / "all" / "main.yml").read_text())
    assert allvars["ntp_association_types"]["0.pool.ntp.org"] == "server"
    assert "poool" in (r.stdout + r.stderr), "silent fallback — operator gets no warning"
