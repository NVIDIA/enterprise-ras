# SPDX-FileCopyrightText: Copyright (c) 2026 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: MIT
"""The two Networking Logical Architecture criteria the validator did not check.

DRB Guidelines slide 12, verbatim:

    "For server-to-server communication on the E/W Network, Private IP addresses
     with no public IPs used internally"

    "For E/W network, there should not be an oversubscription ratio in your
     network design"

Both are endorsement criteria, so they live here rather than in the internal audit
harness: an OEM running `make validate-excel` catches them before submitting, which is
the whole point of ADR-0053 clause 2.

Oversubscription is only defined once an E/W spine tier exists. On a collapsed fabric
every leaf port is a downlink or a peer ISL, so the check stays silent rather than
inventing a verdict — 2-8-9-800 at 1 SU is exactly that shape and must not warn.
"""
import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "scripts"))
import validate_excel as _ve  # noqa: E402  (needs the sys.path insert above)
# Arch models are internal-only (data-models/, ADR-0027) and absent from the
# public tree, so this skips there rather than failing. Defined locally rather
# than imported from conftest: conftest is loaded by pytest, not importable.
_HAVE_ARCH_MODELS = (Path(__file__).resolve().parents[2] / "data-models" / "models").exists()
needs_arch_models = pytest.mark.skipif(
    not _HAVE_ARCH_MODELS,
    reason="arch models are internal-only (data-models/), absent in the public tree")


class _Result:
    """Minimal stand-in for the validator's result object."""

    def __init__(self):
        self.warnings = []
        self.errors = []

    def warn(self, section, message):
        self.warnings.append((section, message))

    def error(self, section, message):
        self.errors.append((section, message))


# --- private-IP helper ----------------------------------------------------------------

def test_rfc1918_subnet_passes():
    assert _ve._ew_subnets_are_private(["192.168.0.0/20", "192.168.16.0/20"]) == []


def test_public_subnet_is_reported():
    assert _ve._ew_subnets_are_private(["8.8.8.0/24"]) == ["8.8.8.0/24"]


def test_mixed_reports_only_the_public_one():
    assert _ve._ew_subnets_are_private(["192.168.0.0/20", "9.9.9.0/24"]) == ["9.9.9.0/24"]


def test_malformed_subnet_is_ignored_not_crashed():
    assert _ve._ew_subnets_are_private(["not-a-subnet"]) == []


def test_blank_and_none_are_ignored():
    """Empty Subnet cells read as None/'' and must not be reported as public."""
    assert _ve._ew_subnets_are_private([None, "", "   "]) == []


def test_ipv6_does_not_crash_the_cgnat_test():
    """`subnet_of` raises across address families. A ULA E/W subnet must be accepted,
    not turned into a TypeError that takes the whole validator down."""
    assert _ve._ew_subnets_are_private(["fd00::/8"]) == []
    # Not 2001:db8::/32 — that is the documentation range, which `ipaddress` reports
    # as private. A public global-unicast prefix is what must be caught.
    assert _ve._ew_subnets_are_private(["2001:4860::/32"]) == ["2001:4860::/32"]


def test_cgnat_is_not_rfc1918():
    """100.64.0.0/10 is carrier-grade NAT (RFC 6598), not private."""
    assert _ve._ew_subnets_are_private(["100.64.0.0/10"]) == ["100.64.0.0/10"]


def test_stdlib_still_agrees_cgnat_is_not_private():
    """Pins the assumption the explicit `_CGNAT` guard exists to survive.

    On Python 3.12 `ipaddress` omits 100.64.0.0/10 from `_private_networks`, so
    `is_private` is already False and the guard is redundant — a mutation test proved
    it: deleting the guard changed no behaviour and no test noticed. It is kept anyway
    because that membership list is a stdlib implementation detail that has moved
    between releases. Asserting the stdlib fact directly is what makes the guard
    meaningful: if an upgrade flips it, this fails and points at why the guard is there,
    instead of the criterion quietly widening.
    """
    import ipaddress
    assert ipaddress.ip_network("100.64.0.0/10").is_private is False


# --- workbook fixtures ----------------------------------------------------------------

def _workbook(vlans, profiles, links):
    """Build the two sheets the checks read. Columns match the shipped layout.

    `profiles` entries are (name, speed) or (name, speed, access_vlan). The VLAN is
    what ties a profile to its VRF, and so what marks it as E/W — a profile without
    one is deliberately invisible to the E/W walk.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    vp = wb.create_sheet("VLANs & Profiles")
    vp.append(["VLANs"])
    vp.append(["VLAN ID", "Name", "Purpose", "Subnet", "Gateway", "VRF", "VNI",
               "DHCP Relay Client"])
    for row in vlans:
        vp.append(row)
    vp.append([])
    vp.append(["Port Profiles"])
    vp.append(["Profile", "Port Mode", "Native/Access VLAN", "Allowed VLANs",
               "Untagged VLAN", "VRF", "LACP Bypass", "Speed", "Breakout", "Lanes"])
    for entry in profiles:
        name, speed = entry[0], entry[1]
        vlan = entry[2] if len(entry) > 2 else None
        vp.append([name, "Access", vlan, None, None, None, "No", speed, 2, 4])

    wm = wb.create_sheet("Wire Map")
    wm.append(["Display in Air", "System Name (A)", "Port (A)", "Port Side (A)",
               "Cable Split (A)", "System Name (B)", "Port (B)", "Port Side (B)",
               "Cable Split (B)", "Network Profile"])
    for a, b, profile in links:
        wm.append(["Yes", a, "swp1", "A", None, b, "swp1", "B", None, profile])
    return wb


_GPU_VLAN = [901, "gpu_rail1", "GPU Rail 1 E/W", "192.168.1.0/24",
             "192.168.1.1", "GPU", 4901, "No"]


def _leaf_spine(down_links, up_links, speed="400G"):
    """One E/W leaf with `down_links` GPU-facing links and `up_links` to a spine."""
    links = [(f"su-01-node-{i:02d}", "gl-plane1-01", "GPU Network")
             for i in range(1, down_links + 1)]
    links += [("gl-plane1-01", "gs-plane1-01", "ISL") for _ in range(up_links)]
    return _workbook(
        vlans=[[901, "gpu_rail1", "GPU Rail 1 E/W", "192.168.1.0/24", "192.168.1.1",
                "GPU", 4901, "No"]],
        profiles=[("GPU Network", speed, 901), ("ISL", speed)],
        links=links,
    )


# --- private-IP check -----------------------------------------------------------------

def test_private_ew_vlans_do_not_warn():
    r = _Result()
    _ve.validate_ew_uses_private_ips(_leaf_spine(4, 4), {}, r)
    assert r.warnings == []


def test_public_ew_vlan_warns():
    wb = _workbook(
        vlans=[[901, "gpu_rail1", "GPU Rail 1 E/W", "8.8.8.0/24", "8.8.8.1",
                "GPU", 4901, "No"]],
        profiles=[("GPU Network", "400G")],
        links=[],
    )
    r = _Result()
    _ve.validate_ew_uses_private_ips(wb, {}, r)
    assert len(r.warnings) == 1
    assert "8.8.8.0/24" in r.warnings[0][1]


def test_public_non_ew_vlan_is_not_this_checks_business():
    """The criterion is about the E/W network. A public range on the EXIT VRF is the
    external edge doing its job, and warning about it would be a false positive on
    every workbook with real customer connectivity.

    Uses 8.8.8.0/24, NOT a 203.0.113.0/24-style documentation range: `ipaddress`
    classifies TEST-NET as private, so a doc range here would make this test pass even
    if the E/W scoping were deleted — it would assert nothing at all.
    """
    wb = _workbook(
        vlans=[[500, "edge", "External Connectivity", "8.8.8.0/24", "8.8.8.1",
                "EXIT", 4500, "No"]],
        profiles=[("Edge Uplink", "200G")],
        links=[],
    )
    r = _Result()
    _ve.validate_ew_uses_private_ips(wb, {}, r)
    assert r.warnings == []


def test_documentation_ranges_are_accepted():
    """Slide 12 asks partners to "use the documentation IP address ranges for reporting
    purposes", so a submission that complies must not be flagged for it. `ipaddress`
    already treats TEST-NET-1/2/3 and 2001:db8::/32 as private — pinned here because
    that is load-bearing behaviour, not an incidental detail of the stdlib."""
    assert _ve._ew_subnets_are_private(
        ["192.0.2.0/24", "198.51.100.0/24", "203.0.113.0/24", "2001:db8::/32"]) == []


# --- oversubscription check -----------------------------------------------------------

def test_one_to_one_does_not_warn():
    r = _Result()
    _ve.validate_ew_not_oversubscribed(_leaf_spine(4, 4), {}, r)
    assert r.warnings == []


def test_oversubscribed_leaf_warns():
    r = _Result()
    _ve.validate_ew_not_oversubscribed(_leaf_spine(8, 2), {}, r)
    assert len(r.warnings) == 1
    msg = r.warnings[0][1]
    assert "gl-plane1-01" in msg
    assert "3200" in msg and "800" in msg


def test_over_provisioned_uplinks_do_not_warn():
    """More up than down is legal. The criterion is 'no oversubscription', not
    'exactly 1:1' — warning here would flag a deliberately generous fabric."""
    r = _Result()
    _ve.validate_ew_not_oversubscribed(_leaf_spine(2, 8), {}, r)
    assert r.warnings == []


def test_mixed_link_speeds_are_summed_not_counted():
    """A 2:1 link COUNT at double the uplink speed is 1:1 in bandwidth. Counting links
    instead of Gbps would warn on a fabric that is not oversubscribed at all."""
    links = [(f"su-01-node-{i:02d}", "gl-plane1-01", "GPU Network") for i in range(1, 5)]
    links += [("gl-plane1-01", "gs-plane1-01", "ISL") for _ in range(2)]
    wb = _workbook(
        vlans=[],
        profiles=[("GPU Network", "200G", 901), ("ISL", "400G")],
        links=links,
    )
    r = _Result()
    _ve.validate_ew_not_oversubscribed(wb, {}, r)
    assert r.warnings == []


def test_collapsed_fabric_is_silent():
    """No spine tier: every leaf port is a downlink or a peer ISL, so the ratio is
    undefined. Reporting 'infinitely oversubscribed' here would fail every collapsed
    design we ship."""
    links = [(f"su-01-node-{i:02d}", "gsl-plane1-01", "GPU Network") for i in range(1, 9)]
    links += [("gsl-plane1-01", "gsl-plane1-02", "ISL") for _ in range(4)]
    wb = _workbook(vlans=[_GPU_VLAN], profiles=[("GPU Network", "400G", 901),
                                              ("ISL", "400G")], links=links)
    r = _Result()
    _ve.validate_ew_not_oversubscribed(wb, {}, r)
    assert r.warnings == []


def test_unsized_profile_is_silent_not_guessed():
    """A profile whose Speed cell cannot be read leaves the ratio unknown. Treating it
    as 0 Gbps would report an oversubscription the workbook does not describe."""
    links = [(f"su-01-node-{i:02d}", "gl-plane1-01", "GPU Network") for i in range(1, 9)]
    links += [("gl-plane1-01", "gs-plane1-01", "ISL") for _ in range(8)]
    wb = _workbook(vlans=[_GPU_VLAN], profiles=[("GPU Network", "400G", 901),
                                              ("ISL", None)], links=links)
    r = _Result()
    _ve.validate_ew_not_oversubscribed(wb, {}, r)
    assert r.warnings == []


def test_oob_links_off_an_ew_leaf_are_neither_up_nor_down():
    """A leaf's OOB management port is not E/W capacity in either direction. Counting
    it as a downlink would manufacture an oversubscription on a 1:1 fabric."""
    links = [(f"su-01-node-{i:02d}", "gl-plane1-01", "GPU Network") for i in range(1, 5)]
    links += [("gl-plane1-01", "gs-plane1-01", "ISL") for _ in range(4)]
    links += [("gl-plane1-01", "oob-switch-01", "OOB / IPMI")]
    wb = _workbook(vlans=[_GPU_VLAN], profiles=[("GPU Network", "400G", 901),
                                              ("ISL", "400G"),
                                              ("OOB / IPMI", "1G")], links=links)
    r = _Result()
    _ve.validate_ew_not_oversubscribed(wb, {}, r)
    assert r.warnings == []


def test_missing_sheets_do_not_crash():
    wb = openpyxl.Workbook()
    r = _Result()
    _ve.validate_ew_not_oversubscribed(wb, {}, r)
    _ve.validate_ew_uses_private_ips(wb, {}, r)
    assert r.warnings == [] and r.errors == []


# --- no regression on the shipped defaults --------------------------------------------

REPO = Path(__file__).resolve().parent.parent
DEFAULTS = sorted(REPO.glob("input/*/default/*.xlsx"))


@pytest.mark.parametrize("path", DEFAULTS, ids=lambda p: p.parent.parent.name)
def test_shipped_defaults_satisfy_every_criterion(path):
    """Our own reference designs are 1:1 non-blocking on RFC1918 E/W subnets and meet
    the slide-10 floor. If one of these ever warns, the design changed or the check is
    wrong — either way it is not something to discover in an OEM's report."""
    arch = path.parent.parent.name
    wb = openpyxl.load_workbook(path, data_only=True)
    r = _Result()
    _ve.validate_ew_not_oversubscribed(wb, {}, r)
    _ve.validate_ew_uses_private_ips(wb, {}, r)
    _ve.validate_ew_bandwidth_per_gpu(wb, {"architecture": arch}, r)
    assert r.warnings == [], f"{path.name}: {r.warnings}"


@pytest.mark.parametrize("path", DEFAULTS, ids=lambda p: p.parent.parent.name)
@needs_arch_models
def test_ew_fabric_resolves_on_every_arch(path):
    """The non-vacuity guard. The first cut of `_ew_leaf_bandwidth` keyed on switch
    ROLE and read zero E/W links on the four collapsed-core archs, where the GPU nodes
    cable to `core` — so the check above passed on those four without measuring
    anything. A silent zero is indistinguishable from a clean design, so assert the
    fabric was actually found.

    The expected value is the `B` field of the arch name (`{CPUs}-{GPUs}-{NICs}-{B}`),
    which is per-GPU E/W bandwidth in Gbps by definition. All seven shipped archs land
    exactly on it, so this doubles as a cross-check that the walk measures what its
    name claims.
    """
    arch = path.parent.parent.name
    expected = int(arch.split("-")[3])
    wb = openpyxl.load_workbook(path, data_only=True)
    down, _up, unsized = _ve._ew_leaf_bandwidth(wb)
    assert down, f"{arch}: no E/W fabric resolved — the check is disarmed here"
    assert not unsized, f"{arch}: an E/W profile has an unreadable Speed cell"

    su = _ve._su_count(wb)
    model = _load_model(arch)
    gpus = su * int(model["gpus_per_su"])
    assert sum(down.values()) / gpus == expected, (
        f"{arch}: measured {sum(down.values()) / gpus} Gb/GPU, nameplate says {expected}"
    )


def _load_model(arch):
    sys.path.insert(0, str(REPO.parent / "data-models"))
    from models import load_arch_model
    return load_arch_model(arch)


# --- slide-10 E/W per-GPU floor -------------------------------------------------------

def test_below_the_floor_warns(monkeypatch):
    """32 GPUs at 100 Gb/GPU is half the floor."""
    links = [(f"su-01-node-{i:02d}", "gl-plane1-01", "GPU Network") for i in range(1, 9)]
    wb = _workbook(
        vlans=[[900, "gpu", "GPU E/W", "192.168.1.0/24", "192.168.1.1", "GPU", 4900,
                "No"]],
        profiles=[("GPU Network", "400G", 900)], links=links)
    monkeypatch.setattr(_ve, "_su_count", lambda _wb: 1)
    monkeypatch.setattr(_ve, "load_arch_model", lambda _a: {"gpus_per_su": 32})
    r = _Result()
    _ve.validate_ew_bandwidth_per_gpu(wb, {"architecture": "2-8-9-800"}, r)
    assert len(r.warnings) == 1
    assert "100.00 Gb/GPU" in r.warnings[0][1]


def test_exactly_on_the_floor_does_not_warn(monkeypatch):
    """2-4-3-200 and 2-8-5-200 sit exactly at 200 Gb/GPU. An epsilon in the wrong
    direction fails two of our own shipped reference designs."""
    links = [(f"su-01-node-{i:02d}", "gl-plane1-01", "GPU Network") for i in range(1, 17)]
    wb = _workbook(
        vlans=[[900, "gpu", "GPU E/W", "192.168.1.0/24", "192.168.1.1", "GPU", 4900,
                "No"]],
        profiles=[("GPU Network", "400G", 900)], links=links)
    monkeypatch.setattr(_ve, "_su_count", lambda _wb: 1)
    monkeypatch.setattr(_ve, "load_arch_model", lambda _a: {"gpus_per_su": 32})
    r = _Result()
    _ve.validate_ew_bandwidth_per_gpu(wb, {"architecture": "2-8-9-800"}, r)
    assert r.warnings == []


def test_cpu_inband_on_a_collapsed_core_is_not_counted_as_ew():
    """The defect that forced the VRF-based rewrite: on a collapsed-core arch the same
    `core` switch carries the GPU rails AND the CPU/In-Band N/S links. Counting by
    switch role folds N/S bandwidth into the E/W total."""
    links = [(f"su-01-node-{i:02d}", "core-01", "GPU Rail 1") for i in range(1, 5)]
    links += [(f"su-01-node-{i:02d}", "core-01", "CPU/In-Band Network")
              for i in range(1, 5)]
    wb = _workbook(
        vlans=[[901, "gpu_rail1", "GPU Rail 1 E/W", "192.168.1.0/24", "192.168.1.1",
                "GPU", 4901, "No"],
               [300, "cpu", "CPU North-South", "172.16.178.0/24", "172.16.178.1",
                "INBAND", 4300, "No"]],
        profiles=[("GPU Rail 1", "400G", 901), ("CPU/In-Band Network", "400G", 300)],
        links=links)
    down, _up, _u = _ve._ew_leaf_bandwidth(wb)
    assert down == {"core-01": 1600}, "N/S bandwidth leaked into the E/W total"


def test_renamed_gpu_profile_still_resolves():
    """Keying on the profile NAME would break on the first submission that renames its
    rails — the real one already does ('GPU Rail N Plane M' -> 'GPU Network')."""
    wb = _workbook(
        vlans=[[900, "whatever", "fabric", "192.168.1.0/24", "192.168.1.1", "GPU",
                4900, "No"]],
        profiles=[("Partner Named This Something Else", "400G", 900)],
        links=[("su-01-node-01", "core-01", "Partner Named This Something Else")])
    assert _ve._ew_leaf_bandwidth(wb)[0] == {"core-01": 400}


def test_slide10_floor_is_silent_when_a_profile_is_unsized(monkeypatch):
    """The `unsized` guard on the slide-10 check had no test — a mutation removing it
    survived. Without it, an unreadable Speed cell counts as 0 Gbps and manufactures a
    below-floor finding out of a cell the workbook simply did not fill in.
    """
    links = [(f"su-01-node-{i:02d}", "gl-plane1-01", "GPU Network") for i in range(1, 17)]
    wb = _workbook(
        vlans=[[900, "gpu", "GPU E/W", "192.168.1.0/24", "192.168.1.1", "GPU", 4900,
                "No"]],
        profiles=[("GPU Network", None, 900)], links=links)
    monkeypatch.setattr(_ve, "_su_count", lambda _wb: 1)
    monkeypatch.setattr(_ve, "load_arch_model", lambda _a: {"gpus_per_su": 32})
    r = _Result()
    _ve.validate_ew_bandwidth_per_gpu(wb, {"architecture": "2-8-9-800"}, r)
    assert r.warnings == [], "an unreadable Speed cell must not read as 0 Gbps"


# --- deployed per-link speed from Breakout/Lanes (ADR-0040) -------------------------

def _profiles_sheet(rows):
    """A workbook with just the Port Profiles block. `rows` = (name, breakout, lanes)."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    vp = wb.create_sheet("VLANs & Profiles")
    vp.append(["Port Profiles"])
    vp.append(["Profile", "Port Mode", "Native/Access VLAN", "Allowed VLANs",
               "Untagged VLAN", "VRF", "LACP Bypass", "Speed", "Breakout", "Lanes"])
    for name, breakout, lanes in rows:
        # Speed cell deliberately says 100G (the per-lane transcription ADR-0040 warns
        # about); the link rate must come from Lanes, not this cell.
        vp.append([name, "L3", None, None, None, "EXIT", "No", "100G", breakout, lanes])
    return wb


def test_deployed_link_gbps_is_lanes_times_100_not_the_speed_cell():
    """ADR-0040: a broken-out sub-port runs at lanes x 100G. The capacity check must
    credit the OEM's DEPLOYED breakout, so an Edge Uplink at breakout-8/lanes-1 is 100G
    (not the RA's 200G), which is what hid half of the exit shortfall."""
    wb = _profiles_sheet([
        ("Edge Uplink", 8, 1),      # 1 x 100G = 100G  (the observed deviation)
        ("Storage Uplink", 4, 2),   # 2 x 100G = 200G
        ("ISL", 2, 4),              # 4 x 100G = 400G
    ])
    got = _ve._deployed_link_gbps(wb)
    assert got == {"Edge Uplink": 100, "Storage Uplink": 200, "ISL": 400}


def test_deployed_link_gbps_skips_blank_lanes_so_caller_falls_back_to_model():
    """A profile with no Lanes must be absent, so the floor check falls back to the
    model profile speed rather than crediting 0."""
    wb = _profiles_sheet([("Storage Uplink", 4, 2)])
    wb["VLANs & Profiles"].append(["Edge Uplink", "L3", None, None, None, "EXIT",
                                   "No", "200G", None, None])  # no Lanes
    got = _ve._deployed_link_gbps(wb)
    assert "Edge Uplink" not in got
    assert got.get("Storage Uplink") == 200
