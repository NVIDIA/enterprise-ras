# SPDX-FileCopyrightText: Copyright (c) 2026 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: MIT
"""Operator-pinned switch eth0 addresses on the air-mgmt plane.

`air_mgmt_subnet` (Air_Only, "Air Management Subnet") is operator-selectable and
already drove the generator (switch eth0 walk, ztp_interfaces, ztp_allow_subnets)
-- but air-deploy.py hardcoded `172.20.0.x` for every address it provisioned. A
changed subnet therefore produced a SPLIT BRAIN: DHCP served on one plane and
listened for on another, so no switch ever got an address. air-deploy.py now
resolves the plane once from the generated inventory.

On top of that, a brownfield operator knows the real management addresses their
switches already answer on and wants the sim to reproduce them. A switch whose
Nodes-tab mgmt IP is INSIDE air_mgmt_subnet is now honoured verbatim; unpinned
switches auto-assign around the pins. Three gates keep that safe:

  1. outside BOTH the OOB VLAN subnets and air_mgmt_subnet -> unreachable, error
     (this is ADR-0041 / ERA-61 preserved -- the question is "reachable?", not
     "on the OOB plane specifically?")
  2. squatting an octet air-deploy provisions for infra (.1/.77/.78/.254) -> error
  3. two switches pinned to the same address -> error

A non-switch host inside air_mgmt_subnet is still an intruder, unchanged.
"""
import importlib.util
import shutil
import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
VALIDATOR = REPO / "scripts" / "validate_excel.py"
BASE = REPO / "input" / "RH-2-8-5-200.xlsx"

AIR_SUBNET = "10.78.255.0/24"

pytestmark = pytest.mark.skipif(not BASE.exists(), reason="brownfield workbook absent")


def _switch_names():
    """Read the two core and two OOB switch names out of the workbook.

    Previously these were four hardcoded constants carrying a customer's
    site-prefixed hostnames. Deriving them keeps the identifier out of shipped
    source, and makes the test independent of whose workbook it runs against —
    which matters because this module skips wherever that workbook is absent
    (CI included), so a hardcoded name is only ever checked on one machine.
    """
    wb = openpyxl.load_workbook(BASE, read_only=True, data_only=True)
    try:
        rows = wb["Nodes"].iter_rows(min_row=2, values_only=True)
        cores, oobs = [], []
        for r in rows:
            name = str((r[1] if len(r) > 1 else "") or "").strip()
            if name.startswith("oob-switch-"):
                oobs.append(name)
            elif name.startswith("core-"):
                cores.append(name)
    finally:
        wb.close()
    assert len(cores) >= 2 and len(oobs) >= 2, (
        f"workbook must declare >=2 core and >=2 oob switches; got {cores}, {oobs}")
    return cores[0], cores[1], oobs[0], oobs[1]


def _a_server_name():
    """Any non-switch node — used to prove the gate fires on servers too."""
    wb = openpyxl.load_workbook(BASE, read_only=True, data_only=True)
    try:
        for r in wb["Nodes"].iter_rows(min_row=2, values_only=True):
            name = str((r[1] if len(r) > 1 else "") or "").strip()
            if name and not name.startswith(("core-", "oob-switch-", "csl-", "gsl-",
                                             "cust-net-edge", "external-", "utility")):
                return name
    finally:
        wb.close()
    raise AssertionError("workbook declares no non-switch node")


CORE1, CORE2, OOB1, OOB2 = _switch_names() if BASE.exists() else ("", "", "", "")
SERVER = _a_server_name() if BASE.exists() else ""


def _air_deploy():
    """Import air-deploy.py (hyphenated, so not a normal import)."""
    sys.path.insert(0, str(REPO / "scripts"))
    sys.path.insert(0, str(REPO / "scripts" / "airlib"))
    spec = importlib.util.spec_from_file_location("_ad", REPO / "scripts" / "air-deploy.py")
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


@pytest.fixture
def workbook(tmp_path):
    """Copy the brownfield workbook, set the air subnet, apply mgmt-IP edits."""
    def _go(mutations=None, air_subnet=AIR_SUBNET):
        dst = tmp_path / "2-8-5-200.xlsx"
        shutil.copy2(BASE, dst)
        wb = openpyxl.load_workbook(dst)
        ws = wb["Air_Only"]
        for r in range(1, ws.max_row + 1):
            if str(ws.cell(r, 1).value or "").strip() == "Air Management Subnet":
                ws.cell(r, 2).value = air_subnet
        nodes = wb["Nodes"]
        hdr = [str(c.value or "").strip() for c in nodes[1]]
        name_c = hdr.index("Name") + 1
        ip_c = hdr.index("Mgmt IP Address") + 1
        for r in range(2, nodes.max_row + 1):
            nm = str(nodes.cell(r, name_c).value or "").strip()
            if nm in (mutations or {}):
                nodes.cell(r, ip_c).value = (mutations or {})[nm]
        wb.save(dst)
        return dst
    return _go


def _errors(xlsx):
    proc = subprocess.run([sys.executable, str(VALIDATOR), str(xlsx)],
                          capture_output=True, text=True, cwd=str(REPO))
    return [ln.strip() for ln in proc.stdout.splitlines() if "❌" in ln]


def _ansible_hosts(xlsx, tmp_path):
    """Generate inventory for `xlsx` and return {switch: ansible_host}."""
    site = "pytest" + tmp_path.name[-8:]
    inp = REPO / "input" / "2-8-5-200" / site
    out = REPO / "output" / "2-8-5-200" / site
    inp.mkdir(parents=True, exist_ok=True)
    shutil.copy2(xlsx, inp / "2-8-5-200.xlsx")
    try:
        subprocess.run([sys.executable, "scripts/excel_parser.py",
                        "--arch", "2-8-5-200", "--site", site],
                       capture_output=True, text=True, cwd=str(REPO), check=True)
        hosts = {}
        for hv in (out / "inventory" / "host_vars").glob("*.yml"):
            for line in hv.read_text().splitlines():
                if line.startswith("ansible_host:"):
                    hosts[hv.stem] = line.split(":", 1)[1].strip()
        return hosts
    finally:
        shutil.rmtree(inp, ignore_errors=True)
        shutil.rmtree(out, ignore_errors=True)


# --- air-deploy.py follows air_mgmt_subnet --------------------------------

def test_air_deploy_defaults_to_172_20_0():
    """Unset/blank keeps every existing deployment byte-identical."""
    ad = _air_deploy()
    assert ad.air_mgmt_network() == "172.20.0.0/24"
    assert ad.air_mgmt_cidr(ad.AIR_MGMT_SVI_OCTET) == "172.20.0.254/24"
    assert ad.air_mgmt_cidr(ad.EXTERNAL_DHCP_OCTET) == "172.20.0.77/24"


def test_air_deploy_follows_a_changed_subnet():
    """The split brain: these used to stay 172.20.0.x no matter the setting."""
    ad = _air_deploy()
    ad.set_air_mgmt_subnet(AIR_SUBNET)
    assert ad.air_mgmt_network() == AIR_SUBNET
    assert ad.air_mgmt_cidr(ad.AIR_MGMT_SVI_OCTET) == "10.78.255.254/24"
    assert ad.air_mgmt_cidr(ad.EXTERNAL_DHCP_OCTET) == "10.78.255.77/24"
    assert ad.air_mgmt_cidr(ad.UTILITY_OCTET) == "10.78.255.78/24"
    assert ad.air_mgmt_ip(ad.AIR_MGMT_SVI_OCTET) == "10.78.255.254"


def test_air_deploy_malformed_subnet_keeps_current_plane():
    ad = _air_deploy()
    ad.set_air_mgmt_subnet("not-a-cidr")
    assert ad.air_mgmt_network() == "172.20.0.0/24"


def test_air_mgmt_owner_map_matches_reserved_set():
    """The two must not drift -- a reserved octet with no owner label would
    produce an error message naming nothing."""
    sys.path.insert(0, str(REPO / "scripts"))
    import oob_reserved as o
    assert set(o.AIR_MGMT_RESERVED_OWNERS) == o.AIR_MGMT_RESERVED_OCTETS


# --- pins are honoured ----------------------------------------------------

def test_real_brownfield_addresses_validate_clean(workbook):
    """The shipping case: customer's live .21-.24 on their own mgmt subnet."""
    assert _errors(workbook()) == []


def test_pins_reach_the_inventory(workbook, tmp_path):
    hosts = _ansible_hosts(workbook(), tmp_path)
    assert hosts[CORE1] == "10.78.255.21"
    assert hosts[CORE2] == "10.78.255.22"
    assert hosts[OOB1] == "10.78.255.23"
    assert hosts[OOB2] == "10.78.255.24"


def test_unpinned_switches_auto_assign_around_pins(workbook, tmp_path):
    """Partial fill -- brownfield operators often know only some addresses."""
    hosts = _ansible_hosts(workbook({OOB1: None, OOB2: None}), tmp_path)
    assert hosts[CORE1] == "10.78.255.21"   # pinned, honoured
    assert hosts[CORE2] == "10.78.255.22"
    assert hosts[OOB1] == "10.78.255.201"   # blank, auto-assigned
    assert hosts[OOB2] == "10.78.255.202"


def test_oob_svi_plane_is_untouched_by_pinning(workbook, tmp_path):
    """Pinning eth0 must not disturb the BMC plane the OOB switches serve."""
    xlsx = workbook()
    inp = REPO / "input" / "2-8-5-200" / "pytestsvi"
    out = REPO / "output" / "2-8-5-200" / "pytestsvi"
    inp.mkdir(parents=True, exist_ok=True)
    shutil.copy2(xlsx, inp / "2-8-5-200.xlsx")
    try:
        subprocess.run([sys.executable, "scripts/excel_parser.py",
                        "--arch", "2-8-5-200", "--site", "pytestsvi"],
                       capture_output=True, text=True, cwd=str(REPO), check=True)
        hv = (out / "inventory" / "host_vars" / f"{OOB1}.yml").read_text()
        assert "svi_ip: 10.78.220.130/25" in hv
        assert "default_gateway: 10.78.220.129" in hv
    finally:
        shutil.rmtree(inp, ignore_errors=True)
        shutil.rmtree(out, ignore_errors=True)


# --- the three gates ------------------------------------------------------

def test_pin_outside_both_planes_errors(workbook):
    """ADR-0041 / ERA-61 preserved: unreachable is still a hard failure."""
    errs = _errors(workbook({CORE1: "10.99.0.5"}))
    assert any("10.99.0.5" in e and "not within any OOB VLAN subnet" in e for e in errs), errs


def test_error_message_names_both_planes_for_a_switch(workbook):
    """An operator told only about the OOB subnet would 'fix' it wrongly."""
    errs = _errors(workbook({CORE1: "10.99.0.5"}))
    assert any("air-mgmt subnet 10.78.255.0/24" in e for e in errs), errs


@pytest.mark.parametrize("octet,owner", [
    (78, "utility"),
    (77, "external-dhcp"),
    (254, "cust-net-edge-01 bridge SVI"),
    # .1 is reserved on this plane but OWNED BY NOTHING -- the gateway is the
    # .254 SVI. It used to be labelled "air-mgmt gateway", which misdirected an
    # operator at the one moment the string is shown.
    (1, "no owner on air-mgmt"),
    # ext-storage eth0 lives here too; omitting it from the reserved set is
    # what gave gs-plane2-08 and ext-storage-01 the same address at SU32.
    (79, "ext-storage-01"),
    (80, "ext-storage-02"),
])
def test_pin_squatting_air_infra_errors(workbook, octet, owner):
    errs = _errors(workbook({CORE1: f"10.78.255.{octet}"}))
    assert any(f"10.78.255.{octet}" in e and owner in e for e in errs), errs


def test_two_switches_pinned_to_one_address_errors(workbook):
    errs = _errors(workbook({CORE1: "10.78.255.30", CORE2: "10.78.255.30"}))
    assert any("pinned by multiple switches" in e for e in errs), errs


def test_non_switch_in_air_mgmt_still_errors(workbook):
    """Servers have no business on the air-mgmt plane -- unchanged behaviour."""
    errs = _errors(workbook({SERVER: "10.78.255.60"}))
    assert any("is inside the air-mgmt subnet" in e for e in errs), errs
