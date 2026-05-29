# SPDX-FileCopyrightText: Copyright (c) 2026 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: MIT
"""Regression coverage for validate_wire_map source-port dedup logic.

The per-side dedup key `(sys_name, nic_port, port_side)` lets dual-plane
breakouts coexist (same NIC, side A + side B → two buckets, no warning).
A side-effect: rows with mixed Port Side (A) values on the same physical
port (e.g. one row with side="NA", another with side="A") hash to
different buckets and used to slip through silently.

This test locks in the stricter cross-check that catches that case
without false-positiving the legitimate breakout pattern.
"""
import sys
from pathlib import Path

import openpyxl
import pytest

SCRIPTS_DIR = Path(__file__).parent.parent / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from validate_excel import validate_wire_map, ValidationResult


WM_HEADERS = [
    "Display in Air", "System Name (A)", "Port (A)", "Port Side (A)",
    "System Name (B)", "Port (B)", "Port Side (B)", "Cable Split",
    "Network Profile",
]


def _wm(rows):
    """Build an openpyxl Wire Map sheet from a list of row tuples."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Wire Map"
    for c, h in enumerate(WM_HEADERS, start=1):
        ws.cell(row=1, column=c, value=h)
    for r, row in enumerate(rows, start=2):
        for c, v in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=v)
    return ws


def _errors(result):
    return [m for m in result.errors if "Physical port reused" in m]


def test_mixed_na_and_side_on_same_port_is_flagged():
    # The exact shape of the bug that hit input/s2-2-8-9-800.xlsx:
    # default row had side="NA", added row had side="A", both
    # cabling the same csl-01:swp63s0 to different peers.
    ws = _wm([
        ("Yes", "csl-01", "swp63s0", "NA", "ext-storage-01", "eth1", "NA", "NA", "Storage Uplink"),
        ("Yes", "csl-01", "swp63s0", "A",  "cust-net-edge-01", "swp52", "NA", 1, "Edge Uplink"),
    ])
    result = ValidationResult()
    validate_wire_map(ws, result)
    flagged = _errors(result)
    assert len(flagged) == 1, f"expected exactly one Physical-port-reused error, got: {result.errors}"
    assert "csl-01" in flagged[0] and "swp63s0" in flagged[0]


def test_two_rows_same_na_side_not_flagged():
    # Pattern (a): dual-homed OOB. server eth0 cables to two OOB
    # switches with empty/NA Port Side (A) on both. Existing warning
    # loop handles this; the new check should stay quiet.
    ws = _wm([
        ("Yes", "support-01", "eth0", "NA", "oob-switch-01", "swp10", "NA", "NA", "OOB / IPMI"),
        ("Yes", "support-01", "eth0", "NA", "oob-switch-02", "swp10", "NA", "NA", "OOB / IPMI"),
    ])
    result = ValidationResult()
    validate_wire_map(ws, result)
    assert _errors(result) == [], f"dual-homed OOB should not error: {result.errors}"


def test_dual_plane_breakout_not_flagged():
    # Pattern (b): same NIC, distinct non-NA sides per row → legit
    # dual-plane breakout. Must not error.
    ws = _wm([
        ("Yes", "gpu-01", "eth1", "A", "gsl-plane1-01", "swp10s0", "A", 1, "GPU Network"),
        ("Yes", "gpu-01", "eth1", "B", "gsl-plane2-01", "swp10s0", "B", 2, "GPU Network"),
    ])
    result = ValidationResult()
    validate_wire_map(ws, result)
    assert _errors(result) == [], f"dual-plane breakout should not error: {result.errors}"


def test_same_non_na_side_repeated_is_flagged():
    # Two rows on (gpu-01, eth1) with side="A" both times — physically
    # impossible (one port-side can't go to two switches).
    ws = _wm([
        ("Yes", "gpu-01", "eth1", "A", "gsl-plane1-01", "swp10s0", "A", 1, "GPU Network"),
        ("Yes", "gpu-01", "eth1", "A", "gsl-plane2-01", "swp11s0", "A", 1, "GPU Network"),
    ])
    result = ValidationResult()
    validate_wire_map(ws, result)
    # The existing per-side warning loop fires (warning), and the new
    # cross-check stays quiet because both sides match → it's pattern
    # (a) by our key. That's acceptable — we don't want both checks
    # double-reporting. The legacy warning still surfaces it.
    flagged = _errors(result)
    # New check shouldn't fire (same single side label → caught by
    # pre-existing warning, no need to duplicate as error).
    assert flagged == []
    # Pre-existing warning loop should still surface it.
    src_warns = [m for m in result.warnings
                 if "Source port used multiple times" in m]
    assert len(src_warns) == 1
