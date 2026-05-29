# SPDX-FileCopyrightText: Copyright (c) 2026 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: MIT
"""
Tests for the header-name-based Wire Map column lookup.

Locks in:
  - Required-column enforcement (missing required header → ValueError)
  - Alias resolution (old `System Role` / new `Function (A)` both work)
  - Column reordering doesn't break the parser
  - Air_Only metadata-only sheet is tolerated (no required columns)
"""
import sys
from pathlib import Path

import openpyxl
import pytest

SCRIPTS_DIR = Path(__file__).parent.parent / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from excel_parser import (
    build_wiremap_column_map,
    _build_wiremap_row_list,
    _WM_HEADER_ALIASES,
    _WM_REQUIRED,
)


def _wb_with_wm(headers, data_rows):
    """Build a workbook with a Wire Map sheet using the given headers + rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Wire Map'
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    for r, row in enumerate(data_rows, start=2):
        for c, v in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=v)
    return wb


# ---------------------------------------------------------------------------
# Required-column enforcement
# ---------------------------------------------------------------------------

def test_missing_required_column_raises():
    """If a required column header is absent, build_wiremap_column_map
    raises ValueError mentioning every missing column."""
    headers = ['Display in Air', 'Function (A)', 'System Name (A)']  # missing 4 required
    wb = _wb_with_wm(headers, [])
    with pytest.raises(ValueError) as exc:
        build_wiremap_column_map(wb['Wire Map'], sheet_kind='wiremap')
    msg = str(exc.value)
    # Each missing column should appear in the error
    for col in ('nic_port', 'network_profile', 'switch_name', 'switch_port'):
        assert col in msg


def test_all_required_present_succeeds():
    """Required-column set met → returns a map."""
    headers = [
        'Display in Air', 'Function (A)', 'System Name (A)', 'Port (A)',
        'Network Profile', 'Function (B)', 'System Name (B)', 'Port (B)',
    ]
    wb = _wb_with_wm(headers, [])
    col_map = build_wiremap_column_map(wb['Wire Map'], sheet_kind='wiremap')
    for k in _WM_REQUIRED:
        assert k in col_map, f"required column {k!r} not in col_map"


# ---------------------------------------------------------------------------
# Alias resolution
# ---------------------------------------------------------------------------

def test_legacy_headers_still_work_via_aliases():
    """Old `System Role` / `Switch Role` / `Switch Name` etc. still
    resolve to the right logical columns."""
    headers = [
        'Display in Air', 'System Role', 'System Name', 'NIC/Port',
        'Port Side (A)', 'Speed', 'Network Profile', 'Mode',
        'Native/Access VLAN', 'Allowed VLANs',
        'Switch Role', 'Switch Name', 'Switch Port',
    ]
    wb = _wb_with_wm(headers, [])
    col_map = build_wiremap_column_map(wb['Wire Map'], sheet_kind='wiremap')
    assert col_map['system_role'] == 2
    assert col_map['system_name'] == 3
    assert col_map['nic_port'] == 4
    assert col_map['network_profile'] == 7
    assert col_map['switch_role'] == 11
    assert col_map['switch_name'] == 12
    assert col_map['switch_port'] == 13


def test_new_ab_headers_resolve():
    """`Function (A)` / `System Name (B)` etc. resolve to the same
    logical columns as the legacy names."""
    headers = [
        'Display in Air', 'Function (A)', 'System Name (A)', 'Port (A)',
        'Port Side (A)', 'Speed', 'Network Profile', 'Mode',
        'Native/Access VLAN', 'Allowed VLANs',
        'Function (B)', 'System Name (B)', 'Port (B)',
    ]
    wb = _wb_with_wm(headers, [])
    col_map = build_wiremap_column_map(wb['Wire Map'], sheet_kind='wiremap')
    assert col_map['system_role'] == 2
    assert col_map['switch_role'] == 11
    assert col_map['switch_name'] == 12


# ---------------------------------------------------------------------------
# Column reordering doesn't break parsing
# ---------------------------------------------------------------------------

def test_reordered_columns_still_parse_correctly():
    """Operator moves columns around → parser still reads the right cells."""
    # Required columns in arbitrary order, plus one extra annotation column.
    headers = [
        'Port (B)', 'System Name (B)', 'Function (B)',  # B side first
        'Network Profile',
        'Port (A)', 'System Name (A)', 'Function (A)',  # A side later
        'Display in Air',
        'Annotation',  # extra non-mapped column
    ]
    data = [
        ['swp5', 'core-01', 'core', 'CPU/In-Band Network',
         'B3220 Port 1', 'gpu-01', 'gpu', 'Yes', 'note'],
    ]
    wb = _wb_with_wm(headers, data)
    rows = _build_wiremap_row_list(wb['Wire Map'], nodes_function_map={})
    assert len(rows) == 1
    r = rows[0]
    assert r['system_name'] == 'gpu-01'
    assert r['nic_port'] == 'B3220 Port 1'
    assert r['net_profile'] == 'CPU/In-Band Network'
    assert r['switch_name'] == 'core-01'
    assert r['switch_port'] == 'swp5'
    assert r['display_in_air'] is True


# ---------------------------------------------------------------------------
# Function columns physically deleted — parser still works via Nodes-tab
# fallback (MR !28) + the column map gracefully handles their absence.
# ---------------------------------------------------------------------------

def test_both_function_columns_deleted_still_parses():
    """Operator deletes both Function (A) AND Function (B) from Wire Map.
    Other columns shift left. Parser must still extract the connection
    via System Name (A)/(B) and Port (A)/(B); category resolves via
    Nodes-tab lookup."""
    # Headers WITHOUT Function (A) or Function (B). Column 1 = Display,
    # 2 = System Name (A), 3 = Port (A), 4 = Network Profile,
    # 5 = System Name (B), 6 = Port (B). All required columns present;
    # role columns absent.
    headers = [
        'Display in Air', 'System Name (A)', 'Port (A)',
        'Network Profile', 'System Name (B)', 'Port (B)',
    ]
    data = [
        ['Yes', 'gpu-01', 'B3140 Slot 1 Port 1', 'GPU Network',
         'core-01', 'swp6s0'],
        ['Yes', 'gpu-01', 'B3140 Slot 2 Port 1', 'GPU Network',
         'core-02', 'swp6s0'],
    ]
    wb = _wb_with_wm(headers, data)
    col_map = build_wiremap_column_map(wb['Wire Map'], sheet_kind='wiremap')
    # The optional Function columns are absent — col_map shouldn't have them.
    assert 'system_role' not in col_map
    assert 'switch_role' not in col_map
    # Required columns all resolved.
    for k in _WM_REQUIRED:
        assert k in col_map

    # _build_wiremap_row_list with a Nodes-map fills in the categories.
    nodes_map = {'gpu-01': 'gpu', 'core-01': 'core', 'core-02': 'core'}
    rows = _build_wiremap_row_list(wb['Wire Map'], nodes_function_map=nodes_map)
    assert len(rows) == 2
    assert rows[0]['system_role'] == 'gpu'  # filled from nodes_map
    assert rows[0]['system_role_raw'] == ''  # cell was absent
    assert rows[0]['switch_role'] == 'core'
    assert rows[0]['switch_role_raw'] == ''


# ---------------------------------------------------------------------------
# Air_Only metadata-only sheet (no required cols) — must not crash
# ---------------------------------------------------------------------------

def test_air_only_metadata_sheet_skips_gracefully():
    """Air_Only used purely for the version-image map should not raise
    when validate-excel inspects it for connection rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Air_Only'
    ws.cell(row=1, column=1, value='Friendly Version')
    ws.cell(row=1, column=2, value='Air Image')
    ws.cell(row=2, column=1, value='5.16.1')
    ws.cell(row=2, column=2, value='cumulus-vx-5.16.1')
    # No wire-map-like columns present.
    with pytest.raises(ValueError):
        # build_wiremap_column_map raises — callers must catch and treat
        # as "no connection rows in this sheet."
        build_wiremap_column_map(ws, sheet_kind='air_only')
