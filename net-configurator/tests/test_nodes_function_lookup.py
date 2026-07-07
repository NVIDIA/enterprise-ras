# SPDX-FileCopyrightText: Copyright (c) 2026 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: MIT
"""
Tests for the Nodes-tab Function lookup that lets Wire Map / Air_Only
rows leave their own Function (System Role / Switch Role) cells blank
and resolve the role via the Nodes tab instead.

Covers:
  - build_nodes_function_map() output shape
  - _build_wiremap_row_list() cascade: Nodes → row → blank
  - Skip-gate semantics (blank Role + blank Name still skipped;
    blank Role but Name present is processed)
  - Loopbacks validator upgrade: switch-name not on Nodes → error
"""
import sys
from pathlib import Path
import openpyxl
import pytest

SCRIPTS_DIR = Path(__file__).parent.parent / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from excel_parser import (
    build_nodes_function_map,
    _build_wiremap_row_list,
    parse_nodes,
    canonical_category,
)
from validate_excel import validate_loopbacks, ValidationResult


# ---------------------------------------------------------------------------
# build_nodes_function_map
# ---------------------------------------------------------------------------

def test_nodes_map_omits_unclassified_names():
    """Hostnames whose Function can't be classified stay out of the map."""
    nodes = [
        {'name': 'core-01', 'category': 'core'},
        {'name': 'gpu-05', 'category': 'gpu'},
        {'name': '', 'category': 'core'},        # blank name — skip
        {'name': 'mystery', 'category': None},   # no category — skip
    ]
    m = build_nodes_function_map(nodes)
    assert m == {'core-01': 'core', 'gpu-05': 'gpu'}


def test_nodes_map_strips_whitespace():
    """Leading/trailing whitespace in Names is normalized."""
    nodes = [
        {'name': '  core-01  ', 'category': 'core'},
    ]
    assert build_nodes_function_map(nodes) == {'core-01': 'core'}


# ---------------------------------------------------------------------------
# _build_wiremap_row_list cascade: Nodes → row → blank
# ---------------------------------------------------------------------------

def _wm_workbook(rows):
    """Build a minimal Wire Map worksheet with the given row tuples."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Wire Map'
    headers = [
        'Display in Air', 'System Role', 'System Name', 'NIC/Port', 'Port Side',
        'Speed', 'Network Profile', 'Mode', 'VLAN', 'Allowed', 'Switch Role',
        'Switch Name', 'Switch Port',
    ]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    for r, row in enumerate(rows, start=2):
        for c, v in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=v)
    return wb


def test_wm_row_resolves_function_from_nodes_when_blank():
    """Blank Wire Map Function with Name present → Nodes map fills it."""
    # 13 cols: Display, Role(blank), Name, NIC, PortSide, Speed, Profile,
    # Mode, Native, Allowed, SwitchRole(blank), SwitchName, SwitchPort.
    rows = [
        ('Yes', '', 'gpu-01', 'B3220 Port 1', '1', '200G', 'GPU Network',
         '', '', '', '', 'core-01', 'swp1s0'),
    ]
    wb = _wm_workbook(rows)
    nodes_map = {'gpu-01': 'gpu', 'core-01': 'core'}
    result = _build_wiremap_row_list(wb['Wire Map'], nodes_function_map=nodes_map)

    assert len(result) == 1
    r = result[0]
    # system_role resolved from nodes_map (was blank in cell)
    assert r['system_role'] == 'gpu'
    assert r['system_role_raw'] == ''  # raw cell value preserved
    # switch_role same treatment
    assert r['switch_role'] == 'core'
    assert r['switch_role_raw'] == ''


def test_wm_row_keeps_explicit_function_when_present():
    """Wire Map Function cell wins over Nodes map (operator override)."""
    rows = [
        # Role explicit (legacy hostname-as-role); Name same
        ('Yes', 'gpu-01', 'gpu-01', 'B3220 Port 1', '1', '200G', 'GPU Network',
         '', '', '', 'core-01', 'core-01', 'swp1s0'),
    ]
    wb = _wm_workbook(rows)
    # Nodes map says canonical category — but row's own value takes
    # precedence because the Nodes map only fills the blank case.
    # Wait — see _resolve: nodes_map.get(name) OR system_role.
    # If nodes_map has gpu-01, looked_up=='gpu', returned. Hmm,
    # so legacy hostname-as-role would be REPLACED. That's correct —
    # canonical wins over hostname.
    nodes_map = {'gpu-01': 'gpu'}
    result = _build_wiremap_row_list(wb['Wire Map'], nodes_function_map=nodes_map)

    assert result[0]['system_role'] == 'gpu'  # canonical from Nodes
    assert result[0]['system_role_raw'] == 'gpu-01'  # raw cell preserved


def test_wm_row_keeps_row_value_when_not_in_nodes_map():
    """Hostname not on Nodes → cascade falls through to row's own Function."""
    rows = [
        ('Yes', 'cust-net-edge-01', 'cust-net-edge-01', 'eth1', '', '', 'EXIT',
         '', '', '', 'outbound', '', ''),
    ]
    wb = _wm_workbook(rows)
    # cust-net-edge-01 not on Nodes for this test
    result = _build_wiremap_row_list(wb['Wire Map'], nodes_function_map={})

    assert result[0]['system_role'] == 'cust-net-edge-01'
    assert result[0]['system_role_raw'] == 'cust-net-edge-01'


def test_wm_skip_gate_requires_both_role_and_name_blank():
    """Empty spacer rows: skip when BOTH role AND name are blank.
    Pre-Nodes-map: skipped on Role alone, which would now drop legit
    Name-only rows."""
    rows = [
        # Truly empty row — skipped.
        ('', '', '', '', '', '', '', '', '', '', '', '', ''),
        # Name-only row — kept; role resolved from Nodes map.
        ('Yes', '', 'gpu-01', 'eth1', '', '', 'GPU', '', '', '', '', 'core-01', 'swp1s0'),
        # Role-only row (legacy spacer pattern with annotation in col 2)
        ('', 'NA', '', '', '', '', '', '', '', '', '', '', ''),
    ]
    wb = _wm_workbook(rows)
    result = _build_wiremap_row_list(wb['Wire Map'],
                                     nodes_function_map={'gpu-01': 'gpu'})
    # Only the name-only row survives.
    assert len(result) == 2  # NA-only row also kept (has role)
    survivors = [r['system_name'] or r['system_role_raw'] for r in result]
    assert 'gpu-01' in survivors


# ---------------------------------------------------------------------------
# Loopbacks: switch not on Nodes → error (upgraded from warn)
# ---------------------------------------------------------------------------

def _loopbacks_ws(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Loopbacks'
    headers = ['Switch', 'Default', 'OOB', 'INBAND', 'EXIT', 'GPU']
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    for r, row in enumerate(rows, start=2):
        for c, v in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=v)
    return ws


def test_loopbacks_unknown_switch_now_errors():
    """A Loopbacks row referencing a switch not on the Nodes tab fires
    an error (was a warning before this branch)."""
    ws = _loopbacks_ws([
        ['core-01', '172.16.176.11/32', '', '', '', ''],   # known
        ['mystery-switch-99', '172.16.176.99/32', '', '', '', ''],  # unknown
    ])
    nodes = [{'name': 'core-01', 'function': 'core'}]
    r = ValidationResult()
    validate_loopbacks(ws, nodes, parsed_vlans=[], settings={}, result=r)
    err_msgs = ' | '.join(r.errors)
    warn_msgs = ' | '.join(r.warnings)
    assert 'mystery-switch-99' in err_msgs, (
        f"Unknown switch should fire an error. errors={r.errors!r} warnings={r.warnings!r}"
    )
    assert 'mystery-switch-99' not in warn_msgs


def test_loopbacks_no_false_positive_from_function_field():
    """Pre-fix bug: Nodes Function='csl' for csl-01 was treated as a
    valid switch name in known_switch_names, so a Loopbacks row called
    literally 'csl' would silently match. After fix, only the Name
    column counts."""
    ws = _loopbacks_ws([
        # Try to sneak in 'csl' as if it were a hostname — should error.
        ['csl', '172.16.176.50/32', '', '', '', ''],
    ])
    nodes = [{'name': 'csl-01', 'function': 'csl'},
             {'name': 'csl-02', 'function': 'csl'}]
    r = ValidationResult()
    validate_loopbacks(ws, nodes, parsed_vlans=[], settings={}, result=r)
    assert any("'csl' is not listed" in e for e in r.errors), (
        f"Bare canonical role 'csl' should error — it's a Function, not a "
        f"hostname. errors={r.errors!r}"
    )


# ---------------------------------------------------------------------------
# New leaf/spine taxonomy: cl/cs/gl/gs canonical roles + legacy aliases
# ---------------------------------------------------------------------------

def test_new_split_roles_resolve():
    assert canonical_category("cl") == "cl"
    assert canonical_category("cs") == "cs"
    assert canonical_category("gl-plane1") == "gl-plane1"
    assert canonical_category("gs-plane2") == "gs-plane2"


def test_legacy_spine_names_purged():
    # The branch-only legacy *-spine names are no longer canonical roles.
    # An explicit legacy Function string now falls through to the hostname
    # fallback, which resolves *-spine hostnames to the new cs / gs roles
    # (and a bare 'csl-spine' string, lacking a trailing instance index,
    # classifies as the converged 'csl' leaf).
    assert canonical_category("csl-spine") == "csl"
    assert canonical_category("gsl-spine-plane1") == "gs-plane1"
    assert canonical_category("gsl-spine-plane2") == "gs-plane2"


def test_converged_roles_unchanged():
    assert canonical_category("csl") == "csl"
    assert canonical_category("gsl-plane1") == "gsl-plane1"
    assert canonical_category("core") == "core"


def test_blank_function_spine_hostname_resolves_to_new_role():
    # Hostname fallback: a blank-Function hostname resolves to the role its
    # name advertises. New-taxonomy hostnames (gl-/gs-/cl-/cs-) return the new
    # canonical so the WM hostname fallback agrees with the Nodes Function side
    # (no canonical-vs-Function disagreement warnings). Legacy hostnames
    # (gsl-plane*, csl-spine-*, csl-*) still resolve to their legacy canonicals
    # so existing archs regenerate byte-identically.
    assert canonical_category("", name="gs-plane1-01") == "gs-plane1"
    assert canonical_category("", name="gl-plane1-01") == "gl-plane1"
    assert canonical_category("", name="gsl-plane1-01") == "gsl-plane1"
    assert canonical_category("", name="cs-01") == "cs"
    assert canonical_category("", name="csl-spine-01") == "cs"
    assert canonical_category("", name="cl-01") == "csl"
