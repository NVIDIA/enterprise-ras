# SPDX-FileCopyrightText: Copyright (c) 2026 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: MIT
"""Tests for scripts/_check_ldap_enabled.py.

The script is called from the Makefile's ``deploy`` recipe as
``_LDAP=$(python3 scripts/_check_ldap_enabled.py "$_ARCH" "$_SITE")``.
It must:

- Print ``1`` iff ``ldap_enabled`` is truthy on the target Excel's
  Settings sheet.
- Print nothing (and exit 0) when the Excel doesn't exist, has no
  Settings sheet, or has no ``ldap_enabled`` row — the Makefile
  substitution treats absence as "LDAP off".
- Re-validate both arguments (arch + site) against the same allowlists
  as ``_era_context.py`` so a future refactor that forgets to call
  the upstream validator still can't traverse paths.
"""
import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest


SCRIPT = Path(__file__).parent.parent / "scripts" / "_check_ldap_enabled.py"


def _run(arch: str, site: str, cwd: Path) -> subprocess.CompletedProcess:
    return subprocess.run(
        [sys.executable, str(SCRIPT), arch, site],
        cwd=str(cwd),
        capture_output=True,
        text=True,
    )


def _make_settings_excel(path: Path, ldap_value: str | None) -> None:
    """Write a tiny Excel with Settings sheet; optionally set ldap_enabled.

    ``ldap_value=None`` means the ldap_enabled row is absent entirely.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Settings"
    ws.cell(row=1, column=1, value="Setting")
    ws.cell(row=1, column=2, value="Value")
    if ldap_value is not None:
        ws.cell(row=2, column=1, value="ldap_enabled")
        ws.cell(row=2, column=2, value=ldap_value)
    wb.save(path)
    wb.close()


@pytest.fixture
def project(tmp_path):
    """Create a dummy project root with the input/<arch>/<site>/ layout."""
    return tmp_path


def _excel_path(project: Path, arch: str, site: str) -> Path:
    d = project / "input" / arch / site
    d.mkdir(parents=True, exist_ok=True)
    return d / f"{arch}.xlsx"


# ---------------------------------------------------------------------------
# ldap_enabled detection
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("value", ["Yes", "yes", "YES", "true", "True", "1"])
def test_truthy_ldap_enabled_prints_one(project, value):
    _make_settings_excel(_excel_path(project, "2-8-5-200", "default"), value)
    result = _run("2-8-5-200", "default", project)
    assert result.returncode == 0, result.stderr
    assert result.stdout.strip() == "1"


@pytest.mark.parametrize("value", ["No", "no", "false", "0", ""])
def test_falsy_ldap_enabled_prints_nothing(project, value):
    _make_settings_excel(_excel_path(project, "2-8-5-200", "default"), value)
    result = _run("2-8-5-200", "default", project)
    assert result.returncode == 0
    assert result.stdout.strip() == ""


def test_ldap_row_absent_prints_nothing(project):
    _make_settings_excel(_excel_path(project, "2-8-5-200", "default"), None)
    result = _run("2-8-5-200", "default", project)
    assert result.returncode == 0
    assert result.stdout.strip() == ""


def test_missing_excel_prints_nothing(project):
    # No input/ layout at all — callers substitute this early in the
    # pipeline before `make import` has run. Must not crash.
    result = _run("2-8-5-200", "default", project)
    assert result.returncode == 0
    assert result.stdout.strip() == ""


def test_missing_settings_sheet_prints_nothing(project):
    path = _excel_path(project, "2-8-5-200", "default")
    wb = openpyxl.Workbook()
    wb.active.title = "Wire Map"  # Settings sheet intentionally absent
    wb.save(path)
    wb.close()
    result = _run("2-8-5-200", "default", project)
    assert result.returncode == 0
    assert result.stdout.strip() == ""


# ---------------------------------------------------------------------------
# Input validation — defence-in-depth against malicious args
# ---------------------------------------------------------------------------

def test_invalid_arch_rejected(project):
    result = _run("1-2-3-456", "default", project)
    assert result.returncode == 2
    assert "invalid arch" in result.stderr


def test_path_traversal_in_site_rejected(project):
    result = _run("2-8-5-200", "../evil", project)
    assert result.returncode == 2
    assert "invalid site" in result.stderr


def test_dotdot_in_site_rejected(project):
    # `..` alone also must not pass — SITE_RE permits single dots but
    # we reject any occurrence of the two-dot sequence outright.
    result = _run("2-8-5-200", "foo..bar", project)
    assert result.returncode == 2


def test_slash_in_site_rejected(project):
    result = _run("2-8-5-200", "a/b", project)
    assert result.returncode == 2


def test_wrong_arg_count_rejected(project):
    result = subprocess.run(
        [sys.executable, str(SCRIPT), "2-8-5-200"],
        cwd=str(project),
        capture_output=True,
        text=True,
    )
    assert result.returncode == 2
    assert "usage" in result.stderr.lower()


def test_validated_args_dont_reach_filesystem_on_rejection(project, tmp_path):
    """If the site fails validation, we must never attempt to
    open a file at the (potentially path-traversing) location."""
    # Set up a sentinel file that a path-traversal attempt might try to
    # read. If the validator is doing its job, we never even try.
    sentinel = tmp_path / "sentinel.xlsx"
    _make_settings_excel(sentinel, "Yes")

    # Point site at a traversal that would resolve to the sentinel.
    result = _run("2-8-5-200", "../sentinel.xlsx", project)
    assert result.returncode == 2
    # Crucially, `1` must NOT appear in stdout — otherwise we leaked
    # the sentinel's ldap_enabled=Yes through a traversed path.
    assert result.stdout.strip() != "1"
