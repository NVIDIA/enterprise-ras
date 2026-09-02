# SPDX-FileCopyrightText: Copyright (c) 2026 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: MIT
"""
ERA-54 / ERA-57 — inputs the generator used to discard without saying so.

Both defects share a shape: the workbook states something authoritative, the
generator quietly ignores it, and the result is a config that applies cleanly and
is wrong. Neither produced an error, and neither shows on the shipped defaults —
they are latent until an operator uses the feature.

  ERA-54  A GSL leaf honoured its OWN Loopbacks-sheet override but computed its
          plane-mate's overlay neighbour from a hardcoded formula. Override one
          leaf and the pair peer at different addresses; the EVPN session never
          establishes. Violates ADR-0033.

  ERA-57  A node cabled to a non-broken-out switch port (`swpN`, no `sN`) was
          dropped on the floor — no role, no bond, no warning.

Driven through the real parser as a subprocess against mutated copies of a
shipped workbook, so they exercise the `make import` path.
"""
import re
import shutil
import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest
import yaml

REPO = Path(__file__).resolve().parent.parent
PARSER = REPO / "scripts" / "excel_parser.py"

ARCH_GSL = "2-8-9-800"
BASE_GSL = REPO / "input" / ARCH_GSL / "default" / f"{ARCH_GSL}.xlsx"

pytestmark = pytest.mark.skipif(
    not BASE_GSL.exists(), reason=f"default {ARCH_GSL} Excel not present")


@pytest.fixture
def gen():
    """Build a scratch site from a shipped workbook, mutate, run the real parser."""
    made = []

    def _run(arch, site, mutate=None):
        in_dir = REPO / "input" / arch / site
        out_dir = REPO / "output" / arch / site
        made.append((in_dir, out_dir))
        in_dir.mkdir(parents=True, exist_ok=True)
        dst = in_dir / f"{arch}.xlsx"
        shutil.copy2(REPO / "input" / arch / "default" / f"{arch}.xlsx", dst)
        if mutate:
            wb = openpyxl.load_workbook(dst)
            mutate(wb)
            wb.save(dst)
        p = subprocess.run(
            [sys.executable, str(PARSER), "--arch", arch, "--site", site,
             "--skip-validate"],
            capture_output=True, text=True, cwd=str(REPO))
        return p, (out_dir / "inventory")

    yield _run
    for in_dir, out_dir in made:
        shutil.rmtree(in_dir, ignore_errors=True)
        shutil.rmtree(out_dir, ignore_errors=True)


def _hv(inv, host):
    return yaml.safe_load((inv / "host_vars" / f"{host}.yml").read_text())


# --------------------------------------------------- ERA-54: plane-mate lo ---

def _set_loopback(wb, switch, value):
    ws = wb["Loopbacks & ASNs"]
    for row in ws.iter_rows(min_row=2):
        if str(row[0].value or "").strip() == switch:
            row[1].value = value
            return
    raise AssertionError(f"{switch} not found on the Loopbacks sheet")


def test_plane_mate_follows_the_loopbacks_sheet(gen):
    """The reported bug: override one leaf, and its mate must peer the new
    address. Previously the mate kept the formula value and the EVPN overlay
    session pointed at an address nobody owned."""
    p, inv = gen(ARCH_GSL, "era54-a",
                 lambda wb: _set_loopback(wb, "gsl-plane1-02", "10.1.1.77"))
    assert p.returncode == 0, p.stdout + p.stderr
    assert _hv(inv, "gsl-plane1-02")["lo_ip"] == "10.1.1.77/32"
    assert _hv(inv, "gsl-plane1-01")["plane_mate_lo_ip"] == "10.1.1.77", \
        "plane-mate address must come from the Loopbacks sheet, not the formula"


def test_plane_mate_override_is_symmetric(gen):
    """Overriding the other member of the pair works the same way."""
    p, inv = gen(ARCH_GSL, "era54-b",
                 lambda wb: _set_loopback(wb, "gsl-plane1-01", "10.1.1.88"))
    assert p.returncode == 0, p.stdout + p.stderr
    assert _hv(inv, "gsl-plane1-02")["plane_mate_lo_ip"] == "10.1.1.88"


def test_plane_mate_falls_back_to_the_formula_when_unset(gen):
    """No override ⇒ the computed address, unchanged. This is what keeps the
    shipped defaults byte-identical."""
    p, inv = gen(ARCH_GSL, "era54-c")
    assert p.returncode == 0, p.stdout + p.stderr
    assert _hv(inv, "gsl-plane1-01")["plane_mate_lo_ip"] == "10.1.1.2"


def test_each_plane_resolves_its_own_mate(gen):
    """Plane 2 must not pick up plane 1's override."""
    p, inv = gen(ARCH_GSL, "era54-d",
                 lambda wb: _set_loopback(wb, "gsl-plane1-02", "10.1.1.77"))
    assert p.returncode == 0, p.stdout + p.stderr
    assert _hv(inv, "gsl-plane2-01")["plane_mate_lo_ip"] == "10.2.1.2"


# ------------------------------------------- ERA-57: bare swpN node ports ---

# The rewired parent port is recorded here so the assertion can target that
# specific port. Asserting only "bond_descriptions is non-empty" is useless —
# the other ~22 nodes keep it populated whether or not the bare port was
# dropped, so such a test passes against the buggy parser too.
_REWIRED = {}


def _rewire_node_to_bare_port(wb):
    """Move ALL of one parent port's node links from swpNsM onto bare swpN."""
    ws = wb["Wire Map"]
    hdr = [str(c.value or "").strip() for c in ws[1]]
    b_port = hdr.index("Port (B)") + 1
    parent = None
    for row in ws.iter_rows(min_row=2):
        val = str(row[b_port - 1].value or "")
        m = re.match(r"^(swp\d+)s\d+$", val)
        if not m:
            continue
        if parent is None:
            parent = m.group(1)
        if m.group(1) == parent:
            row[b_port - 1].value = parent
    assert parent, "no sub-port node link found to rewire"
    _REWIRED["parent"] = parent


def test_bare_swp_node_port_is_not_silently_dropped(gen):
    """A node on a non-broken-out port must still reach `network_roles`.

    Asserts on `network_roles[*].ports` in group_vars, which is what the bond
    loop actually iterates. `bond_descriptions` is NOT a usable signal here: it
    is built by a separate Wire-Map walk that handles bare ports fine, so it
    stays populated either way and the test would pass against the buggy parser.
    Pre-fix the rewired parent is simply absent from the port list."""
    p, inv = gen(ARCH_GSL, "era57-a", _rewire_node_to_bare_port)
    assert p.returncode == 0, p.stdout + p.stderr
    parent_num = int(re.match(r"swp(\d+)", _REWIRED["parent"]).group(1))
    roles = yaml.safe_load((inv / "group_vars" / "csl.yml").read_text())
    cpu_ports = ((roles.get("network_roles") or {}).get("cpu") or {}).get("ports") or []
    assert parent_num in cpu_ports, (
        f"bare port swp{parent_num} was dropped from network_roles; "
        f"cpu ports={cpu_ports}")


def test_unrecognised_switch_port_warns_rather_than_vanishing(gen):
    """Anything that is neither swpN nor swpNsM should say so on stderr."""
    def mutate(wb):
        ws = wb["Wire Map"]
        hdr = [str(c.value or "").strip() for c in ws[1]]
        b_port = hdr.index("Port (B)") + 1
        for row in ws.iter_rows(min_row=2):
            if re.match(r"^swp\d+s\d+$", str(row[b_port - 1].value or "")):
                row[b_port - 1].value = "not-a-port"
                return
    p, _ = gen(ARCH_GSL, "era57-b", mutate)
    assert p.returncode == 0, p.stdout + p.stderr
    assert "not a recognised swp port" in (p.stdout + p.stderr)
