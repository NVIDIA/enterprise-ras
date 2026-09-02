# SPDX-FileCopyrightText: Copyright (c) 2026 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: MIT
"""ERA-73: the ISL Port Profile's breakout must reach gl/gs, not just core/csl.

ADR-0035 says the workbook is the only source of operator intent. The ISL row
governed `core` and `csl` and was ignored by `gl` / `gs` / `gsl`, which took the
template's hardcoded `'2x'` regardless of what the sheet declared.

The mechanism was easy to miss. The plane host_vars already carried
`isl_breakout_parents` — *which* ports break out — but nothing ever set the
*level*, so `gpu_breakout | default('2x')` decided it. An operator editing the
ISL profile changed the N/S fabric and silently did not change the GPU fabric.

Every shipped arch declares Breakout=2 for both the GPU Network and ISL
profiles, which is exactly the hardcoded fallback. So this was latent, not
actively wrong — the generated configs were correct, and the coupling was
missing. That is precisely why it needed a test: nothing in the output revealed
it, and it would have surfaced the first time somebody edited the sheet and
trusted the result.
"""
import re
import shutil
import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest
import yaml

NC = Path(__file__).resolve().parent.parent
ARCH = "2-4-5-800"          # has gl + gs planes and an ISL profile
_COPY_IGNORE = shutil.ignore_patterns(
    "output", "archive", ".git", ".venv*", "__pycache__", "*.pyc",
    ".pytest_cache", ".era-secrets", "legacy",
)
_PLANE_FILES = ("gl_plane1.yml", "gl_plane2.yml", "gs_plane1.yml",
                "gs_plane2.yml", "gsl_plane1.yml", "gsl_plane2.yml")


def _plane_group_vars(tree, arch=ARCH, site="default"):
    d = tree / "output" / arch / site / "inventory" / "group_vars"
    out = {}
    for name in _PLANE_FILES:
        p = d / name
        if p.exists():
            out[name] = yaml.safe_load(p.read_text()) or {}
    return out


def test_plane_group_vars_exist():
    """Fail, do not skip, if the glob finds nothing.

    Every other test here reads the same glob and skips when it is empty, so
    without this a path change would turn the whole file green-by-absence. That
    is the failure mode that let a broken `result.warning()` reach main on the
    same day this was written: a branch nothing executed.
    """
    assert _plane_group_vars(NC), (
        f"no plane group_vars found for {ARCH} — this arch has gl/gs planes, so "
        f"an empty result means the path moved, not that there is nothing to check")


def test_shipped_planes_carry_the_declared_breakout():
    """The level must be present, not left to the template fallback."""
    planes = _plane_group_vars(NC)
    assert planes, f"no plane group_vars for {ARCH} — see test_plane_group_vars_exist"
    missing = [n for n, v in planes.items() if "isl_breakout" not in v]
    assert not missing, (
        f"{missing} carry no isl_breakout — the ISL profile does not reach "
        f"these switches, so editing it would silently not apply")


def _set_isl_breakout(xlsx, value):
    """Set the ISL Port Profile row's Breakout cell, by header name."""
    wb = openpyxl.load_workbook(xlsx)
    ws = wb["VLANs & Profiles"]
    hdr_row = None
    for r in range(1, ws.max_row + 1):
        row = [str(ws.cell(r, c).value or "").strip().lower()
               for c in range(1, ws.max_column + 1)]
        if "breakout" in row and "profile" in row:
            hdr_row = r
            break
    assert hdr_row, "no Port Profiles header row with a Breakout column"
    cols = {str(ws.cell(hdr_row, c).value or "").strip().lower(): c
            for c in range(1, ws.max_column + 1)}
    for r in range(hdr_row + 1, ws.max_row + 1):
        if str(ws.cell(r, cols["profile"]).value or "").strip().upper() == "ISL":
            ws.cell(r, cols["breakout"]).value = value
            wb.save(xlsx)
            return
    pytest.fail("no ISL row in the Port Profiles table")


@pytest.fixture(scope="module")
def edited_tree(tmp_path_factory):
    """A tree whose ISL profile declares 4 instead of the shipped 2."""
    src = NC / "input" / ARCH / "default" / f"{ARCH}.xlsx"
    if not src.exists():
        pytest.skip(f"no workbook for {ARCH}")
    dst = tmp_path_factory.mktemp("nc") / "nc"
    shutil.copytree(NC, dst, ignore=_COPY_IGNORE)
    _set_isl_breakout(dst / "input" / ARCH / "default" / f"{ARCH}.xlsx", 4)

    r = subprocess.run(
        [sys.executable, "scripts/excel_parser.py", "--arch", ARCH,
         "--site", "default", "--skip-validate"],
        cwd=dst, capture_output=True, text=True)
    assert r.returncode == 0, f"parser failed:\n{r.stderr[-2500:]}"
    return dst


def test_editing_the_isl_profile_reaches_the_gpu_planes(edited_tree):
    """The regression itself: 2 -> 4 in the sheet must move gl/gs."""
    planes = _plane_group_vars(edited_tree)
    assert planes, "no plane group_vars generated"
    wrong = {n: v.get("isl_breakout") for n, v in planes.items()
             if v.get("isl_breakout") != "4x"}
    assert not wrong, (
        f"ISL profile set to 4 but these planes still say {wrong} — the sheet "
        f"is not authoritative for the GPU fabric")


def test_gpu_breakout_is_untouched_by_the_isl_edit(edited_tree):
    """The two profiles are independent; editing one must not move the other."""
    planes = _plane_group_vars(edited_tree)
    leaf = {n: v for n, v in planes.items() if n.startswith(("gl_", "gsl_"))}
    if not leaf:
        pytest.skip("no GPU leaf planes")
    wrong = {n: v.get("gpu_breakout") for n, v in leaf.items()
             if v.get("gpu_breakout") != "2x"}
    assert not wrong, (
        f"editing the ISL profile changed gpu_breakout to {wrong}")


def test_divergent_profiles_render_as_two_commands(edited_tree):
    """gl merges GPU and ISL parents into one breakout command when the levels
    agree. They no longer agree here, so it must split — emitting one command
    would silently apply the wrong level to one of the two groups.
    """
    gen = subprocess.run(
        ["ansible-playbook", "playbooks/generate-cli-configs.yml",
         "-i", f"output/{ARCH}/default/inventory/hosts",
         "-e", f"config_output_dir=../output/{ARCH}/default/configs"],
        cwd=edited_tree, capture_output=True, text=True)
    assert gen.returncode == 0, f"generate failed:\n{gen.stderr[-2500:]}"

    cfgs = sorted((edited_tree / "output" / ARCH / "default" / "configs")
                  .glob("gl-plane1-*-config.sh"))
    assert cfgs, "no gl configs rendered"
    txt = cfgs[0].read_text()
    levels = set(re.findall(r"link breakout (\d+x)", txt))
    assert "4x" in levels, (
        f"gl renders {sorted(levels)} — the declared 4x ISL breakout never "
        f"reached the config")
    assert "2x" in levels, (
        f"gl renders {sorted(levels)} — the GPU access ports should still be "
        f"2x; they appear to have been dragged to the ISL level")
