# SPDX-FileCopyrightText: Copyright (c) 2026 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: MIT
"""
Tests for the pre-public-release polish batch:

  * #16 — the OOB-server inventory group is `oob_server` (underscore), not the
          hyphenated `oob-server` that triggers Ansible's "Invalid characters
          in group names" warning. The HOST `oob-server-01` is unaffected.
  * #31 — roles/ldap/templates/users.j2 emits a MISSING_* sentinel instead of a
          blank LDAP base_dn / domain, so a half-filled Excel can't render
          structurally invalid LDIF that silently fails `ldapadd`.
  * #7  — load_workbook_safe() turns a corrupt / non-xlsx file into a friendly
          SystemExit rather than an uncaught openpyxl/zipfile traceback.
"""
import sys
import zipfile
from pathlib import Path

import pytest
from jinja2 import Environment, FileSystemLoader, StrictUndefined

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT / "scripts"))


# ---------------------------------------------------------------------------
# #16 — inventory group name
# ---------------------------------------------------------------------------

SOURCE_HOSTS = sorted((PROJECT_ROOT / "inventories").glob("*/hosts"))


@pytest.mark.parametrize("hosts_file", SOURCE_HOSTS, ids=lambda p: p.parent.name)
def test_source_inventory_uses_underscore_oob_group(hosts_file):
    text = hosts_file.read_text()
    assert "[oob-server]" not in text, (
        f"{hosts_file} still has hyphenated group header [oob-server]"
    )
    # If the arch ships an OOB-server group at all, it must be underscore form.
    if "oob_server" in text or "oob-server-01" in text:
        assert "[oob_server]" in text


def test_generated_inventory_uses_underscore_oob_group():
    """The generator must emit [oob_server], never [oob-server]."""
    import excel_parser  # noqa: E402

    src = Path(excel_parser.__file__).read_text()
    assert 'lines.append("[oob_server]")' in src
    assert 'lines.append("[oob-server]")' not in src


def test_oob_server_host_name_unchanged():
    """The HOST oob-server-01 keeps its hyphen — only the GROUP was renamed.

    The oob-server-01 Air virtual node now lives in the single-home
    scripts/inventory_defaults.yml (host_vars); the seed hosts files are gone and
    the node is only added to the output inventory at Air-deploy time.
    """
    defaults = (PROJECT_ROOT / "scripts" / "inventory_defaults.yml").read_text()
    assert "oob-server-01" in defaults, \
        "expected oob-server-01 vnode in scripts/inventory_defaults.yml"


# ---------------------------------------------------------------------------
# Inventory referential integrity (regression guard for the oob_server rename:
# a renamed group header with a stale `:children` member -> Ansible "undefined
# group" parse failure. Caught live in Air, not by the header-only checks above.)
# ---------------------------------------------------------------------------

COMMITTED_OUTPUT_HOSTS = sorted(
    (PROJECT_ROOT / "output").glob("*/default/inventory/hosts"))
ALL_HOSTS = SOURCE_HOSTS + COMMITTED_OUTPUT_HOSTS


def _parse_ini_inventory(text):
    """Return (defined_groups, children_refs) for an ini-format hosts file.
    defined_groups: every [group] / [group:children] / [group:vars] header.
    children_refs: {child: section} for members listed under a :children block.
    """
    defined, children = set(), {}
    section = None
    for raw in text.splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or line.startswith(";"):
            continue
        if line.startswith("[") and line.endswith("]"):
            name = line[1:-1]
            base = name.split(":", 1)[0]
            defined.add(base)
            section = name
            continue
        if section and section.endswith(":children"):
            # member is a group name (strip any inline host vars, unlikely here)
            children[line.split()[0]] = section
    return defined, children


@pytest.mark.parametrize("hosts_file", ALL_HOSTS,
                         ids=lambda p: str(p.relative_to(PROJECT_ROOT)))
def test_inventory_children_reference_defined_groups(hosts_file):
    """Every group listed under a `[x:children]` block must be a defined group
    header — otherwise Ansible's ini plugin fails with 'undefined group'."""
    defined, children = _parse_ini_inventory(hosts_file.read_text())
    undefined = {c: sec for c, sec in children.items() if c not in defined}
    assert not undefined, (
        f"{hosts_file} references undefined group(s) in :children → "
        f"Ansible ini parse failure: {undefined}")


@pytest.mark.parametrize("hosts_file", ALL_HOSTS,
                         ids=lambda p: str(p.relative_to(PROJECT_ROOT)))
def test_inventory_no_hyphenated_group_headers(hosts_file):
    """No group header may contain a hyphen (Ansible 'Invalid characters'
    warning). Hostnames like oob-server-01 are fine — only headers are checked."""
    bad = [ln.strip() for ln in hosts_file.read_text().splitlines()
           if ln.strip().startswith("[") and ln.strip().endswith("]")
           and "-" in ln.strip().split(":", 1)[0]]
    assert not bad, f"{hosts_file} has hyphenated group header(s): {bad}"


# ---------------------------------------------------------------------------
# #31 — LDAP users.j2 sentinels
# ---------------------------------------------------------------------------

def _render_users_ldif(ldap_ctx):
    template_dir = PROJECT_ROOT / "roles" / "ldap" / "templates"
    env = Environment(
        loader=FileSystemLoader(str(template_dir)),
        undefined=StrictUndefined,
    )
    tmpl = env.get_template("users.j2")
    return tmpl.render(
        ldap=ldap_ctx,
        ldap_users_hashed=[{
            "username": "jdoe", "firstname": "J", "lastname": "Doe",
            "ssha_password": "{SSHA}abc",
        }],
    )


def test_users_ldif_blank_base_dn_emits_sentinel():
    out = _render_users_ldif({"base_dn": "", "domain": ""})
    assert "MISSING_ldap_base_dn_SET_IN_EXCEL" in out
    assert "MISSING_ldap_domain_SET_IN_EXCEL" in out
    # No structurally-invalid trailing-comma DN like "ou=Users," with nothing after.
    assert "ou=Users,\n" not in out
    assert "ou=Groups,\n" not in out
    # No empty mail domain "user@\n".
    assert "@\n" not in out


def test_users_ldif_real_values_have_no_sentinel():
    out = _render_users_ldif({"base_dn": "dc=era,dc=local", "domain": "era.local"})
    assert "MISSING_ldap" not in out
    assert "ou=Users,dc=era,dc=local" in out
    assert "jdoe@era.local" in out


# ---------------------------------------------------------------------------
# #7 — friendly corrupt-workbook error
# ---------------------------------------------------------------------------

def test_load_workbook_safe_corrupt_file(tmp_path):
    import excel_parser  # noqa: E402

    bad = tmp_path / "broken.xlsx"
    bad.write_bytes(b"this is not a real xlsx zip container")
    with pytest.raises(SystemExit) as exc:
        excel_parser.load_workbook_safe(bad)
    msg = str(exc.value)
    assert "Could not open Excel file" in msg
    assert str(bad) in msg


def test_load_workbook_safe_valid_file(tmp_path):
    import excel_parser  # noqa: E402
    import openpyxl

    good = tmp_path / "ok.xlsx"
    openpyxl.Workbook().save(good)
    wb = excel_parser.load_workbook_safe(good)
    assert wb.sheetnames  # opened fine, no SystemExit
    wb.close()


def test_load_workbook_safe_empty_zip_is_caught(tmp_path):
    """A valid zip that isn't an xlsx still yields a friendly error, not a raw
    KeyError/InvalidFileException."""
    import excel_parser  # noqa: E402

    z = tmp_path / "empty.xlsx"
    with zipfile.ZipFile(z, "w") as zf:
        zf.writestr("hello.txt", "not a workbook")
    with pytest.raises(SystemExit):
        excel_parser.load_workbook_safe(z)
