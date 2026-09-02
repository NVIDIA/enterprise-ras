# SPDX-FileCopyrightText: Copyright (c) 2026 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: MIT
"""
OOB VLAN validation guardrails.

Covers:
  - the V1 "Important" finding fix: validate_excel's own Nodes parser must
    read the 'OOB VLAN' column so resolve_oob_vlans() gets per-switch
    mapping data — proven via a real openpyxl worksheet run through
    validate_nodes(), reproducing the exact >1-OOB-VLAN air-mgmt-overlap
    scenario that used to silently no-op.
  - the validate_cross_sheet_data() guardrails: VLAN-id collision,
    invalid OOB-VLAN reference, device-mgmt-IP-outside-OOB-subnet (hard-fail,
    ERA-41), distinct-subnets-require-L3, and OOB-subnet-capacity (hard-fail,
    ERA-41: hosts + auto-derived infra must fit the declared subnet).

Scenario data (VLAN 10 -> 10.10.10.0/24, VLAN 11 -> 10.10.11.0/24) is the
scrubbed multi-OOB-VLAN golden.
"""
import sys
from pathlib import Path

import openpyxl
import pytest

SCRIPTS_DIR = Path(__file__).parent.parent / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from validate_excel import (ValidationResult, validate_cross_sheet_data,
                             validate_nodes, _validate_air_mgmt_overlap,
                             canonical_category)
from excel_parser import resolve_oob_vlans

VLAN_10 = {'id': 10, 'name': 'oob-1', 'subnet': '10.10.10.0/24',
           'gateway': '10.10.10.1', 'vrf': 'OOB', 'network': None,
           'dhcp_relay_client': '', 'row': 3}
VLAN_11 = {'id': 11, 'name': 'oob-2', 'subnet': '10.10.11.0/24',
           'gateway': '10.10.11.1', 'vrf': 'OOB', 'network': None,
           'dhcp_relay_client': '', 'row': 4}


def _oob_node(name, vlan_id, ip, row):
    return {'function': 'oob-switch', 'name': name, 'row': row, 'ip': ip,
            'prefix': 24, 'gateway': None, 'enabled': True,
            'is_air_documentary': False,
            'oob_vlan': '' if vlan_id is None else str(vlan_id)}


def _device_node(function, ip, row):
    return {'function': function, 'name': function, 'row': row, 'ip': ip,
            'prefix': 24, 'gateway': None, 'enabled': True,
            'is_air_documentary': False, 'oob_vlan': ''}


# ---------------------------------------------------------------------------
# V1 "Important" finding: validate_excel must parse the Nodes 'OOB VLAN'
# column, and with >1 OOB VLAN the air-mgmt overlap check must still fire.
# ---------------------------------------------------------------------------

def _build_nodes_ws(rows):
    """Build a minimal Nodes worksheet matching the real shipped header
    layout (Function, Name, Type, MAC Address for ZTP, Mgmt IP Address,
    Prefix, Gateway, ZTP, Enabled, Notes, OOB VLAN)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Nodes"
    headers = ['Function', 'Name', 'Type', 'MAC Address for ZTP',
               'Mgmt IP Address', 'Prefix', 'Gateway', 'ZTP', 'Enabled',
               'Notes', 'OOB VLAN']
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c).value = h
    for r, row in enumerate(rows, start=2):
        for c, v in enumerate(row, start=1):
            ws.cell(row=r, column=c).value = v
    return ws


def test_v1_regression_validate_nodes_parses_oob_vlan_column():
    """validate_nodes() must populate 'oob_vlan' per row (bare id string,
    tolerating float 201.0 -> '201', per excel_parser.py's parse_nodes)."""
    ws = _build_nodes_ws([
        ('oob-switch', 'oob-switch-01', '', '', '10.10.10.5', 24, '10.10.10.1', '', 'Yes', '', 10),
        ('oob-switch', 'oob-switch-02', '', '', '10.10.11.5', 24, '10.10.11.1', '', 'Yes', '', 11.0),
        ('oob-switch', 'oob-switch-03', '', '', '10.10.10.6', 24, '10.10.10.1', '', 'Yes', '', ''),
    ])
    result = ValidationResult()
    parsed_nodes = validate_nodes(ws, result, settings={})
    by_name = {n['name']: n for n in parsed_nodes}
    assert by_name['oob-switch-01']['oob_vlan'] == '10'
    assert by_name['oob-switch-02']['oob_vlan'] == '11'  # float 11.0 -> '11'
    assert by_name['oob-switch-03']['oob_vlan'] == ''    # blank stays blank


def test_v1_regression_two_oob_vlan_air_overlap_fires():
    """Reproduces the exact V1 gap: with 2 OOB VLANs, resolve_oob_vlans()
    used to get nodes without 'oob_vlan' from validate_excel's own parser
    and fall back to the (nonexistent) sole-VLAN default, returning
    subnets=[] — silently no-opping the air-mgmt overlap check. With the
    fix, per-switch resolution works and the SECOND OOB VLAN subnet's
    overlap is caught."""
    ws = _build_nodes_ws([
        ('oob-switch', 'oob-switch-01', '', '', '10.10.10.5', 24, '10.10.10.1', '', 'Yes', '', 10),
        ('oob-switch', 'oob-switch-02', '', '', '10.10.11.5', 24, '10.10.11.1', '', 'Yes', '', 11),
    ])
    result = ValidationResult()
    parsed_nodes = validate_nodes(ws, result, settings={})
    # (Unrelated to this guardrail: a Nodes sheet with only OOB switches and
    # no core/csl trips the separate "no converged-fabric leaf" check — not
    # under test here.)

    parsed_vlans = [VLAN_10, VLAN_11]
    oob_nodes = [n for n in parsed_nodes
                 if canonical_category(n['function'], n['name']) == 'oob-switch']
    mapping = resolve_oob_vlans(parsed_vlans, oob_nodes)

    # This is the crux of the V1 finding: subnets must NOT be empty.
    assert mapping['subnets'] == ['10.10.10.0/24', '10.10.11.0/24']

    # air_mgmt_subnet overlaps ONLY the second OOB VLAN's subnet.
    overlap_result = ValidationResult()
    _validate_air_mgmt_overlap('10.10.11.0/28', mapping['subnets'], overlap_result)
    assert any('10.10.11' in e for e in overlap_result.errors), overlap_result.errors

    # Sanity: an air_mgmt_subnet that overlaps neither must NOT fire.
    clean_result = ValidationResult()
    _validate_air_mgmt_overlap('172.20.0.0/24', mapping['subnets'], clean_result)
    assert not clean_result.errors


# ---------------------------------------------------------------------------
# validate_cross_sheet_data() guardrails
# ---------------------------------------------------------------------------

def test_vlan_id_collision_between_oob_and_non_oob():
    parsed_vlans = [VLAN_10, VLAN_11,
                     {'id': 10, 'name': 'inband-10', 'subnet': '172.16.10.0/24',
                      'gateway': '172.16.10.1', 'vrf': 'INBAND', 'network': None,
                      'dhcp_relay_client': '', 'row': 5}]
    parsed_nodes = [_oob_node('oob-switch-01', 10, '10.10.10.5', 2),
                    _oob_node('oob-switch-02', 11, '10.10.11.5', 3)]
    result = ValidationResult()
    validate_cross_sheet_data({'oob_uplink_mode': 'l3'}, parsed_nodes, parsed_vlans, result)
    assert any('VLAN id 10' in e and 'both an OOB and a non-OOB VLAN' in e
               for e in result.errors), result.errors


def test_invalid_oob_vlan_reference():
    parsed_vlans = [VLAN_10, VLAN_11]
    parsed_nodes = [_oob_node('oob-switch-01', 10, '10.10.10.5', 2),
                    _oob_node('oob-switch-99', 999, '10.10.10.9', 3)]
    result = ValidationResult()
    validate_cross_sheet_data({'oob_uplink_mode': 'l3'}, parsed_nodes, parsed_vlans, result)
    assert any('oob-switch-99' in e and "not a VRF-OOB VLAN" in e
               for e in result.errors), result.errors


def test_blank_oob_vlan_with_multiple_oob_vlans_errors():
    """Final-review fix (Fix 1): with >1 OOB VLAN, a BLANK 'OOB VLAN' cell on
    an OOB switch must NOT silently resolve to None (dropping the switch's
    SVI with no error) — it must fail validation, distinct from the
    non-blank-but-unknown-id case covered by test_invalid_oob_vlan_reference."""
    parsed_vlans = [VLAN_10, VLAN_11]
    parsed_nodes = [_oob_node('oob-switch-01', 10, '10.10.10.5', 2),
                    _oob_node('oob-switch-02', None, '10.10.11.5', 3)]
    result = ValidationResult()
    validate_cross_sheet_data({'oob_uplink_mode': 'l3'}, parsed_nodes, parsed_vlans, result)
    assert any('oob-switch-02' in e and "must name an OOB VLAN" in e
               for e in result.errors), result.errors


def test_device_ip_outside_oob_subnet_errors():
    # ERA-41: a mgmt IP outside every OOB subnet is a hard-fail (was a warn) —
    # it is unreachable (observed: storage-07 assigned .160 outside its /27).
    parsed_vlans = [VLAN_10, VLAN_11]
    parsed_nodes = [
        _oob_node('oob-switch-01', 10, '10.10.10.5', 2),
        _oob_node('oob-switch-02', 11, '10.10.11.5', 3),
        _device_node('gpu-01', '192.168.99.5', 4),
    ]
    result = ValidationResult()
    validate_cross_sheet_data({'oob_uplink_mode': 'l3'}, parsed_nodes, parsed_vlans, result)
    assert any('gpu-01' in e and '192.168.99.5' in e and 'not within any OOB VLAN subnet' in e
               for e in result.errors), result.errors


def test_device_ip_inside_oob_subnet_does_not_error():
    parsed_vlans = [VLAN_10, VLAN_11]
    parsed_nodes = [
        _oob_node('oob-switch-01', 10, '10.10.10.5', 2),
        _oob_node('oob-switch-02', 11, '10.10.11.5', 3),
        _device_node('gpu-01', '10.10.11.20', 4),
    ]
    result = ValidationResult()
    validate_cross_sheet_data({'oob_uplink_mode': 'l3'}, parsed_nodes, parsed_vlans, result)
    assert not any('gpu-01' in e for e in result.errors)


# ---------------------------------------------------------------------------
# ERA-41 capacity: an OOB subnet too small for its hosts + auto-infra hard-fails.
# ---------------------------------------------------------------------------

_VLAN_SMALL = {'id': 20, 'name': 'oob-small', 'subnet': '10.10.20.0/28',
               'gateway': '10.10.20.1', 'vrf': 'OOB', 'network': None,
               'dhcp_relay_client': '', 'row': 3}  # /28 = 14 usable


def _cap_nodes(n_devices):
    # 1 oob-switch (.2) + n_devices (.3 ..) all inside 10.10.20.0/28
    nodes = [_oob_node('oob-switch-01', 20, '10.10.20.2', 2)]
    nodes += [_device_node(f'srv-{i}', f'10.10.20.{i + 3}', i + 3)
              for i in range(n_devices)]
    return nodes


def test_oob_subnet_too_small_hard_fails():
    # 1 switch + 12 devices = 13 on-subnet; +2 L2 infra = 15 > 14 usable.
    # All IPs are INSIDE the /28, so this isolates the capacity check: it must
    # fire even though the containment check finds nothing outside the subnet.
    nodes = _cap_nodes(12)
    result = ValidationResult()
    validate_cross_sheet_data({'oob_uplink_mode': 'l2'}, nodes, [_VLAN_SMALL], result)
    assert any('10.10.20.0/28' in e and 'usable addresses but the deployment needs' in e
               for e in result.errors), result.errors
    # capacity fired in isolation — no containment error (every IP is in-subnet)
    assert not any('not within any OOB VLAN subnet' in e for e in result.errors), result.errors


def test_oob_subnet_large_enough_passes():
    # 1 switch + 5 devices = 6 on-subnet; +2 infra = 8 <= 14 usable.
    nodes = _cap_nodes(5)
    result = ValidationResult()
    validate_cross_sheet_data({'oob_uplink_mode': 'l2'}, nodes, [_VLAN_SMALL], result)
    assert not any('usable addresses but the deployment needs' in e for e in result.errors)


def test_oob_capacity_counts_more_infra_in_l3():
    # 1 switch + 8 devices = 9 on-subnet. L3 infra=6 -> 15 > 14 (fail);
    # L2 infra=2 -> 11 <= 14 (pass). Proves the mode-aware infra count.
    nodes = _cap_nodes(8)
    r_l3 = ValidationResult()
    validate_cross_sheet_data({'oob_uplink_mode': 'l3'}, nodes, [_VLAN_SMALL], r_l3)
    assert any('usable addresses but the deployment needs' in e for e in r_l3.errors), r_l3.errors
    r_l2 = ValidationResult()
    validate_cross_sheet_data({'oob_uplink_mode': 'l2'}, nodes, [_VLAN_SMALL], r_l2)
    assert not any('usable addresses but the deployment needs' in e for e in r_l2.errors)


def test_distinct_subnets_without_l3_errors():
    parsed_vlans = [VLAN_10, VLAN_11]
    parsed_nodes = [_oob_node('oob-switch-01', 10, '10.10.10.5', 2),
                    _oob_node('oob-switch-02', 11, '10.10.11.5', 3)]
    result = ValidationResult()
    validate_cross_sheet_data({'oob_uplink_mode': 'l2'}, parsed_nodes, parsed_vlans, result)
    assert any('require L3 OOB' in e for e in result.errors), result.errors


def test_distinct_subnets_with_l3_does_not_error():
    parsed_vlans = [VLAN_10, VLAN_11]
    parsed_nodes = [_oob_node('oob-switch-01', 10, '10.10.10.5', 2),
                    _oob_node('oob-switch-02', 11, '10.10.11.5', 3)]
    result = ValidationResult()
    validate_cross_sheet_data({'oob_uplink_mode': 'l3'}, parsed_nodes, parsed_vlans, result)
    assert not any('require L3 OOB' in e for e in result.errors), result.errors


def test_single_oob_vlan_happy_path_no_guardrail_errors():
    """Mirrors the shipped single-OOB-VLAN defaults: one OOB VLAN, switches
    with blank 'OOB VLAN' (falls back to the sole default), device inside
    the subnet. None of the four guardrails should fire."""
    parsed_vlans = [VLAN_10,
                     {'id': 300, 'name': 'inband', 'subnet': '172.16.178.0/24',
                      'gateway': '172.16.178.1', 'vrf': 'INBAND', 'network': None,
                      'dhcp_relay_client': '', 'row': 5}]
    parsed_nodes = [
        _oob_node('oob-switch-01', None, '10.10.10.5', 2),
        _oob_node('oob-switch-02', None, '10.10.10.6', 3),
        _device_node('gpu-01', '10.10.10.20', 4),
    ]
    result = ValidationResult()
    validate_cross_sheet_data({'oob_uplink_mode': 'l2'}, parsed_nodes, parsed_vlans, result)
    assert result.errors == []
    assert result.warnings == []
