# SPDX-FileCopyrightText: Copyright (c) 2026 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: MIT
"""One VLAN per OOB switch must reach the access ports, not just the SVI.

The Nodes sheet has carried a per-switch `OOB VLAN` column since ADR-0028, and
`resolve_oob_vlans()` builds a `vlan_by_switch` map from it — but only the L3
side consumed it, to compute `svi_ip`/`gateway`. The access-port bridge VLAN was
resolved globally ("first VLAN named OOB*") and then hardcoded to `access 200`
in the template.

So a switch placed on VLAN 201 got an SVI on the 201 subnet while its host ports
stayed in VLAN 200 — BMCs in a different broadcast domain from their own
gateway. Silently, because every individual piece was valid. A half-wired
feature is worse than an absent one: the workbook accepts the input and the tool
ignores it for the data plane.
"""
import re
import shutil
import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest
import yaml

NC = Path(__file__).resolve().parent.parent
ARCH = "2-8-9-800"
_COPY_IGNORE = shutil.ignore_patterns(
    "output", "archive", ".git", ".venv*", "__pycache__", "*.pyc",
    ".pytest_cache", ".era-secrets", "legacy",
)


def _cell(ws, header):
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(1, c).value or "").strip().lower() == header:
            return c
    return None


@pytest.fixture(scope="module")
def split_vlan_tree(tmp_path_factory):
    """A workbook where oob-switch-02 sits on VLAN 201 instead of 200."""
    src = NC / "input" / ARCH / "default" / f"{ARCH}.xlsx"
    if not src.exists():
        pytest.skip(f"no default workbook for {ARCH}")
    dst = tmp_path_factory.mktemp("nc") / "nc"
    shutil.copytree(NC, dst, ignore=_COPY_IGNORE)
    xlsx = dst / "input" / ARCH / "default" / f"{ARCH}.xlsx"

    wb = openpyxl.load_workbook(xlsx)

    # 1. Declare a second OOB VLAN.
    vp = wb["VLANs & Profiles"]
    hdr = next(r for r in range(1, vp.max_row + 1)
               if str(vp.cell(r, 1).value or "").strip() == "VLAN ID")
    last = hdr + 1
    while str(vp.cell(last + 1, 1).value or "").strip().isdigit():
        last += 1
    vp.insert_rows(last + 1)
    for col, val in enumerate(
        [201, "OOB2", "Out-of-Band Management 2", "192.168.201.0/24",
         "192.168.201.1", "OOB", 4201, "OOB"], start=1
    ):
        vp.cell(last + 1, col).value = val

    # 2. Put oob-switch-02 on it.
    nodes = wb["Nodes"]
    c_name, c_vlan = _cell(nodes, "name"), _cell(nodes, "oob vlan")
    assert c_name and c_vlan, "Nodes sheet lacks Name / OOB VLAN columns"
    moved = 0
    for r in range(2, nodes.max_row + 1):
        if str(nodes.cell(r, c_name).value or "").strip() == "oob-switch-02":
            nodes.cell(r, c_vlan).value = 201
            moved += 1
    assert moved, "no oob-switch-02 row to move"
    wb.save(xlsx)

    r = subprocess.run(
        [sys.executable, "scripts/excel_parser.py", "--arch", ARCH,
         "--site", "default", "--skip-validate"],
        cwd=dst, capture_output=True, text=True)
    assert r.returncode == 0, f"parser failed:\n{r.stderr[-2500:]}"
    return dst


def _host_vars(tree, name):
    return yaml.safe_load(
        (tree / "output" / ARCH / "default" / "inventory" / "host_vars" / f"{name}.yml").read_text())


def test_each_switch_resolves_its_own_access_vlan(split_vlan_tree):
    assert _host_vars(split_vlan_tree, "oob-switch-01")["oob_access_vlan"] == 200
    assert _host_vars(split_vlan_tree, "oob-switch-02")["oob_access_vlan"] == 201


def test_vni_follows_the_same_vlan(split_vlan_tree):
    """`oob_vni` must track the switch's own VLAN, not be looked up separately.

    Resolving them independently is how the two silently disagreed before.
    """
    assert _host_vars(split_vlan_tree, "oob-switch-01")["oob_vni"] == 4200
    assert _host_vars(split_vlan_tree, "oob-switch-02")["oob_vni"] == 4201


def test_access_ports_land_on_the_switchs_own_vlan(split_vlan_tree):
    """The regression itself: host ports must not be pinned to 200."""
    gen = subprocess.run(
        ["ansible-playbook", "playbooks/generate-cli-configs.yml",
         "-i", f"output/{ARCH}/default/inventory/hosts",
         "-e", f"config_output_dir=../output/{ARCH}/default/configs"],
        cwd=split_vlan_tree, capture_output=True, text=True)
    assert gen.returncode == 0, f"generate failed:\n{gen.stderr[-2500:]}"

    cfgdir = split_vlan_tree / "output" / ARCH / "default" / "configs"
    for switch, want in (("oob-switch-01", 200), ("oob-switch-02", 201)):
        txt = (cfgdir / f"{switch}-config.sh").read_text()
        got = {int(m) for m in re.findall(
            r"bridge domain br_default access (\d+)", txt)}
        assert got == {want}, (
            f"{switch}: access ports land on VLAN(s) {sorted(got)}, expected "
            f"{want} — a switch's BMCs would sit in a different broadcast "
            f"domain from their own SVI")


def test_svi_and_access_vlan_agree(split_vlan_tree):
    """The SVI subnet and the access VLAN must describe the same network.

    This is the invariant that was actually violated: the L3 side honoured the
    per-switch VLAN while the L2 side did not.
    """
    hv = _host_vars(split_vlan_tree, "oob-switch-02")
    assert hv["oob_access_vlan"] == 201
    assert str(hv.get("vrr_ip", "")).startswith("192.168.201."), (
        f"SVI {hv.get('vrr_ip')} is not on the VLAN 201 subnet")
