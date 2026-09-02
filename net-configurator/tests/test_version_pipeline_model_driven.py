#!/usr/bin/env python3
# SPDX-FileCopyrightText: Copyright (c) 2026 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: MIT
"""The model → workbook → topology version pipeline must be fully model-driven.

Regression cover for the Cumulus 5.18.0 upgrade, where editing the `versions:`
block in all seven arch models moved exactly ONE row (`core`) and every other
switch role silently kept its old image.

Root cause: the VERSIONS table is keyed by the models' function vocabulary
(`oob-switch`, `gsl-plane1`, `cs`, …) while `_resolve_os()` looks roles up by
the `classify_node()` vocabulary (`oob`, `gsl`, `csl`, …). Only `core` is
spelled the same in both, so only `core` ever matched.

The miss was invisible because an unmatched role fell through to
SWITCH_OS_FALLBACK — a DIFFERENT, OLDER image — with no warning, so a deploy
would report "validated 5.18.0" while running 5.16.0.
"""

import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "scripts"))

from topology_generator import (  # noqa: E402
    SWITCH_OS_FALLBACK,
    TopologyGenerator,
    UnresolvedSwitchVersionError,
)
from utils import reset_mac_registry  # noqa: E402


TARGET = "5.18.0"
TARGET_IMAGE = "cumulus-vx-5.18.0"

# Exactly the vocabulary the arch models use in their `versions:` block.
# See data-models/models/*.yaml — these keys are the contract under test.
MODEL_VERSION_KEYS = [
    "core", "csl", "cs", "gsl-plane1", "gsl-plane2",
    "gs-plane1", "gs-plane2", "gl-plane1", "gl-plane2", "oob-switch",
]

# One switch per model role, named the way the Wire Map names them.
SWITCH_NODES = [
    "core-01", "csl-01", "cs-01",
    "gsl-plane1-01", "gsl-plane2-01",
    "gs-plane1-01", "gs-plane2-01",
    "gl-plane1-01", "gl-plane2-01",
    "oob-switch-01",
]

# DISTINCT version per classify_node() role group. Every model key in a group
# collapses to the same role, so keys within a group must agree — but the
# groups must NOT, or a mis-mapped role can inherit another role's image and
# still look correct. That is exactly how `gsl-plane1` hid: it missed the
# lookup, fell through to core's image, and passed while every role shared
# one version. Distinct versions make the inheritance visible.
ROLE_GROUP_VERSION = {
    "core":        ("5.18.0", "cumulus-vx-5.18.0"),
    "csl":         ("5.17.2", "cumulus-vx-5.17.2"),
    "gsl":         ("5.16.4", "cumulus-vx-5.16.4"),
    "oob":         ("5.15.1", "cumulus-vx-5.15.1"),
}
MODEL_KEY_TO_ROLE = {
    "core": "core",
    "csl": "csl", "cs": "csl",
    "gsl-plane1": "gsl", "gsl-plane2": "gsl",
    "gs-plane1": "gsl", "gs-plane2": "gsl",
    "gl-plane1": "gsl", "gl-plane2": "gsl",
    "oob-switch": "oob",
}
NODE_TO_ROLE = {
    "core-01": "core",
    "csl-01": "csl", "cs-01": "csl",
    "gsl-plane1-01": "gsl", "gsl-plane2-01": "gsl",
    "gs-plane1-01": "gsl", "gs-plane2-01": "gsl",
    "gl-plane1-01": "gsl", "gl-plane2-01": "gsl",
    "oob-switch-01": "oob",
}


@pytest.fixture(autouse=True)
def _reset_macs():
    reset_mac_registry()
    yield
    reset_mac_registry()


def _build_workbook(version_map, image_map_rows, tmp_path, arch="2-8-9-800"):
    """Workbook whose VERSIONS table uses the models' own function vocabulary.

    `version_map` is {model_function_key: version_string}.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_settings = wb.create_sheet("Settings")
    ws_settings.append(["Switch Function", "Cumulus Version"])
    for key, version in version_map.items():
        ws_settings.append([key, version])

    ws = wb.create_sheet("Wire Map")
    ws.append([
        "Display in Air", "System Role", "System Name", "NIC/Port",
        "Speed", "Description", "Network Profile",
        "Disabled by Neighbor", "Speed2", "Description2",
        "Switch Role", "Switch Name", "Switch Port",
    ])
    for i, sw in enumerate(SWITCH_NODES, start=1):
        ws.append([
            "Yes", f"su-01-node-{i:02d}", f"su-01-node-{i:02d}", "NIC1_P1",
            None, None, "CPU/In-Band Network", None, None, None,
            sw, sw, "swp1s0",
        ])

    ws_air = wb.create_sheet("Air_Only")
    ws_air.append([
        "Display in Air", "System Role", "System Name", "NIC/Port",
        "Network Profile", "Switch Role", "Switch Name", "Switch Port",
    ])
    ws_air.append([None] * 8)
    ws_air.append(["Friendly Version", "Air Image"])
    for friendly, image in image_map_rows:
        ws_air.append([friendly, image])

    input_dir = tmp_path / "input" / arch / "default"
    input_dir.mkdir(parents=True, exist_ok=True)
    path = input_dir / f"{arch}.xlsx"
    wb.save(str(path))
    return path


def _uniform_images(tmp_path):
    """All roles pinned to one version — the real 5.18.0 upgrade shape."""
    version_map = {k: TARGET for k in MODEL_VERSION_KEYS}
    path = _build_workbook(version_map, [(TARGET, TARGET_IMAGE)], tmp_path)
    gen = TopologyGenerator(path, arch="2-8-9-800", site="default")
    nodes = gen.generate()["content"]["nodes"]
    return {n: nodes[n]["os"] for n in SWITCH_NODES if n in nodes}


def _distinct_images(tmp_path):
    """Each role group pinned to a DIFFERENT version, so cross-role
    inheritance cannot masquerade as a correct result."""
    version_map = {k: ROLE_GROUP_VERSION[MODEL_KEY_TO_ROLE[k]][0]
                   for k in MODEL_VERSION_KEYS}
    path = _build_workbook(
        version_map, list(ROLE_GROUP_VERSION.values()), tmp_path,
    )
    gen = TopologyGenerator(path, arch="2-8-9-800", site="default")
    nodes = gen.generate()["content"]["nodes"]
    return {n: nodes[n]["os"] for n in SWITCH_NODES if n in nodes}


class TestModelVocabularyReachesTopology:
    """Every role the models can name must resolve to the pinned image."""

    def test_every_switch_role_gets_the_pinned_image(self, tmp_path):
        images = _uniform_images(tmp_path)
        assert images, "no switch nodes were generated — fixture is broken"
        wrong = {n: img for n, img in images.items() if img != TARGET_IMAGE}
        assert not wrong, (
            f"{len(wrong)} switch role(s) did not pick up the pinned "
            f"{TARGET}: {wrong}"
        )

    def test_no_switch_silently_falls_back(self, tmp_path):
        """The specific failure mode: a DIFFERENT, older image, no warning."""
        images = _uniform_images(tmp_path)
        fell_back = {n: img for n, img in images.items()
                     if img == SWITCH_OS_FALLBACK}
        assert not fell_back, (
            f"{len(fell_back)} switch(es) silently fell back to "
            f"{SWITCH_OS_FALLBACK} instead of {TARGET_IMAGE}: {fell_back}"
        )

    def test_oob_specifically(self, tmp_path):
        """`oob-switch` (models) vs `oob` (classify_node) — the exact mismatch."""
        images = _uniform_images(tmp_path)
        assert images.get("oob-switch-01") == TARGET_IMAGE

    def test_each_role_gets_its_OWN_pin_not_another_roles(self, tmp_path):
        """Guards the bug a uniform pin hides.

        With every role on one version, a role that misses the lookup and
        inherits core's image still yields the right string. Pin the groups
        differently and that inheritance becomes a visible mismatch.
        """
        images = _distinct_images(tmp_path)
        expected = {n: ROLE_GROUP_VERSION[NODE_TO_ROLE[n]][1]
                    for n in images}
        wrong = {n: (got, expected[n]) for n, got in images.items()
                 if got != expected[n]}
        assert not wrong, (
            "role(s) resolved to another role's image (got, expected): "
            f"{wrong}"
        )


class TestUnresolvedVersionFailsLoudly:
    """A version with no image-map row must ABORT, never substitute."""

    def test_missing_image_map_row_raises(self, tmp_path):
        # VERSIONS pins 5.18.0 but the image map only knows 5.16.1.
        path = _build_workbook(
            {k: TARGET for k in MODEL_VERSION_KEYS},
            [("5.16.1", "cumulus-vx-5.16.1")], tmp_path,
        )
        with pytest.raises(UnresolvedSwitchVersionError) as exc:
            TopologyGenerator(path, arch="2-8-9-800", site="default")
        msg = str(exc.value)
        assert TARGET in msg, "error must name the unresolved version"
        assert SWITCH_OS_FALLBACK not in msg or "fallback" not in msg.lower(), (
            "must not present the fallback as an acceptable outcome"
        )

    def test_empty_image_map_raises(self, tmp_path):
        path = _build_workbook({k: TARGET for k in MODEL_VERSION_KEYS}, [], tmp_path)
        with pytest.raises(UnresolvedSwitchVersionError):
            TopologyGenerator(path, arch="2-8-9-800", site="default")
