# SPDX-FileCopyrightText: Copyright (c) 2026 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: MIT
"""cust-net-edge switches must stay inside the platform port limit.

An SN5600 has 64 ports. NVUE accepts configuration on swp65+ and `nv config
apply` SUCCEEDS -- the ports simply are not in the datapath. So an
over-subscribed cust-net-edge fails completely silently: at SU32 the air-mgmt
bridge came up with no working members, all 76 switches went unreachable on
172.20.0.0/24, and validate-config reported every switch down while the nodes
themselves were healthy. Nothing errored anywhere in the toolchain.

Two independent defects produced that, and each gets its own guard here:

  1. The Wire Map spread 128 EXIT uplinks over an HA PAIR -- 64 per edge, the
     entire platform budget -- leaving the management star to spill onto
     swp65-swp74.  ->  `cust_edge_count()` now derives the edge count.

  2. Raising the edge count alone was not enough: edges 03/04 appeared in the
     Wire Map but were never DECLARED in the Nodes tab, so topology_generator
     did not treat them as switches. Their rows were skipped, they were wired
     as generic nodes with ethN interfaces instead of swpN, their 32 uplinks
     never landed in `switch_connected`, and the edge-sizing pass -- reading
     that same dict -- saw them as empty and stacked 47 management eth0s on
     top. cust-net-edge-03 finished with 80 interfaces.

The port-budget test walks EVERY committed arch/site pair rather than the two
that happened to break, because the failure is silent and scale-dependent:
whichever arch crosses the line next would otherwise reach Air undetected.
"""
import collections
import importlib.util
import json
import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parents[1]
EDGE_PREFIX = "cust-net-edge"


def _load_topology_generator():
    # scripts/ imports its siblings by bare name (`from utils import ...`).
    sys.path.insert(0, str(REPO / "scripts"))
    spec = importlib.util.spec_from_file_location(
        "topology_generator", REPO / "scripts" / "topology_generator.py"
    )
    module = importlib.util.module_from_spec(spec)
    sys.modules["topology_generator"] = module
    spec.loader.exec_module(module)
    return module


TG = _load_topology_generator()


def _committed_sites():
    out = subprocess.run(
        ["git", "ls-files", "net-configurator/input/*/*/*.xlsx"],
        cwd=REPO.parent, capture_output=True, text=True, check=True,
    ).stdout.split()
    sites = []
    for rel in out:
        parts = Path(rel).parts
        sites.append((parts[-3], parts[-2]))
    return sorted(sites)


SITES = _committed_sites()


def _edge_ports(topology_path: Path):
    """{edge_name: Counter(interface -> times used)} from a topology JSON."""
    content = json.loads(topology_path.read_text())["content"]
    per_edge = collections.defaultdict(collections.Counter)
    for link in content["links"]:
        for endpoint in link:
            if not isinstance(endpoint, dict):
                continue  # "outbound" / "unconnected" sentinel
            node = endpoint["node"]
            if node.startswith(EDGE_PREFIX):
                per_edge[node][endpoint["interface"]] += 1
    return per_edge


def _topology_for(arch: str, site: str) -> Path:
    return REPO / "output" / arch / site / "topology" / f"{arch}-topology.json"


@pytest.mark.parametrize("arch,site", SITES)
def test_no_edge_exceeds_platform_port_limit(arch, site):
    """No cust-net-edge may use more ports than the platform physically has."""
    topology = _topology_for(arch, site)
    if not topology.exists():
        pytest.skip(f"{arch}/{site} not generated in this working tree")
    for edge, ports in sorted(_edge_ports(topology).items()):
        assert len(ports) <= TG.EDGE_PORT_LIMIT, (
            f"{arch}/{site}: {edge} uses {len(ports)} ports, "
            f"limit is {TG.EDGE_PORT_LIMIT}. Ports past the limit accept "
            f"config and apply cleanly but never forward."
        )


@pytest.mark.parametrize("arch,site", SITES)
def test_no_edge_uses_a_port_index_past_the_limit(arch, site):
    """swp65+ is the specific failure mode: it applies, then silently drops."""
    topology = _topology_for(arch, site)
    if not topology.exists():
        pytest.skip(f"{arch}/{site} not generated in this working tree")
    for edge, ports in sorted(_edge_ports(topology).items()):
        for iface in ports:
            if not iface.startswith("swp"):
                continue
            index = int(iface[3:].split("s")[0])
            assert index <= TG.EDGE_PORT_LIMIT, (
                f"{arch}/{site}: {edge}:{iface} is past swp{TG.EDGE_PORT_LIMIT}. "
                f"`nv config apply` will SUCCEED on it and it will not forward."
            )


@pytest.mark.parametrize("arch,site", SITES)
def test_every_wiremap_edge_is_declared_in_nodes(arch, site):
    """Defect 2: a Wire Map edge with no Nodes row is not treated as a switch.

    It is then wired with ethN interfaces instead of swpN and its uplink load
    is invisible to the sizing pass, so the management star lands on top of it.
    """
    workbook = REPO / "input" / arch / site / f"{arch}.xlsx"
    wb = openpyxl.load_workbook(workbook, data_only=True)

    wiremap_edges = set()
    for row in wb["Wire Map"].iter_rows(min_row=2, values_only=True):
        for name in (row[1], row[5]):
            if name and str(name).startswith(EDGE_PREFIX):
                wiremap_edges.add(str(name))

    header = [str(c.value).strip() if c.value else "" for c in wb["Nodes"][1]]
    name_col = header.index("Name")
    declared = {
        str(row[name_col])
        for row in wb["Nodes"].iter_rows(min_row=2, values_only=True)
        if row[name_col] and str(row[name_col]).startswith(EDGE_PREFIX)
    }

    assert wiremap_edges <= declared, (
        f"{arch}/{site}: Wire Map names {sorted(wiremap_edges - declared)} "
        f"but the Nodes tab does not declare them. topology_generator will "
        f"skip their rows and wire them as generic nodes."
    )


@pytest.mark.parametrize("arch,site", SITES)
def test_no_duplicate_interface_on_an_edge(arch, site):
    """Two links on one edge port makes Air reject the whole topology."""
    topology = _topology_for(arch, site)
    if not topology.exists():
        pytest.skip(f"{arch}/{site} not generated in this working tree")
    for edge, ports in sorted(_edge_ports(topology).items()):
        dupes = {iface: n for iface, n in ports.items() if n > 1}
        assert not dupes, f"{arch}/{site}: {edge} has duplicate interfaces {dupes}"


@pytest.mark.parametrize("arch,site", SITES)
def test_every_uplink_carrying_edge_has_an_egress_leg(arch, site):
    """An EXIT edge with no external-conn leg black-holes customer traffic.

    ADR-0002 allowed edges 03+ to exist as mgmt-bridge spokes only "until the
    topology generator deliberately adds more EXIT uplinks plus matching
    external-conn interfaces/subnets". Once an edge carries uplinks it needs
    the leg. Mgmt-only spokes must NOT get one -- that would burn an
    external-conn interface and a /24 for no traffic.
    """
    topology = _topology_for(arch, site)
    if not topology.exists():
        pytest.skip(f"{arch}/{site} not generated in this working tree")
    content = json.loads(topology.read_text())["content"]

    uplink_edges, egress_edges = set(), set()
    for link in content["links"]:
        endpoints = [e for e in link if isinstance(e, dict)]
        if len(endpoints) != 2:
            continue
        for near, far in (endpoints, endpoints[::-1]):
            if not near["node"].startswith(EDGE_PREFIX):
                continue
            if far["node"] == "external-conn":
                egress_edges.add(near["node"])
            # A fabric EXIT uplink: the peer is a cluster switch and the link
            # is NOT that switch's eth0 management leg.
            elif (
                far["interface"] != "eth0"
                and not far["node"].startswith(EDGE_PREFIX)
                and far["interface"].startswith("swp")
            ):
                uplink_edges.add(near["node"])

    assert uplink_edges <= egress_edges, (
        f"{arch}/{site}: {sorted(uplink_edges - egress_edges)} carry EXIT "
        f"uplinks but have no external-conn egress leg; their customer "
        f"traffic has no path out."
    )
