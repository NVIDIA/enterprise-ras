# SPDX-FileCopyrightText: Copyright (c) 2026 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: MIT
"""ERA-44: a storage SERVER access port (L2, bond into VLAN 500) must not be
dropped just because an L3 "Storage Uplink" (VRF=STORAGE) profile is defined.
The discriminator is the profile MODE (L3+VRF=STORAGE = uplink), never the
profile name or the peer's role.

Observed in the field: storage servers on the L2 "Storage" profile lost their bonds +
VLAN 500 membership when a separate L3 Storage Uplink profile existed.
"""
import shutil
import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest

SCRIPTS_DIR = Path(__file__).parent.parent / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from excel_parser import _classify_core_profile  # noqa: E402

NC = Path(__file__).resolve().parent.parent
ARCH = "2-4-3-200"
_COPY_IGNORE = shutil.ignore_patterns(
    "output", "archive", ".git", ".venv*", "__pycache__", "*.pyc",
    ".pytest_cache", ".era-secrets", "legacy",
)


class TestClassifyByMode:
    def test_l2_storage_uplink_named_profile_is_server(self):
        # A profile literally named "Storage Uplink" but NOT L3 is an L2 storage
        # SERVER access port (2-4-3-200 / 2-8-5-200 author their servers this way).
        assert _classify_core_profile("Storage Uplink", ["storage"], set()) == 'storage'

    def test_l3_storage_uplink_is_uplink(self):
        assert _classify_core_profile(
            "Storage Uplink", ["storage"], {"Storage Uplink"}) == 'storage_uplink'

    def test_storage_server_profile_stays_server_even_with_l3_uplink_defined(self):
        # The observed case: servers on "Storage" (L2), an L3 "Storage Uplink" also
        # exists — the server must still classify as 'storage'.
        assert _classify_core_profile("Storage", ["storage"], {"Storage Uplink"}) == 'storage'

    def test_storage_server_peer_is_server(self):
        # 'storage' in sw (the server's own role), no L3 profile -> server.
        assert _classify_core_profile("Storage", ["storage-01"], set()) == 'storage'


def _author_l2_storage(xlsx):
    """Add customer-authored L2 storage to an L3-only workbook.

    ADR-0047 made the shipped samples L3-only: they carry no `Storage` port
    profile, no VLAN 500 and no storage-server Wire Map rows. **L2 storage
    remains fully supported** — it is simply not what we ship — so this test
    builds the L2 scenario explicitly instead of mutating a shipped workbook
    that no longer contains one. That also makes it a direct test of the
    supported-but-not-shipped path, which is the thing at risk of rotting.

    Everything is located by name. The previous revision hardcoded `vp["B26"]`
    to reach the Storage Uplink profile; row indices shift whenever a profile
    is added, and a mis-aimed write silently rewrites a different profile.
    """
    wb = openpyxl.load_workbook(xlsx)
    vp = wb["VLANs & Profiles"]

    def find_row(label):
        for r in range(1, vp.max_row + 1):
            if str(vp.cell(r, 1).value or "").strip() == label:
                return r
        raise AssertionError(f"{label!r} not found in VLANs & Profiles")

    # 1. VLAN 500 — inserted directly after the last VLAN row (VLAN ID header
    #    section), so it lands inside the VLAN block rather than after it.
    vlan_hdr = find_row("VLAN ID")
    last_vlan = vlan_hdr + 1
    while str(vp.cell(last_vlan + 1, 1).value or "").strip().isdigit():
        last_vlan += 1
    vp.insert_rows(last_vlan + 1)
    for col, val in enumerate(
        [500, "Storage", "Storage Network", "172.16.180.0/24",
         "172.16.180.1", "STORAGE", 4500, "No"], start=1
    ):
        vp.cell(last_vlan + 1, col).value = val

    # 2. The L2 `Storage` port profile, immediately after the Profile header.
    prof_hdr = find_row("Profile")
    vp.insert_rows(prof_hdr + 1)
    for col, val in enumerate(
        # Access on VLAN 500, untagged. Storage servers bond untagged (bond1),
        # so the switch port is an access port in the storage VLAN. The shipped
        # workbooks used to carry a Trunk/400,500/untagged-300 Storage row, but
        # nothing ever referenced it (storage servers were never cabled), so
        # that shape was never exercised — it renders `vlan 400,500` and the
        # server's untagged frames land outside VLAN 500.
        ["Storage", "Access", "500", None, None, None, "Yes", "200G", 4, 2],
        start=1,
    ):
        vp.cell(prof_hdr + 1, col).value = val

    # 3. Storage servers, dual-homed to the core pair on the L2 profile.
    nodes = wb["Nodes"]
    wm = wb["Wire Map"]
    for idx in (1, 2):
        name = f"storage-{idx:02d}"
        nodes.append([
            "storage", name, None, "server", None,
            f"192.168.200.{60 + idx}", 24, "192.168.200.1", "No", "Yes", None,
        ])
        for link, core in enumerate(("core-01", "core-02"), start=1):
            wm.append([
                "Yes", name, f"B3220 SL1 P{link}", "NA", None,
                core, f"swp{40 + idx}s{link - 1}", "NA", "1.0", "Storage",
            ])
    wb.save(xlsx)


def test_storage_servers_bond_alongside_l3_uplink(tmp_path):
    """End-to-end: storage servers on the L2 "Storage" profile + a separate L3
    "Storage Uplink" (VRF=STORAGE) profile -> storage servers still get their
    VLAN-500 bonds, and the L3 VRF STORAGE uplink is still emitted.

    The shipped workbook supplies the L3 half (ADR-0047); `_author_l2_storage`
    supplies the L2 half a customer would write. Both must survive together —
    that coexistence is the ERA-44 regression this guards.
    """
    src = NC / "input" / ARCH / "default" / f"{ARCH}.xlsx"
    if not src.exists():
        pytest.skip(f"no committed default workbook for {ARCH}")
    dst = tmp_path / "nc"
    shutil.copytree(NC, dst, ignore=_COPY_IGNORE)

    xlsx = dst / "input" / ARCH / "default" / f"{ARCH}.xlsx"
    _author_l2_storage(xlsx)

    r = subprocess.run(
        [sys.executable, "scripts/excel_parser.py", "--arch", ARCH,
         "--site", "default", "--skip-validate"],
        cwd=dst, capture_output=True, text=True)
    assert r.returncode == 0, f"parser failed:\n{r.stderr[-2000:]}"

    # render one core config via the same ansible path make generate uses
    gen = subprocess.run(
        ["ansible-playbook", "playbooks/generate-cli-configs.yml",
         "-i", f"output/{ARCH}/default/inventory/hosts",
         "-e", f"config_output_dir=../output/{ARCH}/default/configs"],
        cwd=dst, capture_output=True, text=True)
    assert gen.returncode == 0, f"generate failed:\n{gen.stderr[-2000:]}"

    cfg = (dst / "output" / ARCH / "default" / "configs" / "core-01-config.sh").read_text()
    assert "description storage-01" in cfg, "storage server bond dropped"
    assert "bridge domain br_default access 500" in cfg, "storage VLAN 500 membership dropped"
    assert "vlan500 vrf STORAGE" in cfg, "STORAGE VRF uplink missing"
