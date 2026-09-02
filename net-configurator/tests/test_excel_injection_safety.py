# SPDX-FileCopyrightText: Copyright (c) 2026 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: MIT
"""Every Excel cell that renders into a root-executed config is guarded.

The Settings sheet has had this protection since the 2026-06-30 security review
(`SHELL_INJECTION_PRONE_KEYS`). It covered Settings ONLY, so the other
operator-supplied surfaces that reach a template verbatim were unguarded.

This docstring used to claim the guard was already complete ("every Excel cell
that renders UNQUOTED ... is guarded"). It was not, and asserting completeness is
part of why the gap survived a later 89-finding QA audit that concluded
"Excel-to-shell injection defence is thorough and tested". Four columns on the
policy sheets were in neither `_UNQUOTED_EXCEL_CELLS` nor quoted by the template
— Route policy `Rule` (col 2) and `Set value` (col 7), Prefix lists `Max prefix
length` (col 4), and Community lists `Rule` (col 2) — so a payload in any of them
validated cleanly and landed verbatim in a config that ZTP runs as root. They are
in CASES below now, and `core_nvue_cli.j2` quotes the policy values as well, so
the guard is defence-in-depth rather than validator-only.

Do not restate this as complete. State what is covered and let CASES be the list.

Two vectors were confirmed end-to-end before this check existed:

    VRF Name        `OOB; touch /tmp/x`      -> nv set vrf OOB; touch /tmp/x ...
    Auto-Negotiate  `disabled; touch /tmp/x` -> nv set interface ... link
                                                auto-negotiate disabled; touch /tmp/x

Both execute as root on every switch that receives the config.
"""
import shutil
import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest

NC = Path(__file__).resolve().parent.parent
ARCH = "2-4-3-200"
ARCHS = ["2-4-3-200", "2-4-5-800", "2-8-5-200",
         "2-8-9-400", "2-8-9-400-SP", "2-8-9-800"]
_COPY_IGNORE = shutil.ignore_patterns(
    "output", "archive", ".git", ".venv*", "__pycache__", "*.pyc",
    ".pytest_cache", ".era-secrets", "legacy",
)


def _section(ws, label):
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(r, 1).value or "").strip() == label:
            return r
    raise AssertionError(f"section {label!r} not found")


# (id, sheet, locator) — locator(wb) writes the payload and returns nothing.
CASES = [
    ("vrf-name", lambda wb, v: wb["VLANs & Profiles"].cell(
        _section(wb["VLANs & Profiles"], "VRF Name") + 1, 1, v)),
    ("vlan-name", lambda wb, v: wb["VLANs & Profiles"].cell(
        _section(wb["VLANs & Profiles"], "VLAN ID") + 1, 2, v)),
    ("profile-vrf", lambda wb, v: wb["VLANs & Profiles"].cell(
        _section(wb["VLANs & Profiles"], "Profile") + 1, 6, v)),
    ("profile-speed", lambda wb, v: wb["VLANs & Profiles"].cell(
        _section(wb["VLANs & Profiles"], "Profile") + 1, 8, v)),
    ("profile-autoneg", lambda wb, v: wb["VLANs & Profiles"].cell(
        _section(wb["VLANs & Profiles"], "Profile") + 1, 11, v)),
    ("prefix-list-name", lambda wb, v: wb["Prefix lists"].cell(2, 1, v)),
    ("prefix-list-max-len", lambda wb, v: wb["Prefix lists"].cell(2, 4, v)),
    ("route-map-name", lambda wb, v: wb["Route policy"].cell(2, 1, v)),
    ("route-map-rule", lambda wb, v: wb["Route policy"].cell(2, 2, v)),
    ("route-map-set-value", lambda wb, v: wb["Route policy"].cell(2, 7, v)),
    ("community-list-name", lambda wb, v: wb["Community lists"].cell(2, 1, v)),
    ("community-list-rule", lambda wb, v: wb["Community lists"].cell(2, 2, v)),
    ("acl-name", lambda wb, v: wb["ACLs"].cell(2, 1, v)),
]


@pytest.mark.parametrize("name,writer", CASES, ids=[c[0] for c in CASES])
@pytest.mark.parametrize("payload", ["evil; touch /tmp/era_pwned",
                                     "evil`id`", "evil$(id)", "evil|nc"])
def test_shell_metacharacters_are_rejected(name, writer, payload, tmp_path):
    src = NC / "input" / ARCH / "default" / f"{ARCH}.xlsx"
    if not src.exists():
        pytest.skip(f"no default workbook for {ARCH}")
    xlsx = tmp_path / f"{ARCH}.xlsx"
    shutil.copy(src, xlsx)
    wb = openpyxl.load_workbook(xlsx)
    writer(wb, payload)
    wb.save(xlsx)

    r = subprocess.run([sys.executable, "scripts/validate_excel.py", str(xlsx)],
                       cwd=NC, capture_output=True, text=True)
    out = r.stdout + r.stderr
    assert r.returncode != 0, f"{name}: {payload!r} validated cleanly"
    assert "command injection" in out, out[-600:]


@pytest.mark.parametrize("arch", ARCHS)
@pytest.mark.parametrize("site", ["default", "largescale"])
def test_shipped_workbooks_have_no_false_positives(arch, site):
    """A guard that fires on legitimate data would be worse than none.

    Real workbooks carry values with `/`, `:`, `,`, `.`, `-` and spaces
    (subnets, communities, port lists), all of which must remain acceptable.
    """
    xlsx = NC / "input" / arch / site / f"{arch}.xlsx"
    if not xlsx.exists():
        pytest.skip(f"no {site} workbook for {arch}")
    r = subprocess.run([sys.executable, "scripts/validate_excel.py", str(xlsx)],
                       cwd=NC, capture_output=True, text=True)
    assert "command injection" not in (r.stdout + r.stderr), (
        f"{arch}/{site}: injection guard fired on a shipped workbook")


# --- node / system names ----------------------------------------------------
# Names are used three ways, all of which require a hostname charset:
#   * rendered UNQUOTED as `nv set interface <bond> description <name>`;
#   * written verbatim into the generated ansible inventory/hosts;
#   * used as NVUE object names, where a space silently truncates.
# Confirmed vector: renaming a node to `su-01-node-01; touch /tmp/x` in both
# the Nodes and Wire Map sheets put it in bond_descriptions and rendered as
# `nv set interface bond1s0 description su-01-node-01; touch /tmp/x`.

_NAME_PAYLOADS = [
    "su-01-node-01; touch /tmp/era_pwned",
    "su-01-node-01`id`",
    "su-01-node-01 $(id)",
    "su-01 node-01",          # a bare space truncates the NVUE description
]


def _rename_node(xlsx, new_name):
    wb = openpyxl.load_workbook(xlsx)
    ws = wb["Wire Map"]
    h = [str(c.value or "").strip() for c in next(ws.iter_rows())]
    for col in ("System Name (A)", "System Name (B)"):
        if col not in h:
            continue
        i = h.index(col) + 1
        for r in range(2, ws.max_row + 1):
            if str(ws.cell(r, i).value or "").strip() == "su-01-node-01":
                ws.cell(r, i).value = new_name
    nws = wb["Nodes"]
    nh = [str(c.value or "").strip().lower() for c in next(nws.iter_rows())]
    ci = nh.index("name") + 1
    for r in range(2, nws.max_row + 1):
        if str(nws.cell(r, ci).value or "").strip() == "su-01-node-01":
            nws.cell(r, ci).value = new_name
    wb.save(xlsx)


@pytest.mark.parametrize("payload", _NAME_PAYLOADS)
def test_invalid_node_names_are_rejected(payload, tmp_path):
    src = NC / "input" / ARCH / "default" / f"{ARCH}.xlsx"
    if not src.exists():
        pytest.skip(f"no default workbook for {ARCH}")
    xlsx = tmp_path / f"{ARCH}.xlsx"
    shutil.copy(src, xlsx)
    _rename_node(xlsx, payload)
    r = subprocess.run([sys.executable, "scripts/validate_excel.py", str(xlsx)],
                       cwd=NC, capture_output=True, text=True)
    out = r.stdout + r.stderr
    assert r.returncode != 0, f"{payload!r} validated cleanly"
    assert "not a valid hostname" in out, out[-600:]


def test_legitimate_hostnames_are_accepted(tmp_path):
    """Dots, hyphens and underscores are all legal in real names."""
    src = NC / "input" / ARCH / "default" / f"{ARCH}.xlsx"
    if not src.exists():
        pytest.skip(f"no default workbook for {ARCH}")
    for name in ("su-01-node-01a", "node_01", "node.example.com", "SU-01-NODE-01"):
        xlsx = tmp_path / f"{name}.xlsx"
        shutil.copy(src, xlsx)
        _rename_node(xlsx, name)
        r = subprocess.run([sys.executable, "scripts/validate_excel.py", str(xlsx)],
                           cwd=NC, capture_output=True, text=True)
        assert "not a valid hostname" not in (r.stdout + r.stderr), (
            f"legitimate hostname {name!r} was rejected")
