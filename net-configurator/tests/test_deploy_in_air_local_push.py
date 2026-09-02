# SPDX-FileCopyrightText: Copyright (c) 2026 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: MIT
"""ERA-48: with Settings 'Deploy in Air = No', switch inventory must use each
switch's real Nodes-tab mgmt IP (for local push to hardware) instead of an
auto-assigned air-mgmt (172.20.0.x) IP. The default (Deploy in Air = Yes) keeps
the auto-assigned Air IP.

Runs the parser against a throwaway copy of the tool tree (never the live repo),
mirroring test_asn_override_e2e.py.
"""
import shutil
import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest

NC = Path(__file__).resolve().parent.parent  # net-configurator/
ARCH = "2-4-3-200"

_COPY_IGNORE = shutil.ignore_patterns(
    "output", "archive", ".git", ".venv*", "__pycache__", "*.pyc",
    ".pytest_cache", ".era-secrets", "legacy",
)


def _run_parser(dst):
    r = subprocess.run(
        [sys.executable, "scripts/excel_parser.py", "--arch", ARCH,
         "--site", "default", "--skip-validate"],
        cwd=dst, capture_output=True, text=True,
    )
    assert r.returncode == 0, f"parser failed:\n{r.stderr[-2000:]}"
    return dst / "output" / ARCH / "default" / "inventory" / "host_vars"


def _set_deploy_no_with_switch_ips(xlsx):
    """Set Deploy in Air = No and give each switch a real (non-Air) mgmt IP,
    as an operator would for local push. Returns {switch_name: ip}."""
    wb = openpyxl.load_workbook(xlsx)
    for row in wb["Settings"].iter_rows():
        if row[0].value and str(row[0].value).strip() == "deploy_in_air":
            row[1].value = "No"
    ns = wb["Nodes"]
    hdr = [str(c.value or '').strip() for c in next(ns.iter_rows())]
    ni, mi = hdr.index("Name"), hdr.index("Mgmt IP Address")
    ips, octet = {}, 11
    for row in ns.iter_rows(min_row=2):
        name = str(row[ni].value or '')
        if name.startswith(("oob-switch", "core-")):
            ip = f"10.10.0.{octet}"
            row[mi].value = ip
            ips[name] = ip
            octet += 1
    wb.save(xlsx)
    return ips


def test_deploy_in_air_no_uses_real_switch_ip(tmp_path):
    src = NC / "input" / ARCH / "default" / f"{ARCH}.xlsx"
    if not src.exists():
        pytest.skip(f"no committed default workbook for {ARCH}")
    dst = tmp_path / "nc"
    shutil.copytree(NC, dst, ignore=_COPY_IGNORE)
    ips = _set_deploy_no_with_switch_ips(dst / "input" / ARCH / "default" / f"{ARCH}.xlsx")
    assert ips, "expected some switch rows to renumber"

    hv_dir = _run_parser(dst)
    for sw, ip in ips.items():
        hv = hv_dir / f"{sw}.yml"
        assert hv.exists(), f"missing host_vars for {sw}"
        text = hv.read_text()
        assert f"ansible_host: {ip}" in text, \
            f"{sw} should use its real Nodes-tab IP {ip}; got:\n{text}"
        assert "172.20.0." not in text, \
            f"{sw} must NOT get an air-mgmt IP when Deploy in Air = No:\n{text}"
        assert "ip_assignment_mode: static" in text


def test_deploy_in_air_yes_keeps_air_ip(tmp_path):
    # Unmodified default (Deploy in Air = Yes) — switches keep the air-mgmt IP.
    src = NC / "input" / ARCH / "default" / f"{ARCH}.xlsx"
    if not src.exists():
        pytest.skip(f"no committed default workbook for {ARCH}")
    dst = tmp_path / "nc"
    shutil.copytree(NC, dst, ignore=_COPY_IGNORE)

    hv_dir = _run_parser(dst)
    text = (hv_dir / "oob-switch-01.yml").read_text()
    assert "ansible_host: 172.20.0." in text, \
        f"Deploy in Air = Yes must keep the air-mgmt IP:\n{text}"
    assert "ip_assignment_mode: dhcp" in text
