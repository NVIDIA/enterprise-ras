# SPDX-FileCopyrightText: Copyright (c) 2026 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: MIT
"""
Tests for the SU-scaling pipeline:

  * scripts/arch_scaling.py  — single-tier fan-out tables + helpers
  * scripts/validate_excel.py — single-tier-cap + Air-OOB-single-cable
    validators

All tests run end-to-end against the real default Excels under input/
so a schema drift surfaces here before it surfaces in CI.
"""
import shutil
import sys
from pathlib import Path

import openpyxl
import pytest

SCRIPTS_DIR = Path(__file__).parent.parent / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from arch_scaling import (
    ARCH_SCALING,
    MAX_SUPPORTED_SU,
    ScalingTier,
    get_tier,
    is_supported_single_tier,
    max_single_tier_su,
    max_supported_su,
    node_name_to_su,
)
from validate_excel import validate_excel


def _wm_cols(ws):
    """Resolve Wire Map B-side column indices from header row.

    Returns (SWNAME_COL, SWPORT_COL). After the 2026-05-28 schema
    revision added `Cable Split (A)` between Port Side (A) and System
    Name (B), B-side columns shifted from (5, 6) to (6, 7). Header
    lookup keeps the tests resilient to further column changes.
    """
    headers = {}
    for c in range(1, ws.max_column + 1):
        h = (ws.cell(1, c).value or "")
        headers[str(h).strip().lower()] = c
    swname = headers.get("system name (b)",
              headers.get("switch name", 5))
    swport = headers.get("port (b)",
              headers.get("switch port", 6))
    return swname, swport


# ---------------------------------------------------------------------------
# arch_scaling helpers
# ---------------------------------------------------------------------------

class TestNodeNameToSu:
    def test_su_pattern(self):
        assert node_name_to_su("su-1-node-1") == 1
        assert node_name_to_su("su-5-node-3") == 5
        assert node_name_to_su("su-12-node-99") == 12

    def test_gpu_pattern_quarters(self):
        # 4 GPU nodes per SU on 2-8-9-800
        assert node_name_to_su("gpu-1") == 1
        assert node_name_to_su("gpu-2") == 1
        assert node_name_to_su("gpu-3") == 1
        assert node_name_to_su("gpu-4") == 1
        assert node_name_to_su("gpu-5") == 2
        assert node_name_to_su("gpu-8") == 2
        assert node_name_to_su("gpu-9") == 3
        assert node_name_to_su("gpu-16") == 4
        assert node_name_to_su("gpu-17") == 5

    def test_non_compute_returns_none(self):
        for n in ("core-01", "oob-switch-02", "csl-01", "gsl-plane1-02",
                  "dhcp-oob", "oob-server-01", "external-conn-01"):
            assert node_name_to_su(n) is None, f"{n} should not match SU pattern"

    def test_edges(self):
        assert node_name_to_su("") is None
        assert node_name_to_su(None) is None
        # gpu without index → not a compute node
        assert node_name_to_su("gpu") is None
        assert node_name_to_su("gpu-abc") is None

    def test_whitespace_tolerated(self):
        assert node_name_to_su("  su-3-node-2  ") == 3
        assert node_name_to_su("  gpu-5  ") == 2


class TestMaxSingleTierSu:
    def test_each_arch_returns_documented_max(self):
        # These caps are baked in arch_scaling.py from the architecture PDFs;
        # changing them here means the table moved → review with arch docs.
        assert max_single_tier_su("2-4-3-200") == 8
        assert max_single_tier_su("2-8-5-200") == 8
        assert max_single_tier_su("2-8-9-400") == 3
        assert max_single_tier_su("2-8-9-800") == 4

    def test_unknown_arch_returns_none(self):
        assert max_single_tier_su("9-9-9-999") is None
        assert max_single_tier_su("") is None


class TestMaxSupportedSu:
    """The public validator falls back to max_supported_su (single- OR
    multi-tier max) when the internal generator models are absent, so the
    shipped largescale example workbooks must validate."""

    def test_matches_largescale_ceilings(self):
        assert max_supported_su("2-4-3-200") == 8
        assert max_supported_su("2-8-5-200") == 8
        assert max_supported_su("2-8-9-400") == 16   # multi-tier > single-tier max (3)
        assert max_supported_su("2-4-5-800") == 8
        assert max_supported_su("2-8-9-800") == 32
        assert max_supported_su("2-8-9-400-SP") == 32

    def test_all_six_archs_present(self):
        assert set(MAX_SUPPORTED_SU) == {
            "2-4-3-200", "2-8-5-200", "2-8-9-400",
            "2-4-5-800", "2-8-9-800", "2-8-9-400-SP"}

    def test_at_least_single_tier_max(self):
        # Where an arch has a single-tier cap, the validated max must be >= it.
        # (2-4-5-800 / 2-8-9-400-SP have no single-tier mode → None; skip those.)
        for arch in MAX_SUPPORTED_SU:
            st = max_single_tier_su(arch)
            if st is not None:
                assert max_supported_su(arch) >= st

    def test_unknown_arch_returns_none(self):
        assert max_supported_su("9-9-9-999") is None


class TestGetTier:
    @pytest.mark.parametrize(
        "arch, su, expect_oob",
        [
            ("2-4-3-200", 1, 2),
            ("2-4-3-200", 4, 2),
            ("2-4-3-200", 5, 3),
            ("2-4-3-200", 6, 3),
            ("2-4-3-200", 7, 4),
            ("2-4-3-200", 8, 4),
            ("2-8-5-200", 1, 2),
            ("2-8-5-200", 4, 2),
            ("2-8-5-200", 5, 3),
            ("2-8-5-200", 6, 3),
            ("2-8-5-200", 7, 4),
            ("2-8-5-200", 8, 4),
            ("2-8-9-400", 1, 2),
            ("2-8-9-400", 2, 2),
            ("2-8-9-400", 3, 3),
            ("2-8-9-800", 1, 2),
            ("2-8-9-800", 4, 2),
        ],
    )
    def test_tier_oob_switches(self, arch, su, expect_oob):
        tier = get_tier(arch, su)
        assert tier is not None, f"{arch} su={su} should have a tier"
        assert tier.oob_switches == expect_oob

    def test_over_max_returns_none(self):
        assert get_tier("2-4-3-200", 9) is None
        assert get_tier("2-8-5-200", 9) is None
        assert get_tier("2-8-9-400", 4) is None
        assert get_tier("2-8-9-800", 5) is None

    def test_invalid_inputs_return_none(self):
        assert get_tier("2-4-3-200", 0) is None
        assert get_tier("2-4-3-200", -1) is None
        assert get_tier("bogus", 1) is None


class TestIsSupportedSingleTier:
    def test_supported_range(self):
        for su in range(1, 9):
            assert is_supported_single_tier("2-4-3-200", su)

    def test_unsupported_above_cap(self):
        assert not is_supported_single_tier("2-4-3-200", 9)
        assert not is_supported_single_tier("2-8-9-400", 4)

    def test_unsupported_below_one(self):
        assert not is_supported_single_tier("2-4-3-200", 0)


class TestScalingTableStructure:
    """Catch table-edit mistakes before they ship."""

    @pytest.mark.parametrize("arch", list(ARCH_SCALING.keys()))
    def test_tiers_ascending_contiguous(self, arch):
        tiers = ARCH_SCALING[arch]
        assert tiers, f"{arch} has no tiers"
        prev_max = 0
        for t in tiers:
            assert isinstance(t, ScalingTier)
            assert t.min_su == prev_max + 1, (
                f"{arch}: tier min_su={t.min_su} should be {prev_max + 1} "
                f"(no gaps, no overlaps)")
            assert t.max_su >= t.min_su
            assert t.core_or_csl_leaves == 2  # documented invariant
            assert t.oob_switches >= 1
            prev_max = t.max_su


# ---------------------------------------------------------------------------
# validate_excel: single-tier-cap + Air-OOB-single-cable
# ---------------------------------------------------------------------------

REPO_ROOT = Path(__file__).resolve().parent.parent
ALL_ARCHS = ["2-4-3-200", "2-8-5-200", "2-8-9-400", "2-8-9-800"]


class TestSingleTierValidator:
    def test_default_excels_under_cap_pass(self):
        # Each default Excel ships at su=2 (or smaller); should not warn
        # about over-cap. We assert the validator runs to completion.
        for arch in ALL_ARCHS:
            path = REPO_ROOT / "input" / arch / "default" / f"{arch}.xlsx"
            result = validate_excel(str(path))
            # Defaults must validate. (`ok` accepts warnings, only fails
            # on errors.)
            assert result.ok, f"{arch} default failed: {result.summary()}"

    def test_over_cap_emits_error(self, tmp_path):
        # Cook a fake Excel: copy a default, then flip enabled=Yes on a
        # node beyond the generator-supported model rows. Validator must
        # error even when the sheet itself is otherwise structurally valid.
        src = REPO_ROOT / "input" / "2-8-9-400" / "default" / "2-8-9-400.xlsx"
        dst = tmp_path / "over.xlsx"
        shutil.copyfile(src, dst)
        wb = openpyxl.load_workbook(dst)
        ws = wb["Nodes"]
        col = {}
        for c in range(1, ws.max_column + 1):
            h = (ws.cell(1, c).value or "").strip().lower()
            if h:
                col[h] = c
        # Synthesize an SU-24 row. The source docs have that row, but this
        # generator intentionally skips it until E/W super-spine templates
        # exist. The validator cares about *the highest* active SU index, not
        # whether the wire-map is consistent — so just shoehorn one in.
        # Function column is required (parser skips rows with blank function).
        target_row = ws.max_row + 1
        ws.cell(target_row, col["function"]).value = "server"
        ws.cell(target_row, col["name"]).value = "su-24-node-1"
        ws.cell(target_row, col["enabled"]).value = "Yes"
        # Give it a valid IP so the parser doesn't error on it.
        if "mgmt ip address" in col:
            ws.cell(target_row, col["mgmt ip address"]).value = "192.168.200.99"
        if "prefix" in col:
            ws.cell(target_row, col["prefix"]).value = 24
        if "gateway" in col:
            ws.cell(target_row, col["gateway"]).value = "192.168.200.1"
        wb.save(dst)
        wb.close()
        result = validate_excel(str(dst))
        joined = " ".join(result.errors)
        # Over-cap SU must be rejected. With the arch models present the
        # generator-supported check fires; without them (e.g. public release,
        # where scripts/models/ is excluded) the max-validated-SU check rejects
        # it (SU 24 > 16 for 2-8-9-400). Accept either — both are valid errors.
        assert ("is not generator-supported" in joined
                or "exceeds the maximum validated" in joined), (
            f"expected an over-cap SU error, got: {result.summary()}")


class TestAirOobSingleCableValidator:
    def test_two_display_yes_per_node_warns(self, tmp_path):
        src = REPO_ROOT / "input" / "2-4-3-200" / "default" / "2-4-3-200.xlsx"
        dst = tmp_path / "two-yes.xlsx"
        shutil.copyfile(src, dst)
        wb = openpyxl.load_workbook(dst)
        ws = wb["Wire Map"]
        swname, swport = _wm_cols(ws)
        # Find one SU-1 node's OOB rows and force two of them Display=Yes
        # with distinct ports on the same oob-switch.
        target = None
        oob_rows = []
        for r in range(2, ws.max_row + 1):
            sw = ws.cell(r, swname).value or ""
            sysname = ws.cell(r, 2).value or ""
            if "oob-switch" not in str(sw):
                continue
            if node_name_to_su(sysname) != 1:
                continue
            if target is None:
                target = sysname
            if sysname == target:
                oob_rows.append(r)
        # Mark two of the target node's OOB rows Display=Yes w/ different ports
        assert len(oob_rows) >= 2, (
            "default 2-4-3-200 must have ≥2 OOB rows for an SU-1 node")
        # Use high port numbers (SN2201 has 48 ports + 4 uplinks 49-52)
        # that the default doesn't claim, to avoid duplicate-port errors.
        used = set()
        for r in range(2, ws.max_row + 1):
            sw = ws.cell(r, swname).value or ""
            port = str(ws.cell(r, swport).value or "")
            if "oob-switch" in str(sw):
                used.add(port)
        free = [f"swp{p}" for p in range(40, 49) if f"swp{p}" not in used][:2]
        assert len(free) == 2, f"expected 2 free OOB ports, found {free}"
        ws.cell(oob_rows[0], 1).value = "Yes"
        ws.cell(oob_rows[0], swport).value = free[0]
        ws.cell(oob_rows[1], 1).value = "Yes"
        ws.cell(oob_rows[1], swport).value = free[1]
        wb.save(dst)
        wb.close()
        result = validate_excel(str(dst))
        joined = " ".join(result.warnings)
        assert "Air's plain Ubuntu can't bond" in joined, (
            f"expected Air-OOB-single-cable warning, got: "
            f"{result.summary()}")
