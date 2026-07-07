# SPDX-FileCopyrightText: Copyright (c) 2026 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: MIT
"""Gate: an OOB Uplink must land on an SN2201 QSFP28 uplink port (swp49-52),
never a copper host port (swp1-48).

The OOB switch template unconditionally makes swp1-48 L2 bridge access
ports at 1G, while the parser also emits the OOB-uplink port as an L3
unnumbered eBGP neighbor. When an uplink is placed on swp1-48 the port
gets BOTH treatments — a bridged 1G access port can't run unnumbered
eBGP, so the OOB<->CSL underlay (and the EVPN overlay riding over it)
never comes up. (Root-caused on the a live 2-8-9-800 site, where the
Wire Map put the uplinks on swp35/36/37.)

Uplink ports swp49-52 sit outside the hardcoded swp1-48 bridge range, so
they render as clean routed interfaces and BGP establishes — which is
exactly what the known-good 'default' site does.
"""
import sys
from pathlib import Path

import openpyxl

SCRIPTS_DIR = Path(__file__).parent.parent / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from validate_excel import validate_wire_map, ValidationResult


WM_HEADERS = [
    "Display in Air", "System Name (A)", "Port (A)", "Port Side (A)",
    "System Name (B)", "Port (B)", "Port Side (B)", "Cable Split",
    "Network Profile",
]


def _wm(rows):
    """Build an openpyxl Wire Map sheet from a list of row tuples."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Wire Map"
    for c, h in enumerate(WM_HEADERS, start=1):
        ws.cell(row=1, column=c, value=h)
    for r, row in enumerate(rows, start=2):
        for c, v in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=v)
    return ws


def _copper_errors(result):
    return [m for m in result.errors if "copper" in m.lower() and "uplink" in m.lower()]


def test_oob_uplink_on_copper_port_is_error():
    # Exact shape of the live-troubleshooting bug: csl-01 uplink lands on oob-switch
    # copper port swp36 (in the swp1-48 bridged/1G range).
    ws = _wm([
        ("Yes", "csl-01", "swp59s0", "NA", "oob-switch-01", "swp36", "NA", 1, "OOB Uplink"),
    ])
    result = ValidationResult()
    validate_wire_map(ws, result)
    flagged = _copper_errors(result)
    assert len(flagged) == 1, f"expected one copper-uplink error, got: {result.errors}"
    assert "oob-switch-01" in flagged[0] and "swp36" in flagged[0]


def test_oob_uplink_on_sfp28_uplink_port_is_ok():
    # swp49 is an SN2201 QSFP28 uplink port (outside swp1-48) — valid.
    ws = _wm([
        ("Yes", "csl-01", "swp59s0", "NA", "oob-switch-01", "swp49", "NA", 1, "OOB Uplink"),
        ("Yes", "csl-02", "swp59s0", "NA", "oob-switch-01", "swp51", "NA", 1, "OOB Uplink"),
    ])
    result = ValidationResult()
    validate_wire_map(ws, result)
    assert _copper_errors(result) == [], f"unexpected copper-uplink error: {result.errors}"


def test_server_bmc_on_copper_port_is_not_flagged():
    # Control: a server BMC/IPMI row legitimately uses a copper host port.
    # It is NOT an uplink, so the gate must stay quiet.
    ws = _wm([
        ("No", "su-01-node-01", "BMC", "NA", "oob-switch-01", "swp2", "NA", "NA", "OOB / IPMI"),
    ])
    result = ValidationResult()
    validate_wire_map(ws, result)
    assert _copper_errors(result) == [], f"BMC copper row wrongly flagged: {result.errors}"
