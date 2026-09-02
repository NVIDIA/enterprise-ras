# SPDX-FileCopyrightText: Copyright (c) 2026 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: MIT
"""
End-to-end pressure test for the mgmt-IP collision gate.

Unlike test_oob_reserved.py (which unit-tests the pure helpers and the
in-process validate function), this drives the REAL validate_excel.py as a
subprocess against mutated copies of a shipped default Excel — exercising the
full path operators hit via ``make validate-excel``, including the process
exit code that gates a deploy.

Covers the OOB-mode-aware reserved set (L2 vs L3), the air-mgmt-plane guard,
the custom-air-mgmt-subnet override, and the 2026-06-24 maxscale .78 repro.
"""
import shutil
import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
VALIDATE = REPO / "scripts" / "validate_excel.py"
# 2-8-9-400 ships oob_uplink_mode=L3 and is the arch the maxscale bug hit.
BASE_L3 = REPO / "input" / "2-8-9-400" / "default" / "2-8-9-400.xlsx"

# ERA-93: the message now names the DECLARED plane rather than the literal
# "flat OOB subnet", so match on the invariant half of the sentence.
OOB_ERR = "claimed by multiple owners"
AIR_ERR = "air-mgmt subnet"

pytestmark = pytest.mark.skipif(
    not BASE_L3.exists(), reason="default 2-8-9-400 Excel not present"
)


def _norm(s):
    return str(s).strip().lower().replace(" ", "_").replace("-", "_")


def _set_setting(ws, key, value):
    """Set a Settings key, modifying an existing row or inserting among the
    real settings (row 2). Appending would land inside the trailing VERSIONS
    sub-section, which validate_settings skips."""
    for row in ws.iter_rows():
        if row[0].value and _norm(row[0].value) == key:
            ws.cell(row[0].row, 2).value = value
            return
    ws.insert_rows(2)
    ws.cell(2, 1).value = key
    ws.cell(2, 2).value = value


def _set_air_mgmt_subnet(wb, value):
    """Set the air-mgmt subnet where it actually lives — the Air_Only sheet's
    "Air Management Subnet" row (what excel_parser/deploy read), NOT Settings."""
    ws = wb["Air_Only"]
    for row in ws.iter_rows():
        if row[0].value and str(row[0].value).strip().lower() == "air management subnet":
            ws.cell(row[0].row, 2).value = value
            return
    ws.append(["Air Management Subnet", value])


def _server_rows(ws, n):
    hdr = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    fcol = hdr.index("Function") + 1
    ipcol = hdr.index("Mgmt IP Address") + 1
    rows = []
    for r in range(2, ws.max_row + 1):
        f = ws.cell(r, fcol).value
        ip = ws.cell(r, ipcol).value
        if f and ip and str(f).strip().lower() in ("storage", "gpu", "support", "compute"):
            rows.append(r)
            if len(rows) >= n:
                break
    return rows, ipcol


def _set_oob_vlan_subnet(wb, subnet, gateway):
    """Move the OOB VLAN (VRF=OOB) onto a different subnet, and drag the
    oob-switch mgmt IPs along so only the thing under test changes."""
    import ipaddress
    net = ipaddress.IPv4Network(subnet)
    ws = wb["VLANs & Profiles"]
    # Row 1 is the "VLANs" banner; the real header is row 2.
    hdr = [str(ws.cell(2, c).value or "").strip() for c in range(1, ws.max_column + 1)]
    scol, gcol = hdr.index("Subnet") + 1, hdr.index("Gateway") + 1
    vcol = hdr.index("VRF") + 1
    for r in range(3, ws.max_row + 1):
        if str(ws.cell(r, vcol).value or "").strip().upper() == "OOB":
            ws.cell(r, scol).value = subnet
            ws.cell(r, gcol).value = gateway

    nodes = wb["Nodes"]
    nhdr = [nodes.cell(1, c).value for c in range(1, nodes.max_column + 1)]
    fcol = nhdr.index("Function") + 1
    ipcol = nhdr.index("Mgmt IP Address") + 1
    offset = 20
    for r in range(2, nodes.max_row + 1):
        fn = str(nodes.cell(r, fcol).value or "").strip().lower()
        ip = nodes.cell(r, ipcol).value
        if not ip:
            continue
        try:
            addr = ipaddress.IPv4Address(str(ip).strip().split("/")[0])
        except ValueError:
            continue
        if addr in ipaddress.IPv4Network("192.168.200.0/24"):
            nodes.cell(r, ipcol).value = str(net.network_address + offset)
            offset += 1


def _variant(tmp_path, name, *, mode=None, node_ips=None, air_mgmt_subnet=None,
             oob_vlan=None):
    dst = tmp_path / f"{name}.xlsx"
    shutil.copy2(BASE_L3, dst)
    wb = openpyxl.load_workbook(dst)
    if mode is not None:
        _set_setting(wb["Settings"], "oob_uplink_mode", mode)
    if air_mgmt_subnet is not None:
        _set_air_mgmt_subnet(wb, air_mgmt_subnet)
    if oob_vlan is not None:
        _set_oob_vlan_subnet(wb, *oob_vlan)
    if node_ips:
        ws = wb["Nodes"]
        rows, ipcol = _server_rows(ws, len(node_ips))
        for ip, r in zip(node_ips, rows):
            ws.cell(r, ipcol).value = ip
    wb.save(dst)
    return dst


def _run(path):
    p = subprocess.run([sys.executable, str(VALIDATE), str(path)],
                       capture_output=True, text=True)
    out = p.stdout + p.stderr
    return p.returncode, OOB_ERR in out, AIR_ERR in out


def test_l3_default_is_clean(tmp_path):
    rc, oob, air = _run(_variant(tmp_path, "l3_clean"))
    assert rc == 0 and not oob and not air


def test_l2_variant_no_false_positive(tmp_path):
    rc, oob, air = _run(_variant(tmp_path, "l2_clean", mode="L2"))
    assert not oob and not air


def test_l3_server_on_78_caught_and_blocks(tmp_path):
    # The 2026-06-24 maxscale repro: a server colliding with utility .78.
    rc, oob, air = _run(_variant(tmp_path, "l3_78", mode="L3",
                                 node_ips=["192.168.200.78"]))
    assert oob and rc == 1


def test_l2_server_on_78_not_flagged(tmp_path):
    # In L2 there is no utility; .78 is free — must NOT false-positive.
    rc, oob, air = _run(_variant(tmp_path, "l2_78", mode="L2",
                                 node_ips=["192.168.200.78"]))
    assert not oob


def test_dhcp_oob_252_flagged_in_l2(tmp_path):
    # .252 (dhcp-oob) is reserved in BOTH modes.
    rc, oob, air = _run(_variant(tmp_path, "l2_252", mode="L2",
                                 node_ips=["192.168.200.252"]))
    assert oob and rc == 1


def test_two_servers_same_oob_ip_blocks(tmp_path):
    rc, oob, air = _run(_variant(tmp_path, "dup",
                                 node_ips=["192.168.200.60", "192.168.200.60"]))
    assert oob and rc == 1


def test_server_in_air_mgmt_subnet_blocks(tmp_path):
    rc, oob, air = _run(_variant(tmp_path, "airmgmt",
                                 node_ips=["172.20.0.50"]))
    assert air and rc == 1


def test_custom_air_mgmt_subnet_honored(tmp_path):
    # Node inside the operator's custom air-mgmt subnet is flagged...
    rc, oob, air = _run(_variant(tmp_path, "custom_hit",
                                 air_mgmt_subnet="10.50.0.0/24",
                                 node_ips=["10.50.0.9"]))
    assert air and rc == 1


def test_custom_air_mgmt_subnet_not_over_applied(tmp_path):
    # ...and the OLD default 172.20.0.x is no longer treated as air-mgmt.
    rc, oob, air = _run(_variant(tmp_path, "custom_miss",
                                 air_mgmt_subnet="10.50.0.0/24",
                                 node_ips=["172.20.0.9"]))
    assert not air


def _run_out(path):
    p = subprocess.run([sys.executable, str(VALIDATE), str(path)],
                       capture_output=True, text=True)
    return p.returncode, p.stdout + p.stderr


def test_default_air_mgmt_subnet_does_not_overlap(tmp_path):
    # 172.20.0.0/24 (Air_Only) vs 192.168.200.0/24 (OOB VLAN subnet) are disjoint.
    rc, out = _run_out(_variant(tmp_path, "no_overlap"))
    assert "overlaps with" not in out


def test_air_mgmt_subnet_overlapping_mgmt_subnets_errors(tmp_path):
    # Set the Air_Only subnet to collide with the OOB VLAN subnet (S9 check,
    # now driven by the Air_Only-sourced value vs. resolve_oob_vlans).
    rc, out = _run_out(_variant(tmp_path, "overlap",
                                air_mgmt_subnet="192.168.200.0/24"))
    assert "overlaps with" in out and rc == 1


def test_brownfield_oob_vlan_duplicate_is_caught(tmp_path):
    """ERA-93: the gate must see a duplicate on the OOB VLAN the workbook
    DECLARES, not only on 192.168.200.0/24.

    Before the fix `oob_octet()` returned None for every address outside
    192.168.200.0/24, so this workbook produced zero claims and passed while
    inspecting nothing — the same blindness that let ERA-92 (10.78.220.34
    claimed twice, support data plane dead) through.
    """
    dup = "10.78.220.141"
    rc, oob, air = _run(_variant(tmp_path, "brownfield_dup",
                                 oob_vlan=("10.78.220.128/25", "10.78.220.129"),
                                 node_ips=[dup, dup]))
    assert oob, "duplicate on a declared brownfield OOB VLAN was not reported"
    assert rc == 1


def test_brownfield_oob_vlan_clean_has_no_false_positive(tmp_path):
    rc, oob, air = _run(_variant(tmp_path, "brownfield_clean",
                                 oob_vlan=("10.78.220.128/25", "10.78.220.129"),
                                 node_ips=["10.78.220.141", "10.78.220.142"]))
    assert not oob
