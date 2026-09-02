# SPDX-FileCopyrightText: Copyright (c) 2026 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: MIT
"""
Unit tests for prefix list classification — validates that all generated
prefix lists are correctly classified as overridable (global-subnet) or
derived-only (per-switch).
"""
import sys
from pathlib import Path

import pytest

SCRIPTS_DIR = Path(__file__).parent.parent / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from excel_parser import (
    generate_prefix_lists,
    parse_prefix_lists_sheet,
    OVERRIDABLE_PREFIX_LISTS,
    DERIVED_ONLY_PREFIX_LISTS,
)


class TestPrefixListClassification:
    def test_prefix_list_classification_covers_all_generated(self):
        """Verify that all generated prefix lists are classified as exactly one
        of overridable or derived-only."""
        vlans = [
            {
                "id": 300,
                "name": "in-band",
                "subnet": "172.16.178.0/24",
                "vrf": "INBAND",
            },
            {
                "id": 200,
                "name": "oob",
                "subnet": "192.168.200.0/24",
                "vrf": "OOB",
            },
        ]
        prefix_lists = generate_prefix_lists(
            vlans, core_num=1, loopback_base="172.16.176"
        )
        ids = {p["id"] for p in prefix_lists}

        # every generated list is classified in exactly one of the two sets
        assert ids <= (
            OVERRIDABLE_PREFIX_LISTS | DERIVED_ONLY_PREFIX_LISTS
        ), f"Unclassified lists: {ids - (OVERRIDABLE_PREFIX_LISTS | DERIVED_ONLY_PREFIX_LISTS)}"

        # the two sets do not overlap
        assert not (
            OVERRIDABLE_PREFIX_LISTS & DERIVED_ONLY_PREFIX_LISTS
        ), "Sets overlap (should be disjoint)"

    def test_overridable_prefix_lists_includes_era_prefixes(self):
        """Verify ERA_PREFIXES is classified as overridable."""
        assert "ERA_PREFIXES" in OVERRIDABLE_PREFIX_LISTS

    def test_overridable_prefix_lists_includes_global_subnets(self):
        """Verify global-subnet lists are overridable."""
        assert OVERRIDABLE_PREFIX_LISTS >= {
            "ERA_PREFIXES",
            "INBAND_PREFIXES",
            "OOB_PREFIXES",
            "VTEP_PREFIXES",
            "ALL_PREFIXES",
        }

    def test_derived_only_prefix_lists_includes_per_switch_loopbacks(self):
        """Verify per-switch /32 lists are derived-only."""
        assert DERIVED_ONLY_PREFIX_LISTS >= {
            "EXIT_LOCAL_IF",
            "INBAND_LOCAL_IF",
            "OOB_LOCAL_IF",
            "LOCAL_OOB_LOOPBACK",
        }

    def test_sets_are_disjoint(self):
        """Verify the two classification sets don't overlap."""
        assert len(OVERRIDABLE_PREFIX_LISTS & DERIVED_ONLY_PREFIX_LISTS) == 0


class TestParsePrefixListsSheet:
    """'Prefix lists' sheet parser supports Action
    directives (blank/override/add vs. suppress). Locked interface:
    parse_prefix_lists_sheet(ws) -> {'lists': {list_id: [rule, ...]},
    'suppress': set()}. The parser itself is list-set-agnostic — it does
    not decide override-vs-add; that's generate_prefix_lists' job (Task 3),
    based on whether the id already exists.
    """

    def test_parse_prefix_lists_directives(self):
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["List name", "Rule id", "Match", "Max prefix length", "Action"])
        ws.append(["ERA_PREFIXES", "10", "10.0.0.0/8", "24", ""])       # override-style (blank Action)
        ws.append(["CUSTOM_X", "10", "192.0.2.0/24", "32", ""])         # new list (blank Action)
        ws.append(["VTEP_PREFIXES", None, None, None, "suppress"])      # suppress

        d = parse_prefix_lists_sheet(ws)

        assert d["lists"]["ERA_PREFIXES"] == [
            {"id": "10", "match": "10.0.0.0/8", "max_len": "24"}
        ]
        assert d["lists"]["CUSTOM_X"] == [
            {"id": "10", "match": "192.0.2.0/24", "max_len": "32"}
        ]
        assert "VTEP_PREFIXES" in d["suppress"]
        assert "VTEP_PREFIXES" not in d["lists"]

    def test_parse_prefix_lists_suppress_wins_over_rule_rows(self):
        """A suppressed list_id must not also appear in lists, even if it
        also has rule rows elsewhere on the sheet."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["List name", "Rule id", "Match", "Max prefix length", "Action"])
        ws.append(["VTEP_PREFIXES", "1", "10.1.0.0/16", "32", ""])
        ws.append(["VTEP_PREFIXES", None, None, None, "suppress"])

        d = parse_prefix_lists_sheet(ws)

        assert "VTEP_PREFIXES" in d["suppress"]
        assert "VTEP_PREFIXES" not in d["lists"]

    def test_parse_prefix_lists_backward_compat_no_action_column(self):
        """A sheet with the old 4-column layout (no Action column) still
        works: every list_id lands in 'lists', 'suppress' stays empty."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["List name", "Rule id", "Match", "Max prefix length"])
        ws.append(["ERA_PREFIXES", "10", "10.0.0.0/8", "24"])
        ws.append(["CUSTOM_X", "10", "192.0.2.0/24", "32"])

        d = parse_prefix_lists_sheet(ws)

        assert d["lists"]["ERA_PREFIXES"] == [
            {"id": "10", "match": "10.0.0.0/8", "max_len": "24"}
        ]
        assert d["lists"]["CUSTOM_X"] == [
            {"id": "10", "match": "192.0.2.0/24", "max_len": "32"}
        ]
        assert d["suppress"] == set()


class TestApplyPrefixListDirectives:
    """generate_prefix_lists(..., prefix_list_directives=...)
    applies the full {'lists': {...}, 'suppress': set()} directive dict,
    scoped to OVERRIDABLE_PREFIX_LISTS. Directives naming a
    DERIVED_ONLY_PREFIX_LISTS id are ignored, with a warning — those are
    per-switch /32s and must never be user-overridable.
    """

    VLANS = [
        {"id": 300, "name": "in-band", "subnet": "172.16.178.0/24", "vrf": "INBAND"},
        {"id": 200, "name": "oob", "subnet": "192.168.200.0/24", "vrf": "OOB"},
    ]

    def _rule_map(self, prefix_lists):
        # `description` is stripped: ADR-0043 attaches one to every rule, and
        # these tests are about which MATCHES survive an override, not wording.
        return {pl["id"]: [{k: v for k, v in r.items() if k != "description"}
                           for r in pl["rule"]]
                for pl in prefix_lists}

    def test_override_changes_era_prefixes_rules(self):
        directives = {
            "lists": {
                "ERA_PREFIXES": [{"id": "10", "match": "10.0.0.0/8", "max_len": "24"}]
            },
            "suppress": set(),
        }
        prefix_lists = generate_prefix_lists(
            self.VLANS, core_num=1, loopback_base="172.16.176",
            prefix_list_directives=directives,
        )
        rules = self._rule_map(prefix_lists)
        assert rules["ERA_PREFIXES"] == [
            {"id": "10", "match": "10.0.0.0/8", "max_len": "24"}
        ]

    def test_add_yields_new_custom_list(self):
        directives = {
            "lists": {
                "CUSTOM_X": [{"id": "10", "match": "192.0.2.0/24", "max_len": "32"}]
            },
            "suppress": set(),
        }
        prefix_lists = generate_prefix_lists(
            self.VLANS, core_num=1, loopback_base="172.16.176",
            prefix_list_directives=directives,
        )
        rules = self._rule_map(prefix_lists)
        assert rules["CUSTOM_X"] == [
            {"id": "10", "match": "192.0.2.0/24", "max_len": "32"}
        ]

    def test_suppress_removes_vtep_prefixes(self):
        directives = {"lists": {}, "suppress": {"VTEP_PREFIXES"}}
        prefix_lists = generate_prefix_lists(
            self.VLANS, core_num=1, loopback_base="172.16.176",
            prefix_list_directives=directives,
        )
        ids = {pl["id"] for pl in prefix_lists}
        assert "VTEP_PREFIXES" not in ids

    def test_derived_only_override_is_ignored_with_warning(self, capsys):
        original = generate_prefix_lists(
            self.VLANS, core_num=1, loopback_base="172.16.176",
        )
        original_rules = self._rule_map(original)["OOB_LOCAL_IF"]

        directives = {
            "lists": {
                "OOB_LOCAL_IF": [{"id": "99", "match": "203.0.113.99/32", "max_len": "32"}]
            },
            "suppress": set(),
        }
        prefix_lists = generate_prefix_lists(
            self.VLANS, core_num=1, loopback_base="172.16.176",
            prefix_list_directives=directives,
        )
        rules = self._rule_map(prefix_lists)
        assert rules["OOB_LOCAL_IF"] == original_rules

        captured = capsys.readouterr()
        assert "⚠️" in (captured.out + captured.err)
        assert "OOB_LOCAL_IF" in (captured.out + captured.err)

    def test_derived_only_suppress_is_ignored_with_warning(self, capsys):
        directives = {"lists": {}, "suppress": {"OOB_LOCAL_IF"}}
        prefix_lists = generate_prefix_lists(
            self.VLANS, core_num=1, loopback_base="172.16.176",
            prefix_list_directives=directives,
        )
        ids = {pl["id"] for pl in prefix_lists}
        assert "OOB_LOCAL_IF" in ids

        captured = capsys.readouterr()
        assert "⚠️" in (captured.out + captured.err)
        assert "OOB_LOCAL_IF" in (captured.out + captured.err)

    def test_no_directives_matches_directives_none(self):
        """Passing prefix_list_directives=None is equivalent to omitting it
        entirely (derive-by-default)."""
        default_output = generate_prefix_lists(
            self.VLANS, core_num=1, loopback_base="172.16.176",
        )
        none_output = generate_prefix_lists(
            self.VLANS, core_num=1, loopback_base="172.16.176",
            prefix_list_directives=None,
        )
        assert default_output == none_output
