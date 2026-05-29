# SPDX-FileCopyrightText: Copyright (c) 2026 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: MIT
"""Validator coverage for the 8x-breakout-odd-port convention.

Added 2026-05-28 with feedback item #7:
  - 8x breakout must land on an ODD base port (convention)
  - The adjacent (base+1) port must not be live-cabled (its lanes
    are consumed by the 8x breakout)

Detection: any sub-port index >= 4 (`swpNs4`..`swpNs7`) implies the
base port is 8x-broken-out (4x breakout exposes s0-s3 only).
"""
import sys
from pathlib import Path

import openpyxl
import pytest

SCRIPTS_DIR = Path(__file__).parent.parent / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from validate_excel import validate_8x_breakout_odd_ports, ValidationResult


WM_HEADERS = [
    "Display in Air", "System Name (A)", "Port (A)", "Port Side (A)",
    "Cable Split (A)", "System Name (B)", "Port (B)", "Port Side (B)",
    "Cable Split (B)", "Network Profile",
]


def _wm(rows):
    """Build a Wire Map sheet from a list of row tuples (10 columns)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Wire Map"
    for c, h in enumerate(WM_HEADERS, start=1):
        ws.cell(row=1, column=c, value=h)
    for r, row in enumerate(rows, start=2):
        for c, v in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=v)
    return ws


def _run(ws):
    result = ValidationResult()
    validate_8x_breakout_odd_ports(ws, result)
    return result


def _errors_matching(result, needle):
    return [e for e in result.errors if needle in e]


# ─── Even-port 8x breakout is flagged ────────────────────────────────────

def test_8x_on_even_base_port_errors():
    """swp62 with 8x breakout (s7 visible) should error — convention is odd."""
    ws = _wm([
        ("Yes", "core-01", "swp62s0", "NA", "NA", "host-01", "eth1", "NA",
         "NA", "Storage Uplink"),
        ("Yes", "core-01", "swp62s7", "NA", "NA", "host-02", "eth1", "NA",
         "NA", "Storage Uplink"),
    ])
    result = _run(ws)
    assert _errors_matching(result, "swp62"), result.errors
    assert _errors_matching(result, "EVEN"), result.errors


def test_8x_on_odd_base_port_passes():
    """swp63 with 8x breakout — convention-correct."""
    ws = _wm([
        ("Yes", "core-01", "swp63s0", "NA", "NA", "host-01", "eth1", "NA",
         "NA", "Storage Uplink"),
        ("Yes", "core-01", "swp63s7", "NA", "NA", "host-02", "eth1", "NA",
         "NA", "Storage Uplink"),
    ])
    result = _run(ws)
    assert not _errors_matching(result, "swp63"), result.errors


# ─── Adjacent-port collision (base+1 must not be live) ───────────────────

def test_8x_with_adjacent_port_live_errors():
    """8x on swp63 + live cabling on swp64 → conflict."""
    ws = _wm([
        ("Yes", "core-01", "swp63s0", "NA", "NA", "host-01", "eth1", "NA",
         "NA", "Storage Uplink"),
        ("Yes", "core-01", "swp63s7", "NA", "NA", "host-02", "eth1", "NA",
         "NA", "Storage Uplink"),
        # Live row on swp64 (adjacent EVEN port consumed by swp63's 8x)
        ("Yes", "core-01", "swp64", "NA", "NA", "host-03", "eth1", "NA",
         "NA", "Edge Uplink"),
    ])
    result = _run(ws)
    assert _errors_matching(result, "swp64"), result.errors
    assert _errors_matching(result, "consuming the lanes"), result.errors


def test_8x_with_adjacent_disabled_passes():
    """8x on swp63 + swp64 marked Disabled by Neighbor → allowed."""
    ws = _wm([
        ("Yes", "core-01", "swp63s0", "NA", "NA", "host-01", "eth1", "NA",
         "NA", "Storage Uplink"),
        ("Yes", "core-01", "swp63s7", "NA", "NA", "host-02", "eth1", "NA",
         "NA", "Storage Uplink"),
        # Documented as disabled — OK
        ("No", "core-01", "swp64", "NA", "NA", None, None, "NA",
         "NA", "Port Disabled by Neighbor"),
    ])
    result = _run(ws)
    assert not _errors_matching(result, "swp"), result.errors


def test_8x_with_adjacent_absent_passes():
    """8x on swp63 + swp64 not in Wire Map at all → allowed (implicit disable).
    The defaults follow this pattern — adjacent even ports aren't listed."""
    ws = _wm([
        ("Yes", "core-01", "swp63s0", "NA", "NA", "host-01", "eth1", "NA",
         "NA", "Storage Uplink"),
        ("Yes", "core-01", "swp63s7", "NA", "NA", "host-02", "eth1", "NA",
         "NA", "Storage Uplink"),
        # No swp64 row anywhere
    ])
    result = _run(ws)
    assert not _errors_matching(result, "swp"), result.errors


# ─── Lower breakout factors not flagged ──────────────────────────────────

def test_4x_breakout_not_flagged():
    """4x breakout (s0-s3 only) doesn't trigger the 8x rule."""
    ws = _wm([
        ("Yes", "core-01", "swp63s0", "NA", "NA", "host-01", "eth1", "NA",
         "NA", "Edge Uplink"),
        ("Yes", "core-01", "swp63s3", "NA", "NA", "host-02", "eth1", "NA",
         "NA", "Edge Uplink"),
    ])
    result = _run(ws)
    # No sub-port >= 4 — not 8x — should pass
    assert not _errors_matching(result, "swp63"), result.errors


def test_2x_breakout_not_flagged():
    ws = _wm([
        ("Yes", "core-01", "swp62s0", "NA", "NA", "host-01", "eth1", "NA",
         "NA", "Edge Uplink"),
        ("Yes", "core-01", "swp62s1", "NA", "NA", "host-02", "eth1", "NA",
         "NA", "Edge Uplink"),
    ])
    result = _run(ws)
    assert not _errors_matching(result, "swp62"), result.errors


def test_no_breakout_not_flagged():
    """Bare swpN (no sub-port) is not breakout, regardless of odd/even."""
    ws = _wm([
        ("Yes", "core-01", "swp64", "NA", "NA", "host-01", "eth1", "NA",
         "NA", "Edge Uplink"),
    ])
    result = _run(ws)
    assert not _errors_matching(result, "swp64"), result.errors


# ─── Multi-port + multi-switch scenarios ─────────────────────────────────

def test_8x_on_b_side_also_detected():
    """B-side ports must also be checked — e.g. core-to-core ISL rows
    where the 8x breakout shows up in the B columns."""
    ws = _wm([
        # System Name (A) is the host; System Name (B) is the switch
        ("Yes", "host-01", "eth1", "NA", "NA", "core-01", "swp62s0", "NA",
         "NA", "Storage Uplink"),
        ("Yes", "host-02", "eth1", "NA", "NA", "core-01", "swp62s7", "NA",
         "NA", "Storage Uplink"),
    ])
    result = _run(ws)
    # core-01:swp62 is even — should error
    assert _errors_matching(result, "swp62"), result.errors


def test_multiple_8x_ports_each_validated():
    """Two 8x breakouts on the same switch — both checked, only the
    even one errors."""
    ws = _wm([
        # swp63 8x — odd, OK
        ("Yes", "core-01", "swp63s0", "NA", "NA", "h1", "eth1", "NA",
         "NA", "Storage Uplink"),
        ("Yes", "core-01", "swp63s7", "NA", "NA", "h2", "eth1", "NA",
         "NA", "Storage Uplink"),
        # swp62 8x — even, ERRORS
        ("Yes", "core-01", "swp62s0", "NA", "NA", "h3", "eth1", "NA",
         "NA", "Storage Uplink"),
        ("Yes", "core-01", "swp62s7", "NA", "NA", "h4", "eth1", "NA",
         "NA", "Storage Uplink"),
    ])
    result = _run(ws)
    # Match on "core-01: swp62 has" / "core-01: swp63 has" — the
    # error-message preamble. Avoids matching the remediation hint
    # which mentions adjacent odd port names.
    even_errs = _errors_matching(result, "swp62 has 8x")
    odd_errs = _errors_matching(result, "swp63 has 8x")
    assert len(even_errs) == 1, f"expected exactly 1 error on swp62, got: {result.errors}"
    assert len(odd_errs) == 0, f"swp63 should pass, got: {odd_errs}"


def test_8x_on_different_switches_isolated():
    """8x on swp63 of core-01 doesn't conflict with swp64 on core-02
    — adjacent-port check is scoped per switch."""
    ws = _wm([
        ("Yes", "core-01", "swp63s7", "NA", "NA", "h1", "eth1", "NA",
         "NA", "Storage Uplink"),
        # core-02:swp64 is live, but core-01:swp64 is the one that
        # should be flagged if anything. core-02:swp64 has no 8x next
        # door, so it's fine.
        ("Yes", "core-02", "swp64", "NA", "NA", "h2", "eth1", "NA",
         "NA", "Edge Uplink"),
    ])
    result = _run(ws)
    assert not _errors_matching(result, "swp"), result.errors
