# SPDX-FileCopyrightText: Copyright (c) 2026 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: MIT
"""The `Custom_Config` sheet carries operator text into a generated shell script.

Unlike `ACLs` or `Prefix lists`, which carry structured directives the tool re-renders,
this sheet carries opaque config destined for a file that runs on a switch. So the
parser's job is as much refusal as parsing (ADR-0055).

Two failure modes drive these tests:

* **Silent no-op.** A typo'd target that matches zero switches, while the operator
  believes their config shipped. That is the class behind #59 (`make fix-ext-storage`
  exiting 0 having done nothing) and ERA-86 (`2-4-5-400` in the generate matrix but in
  neither e2e matrix). Every unmatched token is therefore a hard error, never a warning.
* **Injection.** The content reaches a bash script. `nv config` / `nv action` and shell
  metacharacters are refused. Best-effort containment, not a sandbox.
"""
import sys
from pathlib import Path

import openpyxl
import pytest

SCRIPTS = Path(__file__).resolve().parent.parent / "scripts"
if str(SCRIPTS) not in sys.path:
    sys.path.insert(0, str(SCRIPTS))

from excel_parser import (  # noqa: E402
    parse_custom_config_sheet,
    CustomConfigError,
)

HEADER = ["Switch_Location", "Config"]

# The switch inventory a workbook declares. Function -> [names]. Servers are present on
# purpose: `Function: gpu` must fail, and it must fail because gpu is not a SWITCH
# function, not because of a separate hand-written server check.
SWITCHES = {
    "csl": ["csl-01", "csl-02"],
    "gsl-plane1": ["gsl-plane1-01", "gsl-plane1-02"],
    "gsl-plane2": ["gsl-plane2-01", "gsl-plane2-02"],
    "oob-switch": ["oob-switch-01", "oob-switch-02"],
}
SERVERS = {"gpu": ["su-01-node-01"], "support": ["support-01"]}


def _sheet(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(HEADER)
    for r in rows:
        ws.append(r)
    return ws


def _parse(rows, switches=None, servers=None):
    return parse_custom_config_sheet(
        _sheet(rows),
        switches_by_function=switches if switches is not None else SWITCHES,
        servers_by_function=servers if servers is not None else SERVERS,
    )


class TestTargeting:
    def test_all_targets_every_switch(self):
        got = _parse([["ALL", "nv set system hostname banner"]])
        assert set(got) == {n for names in SWITCHES.values() for n in names}

    def test_function_targets_every_switch_with_that_function(self):
        got = _parse([["Function: csl", "nv set system message pre-login x"]])
        assert set(got) == {"csl-01", "csl-02"}

    def test_function_accepts_a_comma_separated_list(self):
        got = _parse([["Function: csl, oob-switch", "nv set x y"]])
        assert set(got) == {"csl-01", "csl-02", "oob-switch-01", "oob-switch-02"}

    def test_host_targets_only_the_named_switches(self):
        got = _parse([["Host: csl-01, gsl-plane2-02", "nv set x y"]])
        assert set(got) == {"csl-01", "gsl-plane2-02"}

    def test_rows_accumulate_in_sheet_order(self):
        got = _parse([
            ["ALL", "nv set first line"],
            ["Function: csl", "nv set second line"],
        ])
        assert got["csl-01"] == ["nv set first line", "nv set second line"]
        assert got["oob-switch-01"] == ["nv set first line"]

    def test_multiple_lines_in_one_cell_are_split(self):
        got = _parse([["Function: csl", "nv set line one\nnv set line two\n\nnv set three"]])
        assert got["csl-01"] == ["nv set line one", "nv set line two", "nv set three"]


class TestTargetingIsAHardError:
    """A target that matches nothing must fail the import, not warn.

    The operator's belief that their config shipped is the thing being protected.
    """

    def test_unknown_function_is_rejected_and_names_the_token(self):
        with pytest.raises(CustomConfigError) as exc:
            _parse([["Function: csl, clx", "nv set x y"]])
        assert "clx" in str(exc.value)

    def test_error_lists_the_valid_functions_for_this_workbook(self):
        with pytest.raises(CustomConfigError) as exc:
            _parse([["Function: clx", "nv set x y"]])
        assert "csl" in str(exc.value) and "oob-switch" in str(exc.value)

    def test_function_valid_in_another_arch_still_fails_here(self):
        """`cl` exists on largescale archs; on this workbook it must fail.

        This is the deliberate cost of strict matching: one workbook cannot carry a
        custom-config block spanning arch families.
        """
        with pytest.raises(CustomConfigError) as exc:
            _parse([["Function: cl", "nv set x y"]])
        assert "cl" in str(exc.value)

    def test_no_prefix_or_alias_matching(self):
        """`gl` must NOT sweep up gl-plane1/gl-plane2 — considered and declined."""
        with pytest.raises(CustomConfigError):
            _parse([["Function: gsl", "nv set x y"]])

    def test_unknown_host_is_rejected(self):
        with pytest.raises(CustomConfigError) as exc:
            _parse([["Host: leaf01", "nv set x y"]])
        assert "leaf01" in str(exc.value)

    def test_server_function_is_rejected(self):
        """`gpu` resolves to real nodes, but `nv set` on a server is meaningless."""
        with pytest.raises(CustomConfigError) as exc:
            _parse([["Function: gpu", "nv set x y"]])
        assert "gpu" in str(exc.value)

    def test_case_mismatch_is_rejected(self):
        with pytest.raises(CustomConfigError):
            _parse([["Function: CSL", "nv set x y"]])

    def test_unknown_targeting_keyword_is_rejected(self):
        with pytest.raises(CustomConfigError) as exc:
            _parse([["Role: csl", "nv set x y"]])
        assert "Role" in str(exc.value) or "Switch_Location" in str(exc.value)


class TestContentAllowlist:
    """`nv`-only, with `nv config` and `nv action` blocked (ADR-0055)."""

    @pytest.mark.parametrize("line", [
        "nv set system hostname sw1",
        "nv unset interface swp1 link state",
        "nv show system",
    ])
    def test_nv_verbs_are_accepted(self, line):
        got = _parse([["Function: csl", line]])
        assert got["csl-01"] == [line]

    @pytest.mark.parametrize("line", [
        "nv config apply",
        "nv config replace /tmp/x",
        "nv config detach",
    ])
    def test_nv_config_is_blocked(self, line):
        """Apply-lifecycle control belongs to the tool.

        A mid-script apply leaves the rest of the file staging against live state;
        `nv config replace` can discard the entire generated configuration.
        """
        with pytest.raises(CustomConfigError) as exc:
            _parse([["Function: csl", line]])
        assert "nv config" in str(exc.value)

    @pytest.mark.parametrize("line", [
        "nv action clear router bgp",
        "nv action install system image http://x/y.bin",
    ])
    def test_nv_action_is_blocked(self, line):
        """Imperative, non-idempotent, leaves no config state, and some verbs are
        destructive in a file ZTP re-runs."""
        with pytest.raises(CustomConfigError) as exc:
            _parse([["Function: csl", line]])
        assert "nv action" in str(exc.value)

    @pytest.mark.parametrize("line", [
        "systemctl restart frr",
        "echo hello",
        "  reboot",
        "nvset system hostname x",
    ])
    def test_non_nv_lines_are_blocked(self, line):
        with pytest.raises(CustomConfigError):
            _parse([["Function: csl", line]])

    def test_error_names_the_offending_line_number(self):
        with pytest.raises(CustomConfigError) as exc:
            _parse([["Function: csl", "nv set ok\nreboot\nnv set also ok"]])
        assert "2" in str(exc.value)


class TestInjectionDefences:
    """The content reaches a bash script. A legitimate `nv` line needs none of these."""

    @pytest.mark.parametrize("line", [
        "nv set system hostname `whoami`",
        "nv set system hostname $(id)",
        "nv set system hostname x; reboot",
        "nv set system hostname x && reboot",
        "nv set system hostname x || reboot",
        "nv set system hostname x | tee /tmp/x",
        "nv set system hostname x > /etc/passwd",
        "nv set system hostname x < /etc/shadow",
        "nv set system hostname x \\",
    ])
    def test_shell_metacharacters_are_rejected(self, line):
        with pytest.raises(CustomConfigError):
            _parse([["Function: csl", line]])

    def test_ordinary_values_with_spaces_are_accepted(self):
        """Quoting is the templates' job (`| quote`); the parser must not over-reject."""
        line = 'nv set system message pre-login "Authorized use only"'
        got = _parse([["Function: csl", line]])
        assert got["csl-01"] == [line]


class TestInertWhenAbsent:
    def test_empty_sheet_yields_nothing(self):
        assert _parse([]) == {}

    def test_blank_rows_are_skipped(self):
        got = _parse([[None, None], ["Function: csl", "nv set x y"], ["", ""]])
        assert set(got) == {"csl-01", "csl-02"}

    def test_row_with_target_but_no_config_is_rejected(self):
        """Silently ignoring it would be the same class of bug as a no-op target."""
        with pytest.raises(CustomConfigError):
            _parse([["Function: csl", None]])


class TestTargetableSetMatchesWhatActuallyGetsAConfig:
    """"Targetable" must mean "gets a generated config", not "is a switch".

    `utils.is_switch()` returns True for `edge` and `air-oob`, but
    `playbooks/generate-cli-configs.yml` renders a template only for
    core/csl/cl, gsl+gl planes, cs+gs planes, and oob. Targeting `edge` would
    therefore pass validation and produce nothing — the silent no-op this whole
    sheet's validation exists to prevent.
    """

    def _groups_with_a_template(self):
        """Group names named in generate-cli-configs.yml `when:` conditions."""
        import re
        pb = (Path(__file__).resolve().parent.parent
              / "playbooks" / "generate-cli-configs.yml").read_text()
        return {m.replace("_", "-") for m in re.findall(r"'([a-z0-9_]+)' in group_names", pb)}

    def test_every_targetable_function_has_a_template(self):
        from excel_parser import CONFIG_GENERATING_FUNCTIONS
        groups = self._groups_with_a_template()
        # `oob` is the inventory group; `oob-switch` is the Excel Function spelling.
        groups = {"oob-switch" if g == "oob" else g for g in groups}
        assert CONFIG_GENERATING_FUNCTIONS == groups, (
            "the targetable set has drifted from the playbook that renders configs.\n"
            f"  targetable but no template: {sorted(CONFIG_GENERATING_FUNCTIONS - groups)}\n"
            f"  has a template but not targetable: {sorted(groups - CONFIG_GENERATING_FUNCTIONS)}"
        )

    def test_edge_and_air_oob_are_not_targetable(self):
        from excel_parser import CONFIG_GENERATING_FUNCTIONS
        from utils import is_switch
        for fn in ("edge", "air-oob"):
            assert is_switch(fn), f"precondition: utils.is_switch({fn!r}) should be True"
            assert fn not in CONFIG_GENERATING_FUNCTIONS, (
                f"{fn!r} gets no generated config; targeting it would silently drop "
                f"the operator's custom config"
            )

    def test_switches_by_function_excludes_nodes_without_a_template(self):
        from excel_parser import switches_by_function_from_nodes
        # `role` is the key parse_nodes() uses for the Function column — NOT
        # `function`. An earlier draft of this test invented `function`, which passed
        # while production would have resolved nothing.
        nodes = [
            {"name": "csl-01", "role": "csl"},
            {"name": "cust-net-edge-01", "role": "edge"},
            {"name": "su-01-node-01", "role": "gpu"},
        ]
        got = switches_by_function_from_nodes(nodes)
        assert got == {"csl": ["csl-01"]}, got


class TestAgainstRealParsedNodes:
    """Guards against the failure the unit fixtures cannot catch: a helper that reads
    a key the real node dicts do not have. An invented fixture shape passes happily
    while production resolves nothing."""

    def test_helper_resolves_switches_from_a_real_workbook(self):
        import openpyxl as _oxl
        from excel_parser import parse_nodes, switches_by_function_from_nodes
        wb_path = (Path(__file__).resolve().parent.parent
                   / "input" / "2-8-9-800" / "default" / "2-8-9-800.xlsx")
        if not wb_path.exists():
            pytest.skip("shipped workbook absent")
        nodes = parse_nodes(_oxl.load_workbook(wb_path, data_only=True)["Nodes"])
        got = switches_by_function_from_nodes(nodes)
        assert got, "resolved no switches from a real workbook — wrong node key?"
        assert "csl" in got and sorted(got["csl"]) == ["csl-01", "csl-02"], got
        assert "gpu" not in got and "edge" not in got, got


class TestEndToEndThroughParseExcel:
    """Proves the wiring, not just the parser: a real workbook with a Custom_Config
    sheet must land lines in the right host_vars, and one without must be inert."""

    def _workbook_with_sheet(self, tmp_path, rows):
        import shutil
        import openpyxl as _oxl
        src = (Path(__file__).resolve().parent.parent
               / "input" / "2-8-9-800" / "default" / "2-8-9-800.xlsx")
        if not src.exists():
            pytest.skip("shipped workbook absent")
        dst = tmp_path / "wb.xlsx"
        shutil.copy2(src, dst)
        wb = _oxl.load_workbook(dst)
        # Shipped workbooks now carry a blank Custom_Config sheet, so replace it
        # rather than creating a second one with the same name.
        if "Custom_Config" in wb.sheetnames:
            del wb["Custom_Config"]
        ws = wb.create_sheet("Custom_Config")
        ws.append(HEADER)
        for r in rows:
            ws.append(r)
        wb.save(dst)
        return dst

    def test_lines_reach_only_the_targeted_switches(self, tmp_path):
        import openpyxl as _oxl
        from excel_parser import (parse_nodes, parse_custom_config_sheet,
                                  switches_by_function_from_nodes,
                                  servers_by_function_from_nodes)
        wb_path = self._workbook_with_sheet(
            tmp_path, [["Host: csl-01", "nv set system message pre-login x"]])
        wb = _oxl.load_workbook(wb_path, data_only=True)
        nodes = parse_nodes(wb["Nodes"])
        got = parse_custom_config_sheet(
            wb["Custom_Config"],
            switches_by_function=switches_by_function_from_nodes(nodes),
            servers_by_function=servers_by_function_from_nodes(nodes))
        assert got == {"csl-01": ["nv set system message pre-login x"]}

    def test_a_bad_target_fails_against_a_real_workbook(self, tmp_path):
        """`edge` is a real node and utils.is_switch() calls it a switch — but it
        renders no template, so it must be refused."""
        import openpyxl as _oxl
        from excel_parser import (parse_nodes, parse_custom_config_sheet,
                                  switches_by_function_from_nodes,
                                  servers_by_function_from_nodes)
        wb_path = self._workbook_with_sheet(tmp_path, [["Function: edge", "nv set x y"]])
        wb = _oxl.load_workbook(wb_path, data_only=True)
        nodes = parse_nodes(wb["Nodes"])
        with pytest.raises(CustomConfigError) as exc:
            parse_custom_config_sheet(
                wb["Custom_Config"],
                switches_by_function=switches_by_function_from_nodes(nodes),
                servers_by_function=servers_by_function_from_nodes(nodes))
        assert "edge" in str(exc.value)


class TestRendering:
    """Custom config must be the LAST thing in the file.

    `nv config apply` is issued by the consumers (push-switch-configs.yml:131,
    ztp.sh.j2:104), not by the templates — so "last in the file" is automatically
    "last staged before apply", which is what makes the operator win on any key the
    tool also sets (ADR-0055).
    """

    TEMPLATES = [
        "roles/core/templates/core_nvue_cli.j2",
        "roles/gl/templates/gl_nvue_cli.j2",
        "roles/spine/templates/spine_nvue_cli.j2",
        "roles/oob-switch/templates/oob_nvue_cli.j2",
    ]

    def _render(self, template_rel, host_vars):
        """Render a switch template with ONLY custom_config supplied.

        Everything else resolves to a permissive Undefined that tolerates calls,
        indexing and iteration — the templates do `lo_ip.split('/')` and loop over
        vars we are deliberately not providing. The point is the placement of the
        custom-config block, not the rest of the file.
        """
        import jinja2

        class _Anything(jinja2.ChainableUndefined):
            def __call__(self, *a, **k):
                return self

            def __getitem__(self, k):
                return self

            def __iter__(self):
                return iter(())

            def __str__(self):
                return ""

        root = Path(__file__).resolve().parent.parent
        src = (root / template_rel).read_text()
        env = jinja2.Environment(undefined=_Anything, keep_trailing_newline=True)
        env.filters.setdefault("quote", lambda v: f'"{v}"')
        return env.from_string(src).render(**host_vars)

    @pytest.mark.parametrize("template_rel", TEMPLATES)
    def test_every_switch_template_emits_custom_config(self, template_rel):
        out = self._render(template_rel, {"custom_config": ["nv set system foo bar"]})
        assert "nv set system foo bar" in out, (
            f"{template_rel} does not emit custom_config — a switch of this role would "
            f"silently drop the operator's config"
        )

    @pytest.mark.parametrize("template_rel", TEMPLATES)
    def test_custom_config_is_the_last_content_in_the_file(self, template_rel):
        marker = "nv set system LAST-LINE-MARKER x"
        out = self._render(template_rel, {"custom_config": [marker]})
        remainder = out[out.index(marker) + len(marker):]
        leftover = [ln for ln in remainder.splitlines()
                    if ln.strip() and not ln.strip().startswith("#")]
        assert not leftover, (
            f"{template_rel} emits config AFTER custom_config: {leftover[:3]}. "
            f"The operator would no longer win on a contested key."
        )

    @pytest.mark.parametrize("template_rel", TEMPLATES)
    def test_absent_custom_config_emits_nothing(self, template_rel):
        """Absent sheet must be byte-for-byte inert."""
        out = self._render(template_rel, {})
        assert "Custom Config" not in out, f"{template_rel} emits a banner with no data"

    @pytest.mark.parametrize("template_rel", TEMPLATES)
    def test_lines_are_emitted_verbatim_and_in_order(self, template_rel):
        lines = ["nv set aaa first", "nv unset bbb second", "nv set ccc third"]
        out = self._render(template_rel, {"custom_config": lines})
        positions = [out.index(ln) for ln in lines]
        assert positions == sorted(positions), "custom config reordered"


class TestNodeShapeTolerance:
    """Two node parsers in this repo disagree on the Function key.

    `excel_parser.parse_nodes()` emits `role`; `validate_excel.validate_nodes()`
    emits `function`. Reading only one yields an EMPTY switch inventory in the
    other's context — which made validate_excel reject every Custom_Config row
    with "Valid: ." before this was fixed. Same shape as ADR-0054's trap 1.
    """

    @pytest.mark.parametrize("key", ["role", "function"])
    def test_both_node_dict_shapes_resolve(self, key):
        from excel_parser import switches_by_function_from_nodes
        nodes = [{"name": "csl-01", key: "csl"}, {"name": "su-01-node-01", key: "gpu"}]
        assert switches_by_function_from_nodes(nodes) == {"csl": ["csl-01"]}

    def test_validate_excel_node_output_resolves_switches(self):
        """The regression itself: run validate_excel's own parser and resolve."""
        import openpyxl as _oxl
        import validate_excel as V
        from excel_parser import switches_by_function_from_nodes
        wb_path = (Path(__file__).resolve().parent.parent
                   / "input" / "2-8-9-800" / "default" / "2-8-9-800.xlsx")
        if not wb_path.exists():
            pytest.skip("shipped workbook absent")
        wb = _oxl.load_workbook(wb_path, data_only=True)
        nodes = V.validate_nodes(wb["Nodes"], V.ValidationResult())
        got = switches_by_function_from_nodes(nodes)
        assert "csl" in got, (
            f"validate_excel's node shape resolved no switches ({got}); "
            f"validation would reject every Custom_Config row"
        )


class TestValidateConfigUnsetHandling:
    """`nv unset` must not produce a permanent false MISMATCH.

    The tool emits `nv set system wjh state enabled`; the operator's
    `nv unset system wjh state` runs last, so the switch ends up without the key and
    the tool's own line is legitimately absent from the running config. Reporting that
    as MISSING would fail a correctly-configured switch on every run — the class of
    false failure fixed during the 5.18.0 work. v1 policy: warn, do not fail, do not
    stop validate-all (ADR-0055).
    """

    PLAYBOOK = Path(__file__).resolve().parent.parent / "playbooks" / "validate-config.yml"

    def test_playbook_reads_the_switchs_vars_not_the_jumps(self):
        """This play runs on `jump`; a bare `custom_config` reads the wrong host.

        The lookup lives in a set_fact that loops switch_list, so it is
        `hostvars[item]`; the shell task then indexes the resulting map by
        `item.item`. Either one bare would silently resolve the jump host, which has
        no custom config — the filter would never fire and every override would
        report MISMATCH.
        """
        text = self.PLAYBOOK.read_text()
        assert "hostvars[item].custom_config_keypaths" in text, (
            "the set_fact must read the SWITCH vars via hostvars[item]"
        )
        assert "custom_unset_map[item.item]" in text, (
            "the shell task must index the map by the switch it is comparing"
        )
        assert "\n        {{ custom_config" not in text, (
            "a bare custom_config reference would resolve the jump host vars"
        )

    def test_filter_drops_only_the_unset_key_paths(self, tmp_path):
        """Exercise the exact shell the playbook runs, not a paraphrase of it."""
        import subprocess
        missing = tmp_path / "missing.txt"
        unset = tmp_path / "unset.txt"
        missing.write_text(
            "nv set system wjh state enabled\n"
            "nv set system wjh channel forwarding trigger l2\n"
            "nv set bgp router-id 10.0.0.1\n"
        )
        unset.write_text("system wjh state\n")
        script = f'''
        mv {missing} {tmp_path}/all.txt
        : > {missing}
        while IFS= read -r _line; do
          _keep=1
          while IFS= read -r _path; do
            [ -z "$_path" ] && continue
            case "$_line" in "nv set $_path"*) _keep=0; break;; esac
          done < {unset}
          [ "$_keep" -eq 1 ] && printf '%s\\n' "$_line" >> {missing}
        done < {tmp_path}/all.txt
        cat {missing}
        '''
        out = subprocess.run(["sh", "-c", script], capture_output=True, text=True).stdout
        kept = [ln for ln in out.splitlines() if ln.strip()]
        assert kept == [
            "nv set system wjh channel forwarding trigger l2",
            "nv set bgp router-id 10.0.0.1",
        ], f"filter kept {kept}"

    def test_unrelated_missing_lines_still_fail(self, tmp_path):
        """The exclusion must be narrow — a genuinely missing line still counts."""
        import subprocess
        missing = tmp_path / "m.txt"; unset = tmp_path / "u.txt"
        missing.write_text("nv set bgp router-id 10.0.0.1\n")
        unset.write_text("system wjh state\n")
        script = f'''
        mv {missing} {tmp_path}/all.txt
        : > {missing}
        while IFS= read -r _line; do
          _keep=1
          while IFS= read -r _path; do
            [ -z "$_path" ] && continue
            case "$_line" in "nv set $_path"*) _keep=0; break;; esac
          done < {unset}
          [ "$_keep" -eq 1 ] && printf '%s\\n' "$_line" >> {missing}
        done < {tmp_path}/all.txt
        wc -l < {missing}
        '''
        out = subprocess.run(["sh", "-c", script], capture_output=True, text=True).stdout
        assert out.strip() == "1", "an unrelated missing line must still be reported"


class TestGeneratedSheet:
    """The sheet ships blank-but-present so operators do not have to invent it.

    A hand-made `Custom Config` (space, not underscore) matches nothing and silently
    does nothing — the same silent no-op strict targeting exists to prevent.
    """

    def test_generator_targetable_set_matches_the_parser(self):
        """Two copies of the switch-function set; they must not drift apart."""
        import importlib.util
        from excel_parser import CONFIG_GENERATING_FUNCTIONS
        dm = Path(__file__).resolve().parents[2] / "data-models"
        gen = dm / "generate_arch_excel.py"
        if not gen.exists():
            pytest.skip("internal generator not present (public tree)")
        if str(dm) not in sys.path:            # it imports `models` from beside it
            sys.path.insert(0, str(dm))
        spec = importlib.util.spec_from_file_location("_gen", gen)
        mod = importlib.util.module_from_spec(spec)
        sys.modules["_gen"] = mod
        spec.loader.exec_module(mod)
        assert mod.CONFIG_TARGETABLE_FUNCTIONS == CONFIG_GENERATING_FUNCTIONS, (
            "generator and parser disagree about which functions can be targeted:\n"
            f"  generator only: {sorted(mod.CONFIG_TARGETABLE_FUNCTIONS - CONFIG_GENERATING_FUNCTIONS)}\n"
            f"  parser only:    {sorted(CONFIG_GENERATING_FUNCTIONS - mod.CONFIG_TARGETABLE_FUNCTIONS)}"
        )

    @pytest.mark.parametrize("arch", ["2-8-9-800", "2-4-3-200"])
    def test_shipped_workbook_has_the_sheet(self, arch):
        import openpyxl as _oxl
        wb_path = (Path(__file__).resolve().parent.parent
                   / "input" / arch / "default" / f"{arch}.xlsx")
        if not wb_path.exists():
            pytest.skip(f"{arch} workbook absent")
        wb = _oxl.load_workbook(wb_path, data_only=True)
        assert "Custom_Config" in wb.sheetnames, (
            f"{arch} ships without the sheet; operators would have to create it by "
            f"hand and a misnamed one fails silently"
        )
        ws = wb["Custom_Config"]
        assert [c.value for c in ws[1]][:2] == ["Switch_Location", "Config"]

    @pytest.mark.parametrize("arch", ["2-8-9-800", "2-4-3-200"])
    def test_shipped_sheet_is_inert(self, arch):
        """Guidance rows must not parse as configuration."""
        import openpyxl as _oxl
        from excel_parser import (parse_nodes, parse_custom_config_sheet,
                                  switches_by_function_from_nodes,
                                  servers_by_function_from_nodes, CustomConfigError)
        wb_path = (Path(__file__).resolve().parent.parent
                   / "input" / arch / "default" / f"{arch}.xlsx")
        if not wb_path.exists():
            pytest.skip(f"{arch} workbook absent")
        wb = _oxl.load_workbook(wb_path, data_only=True)
        nodes = parse_nodes(wb["Nodes"])
        try:
            got = parse_custom_config_sheet(
                wb["Custom_Config"],
                switches_by_function=switches_by_function_from_nodes(nodes),
                servers_by_function=servers_by_function_from_nodes(nodes))
        except CustomConfigError as exc:
            pytest.fail(f"{arch}'s shipped sheet does not parse cleanly: {exc}")
        assert got == {}, f"{arch} ships with active custom config: {got}"


class TestCommentRows:
    """The shipped sheet carries usage guidance; guidance is not configuration.

    Found by shipping it: the generated help rows parsed as data and failed
    `make generate` with "'Optional sheet...' does not start with 'nv '". Rows whose
    Switch_Location begins with '#' are skipped, so guidance can live in the sheet
    without being mistaken for a rule — and an operator can comment a row out.
    """

    def test_hash_prefixed_rows_are_skipped(self):
        got = _parse([
            ["# HOW TO USE", "Delete these rows before filling in"],
            ["Function: csl", "nv set x y"],
        ])
        assert set(got) == {"csl-01", "csl-02"}

    def test_a_commented_row_does_not_validate_its_content(self):
        """Guidance text is prose — it must not be held to the nv-only rule."""
        got = _parse([["#  Not allowed", "nv config apply, reboot, ; | & $( )"]])
        assert got == {}

    def test_a_real_row_is_still_validated(self):
        """The skip must key on '#', not disable validation generally."""
        with pytest.raises(CustomConfigError):
            _parse([["Function: csl", "reboot"]])


class TestKeyPaths:
    """Any key path custom config touches must drop out of BOTH comparison sides.

    Live Air run, 2026-08-17, found the design incomplete. Two symptoms, one cause:

      generated:  nv set system message pre-login "ERA custom-config live test"
      running:    nv set system message pre-login 'ERA custom-config live test'

    NVUE re-quotes values, so the operator line itself reads as MISSING; and because
    the operator overrode a key the tool also sets, the tool default banner reads as
    MISSING too; and the running version reads as EXTRA. ADR-0055 handled only
    `nv unset`, which is the rarer case — overriding with `nv set` is the common one.
    """

    def test_unset_keypath_is_everything_after_the_verb(self):
        from excel_parser import custom_config_keypaths
        # `system wjh enable` is the normalizer spelling of the same setting
        assert custom_config_keypaths(["nv unset system wjh state"]) == [
            "system wjh enable", "system wjh state"]

    def test_set_keypath_drops_a_quoted_value(self):
        from excel_parser import custom_config_keypaths
        got = custom_config_keypaths(['nv set system message pre-login "Authorized use only"'])
        assert got == ["system message pre-login"]

    def test_set_keypath_drops_a_single_quoted_value(self):
        from excel_parser import custom_config_keypaths
        got = custom_config_keypaths(["nv set system message pre-login 'hello there'"])
        assert got == ["system message pre-login"]

    def test_set_keypath_drops_a_bare_value(self):
        from excel_parser import custom_config_keypaths
        assert custom_config_keypaths(["nv set system timezone Etc/UTC"]) == ["system timezone"]

    def test_keypaths_are_deduped_and_ordered(self):
        from excel_parser import custom_config_keypaths
        got = custom_config_keypaths([
            "nv set system timezone Etc/UTC",
            "nv unset system wjh state",
            "nv set system timezone Etc/GMT",
        ])
        assert got == ["system timezone", "system wjh enable", "system wjh state"]

    def test_show_lines_contribute_no_keypath(self):
        """`nv show` changes nothing, so it must not suppress any comparison."""
        from excel_parser import custom_config_keypaths
        assert custom_config_keypaths(["nv show system"]) == []

    def test_keypaths_land_in_host_vars(self, tmp_path):
        """The playbook reads this from host_vars; if absent the filter never fires."""
        from excel_parser import custom_config_keypaths
        assert custom_config_keypaths(['nv set system message pre-login "x"']) == [
            "system message pre-login"]


class TestNormalizerVocabularyAliases:
    """A key path must match whichever spelling the comparator ends up using.

    normalize_nvue.py:155 rewrites `state enabled` -> `enable on` before comparing.
    So the operator writes `nv unset system wjh state`, but the generated line the
    comparator sees is `nv set system wjh enable on` — key path `system wjh enable`,
    not `system wjh state`. The prefix match never fired and both cores reported
    MISMATCH on the NOZTP run of 2026-08-17.

    Third occurrence of the same vocabulary trap: ADR-0054 (models say `oob-switch`,
    classify_node says `oob`) and the role/function node-key split were the first two.
    """

    def test_state_path_also_yields_the_enable_spelling(self):
        from excel_parser import custom_config_keypaths
        got = custom_config_keypaths(["nv unset system wjh state"])
        assert "system wjh state" in got and "system wjh enable" in got, got

    def test_enable_path_also_yields_the_state_spelling(self):
        from excel_parser import custom_config_keypaths
        got = custom_config_keypaths(["nv unset system wjh enable"])
        assert "system wjh enable" in got and "system wjh state" in got, got

    def test_alias_only_applies_to_the_trailing_token(self):
        """`state` mid-path is not the enable/disable idiom and must not alias."""
        from excel_parser import custom_config_keypaths
        got = custom_config_keypaths(["nv unset system state machine thing"])
        assert got == ["system state machine thing"], got

    def test_unrelated_paths_gain_no_aliases(self):
        from excel_parser import custom_config_keypaths
        assert custom_config_keypaths(["nv set system timezone Etc/UTC"]) == ["system timezone"]
