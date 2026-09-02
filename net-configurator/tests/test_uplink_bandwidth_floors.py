# SPDX-FileCopyrightText: Copyright (c) 2026 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: MIT
"""ERA-69: external uplink capacity must meet the ERA per-GPU floors.

Every deployment guide states the same two minimums verbatim at every published
scale point: at least 25 Gb/GPU toward the customer network and at least
12.5 Gb/GPU for storage attachment. They are CLUSTER-LEVEL AGGREGATE floors —
a shared pool of uplinks sized against total GPU count, not per-node cabling.
Read per node they reproduce none of the published designs.

The floors and the per-link speed both come from the arch model, so the check
reports a workbook against its own architecture rather than a constant.
"""
import shutil
import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest
# Arch models are internal-only (data-models/, ADR-0027) and absent from the
# public tree, so this skips there rather than failing. Defined locally rather
# than imported from conftest: conftest is loaded by pytest, not importable.
_HAVE_ARCH_MODELS = (Path(__file__).resolve().parents[2] / "data-models" / "models").exists()
needs_arch_models = pytest.mark.skipif(
    not _HAVE_ARCH_MODELS,
    reason="arch models are internal-only (data-models/), absent in the public tree")

NC = Path(__file__).resolve().parent.parent
VALIDATOR = NC / "scripts" / "validate_excel.py"
ARCHS = ["2-4-3-200", "2-4-5-800", "2-8-5-200",
         "2-8-9-400", "2-8-9-400-SP", "2-8-9-800"]
MARKER = "below the ERA floor"


def _run(xlsx):
    return subprocess.run([sys.executable, str(VALIDATOR), str(xlsx)],
                          capture_output=True, text=True, cwd=str(NC))


def _lines(xlsx):
    p = _run(xlsx)
    return [l for l in (p.stdout + p.stderr).splitlines() if MARKER in l]


@pytest.mark.parametrize("arch", ARCHS)
@pytest.mark.parametrize("site", ["default", "largescale"])
def test_shipped_workbooks_meet_the_floors(arch, site):
    """No shipped workbook may ship under the published minimums."""
    xlsx = NC / "input" / arch / site / f"{arch}.xlsx"
    if not xlsx.exists():
        pytest.skip(f"no {site} workbook for {arch}")
    assert not _lines(xlsx), f"{arch}/{site} is under an ERA uplink floor"


def _strip_uplinks(src, dst, profile, keep):
    """Copy `src` to `dst` keeping only `keep` rows of `profile`."""
    shutil.copy(src, dst)
    wb = openpyxl.load_workbook(dst)
    ws = wb["Wire Map"]
    hdr = [str(c.value or "").strip() for c in next(ws.iter_rows())]
    ip = hdr.index("Network Profile") + 1
    seen, drop = 0, []
    for r in range(ws.max_row, 1, -1):
        if str(ws.cell(r, ip).value or "").strip() == profile:
            seen += 1
            if seen > keep:
                drop.append(r)
    for r in drop:
        ws.delete_rows(r)
    wb.save(dst)
    return len(drop)


@pytest.mark.parametrize("profile", ["Edge Uplink", "Storage Uplink"])
@needs_arch_models
def test_under_provisioned_uplinks_are_reported(profile, tmp_path):
    """The guard must actually fire — a check that never fails guards nothing.

    Keeps a single link of the profile, which is far under any floor, and
    asserts the message names the profile and quotes the achieved Gb/GPU.
    """
    arch = "2-8-9-400"
    src = NC / "input" / arch / "default" / f"{arch}.xlsx"
    if not src.exists():
        pytest.skip(f"no default workbook for {arch}")
    dst = tmp_path / f"{arch}.xlsx"
    removed = _strip_uplinks(src, dst, profile, keep=1)
    assert removed, f"no {profile} rows to remove — test would be vacuous"

    lines = _lines(dst)
    assert lines, f"removing {removed} {profile} rows was not reported"
    body = " ".join(lines)
    assert profile in body, body
    assert "Gb/GPU" in body, body


def test_over_provisioned_uplinks_are_not_reported(tmp_path):
    """The floors are minimums. Exceeding them is legal and must stay silent.

    2-4-3-200 ships 18.75 / 37.50 Gb/GPU against floors of 12.5 / 25 — if the
    check compared for equality rather than a minimum it would fail that arch
    on every run.
    """
    arch = "2-4-3-200"
    xlsx = NC / "input" / arch / "default" / f"{arch}.xlsx"
    if not xlsx.exists():
        pytest.skip(f"no default workbook for {arch}")
    assert not _lines(xlsx), "over-provisioned capacity must not be reported"
