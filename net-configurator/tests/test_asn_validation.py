# SPDX-FileCopyrightText: Copyright (c) 2026 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: MIT
"""Tests for the optional Loopbacks ASN column:
validate_excel group/range checks + parser reads the column.
"""
import sys
from pathlib import Path

import openpyxl
import pytest

SCRIPTS_DIR = Path(__file__).parent.parent / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

import validate_excel as ve
import excel_parser as ep


def _loopback_ws(rows, header=("Switch", "ASN")):
    """rows = list of tuples aligned to `header`."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Loopbacks"
    for j, h in enumerate(header, 1):
        ws.cell(1, j, h)
    for i, r in enumerate(rows, 2):
        for j, v in enumerate(r, 1):
            ws.cell(i, j, v)
    return ws


def _nodes(*names):
    return [{"name": n} for n in names]


def _run(ws, nodes, settings=None):
    res = ve.ValidationResult()
    ve.validate_loopbacks(ws, nodes, [], settings or {"ns_tiers": 1}, res)
    return res


# ── happy path ─────────────────────────────────────────────────────────────

def test_valid_converged_core_and_distinct_oob():
    ws = _loopback_ws([("core-01", 65001), ("core-02", 65001),
                       ("oob-switch-01", 65010), ("oob-switch-02", 65011)])
    res = _run(ws, _nodes("core-01", "core-02", "oob-switch-01", "oob-switch-02"))
    assert not res.errors, res.errors


def test_blank_asn_cells_are_ignored():
    ws = _loopback_ws([("core-01", None), ("core-02", "")])
    res = _run(ws, _nodes("core-01", "core-02"))
    assert not res.errors and not res.warnings


# ── range / format ─────────────────────────────────────────────────────────

@pytest.mark.parametrize("bad", [0, -5, 23456, 0xFFFFFFFF + 1, 1.5, "abc"])
def test_range_and_format_rejected(bad):
    ws = _loopback_ws([("oob-switch-01", bad)])
    res = _run(ws, _nodes("oob-switch-01"))
    assert res.errors, f"expected rejection for ASN={bad!r}"


def test_valid_4byte_asn_accepted():
    ws = _loopback_ws([("oob-switch-01", 0xFFFFFFFF)])
    res = _run(ws, _nodes("oob-switch-01"))
    assert not res.errors, res.errors


# ── equal-within ───────────────────────────────────────────────────────────

def test_collapsed_plane_mates_split_is_error():
    # 2-leaf plane -> mates must share one ASN
    ws = _loopback_ws([("gsl-plane1-01", 65001), ("gsl-plane1-02", 65002)])
    res = _run(ws, _nodes("gsl-plane1-01", "gsl-plane1-02"))
    assert any("must share ONE BGP ASN" in e for e in res.errors), res.errors


def test_converged_core_split_is_error():
    ws = _loopback_ws([("core-01", 65001), ("core-02", 65099)])
    res = _run(ws, _nodes("core-01", "core-02"))
    assert any("must share ONE BGP ASN" in e for e in res.errors), res.errors


def test_partial_override_of_shared_group_warns():
    # one mate set, the other blank -> warn (not error)
    ws = _loopback_ws([("gsl-plane1-01", 65001), ("gsl-plane1-02", None)])
    res = _run(ws, _nodes("gsl-plane1-01", "gsl-plane1-02"))
    assert not res.errors, res.errors
    assert any("Set all or none" in w for w in res.warnings), res.warnings


# ── distinct-across ────────────────────────────────────────────────────────

def test_two_groups_same_asn_is_error():
    ws = _loopback_ws([("oob-switch-01", 65010), ("oob-switch-02", 65010)])
    res = _run(ws, _nodes("oob-switch-01", "oob-switch-02"))
    assert any("distinct ASNs" in e for e in res.errors), res.errors


def test_spined_plane_leaf_collides_with_another_leaf():
    names = [f"gl-plane1-{i:02d}" for i in range(1, 5)]  # 4 leaves -> singletons
    ws = _loopback_ws([(names[0], 65001), (names[1], 65001)])  # two leaves same ASN
    res = _run(ws, _nodes(*names))
    assert any("distinct ASNs" in e for e in res.errors), res.errors


# ── header recognition + parser read ───────────────────────────────────────

def test_bgp_asn_header_alias_recognized():
    ws = _loopback_ws([("oob-switch-01", 23456)], header=("Switch", "BGP ASN"))
    res = _run(ws, _nodes("oob-switch-01"))
    assert any("reserved" in e for e in res.errors), res.errors


def _wb_with_loopbacks(rows, header=("Switch", "ASN")):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Loopbacks")
    for j, h in enumerate(header, 1):
        ws.cell(1, j, h)
    for i, r in enumerate(rows, 2):
        for j, v in enumerate(r, 1):
            ws.cell(i, j, v)
    return wb


def test_asn_source_present_when_column_populated():
    wb = _wb_with_loopbacks([("core-01", 65001)])
    assert ve.loopbacks_asn_populated(wb) is True


def test_asn_source_absent_when_column_blank():
    wb = _wb_with_loopbacks([("core-01", None)])
    assert ve.loopbacks_asn_populated(wb) is False


def test_asn_source_absent_when_no_asn_column():
    wb = _wb_with_loopbacks([("core-01", "172.16.176.11")], header=("Switch", "Default"))
    assert ve.loopbacks_asn_populated(wb) is False


def test_asn_source_absent_when_no_loopbacks_sheet():
    wb = openpyxl.Workbook()
    assert ve.loopbacks_asn_populated(wb) is False


def test_parser_reads_asn_column():
    ws = _loopback_ws([("core-01", "172.16.176.11", 65001)],
                      header=("Switch", "Default", "ASN"))
    overrides = ep.parse_loopbacks_sheet(ws)
    assert overrides["core-01"]["asn"] == "65001"
    assert overrides["core-01"]["lo"] == "172.16.176.11"
