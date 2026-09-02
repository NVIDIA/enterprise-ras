# SPDX-FileCopyrightText: Copyright (c) 2026 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: MIT
"""End-to-end: an explicit Loopbacks `ASN` value overrides the derived
per-node BGP ASN in the generated inventory + config, and nothing else changes.

Runs against a throwaway copy of the tool tree (never the live repo), mirroring
test_seedless_generation.py.
"""
import shutil
import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest

NC = Path(__file__).resolve().parent.parent  # net-configurator/
ARCH = "2-8-9-800"
SWITCH = "oob-switch-01"
OVERRIDE = 4260394838  # base(4260394788)+50 — distinct, valid, OOB singleton

_COPY_IGNORE = shutil.ignore_patterns(
    "output", "archive", ".git", ".venv*", "__pycache__", "*.pyc",
    ".pytest_cache", ".era-secrets", "legacy",
)


def _inject_asn(xlsx: Path, switch: str, asn: int):
    """Set `switch`'s ASN in the Loopbacks ASN column (create the column if the
    workbook predates it), overriding just that one switch."""
    wb = openpyxl.load_workbook(xlsx)
    lb = next(s for s in wb.sheetnames if str(s).strip().lower().startswith("loopback"))
    ws = wb[lb]
    hr = next(r for r in range(1, 5)
              if str(ws.cell(r, 1).value or "").strip().startswith("Switch"))
    col = next((c for c in range(1, ws.max_column + 1)
                if str(ws.cell(hr, c).value or "").strip().lower() in
                ("asn", "bgp asn", "bgp_asn", "autonomous system", "as number")), None)
    if col is None:
        col = ws.max_column + 1
        ws.cell(hr, col, "ASN")
    for r in range(hr + 1, ws.max_row + 1):
        if str(ws.cell(r, 1).value or "").strip() == switch:
            ws.cell(r, col, asn)
    wb.save(xlsx)


def test_asn_override_flows_to_config(tmp_path):
    src_xlsx = NC / "input" / ARCH / "default" / f"{ARCH}.xlsx"
    if not src_xlsx.exists():
        pytest.skip(f"no committed default workbook for {ARCH}")

    dst = tmp_path / "nc"
    shutil.copytree(NC, dst, ignore=_COPY_IGNORE)
    _inject_asn(dst / "input" / ARCH / "default" / f"{ARCH}.xlsx", SWITCH, OVERRIDE)

    r = subprocess.run(
        [sys.executable, "scripts/excel_parser.py", "--arch", ARCH,
         "--site", "default", "--skip-validate"],
        cwd=dst, capture_output=True, text=True,
    )
    assert r.returncode == 0, f"parser failed:\n{r.stderr[-2000:]}"

    # The parser generates inventory (host_vars); the NVUE .sh render is a later
    # ansible step (covered by the byte-identical make-generate gate + Air).
    hv_dir = dst / "output" / ARCH / "default" / "inventory" / "host_vars"
    hv = hv_dir / f"{SWITCH}.yml"
    assert hv.exists(), "override switch host_vars missing"
    assert f"bgp_asn: {OVERRIDE}" in hv.read_text(), \
        f"override ASN not in {SWITCH} host_vars"

    # a non-overridden OOB switch keeps its derived ASN (base+idx), not the override
    other = hv_dir / "oob-switch-02.yml"
    assert f"bgp_asn: {OVERRIDE}" not in other.read_text(), \
        "override leaked to a switch that was not overridden"
    assert "bgp_asn: 4260394790" in other.read_text(), \
        "non-overridden OOB switch lost its derived ASN"
