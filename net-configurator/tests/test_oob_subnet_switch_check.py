# SPDX-FileCopyrightText: Copyright (c) 2026 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: MIT
"""
ERA-61 — the OOB-subnet reachability check must cover fabric switches.

`check_vlan_oob_mapping` hard-fails a device whose mgmt IP falls outside every
declared OOB VLAN subnet, because such a device is unreachable after deploy.
It used to skip every category in `_SWITCH_CATEGORIES`, so a switch addressed
off-subnet passed `make validate-excel` clean and then could not be reached.

That was a regression, not an original gap: the predecessor check (removed with
`mgmt_subnets` in d2e29d3) iterated ALL parsed_nodes with no switch exemption.
The replacement in 2b25cbc tightened severity (warn -> error) and silently
narrowed scope at the same time. Only the tightening was intended.

The exemption is now `_OOB_SUBNET_EXEMPT_CATEGORIES` — `edge` (customer-owned
upstream) and `air-oob` (Air virtual node, belongs on air_mgmt_subnet). Every
ERA-managed fabric switch is covered.
"""
import shutil
import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
VALIDATOR = REPO / "scripts" / "validate_excel.py"

# 2-4-5-800 is the only shipped default that populates switch mgmt IPs
# (16 of them, all inside 192.168.200.0/24). The others leave them blank for
# ZTP to assign, so the empty-IP guard skips them and they cannot exercise this.
ARCH = "2-4-5-800"
BASE = REPO / "input" / ARCH / "default" / f"{ARCH}.xlsx"
OFF_SUBNET_IP = "10.99.99.99"
ERR = "not within any OOB VLAN subnet"

pytestmark = pytest.mark.skipif(not BASE.exists(), reason=f"{ARCH} Excel absent")


def _run(xlsx):
    return subprocess.run([sys.executable, str(VALIDATOR), str(xlsx)],
                          capture_output=True, text=True, cwd=str(REPO))


@pytest.fixture
def mutated(tmp_path):
    """Copy the shipped workbook, rewrite one node's mgmt IP, validate it."""
    def _go(match_name, ip):
        dst = tmp_path / f"{ARCH}.xlsx"
        shutil.copy2(BASE, dst)
        wb = openpyxl.load_workbook(dst)
        ws = wb["Nodes"]
        hdr = [str(c.value or "").strip() for c in ws[1]]
        name_c = hdr.index("Name") + 1
        ip_c = hdr.index("Mgmt IP Address") + 1
        hit = False
        for r in range(2, ws.max_row + 1):
            if str(ws.cell(r, name_c).value or "").strip() == match_name:
                ws.cell(r, ip_c).value = ip
                hit = True
                break
        assert hit, f"{match_name} not found in the Nodes tab"
        wb.save(dst)
        return _run(dst)
    return _go


def test_shipped_default_is_clean():
    """The fix must not fail a workbook that was already correct — all 16
    switch mgmt IPs in this arch sit inside 192.168.200.0/24."""
    p = _run(BASE)
    assert ERR not in (p.stdout + p.stderr), \
        "shipped default must not trip the OOB-subnet check"
    assert p.returncode == 0, p.stdout + p.stderr


def test_offsubnet_fabric_switch_is_rejected(mutated):
    """The reported bug. Pre-fix this passed clean and the switch was
    unreachable after deploy."""
    p = mutated("csl-01", OFF_SUBNET_IP)
    out = p.stdout + p.stderr
    assert ERR in out, "an off-subnet switch mgmt IP must be reported"
    assert OFF_SUBNET_IP in out, "the message should name the offending IP"
    assert p.returncode != 0, "an unreachable switch must fail validation"


@pytest.mark.parametrize("switch", ["gl-plane1-01", "gs-plane2-02", "oob-switch-01"])
def test_every_fabric_switch_role_is_covered(mutated, switch):
    """Not just csl — leaf, spine and OOB switches are all ERA-managed and all
    land on the OOB plane."""
    p = mutated(switch, OFF_SUBNET_IP)
    assert ERR in (p.stdout + p.stderr), f"{switch} should be covered"


@pytest.mark.parametrize("device", ["cust-net-edge-01", "air-oob-switch"])
def test_legitimately_offplane_devices_stay_exempt(mutated, device):
    """`edge` is customer-owned and `air-oob` belongs on air_mgmt_subnet.
    Neither is ours to address, so neither may be failed by this check."""
    p = mutated(device, OFF_SUBNET_IP)
    out = p.stdout + p.stderr
    assert ERR not in out, f"{device} must remain exempt from the OOB check"
    assert p.returncode == 0, out


def test_server_coverage_is_unchanged(mutated):
    """Non-switch devices were always covered; this must not regress."""
    p = mutated("su-01-node-01", OFF_SUBNET_IP)
    assert ERR in (p.stdout + p.stderr)
