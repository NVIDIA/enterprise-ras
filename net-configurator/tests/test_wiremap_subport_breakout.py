# SPDX-FileCopyrightText: Copyright (c) 2026 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: MIT
"""
Every sub-port the Wire Map wires must be one the breakout actually creates.

A port broken out 2x has s0-s1 only. Wire swp7s2 and the generated config still
emits it into `type swp`, the bonds and the BGP neighbours — those lists come
from the Wire Map independently of the breakout section. The config generates
and validates cleanly, then fails at `nv config apply`.

Two shipped defaults were doing exactly this before this check existed:

  2-8-9-400   swp57  rendered 4x, config referenced s4-s7
              (Storage Uplink declared 4x2 while wiring swp57s0..s7)
  2-4-5-800   swp7,9-13  rendered 2x, config referenced s2,s3
              (ISL/Edge Uplink declared 2x4 while wiring s0..s3)

The 2-4-5-800 case is the subtler one: swp7 is claimed by CPU/In-Band (4x) AND
Edge Uplink (2x), and the coarser breakout wins. Checking each profile against
its own declared breakout would miss it — every profile looks self-consistent
in isolation. The check has to resolve the per-port winner first, in the same
8x > 2x > 4x last-wins order that core_nvue_cli.j2 emits.
"""
import shutil
import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
VALIDATOR = REPO / "scripts" / "validate_excel.py"
ERR = "renders as"

# Derived, never hand-listed. This was a literal six-arch list that predated
# `2-4-5-400` and silently never covered it — the same pinned-list bug that made
# `generator_parity.py` skip that arch. A list you have to remember to edit is a list
# that stops matching the tree.
SHIPPED_SITES = ("default", "largescale")
SHIPPED = sorted(p for site in SHIPPED_SITES for p in REPO.glob(f"input/*/{site}/*.xlsx"))


def _label(p: Path) -> str:
    return f"{p.parent.parent.name}/{p.parent.name}"


# The workbook used to build mutation fixtures below. Kept as an explicit default-site
# path: these tests assert on specific ports and profile names in that one book.
def _fixture_book(arch: str) -> Path:
    return REPO / "input" / arch / "default" / f"{arch}.xlsx"


def _run(xlsx):
    return subprocess.run([sys.executable, str(VALIDATOR), str(xlsx)],
                          capture_output=True, text=True, cwd=str(REPO))


@pytest.mark.skipif(not SHIPPED, reason="no shipped workbooks present")
@pytest.mark.parametrize("xlsx", SHIPPED, ids=_label)
def test_shipped_workbooks_have_no_impossible_subports(xlsx):
    """Every shipped workbook must be clean, at both sites.

    2-4-5-800 and 2-8-9-400 were not, before the accompanying workbook fix. The
    largescale books were added to this check on 2026-08-13 and were already clean —
    they had simply never been looked at, and they carry the bigger wiremaps.
    """
    p = _run(xlsx)
    offending = [l for l in (p.stdout + p.stderr).splitlines() if ERR in l]
    assert not offending, f"{_label(xlsx)}:\n" + "\n".join(offending)


def test_selection_covers_every_tracked_workbook():
    """The list above must track the tree, not a memory of it.

    Asked against git rather than against `SHIPPED_SITES`, because deriving the
    expectation from the same constant the selection uses is circular: dropping a site
    shrinks both sides and the suite still passes, just smaller. Mirrors the identical
    guard in `test_shipped_workbooks_are_clean.py` — kept separate rather than shared so
    neither suite can be silently narrowed by an edit to the other.
    """
    ls = subprocess.run(["git", "ls-files", "input/*/*/*.xlsx"],
                        cwd=str(REPO), capture_output=True, text=True)
    if ls.returncode != 0 or not ls.stdout.strip():
        pytest.skip("not a git checkout (source tarball) — nothing to cross-check against")

    expected = {"/".join(line.split("/")[1:3]) for line in ls.stdout.split()}
    actual = {_label(p) for p in SHIPPED}
    assert actual == expected, (
        "the subport check does not cover every tracked workbook under input/.\n"
        f"  tracked but NOT checked: {sorted(expected - actual)}\n"
        f"  checked but not tracked: {sorted(actual - expected)}"
    )


@pytest.fixture
def mutated(tmp_path):
    """Copy a shipped workbook, rewrite one Wire Map port cell, validate."""
    def _go(arch, old_port, new_port):
        src = _fixture_book(arch)
        dst = tmp_path / f"{arch}.xlsx"
        shutil.copy2(src, dst)
        wb = openpyxl.load_workbook(dst)
        ws = wb["Wire Map"]
        hdr = [str(c.value or "").strip() for c in ws[1]]
        cols = [hdr.index("Port (A)") + 1, hdr.index("Port (B)") + 1]
        hits = 0
        for r in range(2, ws.max_row + 1):
            for c in cols:
                if str(ws.cell(r, c).value or "").strip() == old_port:
                    ws.cell(r, c).value = new_port
                    hits += 1
        assert hits, f"{old_port} not found in {arch} Wire Map"
        wb.save(dst)
        return _run(dst)
    return _go


def test_subport_beyond_breakout_is_rejected(mutated):
    """swp1 on 2-8-9-800 is CPU/In-Band at 2x — s0,s1 only. Rewiring a link
    onto s3 must be caught."""
    p = mutated("2-8-9-800", "swp1s1", "swp1s3")
    out = p.stdout + p.stderr
    assert ERR in out, "a sub-port beyond the breakout must be reported"
    assert "swp1" in out and "s3" in out, out
    assert p.returncode != 0, "this must fail validation, not warn"


def test_message_names_the_claiming_profiles(mutated):
    """The operator needs to know which profile to change, with its breakout.

    swp1 is a different role on different switches, which is why the message is
    keyed by system name as well as port. The assertion names the profile and
    its breakout factor so a message that degrades to a bare port number fails.

    Updated 2026-08-05: this expected `GPU Network (2x)` when the mutated row
    landed on a GSL-plane switch. Regenerating the workbooks from the models
    moved swp1's first match to a csl CPU/In-Band port. The mechanism under
    test — profile name + breakout, keyed by system — is unchanged; only which
    switch owns swp1 in the shipped book is.
    """
    p = mutated("2-8-9-800", "swp1s1", "swp1s3")
    out = p.stdout + p.stderr
    assert "Claimed by: CPU/In-Band Network (2x)" in out, out
    assert "gsl-plane1-01 swp1" in out, "the switch must be named, not just the port"


def test_legal_subport_is_accepted(mutated):
    """s0 <-> s1 within a 2x port stays clean — the check must not fire on
    every rewire."""
    p = mutated("2-8-9-800", "swp1s1", "swp1s0")
    assert ERR not in (p.stdout + p.stderr)


def test_eight_way_port_allows_high_subports(mutated):
    """An 8x port legitimately reaches s7; the check must not flag it."""
    p = mutated("2-8-9-800", "swp1s1", "swp1s3")
    # sanity: the same index on an 8x port is fine
    q = mutated("2-4-3-200", "swp59s1", "swp59s7")
    assert ERR in (p.stdout + p.stderr)
    assert ERR not in (q.stdout + q.stderr), \
        "s7 on an 8x breakout port is legal and must not be reported"
