# SPDX-FileCopyrightText: Copyright (c) 2026 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: MIT
"""
Tests for the ERA topology generator.

Tests parse_wiremap_excel(), parse_swp_port(), TopologyGenerator, MAC generation
consistency, duplicate endpoint detection, and output JSON structure using
in-memory openpyxl workbooks written to tmp_path.
"""
import json
import re
import sys
from pathlib import Path

import openpyxl
import pytest

# Add scripts/ to path so we can import the modules
SCRIPTS_DIR = Path(__file__).parent.parent / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from utils import generate_mac, reset_mac_registry
from topology_generator import (
    parse_swp_port,
    parse_wiremap_excel,
    parse_air_only_sheet,
    parse_version_image_map,
    parse_switch_versions,
    TopologyGenerator,
    WiremapRow,
    SWITCH_MODELS,
    SWITCH_OS_FALLBACK,
    SERVER_OS,
    NODE_DEFAULTS,
)


@pytest.fixture(autouse=True)
def _reset_mac_reg():
    """Reset MAC registry between tests to avoid cross-test collision errors."""
    reset_mac_registry()
    yield
    reset_mac_registry()


# ---------------------------------------------------------------------------
# Helper: build minimal Excel files on disk (TopologyGenerator needs file paths)
# ---------------------------------------------------------------------------

def _build_wiremap_workbook(rows=None, settings=None, air_only_rows=None):
    """Create an openpyxl Workbook with Wire Map (and optionally Air_Only/Settings)."""
    wb = openpyxl.Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Settings sheet
    ws_settings = wb.create_sheet("Settings")
    if settings:
        for key, value in settings:
            ws_settings.append([key, value])

    # Wire Map sheet with 13 columns
    ws = wb.create_sheet("Wire Map")
    header = [
        "Display in Air", "System Role", "System Name", "NIC/Port",
        "Speed", "Description", "Network Profile",
        "Disabled by Neighbor", "Speed2", "Description2",
        "Switch Role", "Switch Name", "Switch Port",
    ]
    ws.append(header)
    if rows:
        for row in rows:
            padded = list(row) + [None] * (13 - len(row))
            ws.append(padded[:13])

    # Air_Only sheet (optional, new format)
    if air_only_rows is not None:
        ws_air = wb.create_sheet("Air_Only")
        air_header = [
            "Display in Air", "System Role", "System Name", "NIC/Port",
            "Network Profile", "Switch Role", "Switch Name", "Switch Port",
        ]
        ws_air.append(air_header)
        for row in air_only_rows:
            padded = list(row) + [None] * (8 - len(row))
            ws_air.append(padded[:8])

    return wb


def _write_workbook(wb, tmp_path, arch="2-8-5-200"):
    """Write a workbook to the expected path for TopologyGenerator."""
    input_dir = tmp_path / "input" / arch / "default"
    input_dir.mkdir(parents=True, exist_ok=True)
    excel_path = input_dir / f"{arch}.xlsx"
    wb.save(str(excel_path))
    return excel_path


# ---------------------------------------------------------------------------
# parse_swp_port
# ---------------------------------------------------------------------------

class TestParseSwpPort:
    """Tests for parse_swp_port() which extracts port numbers."""

    def test_simple_port(self):
        assert parse_swp_port("swp49") == (49, None)

    def test_subport(self):
        assert parse_swp_port("swp49s3") == (49, 3)

    def test_bare_number(self):
        assert parse_swp_port("50") == (50, None)

    def test_invalid_string(self):
        assert parse_swp_port("eth0") is None

    def test_empty_string(self):
        assert parse_swp_port("") is None

    def test_port_with_prefix_noise(self):
        """Strings that don't match the pattern return None."""
        assert parse_swp_port("bond1s0") is None

    def test_zero_port(self):
        assert parse_swp_port("swp0") == (0, None)


# ---------------------------------------------------------------------------
# parse_wiremap_excel (file-based)
# ---------------------------------------------------------------------------

class TestParseWiremapExcel:
    """Tests for parse_wiremap_excel() which reads Wire Map from an xlsx file."""

    def test_empty_wiremap(self, tmp_path):
        """Wire Map with only header returns empty list."""
        wb = _build_wiremap_workbook()
        path = _write_workbook(wb, tmp_path)
        result = parse_wiremap_excel(path)
        assert result == []

    def test_valid_rows(self, tmp_path):
        """Data rows are parsed into WiremapRow objects."""
        wb = _build_wiremap_workbook(rows=[
            ["Yes", "su-01-node-01", "su-01-node-01", "NIC1_P1", None, None,
             "CPU/In-Band Network", None, None, None,
             "core-01", "core-01", "swp1s0"],
        ])
        path = _write_workbook(wb, tmp_path)

        result = parse_wiremap_excel(path)
        assert len(result) == 1
        assert isinstance(result[0], WiremapRow)
        assert result[0].system_role == "su-01-node-01"
        assert result[0].switch_port == "swp1s0"
        assert result[0].display_in_air is True

    def test_missing_wiremap_sheet_raises(self, tmp_path):
        """Workbook without Wire Map sheet raises ValueError."""
        wb = openpyxl.Workbook()
        wb.active.title = "Not Wire Map"
        path = tmp_path / "test.xlsx"
        wb.save(str(path))

        with pytest.raises(ValueError, match="Wire Map"):
            parse_wiremap_excel(path)

    def test_skip_empty_system_role(self, tmp_path):
        """Rows with empty System Role are skipped."""
        wb = _build_wiremap_workbook(rows=[
            ["Yes", "", "", "NIC1", None, None, "CPU", None, None, None,
             "core-01", "core-01", "swp1s0"],
        ])
        path = _write_workbook(wb, tmp_path)

        result = parse_wiremap_excel(path)
        assert result == []


# ---------------------------------------------------------------------------
# parse_air_only_sheet
# ---------------------------------------------------------------------------

class TestParseAirOnlySheet:
    """Tests for parse_air_only_sheet() which reads the Air_Only sheet."""

    def test_no_air_only_sheet(self):
        """Workbook without Air_Only returns empty list."""
        wb = openpyxl.Workbook()
        result = parse_air_only_sheet(wb)
        assert result == []

    def test_valid_air_only_rows(self):
        """Air_Only rows are correctly parsed."""
        wb = openpyxl.Workbook()
        ws = wb.create_sheet("Air_Only")
        ws.append(["Display in Air", "System Role", "System Name", "NIC/Port",
                    "Network Profile", "Switch Role", "Switch Name", "Switch Port"])
        ws.append(["Yes", "dhcp-oob", "dhcp-oob", "eth1",
                    "Air - Management", "oob-switch-01", "oob-switch-01", "swp44"])

        result = parse_air_only_sheet(wb)
        assert len(result) == 1
        assert result[0].system_role == "dhcp-oob"
        assert result[0].switch_port == "swp44"
        assert result[0].display_in_air is True


# ---------------------------------------------------------------------------
# parse_version_image_map
# ---------------------------------------------------------------------------

class TestParseVersionImageMap:
    """Tests for parse_version_image_map() from Air_Only sheet."""

    def test_no_air_only_sheet(self):
        """Workbook without Air_Only returns empty dict."""
        wb = openpyxl.Workbook()
        result = parse_version_image_map(wb)
        assert result == {}

    def test_valid_version_map(self):
        """Version table is correctly parsed."""
        wb = openpyxl.Workbook()
        ws = wb.create_sheet("Air_Only")
        # Some data rows first (8+ columns, will be skipped by the version parser
        # because col1 won't be 'Friendly Version')
        ws.append(["Display in Air", "System Role", "System Name", "NIC/Port",
                    "Network Profile", "Switch Role", "Switch Name", "Switch Port"])
        ws.append(["Yes", "dhcp-oob", "dhcp-oob", "eth1",
                    "Air - Management", "oob-switch-01", "oob-switch-01", "swp44"])
        # Version table
        ws.append(["Friendly Version", "Air Image Name"])
        ws.append(["5.16.1", "cumulus-linux-vx-amd64-5.16.1.0008.qcow2"])
        ws.append(["5.15.1", "cumulus-linux-vx-amd64-5.15.1.qcow2"])

        result = parse_version_image_map(wb)
        assert result["5.16.1"] == "cumulus-linux-vx-amd64-5.16.1.0008.qcow2"
        assert result["5.15.1"] == "cumulus-linux-vx-amd64-5.15.1.qcow2"


# ---------------------------------------------------------------------------
# parse_switch_versions
# ---------------------------------------------------------------------------

class TestParseSwitchVersions:
    """Tests for parse_switch_versions() from Settings sheet."""

    def test_no_settings_sheet(self):
        """Workbook without Settings returns empty dict."""
        wb = openpyxl.Workbook()
        wb.active.title = "Other"
        result = parse_switch_versions(wb)
        assert result == {}

    def test_valid_versions(self):
        wb = openpyxl.Workbook()
        ws = wb.create_sheet("Settings")
        ws.append(["Switch Function", "Cumulus Version"])
        ws.append(["core", "5.16.1"])
        ws.append(["oob", "5.15.1"])

        result = parse_switch_versions(wb)
        assert result == {"core": "5.16.1", "oob": "5.15.1"}


# ---------------------------------------------------------------------------
# TopologyGenerator — integration with tmp_path Excel files
# ---------------------------------------------------------------------------

class TestTopologyGenerator:
    """Tests for the TopologyGenerator class."""

    def _make_minimal_topology(self, tmp_path, wiremap_rows, settings=None,
                               air_only_rows=None, arch="2-8-5-200"):
        """Helper: write workbook and instantiate TopologyGenerator."""
        wb = _build_wiremap_workbook(
            rows=wiremap_rows,
            settings=settings,
            air_only_rows=air_only_rows,
        )
        path = _write_workbook(wb, tmp_path, arch=arch)
        return TopologyGenerator(path, arch)

    def test_empty_topology(self, tmp_path):
        """Empty wiremap produces topology with no nodes or links."""
        gen = self._make_minimal_topology(tmp_path, [])
        topo = gen.generate()

        assert "content" in topo
        assert "nodes" in topo["content"]
        assert "links" in topo["content"]
        assert topo["content"]["oob"] is False

    def test_basic_node_creation(self, tmp_path):
        """Nodes from wiremap are created in the topology."""
        rows = [
            ["Yes", "su-01-node-01", "su-01-node-01", "NIC1_P1", None, None,
             "CPU/In-Band Network", None, None, None,
             "core-01", "core-01", "swp1s0"],
        ]
        gen = self._make_minimal_topology(tmp_path, rows)
        topo = gen.generate()

        nodes = topo["content"]["nodes"]
        assert "su-01-node-01" in nodes
        assert "core-01" in nodes

    def test_node_os_assignment(self, tmp_path):
        """Switches get Cumulus OS, servers get Ubuntu OS."""
        rows = [
            ["Yes", "su-01-node-01", "su-01-node-01", "NIC1_P1", None, None,
             "CPU/In-Band Network", None, None, None,
             "core-01", "core-01", "swp1s0"],
        ]
        gen = self._make_minimal_topology(tmp_path, rows)
        topo = gen.generate()

        nodes = topo["content"]["nodes"]
        assert nodes["core-01"]["os"] == SWITCH_OS_FALLBACK
        assert nodes["su-01-node-01"]["os"] == SERVER_OS

    def test_link_creation(self, tmp_path):
        """Links are created from wiremap connections."""
        rows = [
            ["Yes", "su-01-node-01", "su-01-node-01", "NIC1_P1", None, None,
             "CPU/In-Band Network", None, None, None,
             "core-01", "core-01", "swp1s0"],
        ]
        gen = self._make_minimal_topology(tmp_path, rows)
        topo = gen.generate()

        links = topo["content"]["links"]
        # Should have at least one connected link plus unconnected stubs
        connected_links = [l for l in links if isinstance(l[1], dict)]
        assert len(connected_links) >= 1

        # Check the connected link endpoints
        link = connected_links[0]
        node_names = {link[0]["node"], link[1]["node"]}
        assert node_names == {"su-01-node-01", "core-01"}

    def test_display_in_air_no_skipped(self, tmp_path):
        """Rows with Display in Air = No do not create their specific nodes.

        Note: air-oob-switch and infra nodes (dhcp-oob, oob-server-01) may still
        be injected by _inject_air_oob_switch(), but wiremap-derived nodes should
        not appear.
        """
        rows = [
            ["No", "su-01-node-01", "su-01-node-01", "NIC1_P1", None, None,
             "CPU/In-Band Network", None, None, None,
             "core-01", "core-01", "swp1s0"],
        ]
        gen = self._make_minimal_topology(tmp_path, rows)
        topo = gen.generate()

        nodes = topo["content"]["nodes"]
        # The wiremap nodes should NOT be present
        assert "su-01-node-01" not in nodes
        assert "core-01" not in nodes

    def test_duplicate_endpoints_deduped(self, tmp_path):
        """Duplicate (node, interface) pairs keep only the first occurrence."""
        rows = [
            ["Yes", "su-01-node-01", "su-01-node-01", "NIC1_P1", None, None,
             "CPU/In-Band Network", None, None, None,
             "core-01", "core-01", "swp1s0"],
            # Same switch port, different system — this should be deduped
            ["Yes", "su-01-node-02", "su-01-node-02", "NIC1_P1", None, None,
             "CPU/In-Band Network", None, None, None,
             "core-01", "core-01", "swp1s0"],
        ]
        gen = self._make_minimal_topology(tmp_path, rows)
        topo = gen.generate()

        # core-01:swp1s0 should only appear in one connected link
        connected_links = [l for l in topo["content"]["links"] if isinstance(l[1], dict)]
        core_swp1s0_count = sum(
            1 for link in connected_links
            for ep in [link[0], link[1]]
            if ep["node"] == "core-01" and ep["interface"] == "swp1s0"
        )
        assert core_swp1s0_count == 1

    def test_outbound_link(self, tmp_path):
        """Rows with switch_role 'outbound' create outbound links."""
        rows = [
            ["Yes", "dhcp-oob", "dhcp-oob", "eth0", None, None,
             "outbound", None, None, None,
             "outbound", "", ""],
        ]
        gen = self._make_minimal_topology(tmp_path, rows)
        topo = gen.generate()

        outbound_links = [l for l in topo["content"]["links"]
                          if isinstance(l[1], str) and l[1] == "outbound"]
        assert len(outbound_links) >= 1
        assert outbound_links[0][0]["node"] == "dhcp-oob"

    def test_mac_in_links(self, tmp_path):
        """Link endpoints include deterministic MAC addresses."""
        rows = [
            ["Yes", "su-01-node-01", "su-01-node-01", "NIC1_P1", None, None,
             "CPU/In-Band Network", None, None, None,
             "core-01", "core-01", "swp1s0"],
        ]
        gen = self._make_minimal_topology(tmp_path, rows)
        topo = gen.generate()

        connected_links = [l for l in topo["content"]["links"] if isinstance(l[1], dict)]
        assert len(connected_links) >= 1
        for ep in [connected_links[0][0], connected_links[0][1]]:
            assert "mac" in ep
            assert re.match(r'^48:b0:2d:[0-9a-f]{2}:[0-9a-f]{2}:[0-9a-f]{2}$', ep["mac"])

    def test_topology_json_structure(self, tmp_path):
        """Generated topology has all required top-level keys."""
        rows = [
            ["Yes", "su-01-node-01", "su-01-node-01", "NIC1_P1", None, None,
             "CPU/In-Band Network", None, None, None,
             "core-01", "core-01", "swp1s0"],
        ]
        gen = self._make_minimal_topology(tmp_path, rows)
        topo = gen.generate()

        # Top-level structure
        assert topo["format"] == "JSON"
        assert "title" in topo
        assert "content" in topo
        assert topo["ztp"] is None

        # Content structure
        content = topo["content"]
        assert "nodes" in content
        assert "links" in content
        assert content["oob"] is False

    def test_node_structure(self, tmp_path):
        """Each node has all required fields."""
        rows = [
            ["Yes", "su-01-node-01", "su-01-node-01", "NIC1_P1", None, None,
             "CPU/In-Band Network", None, None, None,
             "core-01", "core-01", "swp1s0"],
        ]
        gen = self._make_minimal_topology(tmp_path, rows)
        topo = gen.generate()

        node = topo["content"]["nodes"]["core-01"]
        required_fields = ["cpu", "memory", "storage", "positioning", "os",
                           "features", "pxehost", "secureboot", "oob",
                           "emulation_type", "network_pci"]
        for field in required_fields:
            assert field in node, f"Missing field: {field}"
        assert isinstance(node["positioning"], dict)
        assert "x" in node["positioning"]
        assert "y" in node["positioning"]

    def test_unconnected_stubs_generated(self, tmp_path):
        """Switches get unconnected stubs for ports without connections."""
        rows = [
            ["Yes", "su-01-node-01", "su-01-node-01", "NIC1_P1", None, None,
             "CPU/In-Band Network", None, None, None,
             "core-01", "core-01", "swp1s0"],
        ]
        gen = self._make_minimal_topology(tmp_path, rows)
        topo = gen.generate()

        unconnected_links = [l for l in topo["content"]["links"]
                             if isinstance(l[1], str) and l[1] == "unconnected"]
        # core-01 is SN5610 with 64 ports — most should be unconnected
        assert len(unconnected_links) > 0

    def test_write_to_file(self, tmp_path):
        """TopologyGenerator.write() creates a valid JSON file."""
        rows = [
            ["Yes", "su-01-node-01", "su-01-node-01", "NIC1_P1", None, None,
             "CPU/In-Band Network", None, None, None,
             "core-01", "core-01", "swp1s0"],
        ]
        gen = self._make_minimal_topology(tmp_path, rows)

        output_path = tmp_path / "output" / "topology.json"
        gen.write(output_path)

        assert output_path.exists()
        with open(output_path) as f:
            data = json.load(f)
        assert data["format"] == "JSON"

    def test_invalid_hostname_skipped(self, tmp_path):
        """Rows with invalid hostnames (spaces) are not included."""
        rows = [
            ["Yes", "SPARE ISL", "SPARE ISL", "swp60", None, None,
             "ISL", None, None, None,
             "core-02", "core-02", "swp60"],
        ]
        gen = self._make_minimal_topology(tmp_path, rows)
        topo = gen.generate()

        nodes = topo["content"]["nodes"]
        assert "SPARE ISL" not in nodes


# ---------------------------------------------------------------------------
# L3 OOB injection (oob_uplink_mode = l3)
# ---------------------------------------------------------------------------

class TestL3OobInjection:
    """Tests for _inject_l3_oob_nodes — the L3-OOB Air topology path.

    Design doc: docs/plans/2026-05-20-l3-oob-air-topology.md
    """

    def _l3_wiremap(self):
        """Minimal Wire Map with the anchor nodes the L3 path needs."""
        return [
            # A compute node connecting to core-01 (gives us a core node)
            ["Yes", "su-01-node-01", "su-01-node-01", "NIC1_P1", None, None,
             "CPU/In-Band Network", None, None, None,
             "core-01", "core-01", "swp1s0"],
            # A server eth0 landing on oob-switch-01 (gives us the oob node)
            ["Yes", "su-01-node-01", "su-01-node-01", "eth0", None, None,
             "OOB Management", None, None, None,
             "oob-switch-01", "oob-switch-01", "swp1"],
            # cust-net-edge-01 and -02 link to cores (BGP EXIT VRF uplinks)
            ["Yes", "cust-net-edge-01", "cust-net-edge-01", "swp61", None, None,
             "ESL Uplink", None, None, None,
             "core-01", "core-01", "swp64s0"],
            ["Yes", "cust-net-edge-02", "cust-net-edge-02", "swp61", None, None,
             "ESL Uplink", None, None, None,
             "core-02", "core-02", "swp64s0"],
        ]

    def test_l3_mode_injects_three_ubuntu_nodes(self, tmp_path):
        """L3 mode produces external-conn, external-dhcp, utility."""
        wb = _build_wiremap_workbook(
            rows=self._l3_wiremap(),
            settings=[("oob_uplink_mode", "l3")],
        )
        path = _write_workbook(wb, tmp_path, arch="2-8-5-200")
        topo = TopologyGenerator(path, "2-8-5-200").generate()

        nodes = topo["content"]["nodes"]
        assert "external-conn" in nodes
        assert "external-dhcp" in nodes
        assert "utility" in nodes

    def test_l3_mode_does_not_inject_l2_infra(self, tmp_path):
        """L3 mode skips air-oob-switch / dhcp-oob / oob-server-01."""
        wb = _build_wiremap_workbook(
            rows=self._l3_wiremap(),
            settings=[("oob_uplink_mode", "l3")],
        )
        path = _write_workbook(wb, tmp_path, arch="2-8-5-200")
        topo = TopologyGenerator(path, "2-8-5-200").generate()

        nodes = topo["content"]["nodes"]
        assert "air-oob-switch" not in nodes
        assert "dhcp-oob" not in nodes
        assert "oob-server-01" not in nodes

    def test_l3_mode_external_conn_wired_to_edge_swp1(self, tmp_path):
        """external-conn:eth1 lands on cust-net-edge-01:swp1."""
        wb = _build_wiremap_workbook(
            rows=self._l3_wiremap(),
            settings=[("oob_uplink_mode", "l3")],
        )
        path = _write_workbook(wb, tmp_path, arch="2-8-5-200")
        topo = TopologyGenerator(path, "2-8-5-200").generate()

        links = topo["content"]["links"]
        found = False
        for link in links:
            if not (isinstance(link[0], dict) and isinstance(link[1], dict)):
                continue
            endpoints = {(e["node"], e["interface"]) for e in link}
            if ("external-conn", "eth1") in endpoints and \
               ("cust-net-edge-01", "swp1") in endpoints:
                found = True
                break
        assert found, "external-conn:eth1 should link to cust-net-edge-01:swp1"

    def test_l3_mode_external_dhcp_wired_to_edge_swp2(self, tmp_path):
        """external-dhcp:eth1 lands on cust-net-edge-01:swp2."""
        wb = _build_wiremap_workbook(
            rows=self._l3_wiremap(),
            settings=[("oob_uplink_mode", "l3")],
        )
        path = _write_workbook(wb, tmp_path, arch="2-8-5-200")
        topo = TopologyGenerator(path, "2-8-5-200").generate()

        links = topo["content"]["links"]
        found = False
        for link in links:
            if not (isinstance(link[0], dict) and isinstance(link[1], dict)):
                continue
            endpoints = {(e["node"], e["interface"]) for e in link}
            if ("external-dhcp", "eth1") in endpoints and \
               ("cust-net-edge-01", "swp2") in endpoints:
                found = True
                break
        assert found, "external-dhcp:eth1 should link to cust-net-edge-01:swp2"

    def test_l3_mode_switch_eth0s_on_edge_starting_swp3(self, tmp_path):
        """Cluster-switch eth0s land on cust-net-edge-01 at swp3 or higher."""
        wb = _build_wiremap_workbook(
            rows=self._l3_wiremap(),
            settings=[("oob_uplink_mode", "l3")],
        )
        path = _write_workbook(wb, tmp_path, arch="2-8-5-200")
        topo = TopologyGenerator(path, "2-8-5-200").generate()

        links = topo["content"]["links"]
        sw_eth0_to_edge = []
        for link in links:
            if not (isinstance(link[0], dict) and isinstance(link[1], dict)):
                continue
            for sw_ep, edge_ep in ((link[0], link[1]), (link[1], link[0])):
                if (sw_ep.get("interface") == "eth0"
                        and edge_ep.get("node") == "cust-net-edge-01"):
                    sw_eth0_to_edge.append(
                        (sw_ep["node"], edge_ep["interface"])
                    )
        # At least core-01 and oob-switch-01 should be there.
        sw_names = {n for n, _ in sw_eth0_to_edge}
        assert "core-01" in sw_names
        assert "oob-switch-01" in sw_names
        # Ports should all be swp3 or higher (swp1/swp2 reserved).
        for _, port in sw_eth0_to_edge:
            port_num = int(port.replace("swp", ""))
            assert port_num >= 3, f"switch eth0 on reserved port {port}"

    def test_l3_mode_utility_eth1_on_oob_switch_01(self, tmp_path):
        """utility:eth1 lands on oob-switch-01 (eth0 is reserved for outbound)."""
        wb = _build_wiremap_workbook(
            rows=self._l3_wiremap(),
            settings=[("oob_uplink_mode", "l3")],
        )
        path = _write_workbook(wb, tmp_path, arch="2-8-5-200")
        topo = TopologyGenerator(path, "2-8-5-200").generate()

        links = topo["content"]["links"]
        found = False
        for link in links:
            if not (isinstance(link[0], dict) and isinstance(link[1], dict)):
                continue
            endpoints = {(e["node"], e["interface"]) for e in link}
            for (n, i) in endpoints:
                if n == "utility" and i == "eth1":
                    other = next(((nn, ii) for nn, ii in endpoints if nn != "utility"), None)
                    if other and other[0] == "oob-switch-01":
                        found = True
                        break
        assert found, "utility:eth1 should link to oob-switch-01"

    def test_l3_mode_utility_eth0_outbound(self, tmp_path):
        """utility:eth0 → outbound (so Air can expose SSH/HTTP services)."""
        wb = _build_wiremap_workbook(
            rows=self._l3_wiremap(),
            settings=[("oob_uplink_mode", "l3")],
        )
        path = _write_workbook(wb, tmp_path, arch="2-8-5-200")
        topo = TopologyGenerator(path, "2-8-5-200").generate()

        links = topo["content"]["links"]
        outbound_nodes = set()
        for link in links:
            if isinstance(link[0], dict) and link[1] == "outbound":
                if link[0].get("interface") == "eth0":
                    outbound_nodes.add(link[0]["node"])
        assert "utility" in outbound_nodes

    def test_l3_mode_ubuntu_nodes_have_outbound_eth0(self, tmp_path):
        """external-conn and external-dhcp each have eth0 → outbound."""
        wb = _build_wiremap_workbook(
            rows=self._l3_wiremap(),
            settings=[("oob_uplink_mode", "l3")],
        )
        path = _write_workbook(wb, tmp_path, arch="2-8-5-200")
        topo = TopologyGenerator(path, "2-8-5-200").generate()

        links = topo["content"]["links"]
        outbound_nodes = set()
        for link in links:
            if isinstance(link[0], dict) and link[1] == "outbound":
                if link[0].get("interface") == "eth0":
                    outbound_nodes.add(link[0]["node"])
        assert "external-conn" in outbound_nodes
        assert "external-dhcp" in outbound_nodes

    def test_l3_mode_metadata_emitted(self, tmp_path):
        """Topology JSON includes _oob_uplink_mode and _l3_oob metadata."""
        wb = _build_wiremap_workbook(
            rows=self._l3_wiremap(),
            settings=[("oob_uplink_mode", "l3")],
        )
        path = _write_workbook(wb, tmp_path, arch="2-8-5-200")
        topo = TopologyGenerator(path, "2-8-5-200").generate()

        assert topo["_oob_uplink_mode"] == "l3"
        meta = topo["_l3_oob"]
        assert meta["jump_host"] == "utility"
        assert meta["nat_host"] == "external-conn"
        assert meta["dhcp_relay_server"] == "external-dhcp"
        assert meta["edge_switch"] == "cust-net-edge-01"

    def test_default_mode_is_l2(self, tmp_path):
        """No oob_uplink_mode setting → L2 mode (existing behavior)."""
        wb = _build_wiremap_workbook(rows=self._l3_wiremap(), settings=None)
        path = _write_workbook(wb, tmp_path, arch="2-8-5-200")
        topo = TopologyGenerator(path, "2-8-5-200").generate()

        assert topo["_oob_uplink_mode"] == "l2"
        assert "air-oob-switch" in topo["content"]["nodes"]
        assert "external-conn" not in topo["content"]["nodes"]

    def test_l2_mode_explicit(self, tmp_path):
        """Explicit oob_uplink_mode=l2 → L2 mode (unchanged behavior)."""
        wb = _build_wiremap_workbook(
            rows=self._l3_wiremap(),
            settings=[("oob_uplink_mode", "l2")],
        )
        path = _write_workbook(wb, tmp_path, arch="2-8-5-200")
        topo = TopologyGenerator(path, "2-8-5-200").generate()

        assert topo["_oob_uplink_mode"] == "l2"
        assert "air-oob-switch" in topo["content"]["nodes"]
        assert "external-conn" not in topo["content"]["nodes"]

    def test_l3_mode_skips_when_anchor_nodes_missing(self, tmp_path):
        """L3 mode with no cust-net-edge-01 in wiremap → warning, no injection."""
        # Wiremap with cores but NO cust-net-edge
        rows = [
            ["Yes", "su-01-node-01", "su-01-node-01", "NIC1_P1", None, None,
             "CPU/In-Band Network", None, None, None,
             "core-01", "core-01", "swp1s0"],
            ["Yes", "su-01-node-01", "su-01-node-01", "eth0", None, None,
             "OOB Management", None, None, None,
             "oob-switch-01", "oob-switch-01", "swp1"],
        ]
        wb = _build_wiremap_workbook(
            rows=rows,
            settings=[("oob_uplink_mode", "l3")],
        )
        path = _write_workbook(wb, tmp_path, arch="2-8-5-200")
        topo = TopologyGenerator(path, "2-8-5-200").generate()

        nodes = topo["content"]["nodes"]
        # No L3 trio because the anchor is missing
        assert "external-conn" not in nodes
        assert "external-dhcp" not in nodes
        assert "utility" not in nodes


# ---------------------------------------------------------------------------
# MAC consistency between topology_generator and excel_parser
# ---------------------------------------------------------------------------

class TestMacConsistency:
    """Verify MACs are deterministic and consistent across generators."""

    def test_mac_format(self):
        """All MACs start with 48:b0:2d: prefix."""
        mac = generate_mac("core-01", "swp1")
        assert mac.startswith("48:b0:2d:")

    def test_mac_deterministic(self):
        """Same inputs produce same MAC."""
        mac1 = generate_mac("core-01", "swp1")
        reset_mac_registry()
        mac2 = generate_mac("core-01", "swp1")
        assert mac1 == mac2

    def test_mac_unique_per_interface(self):
        """Different interfaces produce different MACs."""
        mac1 = generate_mac("core-01", "swp1")
        mac2 = generate_mac("core-01", "swp2")
        assert mac1 != mac2

    def test_mac_unique_per_node(self):
        """Different nodes produce different MACs."""
        mac1 = generate_mac("core-01", "eth0")
        mac2 = generate_mac("core-02", "eth0")
        assert mac1 != mac2


# ---------------------------------------------------------------------------
# NODE_DEFAULTS and SWITCH_MODELS constants
# ---------------------------------------------------------------------------

class TestConstants:
    """Verify topology generator constants are well-formed."""

    def test_switch_models_have_required_keys(self):
        for role, model in SWITCH_MODELS.items():
            assert "model" in model
            assert "ports" in model
            assert isinstance(model["ports"], int)
            assert model["ports"] > 0

    def test_node_defaults_have_required_keys(self):
        for role, defaults in NODE_DEFAULTS.items():
            assert "cpu" in defaults
            assert "memory" in defaults
            assert "storage" in defaults
            assert defaults["cpu"] > 0
            assert defaults["memory"] > 0

    def test_core_switch_is_sn5610(self):
        assert SWITCH_MODELS["core"]["model"] == "SN5610"
        assert SWITCH_MODELS["core"]["ports"] == 64

    def test_oob_switch_is_sn2201(self):
        assert SWITCH_MODELS["oob"]["model"] == "SN2201"
        assert SWITCH_MODELS["oob"]["ports"] == 48


# ---------------------------------------------------------------------------
# WiremapRow.enabled flag — agent-found C1/C3 regression coverage
# ---------------------------------------------------------------------------

class TestWiremapRowEnabled:
    """Tests for the `enabled` flag that retains disabled-node rows.

    The flag was added (after the Enabled=No filter regression) so that:
      - _build_nodes / _build_connected_links skip disabled rows (no phantom VM)
      - _analyze_breakouts and _build_unconnected_stubs see them
        (so the switch's subports still emit as stubs)
      - TopologyValidator honors the flag too (no false "Missing node")
    """

    def _wiremap_with_disabled_nodes(self):
        """Build a wiremap where su-01-node-01 is implicitly enabled (no Nodes
        entry, so not in disabled_names — survives) and another row that we'll
        mark disabled-equivalent by NOT including it in Nodes sheet at all.
        Disabled nodes are detected via parse_nodes — we simulate disabled by
        not adding entries here; the parser uses status="Disabled"."""
        # 2 cores + 4 OOB switches + some compute, mixed Enabled status
        rows = [
            # Active compute → topology should emit
            ["Yes", "su-01-node-01", "su-01-node-01", "NIC1_P1", None, None,
             "CPU/In-Band Network", None, None, None,
             "core-01", "core-01", "swp1s0"],
            # Disabled compute → MUST keep the switch port info (for breakout)
            # but skip the node itself in build_nodes
            ["Yes", "su-02-node-01", "su-02-node-01", "NIC1_P1", None, None,
             "CPU/In-Band Network", None, None, None,
             "core-01", "core-01", "swp2s0"],
        ]
        return rows

    def test_enabled_default_true(self):
        """WiremapRow.enabled defaults to True for fresh row construction."""
        r = WiremapRow(
            display_in_air=True, system_role="core", system_name="core-01",
            nic_port="swp1s0", net_profile="", switch_role="core",
            switch_name="core-01", switch_port="swp1s0",
        )
        assert r.enabled is True

    def test_enabled_can_be_false(self):
        """Constructing with enabled=False is allowed and tracked."""
        r = WiremapRow(
            display_in_air=True, system_role="compute", system_name="su-02-node-01",
            nic_port="NIC1_P1", net_profile="", switch_role="core",
            switch_name="core-01", switch_port="swp2s0", enabled=False,
        )
        assert r.enabled is False

    def test_parse_wiremap_marks_disabled_rows(self, tmp_path):
        """When the Nodes tab has Enabled=No, the parser still emits the row
        but with enabled=False (keeps switch_port for breakout analysis)."""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        # Settings
        wb.create_sheet("Settings")
        # Nodes: su-02-node-01 is Enabled=No
        ws_nodes = wb.create_sheet("Nodes")
        ws_nodes.append(["Function", "Name", "MAC", "Mgmt IP", "Prefix", "Gateway", "ZTP", "Enabled"])
        ws_nodes.append(["gpu", "su-01-node-01", "", "192.168.200.10", "24", "192.168.200.1", "Yes", "Yes"])
        ws_nodes.append(["gpu", "su-02-node-01", "", "192.168.200.20", "24", "192.168.200.1", "Yes", "No"])
        # Wire Map: both nodes have rows
        ws_wm = wb.create_sheet("Wire Map")
        ws_wm.append([
            "Display in Air", "System Role", "System Name", "NIC/Port",
            "Speed", "Description", "Network Profile",
            "Disabled by Neighbor", "Speed2", "Description2",
            "Switch Role", "Switch Name", "Switch Port",
        ])
        ws_wm.append(["Yes", "su-01-node-01", "su-01-node-01", "NIC1_P1", None, None,
                      "CPU/In-Band Network", None, None, None,
                      "core-01", "core-01", "swp1s0"])
        ws_wm.append(["Yes", "su-02-node-01", "su-02-node-01", "NIC1_P1", None, None,
                      "CPU/In-Band Network", None, None, None,
                      "core-01", "core-01", "swp2s0"])

        path = _write_workbook(wb, tmp_path, arch="2-4-3-200")
        rows = parse_wiremap_excel(path)

        # Find rows by system_name
        r1 = next(r for r in rows if r.system_name == "su-01-node-01")
        r2 = next(r for r in rows if r.system_name == "su-02-node-01")

        # Both rows are retained (not filtered out)
        assert r1.enabled is True
        assert r2.enabled is False
        # Switch ports are retained for both (so breakout analysis sees them)
        assert r1.switch_port == "swp1s0"
        assert r2.switch_port == "swp2s0"

    def test_topology_skips_disabled_node_but_keeps_switch_subport(self, tmp_path):
        """End-to-end: disabled node doesn't appear in nodes, but the switch
        subport it would have connected to still emits (as an unconnected stub)."""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        wb.create_sheet("Settings")
        ws_nodes = wb.create_sheet("Nodes")
        ws_nodes.append(["Function", "Name", "MAC", "Mgmt IP", "Prefix", "Gateway", "ZTP", "Enabled"])
        ws_nodes.append(["gpu", "su-01-node-01", "", "192.168.200.10", "24", "192.168.200.1", "Yes", "Yes"])
        ws_nodes.append(["gpu", "su-02-node-01", "", "192.168.200.20", "24", "192.168.200.1", "Yes", "No"])
        ws_wm = wb.create_sheet("Wire Map")
        ws_wm.append([
            "Display in Air", "System Role", "System Name", "NIC/Port",
            "Speed", "Description", "Network Profile",
            "Disabled by Neighbor", "Speed2", "Description2",
            "Switch Role", "Switch Name", "Switch Port",
        ])
        ws_wm.append(["Yes", "su-01-node-01", "su-01-node-01", "NIC1_P1", None, None,
                      "CPU/In-Band Network", None, None, None,
                      "core-01", "core-01", "swp1s0"])
        ws_wm.append(["Yes", "su-02-node-01", "su-02-node-01", "NIC1_P1", None, None,
                      "CPU/In-Band Network", None, None, None,
                      "core-01", "core-01", "swp2s0"])

        path = _write_workbook(wb, tmp_path, arch="2-4-3-200")
        topo = TopologyGenerator(path, "2-4-3-200").generate()
        nodes = topo["content"]["nodes"]
        # Active node is present
        assert "su-01-node-01" in nodes
        # Disabled node is NOT present (no phantom VM)
        assert "su-02-node-01" not in nodes
        # But swp2s0 still appears on core-01 (as an unconnected stub) so
        # the switch config can reference it without ifreload errors.
        core_ports = {ep["interface"]
                      for link in topo["content"]["links"]
                      if isinstance(link[0], dict) and link[0].get("node") == "core-01"
                      for ep in link if isinstance(ep, dict) and ep.get("node") == "core-01"}
        assert "swp2s0" in core_ports


# ---------------------------------------------------------------------------
# Mode-aware default behavior — parser + topology
# ---------------------------------------------------------------------------

class TestOobUplinkModeNormalization:
    """The parser's _normalize_oob_uplink_mode should handle typos cleanly."""

    def test_normalize_empty_defaults_to_l2(self):
        # Imported lazily to avoid circular issues
        import sys
        from pathlib import Path
        SCRIPTS = Path(__file__).parent.parent / "scripts"
        sys.path.insert(0, str(SCRIPTS))
        from excel_parser import _normalize_oob_uplink_mode

        assert _normalize_oob_uplink_mode({}) == "l2"
        assert _normalize_oob_uplink_mode({"oob_uplink_mode": ""}) == "l2"
        assert _normalize_oob_uplink_mode({"oob_uplink_mode": None}) == "l2"

    def test_normalize_known_modes(self):
        from excel_parser import _normalize_oob_uplink_mode
        assert _normalize_oob_uplink_mode({"oob_uplink_mode": "l2"}) == "l2"
        assert _normalize_oob_uplink_mode({"oob_uplink_mode": "l3"}) == "l3"
        assert _normalize_oob_uplink_mode({"oob_uplink_mode": "L3"}) == "l3"
        assert _normalize_oob_uplink_mode({"oob_uplink_mode": " l3 "}) == "l3"

    def test_normalize_typos_fall_back_to_l2(self, capsys):
        from excel_parser import _normalize_oob_uplink_mode
        assert _normalize_oob_uplink_mode({"oob_uplink_mode": "LAYER3"}) == "l2"
        assert _normalize_oob_uplink_mode({"oob_uplink_mode": "l2/l3"}) == "l2"
        # Warning should be emitted (operator typo signal)
        captured = capsys.readouterr()
        assert "not in" in captured.out


class TestParseCidr:
    """Helper used to safely parse CIDR-style subnet strings."""

    def test_valid_cidr(self):
        import sys
        from pathlib import Path
        SCRIPTS = Path(__file__).parent.parent / "scripts"
        sys.path.insert(0, str(SCRIPTS))
        from excel_parser import _parse_cidr
        assert _parse_cidr("192.168.200.0/24") == ("192.168.200.0", 24)
        assert _parse_cidr("10.0.0.0/8") == ("10.0.0.0", 8)

    def test_strips_whitespace(self):
        from excel_parser import _parse_cidr
        assert _parse_cidr(" 192.168.200.0/24 ") == ("192.168.200.0", 24)

    def test_missing_slash_returns_none(self):
        from excel_parser import _parse_cidr
        assert _parse_cidr("192.168.200.0") is None
        assert _parse_cidr("garbage") is None

    def test_bad_prefix_returns_none(self):
        from excel_parser import _parse_cidr
        assert _parse_cidr("192.168.0.0/notnum") is None
        assert _parse_cidr("192.168.0.0/-1") is None
        assert _parse_cidr("192.168.0.0/33") is None

    def test_empty_or_none_returns_none(self):
        from excel_parser import _parse_cidr
        assert _parse_cidr("") is None
        assert _parse_cidr(None) is None


class TestResolveInfraNodes:
    """air-deploy.py's mode-aware resolver returns the right infra-node names."""

    def _resolver(self):
        import importlib.util
        from pathlib import Path
        spec = importlib.util.spec_from_file_location(
            "air_deploy_mod",
            Path(__file__).parent.parent / "scripts" / "air-deploy.py",
        )
        mod = importlib.util.module_from_spec(spec)
        # The module imports httpx etc — best-effort import; skip if env unavailable.
        try:
            spec.loader.exec_module(mod)
        except ImportError as exc:
            pytest.skip(f"air-deploy.py cannot be imported in this test env: {exc}")
        return mod._resolve_infra_nodes

    def test_resolves_l2(self):
        resolve = self._resolver()
        result = resolve({"_oob_uplink_mode": "l2"})
        assert result["mode"] == "l2"
        assert result["jump_host"] == "dhcp-oob"
        assert result["nat_host"] == "dhcp-oob"
        assert result["dhcp_server"] == "dhcp-oob"
        assert result["status_page_host"] == "dhcp-oob"
        assert "oob-server-01" in result["ssh_service_nodes"]
        assert "dhcp-oob" in result["ssh_service_nodes"]
        assert result["air_bridge"] == "air-oob-switch"

    def test_resolves_l3(self):
        resolve = self._resolver()
        result = resolve({"_oob_uplink_mode": "l3"})
        assert result["mode"] == "l3"
        assert result["jump_host"] == "utility"
        assert result["nat_host"] == "external-conn"
        assert result["dhcp_server"] == "external-dhcp"
        assert result["status_page_host"] == "utility"
        assert result["air_bridge"] is None
        assert set(result["ssh_service_nodes"]) == {"utility", "external-conn", "external-dhcp"}

    def test_defaults_to_l2_on_missing_key(self):
        """Older topology JSONs (pre-mode-flag) don't have _oob_uplink_mode."""
        resolve = self._resolver()
        result = resolve({})  # no _oob_uplink_mode key
        assert result["mode"] == "l2"


class TestSwitchNIPasswordNotLeaked:
    """Regression for the chpasswd xtrace leak in build_switch_ni_commands.

    The first-boot apply.sh runs under `set -ex`, and the systemd unit pipes
    StandardError to the journal. Before the fix, `set -x` echoed the
    `echo cumulus:<password> | chpasswd` line verbatim to stderr, leaking the
    plaintext switch password into the journal. The fix brackets the chpasswd
    line with `set +x` / `set -x`.

    These tests prove three things at once, so a future edit can't silently
    regress any of them:
      1. the password is STILL set (chpasswd receives `cumulus:<pw>`),
      2. the plaintext password no longer appears in the xtrace stream, and
      3. errexit (`set -e`) is still in force.
    """

    # A password with shell-hostile chars: must survive shlex.quote AND must
    # not leak. Exercises the real-world apostrophe/space/bang/at edge cases.
    PASSWORD = "S3cr3t P@ss'w0rd!"

    def _builder(self):
        import importlib.util
        spec = importlib.util.spec_from_file_location(
            "air_deploy_mod_ni",
            Path(__file__).parent.parent / "scripts" / "air-deploy.py",
        )
        mod = importlib.util.module_from_spec(spec)
        try:
            spec.loader.exec_module(mod)
        except ImportError as exc:
            pytest.skip(f"air-deploy.py cannot be imported in this test env: {exc}")
        return mod.build_switch_ni_commands

    def _decode_apply_script(self, commands):
        """Pull the base64-encoded apply.sh out of the NI command list."""
        import base64
        line = next(
            c for c in commands
            if "> /opt/era/apply.sh" in c and "base64 -d" in c
        )
        b64 = re.search(r"echo '([^']+)'", line).group(1)
        return base64.b64decode(b64).decode()

    def test_apply_script_brackets_chpasswd_with_xtrace_off(self):
        """Static structural check: the chpasswd line sits between `set +x`
        and `set -x`, and the script still opens with errexit."""
        build = self._builder()
        cmds = build("oob-switch-01", "# nvue config\n", "172.20.0.5", self.PASSWORD)
        script = self._decode_apply_script(cmds)

        assert script.startswith("#!/bin/bash\nset -ex\n"), \
            "apply.sh must still run with errexit (set -e)"

        lines = script.splitlines()
        chpasswd_idx = next(i for i, l in enumerate(lines) if "chpasswd" in l)
        assert lines[chpasswd_idx - 1] == "set +x", \
            f"chpasswd must be immediately preceded by `set +x`; got {lines[chpasswd_idx - 1]!r}"
        assert lines[chpasswd_idx + 1] == "set -x", \
            f"xtrace must be re-enabled immediately after chpasswd; got {lines[chpasswd_idx + 1]!r}"

    def test_apply_script_is_valid_bash(self):
        """`bash -n` parse check — the set +x/set -x wrap can't break syntax."""
        import subprocess
        build = self._builder()
        cmds = build("oob-switch-01", "# nvue config\n", "172.20.0.5", self.PASSWORD)
        script = self._decode_apply_script(cmds)
        proc = subprocess.run(
            ["bash", "-n"], input=script, capture_output=True, text=True
        )
        assert proc.returncode == 0, f"apply.sh failed bash -n: {proc.stderr}"

    def test_password_set_but_not_traced_when_run(self, tmp_path):
        """Behavioral end-to-end: run apply.sh under `bash -x` with chpasswd /
        nv / systemctl / ip stubbed. Assert the password reaches chpasswd but
        never appears in the xtrace stderr."""
        import os
        import subprocess
        import shutil

        if shutil.which("bash") is None:
            pytest.skip("bash unavailable")

        build = self._builder()
        cmds = build("oob-switch-01", "# nvue config\n", "172.20.0.5", self.PASSWORD)
        script = self._decode_apply_script(cmds)

        # Redirect the hardcoded absolute paths into tmp_path so we don't need
        # root, and so the sourced config file exists.
        era_dir = tmp_path / "opt-era"
        era_dir.mkdir()
        (era_dir / "oob-switch-01-config.sh").write_text("# stub nvue config\n")
        script = script.replace("/opt/era", str(era_dir))

        # Stub every external command the script calls. chpasswd records the
        # user:password pair it receives so we can prove the password was set.
        stub_dir = tmp_path / "bin"
        stub_dir.mkdir()
        marker = tmp_path / "chpasswd.out"
        (stub_dir / "chpasswd").write_text(
            f'#!/bin/bash\ncat > {marker}\n'
        )
        for name in ("nv", "systemctl", "ip"):
            (stub_dir / name).write_text("#!/bin/bash\nexit 0\n")
        for f in stub_dir.iterdir():
            f.chmod(0o755)

        env = dict(os.environ, PATH=f"{stub_dir}:{os.environ['PATH']}")
        proc = subprocess.run(
            ["bash", "-x"], input=script, capture_output=True, text=True, env=env
        )

        assert proc.returncode == 0, (
            f"apply.sh exited {proc.returncode}\nSTDERR:\n{proc.stderr}"
        )
        # 1. Password was actually set.
        assert marker.read_text().strip() == f"cumulus:{self.PASSWORD}", \
            "chpasswd did not receive the expected cumulus:<password> pair"
        # 2. Password did NOT leak into the xtrace stream (the journal target).
        assert self.PASSWORD not in proc.stderr, (
            "plaintext password leaked into xtrace stderr — the set +x/-x "
            "wrap around chpasswd has regressed"
        )
