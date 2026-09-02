# SPDX-FileCopyrightText: Copyright (c) 2026 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: MIT
"""management_switches is retired as a Settings key.

The OOB switch count was already derived from the Nodes tab's oob-switch
rows whenever management_switches was absent or <= 0
(get_oob_nodes_for_inventory's fallback). This task removes the Settings
key entirely: it is no longer REQUIRED, its presence no longer drives or
truncates/pads the OOB switch count, and a stale present value now produces
a soft WARN (not an error) rather than integer-range validation.

management_switches stays in OPTIONAL_SETTINGS_KEYS (not REQUIRED, not
removed from the known-keys union) so its presence is flagged with our
specific "ignored" WARN instead of a generic "unknown key" warning —
mirrors the mgmt_subnets retirement pattern in test_mgmt_subnets_optional.py.
"""
import sys
from pathlib import Path

import openpyxl

SCRIPTS_DIR = Path(__file__).parent.parent / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from validate_excel import (
    validate_settings,
    ValidationResult,
    REQUIRED_SETTINGS_KEYS,
    OPTIONAL_SETTINGS_KEYS,
)


def _settings(pairs):
    """Build a Settings sheet from (key, value) pairs."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Settings"
    ws.cell(row=1, column=1, value="Setting")
    ws.cell(row=1, column=2, value="Value")
    for r, (k, v) in enumerate(pairs, start=2):
        ws.cell(row=r, column=1, value=k)
        ws.cell(row=r, column=2, value=v)
    return ws


def _run(pairs):
    result = ValidationResult()
    validate_settings(_settings(pairs), result)
    return result


_OTHER_REQUIRED_KEYS = [
    ("architecture", "2-4-3-200"),
    ("bgp_asn", 65000),
    ("loopback_base", "172.16.176"),
]


# ─── Membership: management_switches moved from REQUIRED to OPTIONAL ─────

def test_management_switches_not_in_required_keys():
    assert "management_switches" not in REQUIRED_SETTINGS_KEYS


def test_management_switches_in_optional_keys():
    assert "management_switches" in OPTIONAL_SETTINGS_KEYS


# ─── Behavioral: absent management_switches no longer errors ─────────────

def test_workbook_without_management_switches_has_no_missing_key_error():
    result = _run(_OTHER_REQUIRED_KEYS)
    assert not any(
        "management_switches" in e and "Missing required key" in e
        for e in result.errors
    ), result.errors


def test_workbook_without_management_switches_is_clean():
    result = _run(_OTHER_REQUIRED_KEYS)
    assert not any("management_switches" in e for e in result.errors), result.errors
    assert not any("management_switches" in w for w in result.warnings), result.warnings


# ─── Behavioral: presence of management_switches is a soft WARN, not error ─

def test_workbook_with_management_switches_present_warns_not_errors():
    pairs = _OTHER_REQUIRED_KEYS + [("management_switches", 2)]
    result = _run(pairs)
    assert not any("management_switches" in e for e in result.errors), result.errors
    assert any("management_switches is ignored" in w for w in result.warnings), result.warnings


def test_workbook_with_invalid_management_switches_value_still_only_warns():
    # Previously this shape (non-integer / out-of-range) triggered hard
    # integer-range errors. Now presence alone is the only thing checked —
    # the value's shape is irrelevant since it's ignored entirely.
    pairs = _OTHER_REQUIRED_KEYS + [("management_switches", "not-a-number")]
    result = _run(pairs)
    assert not any("management_switches" in e for e in result.errors), result.errors
    assert any("management_switches is ignored" in w for w in result.warnings), result.warnings


def test_management_switches_present_is_not_flagged_as_unknown_key():
    result = _run(_OTHER_REQUIRED_KEYS + [("management_switches", 2)])
    assert not any(
        "Unknown key 'management_switches'" in w for w in result.warnings
    ), result.warnings


def test_management_switches_not_coerced_as_int_type_check():
    # management_switches must no longer be subject to the R4-07
    # integer-type-strictness coercion check (it's ignored, not consumed).
    pairs = _OTHER_REQUIRED_KEYS + [("management_switches", "2")]
    result = _run(pairs)
    assert not any(
        "management_switches" in e and "must be a positive integer" in e
        for e in result.errors
    ), result.errors
