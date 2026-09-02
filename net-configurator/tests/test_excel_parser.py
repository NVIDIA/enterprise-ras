# SPDX-FileCopyrightText: Copyright (c) 2026 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: MIT
"""
Edge case tests for the ERA Excel parser.

Tests parse_settings(), parse_nodes(), parse_vlans(), and wiremap-related
functions with in-memory openpyxl workbooks to verify error handling,
empty data, and valid complete input.
"""
import sys
from pathlib import Path

import openpyxl
import pytest
import yaml

# Add scripts/ to path so we can import the modules
SCRIPTS_DIR = Path(__file__).parent.parent / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from utils import reset_mac_registry
from excel_parser import (
    parse_settings,
    parse_nodes,
    parse_vlans,
    parse_vrfs,
    parse_versions,
    parse_oob_switch_configs,
    parse_core_port_config,
    build_devices,
    classify_host_role,
    _build_wiremap_row_list,
    generate_group_vars,
    generate_prefix_lists,
    categorize_nodes,
    segment_esi_for_node,
    build_per_switch_direct_interfaces,
    build_per_switch_server_roles,
    build_per_switch_gpu_rail_interfaces,
    process_excel_template,
)


def _expand_port_range(range_str):
    """Expand a port range string like 'swp1-3,swp5' into a set of port names."""
    ports = set()
    for token in (range_str or "").split(","):
        token = token.strip()
        if not token:
            continue
        if "-" in token:
            prefix = token.split("-")[0].rstrip("0123456789")
            start = int(token.split("-")[0][len(prefix):])
            end = int(token.split("-")[1])
            for n in range(start, end + 1):
                ports.add(f"{prefix}{n}")
        else:
            ports.add(token)
    return ports


# ---------------------------------------------------------------------------
# Helpers — build minimal workbooks for testing
# ---------------------------------------------------------------------------

def _make_settings_sheet(wb, data=None):
    """Add a Settings sheet with optional key-value rows."""
    ws = wb.create_sheet("Settings")
    if data:
        for key, value in data:
            ws.append([key, value])
    return ws


def _make_nodes_sheet(wb, headers=None, rows=None):
    """Add a Nodes sheet with headers and data rows."""
    ws = wb.create_sheet("Nodes")
    if headers is None:
        headers = ["Function", "Name", "MAC Address for ZTP", "Mgmt IP Address",
                    "Prefix", "Gateway"]
    ws.append(headers)
    if rows:
        for row in rows:
            ws.append(row)
    return ws


def _make_vlans_sheet(wb, vlans=None):
    """Add a VLANs & Profiles sheet with headers and data rows.

    Row 1 = section title, Row 2 = column headers, Row 3+ = data.
    """
    ws = wb.create_sheet("VLANs & Profiles")
    ws.append(["VLANs"])  # row 1: section label
    ws.append(["VLAN_ID", "Name", "Purpose", "Subnet", "VRF"])  # row 2: headers
    if vlans:
        for vlan in vlans:
            ws.append(vlan)
    return ws


def _make_vlans_profiles_sheet(wb, *, vlans=None, profiles=None):
    """Add VLANs & Profiles with VLAN and Port Profiles sections."""
    ws = wb.create_sheet("VLANs & Profiles")
    ws.append(["VLANs"])
    ws.append(["VLAN_ID", "Name", "Purpose", "Subnet", "VRF"])
    for vlan in vlans or []:
        ws.append(vlan)
    ws.append([])
    ws.append(["Port Profiles"])
    ws.append([
        "Profile", "Port Mode", "Native/Access VLAN", "Allowed VLANs",
        "Untagged VLAN", "VRF", "LACP Bypass", "Speed", "Breakout",
        "Lanes",
    ])
    for profile in profiles or []:
        ws.append(profile)
    return ws


def _make_wiremap_sheet(wb, rows=None):
    """Add a Wire Map sheet with 13-column header and data rows."""
    ws = wb.create_sheet("Wire Map")
    header = [
        "Display in Air", "System Role", "System Name", "NIC/Port",
        "Speed", "Description", "Network Profile",
        "Disabled by Neighbor", "Speed2", "Description2",
        "Switch Role", "Switch Name", "Switch Port",
    ]
    ws.append(header)
    if rows:
        for row in rows:
            # Pad short rows
            padded = list(row) + [None] * (13 - len(row))
            ws.append(padded[:13])
    return ws


@pytest.fixture(autouse=True)
def _reset_mac_reg():
    """Reset MAC registry between tests to avoid cross-test collision errors."""
    reset_mac_registry()
    yield
    reset_mac_registry()


# ---------------------------------------------------------------------------
# Per-node ESI + per-switch server-role emission (dedicated cl tier)
# ---------------------------------------------------------------------------

class TestPerNodeESIandPerSwitchRoles:
    def test_segment_esi_is_per_node_and_class_distinct(self):
        # Same node -> same ESI regardless of which leaf/port it lands on.
        assert segment_esi_for_node("su-01-node-06") == segment_esi_for_node("SU-01-NODE-06")
        # Distinct nodes -> distinct ESIs; distinct endpoint classes -> distinct bands.
        ids = {
            segment_esi_for_node("su-01-node-06"),
            segment_esi_for_node("su-02-node-06"),
            segment_esi_for_node("storage-02"),
            segment_esi_for_node("support-04"),
        }
        assert len(ids) == 4
        assert segment_esi_for_node("su-01-node-06") == 1006

    def test_multihomed_node_gets_matching_esi_across_leaves(self):
        """A node dual-homed to two cl leaves on DIFFERENT ports still gets the
        same ESI on both (the cl-07 MH bug). Storage 4-way likewise. And no
        switch emits a bond for a sub-port it does not cable."""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        rows = [
            # su-01-node-06 dual-homed: cl-06 swp1s1, cl-07 swp1s0 (different ports)
            ["Yes", "compute", "su-01-node-06", "B3240 P1", None, None,
             "CPU/In-Band Network", None, None, None, "cl", "cl-06", "swp1s1"],
            ["Yes", "compute", "su-01-node-06", "B3240 P2", None, None,
             "CPU/In-Band Network", None, None, None, "cl", "cl-07", "swp1s0"],
            # storage-02 multi-homed: cl-01 swp12s1, cl-05 swp12s0
            ["Yes", "storage", "storage-02", "STOR P1", None, None,
             "Storage", None, None, None, "cl", "cl-01", "swp12s1"],
            ["Yes", "storage", "storage-02", "STOR P2", None, None,
             "Storage", None, None, None, "cl", "cl-05", "swp12s0"],
            # csl-named leaf: same 'csl' role as cl-*, so on a dedicated tier it
            # is treated identically (gating is by role+tier, never by name).
            ["Yes", "compute", "su-01-node-09", "B3240 P1", None, None,
             "CPU/In-Band Network", None, None, None, "csl", "csl-01", "swp1s0"],
        ]
        _make_wiremap_sheet(wb, rows)
        agg = {"cpu": {"breakout": 4, "lanes": 2, "vlan": 300},
               "storage": {"breakout": 4, "vlan": 500}}
        # Per-switch server bonds/ESI only exist on a DEDICATED N/S tier
        # (ns_tiers > 1); this multi-leaf cl-06/cl-07 fixture is exactly that.
        # Pass a realistic nodes_function_map (switch -> Function) exactly as the
        # production parser does: a split leaf's Function is the bare 'cl'. This
        # guards the regression where the leaf gate classified the Function value
        # ('cl' -> 'cl' != 'csl' -> skip every leaf -> phantom bonds) instead of
        # the hostname. The gate must include cl leaves regardless of this map.
        nfm = {"cl-01": "cl", "cl-05": "cl", "cl-06": "cl",
               "cl-07": "cl", "csl-01": "csl"}
        per_switch = build_per_switch_server_roles(
            wb["Wire Map"], agg, nodes_function_map=nfm, dedicated_ns_tier=True)

        # cl-06 and cl-07 both carry su-01-node-06's ESI (1006), on their own port.
        esi_06 = per_switch["cl-06"]["cpu"]["bond_overrides"]["bond1s1"]["segment_id"]
        esi_07 = per_switch["cl-07"]["cpu"]["bond_overrides"]["bond1s0"]["segment_id"]
        assert esi_06 == esi_07 == 1006

        # 4-way storage: same ESI on each homing leaf, on whichever port.
        s01 = per_switch["cl-01"]["storage"]["bond_overrides"]["bond12s1"]["segment_id"]
        s05 = per_switch["cl-05"]["storage"]["bond_overrides"]["bond12s0"]["segment_id"]
        assert s01 == s05 == segment_esi_for_node("storage-02")

        # Per-switch emission: cl-07 only references its own cabled sub-port.
        assert per_switch["cl-07"]["cpu"]["ports"] == [1]
        assert per_switch["cl-07"]["cpu"]["port_overrides"][1]["subports"] == [0]
        assert "bond1s1" not in per_switch["cl-07"]["cpu"]["bond_overrides"]

        # Name-independence: a csl-* leaf is the same 'csl' role as cl-* and is
        # included on a dedicated tier (this was the cl-vs-csl name-gating bug).
        assert "csl-01" in per_switch

        # Tier-gating: on a CONVERGED tier (ns_tiers == 1) NO leaf gets
        # per-switch roles — the shared group network_roles / golden ESI is kept.
        assert build_per_switch_server_roles(
            wb["Wire Map"], agg, dedicated_ns_tier=False) == {}

    def test_multi_slot_bonds_get_distinct_esi(self):
        """ADR-0061 quad-connected support servers ("ConnectX-7 SL1 P1/P2" +
        "SL3 P1/P2") must get two DISTINCT ESIs, one per adapter slot.

        nic_num alone (the trailing digit) can't tell them apart: "SL1 P1" and
        "SL3 P1" both end in "1", so both bonds landed on bond_idx 0 -> the
        SAME ESI on cl-01/cl-02 -> zebra rejected the second bond with "ESI
        already exists on a different interface" -> nv config apply failed on
        every cl-* leaf carrying a support bond -> the whole converged-leaf
        tier never configured. Reproduced live on 2-4-5-800/largescale (148
        unreachable, cross-ping 0/3) and confirmed on real switch syslog.
        """
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        rows = [
            ["Yes", "support", "support-01", "ConnectX-7 SL1 P1", None, None,
             "Support", None, None, None, "cl", "cl-01", "swp22s0"],
            ["Yes", "support", "support-01", "ConnectX-7 SL1 P2", None, None,
             "Support", None, None, None, "cl", "cl-02", "swp22s0"],
            ["Yes", "support", "support-01", "ConnectX-7 SL3 P1", None, None,
             "Support", None, None, None, "cl", "cl-01", "swp22s1"],
            ["Yes", "support", "support-01", "ConnectX-7 SL3 P2", None, None,
             "Support", None, None, None, "cl", "cl-02", "swp22s1"],
        ]
        _make_wiremap_sheet(wb, rows)
        agg = {"support": {"breakout": 4, "vlan": 400}}
        nfm = {"cl-01": "cl", "cl-02": "cl"}
        per_switch = build_per_switch_server_roles(
            wb["Wire Map"], agg, nodes_function_map=nfm, dedicated_ns_tier=True)

        sl1 = per_switch["cl-01"]["support"]["bond_overrides"]["bond22s0"]["segment_id"]
        sl3 = per_switch["cl-01"]["support"]["bond_overrides"]["bond22s1"]["segment_id"]
        assert sl1 == 800001
        assert sl3 == 1800001

        # Still identical across the mirrored leaf for EACH bond (that's what
        # makes EVPN-MH work at all).
        assert sl1 == per_switch["cl-02"]["support"]["bond_overrides"]["bond22s0"]["segment_id"]
        assert sl3 == per_switch["cl-02"]["support"]["bond_overrides"]["bond22s1"]["segment_id"]

    def test_no_switch_has_duplicate_esi_across_bonds(self):
        """Generator-side guard: no two bond interfaces on the same switch may
        share an `evpn multihoming segment local-id`. FRR treats a duplicate
        ESI as a hard config-apply failure (zebra: "ESI already exists on a
        different interface"), so this must never regenerate. Covers every
        `bond_overrides` entry produced for the sample mixed compute/storage/
        multi-slot-support fixture used elsewhere in this file.
        """
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        rows = [
            ["Yes", "compute", "su-01-node-06", "B3240 P1", None, None,
             "CPU/In-Band Network", None, None, None, "cl", "cl-01", "swp1s0"],
            ["Yes", "compute", "su-01-node-06", "B3240 P2", None, None,
             "CPU/In-Band Network", None, None, None, "cl", "cl-02", "swp1s0"],
            ["Yes", "support", "support-01", "ConnectX-7 SL1 P1", None, None,
             "Support", None, None, None, "cl", "cl-01", "swp22s0"],
            ["Yes", "support", "support-01", "ConnectX-7 SL1 P2", None, None,
             "Support", None, None, None, "cl", "cl-02", "swp22s0"],
            ["Yes", "support", "support-01", "ConnectX-7 SL3 P1", None, None,
             "Support", None, None, None, "cl", "cl-01", "swp22s1"],
            ["Yes", "support", "support-01", "ConnectX-7 SL3 P2", None, None,
             "Support", None, None, None, "cl", "cl-02", "swp22s1"],
        ]
        _make_wiremap_sheet(wb, rows)
        agg = {"cpu": {"breakout": 4, "lanes": 2, "vlan": 300},
               "support": {"breakout": 4, "vlan": 400}}
        nfm = {"cl-01": "cl", "cl-02": "cl"}
        per_switch = build_per_switch_server_roles(
            wb["Wire Map"], agg, nodes_function_map=nfm, dedicated_ns_tier=True)

        for sw, roles in per_switch.items():
            seen = {}
            for role in roles.values():
                for bond_name, override in role.get("bond_overrides", {}).items():
                    esi = override["segment_id"]
                    assert esi not in seen, (
                        f"{sw}: duplicate ESI {esi} on {bond_name!r} and "
                        f"{seen.get(esi)!r}"
                    )
                    seen[esi] = bond_name

    def test_direct_interfaces_are_scoped_to_each_leaf(self):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        rows = [
            ["Yes", "cl", "cl-01", "swp25s0", None, None,
             "ISL", None, None, None, "cs", "cs-01", "swp1s0"],
            ["Yes", "cl", "cl-02", "swp25s1", None, None,
             "ISL", None, None, None, "cs", "cs-01", "swp1s1"],
            ["Yes", "edge", "cust-net-edge-01", "swp1", None, None,
             "Edge Uplink", None, None, None, "cl", "cl-01", "swp63s0"],
            ["Yes", "edge", "cust-net-edge-01", "swp2", None, None,
             "Edge Uplink", None, None, None, "cl", "cl-02", "swp64s1"],
            ["Yes", "oob-switch", "oob-switch-01", "swp49", None, None,
             "OOB Uplink", None, None, None, "cl", "cl-01", "swp61s0"],
        ]
        _make_wiremap_sheet(wb, rows)
        aggregated = {
            "isl_interfaces": {
                "ports": [25], "breakout": 2, "lanes": 4,
                "port_overrides": {},
            },
            "edge_interfaces": {
                "ports": [63, 64], "breakout": 2, "lanes": 4,
                "port_overrides": {},
            },
            "oob_uplink_interfaces": {
                "ports": [61], "breakout": 8, "lanes": 1,
                "port_overrides": {61: {"subports": [0]}},
            },
        }
        result = build_per_switch_direct_interfaces(
            wb["Wire Map"],
            aggregated,
            nodes_function_map={"cl-01": "cl", "cl-02": "cl", "cs-01": "cs"},
            oob_uplink_mode="l3",
        )

        assert result["cl-01"]["isl_interfaces"]["port_overrides"][25]["subports"] == [0]
        assert result["cl-02"]["isl_interfaces"]["port_overrides"][25]["subports"] == [1]
        assert result["cl-01"]["edge_interfaces"]["ports"] == [63]
        assert result["cl-02"]["edge_interfaces"]["ports"] == [64]
        assert result["cl-01"]["oob_uplink_interfaces"]["ports"] == [61]
        assert result["cl-02"]["oob_uplink_interfaces"]["ports"] == []


# ---------------------------------------------------------------------------
# parse_settings
# ---------------------------------------------------------------------------

class TestParseSettings:
    """Tests for parse_settings() edge cases."""

    def test_empty_sheet(self):
        """Empty Settings sheet returns empty dict."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Settings"
        result = parse_settings(ws)
        assert result == {}

    def test_valid_settings(self):
        """Key-value pairs are correctly parsed and snake_cased."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Settings"
        ws.append(["Site Name", "test-site"])
        ws.append(["Domain Name", "example.com"])
        ws.append(["Mgmt Subnets", "192.168.200.0/24"])

        result = parse_settings(ws)
        assert result["site_name"] == "test-site"
        assert result["domain_name"] == "example.com"
        assert result["mgmt_subnets"] == "192.168.200.0/24"

    def test_none_value_skipped(self):
        """Rows with None values are skipped."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Settings"
        ws.append(["Valid Key", "some_value"])
        ws.append(["Missing Value", None])

        result = parse_settings(ws)
        assert "valid_key" in result
        assert "missing_value" not in result

    def test_none_key_skipped(self):
        """Rows with None keys are skipped."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Settings"
        ws.append([None, "orphan_value"])
        ws.append(["Good Key", "good_value"])

        result = parse_settings(ws)
        assert len(result) == 1
        assert "good_key" in result

    def test_key_normalization(self):
        """Keys with spaces and hyphens become snake_case."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Settings"
        ws.append(["Air URL", "https://air.nvidia.com"])
        ws.append(["NTP-Server", "pool.ntp.org"])

        result = parse_settings(ws)
        assert "air_url" in result
        assert "ntp_server" in result

    def test_numeric_values_preserved(self):
        """Numeric values are preserved (not stringified by parse_settings)."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Settings"
        ws.append(["VLAN Count", 10])

        result = parse_settings(ws)
        assert result["vlan_count"] == 10


# ---------------------------------------------------------------------------
# parse_nodes
# ---------------------------------------------------------------------------

class TestParseNodes:
    """Tests for parse_nodes() edge cases."""

    def test_empty_nodes_sheet(self):
        """Nodes sheet with only header returns empty list."""
        wb = openpyxl.Workbook()
        ws = _make_nodes_sheet(wb)
        result = parse_nodes(ws)
        assert result == []

    def test_valid_nodes(self):
        """Complete rows are parsed correctly."""
        wb = openpyxl.Workbook()
        ws = _make_nodes_sheet(wb, rows=[
            ["core-01", "core-01", "48:b0:2d:aa:bb:cc", "192.168.200.2", 24, "192.168.200.1"],
            ["su-01-node-01", "su-01-node-01", "", "192.168.200.11", 24, "192.168.200.1"],
        ])

        result = parse_nodes(ws)
        assert len(result) == 2
        assert result[0]["role"] == "core-01"
        assert result[0]["name"] == "core-01"
        assert result[0]["mac_address"] == "48:b0:2d:aa:bb:cc"
        assert result[0]["mgmt_ip"] == "192.168.200.2"

    def test_missing_mac_empty_string(self):
        """Missing MAC is returned as empty string, not None."""
        wb = openpyxl.Workbook()
        ws = _make_nodes_sheet(wb, rows=[
            ["core-02", "core-02", None, "192.168.200.3", 24, "192.168.200.1"],
        ])

        result = parse_nodes(ws)
        assert result[0]["mac_address"] == ""

    def test_missing_name_falls_back_to_role(self):
        """If Name column is empty, it falls back to the Function column."""
        wb = openpyxl.Workbook()
        ws = _make_nodes_sheet(wb, rows=[
            ["core-01", None, "", "192.168.200.2", 24, "192.168.200.1"],
        ])

        result = parse_nodes(ws)
        assert result[0]["name"] == "core-01"

    def test_skip_rows_without_function(self):
        """Rows with no Function value are skipped."""
        wb = openpyxl.Workbook()
        ws = _make_nodes_sheet(wb, rows=[
            [None, "orphan-name", "", "10.0.0.1", 24, "10.0.0.1"],
            ["core-01", "core-01", "", "192.168.200.2", 24, "192.168.200.1"],
        ])

        result = parse_nodes(ws)
        assert len(result) == 1
        assert result[0]["role"] == "core-01"


class TestGenerateGroupVars:
    """Tests for generated group_vars behavior across source-inventory merges."""

    def test_source_inventory_link_state_is_not_inherited(self, tmp_path):
        """Source inventory must not silently inject link-state defaults."""
        settings = {"architecture": "2-8-5-200"}
        vlans = [
            {
                "id": 200,
                "name": "OOB",
                "subnet": "192.168.200.0/24",
                "vrf": "OOB",
                "vni": 4200,
            },
            {
                "id": 300,
                "name": "CPU/In-Band",
                "subnet": "172.16.178.0/24",
                "vrf": "INBAND",
                "vni": 4300,
            },
        ]
        vrfs = {
            "OOB": {"name": "OOB", "l3_vni": 5001},
            "INBAND": {"name": "INBAND", "l3_vni": 5002},
        }
        output_dir = tmp_path / "inventory"
        output_dir.mkdir()

        generate_group_vars(
            settings=settings,
            vlans=vlans,
            vrfs=vrfs,
            output_dir=output_dir,
            arch="2-8-5-200",
        )

        core_vars = yaml.safe_load((output_dir / "group_vars" / "core.yml").read_text())
        assert "interfaces_up" not in core_vars
        assert "interfaces_down" not in core_vars

    def test_l3_oob_uplink_mode_rewrites_default_vrf_bgp(self, tmp_path):
        """L3 OOB mode should emit direct uplink intent, not bond-based OOB defaults."""
        settings = {
            "architecture": "2-8-9-800",
            "oob_uplink_mode": "L3",
        }
        vlans = [
            {"id": 200, "name": "OOB", "subnet": "10.187.5.0/25", "vrf": "OOB", "vni": 289200},
            {"id": 300, "name": "CPU/In-Band", "subnet": "10.187.5.128/25", "vrf": "INBAND", "vni": 289300},
            {"id": 400, "name": "Support", "subnet": "10.187.4.0/27", "vrf": "INBAND", "vni": 289400},
        ]
        vrfs = {
            "OOB": {"name": "OOB", "l3_vni": 289001},
            "INBAND": {"name": "INBAND", "l3_vni": 289002},
            "EXIT": {"name": "EXIT", "l3_vni": 289004},
        }
        port_config = {
            "oob_uplink_interfaces": {
                "ports": [59],
                "breakout": 8,
                "lanes": 1,
                "port_overrides": {59: {"subports": [0, 1]}},
            },
            "isl_interfaces": {
                "ports": [56, 57, 58],
                "breakout": 2,
                "lanes": 4,
                "port_overrides": {58: {"subports": [0]}},
            },
        }
        nodes = [
            {"role": "oob-switch", "name": "mg-01", "status": "Active", "category": "oob-switch"},
            {"role": "oob-switch", "name": "mg-02", "status": "Active", "category": "oob-switch"},
        ]
        loopback_overrides = {
            "mg-01": {"lo": "10.187.4.35/32"},
            "mg-02": {"lo": "10.187.4.36/32"},
        }
        output_dir = tmp_path / "inventory"
        output_dir.mkdir()

        generate_group_vars(
            settings=settings,
            vlans=vlans,
            vrfs=vrfs,
            output_dir=output_dir,
            arch="2-8-9-800",
            nodes=nodes,
            port_config=port_config,
            loopback_overrides=loopback_overrides,
        )

        all_vars = yaml.safe_load((output_dir / "group_vars" / "all" / "main.yml").read_text())
        air_mgmt_iface = next(
            iface for iface in all_vars["ztp_interfaces"]
            if iface["purpose"] == "air-mgmt"
        )
        assert air_mgmt_iface["gateway"] == "172.20.0.254"

        core_vars = yaml.safe_load((output_dir / "group_vars" / "core.yml").read_text())
        assert "oob_uplink_interfaces" in core_vars
        neighbors = core_vars["default_vrf_bgp"]["neighbors"]
        peer_groups = {pg["id"]: pg for pg in core_vars["default_vrf_bgp"]["peer_groups"]}
        assert any(n["interfaces"] == "isl" and n["peer_group"] == "internal_isl" for n in neighbors)
        assert any(n["interfaces"] == "oob_uplink" and n["peer_group"] == "underlay" for n in neighbors)
        assert any(n["interfaces"] == ["10.187.4.35", "10.187.4.36"] and n["peer_group"] == "overlay"
                   for n in neighbors)
        assert peer_groups["internal_isl"]["remote_as"] == "internal"
        assert peer_groups["underlay"]["remote_as"] == "external"
        assert peer_groups["overlay"]["update_source"] == "lo"
        # W-ECMP only works over eBGP (confirmed via NVIDIA Cumulus Linux
        # docs: "W-ECMP is only supported in EBGP fabrics"). The 'underlay'
        # peer-group (eBGP to MG) qualifies and mirrors the identical,
        # already-correct policy on the OOB/MG switch's own 'underlay'
        # peer-group in oob_nvue_cli.j2. Found via a from-scratch brownfield
        # rebuild of a live 2-8-9-800 site diffed against its live config.
        assert (peer_groups["underlay"]["address_family"]["ipv4_unicast"]["policy_outbound_route_map"]
                == "WEIGHTED_ECMP")

    def test_default_prefix(self):
        """Missing prefix defaults to 24."""
        wb = openpyxl.Workbook()
        ws = _make_nodes_sheet(wb, rows=[
            ["core-01", "core-01", "", "192.168.200.2", None, "192.168.200.1"],
        ])

        result = parse_nodes(ws)
        assert result[0]["prefix"] == 24

    def test_enabled_column_respected(self):
        """Nodes with 'No' in Enabled column are marked Disabled."""
        wb = openpyxl.Workbook()
        headers = ["Function", "Name", "MAC Address for ZTP", "Mgmt IP Address",
                    "Prefix", "Gateway", "Enabled"]
        ws = _make_nodes_sheet(wb, headers=headers, rows=[
            ["core-01", "core-01", "", "192.168.200.2", 24, "192.168.200.1", "Yes"],
            ["core-02", "core-02", "", "192.168.200.3", 24, "192.168.200.1", "No"],
        ])

        result = parse_nodes(ws)
        assert result[0]["status"] == "Active"
        assert result[1]["status"] == "Disabled"

    def test_duplicate_node_names_both_returned(self):
        """Duplicate names are not deduplicated by parse_nodes itself."""
        wb = openpyxl.Workbook()
        ws = _make_nodes_sheet(wb, rows=[
            ["core-01", "core-01", "", "192.168.200.2", 24, "192.168.200.1"],
            ["core-01", "core-01", "", "192.168.200.3", 24, "192.168.200.1"],
        ])

        result = parse_nodes(ws)
        assert len(result) == 2

    def test_missing_column_headers_uses_defaults(self):
        """If headers don't match known names, fallback column indices are used."""
        wb = openpyxl.Workbook()
        # Use completely non-standard headers -- parser falls back to default columns
        ws = _make_nodes_sheet(wb, headers=[
            "Col_A", "Col_B", "Col_C", "Col_D", "Col_E", "Col_F",
        ], rows=[
            ["core-01", "core-01", "48:b0:2d:11:22:33", "192.168.200.2", 24, "192.168.200.1"],
        ])

        result = parse_nodes(ws)
        # Still parsed because default column indices 1-6 match the data layout
        assert len(result) == 1

    def test_parse_nodes_reads_oob_vlan_column(self):
        """Nodes tab reads OOB VLAN column and stores as oob_vlan (str, '' when absent)."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Function", "Name", "MAC Address", "Mgmt IP", "Prefix", "Gateway", "OOB VLAN"])
        ws.append(["OOB Switch", "oob-switch-01", "", "10.0.0.11", 24, "10.0.0.1", 201])
        ws.append(["GPU", "su-01-node-01", "", "192.168.201.21", 24, "", ""])
        nodes = parse_nodes(ws)
        by_name = {n["name"]: n for n in nodes}
        assert by_name["oob-switch-01"]["oob_vlan"] == "201"
        assert by_name["su-01-node-01"]["oob_vlan"] == ""


# ---------------------------------------------------------------------------
# generate_prefix_lists
# ---------------------------------------------------------------------------

class TestGeneratePrefixLists:
    """Tests for OOB_LOCAL_IF in l2 vs l3 oob_uplink_mode.

    Found via a from-scratch brownfield rebuild of a live 2-8-9-800 site:
    in l3 mode the core/CSL switch never owns an SVI
    on the OOB VLAN (the gateway lives on the OOB/MG switches), so the old
    formula computed a throwaway address instead of anything real.

    l2 mode has a real, locally-owned SVI on the OOB VLAN itself (OOB
    switches are dumb L2 bridges in that design) that needs its own
    anti-duplicate protection rule, since OOB_FILTER's other rules
    (INBAND_PREFIXES) don't cover the OOB VLAN's own subnet.

    l3 mode has no such address, so OOB_LOCAL_IF is just the OOB VRF
    loopback with nothing else added. We deliberately do NOT substitute
    this switch's INBAND SVI addresses as a stand-in: OOB_FILTER rule 10
    (INBAND_PREFIXES) already unconditionally denies the whole INBAND
    subnet range before OOB_LOCAL_IF (rule 15) is ever evaluated, so an
    INBAND-SVI /32 entry here would be dead, redundant config -- confirmed
    by era-documentation/guides/csl-routing-policy-analysis.md, which flags
    that exact pattern as a mistake in production's own OOB_LOCAL_IF rule 20.
    """

    _VLANS = [
        {"id": 200, "name": "OOB", "subnet": "10.187.5.0/25",
         "gateway": "10.187.5.1", "vrf": "OOB", "vni": 289200},
        {"id": 300, "name": "CPU/In-Band", "subnet": "10.187.5.128/25",
         "gateway": "10.187.5.129", "vrf": "INBAND", "vni": 289300},
        {"id": 400, "name": "Support", "subnet": "10.187.4.0/27",
         "gateway": "10.187.4.1", "vrf": "INBAND", "vni": 289400},
    ]

    def test_l2_mode_protects_own_oob_svi(self):
        """l2 mode: protect the OOB VLAN's own real, locally-owned SVI --
        nothing else in OOB_FILTER covers it."""
        prefix_lists = generate_prefix_lists(
            self._VLANS, core_num=1, loopback_base="10.187.4",
            oob_uplink_mode="l2",
        )
        oob_local_if = next(pl for pl in prefix_lists if pl["id"] == "OOB_LOCAL_IF")
        # Fields under test only. Exact dict equality breaks whenever a
        # descriptive field is added — ADR-0043 put `description` on every rule
        # — and says nothing about the l2/l3 SVI behaviour these assert.
        assert [{k: v for k, v in r.items() if k != "description"}
                for r in oob_local_if["rule"]] == [
            {"id": "10", "match": "10.187.4.1/32", "max_len": "32"},
            {"id": "20", "match": "10.187.5.2/32", "max_len": "32"},  # own OOB VLAN SVI
        ]

    def test_l3_mode_has_no_svi_to_protect(self):
        """l3 mode: no real OOB SVI on this switch, and INBAND SVIs are
        already covered by an earlier OOB_FILTER rule -- nothing to add."""
        prefix_lists = generate_prefix_lists(
            self._VLANS, core_num=1, loopback_base="10.187.4",
            oob_uplink_mode="l3",
        )
        oob_local_if = next(pl for pl in prefix_lists if pl["id"] == "OOB_LOCAL_IF")
        # Fields under test only. Exact dict equality breaks whenever a
        # descriptive field is added — ADR-0043 put `description` on every rule
        # — and says nothing about the l2/l3 SVI behaviour these assert.
        assert [{k: v for k, v in r.items() if k != "description"}
                for r in oob_local_if["rule"]] == [
            {"id": "10", "match": "10.187.4.1/32", "max_len": "32"},
        ]

    def test_l2_mode_per_switch_ips_differ(self):
        """core_num varies the OOB VLAN SVI host octet per switch, same as
        every other per-switch computed value in this function."""
        prefix_lists = generate_prefix_lists(
            self._VLANS, core_num=2, loopback_base="10.187.4",
            oob_uplink_mode="l2",
        )
        oob_local_if = next(pl for pl in prefix_lists if pl["id"] == "OOB_LOCAL_IF")
        assert oob_local_if["rule"][1]["match"] == "10.187.5.3/32"

    def test_default_oob_uplink_mode_is_l2(self):
        """Omitting oob_uplink_mode falls back to l2 behavior."""
        prefix_lists = generate_prefix_lists(
            self._VLANS, core_num=1, loopback_base="10.187.4",
        )
        oob_local_if = next(pl for pl in prefix_lists if pl["id"] == "OOB_LOCAL_IF")
        assert oob_local_if["rule"][-1]["match"] == "10.187.5.2/32"


def test_generate_prefix_lists_era_prefixes_includes_oob_subnets():
    from excel_parser import generate_prefix_lists
    vlans = [{"id": 300, "name": "in-band", "subnet": "172.16.178.0/24", "vrf": "INBAND"}]
    pls = generate_prefix_lists(vlans, core_num=1, loopback_base="172.16.176",
                                oob_subnets=["192.168.200.0/24", "192.168.201.0/24"])
    era = next(p for p in pls if p["id"] == "ERA_PREFIXES")
    matches = {r["match"] for r in era["rule"]}
    assert "192.168.200.0/24" in matches
    assert "192.168.201.0/24" in matches


# ---------------------------------------------------------------------------
# parse_vlans
# ---------------------------------------------------------------------------

class TestParseVlans:
    """Tests for parse_vlans() edge cases."""

    def test_empty_vlans_sheet(self):
        """VLANs sheet with no data rows returns empty list."""
        wb = openpyxl.Workbook()
        ws = _make_vlans_sheet(wb)
        result = parse_vlans(ws)
        assert result == []

    def test_valid_vlans(self):
        """Standard VLANs are parsed with auto-generated VNI."""
        wb = openpyxl.Workbook()
        ws = _make_vlans_sheet(wb, vlans=[
            [100, "CPU/In-Band", "Compute traffic", "172.16.178.0/24", "default"],
            [200, "OOB", "Management", "192.168.200.0/24", "mgmt"],
        ])

        result = parse_vlans(ws)
        assert len(result) == 2
        assert result[0]["id"] == 100
        assert result[0]["name"] == "CPU/In-Band"
        assert result[0]["subnet"] == "172.16.178.0/24"
        assert result[0]["vni"] == 4100  # 100 + 4000

    def test_vlan_breaks_on_non_integer_id(self):
        """Parsing stops when a non-integer VLAN ID is encountered."""
        wb = openpyxl.Workbook()
        ws = _make_vlans_sheet(wb, vlans=[
            [100, "First", "test", "10.0.0.0/24", "default"],
            ["not-a-number", "Second", "test2", "10.0.1.0/24", "default"],
            [300, "Third", "test3", "10.0.2.0/24", "default"],
        ])

        result = parse_vlans(ws)
        assert len(result) == 1  # Stops at the non-integer row

    def test_vrf_defaults_to_default(self):
        """Missing VRF defaults to 'default'."""
        wb = openpyxl.Workbook()
        ws = _make_vlans_sheet(wb, vlans=[
            [100, "Test", "test", "10.0.0.0/24", None],
        ])

        result = parse_vlans(ws)
        assert result[0]["vrf"] == "default"

    def test_vni_column_overrides_auto(self):
        """Explicit VNI column values override the auto-calculated VNI."""
        wb = openpyxl.Workbook()
        ws = wb.create_sheet("VLANs & Profiles")
        ws.append(["VLANs"])
        ws.append(["VLAN_ID", "Name", "Purpose", "Subnet", "VRF", "VNI"])
        ws.append([100, "Test", "test", "10.0.0.0/24", "default", 99999])

        result = parse_vlans(ws)
        assert result[0]["vni"] == 99999


# ---------------------------------------------------------------------------
# parse_versions
# ---------------------------------------------------------------------------

class TestParseVersions:
    """Tests for parse_versions() which reads the VERSIONS table from Settings."""

    def test_empty_settings(self):
        """Settings sheet without VERSIONS table returns empty dict."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Settings"
        ws.append(["Site Name", "test"])

        result = parse_versions(ws)
        assert result == {}

    def test_valid_versions(self):
        """VERSIONS table is correctly parsed."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Settings"
        ws.append(["Site Name", "test"])
        ws.append(["Switch Function", "Cumulus Version"])
        ws.append(["core", "5.16.1"])
        ws.append(["oob", "5.15.1"])

        result = parse_versions(ws)
        assert result == {"core": "5.16.1", "oob": "5.15.1"}

    def test_versions_stops_at_empty_key_string(self):
        """Parsing stops at a row with empty (non-None) key string."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Settings"
        ws.append(["Switch Function", "Cumulus Version"])
        ws.append(["core", "5.16.1"])
        ws.append(["", None])  # empty string key → stops parsing
        ws.append(["edge", "5.16.0"])  # should NOT be included

        result = parse_versions(ws)
        assert result == {"core": "5.16.1"}

    def test_versions_skips_none_key_rows(self):
        """None-key rows inside the version table are skipped (continue, not break)."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Settings"
        ws.append(["Switch Function", "Cumulus Version"])
        ws.append(["core", "5.16.1"])
        ws.append([None, None])  # None key → continue (skip)
        ws.append(["edge", "5.16.0"])  # still parsed

        result = parse_versions(ws)
        assert result == {"core": "5.16.1", "edge": "5.16.0"}


# ---------------------------------------------------------------------------
# _build_wiremap_row_list
# ---------------------------------------------------------------------------

class TestBuildWiremapRowList:
    """Tests for _build_wiremap_row_list() which reads Wire Map into dicts."""

    def test_empty_wiremap(self):
        """Wire Map with only header returns empty list."""
        wb = openpyxl.Workbook()
        ws = _make_wiremap_sheet(wb)
        result = _build_wiremap_row_list(ws)
        assert result == []

    def test_valid_wiremap_rows(self):
        """Data rows are parsed into dicts with correct keys."""
        wb = openpyxl.Workbook()
        ws = _make_wiremap_sheet(wb, rows=[
            # display, sys_role, sys_name, nic, speed, desc, profile,
            # disabled, speed2, desc2, sw_role, sw_name, sw_port
            ["Yes", "su-01-node-01", "su-01-node-01", "NIC1_P1", None, None,
             "CPU/In-Band Network", None, None, None,
             "core-01", "core-01", "swp1s0"],
        ])

        result = _build_wiremap_row_list(ws)
        assert len(result) == 1
        assert result[0]["system_role"] == "su-01-node-01"
        assert result[0]["system_name"] == "su-01-node-01"
        assert result[0]["switch_role"] == "core-01"
        assert result[0]["switch_port"] == "swp1s0"
        assert result[0]["display_in_air"] is True
        assert result[0]["net_profile"] == "CPU/In-Band Network"

    def test_skip_rows_without_system_role(self):
        """Rows with empty System Role are skipped."""
        wb = openpyxl.Workbook()
        ws = _make_wiremap_sheet(wb, rows=[
            ["Yes", "", "", "NIC1", None, None, "CPU", None, None, None,
             "core-01", "core-01", "swp1s0"],
            ["Yes", "su-01-node-01", "su-01-node-01", "NIC1", None, None, "CPU",
             None, None, None, "core-01", "core-01", "swp2s0"],
        ])

        result = _build_wiremap_row_list(ws)
        assert len(result) == 1

    def test_display_in_air_flag(self):
        """Display in Air 'No' rows are parsed but with display_in_air=False."""
        wb = openpyxl.Workbook()
        ws = _make_wiremap_sheet(wb, rows=[
            ["No", "su-01-node-01", "su-01-node-01", "NIC1", None, None, "CPU",
             None, None, None, "core-01", "core-01", "swp1s0"],
        ])

        result = _build_wiremap_row_list(ws)
        assert len(result) == 1
        assert result[0]["display_in_air"] is False

    def test_system_name_falls_back_to_role(self):
        """Empty System Name falls back to System Role."""
        wb = openpyxl.Workbook()
        ws = _make_wiremap_sheet(wb, rows=[
            ["Yes", "core-01", "", "swp49s0", None, None, "ISL",
             None, None, None, "core-02", "core-02", "swp49s0"],
        ])

        result = _build_wiremap_row_list(ws)
        assert result[0]["system_name"] == "core-01"


def test_per_switch_gpu_rails_include_rows_hidden_from_air():
    """Air display filtering must not remove physical GPU switch config."""
    rows = [
        {
            "switch_name": "core-01",
            "net_profile": "GPU Rail 1",
            "switch_port": "swp9s0",
            "display_in_air": True,
        },
        {
            "switch_name": "core-01",
            "net_profile": "GPU Rail 1",
            "switch_port": "swp9s1",
            "display_in_air": True,
        },
        {
            "switch_name": "core-01",
            "net_profile": "GPU Rail 1",
            "switch_port": "swp11s0",
            "display_in_air": False,
        },
        {
            "switch_name": "core-01",
            "net_profile": "GPU Rail 1",
            "switch_port": "swp11s1",
            "display_in_air": False,
        },
        {
            "switch_name": "core-02",
            "net_profile": "GPU Rail 1",
            "switch_port": "swp12s0",
            "display_in_air": False,
        },
    ]
    vlans = [{"id": 901, "name": "gpu_rail1"}]

    result = build_per_switch_gpu_rail_interfaces(rows, "core-01", vlans)

    assert result["rail1"]["ports"] == [9, 11]
    assert result["rail1"]["vlan"] == 901
    assert result["rail1"]["port_overrides"] == {}


@pytest.mark.parametrize(
    ("arch", "expected_parents"),
    [
        ("2-4-3-200", list(range(9, 25))),
        ("2-8-5-200", list(range(6, 26))),
        ("2-8-9-400", list(range(4, 28))),
    ],
)
def test_converged_inventory_keeps_disabled_scale_out_gpu_ports(
        tmp_path, arch, expected_parents):
    """Active-only Air filtering must not truncate physical switch rails."""
    repo = Path(__file__).parent.parent
    workbook = repo / "input" / arch / "default" / f"{arch}.xlsx"
    output = tmp_path / arch

    process_excel_template(workbook, output)

    for switch in ("core-01", "core-02"):
        host_vars = yaml.safe_load(
            (output / "host_vars" / f"{switch}.yml").read_text()
        )
        configured = sorted({
            parent
            for rail in host_vars["gpu_rail_interfaces"].values()
            for parent in rail["ports"]
        })
        assert configured == expected_parents


# ---------------------------------------------------------------------------
# parse_oob_switch_configs
# ---------------------------------------------------------------------------

class TestParseOobSwitchConfigs:
    """Tests for parse_oob_switch_configs() which derives OOB switch port configs."""

    def test_empty_wiremap(self):
        """Empty Wire Map returns empty dict."""
        wb = openpyxl.Workbook()
        ws = _make_wiremap_sheet(wb)
        result = parse_oob_switch_configs(ws)
        assert result == {}

    def test_access_and_uplink_classification(self):
        """Non-core connections are access, core non-eth0 connections are uplinks."""
        wb = openpyxl.Workbook()
        ws = _make_wiremap_sheet(wb, rows=[
            # su-01-node-01 connected to oob-switch-01 swp1 = access
            ["Yes", "su-01-node-01", "su-01-node-01", "eth0", None, None,
             "OOB / IPMI", None, None, None,
             "oob-switch-01", "oob-switch-01", "swp1"],
            # core-01 swp49 connected to oob-switch-01 swp49 = uplink
            ["Yes", "core-01", "core-01", "swp49s0", None, None,
             "OOB Uplink", None, None, None,
             "oob-switch-01", "oob-switch-01", "swp49"],
        ])

        result = parse_oob_switch_configs(ws)
        assert "oob-switch-01" in result
        cfg = result["oob-switch-01"]
        assert "swp1" in _expand_port_range(cfg["access_ports"])
        assert "swp49" in _expand_port_range(cfg["uplink_ports"])
        assert "swp49" in cfg["spine_bond_members"]

    def test_core_eth0_is_access(self):
        """Core eth0 connection to OOB switch is classified as access, not uplink."""
        wb = openpyxl.Workbook()
        ws = _make_wiremap_sheet(wb, rows=[
            ["Yes", "core-01", "core-01", "eth0", None, None,
             "OOB / IPMI", None, None, None,
             "oob-switch-01", "oob-switch-01", "swp2"],
        ])

        result = parse_oob_switch_configs(ws)
        cfg = result["oob-switch-01"]
        # access_ports is a range string (e.g. "swp1-2") that includes both
        # the explicitly-wired port (swp2) and the auto-allocated air-oob
        # backdoor (first unused port, here swp1). Expand and check membership.
        assert "swp2" in _expand_port_range(cfg["access_ports"])
        assert cfg["spine_bond_members"] == []

    def test_dhcp_oob_port_tracked(self):
        """Port connected to dhcp-oob is tracked in dhcp_oob_port."""
        wb = openpyxl.Workbook()
        ws = _make_wiremap_sheet(wb, rows=[
            ["Yes", "dhcp-oob", "dhcp-oob", "eth1", None, None,
             "Air - Management", None, None, None,
             "oob-switch-01", "oob-switch-01", "swp44"],
        ])

        result = parse_oob_switch_configs(ws)
        assert result["oob-switch-01"]["dhcp_oob_port"] == "swp44"


# ---------------------------------------------------------------------------
# parse_core_port_config
# ---------------------------------------------------------------------------

class TestParseCorePortConfig:
    """Tests for core/CSL port role derivation from Wire Map."""

    def test_l2_oob_uplink_bond_uses_oob_vlan_when_profile_is_l3(self):
        """L2 OOB mode turns OOB Uplink rows into VLAN 200 bonds.

        Generated workbooks keep the OOB Uplink Port Profile in L3 shape
        because the same workbook can be switched to L3 mode. In L2 mode,
        the parser must still emit a concrete OOB access VLAN instead of
        carrying the profile's blank native VLAN through to the template.
        """
        wb = openpyxl.Workbook()
        ws_wiremap = _make_wiremap_sheet(wb, rows=[
            ["Yes", "oob-switch-01", "oob-switch-01", "swp49", None, None,
             "OOB Uplink", None, None, None,
             "core-01", "core-01", "swp59s0"],
            ["Yes", "oob-switch-02", "oob-switch-02", "swp49", None, None,
             "OOB Uplink", None, None, None,
             "core-01", "core-01", "swp59s1"],
        ])
        ws_profiles = _make_vlans_profiles_sheet(
            wb,
            vlans=[
                [200, "OOB", "Out-of-Band", "172.16.2.0/24", "OOB"],
            ],
            profiles=[
                ["OOB Uplink", "L3", None, None, None, "default",
                 "No", "100G", 8, 1],
            ],
        )

        result = parse_core_port_config(
            ws_wiremap,
            ws_profiles,
            vlans=[{"id": 200, "name": "OOB"}],
            oob_uplink_mode="l2",
        )

        oob = result["network_roles"]["oob"]
        assert oob["vlan"] == 200
        assert oob["ports"] == [59]
        assert oob["port_overrides"][59]["subports"] == [0, 1]

    def test_l3_oob_uplink_stays_direct_interface(self):
        """L3 OOB mode must not create L2 OOB bonds."""
        wb = openpyxl.Workbook()
        ws_wiremap = _make_wiremap_sheet(wb, rows=[
            ["Yes", "oob-switch-01", "oob-switch-01", "swp49", None, None,
             "OOB Uplink", None, None, None,
             "core-01", "core-01", "swp59s0"],
        ])
        ws_profiles = _make_vlans_profiles_sheet(
            wb,
            vlans=[
                [200, "OOB", "Out-of-Band", "172.16.2.0/24", "OOB"],
            ],
            profiles=[
                ["OOB Uplink", "L3", None, None, None, "default",
                 "No", "100G", 8, 1],
            ],
        )

        result = parse_core_port_config(
            ws_wiremap,
            ws_profiles,
            vlans=[{"id": 200, "name": "OOB"}],
            oob_uplink_mode="l3",
        )

        assert "oob" not in result["network_roles"]
        assert result["oob_uplink_interfaces"]["ports"] == [59]

    def test_edge_uplinks_forced_into_exit_vrf_when_profile_default(self):
        """Edge uplinks must land in the EXIT VRF even if the Edge
        Uplink profile VRF column says 'default'.

        The EXIT-VRF BGP peers the edge interfaces; for unnumbered eBGP the
        interface must be in the same VRF as the BGP instance, so a profile
        VRF of 'default' (a stale generator default) must be coerced to EXIT
        or every EXIT session sits idle.
        """
        wb = openpyxl.Workbook()
        ws_wiremap = _make_wiremap_sheet(wb, rows=[
            ["Yes", "cust-net-edge-01", "cust-net-edge-01", "swp4", None, None,
             "Edge Uplink", None, None, None,
             "core-01", "core-01", "swp61s0"],
        ])
        ws_profiles = _make_vlans_profiles_sheet(
            wb,
            vlans=[[200, "OOB", "Out-of-Band", "172.16.2.0/24", "OOB"]],
            profiles=[
                ["Edge Uplink", "L3", None, None, None, "default",
                 "No", "400G", 2, 4],
            ],
        )

        result = parse_core_port_config(
            ws_wiremap, ws_profiles,
            vlans=[{"id": 200, "name": "OOB"}],
            oob_uplink_mode="l3",
        )

        assert "edge_interfaces" in result
        assert result["edge_interfaces"]["vrf"] == "EXIT"

    def test_edge_uplinks_respect_explicit_nondefault_vrf(self):
        """An explicit non-default profile VRF is preserved (not overridden)."""
        wb = openpyxl.Workbook()
        ws_wiremap = _make_wiremap_sheet(wb, rows=[
            ["Yes", "cust-net-edge-01", "cust-net-edge-01", "swp4", None, None,
             "Edge Uplink", None, None, None,
             "core-01", "core-01", "swp61s0"],
        ])
        ws_profiles = _make_vlans_profiles_sheet(
            wb,
            vlans=[[200, "OOB", "Out-of-Band", "172.16.2.0/24", "OOB"]],
            profiles=[
                ["Edge Uplink", "L3", None, None, None, "EXIT",
                 "No", "400G", 2, 4],
            ],
        )

        result = parse_core_port_config(
            ws_wiremap, ws_profiles,
            vlans=[{"id": 200, "name": "OOB"}],
            oob_uplink_mode="l3",
        )

        assert result["edge_interfaces"]["vrf"] == "EXIT"


# ---------------------------------------------------------------------------
# build_devices
# ---------------------------------------------------------------------------

class TestBuildDevices:
    """Tests for build_devices() which generates device dicts for dnsmasq."""

    def test_empty_nodes(self):
        """Empty nodes list returns empty devices dict."""
        result = build_devices([], [])
        assert result == {}

    def test_switches_excluded(self):
        """Switch nodes are not included in devices."""
        nodes = [
            {"name": "core-01", "role": "core-01", "mgmt_ip": "192.168.200.2", "enabled": True},
        ]
        result = build_devices(nodes, [])
        assert "core-01" not in result

    def test_infra_excluded(self):
        """Infrastructure nodes (dhcp-oob, oob-server-01) are not included."""
        nodes = [
            {"name": "dhcp-oob", "role": "dhcp-oob", "mgmt_ip": "192.168.200.252", "enabled": True},
        ]
        result = build_devices(nodes, [])
        assert "dhcp-oob" not in result

    def test_compute_node_included(self):
        """Compute nodes get eth0_ip and auto-generated mac."""
        nodes = [
            {"name": "su-01-node-01", "role": "su-01-node-01",
             "mgmt_ip": "192.168.200.11", "enabled": True},
        ]
        result = build_devices(nodes, [])
        assert "su-01-node-01" in result
        assert result["su-01-node-01"]["eth0_ip"] == "192.168.200.11"
        assert result["su-01-node-01"]["mac"].startswith("48:b0:2d:")

    def test_explicit_mac_preserved(self):
        """MAC from Excel is used when present."""
        nodes = [
            {"name": "su-01-node-01", "role": "su-01-node-01",
             "mac": "48:b0:2d:aa:bb:cc",
             "mgmt_ip": "192.168.200.11", "enabled": True},
        ]
        result = build_devices(nodes, [])
        assert result["su-01-node-01"]["mac"] == "48:b0:2d:aa:bb:cc"

    def test_disabled_nodes_excluded(self):
        """Nodes with enabled=False are excluded."""
        nodes = [
            {"name": "su-01-node-01", "role": "su-01-node-01",
             "mgmt_ip": "192.168.200.11", "enabled": False},
        ]
        result = build_devices(nodes, [])
        assert "su-01-node-01" not in result

    def test_compute_data_plane_ips(self):
        """Compute nodes get bond_ip and gpu_ips when subnets available."""
        nodes = [
            {"name": "su-01-node-01", "role": "su-01-node-01",
             "mgmt_ip": "192.168.200.11", "enabled": True},
        ]
        vlans = [
            {"name": "CPU/In-Band", "subnet": "172.16.178.0/24"},
            {"name": "GPU Network", "subnet": "172.16.179.0/24"},
        ]
        result = build_devices(nodes, vlans)
        dev = result["su-01-node-01"]
        assert "bond_ip" in dev
        assert dev["bond_ip"].startswith("172.16.178.")

    def test_inband_vlan_name_maps_to_compute_cpu_subnet(self):
        """Generated workbooks name the CPU subnet INBAND."""
        nodes = [
            {"name": "su-01-node-01", "role": "su-01-node-01",
             "mgmt_ip": "192.168.200.11", "enabled": True},
        ]
        vlans = [
            {"name": "INBAND", "subnet": "172.16.3.0/24"},
            {"name": "GPU Network", "subnet": "192.168.0.0/20"},
        ]
        result = build_devices(nodes, vlans)
        assert result["su-01-node-01"]["bond_ip"].startswith("172.16.3.")

    def test_support_nodes_fall_back_to_inband_subnet(self):
        """Generated support rows use CPU/In-Band Network when no SUPPORT VLAN exists."""
        nodes = [
            {"name": "support-01", "role": "support-01",
             "mgmt_ip": "192.168.200.61", "enabled": True},
        ]
        vlans = [
            {"name": "INBAND", "subnet": "172.16.3.0/24"},
        ]
        result = build_devices(nodes, vlans)
        dev = result["support-01"]
        assert dev["bond_ip1"].startswith("172.16.3.")
        assert dev["bond_ip2"].startswith("172.16.3.")

    def test_storage_data_plane_ips(self):
        """Storage nodes get bond_ip1 and bond_ip2 when subnet available."""
        nodes = [
            {"name": "storage-01", "role": "storage-01",
             "mgmt_ip": "192.168.200.61", "enabled": True},
        ]
        vlans = [
            {"name": "Storage", "subnet": "172.16.180.0/24"},
        ]
        result = build_devices(nodes, vlans)
        dev = result["storage-01"]
        assert "bond_ip1" in dev
        assert dev["bond_ip1"].startswith("172.16.180.")
        assert "bond_ip2" in dev


# ---------------------------------------------------------------------------
# classify_host_role (additional edge cases)
# ---------------------------------------------------------------------------

class TestClassifyHostRoleEdgeCases:
    """Additional edge cases beyond what test_parser_functions.py covers."""

    def test_air_oob_switch_is_air_oob(self):
        """air-oob-switch is classified as 'air-oob' (not mapped to 'switch')."""
        # classify_host_role only maps core/oob/edge → switch; air-oob passes through
        assert classify_host_role("air-oob-switch") == "air-oob"

    def test_unknown_name(self):
        """Unknown names return 'unknown'."""
        assert classify_host_role("totally-weird-name") == "unknown"

    def test_k8s_not_switch(self):
        """k8s nodes are not switches."""
        assert classify_host_role("k8s-01") == "k8s"


# ---------------------------------------------------------------------------
# parse_dhcp_relay_table + parse_vlans DHCP Relay Client column
# ---------------------------------------------------------------------------

from excel_parser import parse_dhcp_relay_table


def _make_vlans_profiles_with_dhcp(wb, vlans=None, dhcp_relay_rows=None,
                                   add_relay_client_col=True):
    """Build a minimal VLANs & Profiles sheet with optional DHCP Relay table
    and per-VLAN DHCP Relay Client column.

    vlans: list of (id, name, purpose, subnet, gateway, vrf, vni, relay_client).
    dhcp_relay_rows: list of (server_ip_raw, vrf, upstream_interface).
    """
    ws = wb.create_sheet("VLANs & Profiles")
    ws.append(["VLANs"])
    header = ["VLAN ID", "Name", "Purpose", "Subnet", "Gateway", "VRF", "VNI"]
    if add_relay_client_col:
        header.append("DHCP Relay Client")
    ws.append(header)
    for v in (vlans or []):
        ws.append(list(v))
    # Empty row gap
    ws.append([])
    if dhcp_relay_rows is not None:
        ws.append(["DHCP Relay"])
        ws.append(["Server IP", "VRF", "Upstream Interface"])
        for row in dhcp_relay_rows:
            ws.append(list(row))
    return ws


class TestParseDhcpRelayTable:
    """parse_dhcp_relay_table() reads the DHCP Relay table from VLANs & Profiles."""

    def test_no_section(self):
        wb = openpyxl.Workbook()
        ws = _make_vlans_profiles_with_dhcp(wb, dhcp_relay_rows=None)
        assert parse_dhcp_relay_table(ws) == []

    def test_empty_table(self):
        wb = openpyxl.Workbook()
        ws = _make_vlans_profiles_with_dhcp(wb, dhcp_relay_rows=[])
        assert parse_dhcp_relay_table(ws) == []

    def test_single_row(self):
        wb = openpyxl.Workbook()
        ws = _make_vlans_profiles_with_dhcp(
            wb, dhcp_relay_rows=[("192.168.200.252", "OOB", "vlan200")])
        result = parse_dhcp_relay_table(ws)
        assert result == [{
            'servers': ['192.168.200.252'],
            'vrf': 'OOB',
            'upstream_interfaces': ['vlan200'],
        }]

    def test_comma_separated_servers(self):
        wb = openpyxl.Workbook()
        ws = _make_vlans_profiles_with_dhcp(
            wb,
            dhcp_relay_rows=[("192.168.200.252,192.168.200.253", "OOB", "vlan200")],
        )
        result = parse_dhcp_relay_table(ws)
        assert result[0]['servers'] == ['192.168.200.252', '192.168.200.253']

    def test_comma_separated_upstream_interfaces(self):
        """NVUE supports multiple upstream-interface entries per server-group
        (deck slide 16); operator declares them comma-separated."""
        wb = openpyxl.Workbook()
        ws = _make_vlans_profiles_with_dhcp(
            wb,
            dhcp_relay_rows=[("192.168.200.252", "OOB", "swp64s0,swp64s1")],
        )
        result = parse_dhcp_relay_table(ws)
        assert result[0]['upstream_interfaces'] == ['swp64s0', 'swp64s1']

    def test_multiple_rows(self):
        wb = openpyxl.Workbook()
        ws = _make_vlans_profiles_with_dhcp(
            wb,
            dhcp_relay_rows=[
                ("192.168.200.252", "OOB", "vlan200"),
                ("10.0.0.100", "EXIT", "swp61s0"),
            ],
        )
        result = parse_dhcp_relay_table(ws)
        assert len(result) == 2
        assert result[0]['vrf'] == 'OOB'
        assert result[1]['vrf'] == 'EXIT'
        assert result[1]['upstream_interfaces'] == ['swp61s0']

    def test_vrf_normalized_uppercase(self):
        wb = openpyxl.Workbook()
        ws = _make_vlans_profiles_with_dhcp(
            wb, dhcp_relay_rows=[("192.168.200.252", "oob", "vlan200")])
        result = parse_dhcp_relay_table(ws)
        assert result[0]['vrf'] == 'OOB'


class TestParseVlansDhcpRelayClient:
    """parse_vlans() reads the new DHCP Relay Client column."""

    def test_column_present_with_no_values(self):
        wb = openpyxl.Workbook()
        ws = _make_vlans_profiles_with_dhcp(
            wb,
            vlans=[(300, 'inband', 'CPU', '192.168.45.0/24',
                    '192.168.45.1', 'INBAND', 4300, 'No')],
        )
        vlans = parse_vlans(ws)
        assert vlans[0]['dhcp_relay_client'] == 'No'

    def test_column_present_with_oob_value(self):
        wb = openpyxl.Workbook()
        ws = _make_vlans_profiles_with_dhcp(
            wb,
            vlans=[(300, 'inband', 'CPU', '192.168.45.0/24',
                    '192.168.45.1', 'INBAND', 4300, 'OOB')],
        )
        vlans = parse_vlans(ws)
        assert vlans[0]['dhcp_relay_client'] == 'OOB'

    def test_column_absent(self):
        """Backward-compat: VLANs sheet without the column → blank value."""
        wb = openpyxl.Workbook()
        ws = _make_vlans_profiles_with_dhcp(
            wb,
            vlans=[(300, 'inband', 'CPU', '192.168.45.0/24',
                    '192.168.45.1', 'INBAND', 4300)],
            add_relay_client_col=False,
        )
        vlans = parse_vlans(ws)
        assert vlans[0]['dhcp_relay_client'] == ''


# ---------------------------------------------------------------------------
# Per-rail-per-plane GPU VLAN mode
# ---------------------------------------------------------------------------

class TestPerRailPerPlaneParser:
    """build_devices() recognizes gpu_rail<R>_plane<P> VLAN rows and
    "GPU Rail R Plane P" Wire Map profiles when gpu_vlan_mode = per_rail_per_plane."""

    def _make_test_vlans(self):
        """Standard 2 rails × 2 planes VLAN set + CPU/OOB."""
        return [
            {'id': 200, 'name': 'oob',  'subnet': '192.168.200.0/24', 'vrf': 'OOB'},
            {'id': 300, 'name': 'CPU',  'subnet': '172.16.178.0/24',  'vrf': 'INBAND'},
            # Reused VLAN IDs across planes (operator's choice)
            {'id': 901, 'name': 'gpu_rail1_plane1', 'subnet': '192.168.0.0/24',  'gateway': '192.168.0.1',  'vrf': 'GPU'},
            {'id': 902, 'name': 'gpu_rail2_plane1', 'subnet': '192.168.1.0/24',  'gateway': '192.168.1.1',  'vrf': 'GPU'},
            {'id': 901, 'name': 'gpu_rail1_plane2', 'subnet': '192.168.16.0/24', 'gateway': '192.168.16.1', 'vrf': 'GPU'},
            {'id': 902, 'name': 'gpu_rail2_plane2', 'subnet': '192.168.17.0/24', 'gateway': '192.168.17.1', 'vrf': 'GPU'},
        ]

    def _make_test_wiremap(self, host='gpu-01'):
        """Wire map with 4 GPU NICs — 2 in plane 1, 2 in plane 2."""
        return [
            {'sys_role': host, 'sys_name': host, 'nic_port': 'eth3',
             'network_profile': 'GPU Rail 1 Plane 1',
             'sw_role': 'gsl-plane1', 'sw_name': 'gsl-plane1-01', 'sw_port': 'swp1s0'},
            {'sys_role': host, 'sys_name': host, 'nic_port': 'eth4',
             'network_profile': 'GPU Rail 2 Plane 1',
             'sw_role': 'gsl-plane1', 'sw_name': 'gsl-plane1-01', 'sw_port': 'swp2s0'},
            {'sys_role': host, 'sys_name': host, 'nic_port': 'eth5',
             'network_profile': 'GPU Rail 1 Plane 2',
             'sw_role': 'gsl-plane2', 'sw_name': 'gsl-plane2-01', 'sw_port': 'swp1s0'},
            {'sys_role': host, 'sys_name': host, 'nic_port': 'eth6',
             'network_profile': 'GPU Rail 2 Plane 2',
             'sw_role': 'gsl-plane2', 'sw_name': 'gsl-plane2-01', 'sw_port': 'swp2s0'},
        ]

    # NOTE: full end-to-end tests for IP allocation per (rail, plane) require a
    # built wiremap_rows list shaped by _build_wiremap_row_list (with system_role,
    # nic_port, network_profile, dst_switch, etc. populated correctly). Hand-rolled
    # dicts don't suffice. End-to-end verification is covered by smoke-running a
    # synthetic 2-8-9-800-derived Excel through `make generate` — see
    # docs/plans/2026-05-18-gpu-plane-per-rail.md.

    def test_mode_single_ignores_per_rail_plane_rows(self):
        """When mode is single, gpu_rail*_plane* rows shouldn't get used for IP allocation."""
        nodes = [{'name': 'gpu-01', 'role': 'gpu', 'mac': '', 'mgmt_ip': '192.168.200.10', 'enabled': True}]
        vlans = self._make_test_vlans()
        wiremap = self._make_test_wiremap()
        result = build_devices(nodes, vlans, wiremap_rows=wiremap,
                               gpu_vlan_mode='single')
        # No gpu_interfaces emitted since the per-rail-per-plane path is gated off
        gpus = result.get('gpu-01', {}).get('gpu_interfaces', [])
        assert gpus == []


def _rolenode(name, cat, status="Enabled"):
    return {"name": name, "category": cat, "status": status}


def test_split_roles_bucket_into_new_groups():
    nodes = [_rolenode("cl-01", "cl"), _rolenode("cs-01", "cs"),
             _rolenode("gl-plane1-01", "gl-plane1"), _rolenode("gs-plane1-01", "gs-plane1")]
    c = categorize_nodes(nodes, {})
    assert any(n["name"] == "cl-01" for n in c["cl"])
    assert any(n["name"] == "cs-01" for n in c["cs"])
    assert any(n["name"] == "gl-plane1-01" for n in c["gl_plane1"])
    assert any(n["name"] == "gs-plane1-01" for n in c["gs_plane1"])


def test_build_devices_signature_has_no_mgmt_subnets():
    import inspect
    from scripts.excel_parser import build_devices
    assert "mgmt_subnets" not in inspect.signature(build_devices).parameters


def test_ztp_interfaces_from_oob_vlans(tmp_path):
    # Build a minimal parsed context and assert ztp_interfaces count == distinct OOB subnets.
    from scripts.excel_parser import resolve_oob_vlans
    vlans = [
        {"id": 200, "name": "oob-1", "subnet": "192.168.200.0/24", "gateway": "192.168.200.1", "vrf": "OOB"},
        {"id": 201, "name": "oob-2", "subnet": "192.168.201.0/24", "gateway": "192.168.201.1", "vrf": "OOB"},
    ]
    oob = [{"name": "oob-switch-01", "oob_vlan": "200"}, {"name": "oob-switch-02", "oob_vlan": "201"}]
    subnets = resolve_oob_vlans(vlans, oob)["subnets"]
    assert len(subnets) == 2  # one ztp interface per distinct OOB subnet


def test_no_mgmt_subnets_left_in_excel_parser():
    from pathlib import Path
    src = Path("scripts/excel_parser.py").read_text()
    assert "mgmt_subnets" not in src, "mgmt_subnets must be fully removed"


def test_air_scripts_have_no_mgmt_subnets():
    from pathlib import Path
    for p in ["scripts/air-deploy.py", "scripts/generate-node-instructions.py"]:
        assert "mgmt_subnets" not in Path(p).read_text(), p
