# SPDX-FileCopyrightText: Copyright (c) 2026 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: MIT
"""
The ISL must have the number of links AND the per-link bandwidth the arch
model specifies.

Both are published. The ERA deployment guides carry an "ISL Ports (both ends)"
column per scale point, and an ISL port-range row giving geometry:

    ISL (Interlink)  swp28s0  swp51s1  Breakout port to 2x 400G ports with 4 lanes
        -- ERA-00010-001 v03, 2-8-9-400 port table

Both are transcribed into the models:

    network.port_profiles.isl            -> {speed, breakout, lanes}   (bandwidth)
    fabrics.north_south.allocated_ports.isl[su] -> count, BOTH ENDS    (links)

`'default': 0` in allocated_ports.isl is not a defect — it encodes the guides'
literal `N/A` for the collapsed configurations, where there is no separate N/S
ISL because the leaf IS the E/W switch. Those SU counts are skipped, not failed.

Unit: the model's `allocated_ports.isl` is ERA-00011-001 v04 Table 15's
"ISL Ports (both ends)" column, published there as the count of twin-port OSFP
TRANSCEIVERS both ends (equivalently the MPO fibre count for a 2x-breakout ISL).
So 2-8-9-800 SU=2 -> 10 = 5 physical ports per core switch = 10 sub-links. The
check counts DISTINCT PHYSICAL ports at both ends (collapsing swpNsX breakout
sub-ports onto their parent), NOT sub-port endpoints: counting endpoints
double-counted every 2x-broken-out ISL and wrongly flagged correctly-cabled
workbooks (a real 2-8-9-800 4 SU submission scored 28 against a model of 14).
The generator's matching `// 2` under-cabling was corrected in the same change.

Known drift, pinned below: 2-4-5-800 disagrees with its model on both geometry
(`200G 4x2` vs `400G 2x4`) and count (40 ends vs 20). ADR-0042 moved the sheet
to fix an apply-breaking defect and left the model alone -- but ERA-00012
Table 15 backs the MODEL ("Twin-port OSFP, 2x400 Gbps", 20 at SU=1), so which
side is correct is open, not a simple stale model. Neither is auto-corrected.
"""
import re
import subprocess
import sys
from pathlib import Path

import pytest
# Arch models are internal-only (data-models/, ADR-0027) and absent from the
# public tree, so this skips there rather than failing. Defined locally rather
# than imported from conftest: conftest is loaded by pytest, not importable.
_HAVE_ARCH_MODELS = (Path(__file__).resolve().parents[2] / "data-models" / "models").exists()
needs_arch_models = pytest.mark.skipif(
    not _HAVE_ARCH_MODELS,
    reason="arch models are internal-only (data-models/), absent in the public tree")

REPO = Path(__file__).resolve().parent.parent
VALIDATOR = REPO / "scripts" / "validate_excel.py"
ARCHS = ["2-4-3-200", "2-4-5-800", "2-8-5-200",
         "2-8-9-400", "2-8-9-400-SP", "2-8-9-800"]

GEOM = "ISL geometry"
COUNT = "ISL link count"

# Exact sets. Closing one of these must fail the suite rather than silently
# shrink it (ADR-0040's KNOWN_BAD_SPEED_CELLS discipline).
# Both emptied 2026-08-05 by regenerating the workbooks from the models:
# 2-4-5-800's ISL now carries the model's 400G/2x/4 geometry instead of the
# sheet's stale 200G/4x/2, and its link count follows the model too.
KNOWN_GEOMETRY_DRIFT = set()
KNOWN_COUNT_DRIFT = set()


def _run(xlsx):
    return subprocess.run([sys.executable, str(VALIDATOR), str(xlsx)],
                          capture_output=True, text=True, cwd=str(REPO))


def _lines(arch, marker):
    xlsx = REPO / "input" / arch / "default" / f"{arch}.xlsx"
    if not xlsx.exists():
        pytest.skip(f"{arch} Excel absent")
    p = _run(xlsx)
    return [l for l in (p.stdout + p.stderr).splitlines() if marker in l]


@pytest.mark.parametrize("arch", ARCHS)
def test_isl_geometry_matches_model(arch):
    """Workbook ISL speed/breakout/lanes must equal network.port_profiles.isl."""
    found = bool(_lines(arch, GEOM))
    assert found == (arch in KNOWN_GEOMETRY_DRIFT), (
        f"{arch}: geometry drift reported={found}, "
        f"expected={arch in KNOWN_GEOMETRY_DRIFT}")


@pytest.mark.parametrize("arch", ARCHS)
def test_isl_link_count_matches_model(arch):
    """Wired ISL physical ports (both ends) must equal allocated_ports.isl[su]."""
    found = bool(_lines(arch, COUNT))
    assert found == (arch in KNOWN_COUNT_DRIFT), (
        f"{arch}: count drift reported={found}, "
        f"expected={arch in KNOWN_COUNT_DRIFT}")


def test_collapsed_su_counts_are_skipped_not_failed():
    """`'default': 0` encodes the guides' N/A; it must not read as "expect zero".

    2-4-3-200 (SU=8), 2-8-5-200 (SU=5) and 2-8-9-400 (SU=3) all resolve to the
    `'default'` entry and all have real ISLs in the Wire Map. Treating 0 as an
    expected count would fail all three.
    """
    for arch in ("2-4-3-200", "2-8-5-200", "2-8-9-400"):
        assert not _lines(arch, COUNT), f"{arch} should be skipped, not failed"


@needs_arch_models
def test_geometry_message_names_both_sides(tmp_path):
    """The drift message has to say what the model wants and what the sheet has.

    Drift is INJECTED rather than borrowed from a shipped workbook. This used to
    read 2-4-5-800's real 200G-vs-400G divergence; regenerating from the models
    removed it, and a message test that depends on a shipped defect goes silent
    the moment the defect is fixed — passing while proving nothing, or failing
    for a good reason as it did here.
    """
    import shutil

    import openpyxl

    arch = "2-4-5-800"
    src = REPO / "input" / arch / "default" / f"{arch}.xlsx"
    if not src.exists():
        pytest.skip(f"{arch} Excel absent")
    xlsx = tmp_path / f"{arch}.xlsx"
    shutil.copy(src, xlsx)

    # Rewrite the ISL row's geometry to disagree with the model (400G 2x4).
    wb = openpyxl.load_workbook(xlsx)
    ws = wb["VLANs & Profiles"]
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(r, 1).value or "").strip().upper() == "ISL":
            ws.cell(r, 8).value = "200G"
            ws.cell(r, 9).value = 4
            ws.cell(r, 10).value = 2
            break
    else:
        pytest.fail("no ISL port-profile row to perturb")
    wb.save(xlsx)

    p = _run(xlsx)
    lines = [l for l in (p.stdout + p.stderr).splitlines() if GEOM in l]
    assert lines, "injected ISL geometry drift was not reported"
    body = " ".join(lines)
    assert "400G" in body and "200G" in body, body
