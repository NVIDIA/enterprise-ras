# SPDX-FileCopyrightText: Copyright (c) 2026 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: MIT
"""
ERA-55 — the Port Profiles `Speed` column is read and reaches the inventory.

Column H of the `Port Profiles` block (inside `VLANs & Profiles`) has been
authored in every shipped workbook since the column existed, and was never
read: the parser took column 7 (LACP Bypass) then jumped to column 9
(Breakout). Nothing downstream could see it.

This round wires it in as far as the inventory ONLY. No template consumes
`speed` yet, and that is deliberate — see `test_speed_matches_lanes_rule`
below: 18 cells across the six shipped workbooks currently declare the
per-lane rate (100G) where the per-sub-port rate (200G) belongs. Emitting the
column before correcting those values would push `link speed 100G` onto ports
that production runs at 200G, which is worse than the column being dead.
"""
import re
import shutil
import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest
import yaml

REPO = Path(__file__).resolve().parent.parent
PARSER = REPO / "scripts" / "excel_parser.py"
ARCHS = ["2-4-3-200", "2-4-5-800", "2-8-5-200",
         "2-8-9-400", "2-8-9-400-SP", "2-8-9-800"]

# Every role dict that should carry the profile's Speed through to inventory.
_ROLE_KEYS = ("isl_interfaces", "oob_uplink_interfaces", "edge_interfaces",
              "gpu_interfaces", "storage_interfaces")


def _profile_rows(arch):
    """Yield (profile_name, speed, breakout, lanes) from a shipped workbook."""
    xl = REPO / "input" / arch / "default" / f"{arch}.xlsx"
    ws = openpyxl.load_workbook(xl, data_only=True)["VLANs & Profiles"]
    hdr = next((r for r in range(1, 60)
                if str(ws.cell(r, 1).value).strip() == "Profile"), None)
    assert hdr, f"{arch}: no Port Profiles header row found"
    for r in range(hdr + 1, hdr + 20):
        name = ws.cell(r, 1).value
        if not name:
            break
        yield (str(name).strip(), ws.cell(r, 8).value,
               ws.cell(r, 9).value, ws.cell(r, 10).value)


def _core_vars(arch):
    p = REPO / "output" / arch / "default" / "inventory" / "group_vars" / "core.yml"
    if not p.exists():
        pytest.skip(f"{arch} not generated")
    return yaml.safe_load(p.read_text())


# ------------------------------------------------------- the column is read ---

@pytest.mark.parametrize("arch", ARCHS)
def test_every_role_carries_a_speed(arch):
    """Pre-fix this is the discriminating assertion: `speed` was absent from
    every role dict because column 8 was never read."""
    core = _core_vars(arch)
    roles = {f"network_roles.{k}": v
             for k, v in (core.get("network_roles") or {}).items()}
    roles.update({k: core[k] for k in _ROLE_KEYS if core.get(k)})
    missing = [n for n, v in roles.items() if not v.get("speed")]
    assert not missing, f"{arch}: roles with no speed: {missing}"


@pytest.mark.parametrize("arch", ARCHS)
def test_inventory_speed_matches_the_workbook(arch):
    """The value in the inventory is the value in the cell — not a default."""
    by_name = {n: s for n, s, _, _ in _profile_rows(arch)}
    core = _core_vars(arch)
    # GPU Network and ISL have unambiguous 1:1 role mappings.
    for profile, key in (("GPU Network", "gpu_interfaces"), ("ISL", "isl_interfaces")):
        if by_name.get(profile) and core.get(key):
            assert core[key]["speed"] == str(by_name[profile]).strip(), (
                f"{arch}: {key} speed should come from the '{profile}' row")


def test_unrecognised_speed_is_rejected_not_propagated(tmp_path):
    """A malformed Speed warns and is dropped rather than reaching a config."""
    arch, site = "2-4-3-200", "era55-badspeed"
    in_dir = REPO / "input" / arch / site
    out_dir = REPO / "output" / arch / site
    in_dir.mkdir(parents=True, exist_ok=True)
    dst = in_dir / f"{arch}.xlsx"
    try:
        shutil.copy2(REPO / "input" / arch / "default" / f"{arch}.xlsx", dst)
        wb = openpyxl.load_workbook(dst)
        ws = wb["VLANs & Profiles"]
        hdr = next(r for r in range(1, 60)
                   if str(ws.cell(r, 1).value).strip() == "Profile")
        for r in range(hdr + 1, hdr + 20):
            if str(ws.cell(r, 1).value or "").strip() == "ISL":
                ws.cell(r, 8).value = "very-fast"
                break
        wb.save(dst)
        p = subprocess.run(
            [sys.executable, str(PARSER), "--arch", arch, "--site", site,
             "--skip-validate"], capture_output=True, text=True, cwd=str(REPO))
        assert p.returncode == 0, p.stdout + p.stderr
        assert "unrecognised Speed" in (p.stdout + p.stderr)
        core = yaml.safe_load(
            (out_dir / "inventory" / "group_vars" / "core.yml").read_text())
        assert "speed" not in core.get("isl_interfaces", {}), \
            "a malformed speed must be dropped, not passed through"
    finally:
        shutil.rmtree(in_dir, ignore_errors=True)
        shutil.rmtree(out_dir, ignore_errors=True)


# ------------------------------------------------ the values are still wrong ---

# ERA-55: sub-port speed == lanes x 100G, confirmed against ten July-2026
# production captures (e.g. mdf-c09r36-2894-spine-1: `swp1s0-3,swp2s0-2 link
# speed 200G` on breakout-4/lanes-2 ports) and independently by the NIC-model
# port_profiles block. Every row below declares the PER-LANE rate instead.
#
# This list is the reason no template emits `speed` yet. When the workbook
# cells are corrected, this test fails loudly and the list should empty out —
# that failure is the signal the data fix landed, not a regression.
KNOWN_BAD_SPEED_CELLS = set()
# Emptied 2026-08-05. All seventeen entries cleared when the workbooks were
# regenerated from the arch models:
#   * twelve corrected — CPU/In-Band and Support on all six archs, plus
#     2-4-3-200 Storage Uplink;
#   * five removed — the L2 `Storage` profile left the shipped workbooks with
#     ADR-0047 (storage is L3-only and external), taking its 100G/2-lane cell
#     with it. Note this is removal, not correction: `port_profiles.storage`
#     still declares 100G at 2 lanes in any model that re-adds L2 storage, so
#     fix that value before shipping an L2 storage sample.
#
# Keep this empty. A new entry means the rule regressed, not that the debt grew.


def test_speed_matches_lanes_rule():
    """`speed == lanes * 100G` on every profile that declares lanes.

    Rows still carrying the per-lane value are pinned in KNOWN_BAD_SPEED_CELLS
    so the debt is explicit and bounded. Asserting on the exact set (rather
    than 'no new failures') means correcting the data forces this list to be
    updated deliberately.
    """
    violations = set()
    for arch in ARCHS:
        for name, speed, _breakout, lanes in _profile_rows(arch):
            if not lanes or not speed:
                continue          # access ports (1G, no lanes) are exempt
            m = re.match(r"^(\d+)G$", str(speed).strip())
            assert m, f"{arch}/{name}: unparseable speed {speed!r}"
            if int(m.group(1)) != int(lanes) * 100:
                violations.add((arch, name))

    unexpected = violations - KNOWN_BAD_SPEED_CELLS
    assert not unexpected, (
        f"new speed/lanes mismatches (not in the ERA-55 known list): "
        f"{sorted(unexpected)}")

    fixed = KNOWN_BAD_SPEED_CELLS - violations
    assert not fixed, (
        f"these cells were corrected — remove them from KNOWN_BAD_SPEED_CELLS "
        f"and wire `speed` into the templates: {sorted(fixed)}")
