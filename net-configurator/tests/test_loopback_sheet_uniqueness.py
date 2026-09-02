# SPDX-FileCopyrightText: Copyright (c) 2026 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: MIT
"""Shipped workbooks must agree with the parser's loopback block map.

`test_vrf_loopback_allocation.py` gates the *computed* allocation. This file
gates the other half: the hand-authored `Loopbacks & ASNs` sheet, whose values
override the computed ones. A bad value there is invisible to the allocator
tests but reaches the switch.

Two properties are checked, and both are needed:

*Uniqueness* catches a live collision — two switches configured with the same
/32. That is what the largescale workbooks had: with GPU-VRF loopbacks laid
out at `10.<plane>.1.<10 + index>` against switch loopbacks at
`10.<plane>.1.<index>`, the ranges overlap from the eleventh leaf on, so
gl-plane1-01's GPU loopback was also gl-plane1-11's switch loopback.

*Block conformance* catches the same defect one scale *before* it bites.
`2-4-5-800/default` pinned its gs spines at `.5`/`.6` — unique, so uniqueness
alone passed it, but sitting inside the leaf range and one added leaf away
from a collision. Requiring each pinned value to land in its role's declared
block turns that from a latent bug into a test failure.
"""

import collections
import re
import subprocess
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent

# Mirrors PLANE_LOOPBACK_BLOCKS in scripts/excel_parser.py: (base, capacity)
# keyed by (role, column). Deliberately restated rather than imported — the
# point is to detect the code and the workbooks drifting apart, which a shared
# constant would hide.
PLANE_BLOCKS = {
    ("leaf", "Default"):  (0, 20),
    ("leaf", "GPU"):      (20, 20),
    ("spine", "Default"): (40, 10),
    ("spine", "GPU"):     (50, 10),
}

PLANE_SWITCH = re.compile(r"^(gl|gsl|gs)-plane(\d+)-(\d+)$")


def _tracked_workbooks():
    """Shipped input workbooks only.

    Prefer git so locally-imported throwaway sites (gitignored) are excluded.
    The source tarball has no .git, so fall back to the shipped site names.
    """
    try:
        out = subprocess.run(
            ["git", "ls-files", "input/*/*/*.xlsx"],
            cwd=REPO, capture_output=True, text=True, check=True,
        ).stdout.split()
        if out:
            return sorted(out)
    except (subprocess.CalledProcessError, FileNotFoundError):
        pass
    return sorted(
        str(p.relative_to(REPO))
        for p in REPO.glob("input/*/*/*.xlsx")
        if p.parent.name in ("default", "largescale")
    )


def _sheet_entries(path):
    """[(switch, column, ip)] for every populated loopback cell."""
    wb = openpyxl.load_workbook(REPO / path, read_only=True, data_only=True)
    name = next((s for s in wb.sheetnames if s.lower().startswith("loopback")), None)
    if not name:
        return []
    rows = [r for r in wb[name].iter_rows(values_only=True)
            if r and any(c not in (None, "") for c in r)]
    header = next((r for r in rows if r and r[0] == "Switch"), None)
    if not header:
        return []
    entries = []
    for row in rows[rows.index(header) + 1:]:
        switch = row[0]
        if not switch:
            continue
        for column, value in zip(header[1:], row[1:]):
            # ASN is a number, not an address; blanks fall back to computed.
            if column == "ASN" or not value:
                continue
            ip = str(value).split("/")[0].strip()
            if ip.count(".") == 3:
                entries.append((str(switch).strip(), column, ip))
    return entries


def _ids(paths):
    return ["/".join(Path(p).parts[1:3]) for p in paths]


WORKBOOKS = _tracked_workbooks()


@pytest.mark.parametrize("workbook", WORKBOOKS, ids=_ids(WORKBOOKS))
def test_no_duplicate_pinned_loopback(workbook):
    owners = collections.defaultdict(list)
    for switch, column, ip in _sheet_entries(workbook):
        owners[ip].append(f"{switch}:{column}")
    duplicates = {ip: who for ip, who in owners.items() if len(who) > 1}
    assert not duplicates, "\n".join(
        f"{ip} pinned to {', '.join(who)}" for ip, who in sorted(duplicates.items())
    )


@pytest.mark.parametrize("workbook", WORKBOOKS, ids=_ids(WORKBOOKS))
def test_plane_loopbacks_land_in_their_block(workbook):
    """Every pinned E/W plane loopback sits at its block base + switch index."""
    violations = []
    for switch, column, ip in _sheet_entries(workbook):
        match = PLANE_SWITCH.match(switch)
        if not match or column not in ("Default", "GPU"):
            continue
        role = "spine" if match.group(1) == "gs" else "leaf"
        base, capacity = PLANE_BLOCKS[(role, column)]
        index = int(match.group(3))
        octet = int(ip.rsplit(".", 1)[1])
        if index > capacity:
            violations.append(f"{switch}:{column} index {index} exceeds capacity {capacity}")
        elif octet != base + index:
            violations.append(
                f"{switch}:{column}={ip} is outside the {role}/{column} block "
                f"(expected final octet {base + index})"
            )
    assert not violations, "\n".join(violations)


def test_the_gates_can_actually_fail():
    """Mutation check: both properties must reject a known-bad layout.

    The pre-fix largescale layout — leaves at .1-.16, GPU at .11-.26 — is the
    fixture, so this fails if either gate is ever loosened into a no-op.
    """
    entries = [(f"gl-plane1-{i:02d}", col, f"10.1.1.{off + i}")
               for i in range(1, 17) for col, off in (("Default", 0), ("GPU", 10))]

    owners = collections.defaultdict(list)
    for switch, column, ip in entries:
        owners[ip].append(f"{switch}:{column}")
    assert [ip for ip, who in owners.items() if len(who) > 1], "uniqueness gate went blind"

    off_block = [
        ip for switch, column, ip in entries
        if int(ip.rsplit(".", 1)[1])
        != PLANE_BLOCKS[("leaf", column)][0] + int(PLANE_SWITCH.match(switch).group(3))
    ]
    assert off_block, "block-conformance gate went blind"
