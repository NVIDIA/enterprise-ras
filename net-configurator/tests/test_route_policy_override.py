# SPDX-FileCopyrightText: Copyright (c) 2026 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: MIT
"""
Unit tests for the optional 'Route policy' / 'Community lists' Excel sheets
. Same derive-by-default + full-override pattern as the
'Prefix lists' sheet (Task 2/3, see test_prefix_list_override.py): the
sheets are parser-agnostic to override-vs-add — that decision is made by
`_apply_named_overrides`, which the caller (`generate_group_vars`) applies
to the source-inventory `route_map` / `community_list` lists it already
merged in.
"""
import sys
from pathlib import Path

import openpyxl
import pytest

SCRIPTS_DIR = Path(__file__).parent.parent / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from excel_parser import (
    parse_route_policy_sheet,
    parse_community_lists_sheet,
    _apply_named_overrides,
)


class TestParseRoutePolicySheet:
    """Columns: Route-map | Rule | Action | Match type | Match value |
    Set type | Set value. A rule may span multiple rows to add additional
    match/set entries."""

    def test_parse_exit_filter_rules(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Route-map", "Rule", "Action", "Match type", "Match value",
                   "Set type", "Set value"])
        ws.append(["EXIT_FILTER", "10", "deny", "ip-prefix-list", "EXIT_LOCAL_IF", "", ""])
        ws.append(["EXIT_FILTER", "20", "permit", "", "", "community", "11:11"])

        d = parse_route_policy_sheet(ws)

        assert d["EXIT_FILTER"] == [
            {
                "id": "10",
                "action": "deny",
                "match": [{"type": "ip-prefix-list", "value": "EXIT_LOCAL_IF"}],
            },
            {
                "id": "20",
                "action": "permit",
                "set": [{"type": "community", "value": "11:11"}],
            },
        ]

    def test_rule_spanning_multiple_rows_collects_multiple_matches(self):
        """A single rule (same route-map + rule id) spread across multiple
        rows accumulates all match/set entries in row order."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Route-map", "Rule", "Action", "Match type", "Match value",
                   "Set type", "Set value"])
        ws.append(["BLOCK_VTEPS", "10", "deny", "ip-prefix-list", "VTEP_PREFIXES", "", ""])
        ws.append(["BLOCK_VTEPS", "10", "deny", "type", "ipv4", "", ""])

        d = parse_route_policy_sheet(ws)

        assert d["BLOCK_VTEPS"] == [
            {
                "id": "10",
                "action": "deny",
                "match": [
                    {"type": "ip-prefix-list", "value": "VTEP_PREFIXES"},
                    {"type": "type", "value": "ipv4"},
                ],
            }
        ]

    def test_blank_match_and_set_omit_keys(self):
        """A rule with neither match nor set columns filled gets no
        'match'/'set' key at all (matches the source-inventory shape, which
        never emits empty lists)."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Route-map", "Rule", "Action", "Match type", "Match value",
                   "Set type", "Set value"])
        ws.append(["PLAIN_PERMIT", "10", "permit", "", "", "", ""])

        d = parse_route_policy_sheet(ws)

        assert d["PLAIN_PERMIT"] == [{"id": "10", "action": "permit"}]

    def test_blank_route_map_row_skipped(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Route-map", "Rule", "Action", "Match type", "Match value",
                   "Set type", "Set value"])
        ws.append([None, None, None, None, None, None, None])
        ws.append(["EXIT_FILTER", "10", "deny", "ip-prefix-list", "EXIT_LOCAL_IF", "", ""])

        d = parse_route_policy_sheet(ws)

        assert list(d.keys()) == ["EXIT_FILTER"]

    def test_empty_sheet_returns_empty_dict(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Route-map", "Rule", "Action", "Match type", "Match value",
                   "Set type", "Set value"])

        assert parse_route_policy_sheet(ws) == {}


class TestParseCommunityListsSheet:
    """Columns: Community-list | Rule | Action | Community."""

    def test_parse_community_list_11(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Community-list", "Rule", "Action", "Community"])
        ws.append(["11", "100", "permit", "11:11"])

        d = parse_community_lists_sheet(ws)

        assert d["11"] == [{"id": "100", "action": "permit", "community": "11:11"}]

    def test_multiple_rules_same_community_list(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Community-list", "Rule", "Action", "Community"])
        ws.append(["11", "100", "permit", "11:11"])
        ws.append(["11", "200", "deny", "22:22"])

        d = parse_community_lists_sheet(ws)

        assert d["11"] == [
            {"id": "100", "action": "permit", "community": "11:11"},
            {"id": "200", "action": "deny", "community": "22:22"},
        ]

    def test_blank_row_skipped(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Community-list", "Rule", "Action", "Community"])
        ws.append([None, None, None, None])
        ws.append(["11", "100", "permit", "11:11"])

        d = parse_community_lists_sheet(ws)

        assert list(d.keys()) == ["11"]

    def test_empty_sheet_returns_empty_dict(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Community-list", "Rule", "Action", "Community"])

        assert parse_community_lists_sheet(ws) == {}


class TestApplyNamedOverrides:
    """`_apply_named_overrides(base_list, overrides)` — shared override-merge
    helper used for both `route_map` and `community_list` (both are lists of
    `{'id': ..., 'rule': [...]}` entries)."""

    BASE_ROUTE_MAP = [
        {"id": "BLOCK_VTEPS", "rule": [{"id": "10", "action": "deny"}]},
        {"id": "EXIT_FILTER", "rule": [{"id": "10", "action": "deny"}]},
        {"id": "INBAND_FILTER", "rule": [{"id": "5", "action": "deny"}]},
    ]

    def test_same_named_entry_is_replaced(self):
        overrides = {"EXIT_FILTER": [{"id": "10", "action": "permit"}]}
        result = _apply_named_overrides(self.BASE_ROUTE_MAP, overrides)

        by_id = {e["id"]: e for e in result}
        assert by_id["EXIT_FILTER"] == {
            "id": "EXIT_FILTER",
            "rule": [{"id": "10", "action": "permit"}],
        }
        # Others untouched
        assert by_id["BLOCK_VTEPS"] == self.BASE_ROUTE_MAP[0]
        assert by_id["INBAND_FILTER"] == self.BASE_ROUTE_MAP[2]

    def test_replace_preserves_list_order(self):
        overrides = {"EXIT_FILTER": [{"id": "99", "action": "permit"}]}
        result = _apply_named_overrides(self.BASE_ROUTE_MAP, overrides)

        assert [e["id"] for e in result] == ["BLOCK_VTEPS", "EXIT_FILTER", "INBAND_FILTER"]

    def test_new_entry_is_appended(self):
        overrides = {"CUSTOM_MAP": [{"id": "10", "action": "permit"}]}
        result = _apply_named_overrides(self.BASE_ROUTE_MAP, overrides)

        assert [e["id"] for e in result] == [
            "BLOCK_VTEPS", "EXIT_FILTER", "INBAND_FILTER", "CUSTOM_MAP",
        ]
        by_id = {e["id"]: e for e in result}
        assert by_id["CUSTOM_MAP"] == {
            "id": "CUSTOM_MAP",
            "rule": [{"id": "10", "action": "permit"}],
        }

    def test_base_list_not_mutated(self):
        overrides = {"EXIT_FILTER": [{"id": "10", "action": "permit"}]}
        original = [dict(e) for e in self.BASE_ROUTE_MAP]

        _apply_named_overrides(self.BASE_ROUTE_MAP, overrides)

        assert self.BASE_ROUTE_MAP == original

    def test_community_list_shape_also_works(self):
        base_community_list = [
            {"id": "11", "rule": [{"id": "100", "action": "permit", "community": "11:11"}]},
        ]
        overrides = {"11": [{"id": "100", "action": "deny", "community": "99:99"}]}

        result = _apply_named_overrides(base_community_list, overrides)

        assert result == [
            {"id": "11", "rule": [{"id": "100", "action": "deny", "community": "99:99"}]},
        ]

    def test_no_overrides_returns_base_list_unchanged(self):
        assert _apply_named_overrides(self.BASE_ROUTE_MAP, {}) is self.BASE_ROUTE_MAP
        assert _apply_named_overrides(self.BASE_ROUTE_MAP, None) is self.BASE_ROUTE_MAP


class TestGenerateGroupVarsDeriveByDefault:
    """Absent 'Route policy' / 'Community lists' sheets ⇒ the merged
    route_map/community_list stay exactly what the source inventory
    provides — no regression to the pre-Task-5 behavior."""

    def test_generate_group_vars_without_directives_leaves_route_map_and_community_list_untouched(self, tmp_path):
        from excel_parser import generate_group_vars

        settings = {"architecture": "2-8-9-400"}
        vlans = []
        vrfs = {}
        output_dir = tmp_path / "inventory"
        output_dir.mkdir(parents=True)

        generate_group_vars(
            settings, vlans, vrfs, output_dir, "2-8-9-400",
            route_policy_directives=None, community_list_directives=None,
        )

        import yaml
        core_file = output_dir / "group_vars" / "core.yml"
        with open(core_file) as f:
            core_vars = yaml.safe_load(f)

        # The derive-by-default source is now the single-home
        # consolidated defaults (inventory_defaults), not the per-arch seed.
        from inventory_defaults import arch_group_vars
        source_vars = arch_group_vars("core", "2-8-9-400")

        assert core_vars["route_map"] == source_vars["route_map"]
        assert core_vars["community_list"] == source_vars["community_list"]

    def test_generate_group_vars_with_directives_overrides_route_map_and_community_list(self, tmp_path):
        from excel_parser import generate_group_vars

        settings = {"architecture": "2-8-9-400"}
        vlans = []
        vrfs = {}
        output_dir = tmp_path / "inventory"
        output_dir.mkdir(parents=True)

        route_policy_directives = {
            "EXIT_FILTER": [{"id": "10", "action": "permit"}],
        }
        community_list_directives = {
            "11": [{"id": "100", "action": "deny", "community": "99:99"}],
        }

        generate_group_vars(
            settings, vlans, vrfs, output_dir, "2-8-9-400",
            route_policy_directives=route_policy_directives,
            community_list_directives=community_list_directives,
        )

        import yaml
        core_file = output_dir / "group_vars" / "core.yml"
        with open(core_file) as f:
            core_vars = yaml.safe_load(f)

        route_map_by_id = {e["id"]: e for e in core_vars["route_map"]}
        assert route_map_by_id["EXIT_FILTER"] == {
            "id": "EXIT_FILTER",
            "rule": [{"id": "10", "action": "permit"}],
        }
        # BLOCK_VTEPS was untouched
        assert "BLOCK_VTEPS" in route_map_by_id

        community_list_by_id = {e["id"]: e for e in core_vars["community_list"]}
        assert community_list_by_id["11"] == {
            "id": "11",
            "rule": [{"id": "100", "action": "deny", "community": "99:99"}],
        }
