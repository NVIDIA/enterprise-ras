# SPDX-FileCopyrightText: Copyright (c) 2026 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: MIT
"""Validator coverage for the Nodes-tab Type column + Enabled=Air semantics.

Added 2026-05-28 with the schema change:
  - Type column (optional) — {switch, node}
  - Enabled=Air — documentary row for auto-injected Air infrastructure
  - Old Excels without these columns must continue to validate (back-compat).
"""
import sys
from pathlib import Path

import openpyxl
import pytest

SCRIPTS_DIR = Path(__file__).parent.parent / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

from validate_excel import validate_nodes, ValidationResult


NODES_HEADERS_LEGACY = [
    "Function", "Name", "MAC Address for ZTP", "Mgmt IP Address",
    "Prefix", "Gateway", "ZTP", "Enabled",
]
NODES_HEADERS_NEW = NODES_HEADERS_LEGACY + ["Type", "Notes"]


def _nodes(rows, headers=NODES_HEADERS_NEW):
    """Build an openpyxl Nodes sheet from rows. Headers default to new schema."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Nodes"
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    for r, row in enumerate(rows, start=2):
        for c, v in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=v)
    return ws


def _run(ws, settings=None):
    """Run validate_nodes against the sheet; return (result, parsed_nodes)."""
    result = ValidationResult()
    parsed_nodes = validate_nodes(ws, result, settings=settings or {})
    return result, parsed_nodes


def _errors_matching(result, needle):
    return [e for e in result.errors if needle in e]


# ─── Type column basic validation ────────────────────────────────────────

def test_legacy_excel_without_type_column_validates():
    """Old Excels that don't have a Type column must still validate."""
    ws = _nodes(
        [("core", "core-01", None, "192.168.200.2", 24, "192.168.200.1", "Yes", "Yes")],
        headers=NODES_HEADERS_LEGACY,
    )
    result, _ = _run(ws)
    # Should NOT complain about a missing Type column
    assert not _errors_matching(result, "Type"), result.errors


def test_invalid_type_value_is_flagged():
    ws = _nodes([
        ("core", "core-01", None, "192.168.200.2", 24, "192.168.200.1",
         "Yes", "Yes", "router", None),  # 'router' is not switch or node
    ])
    result, _ = _run(ws)
    bad = _errors_matching(result, "Type 'router'")
    assert len(bad) == 1, result.errors


def test_blank_type_value_is_silent():
    """Type column present but cell empty should not error (gives operators
    a graceful path while transitioning Excels)."""
    ws = _nodes([
        ("core", "core-01", None, "192.168.200.2", 24, "192.168.200.1",
         "Yes", "Yes", "", None),
    ])
    result, _ = _run(ws)
    assert not _errors_matching(result, "Type"), result.errors


# ─── Function ↔ Type consistency (Enabled=Yes/No) ────────────────────────

def test_switch_function_with_node_type_errors():
    ws = _nodes([
        ("core", "core-01", None, "192.168.200.2", 24, "192.168.200.1",
         "Yes", "Yes", "node", None),
    ])
    result, _ = _run(ws)
    assert _errors_matching(result, "is a switch role but Type='node'"), result.errors


def test_server_function_with_switch_type_errors():
    ws = _nodes([
        ("gpu", "gpu-01", None, "192.168.200.10", 24, "192.168.200.1",
         "Yes", "Yes", "switch", None),
    ])
    result, _ = _run(ws)
    assert _errors_matching(result, "is a server/node role but Type='switch'"), result.errors


def test_consistent_switch_function_and_type_passes():
    ws = _nodes([
        ("core", "core-01", None, "192.168.200.2", 24, "192.168.200.1",
         "Yes", "Yes", "switch", None),
    ])
    result, _ = _run(ws)
    assert not _errors_matching(result, "Type"), result.errors


def test_consistent_node_function_and_type_passes():
    ws = _nodes([
        ("gpu", "gpu-01", None, "192.168.200.10", 24, "192.168.200.1",
         "Yes", "Yes", "node", None),
    ])
    result, _ = _run(ws)
    assert not _errors_matching(result, "Type"), result.errors


def test_type_value_case_insensitive():
    """SWITCH and Switch should both work — keep operators from being
    burned by case-sensitivity on a free-text field."""
    ws = _nodes([
        ("core", "core-01", None, "192.168.200.2", 24, "192.168.200.1",
         "Yes", "Yes", "Switch", None),
    ])
    result, _ = _run(ws)
    assert not _errors_matching(result, "Type"), result.errors


# ─── Enabled=Air semantics ───────────────────────────────────────────────

def test_air_documentary_row_with_known_name_passes():
    """An Enabled=Air row pointing at a known Air-only node validates."""
    ws = _nodes([
        ("edge", "cust-net-edge-01", None, None, None, None,
         "No", "Air", "switch", "Air-only. Customer edge sim"),
    ])
    result, _ = _run(ws)
    assert not _errors_matching(result, "cust-net-edge-01"), result.errors


def test_air_documentary_row_with_unknown_name_errors():
    """Enabled=Air on an unknown name should hard-error."""
    ws = _nodes([
        ("edge", "my-random-node", None, None, None, None,
         "No", "Air", "switch", "made up name"),
    ])
    result, _ = _run(ws)
    assert _errors_matching(result, "not a known Air-only node"), result.errors


def test_air_unknown_name_errors_on_legacy_excel_without_type_column():
    """Regression: Enabled=Air must validate the Name even on a legacy
    (Type-less) Excel. A real node typo'd as Enabled=Air would otherwise
    silently drop out of provisioning with zero diagnostics."""
    ws = _nodes(
        [("edge", "totally-made-up-node", None, None, None, None, "No", "Air")],
        headers=NODES_HEADERS_LEGACY,
    )
    result, _ = _run(ws)
    assert _errors_matching(result, "not a known Air-only node"), result.errors


def test_air_known_name_passes_on_legacy_excel_without_type_column():
    """A known Air-only node marked Enabled=Air still validates with no
    Type column present (T3 type-match is skipped, T2 name check passes)."""
    ws = _nodes(
        [("utility", "utility", None, None, None, None, "No", "Air")],
        headers=NODES_HEADERS_LEGACY,
    )
    result, _ = _run(ws)
    assert not _errors_matching(result, "not a known Air-only node"), result.errors


def test_air_documentary_row_with_wrong_type_errors():
    """utility is a node — declaring it as a switch is wrong."""
    ws = _nodes([
        ("utility", "utility", None, None, None, None,
         "No", "Air", "switch", None),
    ])
    result, _ = _run(ws)
    assert _errors_matching(result, "utility' is a node"), result.errors


def test_reserved_name_with_enabled_air_is_allowed():
    """`air-oob-switch` is in the reserved-names list. Without Enabled=Air
    that's a collision error; WITH Enabled=Air it's a documentary row."""
    # Without Enabled=Air → collision
    ws_bad = _nodes([
        ("air-oob", "air-oob-switch", None, "192.168.200.99", 24,
         "192.168.200.1", "Yes", "Yes", "switch", None),
    ])
    bad_result, _ = _run(ws_bad)
    assert _errors_matching(bad_result, "reserved for Air-injected"), bad_result.errors

    # With Enabled=Air → no collision error
    ws_good = _nodes([
        ("air-oob", "air-oob-switch", None, None, None, None,
         "No", "Air", "switch", "Air-only L2-mode legacy"),
    ])
    good_result, _ = _run(ws_good)
    assert not _errors_matching(good_result, "reserved for Air-injected"), \
        good_result.errors


def test_air_documentary_row_skips_mgmt_ip_required():
    """Enabled=Air rows don't need a Mgmt IP — they document auto-injected
    infra that isn't operator-provisioned."""
    ws = _nodes([
        ("core", "core-01", None, "192.168.200.2", 24, "192.168.200.1",
         "Yes", "Yes", "switch", None),  # real node, has IP, passes
        ("utility", "utility", None, None, None, None,
         "No", "Air", "node", "Air-only jumpbox"),  # no IP, but Air row
    ])
    result, _ = _run(ws, settings={'architecture': '2-4-3-200'})
    assert not _errors_matching(result, "utility): Missing management IP"), \
        result.errors


# ─── Arch-restricted Function values ─────────────────────────────────────

def test_ext_storage_on_collapsed_core_arch_errors():
    """ext-storage is only valid on 2-8-9-800. Putting it on 2-4-3-200
    would deploy an inert Air VM (no CSL templates peering to it)."""
    ws = _nodes([
        ("ext-storage", "ext-storage-01", None, None, None, None,
         "Yes", "Yes", "node", None),
    ])
    result, _ = _run(ws, settings={'architecture': '2-4-3-200'})
    assert _errors_matching(result, "is only valid on arch"), result.errors


def test_ext_storage_on_dual_plane_arch_passes():
    ws = _nodes([
        ("ext-storage", "ext-storage-01", None, None, None, None,
         "Yes", "Yes", "node", None),
    ])
    result, _ = _run(ws, settings={'architecture': '2-8-9-800'})
    assert not _errors_matching(result, "is only valid on arch"), result.errors


def test_csl_on_collapsed_core_arch_errors():
    """CSL switches don't exist in collapsed-core archs."""
    ws = _nodes([
        ("csl", "csl-01", None, "192.168.200.5", 24, "192.168.200.1",
         "Yes", "Yes", "switch", None),
    ])
    result, _ = _run(ws, settings={'architecture': '2-8-5-200'})
    assert _errors_matching(result, "is only valid on arch"), result.errors


def test_core_on_dual_plane_arch_errors():
    """core switches don't exist in dual-plane (it uses csl/gsl instead)."""
    ws = _nodes([
        ("core", "core-01", None, "192.168.200.2", 24, "192.168.200.1",
         "Yes", "Yes", "switch", None),
    ])
    result, _ = _run(ws, settings={'architecture': '2-8-9-800'})
    assert _errors_matching(result, "is only valid on arch"), result.errors


def test_arch_check_skipped_for_air_documentary_rows():
    """Documentary Air rows can carry reserved Function values (e.g.
    cust-net-edge with Function=edge) regardless of arch. The arch
    check must not fire on them."""
    ws = _nodes([
        ("edge", "cust-net-edge-01", None, None, None, None,
         "No", "Air", "switch", "Air-only"),
    ])
    result, _ = _run(ws, settings={'architecture': '2-4-3-200'})
    # 'edge' isn't in ARCH_RESTRICTED_FUNCTIONS anyway, but also confirm
    # the check is suppressed for Air-doc rows generally.
    assert not _errors_matching(result, "is only valid on arch"), result.errors
