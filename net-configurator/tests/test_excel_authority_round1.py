# SPDX-FileCopyrightText: Copyright (c) 2026 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: MIT
"""
Forcing-function tests for the ADR-0035 "Excel is authoritative" round.

Each test pins a behaviour that was previously *silent*: a value the workbook
could not influence, or a fallback that fired without saying so. They drive the
REAL entry points as subprocesses (excel_parser.py / validate_excel.py) against
mutated copies of a shipped default workbook, so they exercise the path an
operator hits via ``make import`` / ``make validate-excel``.

Discriminating-by-construction: the oob_vlan cases are built so the pre-ADR-0035
parser (which keyed off ``name.startswith('OOB')``) would return a *different*
answer, not merely skip a warning.
"""
import re
import shutil
import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest

REPO = Path(__file__).resolve().parent.parent
PARSER = REPO / "scripts" / "excel_parser.py"
VALIDATE = REPO / "scripts" / "validate_excel.py"

ARCH = "2-8-9-800"
BASE = REPO / "input" / ARCH / "default" / f"{ARCH}.xlsx"

SCRIPTS_DIR = REPO / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

pytestmark = pytest.mark.skipif(
    not BASE.exists(), reason=f"default {ARCH} Excel not present"
)


# ---------------------------------------------------------------- harness ---

def _site_dirs(site):
    return (REPO / "input" / ARCH / site), (REPO / "output" / ARCH / site)


@pytest.fixture
def gen(request):
    """Build a scratch site from the shipped workbook, mutate it, run the real
    parser, and hand back (stdout+stderr, inventory_dir). Cleans up both the
    input and output site trees afterwards (non-default sites are gitignored)."""
    made = []

    def _run(site, mutate=None):
        in_dir, out_dir = _site_dirs(site)
        made.append((in_dir, out_dir))
        in_dir.mkdir(parents=True, exist_ok=True)
        dst = in_dir / f"{ARCH}.xlsx"
        shutil.copy2(BASE, dst)
        if mutate is not None:
            wb = openpyxl.load_workbook(dst)
            mutate(wb)
            wb.save(dst)
        p = subprocess.run(
            [sys.executable, str(PARSER), "--arch", ARCH, "--site", site,
             "--skip-validate"],
            capture_output=True, text=True, cwd=str(REPO))
        return p, (out_dir / "inventory")

    yield _run

    for in_dir, out_dir in made:
        shutil.rmtree(in_dir, ignore_errors=True)
        shutil.rmtree(out_dir, ignore_errors=True)


def _yaml(path):
    import yaml
    return yaml.safe_load(path.read_text())


def _vlan_rows(ws):
    """Row indices of the VLAN table (a numeric VLAN ID in column A)."""
    return [r for r in range(1, ws.max_row + 1)
            if str(ws.cell(r, 1).value or "").strip().isdigit()]


# ------------------------------------------------------- D-9: oob_vlan -----

def test_oob_vlan_follows_vrf_column_not_the_name(gen):
    """The OOB VLAN is identified by VRF == 'OOB', so a lowercase / non-'OOB'-
    prefixed Name still resolves — and a non-200 ID is honoured.

    Pre-ADR-0035 this returned 200 (name.startswith('OOB') missed 'mgmt-oob',
    so the silent fallback fired), which is the wrong VLAN."""
    def mutate(wb):
        ws = wb["VLANs & Profiles"]
        for r in _vlan_rows(ws):
            if str(ws.cell(r, 6).value or "").strip().upper() == "OOB":
                ws.cell(r, 1).value = 250          # not the fallback value
                ws.cell(r, 2).value = "mgmt-oob"   # lowercase, no 'OOB' prefix
                break

    p, inv = gen("authz-vrf", mutate)
    assert p.returncode == 0, p.stdout + p.stderr
    oob = _yaml(inv / "group_vars" / "oob.yml")
    assert oob["oob_vlan"] == "250", (
        f"oob_vlan should follow the VRF column to VLAN 250, got {oob['oob_vlan']!r}")
    assert "No VLAN with VRF 'OOB'" not in (p.stdout + p.stderr)


def test_missing_oob_vrf_warns_instead_of_silently_defaulting(gen):
    """With no VRF=='OOB' VLAN the parser still falls back to 200, but says so."""
    def mutate(wb):
        ws = wb["VLANs & Profiles"]
        for r in _vlan_rows(ws):
            if str(ws.cell(r, 6).value or "").strip().upper() == "OOB":
                ws.cell(r, 6).value = "INBAND"   # no OOB-VRF VLAN remains
                break

    p, inv = gen("authz-novrf", mutate)
    assert p.returncode == 0, p.stdout + p.stderr
    assert "No VLAN with VRF 'OOB'" in (p.stdout + p.stderr)
    assert _yaml(inv / "group_vars" / "oob.yml")["oob_vlan"] == "200"


# -------------------------------------------- D-8: OOB port fallback -------

def test_oob_switch_without_wiremap_rows_warns(gen):
    """A switch with no Wire Map rows gets reference port pins — loudly."""
    def mutate(wb):
        ws = wb["Wire Map"]
        for r in range(ws.max_row, 1, -1):
            if any("oob-switch-02" == str(ws.cell(r, c).value or "").strip()
                   for c in (2, 6)):
                ws.delete_rows(r)

    p, inv = gen("authz-nowm", mutate)
    assert p.returncode == 0, p.stdout + p.stderr
    out = p.stdout + p.stderr
    assert "No Wire Map rows for OOB switch" in out and "oob-switch-02" in out
    hv = _yaml(inv / "host_vars" / "oob-switch-02.yml")
    assert hv["access_ports"] == "swp1-48"   # the reference pins, as warned


# ----------------------------------------------- C-7: scale facts gone -----

def test_scale_facts_are_not_emitted(gen):
    """scalable_units / nodes_per_su had no Excel source and no reader; they
    must not reappear as invented facts in the generated inventory."""
    p, inv = gen("authz-scale")
    assert p.returncode == 0, p.stdout + p.stderr
    all_vars = _yaml(inv / "group_vars" / "all" / "main.yml")
    assert "scalable_units" not in all_vars
    assert "nodes_per_su" not in all_vars


# ------------------------------------------ E-*: stale defaults cleared ----

def test_inventory_defaults_carry_no_stale_shadowed_keys():
    """Keys that were shadowed by host_vars, or had zero consumers, are gone."""
    from inventory_defaults import arch_group_vars

    archs = ["2-4-3-200", "2-4-5-800", "2-8-5-200",
             "2-8-9-400", "2-8-9-400-SP", "2-8-9-800"]
    for arch in archs:
        oob = arch_group_vars("oob", arch)
        for dead in ("pre_login_message", "spine_bond_members", "overlay_peers",
                     "oob_vni", "oob_vrf", "vrf_oob_evpn_vlan", "vrf_oob_vni",
                     "spine_uplink_ports", "access_ports"):
            assert dead not in oob, f"oob/{arch} still carries {dead}"
        assert "pre_login_message" not in arch_group_vars("core", arch)
        # all_shared.ldap was unreachable — excel_parser always sets all_vars['ldap']
        assert "ldap" not in arch_group_vars("all", arch)


def test_banner_comes_only_from_common(gen):
    """The login banner is Excel-sourced via `common`; no bare pre_login_message
    key should be emitted into core.yml / oob.yml to look like the source."""
    p, inv = gen("authz-banner")
    assert p.returncode == 0, p.stdout + p.stderr
    for f in ("core.yml", "oob.yml"):
        assert "pre_login_message" not in _yaml(inv / "group_vars" / f)
    common = _yaml(inv / "group_vars" / "all" / "main.yml")["common"]
    assert common["pre_login_message"].strip(), "Excel banner should still flow"


# ------------------------------------- E-12: LDAP survived the removal ----

# Every key the deleted `all_shared.ldap` block carried. The parser must still
# produce all of them, or removing that block silently amputated LDAP.
_LDAP_KEYS = {"enabled", "domain", "organization", "admin_password",
              "base_dn", "root_dn", "servers", "users"}


def test_ldap_still_fully_generated_when_enabled(gen):
    """`all_shared.ldap` was removed as unreachable (the parser always sets
    all_vars['ldap']). Prove LDAP is intact end-to-end when an operator turns it
    on: every key present, users generated, servers Excel-driven."""
    def mutate(wb):
        ws = wb["Settings"]
        for row in ws.iter_rows():
            if str(row[0].value or "").strip() == "ldap_enabled":
                ws.cell(row[0].row, 2).value = "Yes"
                return

    p, inv = gen("authz-ldapon", mutate)
    assert p.returncode == 0, p.stdout + p.stderr
    ldap = _yaml(inv / "group_vars" / "all" / "main.yml")["ldap"]

    assert set(ldap) == _LDAP_KEYS, f"LDAP key parity lost: {set(ldap) ^ _LDAP_KEYS}"
    assert ldap["enabled"] is True
    # Users are still generated (roles/ldap/tasks/main.yml loops over ldap.users).
    assert [u["username"] for u in ldap["users"]] == ["jdoe", "asmith"]
    for u in ldap["users"]:
        assert {"firstname", "lastname", "username", "password"} <= set(u)
    # Servers come from the Excel ldap_servers field, not a hardcoded default.
    assert ldap["servers"] and ldap["servers"][0]["ip"] == "172.20.0.78"
    assert ldap["base_dn"] and ldap["root_dn"]


def test_ldap_disabled_by_default_in_shipped_workbooks(gen):
    """The shipped workbooks ship ldap_enabled=No; the removed defaults block
    had `enabled: true`, so a regression that reinstated it would flip this."""
    p, inv = gen("authz-ldapoff")
    assert p.returncode == 0, p.stdout + p.stderr
    ldap = _yaml(inv / "group_vars" / "all" / "main.yml")["ldap"]
    assert ldap["enabled"] is False
    assert set(ldap) == _LDAP_KEYS, "keys must exist even when disabled"


# ------------------------------- inventory header describes itself truly ---

_SWITCH_GROUPS = ("core", "csl", "cl", "cs", "gsl_plane1", "gsl_plane2",
                  "gl_plane1", "gl_plane2", "gs_plane1", "gs_plane2", "oob")


def _group_members(hosts_text, group):
    m = re.search(rf'^\[{group}\]\n((?:[^\[\n].*\n?)*)', hosts_text, re.M)
    return [x for x in m.group(1).splitlines() if x.strip()] if m else []


def test_inventory_header_switch_count_matches_the_groups(gen):
    """The header census must equal the switches actually in the inventory.

    It previously counted only categories['core'] + ['oob'], so 2-8-9-800 read
    "2 Core switches, 2 OOB switches" while carrying 2 CSL + 4 GSL — CSL was
    mislabelled Core and every GPU-fabric switch was omitted."""
    p, inv = gen("authz-hdr")
    assert p.returncode == 0, p.stdout + p.stderr
    txt = (inv / "hosts").read_text()

    header = next(l for l in txt.splitlines()
                  if l.startswith("# ") and "switches" in l)
    claimed = int(re.search(r"# (\d+) switches", header).group(1))
    actual = sum(len(_group_members(txt, g)) for g in _SWITCH_GROUPS)
    assert claimed == actual, f"header claims {claimed}, inventory has {actual}: {header}"

    # 2-8-9-800 is a dedicated-GPU arch: CSL must be named as CSL, GSL counted,
    # and "Core" must not appear (there are no core-category switches).
    assert "2 CSL" in header and "4 GSL" in header, header
    assert "Core" not in header, f"CSL mislabelled as Core: {header}"


# ---------------------------------------------- A-1: dead key surfaced -----

def test_ztp_enabled_is_reported_dead_not_silently_accepted():
    """ztp_enabled has no consumer; validate must say so rather than accept it."""
    p = subprocess.run([sys.executable, str(VALIDATE), str(BASE)],
                       capture_output=True, text=True)
    out = p.stdout + p.stderr
    assert "ztp_enabled" in out and "no effect" in out
    assert p.returncode == 0, "a dead key is a warning, not a hard failure"
