# SPDX-FileCopyrightText: Copyright (c) 2026 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: MIT
"""validate_excel warns when a pinned E/W plane loopback leaves its block.

`test_loopback_sheet_uniqueness.py` gates the workbooks *we* ship. This gates
the customer path: an operator's own workbook only ever meets `make
validate-excel`, so the block check has to live there too.

The duplicate-IP check already catches a plane that has *already* collided.
The value of this one is the scale before that — `2-4-5-800/default` pinned
its gs spines at .5/.6, which is unique and so passed every existing check,
while sitting in the leaf range one added leaf away from a collision.
"""
import sys
from pathlib import Path

import openpyxl

SCRIPTS_DIR = Path(__file__).parent.parent / "scripts"
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

import validate_excel as ve  # noqa: E402

HEADER = ("Switch", "Default", "GPU")


def _ws(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Loopbacks"
    for j, h in enumerate(HEADER, 1):
        ws.cell(1, j, h)
    for i, r in enumerate(rows, 2):
        for j, v in enumerate(r, 1):
            ws.cell(i, j, v)
    return ws


def _warnings(rows):
    res = ve.ValidationResult()
    nodes = [{"name": r[0]} for r in rows]
    ve.validate_loopbacks(_ws(rows), nodes, [], {"ns_tiers": 1}, res)
    return [w for w in res.warnings if "block" in str(w)]


def test_on_block_layout_is_silent():
    rows = [
        ("gl-plane1-01", "10.1.1.1", "10.1.1.21"),
        ("gl-plane1-02", "10.1.1.2", "10.1.1.22"),
        ("gs-plane1-01", "10.1.1.41", "10.1.1.51"),
        ("gs-plane1-02", "10.1.1.42", "10.1.1.52"),
    ]
    assert _warnings(rows) == []


def test_old_gpu_offset_is_flagged():
    """The `10 + index` layout that collided at sixteen leaves."""
    rows = [("gl-plane1-01", "10.1.1.1", "10.1.1.11")]
    warnings = _warnings(rows)
    assert len(warnings) == 1
    assert "gl-plane1-01" in str(warnings[0])
    assert "21" in str(warnings[0])


def test_spine_inside_leaf_range_is_flagged_before_it_collides():
    """The latent case: unique today, colliding at five leaves."""
    rows = [
        ("gl-plane1-01", "10.1.1.1", "10.1.1.21"),
        ("gs-plane1-01", "10.1.1.5", "10.1.1.51"),
    ]
    warnings = _warnings(rows)
    assert len(warnings) == 1
    assert "gs-plane1-01" in str(warnings[0])
    assert "41" in str(warnings[0])


def test_plane_prefix_is_not_assumed():
    """Only the final octet is checked — planes own different /24s."""
    assert _warnings([("gl-plane2-01", "10.2.1.1", "10.2.1.21")]) == []
    assert _warnings([("gl-plane2-01", "10.1.2.1", "10.1.2.21")]) == []


def test_index_beyond_block_capacity_is_flagged():
    rows = [("gs-plane1-11", "10.1.1.51", "10.1.1.61")]
    assert any("capacity" in str(w) for w in _warnings(rows))


def test_non_plane_switches_are_ignored():
    """core / cl / oob-switch rows follow the N/S map, not this one."""
    rows = [
        ("core-01", "172.16.176.11", "192.168.110.5"),
        ("oob-switch-01", "172.16.176.101", None),
    ]
    assert _warnings(rows) == []
