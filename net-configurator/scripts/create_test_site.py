#!/usr/bin/env python3
# SPDX-FileCopyrightText: Copyright (c) 2026 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: MIT
"""
Create modified Excel copies for site validation testing.

For each architecture, copies the default Excel, changes IPs/subnets/settings
to unique values, and saves with a new site name. Used to verify the parser
and generate pipeline handles non-default sites correctly.
"""

import copy
import re
import shutil
from pathlib import Path

import openpyxl

SITE_NAME = "testabc"
ARCHS = ["2-4-3-200", "2-8-5-200", "2-8-9-400"]

# IP remapping: shift the major octets to a test range
# 192.168.200.x -> 10.100.200.x, 192.168.210.x -> 10.100.210.x, etc.
# 172.16.17x.x  -> 10.200.17x.x
# 192.168.110.x -> 10.100.110.x

def remap_ip(ip_str):
    """Remap an IP address to the test range."""
    if ip_str is None:
        return None
    ip = str(ip_str).strip()
    if not ip:
        return ip_str
    # 192.168.X.Y -> 10.100.X.Y
    ip = re.sub(r'^192\.168\.', '10.100.', ip)
    # 172.16.X.Y -> 10.200.X.Y
    ip = re.sub(r'^172\.16\.', '10.200.', ip)
    return ip


def remap_subnet(subnet_str):
    """Remap a subnet CIDR to the test range."""
    if subnet_str is None:
        return None
    s = str(subnet_str).strip()
    if not s:
        return subnet_str
    s = re.sub(r'^192\.168\.', '10.100.', s)
    s = re.sub(r'^172\.16\.', '10.200.', s)
    return s


def remap_csv_subnets(csv_str):
    """Remap comma-separated subnets."""
    if csv_str is None:
        return None
    parts = str(csv_str).split(',')
    return ', '.join(remap_subnet(p.strip()) for p in parts)


def remap_csv_ips(csv_str):
    """Remap comma-separated IPs."""
    if csv_str is None:
        return None
    parts = str(csv_str).split(',')
    return ','.join(remap_ip(p.strip()) for p in parts)


def modify_settings(ws):
    """Modify Settings sheet with test values."""
    changes = {}
    for row in range(1, ws.max_row + 1):
        key = ws.cell(row=row, column=1).value
        val = ws.cell(row=row, column=2).value
        if key is None or val is None:
            continue

        key_lower = str(key).strip().lower()
        new_val = None

        if key_lower == 'site_name':
            new_val = SITE_NAME
        elif key_lower == 'mgmt_subnets':
            new_val = remap_csv_subnets(val)
        elif key_lower == 'loopback_base':
            # 172.16.176 -> 10.200.176
            new_val = str(val).replace('172.16.', '10.200.')
        elif key_lower == 'bgp_asn':
            new_val = 4260395000  # different ASN
        elif key_lower == 'timezone':
            new_val = 'America/New_York'
        elif key_lower == 'ldap_domain':
            new_val = 'test.example.com'
        elif key_lower == 'ldap_base_dn':
            new_val = 'dc=test,dc=example,dc=com'
        elif key_lower == 'ldap_root_dn':
            new_val = 'cn=admin,ou=Users,dc=test,dc=example,dc=com'
        elif key_lower == 'ldap_servers':
            new_val = remap_csv_ips(val)
        elif key_lower == 'ztp_server':
            new_val = remap_ip(val)
        elif key_lower == 'exit_dhcp_servers' and val:
            new_val = remap_csv_subnets(val)
        elif key_lower == 'mh_mac':
            new_val = '44:38:39:FF:00:BB'
        elif key_lower == 'anycast_mac':
            new_val = '44:38:39:FF:00:EE'
        elif key_lower == 'ntp_servers':
            new_val = '0.pool.ntp.org\n1.pool.ntp.org\n2.pool.ntp.org\n3.pool.ntp.org'

        if new_val is not None:
            ws.cell(row=row, column=2).value = new_val
            changes[key_lower] = f'{val} -> {new_val}'

    return changes


def modify_nodes(ws):
    """Modify Nodes sheet - remap management IPs and gateways."""
    changes = 0
    for row in range(2, ws.max_row + 1):
        func = ws.cell(row=row, column=1).value
        if func is None:
            continue

        # Column 4: Mgmt IP Address
        old_ip = ws.cell(row=row, column=4).value
        if old_ip:
            new_ip = remap_ip(old_ip)
            if new_ip != old_ip:
                ws.cell(row=row, column=4).value = new_ip
                changes += 1

        # Column 6: Gateway
        old_gw = ws.cell(row=row, column=6).value
        if old_gw:
            new_gw = remap_ip(old_gw)
            if new_gw != old_gw:
                ws.cell(row=row, column=6).value = new_gw
                changes += 1

    return changes


def modify_vlans(ws):
    """Modify VLANs & Profiles sheet - remap VLAN subnets and gateways."""
    changes = 0
    for row in range(3, 8):  # VLAN rows are 3-7
        vlan_id = ws.cell(row=row, column=1).value
        if vlan_id is None:
            continue

        # Column 4: Subnet
        old_subnet = ws.cell(row=row, column=4).value
        if old_subnet:
            new_subnet = remap_subnet(old_subnet)
            if new_subnet != old_subnet:
                ws.cell(row=row, column=4).value = new_subnet
                changes += 1

        # Column 5: Gateway
        old_gw = ws.cell(row=row, column=5).value
        if old_gw:
            new_gw = remap_ip(old_gw)
            if new_gw != old_gw:
                ws.cell(row=row, column=5).value = new_gw
                changes += 1

    return changes


def process_architecture(arch):
    """Create a modified Excel for one architecture."""
    src = Path(f'input/{arch}/default/{arch}.xlsx')
    dst = Path(f'/tmp/era-test-{arch}-{SITE_NAME}.xlsx')

    if not src.exists():
        print(f'  ERROR: {src} not found!')
        return False

    # Copy and load
    shutil.copy2(src, dst)
    wb = openpyxl.load_workbook(dst)

    print(f'\n{"="*60}')
    print(f'Architecture: {arch}')
    print(f'{"="*60}')

    # Modify Settings
    settings_changes = modify_settings(wb['Settings'])
    print(f'\nSettings changes ({len(settings_changes)}):')
    for k, v in settings_changes.items():
        print(f'  {k}: {v}')

    # Modify Nodes
    node_changes = modify_nodes(wb['Nodes'])
    print(f'\nNodes: {node_changes} IP/gateway values remapped')

    # Modify VLANs
    vlan_changes = modify_vlans(wb['VLANs & Profiles'])
    print(f'VLANs: {vlan_changes} subnet/gateway values remapped')

    # Save
    wb.save(dst)
    print(f'\nSaved: {dst}')
    return True


def main():
    print(f'Creating test site "{SITE_NAME}" Excel files for all architectures')
    print(f'IP remapping: 192.168.x.y -> 10.100.x.y, 172.16.x.y -> 10.200.x.y')

    for arch in ARCHS:
        if not process_architecture(arch):
            return 1

    print(f'\n{"="*60}')
    print('All files created. Run import with:')
    for arch in ARCHS:
        print(f'  make import EXCEL=/tmp/era-test-{arch}-{SITE_NAME}.xlsx')
    print(f'Then generate with:')
    for arch in ARCHS:
        print(f'  make generate ARCH={arch} SITE={SITE_NAME}')

    return 0


if __name__ == '__main__':
    exit(main())
