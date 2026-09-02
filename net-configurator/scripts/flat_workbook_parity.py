#!/usr/bin/env python3
# SPDX-FileCopyrightText: Copyright (c) 2026 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: MIT
"""Are the flat sample workbooks still what the per-arch sites say they are?

`input/sample-<arch>.xlsx` and `input/largescale-<arch>.xlsx` are DERIVED from
the per-arch site workbooks under `input/<arch>/`. They are not authored. A
`sample-*` file is its `default/` source with one cell changed — `site_name`,
`default` -> `sample`; a `largescale-*` file is a plain copy.

Because they are copies, they go stale silently: regenerate a per-arch workbook
and the flat one still opens, still validates, and still deploys — it just
describes an older fabric. Nothing about the file looks wrong.

This module is the single definition of "still matching", so that the two places
that enforce it cannot drift apart from each other. The flat workbooks are a
derived artifact: they are regenerated from the per-arch sources by the
maintainers and should never be hand-edited to satisfy this check.

Usage:
    python3 scripts/flat_workbook_parity.py --check input
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

import openpyxl

# (flat filename prefix, per-arch scale directory it is derived from)
FAMILIES = (("sample", "default"), ("largescale", "largescale"))

SITE_NAME_KEY = "site_name"

# What a reader of a failure should do about it. These files are not authored:
# they are regenerated from the per-arch sources by the maintainers, so the fix
# is never a hand-edit of the flat workbook.
REMEDY = ("The flat workbooks are derived, not authored — they are regenerated "
          "from the per-arch sources by the maintainers; do not edit them by hand.")


def discover_archs(input_dir: Path) -> tuple[str, ...]:
    """Arch names present as site directories, derived — never hardcoded.

    A hardcoded list is how 2-4-5-400 shipped with no flat files and stayed
    invisible to the parity harness for four days.

    The UNION of both scales, not just ``default/``. Globbing only ``default/``
    made a missing source self-concealing: the very file whose absence should be
    reported was also the thing that made the arch visible, so deleting
    ``<arch>/default/<arch>.xlsx`` silently dropped that arch — and its
    ``largescale-`` flat file, which does not derive from ``default/`` at all —
    from the check, which then passed. With the union, an arch missing either
    scale is still discovered and falls through to the "source workbook missing"
    report in :func:`check`.
    """
    return tuple(sorted({
        p.parent.parent.name
        for scale in ("default", "largescale")
        for p in input_dir.glob(f"*/{scale}/*.xlsx")
        if p.stem == p.parent.parent.name
    }))


def discover_flat_workbooks(input_dir: Path) -> tuple[tuple[str, str, Path], ...]:
    """Every flat workbook on disk as (prefix, arch, path)."""
    return tuple(sorted(
        (prefix, p.stem[len(prefix) + 1:], p)
        for prefix, _scale in FAMILIES
        for p in input_dir.glob(f"{prefix}-*.xlsx")
    ))


def orphan_problems(input_dir: Path, archs: tuple[str, ...]) -> list[str]:
    """Flat workbooks whose arch no longer exists as a site directory.

    Discovery walks arch dirs -> flat files, so a renamed or removed arch
    directory takes its flat files out of the check entirely instead of
    reporting them. This walks the other direction to close that hole.
    """
    known = set(archs)
    return [
        f"{path.name}: orphaned — no site directory {arch}/ under {input_dir}"
        for _prefix, arch, path in discover_flat_workbooks(input_dir)
        if arch not in known
    ]


def source_for(input_dir: Path, arch: str, scale: str) -> Path:
    return input_dir / arch / scale / f"{arch}.xlsx"


def flat_for(input_dir: Path, prefix: str, arch: str) -> Path:
    return input_dir / f"{prefix}-{arch}.xlsx"


def workbook_cells(path: Path) -> dict[str, tuple[tuple, ...]]:
    """Every sheet's cell values. Content, not bytes — a re-zip is not a defect."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        return {name: tuple(tuple(r) for r in wb[name].iter_rows(values_only=True))
                for name in wb.sheetnames}
    finally:
        wb.close()


def expected_cells(source: Path, site_name: str) -> dict[str, tuple[tuple, ...]]:
    """The source's cells as they should appear in the derived flat workbook."""
    cells = workbook_cells(source)
    settings = cells.get("Settings")
    if settings is None:
        return cells
    patched = []
    for row in settings:
        if row and row[0] == SITE_NAME_KEY and len(row) > 1:
            row = (row[0], site_name) + tuple(row[2:])
        patched.append(tuple(row))
    cells["Settings"] = tuple(patched)
    return cells


def _site_name_of(cells: dict[str, tuple[tuple, ...]]) -> str | None:
    for row in cells.get("Settings", ()):
        if row and row[0] == SITE_NAME_KEY and len(row) > 1:
            return row[1]
    return None


def _first_difference(expected, actual) -> str:
    for sheet in sorted(set(expected) | set(actual)):
        exp, act = expected.get(sheet), actual.get(sheet)
        if exp is None or act is None:
            return f"sheet {sheet!r} present on only one side"
        if exp == act:
            continue
        for i, (e, a) in enumerate(zip(exp, act), start=1):
            if e != a:
                return f"{sheet} row {i}: expected {e!r}, found {a!r}"
        return f"{sheet}: row count {len(exp)} vs {len(act)}"
    return "unknown difference"


def check(input_dir: Path) -> list[str]:
    """Problems found; empty list means every flat workbook is current."""
    problems: list[str] = []
    archs = discover_archs(input_dir)
    if not archs:
        # Reported, not returned: the orphan pass below still has something to
        # say when every arch directory is gone but the flat files remain.
        problems.append(f"no arch site directories found under {input_dir} — "
                        f"refusing to report clean on an empty tree")
    for arch in archs:
        for prefix, scale in FAMILIES:
            source = source_for(input_dir, arch, scale)
            flat = flat_for(input_dir, prefix, arch)
            if not source.exists():
                problems.append(f"{source.name}: source workbook missing at {source}")
                continue
            if not flat.exists():
                problems.append(
                    f"{flat.name}: missing — derived from {arch}/{scale}. "
                    f"{REMEDY}")
                continue
            want = expected_cells(source, prefix if prefix == "sample" else scale)
            got = workbook_cells(flat)
            site = _site_name_of(got)
            wanted_site = "sample" if prefix == "sample" else scale
            if site != wanted_site:
                problems.append(
                    f"{flat.name}: site_name is {site!r}, expected {wanted_site!r}")
                continue
            if want != got:
                problems.append(
                    f"{flat.name}: stale vs {arch}/{scale} — "
                    f"{_first_difference(want, got)}. {REMEDY}")
    problems.extend(orphan_problems(input_dir, archs))
    return problems


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("--check", metavar="INPUT_DIR", required=True, type=Path,
                    help="input/ directory to verify")
    args = ap.parse_args(argv)
    problems = check(args.check)
    for p in problems:
        print(f"  {p}")
    if problems:
        print(f"{len(problems)} flat workbook(s) out of date.")
        return 1
    print("✓ flat workbooks match their per-arch sources")
    return 0


if __name__ == "__main__":
    sys.exit(main())
