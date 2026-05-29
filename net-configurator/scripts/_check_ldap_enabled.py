#!/usr/bin/env python3
# SPDX-FileCopyrightText: Copyright (c) 2026 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: MIT
"""Print ``1`` if ldap_enabled=Yes in the given arch/site Excel, else nothing.

Usage:

    python3 scripts/_check_ldap_enabled.py <arch> <site>

Opens ``input/<arch>/<site>/<arch>.xlsx``, scans the Settings sheet,
and prints ``1`` to stdout if the ``ldap_enabled`` row is set to a
truthy value. Prints nothing and exits 0 in every other case —
including missing file, malformed sheet, or ``ldap_enabled`` absent —
so callers can safely use ``$(...)`` substitution without handling
errors explicitly.

Both arguments are re-validated here even though callers are expected
to pre-validate via ``_era_context.py`` — defence-in-depth against a
future refactor that forgets the upstream check.

Replaces a 200-plus-char inline Python one-liner in the Makefile
that built the Excel path by interpolating ``$$_ARCH`` and ``$$_SITE``
from ``.era-context`` directly into a Python string literal, which
could have executed arbitrary Python if the context file was ever
tampered with.
"""

import re
import sys
from pathlib import Path

VALID_ARCHS = {"2-4-3-200", "2-8-5-200", "2-8-9-400", "2-8-9-800"}
SITE_RE = re.compile(r"^[A-Za-z0-9_-][A-Za-z0-9._-]*$")


def main() -> int:
    if len(sys.argv) != 3:
        sys.stderr.write("usage: _check_ldap_enabled.py <arch> <site>\n")
        return 2

    arch, site = sys.argv[1], sys.argv[2]
    if arch not in VALID_ARCHS:
        sys.stderr.write(f"_check_ldap_enabled: invalid arch {arch!r}\n")
        return 2
    if not SITE_RE.match(site) or ".." in site:
        sys.stderr.write(f"_check_ldap_enabled: invalid site {site!r}\n")
        return 2

    path = Path("input") / arch / site / f"{arch}.xlsx"
    if not path.is_file():
        return 0  # no Excel yet (common early in the pipeline)

    try:
        import openpyxl
    except ImportError:
        return 0

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception:
        return 0

    try:
        if "Settings" not in wb.sheetnames:
            return 0
        ws = wb["Settings"]
        for row in range(1, ws.max_row + 1):
            key = str(ws.cell(row, 1).value or "").strip().lower().replace(" ", "_")
            value = str(ws.cell(row, 2).value or "").strip().lower()
            if key == "ldap_enabled" and value in ("yes", "true", "1"):
                print("1")
                return 0
    finally:
        wb.close()
    return 0


if __name__ == "__main__":
    sys.exit(main())
