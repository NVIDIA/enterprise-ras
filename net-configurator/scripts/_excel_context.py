#!/usr/bin/env python3
# SPDX-FileCopyrightText: Copyright (c) 2026 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: MIT
"""Read arch and site directly from an Excel file's Settings tab.

Prints shell-safe assignments identical to _era_context.py so the
Makefile deploy target can use the same eval pattern:

    eval "$(python3 scripts/_excel_context.py path/to/file.xlsx)"
    # sets: _ARCH and _SITE

This avoids the .era-context race condition when multiple make deploy
invocations run in parallel — each reads from its own Excel file rather
than from the shared .era-context that import overwrites.
"""

import re
import shlex
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.stderr.write("_excel_context: openpyxl not installed\n")
    sys.exit(1)

VALID_ARCHS = {"2-4-3-200", "2-4-5-400", "2-4-5-800", "2-8-5-200", "2-8-9-400", "2-8-9-800", "2-8-9-400-SP"}
SITE_RE = re.compile(r"^[A-Za-z0-9_-][A-Za-z0-9._-]*$")


def die(msg: str) -> None:
    sys.stderr.write(f"_excel_context error: {msg}\n")
    sys.exit(1)


def main() -> int:
    if len(sys.argv) != 2:
        die(f"usage: {sys.argv[0]} path/to/file.xlsx")

    path = Path(sys.argv[1])
    if not path.exists():
        die(f"file not found: {path}")

    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        die(f"cannot open {path}: {exc}")

    if "Settings" not in wb.sheetnames:
        die(f"no Settings sheet in {path}")

    ws = wb["Settings"]
    settings = {}
    for row in ws.iter_rows(values_only=True):
        if row[0] and row[1] is not None:
            settings[str(row[0]).strip()] = str(row[1]).strip()

    arch = settings.get("architecture", "")
    site = settings.get("site_name", "") or "default"

    if arch not in VALID_ARCHS:
        die(f"invalid architecture {arch!r} in {path} — must be one of: " +
            ", ".join(sorted(VALID_ARCHS)))
    if not SITE_RE.match(site):
        die(f"invalid site_name {site!r} in {path}")
    if ".." in site or site in ("", ".", ".."):
        die(f"invalid site_name {site!r} in {path}")

    print(f"_ARCH={shlex.quote(arch)}")
    print(f"_SITE={shlex.quote(site)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
