#!/usr/bin/env python3
# SPDX-FileCopyrightText: Copyright (c) 2026 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: MIT
"""
ERA Excel Parser — generates Ansible inventory from wiremap Excel templates.

Parses the Nodes and Wire Map sheets from each architecture's Excel file and
produces host_vars, group_vars, and hosts files under output/<arch>/<site>/inventory/.

Usage:
    # Generate inventory for a single architecture
    python3 scripts/excel_parser.py --arch 2-8-5-200

    # Generate inventory for a specific site
    python3 scripts/excel_parser.py --arch 2-8-5-200 --site customer-a

    # Discover and generate all architectures
    python3 scripts/excel_parser.py
"""

import ipaddress
import openpyxl
import hashlib
import re
import sys
import yaml
import zipfile
from collections import Counter, defaultdict
from pathlib import Path

from utils import (
    generate_mac,
    reset_mac_registry,
    classify_node as _classify_node,
    is_switch,
    build_interface_map,
    build_nic_destinations,
    plane_for_switch,
    loopbacks_sheet_name,
)
from oob_reserved import AIR_MGMT_RESERVED_OCTETS, OOB_RESERVED_OCTETS
import asn_allocation as asn_alloc
import loopback_allocation as lb_alloc
from inventory_defaults import (
    SERVERS_GROUP_VARS, SWITCHES_GROUP_VARS, arch_group_vars, air_vnode_host_vars)

# RFC1123-ish: node name used as a host_vars filename + a bare INI hosts token.
_INV_HOSTNAME_RE = re.compile(r'[A-Za-z0-9][A-Za-z0-9._-]*')


def assert_valid_inv_hostname(name):
    """Reject node names that aren't safe as a filesystem path component / INI
    inventory token (security review #10). This is a hard gate independent of
    validate_excel, which `--skip-validate` bypasses on the documented
    switch-ztp-deploy/import paths. Returns the name on success.
    """
    if not _INV_HOSTNAME_RE.fullmatch(name or ''):
        raise SystemExit(
            f"❌ Invalid node name {name!r} (Nodes sheet 'Name'): only letters, "
            f"digits, dot, hyphen, and underscore are allowed, starting with a "
            f"letter or digit. Path separators, newlines, and shell/INI "
            f"metacharacters are rejected — this value becomes a file path and an "
            f"Ansible inventory entry. Fix the Name cell."
        )
    return name


def load_workbook_safe(path, **kwargs):
    """openpyxl.load_workbook wrapper that turns a corrupt/non-xlsx file into a
    friendly SystemExit instead of an uncaught traceback (issue #7)."""
    try:
        return openpyxl.load_workbook(path, **kwargs)
    except (openpyxl.utils.exceptions.InvalidFileException,
            zipfile.BadZipFile, KeyError, OSError) as exc:
        raise SystemExit(
            f"❌ Could not open Excel file: {path}\n"
            f"   The file appears to be corrupt, not a valid .xlsx workbook, "
            f"or unreadable.\n"
            f"   ({type(exc).__name__}: {exc})\n"
            f"   → Re-export it from the ERA template and try again."
        )


# Default disabled interfaces (fallback if not in Settings)
DEFAULT_DISABLED_INTERFACES = {
    '2-4-3-200': [60, 62, 64],
    '2-8-5-200': [50, 52, 60, 62, 64],
    '2-8-9-400': [54, 56, 58, 60, 62, 64],
}

# Loopback network base
LOOPBACK_BASE = '172.16.176'

# Loopback allocation (N/S per-VRF blocks + E/W plane blocks) moved to
# loopback_allocation.py — the single source shared with the arch generator,
# which writes these same values into the Loopbacks tab. Keeping one copy is
# what stops the generator and the parser drifting apart; they had, and every
# regenerated workbook disagreed with the shipped one. Re-exported here so the
# in-module references and existing importers keep working.
VRF_LOOPBACK_BLOCK_SIZE = lb_alloc.VRF_LOOPBACK_BLOCK_SIZE
VRF_LOOPBACK_BLOCKS = lb_alloc.VRF_LOOPBACK_BLOCKS
PLANE_LOOPBACK_BLOCKS = lb_alloc.PLANE_LOOPBACK_BLOCKS
plane_loopback_octet = lb_alloc.plane_loopback_octet
vrf_loopback_octet = lb_alloc.vrf_loopback_octet

# Per-tier BGP ASN allocation moved to asn_allocation.py — the single
# source shared with the generator + validator. Per-node ASNs are now explicit on
# the Excel Loopbacks tab (`ASN` column); the helpers below are the fallback used
# when a workbook predates that column. Offset constants re-exported for any
# in-module reference. See asn_allocation.py for the block layout.
DEFAULT_BASE_ASN = asn_alloc.DEFAULT_BASE_ASN
OOB_ASN_OFFSET = asn_alloc.OOB_ASN_OFFSET
CSL_LEAF_ASN_OFFSET = asn_alloc.CSL_LEAF_ASN_OFFSET
CSL_SPINE_ASN_OFFSET = asn_alloc.CSL_SPINE_ASN_OFFSET
GSL_PLANE_ASN_STRIDE = asn_alloc.GSL_PLANE_ASN_STRIDE
GSL_SPINE_ASN_OFFSET = asn_alloc.GSL_SPINE_ASN_OFFSET
GSL_LEAF_ASN_OFFSET = asn_alloc.GSL_LEAF_ASN_OFFSET

# Role-based host octet ranges for management IP assignment.
# Ranges are spaced to avoid overlap even with large compute counts (up to ~40 nodes).
#   compute: .11+   support: .51+   storage: .61+   k8s: .71+   bcme: .81+
ROLE_HOST_BASE = {
    'compute': 11, 'support': 51, 'storage': 61,
    'k8s': 71, 'bcme': 81, 'unknown': 91,
}


def ports_to_range_string(port_nums):
    """Convert a set of port numbers to compact NVUE swp range notation.

    e.g. {1,2,3,5,7,8} -> 'swp1-3,swp5,swp7-8'
    """
    if not port_nums:
        return ''
    nums = sorted(port_nums)
    ranges = []
    start = end = nums[0]
    for n in nums[1:]:
        if n == end + 1:
            end = n
        else:
            ranges.append((start, end))
            start = end = n
    ranges.append((start, end))
    parts = []
    for s, e in ranges:
        parts.append(f'swp{s}-{e}' if s != e else f'swp{s}')
    return ','.join(parts)


# generate_mac() imported from utils.py — shared with topology_generator.py


# Canonical role vocabulary (see docs/ROLES.md). When the Excel Nodes-tab
# Function (or Wire Map System Role) cell holds one of these strings, the
# parser uses it directly instead of pattern-matching on the hostname.
# Backward compat: legacy Excels with hostname-as-role keep working through
# the classify_node() fallback below.
#
# GSL is split per plane (gsl-plane1, gsl-plane2) because plane membership is
# part of the routing-design identity — each plane is its own L3 fabric. Bare
# `gsl` is still accepted for backward-compat with earlier Excels but the
# validator emits a warning when it's used in a dual-plane arch.
CANONICAL_ROLES = frozenset({
    'gpu', 'support', 'storage', 'core', 'csl',
    'gsl', 'gsl-plane1', 'gsl-plane2',
    # Leaf/spine taxonomy: cl=compute leaf, cs=compute spine,
    # gl=GPU leaf, gs=GPU spine. The branch-only legacy spine names
    # csl-spine / gsl-spine-plane* have been purged — no arch references
    # them anymore.
    'cl', 'cs', 'gl-plane1', 'gl-plane2', 'gs-plane1', 'gs-plane2',
    'oob-switch', 'oob-server', 'dhcp', 'edge', 'air-oob',
    'ext-storage',
    # L3 OOB Ubuntu trio — singleton Air-only nodes. Each is its own
    # category-of-one (no -01/-02 instance index), so Function == Name.
    # Auto-injected by topology_generator; documentary rows on the
    # Nodes tab carry Enabled=Air.
    'external-conn', 'external-dhcp', 'utility',
})

# Map canonical role string -> internal classification used by the rest of
# the parser. Some canonical strings collapse into 'switch'/'infra' so the
# downstream "skip switches/infra from devices" filter still works.
_CANONICAL_TO_INTERNAL = {
    'gpu':         'compute',
    'support':     'support',
    'storage':     'storage',
    'core':        'switch',
    'csl':         'csl',
    'gsl':         'gsl',
    'gsl-plane1':  'gsl',
    'gsl-plane2':  'gsl',
    # Leaf/spine taxonomy — same internal classification as the
    # converged names they correspond to.
    'cl':          'csl',
    'cs':          'csl-spine',
    'gl-plane1':   'gsl',
    'gl-plane2':   'gsl',
    'gs-plane1':   'gsl-spine',
    'gs-plane2':   'gsl-spine',
    'oob-switch':  'switch',
    'oob-server':  'infra',
    'dhcp':        'infra',
    'edge':        'switch',
    'air-oob':     'air-oob',
    # ext-storage = customer-side simulated storage aggregate (Ubuntu + FRR
    # running BGP unnumbered eBGP back to CSL STORAGE VRF). Air-only node;
    # not provisioned with cluster IPs, eth0 OOB, or netplan beyond what
    # air-deploy.py injects as Node Instructions.
    'ext-storage': 'infra',
    # L3 OOB Ubuntu trio — singleton Air-only nodes auto-injected by
    # topology_generator. Documentary rows on the Nodes tab carry
    # Enabled=Air; parser/validator never provision from them.
    'external-conn':  'infra',
    'external-dhcp':  'infra',
    'utility':        'infra',
}


# Map canonical role -> hostname-pattern category used by legacy
# `role.startswith(...)` checks. Used to translate "what category is this
# node?" questions into the pre-canonical hostname-prefix world.
_CANONICAL_TO_PATTERN = {
    'gpu':         'compute',     # name starts with 'gpu-' or 'su-NN-node-NN'
    'support':     'support',
    'storage':     'storage',
    'core':        'core',
    'csl':         'csl',
    'gsl':         'gsl',
    'gsl-plane1':  'gsl-plane1',
    'gsl-plane2':  'gsl-plane2',
    # Leaf/spine taxonomy — each maps to its own pattern string.
    'cl':          'cl',
    'cs':          'cs',
    'gl-plane1':   'gl-plane1',
    'gl-plane2':   'gl-plane2',
    'gs-plane1':   'gs-plane1',
    'gs-plane2':   'gs-plane2',
    'oob-switch':  'oob-switch',
    'oob-server':  'oob-server',
    'dhcp':        'dhcp',
    'edge':        'edge',
    'air-oob':     'air-oob',
    'ext-storage': 'ext-storage',
    'external-conn': 'external-conn',
    'external-dhcp': 'external-dhcp',
    'utility':       'utility',
}


def canonical_category(function_value, name=None):
    """Resolve a Function/System Role cell value to a canonical category.

    Excel-first: if the cell holds a canonical role string, return it
    directly. Otherwise fall back to hostname-prefix classification on
    the optional `name` arg (or on the cell value itself when it looks
    like a hostname).

    Returns the canonical role string (e.g. 'core', 'gsl-plane1') or
    None if no classification matched. The bare 'gsl' canonical is
    promoted to 'gsl-plane1' / 'gsl-plane2' when `name` carries the
    plane suffix; otherwise it stays bare.
    """
    # An explicit Function value that is already a canonical role wins as-is.
    # New roles (cl / cs / gl-plane* / gs-plane*) and retained legacy roles
    # (csl / gsl-plane*) each resolve to
    # themselves. Adopting the new taxonomy for an arch is done by declaring
    # the new Function value — NOT by silently rewriting the old one — so
    # existing archs stay byte-identical until they opt in.
    v = (function_value or '').strip().lower()
    if v in CANONICAL_ROLES:
        # Promote bare 'gsl' if the hostname tells us which plane it is.
        if v == 'gsl' and name:
            n = name.strip().lower()
            if 'plane1' in n:
                return 'gsl-plane1'
            if 'plane2' in n:
                return 'gsl-plane2'
        return v
    # Legacy fallback: classify by hostname pattern. Try `name` first
    # (more specific), then `function_value` (which often holds the
    # hostname on legacy Excels). The hostname fallback returns the retained
    # converged role names so existing archs (blank Function) stay
    # byte-identical; spine hostnames now resolve to the new cs / gs roles.
    target = (name or function_value or '').strip().lower()
    if target.startswith('gsl-spine-plane1') or target.startswith('gs-plane1'):
        return 'gs-plane1'
    if target.startswith('gsl-spine-plane2') or target.startswith('gs-plane2'):
        return 'gs-plane2'
    if target.startswith('gl-plane1'):
        return 'gl-plane1'
    if target.startswith('gl-plane2'):
        return 'gl-plane2'
    if target.startswith('gsl-plane1'):
        return 'gsl-plane1'
    if target.startswith('gsl-plane2'):
        return 'gsl-plane2'
    if target.startswith('gsl-') or target.startswith('gl-') or target.startswith('gs-'):
        return 'gsl'
    if target.startswith('core-'):
        return 'core'
    if target.startswith('csl-spine-') or target.startswith('cs-'):
        return 'cs'
    if target.startswith('csl-') or target.startswith('cl-'):
        return 'csl'
    if target.startswith('oob-switch'):
        return 'oob-switch'
    if target.startswith('oob-server'):
        return 'oob-server'
    if target.startswith('dhcp'):
        return 'dhcp'
    if target.startswith('air-oob') or target == 'air-oob-switch':
        return 'air-oob'
    # 'edge' uses substring match to catch OEM-specific patterns like
    # 'cust-net-edge-01' that legacy `classify_node` already recognized.
    # Must come AFTER air-oob/dhcp/oob-server prefix checks above.
    if 'edge' in target and 'dhcp' not in target and 'oob-server' not in target:
        return 'edge'
    if target.startswith('gpu-') or (target.startswith('su-') and 'node' in target):
        return 'gpu'
    if target.startswith('support-') or target.startswith('bcm-') or target.startswith('bcme-') \
            or target.startswith('k8s-') or target.startswith('slurm-'):
        return 'support'
    if target.startswith('storage-'):
        return 'storage'
    return None


def extract_role_index(name):
    """Return the trailing-digit index from a hostname, e.g. 'core-01' -> 1,
    'dog10' -> 10. Returns None if no digits trail.

    Callers that need a per-node order-among-its-category fallback for
    name-less / digitless hostnames must supply that themselves
    (typically via enumerate() over a category-filtered list).
    """
    import re as _re
    m = _re.search(r'(\d+)$', (name or '').strip())
    return int(m.group(1)) if m else None


# --------------------------------------------------------------------------
# Wire Map / Air_Only header-name-based column lookup
# --------------------------------------------------------------------------
#
# Previously the parser read Wire Map and Air_Only cells by hardcoded
# column index (`row[10]`, `row[12]`, etc.). That meant adding or
# reordering any column in the spreadsheet silently shifted every
# downstream read.
#
# Now: read the header row once, build a `{logical_name: col_idx}` map
# with alias support, and look up cells by logical name. Operators can
# reorder columns, insert annotation columns, or rename headers
# (e.g. "System Role" → "Function") without touching code.
#
# Required vs optional columns per the parser's actual consumption:
#
#   REQUIRED                       The parser fails to run without these.
#   - system_name                  Host identity (Nodes lookup, MAC, dict keys)
#   - nic_port                     Per-host interface
#   - network_profile              CPU/GPU/OOB/Storage classification
#   - switch_name                  Peer identity (dict keys, topology links)
#   - switch_port                  Peer port for topology links
#
#   OPTIONAL (parser falls back)
#   - display_in_air               Defaults to False if blank
#   - system_role / function       Falls back to Nodes-tab lookup (MR !28)
#   - switch_role / switch_function   Same
#
#   OPERATOR DOCUMENTATION (parser ignores; kept for cabling crew)
#   - port_side, cable_split
#
# Removed 2026-05-19 (port settings are now derived exclusively from
# Port Profiles + VLANs sheet — single source of truth):
#   - speed, mode, native_vlan, allowed_vlans

_WM_REQUIRED = frozenset({'system_name', 'nic_port', 'network_profile',
                          'switch_name', 'switch_port'})

# Accepted header text per logical column (case-insensitive, whitespace
# normalized). First listed name is the canonical / preferred name for
# new templates; the rest are aliases accepted for backward-compat.
#
# Naming convention: each row of the Wire Map describes a cable between
# two ports. The "A side" / "B side" framing keeps the headers honest
# (the B side isn't always a switch — outbound, customer-edge router,
# etc.). Older Excels used `System Role` / `Switch Role` etc. and those
# names are still accepted via the alias tuples below.
_WM_HEADER_ALIASES = {
    'display_in_air':  ('display in air', 'display', 'air'),
    'system_role':     ('function (a)', 'function(a)', 'a-side function',
                        'a side function',
                        'function', 'system role', 'role'),
    'system_name':     ('system name (a)', 'system name(a)', 'a-side system name',
                        'a side system name', 'a-side name', 'a side name',
                        'system name', 'hostname', 'name', 'host'),
    'nic_port':        ('port (a)', 'port(a)', 'a-side port', 'a side port',
                        'nic/port', 'nic/port/breakout', 'nic port', 'nic',
                        'port', 'host port'),
    'port_side':       ('port side (a)', 'port side', 'side'),
    'network_profile': ('network profile', 'profile', 'profile/vlan'),
    'switch_role':     ('function (b)', 'function(b)', 'b-side function',
                        'b side function',
                        'switch function', 'switch role', 'switch type'),
    'switch_name':     ('system name (b)', 'system name(b)', 'b-side system name',
                        'b side system name', 'b-side name', 'b side name',
                        'switch name', 'switch hostname', 'peer'),
    'switch_port':     ('port (b)', 'port(b)', 'b-side port', 'b side port',
                        'switch port', 'peer port', 'switch interface'),
}

# Air_Only sheet column aliases. Compact layout — same logical set as
# Wire Map but fewer columns. Accepts (A)/(B) headers too.
_AIR_HEADER_ALIASES = {
    'display_in_air':  ('display in air', 'display', 'air'),
    'system_role':     ('function (a)', 'function(a)', 'a-side function',
                        'function', 'system role', 'role'),
    'system_name':     ('system name (a)', 'system name(a)', 'a-side system name',
                        'a-side name', 'system name', 'hostname', 'name'),
    'nic_port':        ('port (a)', 'port(a)', 'a-side port',
                        'nic/port', 'nic/port/breakout', 'nic'),
    'network_profile': ('network profile', 'profile'),
    'switch_role':     ('function (b)', 'function(b)', 'b-side function',
                        'switch function', 'switch role'),
    'switch_name':     ('system name (b)', 'system name(b)', 'b-side system name',
                        'b-side name', 'switch name', 'switch hostname'),
    'switch_port':     ('port (b)', 'port(b)', 'b-side port',
                        'switch port', 'peer port'),
}
_AIR_REQUIRED = _WM_REQUIRED


def _normalize_header(s):
    """Lowercase + collapse whitespace, for alias matching."""
    return ' '.join(str(s or '').strip().lower().split())


def build_wiremap_column_map(ws, sheet_kind='wiremap'):
    """Read row 1 of a Wire Map or Air_Only worksheet and return
    {logical_name: col_idx_1based}.

    sheet_kind: 'wiremap' or 'air_only' — selects the alias table.

    Raises ValueError if any REQUIRED logical column is missing — caught
    upstream and surfaced to the operator with a clear "this header is
    missing" message.
    """
    aliases = _AIR_HEADER_ALIASES if sheet_kind == 'air_only' else _WM_HEADER_ALIASES
    required = _AIR_REQUIRED if sheet_kind == 'air_only' else _WM_REQUIRED

    col_map = {}
    for c in range(1, ws.max_column + 1):
        header = _normalize_header(ws.cell(row=1, column=c).value)
        if not header:
            continue
        for logical, alias_tuple in aliases.items():
            if header in alias_tuple:
                col_map.setdefault(logical, c)
                break

    missing = required - col_map.keys()
    if missing:
        raise ValueError(
            f"Sheet is missing required column(s): {sorted(missing)}. "
            f"Required headers (any alias works): "
            + ", ".join(
                f"{k}={list(aliases[k])}" for k in sorted(missing)
            )
        )
    return col_map


def _wm_cell(row, col_map, key):
    """Fetch a value from a Wire Map row tuple by logical column name.

    `row` is the tuple from ws.iter_rows(values_only=True).
    Returns the stripped string, or '' if the column isn't mapped
    (optional column absent) or the cell is blank.
    """
    idx = col_map.get(key)
    if idx is None:
        return ''
    if idx - 1 >= len(row):
        return ''
    v = row[idx - 1]
    return str(v).strip() if v is not None else ''


def _wm_cell_ws(ws, row_idx, col_map, key):
    """Same as _wm_cell but reads directly via ws.cell() (1-based row).

    Used by parsers that iterate cell-by-cell rather than via iter_rows.
    """
    idx = col_map.get(key)
    if idx is None:
        return ''
    v = ws.cell(row=row_idx, column=idx).value
    return str(v).strip() if v is not None else ''


def build_nodes_function_map(nodes):
    """Return {hostname: canonical_function} from the parsed Nodes tab.

    Used by Wire Map / Air_Only row parsing to resolve a row's Function
    when the row's own col 2 (System Role) cell is blank or stale — the
    Nodes tab is the single source of truth for "what is this host."

    Cascade in callers should be:
        nodes_map.get(system_name)            # Nodes-tab lookup
        or row's own system_role value         # legacy / explicit override
        or canonical_category(name)            # hostname-prefix fallback

    Values are the canonical role strings already computed by
    parse_nodes() (node['category']). Hostnames whose Function couldn't
    be classified are omitted.
    """
    out = {}
    for n in nodes or []:
        name = (n.get('name') or '').strip()
        cat = n.get('category')
        if name and cat:
            out[name] = cat
    return out


def classify_host_role(name: str) -> str:
    """Determine a host's data-plane role from its Function / Name.

    Returns one of: 'compute', 'storage', 'support', 'k8s', 'bcme',
    'switch', 'csl', 'gsl', 'air-oob', 'infra', 'unknown'.

    Excel-first behavior (per docs/ROLES.md): if the input matches a
    canonical role string (case-insensitive), map directly to the internal
    classification. Otherwise fall back to legacy name-pattern matching
    via classify_node().
    """
    s = (name or '').strip().lower()
    if s in CANONICAL_ROLES:
        return _CANONICAL_TO_INTERNAL[s]
    role = _classify_node(name)
    if role in ('core', 'oob', 'edge'):
        return 'switch'
    return role


def _build_wiremap_row_list(ws_wiremap, ws_air_only=None, nodes_function_map=None,
                            disabled_names=None):
    """Read Wire Map (and optionally Air_Only) into a list of dicts.

    Returns rows in the same order the topology generator processes them:
    Air_Only rows first, then Wire Map rows.  Each row is a dict with keys
    matching what build_interface_map() expects.

    Column mapping (0-based from iter_rows):
      Wire Map:  [0]=Display in Air, [1]=System Role, [2]=System Name,
                 [3]=NIC/Port, [6]=Network Profile, [10]=Switch Role,
                 [11]=Switch Name, [12]=Switch Port
      Air_Only:  [0]=Display in Air, [1]=System Role, [2]=System Name,
                 [3]=NIC/Port, [4]=Network Profile, [5]=Switch Role,
                 [6]=Switch Name, [7]=Switch Port

    `disabled_names`: optional set of node hostnames marked Enabled=No on
    the Nodes tab. Rows where the A-side OR B-side matches a disabled
    name are excluded — otherwise the switch's NVUE config emits port
    config for peers that won't come up. (R3-6.)
    """
    nodes_map = nodes_function_map or {}
    disabled_names = disabled_names or set()

    def _resolve(system_role, system_name):
        """Cascading resolution: Nodes-map → row's own value → unchanged."""
        looked_up = nodes_map.get(system_name) if system_name else None
        return looked_up or system_role

    def _parse_sheet(ws, sheet_kind):
        try:
            col_map = build_wiremap_column_map(ws, sheet_kind=sheet_kind)
        except ValueError:
            # Air_Only is sometimes used purely for the version-image map and
            # Air-mgmt-subnet metadata, with no wire-map-style columns. Treat
            # that as "no connection rows" rather than failing the parser.
            if sheet_kind == 'air_only':
                return []
            raise
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            if i == 1:
                continue  # skip header

            display_raw = _wm_cell(row, col_map, 'display_in_air').lower()
            system_role_raw = _wm_cell(row, col_map, 'system_role')
            system_name = _wm_cell(row, col_map, 'system_name') or system_role_raw
            nic_port = _wm_cell(row, col_map, 'nic_port')
            net_profile = _wm_cell(row, col_map, 'network_profile')
            switch_role_raw = _wm_cell(row, col_map, 'switch_role')
            switch_name = _wm_cell(row, col_map, 'switch_name') or switch_role_raw
            switch_port = _wm_cell(row, col_map, 'switch_port')

            # Skip empty/spacer rows: both Role AND Name cells blank.
            if not system_role_raw and not system_name:
                continue

            # R3-6: skip rows whose A-side or B-side is a disabled node.
            # Otherwise switch port/bond config gets emitted for peers
            # that won't come up.
            if (system_name and system_name in disabled_names) or \
               (switch_name and switch_name in disabled_names):
                continue

            # Resolve roles via Nodes-tab lookup (cascading fallback).
            system_role = _resolve(system_role_raw, system_name)
            switch_role = _resolve(switch_role_raw, switch_name)

            rows.append({
                'display_in_air': display_raw == 'yes',
                'system_role': system_role,
                'system_role_raw': system_role_raw,
                'system_name': system_name,
                'nic_port': nic_port,
                'net_profile': net_profile,
                'switch_role': switch_role,
                'switch_role_raw': switch_role_raw,
                'switch_name': switch_name,
                'switch_port': switch_port,
            })
        return rows

    result = []
    if ws_air_only is not None:
        result.extend(_parse_sheet(ws_air_only, 'air_only'))
    result.extend(_parse_sheet(ws_wiremap, 'wiremap'))
    return result


def _svi_reserved_offsets(nodes):
    """Addresses the FABRIC itself claims at the bottom of every data VLAN.

    `network + 1` is the VRR/anycast gateway, and each SVI-bearing switch then
    takes one more (`_svi_switch_ip()` walks the host range with the gateway
    excluded: core-01 -> network+2, core-02 -> network+3, ...). Servers must
    start above ALL of them.

    Reserving only the gateway — as the original ERA-42 packing did — handed
    the first support/storage node the same address as core-01's SVI on any
    non-/24 VLAN. Observed live on a Support VLAN 10.78.220.32/27: k8s-01 got
    .34/.35, identical to the two cores' SVIs, so the gateway .33 never ARPed
    and the whole support data plane was dead while /24 archs were unaffected.
    """
    n = sum(1 for x in (nodes or [])
            if str(x.get('status', '')).strip().lower() == 'active'
            and x.get('category') in ('core', 'csl', 'cl'))
    return 2 + (n if n else 2)


def _dataplane_host_ips(subnet_str, idx, count, legacy_start, stride, reserved=2,
                        total=None):
    """Allocate `count` consecutive data-plane host IPs for node `idx` inside the
    declared VLAN subnet (ERA-42).

    The legacy code hardcoded `{first-three-octets}.{offset}/24`, which put hosts
    OUTSIDE the switch SVI subnet on any non-/24 or non-.0-network subnet
    (observed: Support VLAN 100.82.255.128/27 → hosts assigned .101/24, unreachable).

    - **/24 subnet:** preserve the legacy octet offset (`legacy_start + stride*idx`)
      so existing /24 output stays byte-identical — but repack from `reserved`
      when `total` is given and the legacy layout would not fit (see below).
    - **any other prefix:** pack hosts compactly from `network + reserved`,
      where `reserved` covers the anycast gateway AND one SVI per SVI-bearing
      switch (see _svi_reserved_offsets), and honor the subnet's real prefix +
      network offset.

    ``total`` is the node count the caller is allocating for. A /24 has 254
    usable addresses, but `legacy_start=201` leaves only 54 above it, so any
    site with more than 54 compute nodes silently got no address at all past
    the 54th — 2-8-9-800 and 2-8-9-400-SP largescale addressed 54 of 128,
    2-4-5-800 54 of 144, 2-8-9-400 54 of 64. The identical 54 across four
    different architectures is the tell: it is this constant, not four
    independently undersized subnets. The subnets are fine — only `.0` and `.1`
    are in use below `.201`, leaving 197 free addresses the allocator never
    reached for.

    This mirrors the fix already applied to switch eth0 air-mgmt allocation
    (see the `_AIR_MGMT_RESERVED` comment in generate_group_vars), which had
    the same `.201`-base overflow: keep the legacy base when it fits so small
    and default sites stay byte-identical, and repack from the bottom only
    when it does not. Callers that pass no ``total`` keep the old behaviour.

    Returns a list of ``'ip/prefix'`` strings, or None if they don't fit the
    subnet (caller then assigns no IP, same as the old over-.254 guard).
    """
    if not subnet_str:
        return None
    try:
        net = ipaddress.ip_network(subnet_str, strict=False)
    except ValueError:
        return None
    if net.prefixlen == 24:
        # Repack from the bottom only when the legacy base cannot seat everyone;
        # the decision depends solely on `total`, so it is identical for every
        # idx in the run and never produces a half-legacy, half-repacked subnet.
        legacy_fits = True
        if total:
            last = legacy_start + stride * (total - 1) + (count - 1)
            legacy_fits = last < int(net.num_addresses) - 1
        start = ((legacy_start + stride * idx) if legacy_fits
                 else (reserved + count * idx))
    else:
        start = reserved + count * idx
    ips = []
    for k in range(count):
        off = start + k
        if off >= int(net.num_addresses) - 1:   # would hit/exceed broadcast
            return None
        ips.append(f"{net.network_address + off}/{net.prefixlen}")
    return ips


# Role -> data-plane address pool. Several roles share one pool, so anything
# deciding "how many nodes draw from this subnet" has to group them the way the
# allocation loop does. Kept as one table so the two cannot drift;
# validate_excel.py encodes the same grouping in its `checks` list.
_ROLE_ALLOC_POOL = {
    'compute': 'compute',
    'storage': 'storage',
    'support': 'support',
    'k8s': 'support',
    'bcme': 'support',
}


def build_devices(nodes, vlans, node_oob_mapping=None, wiremap_rows=None, gpu_vlan_mode='single'):
    """Build the devices dict for dnsmasq DHCP reservations and server netplan config.

    Generates:
      - eth0_ip + mac for ALL non-switch hosts (for DHCP reservations)
      - Data-plane IPs for known roles:
        - compute: bond_ip (CPU subnet), gpu_ip1/gpu_ip2 (GPU subnet)
        - storage: bond_ip1/bond_ip2 (storage subnet)
        - support: bond_ip1/bond_ip2 (support subnet)

    eth0_ip always comes from the Nodes-tab Mgmt IP cell (the spreadsheet is the
    source of truth for the deployed management IP; see the build loop below
    where ``eth0_ip = node.get('mgmt_ip', '')``). The node_oob_mapping argument
    (from Wire Map 'Air - Management' rows) is accepted for backward
    compatibility but is NOT used here to derive eth0_ip — an earlier design
    derived the IP from the OOB switch's subnet; the live code does not.

    MACs are auto-generated deterministically (matching topology_generator.py)
    unless already present in the Excel Nodes tab.

    Args:
        nodes: list of node dicts from parse_nodes() [{name, mac, mgmt_ip, ...}]
        vlans: list of vlan dicts from parse_vlans() [{id, name, subnet, ...}]
        node_oob_mapping: dict {node_name: oob_switch_name} from Wire Map
        wiremap_rows: list of dicts from _build_wiremap_row_list() for interface mapping
        gpu_vlan_mode: 'single' (default), 'per_rail', or 'per_rail_per_plane'.
            - single: one GPU VLAN; legacy behavior.
            - per_rail: one VLAN per rail (`gpu_rail<R>` rows, `GPU Rail R`
              Wire Map profiles). See docs/plans/2026-05-17-vlan-per-gpu.md.
            - per_rail_per_plane: one VLAN per (rail, plane) combination
              (`gpu_rail<R>_plane<P>` rows, `GPU Rail R Plane P` Wire Map
              profiles). See docs/plans/2026-05-18-gpu-plane-per-rail.md.
    """
    devices = {}
    node_oob_mapping = node_oob_mapping or {}

    # Build subnet lookup by VLAN name (lowercase), with normalized aliases
    subnet_map = {}
    # Per-plane GPU subnets (dual-plane mode): collected from gpu_plane<N> rows
    # so the compute IP allocator can pick the right subnet per NIC.
    gpu_planes = {}
    # Per-rail GPU subnets (gpu_vlan_mode=per_rail): collected from gpu_rail<N>
    # rows. One VLAN per GPU NIC. Keyed by rail index ('rail1', 'rail2', ...).
    gpu_rails = {}
    # Per-rail-per-plane GPU subnets (gpu_vlan_mode=per_rail_per_plane):
    # collected from gpu_rail<R>_plane<P> rows. Keyed by (rail_key, plane_key)
    # tuple e.g. ('rail1', 'plane1'). One VLAN per (rail, plane) combination.
    gpu_rail_planes = {}
    for vlan in vlans:
        if vlan.get('name') and vlan.get('subnet'):
            key = vlan['name'].lower()
            subnet_map[key] = vlan['subnet']
            # Add short aliases for multi-word VLAN names
            if 'cpu' in key or 'in-band' in key or 'inband' in key:
                subnet_map['cpu'] = vlan['subnet']
            if 'gpu' in key:
                subnet_map['gpu'] = vlan['subnet']
            if 'storage' in key:
                subnet_map['storage'] = vlan['subnet']
            if 'support' in key:
                subnet_map['support'] = vlan['subnet']

            # Multi-plane capture: 'gpu_plane1' -> plane='plane1'
            m = re.match(r'^gpu_plane(\d+)$', key)
            if m:
                plane = f"plane{m.group(1)}"
                base = vlan['subnet'].split('/')[0].rsplit('.', 1)[0]
                gpu_planes[plane] = {
                    'subnet': vlan['subnet'],
                    # Prefer the Excel Gateway column; fall back to base.1 only
                    # if the cell is empty.
                    'gateway': vlan.get('gateway') or f"{base}.1",
                    'vlan_id': vlan.get('id'),
                    # Per-plane PBR table BASE. Each NIC in the plane gets a
                    # unique table = base + nic_idx_in_plane (see the
                    # gpu_interfaces emitter further down). Plane offset of
                    # 100 leaves headroom for up to ~99 NICs per plane per
                    # host before colliding with the next plane — well above
                    # the 16-NIC headroom the IP allocator enforces.
                    # Without per-NIC tables, all NICs in a plane funnel
                    # into one table with N ECMP default routes; source-IP
                    # policy routing can't pin per-NIC egress, and the
                    # resulting ARP/MAC ambiguity caused ~85% loss on the
                    # GPU VRR ping in dual-plane Air sims.
                    'table': (vlan.get('id') or 0) + (int(m.group(1)) - 1) * 100,
                }

            # Per-rail capture: 'gpu_rail1' -> rail='rail1'
            m_rail = re.match(r'^gpu_rail(\d+)$', key)
            if m_rail:
                rail = f"rail{m_rail.group(1)}"
                base = vlan['subnet'].split('/')[0].rsplit('.', 1)[0]
                gpu_rails[rail] = {
                    'subnet': vlan['subnet'],
                    'gateway': vlan.get('gateway') or f"{base}.1",
                    'vlan_id': vlan.get('id'),
                    # PBR table = VLAN ID (each rail has its own VLAN ID, so
                    # tables are inherently unique — no offset needed).
                    'table': vlan.get('id') or 0,
                    'rail_index': int(m_rail.group(1)),
                }

            # Per-rail-per-plane capture: 'gpu_rail1_plane2' -> ('rail1','plane2')
            m_rp = re.match(r'^gpu_rail(\d+)_plane(\d+)$', key)
            if m_rp:
                rail = f"rail{m_rp.group(1)}"
                plane = f"plane{m_rp.group(2)}"
                base = vlan['subnet'].split('/')[0].rsplit('.', 1)[0]
                gpu_rail_planes[(rail, plane)] = {
                    'subnet': vlan['subnet'],
                    'gateway': vlan.get('gateway') or f"{base}.1",
                    'vlan_id': vlan.get('id'),
                    # PBR table: vlan_id + plane_offset keeps tables unique
                    # even when operators reuse VLAN IDs across planes (like
                    # the dual-plane VLAN 900 convention extended).
                    'table': (vlan.get('id') or 0) + (int(m_rp.group(2)) - 1) * 100,
                    'rail_index': int(m_rp.group(1)),
                    'plane_index': int(m_rp.group(2)),
                }

    # Extract base IPs from subnets (e.g., '172.16.178.0/24' → '172.16.178')
    def subnet_base(subnet_str):
        return subnet_str.split('/')[0].rsplit('.', 1)[0] if subnet_str else None

    cpu_base = subnet_base(subnet_map.get('cpu'))
    gpu_base = subnet_base(subnet_map.get('gpu'))
    storage_base = subnet_base(subnet_map.get('storage'))
    support_base = subnet_base(subnet_map.get('support'))

    # Track per-role indices for sequential IP assignment
    role_index = {'compute': 0, 'storage': 0, 'support': 0}
    # ...and the per-role TOTALS, known before the loop so the allocator can
    # decide once whether the legacy `.201` base seats everyone. Counted with
    # the same role classification the loop uses, so the two cannot disagree.
    # The predicate below is the loop's own skip sequence, kept in the same
    # order so a change to one is visibly a change to the other. A pre-count
    # that disagreed with the loop would hand the allocator a `total` that does
    # not match the nodes it actually allocates for, which is worse than the
    # bug being fixed.
    role_total = {'compute': 0, 'storage': 0, 'support': 0}
    for _n in nodes:
        if not _n.get('name', ''):
            continue
        if _n.get('status', 'Active') in ('Disabled', 'Air'):
            continue
        if _n.get('enabled', True) is False:
            continue
        # Bucket by POOL, not by role name. 'support', 'k8s' and 'bcme' are
        # three distinct roles that all draw from ONE pool (role_index
        # ['support'] below), so counting only the literal 'support' role
        # undercounts it — the repack guard then believes the legacy .101 base
        # fits when it does not, silently stranding exactly the nodes this
        # change exists to address. Caught in review, MR !262.
        _pool = _ROLE_ALLOC_POOL.get(
            classify_host_role(_n.get('role', _n.get('name', ''))))
        if _pool:
            role_total[_pool] += 1
    # Bottom-of-subnet addresses the fabric claims (gateway + per-switch SVI).
    _svi_reserved = _svi_reserved_offsets(nodes)

    for node in nodes:
        name = node.get('name', '')
        if not name:
            continue
        # parse_nodes() emits 'status' as 'Active'|'Disabled'|'Air'; an
        # older caller might pass 'enabled' bool. Honor either. Default
        # to active so callers that pass neither still work. Air status
        # = documentary row for auto-injected Air infra; same treatment
        # as Disabled here (don't provision from the Nodes-tab entry).
        if node.get('status', 'Active') in ('Disabled', 'Air'):
            continue
        if node.get('enabled', True) is False:
            continue

        # Use Function column ('role') for switch/infra classification — the Name column
        # may be an OEM hostname (e.g. 'nw-switch01') that doesn't match known patterns
        role = classify_host_role(node.get('role', name))

        # Skip switches and Air infrastructure nodes
        if role in ('switch', 'infra'):
            continue

        # MAC: use Excel value if present, otherwise auto-generate
        mac = node.get('mac') or generate_mac(name, "eth0")

        # eth0_ip: always use Nodes tab mgmt_ip as the authoritative source
        eth0_ip = node.get('mgmt_ip', '')

        entry = {
            'eth0_ip': eth0_ip,
            'mac': mac,
            # Excel `Function` is the source of truth for what a device IS
            # (docs/ROLES.md); the hostname is the operator's to choose. Carried
            # into devices so the Air Node-Instruction renderer can dispatch on
            # role instead of guessing from a name prefix — a customer whose
            # hosts are site-prefixed (`<site>-gpu-01`) matches no prefix
            # and silently received NO netplan at all.
            'role': role,
        }

        # Build interface mapping from Wire Map (if available)
        iface_map = {}
        if wiremap_rows:
            iface_map = build_interface_map(wiremap_rows, name)
            if iface_map:
                entry['interfaces'] = iface_map

        # Compute data-plane IPs based on role
        if role == 'compute' and cpu_base and gpu_base:
            idx = role_index['compute']

            if gpu_vlan_mode == 'per_rail_per_plane' and gpu_rail_planes and wiremap_rows:
                # Per-rail-per-plane mode: each GPU NIC belongs to a (rail,
                # plane) combination. Wire Map Network Profile carries both
                # ("GPU Rail R Plane P"). Same host octet across every
                # (rail, plane) — gpu-01 → .201 in every combination.
                host_offset = 201 + idx
                _bond = _dataplane_host_ips(subnet_map.get('cpu'), idx, 1, 201, 1,
                                            reserved=_svi_reserved,
                                            total=role_total['compute'])
                if _bond:
                    entry['bond_ip'] = _bond[0]
                gpu_ifaces_in_order = iface_map.get('gpu', [])
                nic_dests = build_nic_destinations(wiremap_rows, name)
                gpu_interfaces = []
                for nic in gpu_ifaces_in_order:
                    profile = (nic_dests.get(nic, {}).get('raw_profile') or '').strip()
                    # Match "GPU Rail R Plane P" in any whitespace/underscore form
                    m_prof = re.match(
                        r'^gpu[\s_-]*rail[\s_-]*(\d+)[\s_-]*plane[\s_-]*(\d+)$',
                        profile.lower())
                    if not m_prof:
                        continue
                    rail_key = f"rail{m_prof.group(1)}"
                    plane_key = f"plane{m_prof.group(2)}"
                    rp_info = gpu_rail_planes.get((rail_key, plane_key))
                    if not rp_info:
                        continue
                    net = ipaddress.ip_network(rp_info['subnet'], strict=False)
                    if host_offset >= net.num_addresses - 1:
                        continue
                    ip = f"{net.network_address + host_offset}/{net.prefixlen}"
                    gpu_interfaces.append({
                        'iface': nic,
                        # Combined key reused by template/netplan iterator
                        'plane': f"{rail_key}_{plane_key}",
                        'rail': rail_key,
                        'rail_plane': plane_key,
                        'ip': ip,
                        'gateway': rp_info['gateway'],
                        'table': rp_info['table'],
                    })

                if gpu_interfaces:
                    entry['gpu_interfaces'] = gpu_interfaces
            elif gpu_vlan_mode == 'per_rail' and gpu_rails and wiremap_rows:
                # Per-rail mode: each GPU NIC has its own VLAN + subnet. NIC
                # → rail mapping comes from the Wire Map's Network Profile
                # column ("GPU Rail N" → rail N). Same host octet on every
                # rail (gpu-01 → .201 across all rails, gpu-02 → .202, etc.).
                host_offset = 201 + idx
                _bond = _dataplane_host_ips(subnet_map.get('cpu'), idx, 1, 201, 1,
                                            reserved=_svi_reserved,
                                            total=role_total['compute'])
                if _bond:
                    entry['bond_ip'] = _bond[0]
                gpu_ifaces_in_order = iface_map.get('gpu', [])
                nic_dests = build_nic_destinations(wiremap_rows, name)
                gpu_interfaces = []
                for nic in gpu_ifaces_in_order:
                    # Use raw_profile (e.g. "GPU Rail 1") not the classified
                    # 'gpu' bucket, so we can extract the rail index.
                    profile = (nic_dests.get(nic, {}).get('raw_profile') or '').strip()
                    # Match "GPU Rail 1", "gpu rail 1", "GPU_Rail_1", etc.
                    m_prof = re.match(r'^gpu[\s_-]*rail[\s_-]*(\d+)$',
                                       profile.lower())
                    if not m_prof:
                        continue
                    rail_key = f"rail{m_prof.group(1)}"
                    rail_info = gpu_rails.get(rail_key)
                    if not rail_info:
                        continue
                    net = ipaddress.ip_network(rail_info['subnet'], strict=False)
                    if host_offset >= net.num_addresses - 1:
                        continue
                    ip = f"{net.network_address + host_offset}/{net.prefixlen}"
                    gpu_interfaces.append({
                        'iface': nic,
                        # 'plane' field is reused by the template/netplan
                        # iterator; for per-rail it holds 'rail1'/'rail2'/etc.
                        'plane': rail_key,
                        'ip': ip,
                        'gateway': rail_info['gateway'],
                        'table': rail_info['table'],
                    })

                if gpu_interfaces:
                    entry['gpu_interfaces'] = gpu_interfaces
            elif gpu_planes and wiremap_rows:
                # Dual-plane: pick each GPU NIC's IP from the plane its
                # Wire Map destination belongs to.
                gpu_ifaces_in_order = iface_map.get('gpu', [])
                # bond_ip is on the CPU /24. In dual-plane mode GPU IPs come
                # from per-plane /20s (not the CPU /24), so bond_ip stride is
                # 1 per node — single-plane retains its gpu_count*idx stride
                # because GPU IPs share the /24.
                _bond = _dataplane_host_ips(subnet_map.get('cpu'), idx, 1, 201, 1,
                                            reserved=_svi_reserved,
                                            total=role_total['compute'])
                if _bond:
                    entry['bond_ip'] = _bond[0]

                nic_dests = build_nic_destinations(wiremap_rows, name)
                gpu_interfaces = []
                plane_seen = {}  # plane -> nic count assigned within this node
                for nic in gpu_ifaces_in_order:
                    dst_switch = nic_dests.get(nic, {}).get('dst_switch', '')
                    plane = plane_for_switch(dst_switch)
                    if not plane or plane not in gpu_planes:
                        continue
                    plane_info = gpu_planes[plane]
                    net = ipaddress.ip_network(plane_info['subnet'], strict=False)
                    nic_idx_in_plane = plane_seen.get(plane, 0)
                    # 16 NIC headroom per node per plane (plenty in a /20).
                    plane_offset = 201 + 16 * idx + nic_idx_in_plane
                    if plane_offset >= net.num_addresses - 1:
                        continue
                    ip = f"{net.network_address + plane_offset}/{net.prefixlen}"
                    gpu_interfaces.append({
                        'iface': nic,
                        'plane': plane,
                        'ip': ip,
                        'gateway': plane_info['gateway'],
                        # Per-NIC PBR table: every NIC in the plane gets a
                        # unique routing table. Each table holds exactly one
                        # default route (this NIC's), so source-IP policy
                        # routing (`from <NIC IP>/32 lookup <table>`) pins
                        # outbound traffic per-NIC unambiguously. With the
                        # previous per-plane shared table, all 8 plane1 NICs
                        # ECMP'd through the same table — kernel couldn't
                        # tell which NIC owned `.201`, ARP responses leaked
                        # across NICs, EVPN saw the MAC at multiple VTEPs,
                        # and the gpu-gw VRR ping had ~85% loss.
                        'table': plane_info['table'] + nic_idx_in_plane,
                    })
                    plane_seen[plane] = nic_idx_in_plane + 1

                if gpu_interfaces:
                    entry['gpu_interfaces'] = gpu_interfaces
            else:
                # Single-plane path — bond_ip lives on the CPU /24, gpu_ips
                # live on the GPU /24. They're separate subnets, so stride
                # them independently: bond_ip steps by 1 per node (max ~53
                # nodes/24) while gpu_ips step by gpu_count per node (need
                # one address per GPU NIC).
                gpu_count = max(len(iface_map.get('gpu', [])), 2)
                _bond = _dataplane_host_ips(subnet_map.get('cpu'), idx, 1, 201, 1,
                                            reserved=_svi_reserved,
                                            total=role_total['compute'])
                if _bond:
                    entry['bond_ip'] = _bond[0]
                # One GPU IP per GPU interface, on the GPU subnet.
                gpu_ips = _dataplane_host_ips(
                    subnet_map.get('gpu'), idx, gpu_count, 201, gpu_count,
                    reserved=_svi_reserved)
                if gpu_ips:
                    entry['gpu_ips'] = gpu_ips
                    # Backward compatibility
                    if len(gpu_ips) >= 1:
                        entry['gpu_ip1'] = gpu_ips[0]
                    if len(gpu_ips) >= 2:
                        entry['gpu_ip2'] = gpu_ips[1]
            role_index['compute'] += 1

        elif role == 'storage' and storage_base:
            idx = role_index['storage']
            _bonds = _dataplane_host_ips(subnet_map.get('storage'), idx, 2, 101, 2,
                                         reserved=_svi_reserved,
                                         total=role_total['storage'])
            if _bonds:
                entry['bond_ip1'] = _bonds[0]
                entry['bond_ip2'] = _bonds[1]
            role_index['storage'] += 1

        elif role in ('support', 'k8s', 'bcme') and (support_base or cpu_base):
            idx = role_index['support']
            _bonds = _dataplane_host_ips(
                subnet_map.get('support') or subnet_map.get('cpu'), idx, 2, 101, 2,
                reserved=_svi_reserved, total=role_total['support'])
            if _bonds:
                entry['bond_ip1'] = _bonds[0]
                entry['bond_ip2'] = _bonds[1]
            role_index['support'] += 1

        devices[name] = entry

    # Note: switches are NOT added to devices — they get DHCP reservations
    # from the ZTP section of the dnsmasq template (using host_vars mac_address).

    return devices


def parse_oob_switch_configs(ws, ws_air_only=None, nodes_function_map=None):
    """Parse Wire Map (and optionally Air_Only) sheet to derive OOB switch port configs.

    Wire Map columns (0-based):
      1  = System Role  (the node/server side of the connection)
      3  = NIC/Port
      10 = Switch Role  (which switch this cable plugs into)
      12 = Switch Port  (which port on that switch)

    Air_Only columns (0-based):
      1  = System Role
      5  = Switch Role
      7  = Switch Port

    Logic:
      - Rows where Switch Role is 'oob-switch-*' and Switch Port is a plain swpN
        (no sub-port) are OOB switch port assignments.
      - Uplink detection uses two independent signals (either is sufficient):
        1. Network Profile is ISL / L3 / OOB Uplink (profile-first, hostname-agnostic)
        2. Peer Function resolves to core/csl (Function-first via nodes_function_map)
      - All other ports are access ports serving BMC/IPMI connections.
      - Air_Only rows (dhcp-oob, oob-server-01, dhcp-edge) are always access ports.
      - The port on each OOB switch connected to dhcp-oob is tracked as dhcp_oob_port
        so the ZTP MAC can be computed as generate_mac(function_name, dhcp_oob_port).

    Returns:
      {oob_switch_name: {
          'access_ports': 'swp1-43',
          'uplink_ports': 'swp1-43,swp49,swp51',
          'spine_bond_members': ['swp49', 'swp51'],
          'dhcp_oob_port': 'swp44',   # port connected to dhcp-oob (ZTP port)
      }}
    """
    nfm = nodes_function_map or {}
    oob_access = defaultdict(set)
    oob_uplink = defaultdict(set)
    dhcp_oob_ports: dict = {}  # {oob_switch_function_name: port_str}

    _UPLINK_PROFILES = frozenset({'isl', 'l3', 'oob uplink', 'sn2201 uplink'})

    col_map = build_wiremap_column_map(ws, sheet_kind='wiremap')
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i == 1:
            continue  # header row
        display_raw = _wm_cell(row, col_map, 'display_in_air').lower()
        system_role = _wm_cell(row, col_map, 'system_role')
        system_name = _wm_cell(row, col_map, 'system_name') or system_role
        nic_port = _wm_cell(row, col_map, 'nic_port')
        switch_role = _wm_cell(row, col_map, 'switch_role')
        switch_name = _wm_cell(row, col_map, 'switch_name') or switch_role
        switch_port = _wm_cell(row, col_map, 'switch_port')
        net_prof = (_wm_cell(row, col_map, 'network_profile') or '').strip().lower()
        is_air_visible = (display_raw == 'yes')

        # Resolve via Nodes-tab Function map first, then hostname fallback.
        sys_func = nfm.get(system_name, system_role)
        sw_func  = nfm.get(switch_name, switch_role)
        sys_cat = canonical_category(sys_func, system_name)
        sw_cat  = canonical_category(sw_func, switch_name)

        # Branch 1: OOB switch is the right-hand side (typical case — node →
        # oob-switch). Captures access ports and dhcp-oob port reservations.
        # Use switch_name (col 12) as the unique key — switch_role becomes
        # ambiguous after canonical conversion (every OOB switch has
        # switch_role='oob-switch').
        if sw_cat == 'oob-switch' and switch_port and switch_port != 'None':
            # Track which port dhcp-oob connects to (this is the ZTP interface)
            if system_role == 'dhcp-oob' or system_name == 'dhcp-oob':
                dhcp_oob_ports[switch_name] = switch_port

            # Only plain swpN ports (no sub-ports like swp49s0 on OOB switches)
            m = re.match(r'^swp(\d+)$', switch_port)
            if m:
                port_num = int(m.group(1))
                # Uplink detection: Network Profile OR peer role (Function-first).
                is_uplink = (
                    (net_prof in _UPLINK_PROFILES) or
                    (sys_cat in ('core', 'csl', 'cl') and nic_port != 'eth0')
                )
                if is_uplink:
                    if is_air_visible:
                        oob_uplink[switch_name].add(port_num)
                elif is_air_visible:
                    oob_access[switch_name].add(port_num)

        # Branch 2: OOB switch is the LEFT side (oob → core/csl uplinks). For
        # these rows the OOB-side port is in nic_port. Used in the 2-8-9-800
        # OEM wiremap layout where OOB→CSL uplinks are written from the OOB
        # perspective.
        elif sys_cat == 'oob-switch' and nic_port:
            is_uplink = (
                (net_prof in _UPLINK_PROFILES) or
                sw_cat in ('core', 'csl', 'cl')
            )
            if is_uplink:
                if is_air_visible:
                    m = re.match(r'^swp(\d+)$', nic_port)
                    if m:
                        oob_uplink[system_name].add(int(m.group(1)))

    # Also include Air_Only rows (virtual nodes: dhcp-oob, oob-server-01, dhcp-edge).
    # Some Excels use Air_Only purely for the version-image map + Air mgmt subnet
    # metadata and have no connection rows. Treat missing required columns as
    # "no rows to read" rather than failing the whole import.
    try:
        ao_col_map = build_wiremap_column_map(ws_air_only, sheet_kind='air_only') if ws_air_only is not None else None
    except ValueError:
        ao_col_map = None
    if ws_air_only is not None and ao_col_map is not None:
        for i, row in enumerate(ws_air_only.iter_rows(values_only=True), 1):
            if i == 1:
                continue  # header row
            system_role_ao = _wm_cell(row, ao_col_map, 'system_role')
            system_name_ao = _wm_cell(row, ao_col_map, 'system_name') or system_role_ao
            switch_role = _wm_cell(row, ao_col_map, 'switch_role')
            switch_name = _wm_cell(row, ao_col_map, 'switch_name') or switch_role
            switch_port = _wm_cell(row, ao_col_map, 'switch_port')

            if canonical_category(switch_role, switch_name) != 'oob-switch' \
                    or not switch_port or switch_port == 'None':
                continue

            # Track dhcp-oob port (Air_Only takes priority over Wire Map)
            if system_role_ao == 'dhcp-oob' or system_name_ao == 'dhcp-oob':
                dhcp_oob_ports[switch_name] = switch_port

            m = re.match(r'^swp(\d+)$', switch_port)
            if not m:
                continue
            oob_access[switch_name].add(int(m.group(1)))

    result = {}
    for sw in set(oob_access) | set(oob_uplink):
        access_nums = oob_access[sw]
        uplink_nums = oob_uplink[sw]
        # SN2201 has 48 host ports (swp1-48) and 4 uplinks (swp49-52).
        # All 48 host ports stay in the VLAN 200 access bridge regardless of
        # which are wired. This is DELIBERATE, not a shortcut — an OOB switch
        # is a plug-and-play access pool, so an operator racking a server later
        # must be able to use any free port and land on VLAN 200 without
        # regenerating and re-applying config. Compute-fabric ports are
        # precisely wired and derived from the Wire Map; OOB host ports are not.
        #
        # Do NOT "fix" this by deriving the range from the Wire Map (ERA-58).
        # Production captures show shorter ranges — swp1-46 on sw2432/nw2432,
        # swp1-34 on svc, all on identical 52-port SN2201s — but those are
        # sites that scoped the bridge to what they had cabled, which trades
        # away the plug-and-play property.
        #
        # The one case that genuinely needs a narrower range is DIFFERENT VLANs
        # on the same OOB switch: nw2432-oob-1 keeps swp47 `link state up` and
        # `type swp` but deliberately OUTSIDE the VLAN 200 bridge. Multi-VLAN
        # OOB is deferred, so the blanket is correct until it lands; when it
        # does, per-port VLAN assignment replaces this, not a cabling-derived
        # range. ERA-58 is deferred on exactly that dependency.
        #
        # Air also injects a side-channel link from each OOB switch to
        # air-oob-switch on the first unused port in 1..52 (see
        # topology_generator._inject_air_oob_switch, pass 3). That port
        # carries VLAN 200 traffic from dhcp-oob / oob-server-01 to the
        # hosts, so it must also be in the access bridge. If it falls
        # outside the swp1-48 host range (e.g., swp50 when all 48 host
        # ports are wired), we add it explicitly.
        bridge_nums = set(range(1, 49))
        air_oob_port_num = next(
            (p for p in range(1, 53)
             if p not in access_nums and p not in uplink_nums),
            None,
        )
        if air_oob_port_num is not None:
            bridge_nums.add(air_oob_port_num)
        entry = {
            'access_ports': ports_to_range_string(bridge_nums),
            'uplink_ports': ports_to_range_string(bridge_nums | uplink_nums),
            'spine_bond_members': [f'swp{n}' for n in sorted(uplink_nums)],
        }
        if sw in dhcp_oob_ports:
            entry['dhcp_oob_port'] = dhcp_oob_ports[sw]
        result[sw] = entry
    return result


# Air virtual nodes — these exist in Air simulations but not in physical deployments
AIR_VIRTUAL_NODES = {"dhcp-oob", "oob-server-01", "dhcp-edge"}


def parse_air_virtual_nodes(ws, new_format=False):
    """Detect Air virtual nodes from rows with 'Air - ' network profiles.

    For the old format, reads Wire Map (col 7=profile, col 2=system_role, col 11=switch_role).
    For the new format, reads Air_Only sheet (col 5=profile, col 2=system_role, col 6=switch_role).

    Returns a set of virtual node names found (e.g. {'dhcp-oob', 'oob-server-01'}).
    """
    virtual_nodes = set()
    # New Air_Only sheet: System Role=col2, Network Profile=col5, Switch Role=col6
    # Old Wire Map sheet: System Role=col2, Network Profile=col7, Switch Role=col11
    profile_col = 4 if new_format else 6   # 0-based
    switch_col  = 5 if new_format else 10  # 0-based

    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i == 1:
            continue
        if len(row) <= profile_col:
            continue  # skip short rows (e.g., version mapping table)
        profile = str(row[profile_col]).strip() if row[profile_col] else ''
        if not profile.startswith('Air - '):
            continue
        system_role = str(row[1]).strip() if row[1] else ''
        switch_role = str(row[switch_col]).strip() if row[switch_col] else ''
        if system_role in AIR_VIRTUAL_NODES:
            virtual_nodes.add(system_role)
        if switch_role in AIR_VIRTUAL_NODES:
            virtual_nodes.add(switch_role)
    return virtual_nodes


def parse_air_settings(ws_air_only):
    """Read key-value settings from the Air_Only sheet.

    Scans for rows where col1 is a known setting name.
    Returns dict: {'air_mgmt_subnet': '172.20.0.0/24', ...}
    """
    _KNOWN_KEYS = {
        'air management subnet': 'air_mgmt_subnet',
    }
    settings = {}
    for i, row in enumerate(ws_air_only.iter_rows(values_only=True), 1):
        if len(row) < 2 or not row[0]:
            continue
        key = str(row[0]).strip().lower()
        if key in _KNOWN_KEYS and row[1]:
            settings[_KNOWN_KEYS[key]] = str(row[1]).strip()
    return settings


def parse_node_mgmt_mapping(ws, new_format=False):
    """Parse Wire Map eth0 rows to find node → OOB switch mappings.

    Old format: detects rows where profile contains 'Air - Management'.
    New format: detects rows where nic_port=='eth0' and switch connects to an OOB switch
                (Air - Management profile was replaced by OOB / IPMI in the Wire Map).

    Returns dict: {node_name: oob_switch_name}
    e.g. {'su-01-node-01': 'oob-switch-03', 'support-01': 'oob-switch-01'}
    """
    mapping = {}
    col_map = build_wiremap_column_map(ws, sheet_kind='wiremap')
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i == 1:
            continue
        nic_port    = _wm_cell(row, col_map, 'nic_port')
        profile     = _wm_cell(row, col_map, 'network_profile')
        switch_role = _wm_cell(row, col_map, 'switch_role')
        switch_name = _wm_cell(row, col_map, 'switch_name') or switch_role
        system_role = _wm_cell(row, col_map, 'system_role')
        system_name = _wm_cell(row, col_map, 'system_name') or system_role

        # Identify OOB-switch peer via canonical_category so this works
        # with both legacy hostname-as-role and post-step-4b canonical
        # cells. Use switch_name as the dict value so the mapping survives
        # canonical conversion (where every OOB row has the same role).
        if nic_port != 'eth0' or canonical_category(switch_role, switch_name) != 'oob-switch':
            continue

        key = system_name or system_role
        if new_format:
            # New format: eth0 OOB connections use OOB / IPMI profile in Wire Map
            if key:
                mapping[key] = switch_name
        else:
            # Old format: only rows explicitly tagged 'Air - Management'
            if 'Air' in profile and 'Management' in profile and key:
                mapping[key] = switch_name
    return mapping


# Hardware constants for SN5610/SN5600 per role type. These are FALLBACKS —
# the Port Profiles sheet is authoritative and overrides all of them.
#
# breakout = number of sub-ports per physical port.
# lanes    = physical lanes per sub-port.
# The ports are 800G over 8 lanes, so the per-lane rate is 100G and a
# sub-port runs at `lanes x 100G`. breakout x lanes == 8 on every row.
# Speed itself is NOT defaulted here: it comes from the sheet or is absent
# (ERA-55). The comments below record what each geometry implies.
_ROLE_HW = {
    'cpu':     {'breakout': 4, 'lanes': 2},   # 4x200G
    'gpu':     {'breakout': 2, 'lanes': 4},   # 2x400G direct
    'support': {'breakout': 4, 'lanes': 2},   # 4x200G, same as cpu
    'storage': {'breakout': 4, 'lanes': 2},   # 4x200G, bonded storage nodes
    'isl':     {'breakout': 2, 'lanes': 4},   # 2x400G direct
    'oob':     {'breakout': 8, 'lanes': 1},   # 8x100G uplinks to OOB switches
    'edge':    {'breakout': 8, 'lanes': 1},   # 8x100G uplinks to customer edge
    'storage_uplink': {'breakout': 8, 'lanes': 1},  # 8x100G uplinks to storage switches
}

# Map role type -> VLANs & Profiles sheet profile name (for config lookup)
# NVUE association types for `nv set system ntp server <s> association-type <t>`.
# Closed set: the value is rendered into a root-executed config script, so an
# unrecognised type falls back to 'server' rather than passing through.
# NVUE values for `link auto-negotiate`. CLOSED SET — this cell is rendered
# UNQUOTED into a root-executed switch config, so anything outside it is
# refused rather than passed through. `SHELL_INJECTION_PRONE_KEYS` in
# validate_excel guards the Settings sheet only; Port Profiles cells are not
# covered by it, and this column is one of two that reach the config verbatim
# (the other, Speed, is regex-validated below for the same reason).
_VALID_AUTO_NEGOTIATE = frozenset({'on', 'off', 'enabled', 'disabled'})


def _validated_auto_negotiate(value, profile_name):
    """Return a whitelisted `link auto-negotiate` value, or None.

    Refuses anything not in _VALID_AUTO_NEGOTIATE. A value like
    `disabled; touch /tmp/x` would otherwise render as
    `nv set interface ... link auto-negotiate disabled; touch /tmp/x`
    and execute as root on every switch carrying that profile.
    """
    if value in (None, ''):
        return None
    text = str(value).strip().lower()
    if text in _VALID_AUTO_NEGOTIATE:
        return text
    print(f"  ⚠️  Port Profiles: profile '{profile_name}' has an unrecognised "
          f"Auto-Negotiate value {str(value).strip()!r} — expected one of "
          f"{', '.join(sorted(_VALID_AUTO_NEGOTIATE))}. Ignoring.",
          file=sys.stderr)
    return None


_VALID_NTP_ASSOCIATION_TYPES = frozenset({'server', 'pool', 'peer'})


_ROLE_PROFILE_NAME = {
    'cpu':             'CPU/In-Band Network',
    'gpu':             'GPU Network',
    'support':         'Support',
    'storage':         'Storage',
    'oob':             'OOB Uplink',
    'oob_access':      'OOB / IPMI',
    'edge':            'Edge Uplink',
    'isl':             'ISL',
    'storage_uplink':  'Storage Uplink',
}


def _classify_node_profile(net_prof):
    """Return role type for a node-to-core Wire Map profile.

    Known categories (cpu/gpu/support/storage) keep their canonical
    short role keys for template back-compat. Any other profile name
    falls through to a slugified role key derived from the profile
    name itself — e.g. "BCM Network" → "bcm", "K8s" → "k8s" — so the
    operator can declare custom server roles in the Port Profiles
    section (each with its own Allowed VLANs / Untagged / VRF /
    LACP Bypass) and the parser emits a separate bond group per
    profile. The customer golden config uses this shape for bcm vs
    slurm vs k8s with different VLAN allow lists.

    Returns None only for empty input.
    """
    if not net_prof:
        return None
    p = net_prof.lower()
    if 'cpu' in p or 'in-band' in p:
        return 'cpu'
    if 'gpu' in p:
        return 'gpu'
    if 'oob' in p and 'uplink' in p:
        return 'oob'
    if 'support' in p:
        return 'support'
    if 'storage' in p:
        return 'storage'
    return _slugify_role_name(net_prof)


def _slugify_role_name(profile_name):
    """Slugify a Port Profile name into a role key.

    "BCM Network"  → "bcm"
    "Slurm"        → "slurm"
    "K8s Network"  → "k8s"
    "Test Compute" → "test_compute"
    """
    s = (profile_name or '').strip().lower()
    # Drop trailing " network" so "BCM Network" and "BCM" both → "bcm"
    if s.endswith(' network'):
        s = s[: -len(' network')]
    return re.sub(r'[^a-z0-9]+', '_', s).strip('_')


def _classify_core_profile(net_prof, sw_roles, l3_storage_profiles=None):
    """Return role type for a core-as-system Wire Map profile.

    Storage split (ERA-44): a profile in `l3_storage_profiles` (Mode=L3 AND
    VRF=STORAGE) is a genuine external uplink -> 'storage_uplink' (bound to the
    VRF, never bonded). Any other storage-y profile -- even one literally named
    "Storage Uplink" but with an L2 Port Mode -- is a storage SERVER access
    port -> 'storage' (L2 bond into the storage VLAN). Discriminator is the
    profile MODE, never the profile name or the peer's role.
    """
    p = net_prof.lower()
    sw = ' '.join(r.lower() for r in sw_roles)
    if 'isl' in p or 'peer' in p:
        return 'isl'
    # OOB uplinks: "OOB Uplink", "OOB...Uplink", "SN2201 Uplink", or uplink to oob-switch
    if ('oob' in p and 'uplink' in p) or ('sn2201' in p) or ('uplink' in p and 'oob-switch' in sw):
        return 'oob'
    # Storage: genuine L3/VRF=STORAGE uplink vs L2 storage-server access port.
    if ('storage' in p and 'uplink' in p) or ('storage' in sw and 'oob-switch' not in sw and 'edge' not in sw):
        if l3_storage_profiles and net_prof in l3_storage_profiles:
            return 'storage_uplink'
        return 'storage'
    # Edge/EXIT uplinks: "Edge Uplink", "ESL Uplink", or generic uplink
    if 'edge' in p or 'uplink' in p or 'esl' in p:
        return 'edge'
    return None


def _normalize_oob_uplink_mode(settings):
    """Return the Excel-driven OOB uplink mode, defaulting to l2 (the
    existing/tested default for all four ERA architectures). Topology
    generator and parser both default to 'l2' to stay in sync.

    Warns once if the operator typed something other than 'l2'/'l3'."""
    raw = (settings or {}).get('oob_uplink_mode', None)
    if raw is None or str(raw).strip() == '':
        return 'l2'
    mode = str(raw).strip().lower()
    if mode not in ('l2', 'l3'):
        print(f"  ⚠️  Settings.oob_uplink_mode='{raw}' not in ('l2','l3') — "
              f"falling back to l2.")
        return 'l2'
    return mode


def _expand_subport_names(iface_def):
    """Expand an iface_def into a sorted list of sub-port names."""
    names = []
    port_overrides = iface_def.get('port_overrides', {}) if iface_def else {}
    for port in iface_def.get('ports', []):
        breakout = port_overrides.get(port, {}).get('breakout', iface_def['breakout'])
        subports = port_overrides.get(port, {}).get('subports', range(breakout))
        for sub in subports:
            names.append(f"swp{port}s{sub}")
    return names


def _supplement_vrf_config_from_excel(core_vars, vrfs, dhcp_relay_table, vlans):
    """Append vrf_config entries for Excel-declared VRFs that the source
    inventory omitted.

    Level-1 of "Excel as source of truth" for VRF declarations. The source
    inventory's vrf_config remains authoritative for whatever it declares
    (preserving byte-identity on existing archs), and this function fills
    the gaps that newer archs leave behind. A VRF gets a supplemental entry
    when the Excel declares it AND something on this switch references it
    (a VLAN's VRF assignment OR a DHCP Relay scope). Default VRF and GPU
    VRF are skipped — they're emitted through other paths.

    The generated entry mirrors the established ERA pattern: ipv4_unicast
    + l2vpn_evpn enables, redistribute_connected, route_export_to_evpn,
    and a route_import_from_vrf clause that imports from every other VRF
    that carries a DHCP Relay scope (OOB ↔ EXIT ↔ INBAND topology). The
    referenced <VRF>_FILTER route-maps must exist in the rendered
    inventory — they do, because the source-inventory csl.yml carries
    them as the ERA standard policy template.

    Level-2 TODO (tracked in docs/internal/TODO-audit-findings.md):
    move route_map / community_list / peer_groups generation into the
    parser too so the source inventory can be retired entirely.
    """
    if not vrfs:
        return
    existing = core_vars.setdefault('vrf_config', [])
    if not isinstance(existing, list):
        return
    existing_ids = {e.get('id') for e in existing if isinstance(e, dict)}

    referenced = set()
    for v in (vlans or []):
        vrf = (v.get('vrf') or '').strip().upper()
        if vrf:
            referenced.add(vrf)
    relay_vrfs = set()
    for entry in (dhcp_relay_table or []):
        vrf = (entry.get('vrf') or '').strip().upper()
        if vrf:
            relay_vrfs.add(vrf)
            referenced.add(vrf)

    for vrf_name, vrf_data in vrfs.items():
        vrf_id = str(vrf_name).strip().upper()
        # STORAGE is emitted via per-host vrf_config_extra by the
        # external-storage-uplink path (`_build_storage_uplink_host_vars`
        # near line 3819). Skipping here avoids a duplicate `nv set vrf
        # STORAGE ...` block on archs that already use that path.
        if vrf_id in ('', 'DEFAULT', 'GPU', 'STORAGE'):
            continue
        if vrf_id in existing_ids:
            continue
        if vrf_id not in referenced:
            continue

        l3_vni = vrf_data.get('l3_vni')
        l3_vlan = vrf_data.get('vlan')
        if not l3_vni or not l3_vlan:
            continue

        ipv4_af = {
            'enable': 'on',
            'redistribute_connected': 'on',
            'route_export_to_evpn': 'on',
        }
        if vrf_id in relay_vrfs:
            peers = sorted(relay_vrfs - {vrf_id})
            if peers:
                ipv4_af['route_import_from_vrf'] = {
                    'enable': 'on',
                    'list': peers,
                    'route_map': f'{vrf_id}_FILTER',
                }
        entry = {
            'id': vrf_id,
            'vlan': int(l3_vlan) if str(l3_vlan).isdigit() else l3_vlan,
            'vni': str(l3_vni),
            'route_export': True,
            'bgp': {
                'address_family': {
                    'ipv4_unicast': ipv4_af,
                    'l2vpn_evpn': {'enable': 'on'},
                },
            },
        }

        # If this VRF terminates an external eBGP session — today only EXIT
        # carries customer-net traffic via the Edge Uplink port profile —
        # stitch in the `exit` peer-group so the L3 uplink ports actually
        # peer. `_sync_edge_vrf_neighbors` (called from the
        # same `_apply_oob_l3_uplink_mode` neighborhood) will overwrite the
        # `interfaces` list with the Excel-driven edge_interfaces values.
        # If a future arch adds a second external-peering VRF, generalize
        # by detecting "Port Profile mode=L3 + vrf=<this>" from the Excel
        # profile_config rather than hard-coding EXIT.
        # Only seed the external eBGP neighbor when `edge_interfaces`
        # actually expands to at least one subport name. If `ports: []`
        # made it through but `_sync_edge_vrf_neighbors` early-returns on
        # empty subports, we'd leave an `interfaces: []` placeholder that
        # renders as a malformed `nv set vrf EXIT router bgp neighbor`
        # block with zero peers.
        if (vrf_id == 'EXIT'
                and core_vars.get('edge_interfaces')
                and _expand_subport_names(core_vars['edge_interfaces'])):
            entry['bgp']['neighbors'] = [{
                'interfaces': [],
                'peer_group': 'exit',
                'type': 'unnumbered',
            }]
            # ADR-0043: named `exit` for the role it serves, not
            # `underlay_esl_external` for where it sits. Verified 1:1 with that
            # role before renaming — all 637 occurrences of the old name lived
            # in `vrf EXIT`, on cl/core/csl edge-uplink ports only, so this is
            # a pure rename with no change of membership, VRF or policy.
            # `external` on cust-net-edge is a different peer-group (default
            # VRF, far side of this peering) and is deliberately untouched.
            entry['bgp']['peer_groups'] = [{
                'id': 'exit',
                'remote_as': 'external',
                'address_family': {
                    'ipv4_unicast': {
                        'enable': 'on',
                        'policy': {
                            'outbound_route_map': 'OUTBOUND_ERA_PREFIXES',
                        },
                    },
                },
            }]

        existing.append(entry)
        existing_ids.add(vrf_id)


def _sync_edge_vrf_neighbors(core_vars):
    """Keep inherited EXIT VRF neighbors aligned with Excel edge ports.

    Source inventory group_vars carry the routing policy and peer-group
    details, but their neighbor interface lists are tied to the reference
    workbook. When the generated XLSX changes edge ports, the generated
    edge_interfaces must be authoritative.
    """
    edge_interfaces = core_vars.get('edge_interfaces')
    vrf_config = core_vars.get('vrf_config')
    if not edge_interfaces or not isinstance(vrf_config, list):
        return

    edge_subports = _expand_subport_names(edge_interfaces)
    if not edge_subports:
        return

    for vrf in vrf_config:
        if not isinstance(vrf, dict):
            continue
        bgp = vrf.get('bgp')
        if not isinstance(bgp, dict):
            continue
        for neighbor in bgp.get('neighbors') or []:
            if not isinstance(neighbor, dict):
                continue
            if neighbor.get('peer_group') == 'exit':
                neighbor['interfaces'] = list(edge_subports)


def parse_core_port_config(ws_wiremap, ws_vlans_profiles, nodes_function_map=None,
                           vlans=None, oob_uplink_mode='l3'):
    """Derive core switch port configuration from the Wire Map sheet.

    Reads two classes of Wire Map rows:
      A. Node-to-core connections: sys_role != core-*, sw_role == core-*
         Switch Port (swpNsX) tells us which core port serves which network profile.
      B. Core-as-system connections: sys_role == core-01
         NIC/Port (swpNsX) tells us the core's own uplinks (OOB, ISL, Edge, Storage).
         Plain integers indicate 'Port Disabled by Neighbor' (adjacent disabled port).

    `nodes_function_map` (optional) maps `hostname → canonical function` from
    the Nodes tab. Used as a fallback when the Wire Map's `Function (A/B)`
    columns are blank or deleted — without it, rows with no role on either
    side get skipped entirely, leaving the source-inventory fallback in
    place (which may carry stale data like the wrong storage `vlan_untagged`).

    `vlans` (optional) is the parsed VLAN list from the VLANs sheet. When
    provided AND the Wire Map references `GPU Rail <N>` profiles, the
    return dict also includes `gpu_rail_interfaces` keyed by rail index,
    with each rail's port list + VLAN ID. The core template uses this to
    emit one `bridge domain br_default access <vid>` line per rail
    instead of bundling all GPU ports into a single VLAN.

    Returns a dict suitable for merging into group_vars/core.yml:
      network_roles, gpu_interfaces, isl_interfaces, edge_interfaces,
      storage_interfaces (if any), interfaces_disabled,
      gpu_rail_interfaces (per-rail mode only).
    """
    nodes_function_map = nodes_function_map or {}
    # Per-rail mode: map "GPU Rail N" profile names → rail VLAN IDs.
    # Per-rail-per-plane mode: map "GPU Rail R Plane P" profile names →
    #   per-(rail, plane) VLAN IDs. Keys are (rail_idx, plane_idx) tuples.
    # Both empty when not in the respective mode.
    gpu_rail_vlan_ids: dict = {}
    gpu_rail_plane_vlan_ids: dict = {}
    oob_vlan_id = None
    if vlans:
        for vlan in vlans:
            name = (vlan.get('name') or '').lower()
            if oob_vlan_id is None and name.startswith('oob') and vlan.get('id'):
                oob_vlan_id = vlan['id']
            m_rp = re.match(r'^gpu_rail(\d+)_plane(\d+)$', name)
            if m_rp and vlan.get('id'):
                gpu_rail_plane_vlan_ids[(int(m_rp.group(1)), int(m_rp.group(2)))] = vlan['id']
                continue
            m = re.match(r'^gpu_rail(\d+)$', name)
            if m and vlan.get('id'):
                gpu_rail_vlan_ids[int(m.group(1))] = vlan['id']
    # --- Step 1: Build profile config from VLANs & Profiles sheet ---
    # Find the "Port Profiles" section header, then read rows below it.
    # Columns (within Port Profiles section):
    #   1=Profile, 2=Port Mode, 3=Native/Access VLAN, 4=Allowed VLANs,
    #   5=Untagged VLAN, 6=VRF, 7=LACP Bypass, 8=Port Speed
    profile_config = {}   # profile_name -> {mode, native, allowed, untagged, vrf, lacp_bypass}
    in_profiles_section = False
    for row in range(1, ws_vlans_profiles.max_row + 1):
        cell_val = ws_vlans_profiles.cell(row, 1).value
        if isinstance(cell_val, str) and cell_val.strip() == 'Port Profiles':
            in_profiles_section = True
            continue  # skip the section header row
        if not in_profiles_section:
            continue
        name = cell_val
        mode = ws_vlans_profiles.cell(row, 2).value
        if isinstance(name, str) and name.strip() == 'Profile':
            continue  # skip column header row
        if not isinstance(name, str) or not mode:
            continue
        native = ws_vlans_profiles.cell(row, 3).value
        allowed = ws_vlans_profiles.cell(row, 4).value
        untagged = ws_vlans_profiles.cell(row, 5).value
        vrf = ws_vlans_profiles.cell(row, 6).value
        lacp_bypass_val = ws_vlans_profiles.cell(row, 7).value
        # Column 8 is Speed — the per-sub-port link rate (lanes x 100G on the
        # 800G SN5600/SN5610 ports). It was authored in every shipped workbook
        # but skipped here for as long as the column has existed, so nothing
        # downstream could see it (ERA-55).
        speed_val = ws_vlans_profiles.cell(row, 8).value
        breakout_val = ws_vlans_profiles.cell(row, 9).value
        lanes_val = ws_vlans_profiles.cell(row, 10).value
        # Column 11 is Auto-Negotiate (ERA-58). Appended after Lanes rather
        # than inserted, so none of the positional reads above shift — the
        # failure mode ERA-81 records.
        autoneg_val = ws_vlans_profiles.cell(row, 11).value
        speed = str(speed_val).strip() if speed_val not in (None, '') else None
        if speed and not re.match(r'^\d+(\.\d+)?[MGT]$', speed):
            print(f"  ⚠️  Port Profiles: profile '{name}' has an unrecognised "
                  f"Speed '{speed}' — expected a value like '200G'. Ignoring.",
                  file=sys.stderr)
            speed = None
        profile_config[name] = {
            'mode': mode,
            'native': native,
            'allowed': allowed,
            'untagged': untagged,
            'vrf': str(vrf).strip() if vrf else None,
            'lacp_bypass': str(lacp_bypass_val).strip().lower() in ('yes', 'true', '1')
                          if lacp_bypass_val else False,
            'breakout': int(breakout_val) if breakout_val else None,
            'lanes': int(lanes_val) if lanes_val else None,
            'speed': speed,
            'auto_negotiate': _validated_auto_negotiate(autoneg_val, name),
        }
    # ERA-44: profiles that are genuine L3 external storage uplinks
    # (Mode=L3 AND VRF=STORAGE). Everything else that looks storage-y is an
    # L2 storage SERVER access port. Same rule as find_l3_storage_profiles().
    l3_storage_profiles = {
        name for name, cfg in profile_config.items()
        if str(cfg.get('mode') or '').strip().lower() == 'l3'
        and str(cfg.get('vrf') or '').strip().upper() == 'STORAGE'
    }

    # Build reverse map (slug → original profile name) for custom roles.
    # Known roles (cpu/gpu/support/storage/etc.) still resolve via the
    # canonical _ROLE_PROFILE_NAME table; custom roles use their slug.
    slug_to_profile_name = {}
    for pname in profile_config.keys():
        slug = _slugify_role_name(pname)
        if slug not in slug_to_profile_name:
            slug_to_profile_name[slug] = pname

    def _profile_for_role(role_type):
        """Return the full profile config dict for a role type, or empty dict."""
        prof = _ROLE_PROFILE_NAME.get(role_type)
        if prof and prof in profile_config:
            return profile_config[prof]
        # Custom role: look up by slug → original profile name.
        prof = slug_to_profile_name.get(role_type)
        if prof and prof in profile_config:
            return profile_config[prof]
        return {}

    def _vlan_for_role(role_type):
        return _profile_for_role(role_type).get('native')

    # --- Step 2: Collect Wire Map rows ---
    # node_profiles[prof][base_port] = {subports}
    node_profiles = defaultdict(lambda: defaultdict(set))
    # core_ports[prof] = [(base_port, subport, sw_role), ...]
    core_ports = defaultdict(list)
    # role_core_entries also accepts switch-owned direct-interface rows that
    # arrive with the external peer on side A and the core/CSL switch on side B.
    role_core_entries = defaultdict(list)  # role_type→[(base, sub, sw_role)]
    disabled_ports = []

    wm_col_map = build_wiremap_column_map(ws_wiremap, sheet_kind='wiremap')
    for row in range(2, ws_wiremap.max_row + 1):
        sys_role = _wm_cell_ws(ws_wiremap, row, wm_col_map, 'system_role')
        sys_name = _wm_cell_ws(ws_wiremap, row, wm_col_map, 'system_name') or sys_role
        nic_port = _wm_cell_ws(ws_wiremap, row, wm_col_map, 'nic_port')
        net_prof = _wm_cell_ws(ws_wiremap, row, wm_col_map, 'network_profile')
        sw_role  = _wm_cell_ws(ws_wiremap, row, wm_col_map, 'switch_role')
        sw_name  = _wm_cell_ws(ws_wiremap, row, wm_col_map, 'switch_name') or sw_role
        sw_port  = _wm_cell_ws(ws_wiremap, row, wm_col_map, 'switch_port')

        # Cascade missing roles from the Nodes tab — covers Excels where
        # the Function (A/B) columns have been deleted in favour of
        # System Name + Nodes-tab lookup.
        if not sys_role and sys_name:
            sys_role = nodes_function_map.get(sys_name, '')
        if not sw_role and sw_name:
            sw_role = nodes_function_map.get(sw_name, '')

        if not sys_role or not net_prof:
            continue

        sys_cat = canonical_category(sys_role, sys_name)
        sw_cat  = canonical_category(sw_role, sw_name)

        # In dedicated_gpu designs (e.g. 2-8-9-800) the converged-fabric leaf
        # is named csl-* rather than core-*. The wiremap rows are otherwise
        # identical in shape, so we treat the first csl the same way as
        # the first core. Detect "first-of-its-kind" via index extraction so
        # hostname-as-role legacy AND canonical-with-arbitrary-name both work.
        if sys_cat in ('core', 'csl', 'cl') and extract_role_index(sys_name) == 1:
            # Core/CSL is the "system" side — nic_port is the switch's own port
            prof_lower = net_prof.lower()
            if 'disabled' in prof_lower or 'neighbor' in prof_lower:
                m = re.match(r'^(?:swp)?(\d+)$', nic_port)
                if m:
                    disabled_ports.append(int(m.group(1)))
            elif 'unused' in prof_lower or 'usused' in prof_lower:
                pass  # intentionally unused sub-ports — skip
            else:
                m = re.match(r'^swp(\d+)s(\d+)$', nic_port)
                if m:
                    core_ports[net_prof].append((int(m.group(1)), int(m.group(2)), sw_role))

        elif sys_cat not in ('core', 'csl', 'cl') and sw_cat in ('core', 'csl', 'cl'):
            # Node/server connecting to core OR csl switch. L3 uplink profiles
            # (edge/OOB/storage-uplink) are switch-owned direct interfaces even
            # when authored with the external device on side A and the switch on
            # side B, so bucket those into the same path as core-as-system rows.
            core_rt = _classify_core_profile(net_prof, [sys_role, sys_name], l3_storage_profiles)
            prof = _profile_for_role(core_rt) if core_rt else {}
            prof_mode = str(prof.get('mode') or '').strip().lower()
            direct_l3 = (
                core_rt == 'edge'
                or (core_rt == 'oob' and oob_uplink_mode == 'l3')
                or (core_rt == 'storage_uplink' and prof_mode == 'l3')
            )
            if direct_l3:
                m = re.match(r'^swp(\d+)s(\d+)$', sw_port)
                if m:
                    role_core_entries[core_rt].append((int(m.group(1)), int(m.group(2)), sw_role))
                    continue
                m = re.match(r'^swp(\d+)$', sw_port)
                if m:
                    role_core_entries[core_rt].append((int(m.group(1)), 0, sw_role))
                    continue

            # Regular node/server bond into core/CSL.
            m = re.match(r'^swp(\d+)s(\d+)$', sw_port)
            if m:
                node_profiles[net_prof][int(m.group(1))].add(int(m.group(2)))
                continue
            # Non-broken-out switch port (bare `swpN`). Treated as sub-port 0,
            # mirroring the direct-L3 path above. Without this the row was
            # silently discarded: no role, no bond, no config, no warning — and
            # on a fabric with no breakout at all `network_roles` came out empty
            # and generation hard-failed far downstream with nothing pointing
            # back here.
            m = re.match(r'^swp(\d+)$', sw_port)
            if m:
                node_profiles[net_prof][int(m.group(1))].add(0)
                continue
            if sw_port:
                print(f"  ⚠️  Wire Map: switch port '{sw_port}' on profile "
                      f"'{net_prof}' is not a recognised swp port — row ignored.",
                      file=sys.stderr)

    # --- Step 3: Merge profiles of the same role type ---
    # Capture per-rail port lists BEFORE merging GPU profiles together,
    # so per-rail mode can emit one `access <vid>` line per rail instead
    # of bundling all GPU ports into the merged `gpu` bucket.
    rail_node_ports = defaultdict(lambda: defaultdict(set))  # rail_idx→base→{subs}
    if gpu_rail_vlan_ids:
        for prof, ports in node_profiles.items():
            m = re.match(r'^gpu[\s_-]*rail[\s_-]*(\d+)$', (prof or '').lower())
            if m:
                rail_idx = int(m.group(1))
                for base, subs in ports.items():
                    rail_node_ports[rail_idx][base] |= subs

    # Per-rail-per-plane variant: capture (rail, plane) → port lists from
    # "GPU Rail R Plane P" profile names. Used to emit one bridge-access
    # line per (rail, plane) VLAN on the switches.
    rail_plane_node_ports = defaultdict(lambda: defaultdict(set))  # (r,p)→base→{subs}
    if gpu_rail_plane_vlan_ids:
        for prof, ports in node_profiles.items():
            m = re.match(
                r'^gpu[\s_-]*rail[\s_-]*(\d+)[\s_-]*plane[\s_-]*(\d+)$',
                (prof or '').lower())
            if m:
                key = (int(m.group(1)), int(m.group(2)))
                for base, subs in ports.items():
                    rail_plane_node_ports[key][base] |= subs

    # node side: merge base_port→subport maps per role type
    role_node_ports = defaultdict(lambda: defaultdict(set))  # role_type→base→{subs}
    for prof, ports in node_profiles.items():
        rt = _classify_node_profile(prof)
        if rt:
            for base, subs in ports.items():
                role_node_ports[rt][base] |= subs

    # core side: merge entries per role type
    storage_l2_prof = None          # actual profile name for L2 storage-server ports
    storage_l2_entries = []
    for prof, entries in core_ports.items():
        rt = _classify_core_profile(prof, [e[2] for e in entries], l3_storage_profiles)
        if rt == 'storage':
            # L2 storage SERVER access ports authored core-as-system (switch on
            # side A). Keep the ACTUAL profile so breakout/native come from the
            # profile the row uses (e.g. 2-8-5-200 "Storage Uplink" Trunk
            # breakout 8, vs "Storage" breakout 4), not a role->profile remap.
            if storage_l2_prof is None:
                storage_l2_prof = prof
            storage_l2_entries.extend(entries)
        elif rt:
            role_core_entries[rt].extend(entries)

    # --- Step 4: Build output structures ---
    def _port_overrides(port_data, breakout):
        """Return port_overrides dict for ports with fewer active subports than breakout."""
        overrides = {}
        for base, subs in port_data.items():
            active = sorted(subs)
            if active != list(range(breakout)):
                overrides[base] = {'subports': active}
        return overrides

    network_roles = {}
    gpu_interfaces = None
    isl_interfaces = None
    edge_interfaces = None
    storage_interfaces = None
    oob_uplink_interfaces = None

    # Process node-to-core roles
    # Custom server roles (anything beyond cpu/gpu/support/storage) use
    # the Port Profile's own breakout/lanes columns when present, else
    # the same shape as a generic server bond (4x lanes-per-port 2).
    _CUSTOM_ROLE_HW_DEFAULT = {'breakout': 4, 'lanes': 2}
    for rt, port_data in role_node_ports.items():
        hw = _ROLE_HW.get(rt, _CUSTOM_ROLE_HW_DEFAULT)
        prof = _profile_for_role(rt)
        breakout = prof.get('breakout') or hw['breakout']
        lanes = prof.get('lanes') or hw['lanes']
        # Speed has no hardware default: the Excel column is the only source
        # (ADR-0035). Left unset when the cell is blank rather than derived,
        # so a blank cell stays visibly blank instead of being invented.
        speed = prof.get('speed')
        base_ports = sorted(port_data.keys())
        vlan = prof.get('native')
        if rt == 'oob' and not vlan:
            vlan = oob_vlan_id or 200

        if rt == 'gpu':
            gpu_interfaces = {
                'ports': base_ports,
                'breakout': breakout,
                'lanes': lanes,
                'vlan': vlan,
                'state': 'up',
                'port_overrides': {},
            }
            if speed:
                gpu_interfaces['speed'] = speed
        else:
            overrides = _port_overrides(port_data, breakout)
            # Trunk-bond multi-VLAN support: if the Port Profile has an
            # `Allowed VLANs` value (e.g. "200,400"), use that as the
            # tagged VLAN list and fall back to the native VLAN when
            # blank. The Excel column has been there forever but its
            # value was previously dropped, so trunk profiles always
            # rendered as `access <native>` even when allowed VLANs
            # were configured. The template emits `vlan <list>` when
            # both `vlan` and `vlan_untagged` are set.
            allowed_str = str(prof.get('allowed') or '').strip()
            allowed_list = [v.strip() for v in allowed_str.split(',') if v.strip()]
            if allowed_list:
                role_vlan = ','.join(allowed_list)
            else:
                role_vlan = vlan
            role_cfg = {
                'ports': base_ports,
                'breakout': breakout,
                'lanes': lanes,
                'vlan': role_vlan,
                'lacp_bypass': prof.get('lacp_bypass', False),
                'port_overrides': overrides,
                'bond_overrides': {},
            }
            # If the profile has an explicit `Untagged VLAN`, use it.
            # Otherwise, when allowed VLANs are present (trunk mode) the
            # native VLAN becomes the untagged value — this is how prod
            # configs express "trunk + native". When allowed is empty,
            # leave vlan_untagged unset so the template falls back to
            # the access-mode path.
            if prof.get('untagged'):
                role_cfg['vlan_untagged'] = int(prof['untagged'])
            elif allowed_list and vlan:
                role_cfg['vlan_untagged'] = int(vlan)
            if prof.get('vrf'):
                role_cfg['vrf'] = prof['vrf']
            if speed:
                role_cfg['speed'] = speed
            if prof.get('auto_negotiate'):
                role_cfg['auto_negotiate'] = prof['auto_negotiate']
            network_roles[rt] = role_cfg

    # Process core-as-system roles
    for rt, entries in role_core_entries.items():
        hw = _ROLE_HW[rt]
        prof = _profile_for_role(rt)
        breakout = prof.get('breakout') or hw['breakout']
        lanes = prof.get('lanes') or hw['lanes']
        speed = prof.get('speed')
        port_data = defaultdict(set)
        for base, sub, _ in entries:
            port_data[base].add(sub)
        base_ports = sorted(port_data.keys())
        overrides = _port_overrides(port_data, breakout)

        if rt == 'isl':
            isl_cfg = {
                'ports': base_ports,
                'breakout': breakout,
                'lanes': lanes,
                'port_overrides': overrides if overrides else {},
            }
            if prof.get('vrf'):
                isl_cfg['vrf'] = prof['vrf']
            if speed:
                isl_cfg['speed'] = speed
            isl_interfaces = isl_cfg
        elif rt == 'oob':
            if oob_uplink_mode == 'l3':
                oob_uplink_interfaces = {
                    'ports': base_ports,
                    'breakout': breakout,
                    'lanes': lanes,
                    'port_overrides': overrides if overrides else {},
                }
                if speed:
                    oob_uplink_interfaces['speed'] = speed
            else:
                network_roles['oob'] = {
                    'ports': base_ports,
                    'breakout': breakout,
                    'lanes': lanes,
                    'vlan': prof.get('native') or oob_vlan_id or 200,
                    'lacp_bypass': prof.get('lacp_bypass', False),
                    'port_overrides': overrides,
                    'bond_overrides': {},
                }
                if speed:
                    network_roles['oob']['speed'] = speed
        elif rt == 'edge':
            edge_cfg = {
                'ports': base_ports,
                'breakout': breakout,
                'lanes': lanes,
                'port_overrides': overrides,
            }
            # Edge uplinks are the EXIT-VRF customer-edge links — the EXIT VRF
            # BGP peers them (see the EXIT neighbor seeding in
            # _supplement_vrf_config_from_excel / _sync_edge_vrf_neighbors). For
            # unnumbered eBGP the interface must live in the SAME VRF as the BGP
            # instance, so an edge profile VRF of 'default' (a stale generator
            # default) leaves every EXIT session idle. Force EXIT unless the
            # profile names a real non-default VRF.
            _edge_vrf = (prof.get('vrf') or '').strip()
            edge_cfg['vrf'] = _edge_vrf if (_edge_vrf and _edge_vrf.lower() != 'default') else 'EXIT'
            if speed:
                edge_cfg['speed'] = speed
            edge_interfaces = edge_cfg
        elif rt == 'storage_uplink':
            # If any storage_uplink port is already claimed by another role
            # (e.g. support on port 26 in 2-4-3-200), inherit that role's
            # breakout/lanes — a physical port can only have one breakout mode.
            effective_breakout = breakout
            effective_lanes = lanes
            effective_speed = speed
            for p in base_ports:
                for existing_role in network_roles.values():
                    if p in existing_role.get('ports', []):
                        effective_breakout = existing_role['breakout']
                        effective_lanes = existing_role['lanes']
                        # Speed follows the inherited geometry for the same
                        # reason breakout does — one physical port has one
                        # sub-port rate, so the sharing role's speed wins.
                        effective_speed = existing_role.get('speed') or speed
                        break
            overrides = _port_overrides(port_data, effective_breakout)
            storage_interfaces = {
                'ports': base_ports,
                'breakout': effective_breakout,
                'lanes': effective_lanes,
                'vlan': prof.get('native'),
                'lacp_bypass': prof.get('lacp_bypass', False),
                'port_overrides': overrides,
                'bond_overrides': {},
            }
            if effective_speed:
                storage_interfaces['speed'] = effective_speed

    # Add storage_uplink into network_roles as 'storage' — but only when
    # the Storage Uplink Port Profile is L2 (Trunk/Access/Hybrid). When
    # the profile is L3 (STORAGE VRF external uplink), the ports are
    # bound to a VRF as direct L3 interfaces instead of bonded into the
    # bridge. Skip the role assignment so we don't generate `bondNs0`
    # bonds + `swpNs0 vrf STORAGE` conflicting emissions.
    # See docs/plans/2026-05-19-storage-vrf-design.md (PR-c).
    # ERA-44: L2 storage SERVER access ports authored core-as-system (switch on
    # side A) get the SAME L2 treatment as 'support' -- bond into the storage
    # VLAN with breakout/native/lacp from the profile the rows actually use.
    # Genuine L3 storage uplinks (profile in l3_storage_profiles -> role
    # 'storage_uplink') are bound to VRF STORAGE via the host_vars path and are
    # intentionally NOT bonded here. Node-loop storage (set above) wins if present.
    if storage_l2_entries and 'storage' not in network_roles:
        s_prof = profile_config.get(storage_l2_prof, {})
        s_hw = _ROLE_HW.get('storage', {'breakout': 4, 'lanes': 2})
        s_breakout = s_prof.get('breakout') or s_hw['breakout']
        s_lanes = s_prof.get('lanes') or s_hw['lanes']
        s_speed = s_prof.get('speed')
        s_port_data = defaultdict(set)
        for base, sub, _ in storage_l2_entries:
            s_port_data[base].add(sub)
        s_base_ports = sorted(s_port_data.keys())
        # Shared-port breakout inheritance (e.g. support on port 26 in
        # 2-4-3-200): a physical port can only carry one breakout mode.
        for p in s_base_ports:
            for existing_role in network_roles.values():
                if p in existing_role.get('ports', []):
                    s_breakout = existing_role['breakout']
                    s_lanes = existing_role['lanes']
                    s_speed = existing_role.get('speed') or s_speed
                    break
        network_roles['storage'] = {
            'ports': s_base_ports,
            'breakout': s_breakout,
            'lanes': s_lanes,
            'vlan': s_prof.get('native'),
            'lacp_bypass': s_prof.get('lacp_bypass', False),
            'port_overrides': _port_overrides(s_port_data, s_breakout),
            'bond_overrides': {},
        }
        if s_speed:
            network_roles['storage']['speed'] = s_speed

    result = {
        'network_roles': network_roles,
        'interfaces_disabled': sorted(set(disabled_ports)),
    }
    if gpu_interfaces:
        result['gpu_interfaces'] = gpu_interfaces
    if isl_interfaces:
        result['isl_interfaces'] = isl_interfaces
    if edge_interfaces:
        result['edge_interfaces'] = edge_interfaces
    if oob_uplink_interfaces:
        result['oob_uplink_interfaces'] = oob_uplink_interfaces

    # Per-rail GPU interfaces — one entry per rail (or per rail+plane) with
    # its own port list and VLAN ID. Derived from Wire Map "GPU Rail <N>"
    # or "GPU Rail R Plane P" profiles + the matching VLAN rows. The core
    # template iterates this dict and emits one `bridge domain br_default
    # access <vid>` line per (rail, plane).
    gpu_hw = _ROLE_HW.get('gpu', {})
    gpu_prof = _profile_for_role('gpu')
    rail_breakout = gpu_prof.get('breakout') or gpu_hw.get('breakout', 1)
    rail_lanes = gpu_prof.get('lanes') or gpu_hw.get('lanes', 1)
    # Per-rail GPU ports carry the same per-lane rate as the single-VLAN GPU
    # profile they were split out of (ERA-55) — without it every rail renders
    # breakout with no `link speed`.
    rail_speed = gpu_prof.get('speed') or gpu_hw.get('speed')
    gpu_rail_interfaces = {}

    # Per-rail entries: key = 'rail<R>'
    if rail_node_ports and gpu_rail_vlan_ids:
        for rail_idx, port_data in sorted(rail_node_ports.items()):
            vid = gpu_rail_vlan_ids.get(rail_idx)
            if not vid:
                continue
            base_ports = sorted(port_data.keys())
            gpu_rail_interfaces[f'rail{rail_idx}'] = {
                'ports': base_ports,
                'breakout': rail_breakout,
                'lanes': rail_lanes,
                'vlan': vid,
                'state': 'up',
                'port_overrides': _port_overrides(port_data, rail_breakout),
            }
            if rail_speed:
                gpu_rail_interfaces[f'rail{rail_idx}']['speed'] = rail_speed

    # Per-rail-per-plane entries: key = 'rail<R>_plane<P>'
    if rail_plane_node_ports and gpu_rail_plane_vlan_ids:
        for (rail_idx, plane_idx), port_data in sorted(rail_plane_node_ports.items()):
            vid = gpu_rail_plane_vlan_ids.get((rail_idx, plane_idx))
            if not vid:
                continue
            base_ports = sorted(port_data.keys())
            gpu_rail_interfaces[f'rail{rail_idx}_plane{plane_idx}'] = {
                'ports': base_ports,
                'breakout': rail_breakout,
                'lanes': rail_lanes,
                'vlan': vid,
                'state': 'up',
                'port_overrides': _port_overrides(port_data, rail_breakout),
            }
            if rail_speed:
                gpu_rail_interfaces[f'rail{rail_idx}_plane{plane_idx}']['speed'] = rail_speed

    if gpu_rail_interfaces:
        result['gpu_rail_interfaces'] = gpu_rail_interfaces

    return result


def segment_esi_for_node(node_name):
    """Globally-unique EVPN-MH segment local-id for a multihomed endpoint.

    On the dedicated multi-leaf N/S tier (cl/cs), the
    ESI is derived from the ENDPOINT identity — not the switch port — so a
    multihomed node carries the same ESI on every leaf that homes it (port-based
    `port*10+sub` would mismatch across leaves and collide across leaf-pairs).
    The shared `es-sys-mac` (group var `mh_mac`) plus this per-node local-id form
    the Type-3 ESI. Distinct numeric bands per endpoint class avoid collisions;
    all values stay well within NVUE's local-id range (1..16777215).
    """
    n = (node_name or '').strip().lower()
    m = re.match(r'su-(\d+)-node-(\d+)', n)
    if m:
        # compute: su*1000+node -> 1001..N, unique per (su, node)
        return int(m.group(1)) * 1000 + int(m.group(2))
    m = re.match(r'storage-(\d+)', n)
    if m:
        return 900000 + int(m.group(1))
    m = re.match(r'support-(\d+)', n)
    if m:
        return 800000 + int(m.group(1))
    m = re.match(r'gpu-(\d+)', n)   # flat gpu-NN naming fallback
    if m:
        return 100000 + int(m.group(1))
    # Last-resort stable mapping for any other endpoint name.
    return 700000 + (int(hashlib.sha1(n.encode(), usedforsecurity=False).hexdigest(), 16) % 90000)


def build_per_switch_server_roles(ws_wiremap, aggregated_network_roles,
                                  nodes_function_map=None,
                                  dedicated_ns_tier=False):
    """Per-switch `network_roles` for the dedicated N/S compute leaf tier.

    The shared group `network_roles` (group_vars/csl.yml) is the UNION of every
    leaf's server sub-ports; at scale a leaf that does not cable a given sub-port
    still gets a bond enslaving it -> ifreload rolls the whole apply back. And
    the port-derived ESI mismatches across a node's leaves -> EVPN-MH breaks.

    This returns {switch_name: network_roles} carrying, for each dedicated leaf,
    ONLY the server-facing (cpu/storage/support) bonds it ACTUALLY cables, each
    with a per-node ESI (`bond_overrides[<bond>].segment_id`). Role metadata
    (breakout, vlan, lanes, lacp_bypass, vrf, ...) is copied from the aggregated
    config (uniform per role). Written into host_vars, overriding the group
    `network_roles`.

    Gating is by ROLE + TIER. Per-switch bonds/ESI apply ONLY on a dedicated N/S
    tier (`dedicated_ns_tier`, i.e. `ns_tiers > 1`); this returns `{}` immediately
    otherwise. Within a dedicated tier the role is derived by classifying each
    switch's own HOSTNAME through `canonical_category(sw_name)`, and only the N/S
    compute leaf role is selected — `canonical_category(sw_name) == 'csl'`, which
    covers both the `cl-*` (split) and legacy `csl-*` (converged-name) hostnames
    for the same role, so a leaf behaves identically whichever name it carries.
    Do NOT gate on the raw Function column: a split leaf's Function is the bare
    `cl`, and `canonical_category('cl')` returns `cl` (only the `cl-*`/`csl-*`
    *prefixed* forms resolve to `csl`), which would skip every leaf. A converged
    leaf (`ns_tiers == 1`, whether named `cl-*` like 2-4-5-800 or `csl-*` like
    2-8-9-800) keeps the shared group `network_roles` byte-identical. `cs-`/`gs-`
    spines resolve to other categories, carry no server bonds, and never match.
    """
    nodes_function_map = nodes_function_map or {}
    aggregated_network_roles = aggregated_network_roles or {}
    # Per-switch server bonds + per-node ESI apply ONLY on a DEDICATED N/S tier
    # (ns_tiers > 1). A converged tier (ns_tiers == 1) keeps the shared group
    # network_roles — byte-identical golden — regardless of whether its leaves
    # are named cl-* or csl-*. Gate on tier + role, never on hostname.
    if not dedicated_ns_tier:
        return {}
    wm_col_map = build_wiremap_column_map(ws_wiremap, sheet_kind='wiremap')
    # per_switch[sw][rt][base][sub] = connected node name
    per_switch = defaultdict(lambda: defaultdict(lambda: defaultdict(dict)))
    for row in range(2, ws_wiremap.max_row + 1):
        sys_role = _wm_cell_ws(ws_wiremap, row, wm_col_map, 'system_role')
        sys_name = _wm_cell_ws(ws_wiremap, row, wm_col_map, 'system_name') or sys_role
        net_prof = _wm_cell_ws(ws_wiremap, row, wm_col_map, 'network_profile')
        sw_name = _wm_cell_ws(ws_wiremap, row, wm_col_map, 'switch_name') \
            or _wm_cell_ws(ws_wiremap, row, wm_col_map, 'switch_role')
        sw_port = _wm_cell_ws(ws_wiremap, row, wm_col_map, 'switch_port')
        nic_port = _wm_cell_ws(ws_wiremap, row, wm_col_map, 'nic_port') or ''
        _nm = re.search(r'(\d+)\s*$', str(nic_port))
        nic_num = int(_nm.group(1)) if _nm else 1
        # Some labels carry an explicit adapter-slot marker (e.g. ADR-0061's
        # "ConnectX-7 SL1 P1" / "SL3 P1") where P1/P2 RESTART per slot instead
        # of numbering sequentially across the whole node. nic_num alone can't
        # tell SL1 P1 from SL3 P1 apart (both end in "1") -- see bond_idx below.
        _slot = re.search(r'SL(\d+)', str(nic_port), re.IGNORECASE)
        slot_num = int(_slot.group(1)) if _slot else None
        if not net_prof or not sw_name:
            continue
        # Dedicated N/S compute leaves only. Classify the switch's own hostname
        # via canonical_category: 'csl' covers the cl-* (split) and legacy csl-*
        # (converged-name) leaf hostnames, while cs-* spines resolve to 'cs' and
        # gl-/gs- to their plane categories — so spines/GPU carry no server bonds.
        # NOTE: pass sw_name as the value to classify. Passing a bare Function
        # like 'cl' would return 'cl' (canonical_category only maps the cl-*/csl-*
        # *prefixed* forms to 'csl'), silently skipping every leaf and stranding
        # them on the shared group-union network_roles → phantom bonds on
        # uncabled ports → ifreload rollback.
        if canonical_category(sw_name) != 'csl':
            continue
        rt = _classify_node_profile(net_prof)
        if rt not in ('cpu', 'storage', 'support'):
            continue
        m = re.match(r'^swp(\d+)s(\d+)$', sw_port)
        if not m:
            continue
        base, sub = int(m.group(1)), int(m.group(2))
        per_switch[sw_name][rt][base][sub] = (sys_name, nic_num, slot_num)

    result = {}
    for sw, roles in per_switch.items():
        slots_by_node = defaultdict(set)
        for ports in roles.values():
            for submap in ports.values():
                for node, _nic_num, slot_num in submap.values():
                    if slot_num is not None:
                        slots_by_node[node].add(slot_num)
        slot_indices = {
            node: {slot: index for index, slot in enumerate(sorted(slots))}
            for node, slots in slots_by_node.items()
        }
        network_roles = {}
        for rt, ports in roles.items():
            meta = aggregated_network_roles.get(rt)
            if not meta:
                continue
            breakout = meta.get('breakout') or 4
            base_ports = sorted(ports.keys())
            port_overrides = {}
            bond_overrides = {}
            for base, submap in ports.items():
                active = sorted(submap.keys())
                if active != list(range(breakout)):
                    port_overrides[base] = {'subports': active}
                for sub, (node, nic_num, slot_num) in submap.items():
                    # A multi-bond node (e.g. storage with bond0+bond1) lands on
                    # several sub-ports; keying the ESI on node name alone gave all
                    # its bonds the SAME local-id -> identical ESI -> FRR rejects
                    # ("ESI already exists on a different interface") -> 0 Ethernet
                    # Segments -> EVPN-MH protodowns every server bond -> in-band dead.
                    # Offset by the node-local bond index so it is identical on both
                    # leaves of the pair but distinct per bond. bond_idx>0 only for
                    # multi-bond nodes, so single-bond cpu/support/gpu ESIs are
                    # unchanged.
                    #
                    # Two label conventions exist for "which physical NIC pair is
                    # this port part of":
                    #  - Sequential across the whole node (storage's real labels:
                    #    "NIC Port 1".."NIC Port 4") -> nic_num 1-2 is bond0,
                    #    3-4 is bond1. bond_idx = (nic_num-1)//2 handles this.
                    #  - Per-adapter-slot, restarting P1/P2 on each slot (ADR-0061
                    #    support: "ConnectX-7 SL1 P1/P2", "SL3 P1/P2"). Here
                    #    nic_num alone repeats (SL1 P1 and SL3 P1 both end in "1"),
                    #    so SL1's and SL3's bonds collided on the same ESI until
                    #    this was root-caused live on 2-4-5-800/largescale
                    #    (cl-01..cl-07 all rejected "ESI already exists on a
                    #    different interface" and never configured). Rank the
                    #    node's declared slots so its first slot remains bond 0
                    #    and later slots receive stable, distinct offsets.
                    bond_idx = slot_indices[node][slot_num] if slot_num is not None else (nic_num - 1) // 2
                    bond_overrides[f'bond{base}s{sub}'] = {
                        'segment_id': segment_esi_for_node(node) + bond_idx * 1_000_000,
                    }
            role_cfg = dict(meta)            # copy uniform role metadata
            role_cfg['ports'] = base_ports
            role_cfg['port_overrides'] = port_overrides
            role_cfg['bond_overrides'] = bond_overrides
            network_roles[rt] = role_cfg
        result[sw] = network_roles
    return result


def build_per_switch_direct_interfaces(
        ws_wiremap, aggregated_port_config, nodes_function_map=None,
        oob_uplink_mode='l2'):
    """Return actual direct-interface blocks for each core/N/S leaf switch.

    Group port blocks are unions across the Wire Map.  That is safe only when
    every switch has identical cabling.  Source topologies with an odd number
    of N/S leaves distribute edge and OOB links unevenly, so applying the union
    creates BGP neighbors on uncabled ports.  Host vars override those unions
    with each switch's real ISL, edge, and L3 OOB interfaces.
    """
    aggregated_port_config = aggregated_port_config or {}
    nodes_function_map = nodes_function_map or {}
    block_by_role = {
        'isl': 'isl_interfaces',
        'edge': 'edge_interfaces',
        'oob': 'oob_uplink_interfaces',
    }
    enabled_blocks = {
        role: key for role, key in block_by_role.items()
        if isinstance(aggregated_port_config.get(key), dict)
    }
    if not enabled_blocks:
        return {}

    target_switches = {
        name for name in nodes_function_map
        if canonical_category(name) in ('core', 'csl')
    }
    per_switch = {
        switch: {
            key: defaultdict(set)
            for key in enabled_blocks.values()
        }
        for switch in target_switches
    }
    rows = _build_wiremap_row_list(
        ws_wiremap, None, nodes_function_map=nodes_function_map)
    subport_re = re.compile(r'^swp(\d+)s(\d+)$')
    bare_re = re.compile(r'^swp(\d+)$')

    for row in rows:
        profile = row.get('net_profile')
        if not profile:
            continue
        role = _classify_core_profile(
            profile,
            [row.get('system_role'), row.get('switch_role')],
            set(),
        )
        if role == 'oob' and str(oob_uplink_mode).lower() != 'l3':
            continue
        key = enabled_blocks.get(role)
        if not key:
            continue

        endpoints = (
            (row.get('system_name'), row.get('system_role'), row.get('nic_port')),
            (row.get('switch_name'), row.get('switch_role'), row.get('switch_port')),
        )
        for name, function, port in endpoints:
            switch = name or function
            if switch not in per_switch:
                continue
            match = subport_re.fullmatch(str(port or ''))
            if match:
                parent, subport = (int(value) for value in match.groups())
            else:
                match = bare_re.fullmatch(str(port or ''))
                if not match:
                    continue
                parent, subport = int(match.group(1)), 0
            per_switch[switch][key][parent].add(subport)

    result = {}
    for switch, blocks in per_switch.items():
        result[switch] = {}
        for key, port_data in blocks.items():
            meta = dict(aggregated_port_config[key])
            breakout = int(meta.get('breakout') or 1)
            meta['ports'] = sorted(port_data)
            meta['port_overrides'] = {
                parent: {'subports': sorted(subports)}
                for parent, subports in port_data.items()
                if sorted(subports) != list(range(breakout))
            }
            result[switch][key] = meta
    return result


def parse_gsl_port_config(ws_wiremap, nodes_function_map=None):
    """Derive per-host port configuration for GSL switches from the Wire Map.

    GSLs are GPU spine/leaf switches in dedicated_gpu designs (e.g. 2-8-9-800).
    Per-plane independent fabric: GSL hosts sw_role values are gsl-plane1-NN
    and gsl-plane2-NN. Wire Map rows for GSL fall into two shapes:

      A. Node→GSL GPU connections (sys_role=gpu-NN, sw_role=gsl-planeN-NN):
         The Switch Port column is the GSL's own port (often a sub-port like
         swp1s0). These ARE the GPU access ports we need to bridge.
      B. GSL-to-GSL ISL connections (sys_role=gsl-planeN-NN, sw_role=gsl-planeN-NM):
         If present, they list internal-ISL trunk ports — drive BGP unnumbered
         peer-group "internal_isl" plus a 2x breakout on those parents.

    For each GSL hostname we return a dict with the following keys (all
    optional; absent keys mean "no config of that kind for this host"):
      gpu_subports          — comma-separated list of breakout sub-ports to bridge
                              (e.g. "swp1s0,swp2s0,swp3s0,swp4s0")
      gpu_breakout_parents  — comma-separated parent ports that need 2x breakout
                              (e.g. "swp1,swp2,swp3,swp4"). Only parents that
                              actually have sub-port wiremap rows.
      isl_subports          — comma-separated ISL sub-ports (BGP unnumbered)
      isl_breakout_parents  — comma-separated ISL parents needing 2x breakout

    Only sub-ports that are present in the wiremap (and therefore in the Air
    topology) are emitted. The rendered NVUE config never references a port
    that doesn't exist in topology, so `nv config apply` won't roll back.
    """
    # host -> { 'gpu': {parent: set(subs)}, 'isl': {parent: set(subs)},
    #           'rail': {rail_idx: {parent: set(subs)}},
    #           'rail_plane': {(rail_idx, plane_idx): {parent: set(subs)}} }
    by_host = defaultdict(lambda: {
        'gpu': defaultdict(set),
        'isl': defaultdict(set),
        'rail': defaultdict(lambda: defaultdict(set)),
        'rail_plane': defaultdict(lambda: defaultdict(set)),
    })

    sub_re = re.compile(r'^swp(\d+)s(\d+)$')
    bare_re = re.compile(r'^swp(\d+)$')
    rail_prof_re = re.compile(
        r'^gpu[\s_-]*rail[\s_-]*(\d+)$', re.IGNORECASE)
    rail_plane_prof_re = re.compile(
        r'^gpu[\s_-]*rail[\s_-]*(\d+)[\s_-]*plane[\s_-]*(\d+)$', re.IGNORECASE)

    # Use _build_wiremap_row_list so we get the same nodes-function-map
    # cascade that fills in System Role / Switch Role from the Nodes tab
    # when those columns are blank or missing. Without it, Wire Maps that
    # rely on cascade (e.g. 2-8-9-800) skip every row.
    rows = _build_wiremap_row_list(ws_wiremap, None, nodes_function_map=nodes_function_map)

    for row_dict in rows:
        sys_role = row_dict.get('system_role')
        sys_name = row_dict.get('system_name') or sys_role
        nic_port = row_dict.get('nic_port')
        net_prof = row_dict.get('net_profile')
        sw_role  = row_dict.get('switch_role')
        sw_name  = row_dict.get('switch_name') or sw_role
        sw_port  = row_dict.get('switch_port')

        if not sys_role or not net_prof:
            continue

        # Skip Air-only rows (mgmt, etc.) and disabled rows
        prof_lower = net_prof.lower()
        if prof_lower.startswith('air -'):
            continue
        if 'disabled' in prof_lower or 'unused' in prof_lower:
            continue

        # Category checks via canonical_category — accepts both legacy
        # hostname-as-role and post-step-4b canonical role strings.
        sys_cat = canonical_category(sys_role, sys_name)
        sw_cat  = canonical_category(sw_role, sw_name)
        is_gsl = lambda c: c in ('gsl-plane1', 'gsl-plane2', 'gsl',
                              'gl-plane1', 'gl-plane2',
                              'gs-plane1', 'gs-plane2')

        # Use sw_name / sys_name as the per-host keys — sw_role becomes
        # ambiguous after canonical conversion (every GSL on a plane shares
        # the same canonical role string).

        # Case A: node→GSL (sys is not gsl-, switch is gsl-)
        if is_gsl(sw_cat) and not is_gsl(sys_cat):
            m_sub = sub_re.match(sw_port)
            m_bare = bare_re.match(sw_port)
            parent_int = None
            sub_int = None
            if m_sub:
                parent_int = int(m_sub.group(1))
                sub_int = int(m_sub.group(2))
                by_host[sw_name]['gpu'][parent_int].add(sub_int)
            elif m_bare:
                # Treat unbroken parent as its own port (sub = -1 sentinel)
                parent_int = int(m_bare.group(1))
                sub_int = -1
                by_host[sw_name]['gpu'][parent_int].add(sub_int)

            # Per-rail modes: also bucket this port by rail so host_vars can
            # produce one bridge-access line per rail VLAN rather than lumping
            # all GPU ports into the legacy vlan900 block.
            if parent_int is not None and net_prof:
                profile = net_prof.strip()
                m_rp = rail_plane_prof_re.match(profile)
                if m_rp:
                    rp_key = (int(m_rp.group(1)), int(m_rp.group(2)))
                    by_host[sw_name]['rail_plane'][rp_key][parent_int].add(sub_int)
                else:
                    m_r = rail_prof_re.match(profile)
                    if m_r:
                        rail_key = int(m_r.group(1))
                        by_host[sw_name]['rail'][rail_key][parent_int].add(sub_int)

        # Case B: GSL→GSL (both sides gsl-plane*); ISL trunk
        elif is_gsl(sys_cat) and is_gsl(sw_cat):
            # The "system side" is one GSL (uses NIC/Port column); record there.
            m_sub = sub_re.match(nic_port)
            m_bare = bare_re.match(nic_port)
            if m_sub:
                parent = int(m_sub.group(1))
                sub = int(m_sub.group(2))
                by_host[sys_name]['isl'][parent].add(sub)
            elif m_bare:
                parent = int(m_bare.group(1))
                by_host[sys_name]['isl'][parent].add(-1)
            # Also record the switch-side parent
            m_sub2 = sub_re.match(sw_port)
            m_bare2 = bare_re.match(sw_port)
            if m_sub2:
                parent = int(m_sub2.group(1))
                sub = int(m_sub2.group(2))
                by_host[sw_name]['isl'][parent].add(sub)
            elif m_bare2:
                parent = int(m_bare2.group(1))
                by_host[sw_name]['isl'][parent].add(-1)

    # Build per-host output dicts
    result = {}
    for host, data in by_host.items():
        host_cfg = {}

        # GPU access section
        gpu = data['gpu']
        if gpu:
            gpu_subports = []
            gpu_breakout_parents = []
            for parent in sorted(gpu.keys()):
                subs = gpu[parent]
                # If only the bare port (-1) is present, it's not a breakout.
                if subs == {-1}:
                    gpu_subports.append(f'swp{parent}')
                else:
                    gpu_breakout_parents.append(f'swp{parent}')
                    for s in sorted(x for x in subs if x != -1):
                        gpu_subports.append(f'swp{parent}s{s}')
            if gpu_subports:
                host_cfg['gpu_subports'] = ','.join(gpu_subports)
            if gpu_breakout_parents:
                host_cfg['gpu_breakout_parents'] = ','.join(gpu_breakout_parents)

        # ISL section
        isl = data['isl']
        if isl:
            isl_subports = []
            isl_breakout_parents = []
            # A physical port may carry BOTH GPU and ISL traffic on different
            # sub-ports (mixed-role port). The `link breakout` command must be
            # emitted exactly once per port — so suppress ISL parents that
            # already appear in gpu_breakout_parents to avoid the template
            # rendering a duplicate `swpN` in the breakout list (NVUE last-wins,
            # silently dropping one of the breakout intents).
            already_gpu = set(host_cfg.get('gpu_breakout_parents', '').split(',')) \
                          if host_cfg.get('gpu_breakout_parents') else set()
            for parent in sorted(isl.keys()):
                subs = isl[parent]
                if subs == {-1}:
                    isl_subports.append(f'swp{parent}')
                else:
                    if f'swp{parent}' not in already_gpu:
                        isl_breakout_parents.append(f'swp{parent}')
                    for s in sorted(x for x in subs if x != -1):
                        isl_subports.append(f'swp{parent}s{s}')
            if isl_subports:
                host_cfg['isl_subports'] = ','.join(isl_subports)
            if isl_breakout_parents:
                host_cfg['isl_breakout_parents'] = ','.join(isl_breakout_parents)

        # Per-(rail, plane) sub-port lists. One entry per (rail, plane)
        # touching this GSL switch, used by the template to emit one
        # bridge-access line per rail. Key shape: 'rail<R>_plane<P>'.
        r = data['rail']
        if r:
            rail_subports = {}
            for rail_idx, ports in r.items():
                sub_strs = []
                for parent in sorted(ports.keys()):
                    subs = ports[parent]
                    if subs == {-1}:
                        sub_strs.append(f'swp{parent}')
                    else:
                        for s in sorted(x for x in subs if x != -1):
                            sub_strs.append(f'swp{parent}s{s}')
                if sub_strs:
                    rail_subports[f'rail{rail_idx}'] = ','.join(sub_strs)
            if rail_subports:
                host_cfg['gpu_rail_subports'] = rail_subports

        # Per-(rail, plane) sub-port lists. One entry per (rail, plane)
        # touching this GSL switch, used by the template to emit one
        # bridge-access line per rail-plane VLAN. Key shape:
        # 'rail<R>_plane<P>'.
        rp = data['rail_plane']
        if rp:
            rail_plane_subports = {}
            for (rail_idx, plane_idx), ports in rp.items():
                sub_strs = []
                for parent in sorted(ports.keys()):
                    subs = ports[parent]
                    if subs == {-1}:
                        sub_strs.append(f'swp{parent}')
                    else:
                        for s in sorted(x for x in subs if x != -1):
                            sub_strs.append(f'swp{parent}s{s}')
                if sub_strs:
                    rail_plane_subports[f'rail{rail_idx}_plane{plane_idx}'] = ','.join(sub_strs)
            if rail_plane_subports:
                host_cfg['gpu_rail_plane_subports'] = rail_plane_subports

        if host_cfg:
            result[host] = host_cfg

    return result


def _port_groups_to_breakout_config(port_groups):
    subports = []
    breakout_parents = []
    for parent in sorted(port_groups.keys()):
        subs = port_groups[parent]
        if subs == {-1}:
            subports.append(f'swp{parent}')
        else:
            breakout_parents.append(f'swp{parent}')
            for sub in sorted(x for x in subs if x != -1):
                subports.append(f'swp{parent}s{sub}')

    result = {}
    if subports:
        result['isl_subports'] = ','.join(subports)
    if breakout_parents:
        result['isl_breakout_parents'] = ','.join(breakout_parents)
    return result


def _derive_host_isl_port_config(wiremap_rows, host_name, profile_names=None):
    """Derive one host's ISL port config from normalized Wire Map rows."""
    if not wiremap_rows or not host_name:
        return {}

    profiles = {p.strip().lower() for p in (profile_names or ['isl'])}
    sub_re = re.compile(r'^swp(\d+)s(\d+)$')
    bare_re = re.compile(r'^swp(\d+)$')
    port_groups = defaultdict(set)

    for row in wiremap_rows:
        prof = (row.get('net_profile') or '').strip().lower()
        if prof not in profiles:
            continue
        if prof.startswith('air -') or 'disabled' in prof or 'unused' in prof:
            continue

        port = ''
        if (row.get('system_name') or '').strip() == host_name:
            port = (row.get('nic_port') or '').strip()
        elif (row.get('switch_name') or '').strip() == host_name:
            port = (row.get('switch_port') or '').strip()
        if not port:
            continue

        m_sub = sub_re.match(port)
        m_bare = bare_re.match(port)
        if m_sub:
            port_groups[int(m_sub.group(1))].add(int(m_sub.group(2)))
        elif m_bare:
            port_groups[int(m_bare.group(1))].add(-1)

    return _port_groups_to_breakout_config(port_groups)


# Nodes-tab management IPs start at the 9th host of the OOB subnet, so only the
# first 8 low addresses are safe for switch SVIs.
_OOB_SVI_LOW_SLOTS = 8


def _oob_switch_svi_ip(base, net_last, idx, prefix):
    """SVI address for the ``idx``-th OOB switch sharing one OOB subnet.

    The first 8 keep the historical low walk (``.2``, ``.3``, ... on a
    ``.0``-aligned /24). Past that, allocate DOWNWARD from the top of the
    subnet instead of continuing up.

    Why: this walk and the Nodes-tab management-IP walk share the subnet and
    used to advance head-on. Node mgmt IPs begin at the 9th host (``.10`` on a
    ``.0``-aligned /24), so the 9th switch onward landed on a real host's eth0.
    Four largescale sites each shipped 8 such duplicates -- e.g. ``cl-01``'s
    eth0 and ``oob-switch-09``'s SVI both ``192.168.200.10``. A duplicate on
    the OOB VLAN is an ARP/DAD war and ~60% packet loss to the colliding host,
    the exact failure ``scripts/oob_reserved.py`` exists to prevent for the
    fixed infra octets; it just never covered the SVIs.

    Growing downward makes the two walks recede from each other, reserves
    nothing up front (so small sites are untouched), and skips the reserved
    service octets so an SVI can never land on ``dhcp-oob`` (.252) or the
    air-mgmt SVI (.254).
    """
    if idx < _OOB_SVI_LOW_SLOTS:
        return f"{base}.{net_last + 2 + idx}"
    try:
        net = ipaddress.ip_network(f"{base}.{net_last}/{prefix}", strict=False)
    except ValueError:
        return f"{base}.{net_last + 2 + idx}"
    reserved = {str(net.network_address + o) for o in OOB_RESERVED_OCTETS}
    gw = str(net.network_address + net_last + 1)
    top = [str(h) for h in reversed(list(net.hosts()))
           if str(h) not in reserved and str(h) != gw]
    j = idx - _OOB_SVI_LOW_SLOTS
    # Exhausted: fall back to the legacy value rather than crash. validate_excel
    # gates duplicates, so this surfaces as an error instead of a silent overlap.
    return top[j] if j < len(top) else f"{base}.{net_last + 2 + idx}"


def _parse_cidr(subnet_str, *, context=""):
    """Safely parse a CIDR-style subnet string into (net_ip, prefix_int).

    Returns None if subnet_str doesn't look like a real CIDR. Helper exists
    because we have several spots that call subnet_str.split('/') unguarded
    — if a user types an OOB VLAN subnet as `'garbage'` (no `/`) the parser
    crashes with a confusing tuple-unpack error. Use this helper at all sites
    that need to parse subnet strings from Settings.
    """
    if not subnet_str or not isinstance(subnet_str, str):
        return None
    s = subnet_str.strip()
    if '/' not in s:
        if context:
            print(f"  ⚠️  {context}: '{s}' is not a valid CIDR (no '/'); skipping")
        return None
    parts = s.split('/')
    if len(parts) != 2:
        return None
    net_ip = parts[0].strip()
    try:
        prefix = int(parts[1].strip())
    except (ValueError, TypeError):
        if context:
            print(f"  ⚠️  {context}: prefix '{parts[1]}' is not an integer; skipping")
        return None
    if prefix < 0 or prefix > 32:
        return None
    return (net_ip, prefix)


def parse_settings(ws):
    """Parse the Settings sheet into a dictionary."""
    settings = {}
    for row in range(1, ws.max_row + 1):
        key = ws.cell(row=row, column=1).value
        value = ws.cell(row=row, column=2).value
        if key and value is not None:
            # Normalize key to snake_case
            key_clean = key.lower().replace(' ', '_').replace('-', '_')
            settings[key_clean] = value
    return settings


def resolve_role_version(versions: dict, role: str) -> str | None:
    """Look up a role's Cumulus version, tolerating the models' vocabulary.

    The VERSIONS table is keyed by the arch models' function names
    (`oob-switch`, `gsl-plane1`, `cs`, …) while callers here ask by
    classify_node() role (`oob`, `gsl`, `csl`). Only `core` is spelled the
    same in both, so a plain `versions.get('oob')` silently returned None and
    `cumulus_target_version` never reached the oob group_vars — the group_vars
    half of the same bug that left the topology on a stale image.
    """
    if not versions:
        return None
    direct = versions.get(role)
    if direct:
        return direct
    for func, version in versions.items():
        if _classify_node(str(func).strip().lower()) == role:
            return version
    return None


def parse_versions(ws):
    """Parse the VERSIONS table from the Settings sheet.

    Looks for a row with 'Switch Function' in column 1 and reads the rows
    below it as function → cumulus_version pairs.

    Returns dict: {'core': '5.16.1', 'oob': '5.15.1'} or {} if not found.
    """
    versions = {}
    in_versions = False
    for row in range(1, ws.max_row + 1):
        key = ws.cell(row=row, column=1).value
        if key is None:
            continue
        key_str = str(key).strip()
        if key_str.lower() == 'switch function':
            in_versions = True
            continue
        if in_versions:
            # Stop at blank row or a new section header (no value in col 2)
            value = ws.cell(row=row, column=2).value
            if not key_str or value is None:
                break
            versions[key_str.lower()] = str(value).strip()
    return versions


def parse_nodes(ws):
    """Parse the Nodes sheet into a list of node dictionaries.

    Each node gets:
      - role:     raw Function cell value (verbatim, for backward-compat
                  and for downstream code that still keys by hostname).
      - name:     hostname from Name column (falls back to role if blank).
      - category: canonical role category derived via canonical_category()
                  — preferred over `role` for category checks. Excel-first
                  (recognises canonical strings), with hostname-pattern
                  fallback for legacy Excels.
      - index:    integer instance index. Trailing-digit extraction from
                  Name first; falls back to position-among-same-category
                  for digitless names so `mycore`/`dog10` both work.
      - status, mac_address, mgmt_ip, prefix, gateway: as before.
    """
    nodes = []

    # Build column map from header row (row 1)
    col_map = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=c).value
        if val:
            col_map[str(val).strip().lower()] = c

    def _col(key, default):
        return col_map.get(key, default)

    func_col    = _col('function', 1)
    name_col    = _col('name', 2)
    mac_col     = _col('mac address for ztp', _col('mac address', _col('mac', 3)))
    mgmt_col    = _col('mgmt ip address', _col('mgmt ip', _col('management ip', 4)))
    prefix_col  = _col('prefix', 5)
    gateway_col = _col('gateway', 6)
    ztp_col     = _col('ztp', None)
    enabled_col = _col('enabled', None)
    oob_vlan_col = _col('oob vlan', None)

    for row in range(2, ws.max_row + 1):
        role = ws.cell(row=row, column=func_col).value
        if not role:
            continue

        # Check Enabled column — default to Active if column missing.
        # Three states:
        #   Yes / True / 1 / blank → Active (drives provisioning)
        #   No / False / 0         → Disabled (excluded from topology too)
        #   Air                    → Air-documentary (NOT provisioned, but
        #                            topology generator must still auto-
        #                            inject the corresponding Air-only
        #                            infrastructure — i.e. don't suppress
        #                            via disabled_names)
        enabled_val = str(ws.cell(row=row, column=enabled_col).value or 'Yes').strip().lower() if enabled_col else 'yes'
        is_active = enabled_val in ('yes', 'true', '1', '')
        is_air_documentary = (enabled_val == 'air')
        if is_active:
            status = 'Active'
        elif is_air_documentary:
            # Distinct from 'Disabled' so topology_generator's
            # disabled_names filter doesn't suppress the auto-injection
            # of the matching Air-only infra.
            status = 'Air'
        else:
            status = 'Disabled'

        name_val = ws.cell(row=row, column=name_col).value or role

        node = {
            'role': str(role).strip(),
            'name': name_val,
            'category': canonical_category(role, name_val),
            'index': extract_role_index(name_val),
            'status': status,
            'mac_address': ws.cell(row=row, column=mac_col).value or '',
            'mgmt_ip': ws.cell(row=row, column=mgmt_col).value or '',
            'prefix': ws.cell(row=row, column=prefix_col).value or 24,
            'gateway': ws.cell(row=row, column=gateway_col).value or '',
        }
        if ztp_col:
            node['ztp'] = ws.cell(row=row, column=ztp_col).value or ''
        oob_vlan_raw = ws.cell(row=row, column=oob_vlan_col).value if oob_vlan_col else None
        # Normalise to a bare id string ('201'), tolerating floats from Excel (201.0).
        if oob_vlan_raw is None or str(oob_vlan_raw).strip() == '':
            node['oob_vlan'] = ''
        elif isinstance(oob_vlan_raw, float) and oob_vlan_raw.is_integer():
            node['oob_vlan'] = str(int(oob_vlan_raw))
        else:
            node['oob_vlan'] = str(oob_vlan_raw).strip()
        nodes.append(node)

    # Second pass: assign index by order-among-same-category for nodes
    # whose Name had no trailing digits (e.g. `mycore`). 1-based.
    per_category_seq = {}
    for n in nodes:
        cat = n.get('category')
        if cat is None:
            continue
        per_category_seq.setdefault(cat, 0)
        per_category_seq[cat] += 1
        if n['index'] is None:
            n['index'] = per_category_seq[cat]

    return nodes


def _sanitize_scalar(value):
    """Collapse embedded control characters (newlines, CR, tab) in a free-text
    cell to single spaces and strip the ends.

    Free-text Excel fields (VLAN name, purpose, VRF) are rendered into
    generated config files such as dnsmasq.conf. An embedded newline would let
    an attacker who can edit the workbook break out of a comment line and inject
    a real directive. Sanitizing at read time neutralizes that without
    rejecting otherwise-valid input.
    """
    if value is None:
        return None
    if not isinstance(value, str):
        return value
    return re.sub(r'[\x00-\x1f\x7f]+', ' ', value).strip()


def parse_vlans(ws):
    """Parse VLANs from the VLANs & Profiles sheet (now with VRF and VNI columns)."""
    vlans = []
    
    # Find column indices from header row (row 2)
    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=2, column=col).value
        if val:
            # Header may contain multi-line help text (e.g. "DHCP Relay Client\n
            # (No or comma-separated VRF list)"). Match on the first line only
            # so operators can add tooltip-style descriptions without breaking
            # column lookup.
            first_line = str(val).splitlines()[0].strip()
            headers[first_line.lower().replace(' ', '_')] = col
    
    # Get column indices with defaults
    id_col = headers.get('vlan_id', 1)
    name_col = headers.get('name', 2)
    purpose_col = headers.get('purpose', 3)
    subnet_col = headers.get('subnet', 4)
    gateway_col = headers.get('gateway', 5)
    vrf_col = headers.get('vrf', 6)
    vni_col = headers.get('vni', None)
    relay_client_col = headers.get('dhcp_relay_client', None)

    # VLANs section starts at row 3 (after header row 2)
    for row in range(3, ws.max_row + 1):
        vlan_id = ws.cell(row=row, column=id_col).value
        if vlan_id is None or not isinstance(vlan_id, int):
            break  # End of VLAN section

        gw_cell = ws.cell(row=row, column=gateway_col).value
        vlan = {
            'id': vlan_id,
            'name': _sanitize_scalar(ws.cell(row=row, column=name_col).value),
            'purpose': _sanitize_scalar(ws.cell(row=row, column=purpose_col).value),
            'subnet': ws.cell(row=row, column=subnet_col).value,
            'gateway': str(gw_cell).strip() if gw_cell else None,
            'vrf': _sanitize_scalar(ws.cell(row=row, column=vrf_col).value) or 'default',
        }

        # VNI: use column value if present, else derive as VLAN_ID + 4000
        if vni_col:
            vni = ws.cell(row=row, column=vni_col).value
            vlan['vni'] = int(vni) if vni else (vlan_id + 4000)
        else:
            vlan['vni'] = vlan_id + 4000

        # DHCP Relay Client: comma-list of VRF names this VLAN relays to.
        # Blank/'No' means no relay for this VLAN.
        if relay_client_col:
            raw = ws.cell(row=row, column=relay_client_col).value
            vlan['dhcp_relay_client'] = str(raw).strip() if raw else ''
        else:
            vlan['dhcp_relay_client'] = ''

        vlans.append(vlan)

    return vlans


def parse_vrfs(ws):
    """Parse VRFs section from the VLANs & Profiles sheet."""
    vrfs = {}

    # Find VRFs section
    vrfs_row = None
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == 'VRFs':
            vrfs_row = row
            break

    if vrfs_row is None:
        return vrfs

    # Parse VRF data (starts 2 rows after header)
    for row in range(vrfs_row + 2, ws.max_row + 1):
        vrf_name = ws.cell(row=row, column=1).value
        if vrf_name is None or vrf_name == 'Port Profiles' or vrf_name == 'DHCP Relay':
            break

        vrfs[vrf_name] = {
            'name': vrf_name,
            'description': ws.cell(row=row, column=2).value,
            'l3_vni': ws.cell(row=row, column=3).value,
            'vlan': ws.cell(row=row, column=4).value,
        }

    return vrfs


def find_l3_storage_profiles(ws_vlans_profiles):
    """Return the set of Port Profile names with Port Mode=L3 AND VRF=STORAGE.

    Used by STORAGE VRF rollout (PR-b) to identify which Wire Map rows
    are external storage uplinks. Empty set when no such profile is
    declared — caller treats that as "no STORAGE VRF on this site."

    See docs/plans/2026-05-19-storage-vrf-design.md.
    """
    profiles = set()
    in_section = False
    header_row = None
    col_map = {}
    for row in range(1, ws_vlans_profiles.max_row + 1):
        val = ws_vlans_profiles.cell(row, 1).value
        if isinstance(val, str) and val.strip() == 'Port Profiles':
            in_section = True
            continue
        if not in_section:
            continue
        if not val:
            break
        val_str = str(val).strip()
        if val_str == 'DHCP Relay':
            break
        if val_str == 'Profile':
            header_row = row
            for c in range(1, ws_vlans_profiles.max_column + 1):
                h = ws_vlans_profiles.cell(row, c).value
                if h:
                    col_map[str(h).strip().lower()] = c
            continue
        if header_row is None:
            continue
        mode_val = ws_vlans_profiles.cell(row, col_map.get('port mode', 2)).value
        vrf_val = ws_vlans_profiles.cell(row, col_map.get('vrf', 6)).value
        if (mode_val and str(mode_val).strip().lower() == 'l3'
                and vrf_val and str(vrf_val).strip().upper() == 'STORAGE'):
            profiles.add(val_str)
    return profiles


MAX_SWITCH_PORT = 64


def collect_8x_ports(host_vars):
    """Every base port this switch puts into 8-way breakout, with its source.

    Three contributors, and they must all be counted: network_roles,
    edge_interfaces and storage_interfaces. The template's own adjacency check
    walks only the first two, which is why an 8x storage uplink landing beside
    the edge uplink went unnoticed.
    """
    found = {}

    def _scan(label, block, *, odd_only=False):
        if not isinstance(block, dict):
            return
        overrides = block.get('port_overrides') or {}
        for port in block.get('ports') or []:
            eff = (overrides.get(port) or {}).get('breakout', block.get('breakout'))
            if eff == 8 and (not odd_only or int(port) % 2 == 1):
                found.setdefault(int(port), label)

    for role_name, role in (host_vars.get('network_roles') or {}).items():
        _scan(f"network_roles.{role_name}", role)
    # Old shipped workbooks model OOB as 8x on even parents (for example
    # 2-4-5-800 swp8) and must remain parseable.  New hardware allocation emits
    # odd-parent 8x OOB blocks; derive their consumed neighbor without turning
    # the legacy workbook into a new hard failure.
    _scan(
        'oob_uplink_interfaces',
        host_vars.get('oob_uplink_interfaces'),
        odd_only=True,
    )
    _scan('edge_interfaces', host_vars.get('edge_interfaces'))
    _scan('storage_interfaces', host_vars.get('storage_interfaces'))
    return found


def assert_valid_8x_breakout(hostname, host_vars, group_port_blocks=None):
    """Hard-fail on an illegal 8x breakout layout.

    On Spectrum hardware an 8-way breakout consumes the lanes of the adjacent
    next-higher cage, so the layout is: ODD base broken out 8x, EVEN base+1
    disabled and never independently cabled.

    This is an exception, not a warning. The template used to emit the problem
    as a `# WARNING:` comment INTO the generated config, which nothing reads —
    so a config that silently kills a port shipped as if it were fine. The case
    that surfaced it: swp63 (storage) 8x beside swp64 (edge uplinks, EXIT VRF).
    swp63's breakout eats swp64's lanes, taking out the eBGP session to the
    customer edge — invisible until the fabric has no route off-cluster.
    """
    # network_roles / edge_interfaces / isl_interfaces live in the CORE group
    # vars; only storage_interfaces is per-host. The switch sees the merge, so
    # the check has to as well — host wins on conflict, as Ansible does.
    merged = dict(group_port_blocks or {})
    merged.update(host_vars)
    eightx = collect_8x_ports(merged)
    if not eightx:
        return
    disabled = {int(p) for p in (merged.get('interfaces_disabled') or [])}
    problems = []

    for port, source in sorted(eightx.items()):
        if port % 2 == 0:
            problems.append(
                f"swp{port} ({source}) is 8x on an EVEN port. 8x must sit on an "
                f"ODD base so the adjacent even port can be disabled; swp{port} "
                f"would consume swp{port + 1}"
                + (", which does not exist on this switch"
                   if port >= MAX_SWITCH_PORT else "")
                + ". Move it to an odd port."
            )
            continue
        neighbour = port + 1
        if neighbour in eightx:
            problems.append(
                f"swp{port} ({source}) and swp{neighbour} ({eightx[neighbour]}) "
                f"are both 8x. swp{port}'s breakout consumes swp{neighbour}'s "
                f"lanes, so swp{neighbour} cannot be broken out or cabled."
            )
        elif neighbour not in disabled:
            problems.append(
                f"swp{port} ({source}) is 8x but swp{neighbour} is not in "
                f"interfaces_disabled. The adjacent even port is consumed by the "
                f"breakout and must be explicitly disabled."
            )

    if problems:
        detail = "\n  - ".join(problems)
        raise ValueError(
            f"{hostname}: invalid 8x breakout layout\n  - {detail}\n"
            f"Fix the Wire Map / Port Profiles so every 8x port is an odd base "
            f"with its even neighbour disabled."
        )


def find_port_profile_hw(ws_vlans_profiles, profile_name):
    """{'breakout','lanes','speed'} for a named Port Profiles row, or {}.

    Sibling of find_storage_uplink_hw for callers that need one profile's
    electrical shape without the full parse. Columns are resolved by header
    name so an inserted column cannot silently shift the read.
    """
    header_row = None
    col_map = {}
    for row in range(1, ws_vlans_profiles.max_row + 1):
        val = ws_vlans_profiles.cell(row, 1).value
        if val and str(val).strip().lower() == 'profile':
            header_row = row
            for c in range(1, ws_vlans_profiles.max_column + 1):
                h = ws_vlans_profiles.cell(header_row, c).value
                if h:
                    col_map[str(h).strip().lower()] = c
            break
    if header_row is None:
        return {}
    for row in range(header_row + 1, ws_vlans_profiles.max_row + 1):
        name = ws_vlans_profiles.cell(row, 1).value
        if name in (None, ''):
            break
        if str(name).strip().lower() != str(profile_name).strip().lower():
            continue
        brk = ws_vlans_profiles.cell(row, col_map.get('breakout', 9)).value
        lanes = ws_vlans_profiles.cell(row, col_map.get('lanes', 10)).value
        spd = ws_vlans_profiles.cell(row, col_map.get('speed', 8)).value
        hw = {}
        if brk:
            hw['breakout'] = int(brk)
        if lanes:
            hw['lanes'] = int(lanes)
        if spd and str(spd).strip():
            hw['speed'] = str(spd).strip()
        return hw
    return {}


def find_storage_uplink_hw(ws_vlans_profiles):
    """Breakout/lanes for the L3 STORAGE uplink profile, straight from the sheet.

    The Port Profiles row is the declared shape of the port and must win. It
    had not been: generate_host_vars hardcoded 8x/1-lane with a comment saying a
    later PR would read this column, and that never landed. Where the sheet says
    `Storage Uplink | L3 | STORAGE | 100G | 4 | 2`, the emitted config said 8x —
    which on 2-8-9-800 put swp63 into 8-way breakout beside swp64's 8x edge
    uplinks. An 8x port consumes its even neighbour's lanes, so that silently
    took the cage carrying the EXIT-VRF eBGP session to the customer edge.

    Falls back to the _ROLE_HW default only when the sheet leaves the cells
    blank, so a workbook that declares nothing behaves as before.
    """
    default = dict(_ROLE_HW['storage_uplink'])
    in_section = False
    header_row = None
    col_map = {}
    for row in range(1, ws_vlans_profiles.max_row + 1):
        val = ws_vlans_profiles.cell(row, 1).value
        if isinstance(val, str) and val.strip() == 'Port Profiles':
            in_section = True
            continue
        if not in_section or not val:
            if in_section and not val:
                break
            continue
        val_str = str(val).strip()
        if val_str == 'DHCP Relay':
            break
        if val_str == 'Profile':
            header_row = row
            for c in range(1, ws_vlans_profiles.max_column + 1):
                h = ws_vlans_profiles.cell(row, c).value
                if h:
                    col_map[str(h).strip().lower()] = c
            continue
        if header_row is None:
            continue
        mode_val = ws_vlans_profiles.cell(row, col_map.get('port mode', 2)).value
        vrf_val = ws_vlans_profiles.cell(row, col_map.get('vrf', 6)).value
        if not (mode_val and str(mode_val).strip().lower() == 'l3'
                and vrf_val and str(vrf_val).strip().upper() == 'STORAGE'):
            continue
        brk = ws_vlans_profiles.cell(row, col_map.get('breakout', 9)).value
        lanes = ws_vlans_profiles.cell(row, col_map.get('lanes', 10)).value
        # Speed rides along with the geometry (ERA-55). Without it the STORAGE
        # uplink is the one broken-out role that reaches the template with no
        # per-lane rate, so its sub-ports render breakout but never `link speed`.
        spd = ws_vlans_profiles.cell(row, col_map.get('speed', 8)).value
        hw = {'breakout': int(brk) if brk else default['breakout'],
              'lanes': int(lanes) if lanes else default['lanes']}
        if spd and str(spd).strip():
            hw['speed'] = str(spd).strip()
        return hw
    return default


def get_storage_uplink_ports_per_switch(ws_wiremap, l3_storage_profiles,
                                          nodes_function_map=None,
                                          disabled_names=None):
    """Scan Wire Map for rows using an L3 STORAGE port profile.

    Returns {switch_name: [port_string, ...]} — e.g.
    {'csl-01': ['swp63s0', 'swp63s1'], 'csl-02': ['swp63s0', 'swp63s1']}.
    Empty dict if no rows match (any switch with zero storage uplinks
    is simply absent from the result, not present with an empty list).

    See docs/plans/2026-05-19-storage-vrf-design.md (STORAGE VRF rollout
    PR-b). Operates against the same `_build_wiremap_row_list` cascade
    so it honors nodes_function_map fallback + disabled-host filtering.
    """
    if not l3_storage_profiles:
        return {}
    rows = _build_wiremap_row_list(ws_wiremap, None,
                                    nodes_function_map=nodes_function_map,
                                    disabled_names=disabled_names)
    per_switch = defaultdict(list)
    # Storage uplink rows are commonly written with the CSL on the A-side
    # (system_name) and the external storage device on the B-side
    # (switch_name = ext-storage-NN). Detect the "switch we configure" by
    # picking whichever side is NOT an external sentinel.
    ext_re = re.compile(r'^ext[-_]', re.IGNORECASE)
    for r in rows:
        prof = (r.get('net_profile') or '').strip()
        if prof not in l3_storage_profiles:
            continue
        a_name, a_port = r.get('system_name'), r.get('nic_port')
        b_name, b_port = r.get('switch_name'), r.get('switch_port')
        # Pick whichever side is NOT an external sentinel.
        if b_name and not ext_re.match(b_name):
            switch, port = b_name, b_port      # normal: B-side is the switch
        elif a_name and not ext_re.match(a_name):
            switch, port = a_name, a_port      # reversed: A-side is the switch
        else:
            continue                            # both sides external — skip
        if switch and port and port not in per_switch[switch]:
            per_switch[switch].append(port)
    return dict(per_switch)


def get_bond_descriptions_per_switch(ws_wiremap, nodes_function_map=None,
                                       disabled_names=None, known_node_names=None):
    """Build {switch_name: {bond_name: peer_node_name}} from Wire Map.

    For each Wire Map row that pairs a switch port with a node port:
      - Switch side gives us the swp port → bond name (swpNsM → bondNsM)
      - Node side gives us the peer hostname → description

    Only emits descriptions when the peer is a real declared node
    (`known_node_names`). External-device sentinels (ext-*,
    cust-net-edge-*, SPARE*, outbound) get their own description via
    other code paths (e.g. PR-c emits 'External Uplink - <VRF> VRF'
    for storage uplinks). Virtual Air nodes (dhcp-oob, oob-server-01,
    dhcp-edge) are skipped too — they don't have a meaningful bond
    description.

    See docs/plans/2026-05-19-2-8-9-800-prod-feedback.md (DV4).
    """
    if not known_node_names:
        return {}
    nodes_function_map = nodes_function_map or {}
    rows = _build_wiremap_row_list(ws_wiremap, None,
                                    nodes_function_map=nodes_function_map,
                                    disabled_names=disabled_names)
    sub_re = re.compile(r'^swp(\d+)s(\d+)$')
    bare_re = re.compile(r'^swp(\d+)$')

    def _bond_name_for_port(port):
        m = sub_re.match(port or '')
        if m:
            return f'bond{m.group(1)}s{m.group(2)}'
        m = bare_re.match(port or '')
        if m:
            return f'bond{m.group(1)}'
        return None

    # Build "is a switch" predicate from nodes function map. Switches
    # (core/csl/gsl-planeN/oob-switch/edge) are in known_node_names too,
    # but for bond descriptions we want the SERVER hostname, never the
    # switch hostname.
    _SWITCH_FUNCS = {'core', 'csl', 'gsl', 'gsl-plane1', 'gsl-plane2',
                     'oob-switch', 'edge',
                     'cl', 'cs', 'gl-plane1', 'gl-plane2', 'gs-plane1', 'gs-plane2'}

    def _is_switch(name):
        return canonical_category(nodes_function_map.get(name, ''), name) in _SWITCH_FUNCS

    per_switch: dict = defaultdict(dict)
    for r in rows:
        a_name, a_port = r.get('system_name'), r.get('nic_port')
        b_name, b_port = r.get('switch_name'), r.get('switch_port')
        a_is_switch = _is_switch(a_name)
        b_is_switch = _is_switch(b_name)
        a_is_known_server = a_name in known_node_names and not a_is_switch
        b_is_known_server = b_name in known_node_names and not b_is_switch
        # Pair must be (switch, server) — one each side
        if a_is_switch and b_is_known_server:
            switch, sw_port, node = a_name, a_port, b_name
        elif b_is_switch and a_is_known_server:
            switch, sw_port, node = b_name, b_port, a_name
        else:
            continue
        bond = _bond_name_for_port(sw_port)
        if switch and bond and node:
            per_switch[switch].setdefault(bond, node)
    return dict(per_switch)


def parse_dhcp_relay_table(ws):
    """Parse the DHCP Relay table from the VLANs & Profiles sheet.

    Looks for a row with 'DHCP Relay' in column 1, then a header row with
    'Server IP', 'VRF', 'Upstream Interface'. Data rows follow until blank
    or a new named section.

    Returns list of dicts:
        [{'servers': ['192.168.200.252'], 'vrf': 'OOB',
          'upstream_interface': 'vlan200'}, ...]

    Empty list if section not found or has no data rows. Comma-separated
    server IPs in column 1 are split into the 'servers' list.
    """
    table = []

    # Find DHCP Relay section header
    header_row = None
    for row in range(1, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val and str(val).strip() == 'DHCP Relay':
            header_row = row
            break

    if header_row is None:
        return table

    # Data rows start 2 rows after the section header (skipping subheader)
    for row in range(header_row + 2, ws.max_row + 1):
        ip_cell = ws.cell(row=row, column=1).value
        if ip_cell is None or not str(ip_cell).strip():
            break
        # Stop at the next named section
        if str(ip_cell).strip() in ('VRFs', 'Port Profiles', 'VLANs'):
            break
        vrf_cell = ws.cell(row=row, column=2).value
        up_cell = ws.cell(row=row, column=3).value

        servers = [s.strip() for s in str(ip_cell).split(',') if s.strip()]
        if not servers or not vrf_cell:
            continue

        # Upstream Interface may be a comma-list — NVUE supports multiple
        # upstream-interface entries per server-group.
        up_raw = str(up_cell or '').strip()
        upstreams = [u.strip() for u in up_raw.split(',') if u.strip()]

        table.append({
            'servers': servers,
            'vrf': str(vrf_cell).strip().upper(),
            'upstream_interfaces': upstreams,
        })

    return table


_LOOPBACK_VRFS = ('OOB', 'INBAND', 'EXIT', 'GPU', 'STORAGE')


def _classify_loopback_header(header_text):
    """Return the canonical key for a Loopbacks-sheet column header, or None.

    Recognized (case-insensitive):
        Switch / Switch name / Hostname -> 'switch'
        Default / lo / Loopback         -> 'lo'
        OOB / OOB VRF                   -> 'OOB'
        INBAND / In-Band                -> 'INBAND'
        EXIT                            -> 'EXIT'
        GPU                             -> 'GPU'
    """
    if header_text is None:
        return None
    key = str(header_text).strip().lower()
    if not key:
        return None
    if key in ('switch', 'switch name', 'hostname'):
        return 'switch'
    if key in ('default', 'default (lo)', 'lo', 'lo_ip', 'underlay', 'loopback'):
        return 'lo'
    if key in ('asn', 'as', 'bgp asn', 'bgp_asn', 'autonomous system', 'as number'):
        return 'asn'
    for vrf in _LOOPBACK_VRFS:
        v = vrf.lower()
        if key == v or key == f'{v} vrf':
            return vrf
    if key == 'in-band':
        return 'INBAND'
    return None


def parse_loopbacks_sheet(ws):
    """Parse the optional 'Loopbacks' sheet — per-switch / per-VRF overrides.

        Switch | Default | OOB | INBAND | EXIT | GPU

    One row per switch. Returns a dict keyed by switch name:
        {
            'core-01': {
                'lo': '172.16.176.11/32',
                'OOB':    '172.16.176.1/32',
                'INBAND': '172.16.176.3/32',
                'EXIT':   '172.16.176.5/32',
                'GPU':    '192.168.110.5/32',
            },
            ...
        }

    Missing cells fall back to the parser's computed defaults. Returns {}
    if the sheet is absent or has no usable header row.
    """
    overrides = {}
    if ws is None or ws.max_row < 2:
        return overrides

    header_row = None
    for r in range(1, min(ws.max_row + 1, 5)):
        val = ws.cell(row=r, column=1).value
        if val and str(val).strip().lower().startswith('switch'):
            header_row = r
            break
    if header_row is None:
        return overrides

    col_map = {}
    for c in range(1, ws.max_column + 1):
        key = _classify_loopback_header(ws.cell(row=header_row, column=c).value)
        if key is not None:
            col_map[key] = c

    if 'switch' not in col_map:
        return overrides

    for row in range(header_row + 1, ws.max_row + 1):
        sw_cell = ws.cell(row=row, column=col_map['switch']).value
        if not sw_cell or not str(sw_cell).strip():
            continue
        sw_name = str(sw_cell).strip()
        entry = {}
        for key, col in col_map.items():
            if key == 'switch':
                continue
            v = ws.cell(row=row, column=col).value
            if v is not None and str(v).strip():
                entry[key] = str(v).strip()
        if entry:
            overrides[sw_name] = entry

    return overrides


def parse_prefix_lists_sheet(ws):
    """
    Parse the 'Prefix lists' sheet.
    Columns: List name, Rule id, Match (CIDR), Max prefix length, Action.
    Row 1 may be a merged note row; header is 'List name' row; data below.
    The Action column is optional (backward-compat with the older
    4-column layout, which had no directives at all).

    Action semantics:
      - blank ⇒ the row's rule is collected under its list_id in ``lists``.
      - 'suppress' (case-insensitive) ⇒ list_id is added to ``suppress``
        (any Match on that row is ignored); a suppressed list_id is
        removed from / never added to ``lists`` — suppress always wins,
        regardless of row order.

    This parser is list-set-agnostic: it does NOT decide override vs. add
    for a given list_id — that's generate_prefix_lists' job (based on
    whether the id already exists among the generated defaults, gated on
    the OVERRIDABLE_PREFIX_LISTS / DERIVED_ONLY_PREFIX_LISTS classification
    below). This sheet is entirely optional; an absent workbook, or an
    absent/empty sheet, produces byte-identical output to a workbook that
    never had it. See docs/EXCEL_CONFIGURATION_GUIDE.md ("Sheet 5: Routing
    Policy Overrides") for the operator-facing column reference.

    Returns: {'lists': {list_id: [ {id, match, max_len?}, ... ]},
              'suppress': set()}
    """
    result = {'lists': {}, 'suppress': set()}
    lists = result['lists']
    suppress = result['suppress']
    if ws.max_row < 2:
        return result
    # Find header row (row with "List name" in col 1); data starts next row
    header_row = 1
    for r in range(1, min(ws.max_row + 1, 5)):
        val = ws.cell(row=r, column=1).value
        if val and str(val).strip().lower() == 'list name':
            header_row = r
            break
    data_start = header_row + 1
    for row in range(data_start, ws.max_row + 1):
        list_name = ws.cell(row=row, column=1).value
        rule_id = ws.cell(row=row, column=2).value
        match_val = ws.cell(row=row, column=3).value
        if not list_name or not str(list_name).strip():
            continue
        list_id = str(list_name).strip()
        if list_id.lower() == 'list name':
            continue  # skip header row if it appears in data range
        action_val = ws.cell(row=row, column=5).value
        action = str(action_val).strip().lower() if action_val else ''
        if action == 'suppress':
            suppress.add(list_id)
            lists.pop(list_id, None)
            continue
        rule = {
            'id': str(rule_id).strip() if rule_id else str(row - 1),
            'match': str(match_val).strip() if match_val else '',
        }
        max_len_val = ws.cell(row=row, column=4).value
        if max_len_val is not None and str(max_len_val).strip():
            rule['max_len'] = str(int(max_len_val)) if isinstance(max_len_val, (int, float)) else str(max_len_val).strip()
        if not rule['match']:
            continue
        if list_id in suppress:
            continue  # suppress wins even if the suppress row came first
        if list_id not in lists:
            lists[list_id] = []
        lists[list_id].append(rule)
    return result


def parse_route_policy_sheet(ws):
    """
    Parse the optional 'Route policy' sheet.
    Columns: Route-map | Rule | Action | Match type | Match value |
             Set type | Set value.
    Row 1 may be a merged note row; header is the 'Route-map' row; data
    starts below it. A rule may span multiple rows (repeated Route-map /
    Rule / Action) to add additional match/set entries — rows for the same
    (route-map, rule id) are collected into one rule entry regardless of
    row order. Blank Match type/value ⇒ no match entry is appended for that
    row; same for Set type/value. Fully-blank rows and the header row (if it
    reappears in the data range) are skipped.

    This parser is route-map-agnostic — it does NOT decide override vs. add
    for a given route-map id; that's the caller's job via
    `_apply_named_overrides` (mirrors `parse_prefix_lists_sheet`). Unlike
    `parse_prefix_lists_sheet`, there is no suppress action and no
    derived-only protection — any route-map id can be overridden. Directives
    from this sheet are merged into the Core/CSL `group_vars` only (see
    `generate_group_vars`). Absent/empty sheet ⇒ byte-identical output. See
    docs/EXCEL_CONFIGURATION_GUIDE.md ("Sheet 5: Routing Policy Overrides").

    Returns: {route_map_id: [ {id, action, match: [...], set: [...]}, ... ]}
    matching the source-inventory `route_map` YAML shape exactly (a `match`
    or `set` key is present only when at least one row supplied it).
    """
    result = {}
    rule_index = {}  # (route_map_id, rule_id) -> rule entry dict
    if ws.max_row < 2:
        return result
    header_row = 1
    for r in range(1, min(ws.max_row + 1, 5)):
        val = ws.cell(row=r, column=1).value
        if val and str(val).strip().lower() == 'route-map':
            header_row = r
            break
    data_start = header_row + 1
    for row in range(data_start, ws.max_row + 1):
        rm_val = ws.cell(row=row, column=1).value
        if not rm_val or not str(rm_val).strip():
            continue
        rm_id = str(rm_val).strip()
        if rm_id.lower() == 'route-map':
            continue  # skip header row if it appears in data range
        rule_val = ws.cell(row=row, column=2).value
        rule_id = str(rule_val).strip() if rule_val else ''
        action_val = ws.cell(row=row, column=3).value
        action = str(action_val).strip().lower() if action_val else ''

        key = (rm_id, rule_id)
        rule_entry = rule_index.get(key)
        if rule_entry is None:
            rule_entry = {'id': rule_id, 'action': action}
            rule_index[key] = rule_entry
            result.setdefault(rm_id, []).append(rule_entry)

        match_type = ws.cell(row=row, column=4).value
        match_value = ws.cell(row=row, column=5).value
        if match_type and str(match_type).strip():
            rule_entry.setdefault('match', []).append({
                'type': str(match_type).strip(),
                'value': str(match_value).strip() if match_value else '',
            })
        set_type = ws.cell(row=row, column=6).value
        set_value = ws.cell(row=row, column=7).value
        if set_type and str(set_type).strip():
            rule_entry.setdefault('set', []).append({
                'type': str(set_type).strip(),
                'value': str(set_value).strip() if set_value else '',
            })
    return result


def parse_community_lists_sheet(ws):
    """
    Parse the optional 'Community lists' sheet.
    Columns: Community-list | Rule | Action | Community.
    Row 1 may be a merged note row; header is the 'Community-list' row;
    data starts below it. Fully-blank rows and the header row (if it
    reappears in the data range) are skipped.

    Same override/add-only semantics as `parse_route_policy_sheet` via
    `_apply_named_overrides` (no suppress, no derived-only protection),
    merged into the Core/CSL `group_vars` only. Absent/empty sheet ⇒
    byte-identical output. See docs/EXCEL_CONFIGURATION_GUIDE.md
    ("Sheet 5: Routing Policy Overrides").

    Returns: {community_list_id: [ {id, action, community}, ... ]}
    matching the source-inventory `community_list` YAML shape exactly.
    """
    result = {}
    if ws.max_row < 2:
        return result
    header_row = 1
    for r in range(1, min(ws.max_row + 1, 5)):
        val = ws.cell(row=r, column=1).value
        if val and str(val).strip().lower() == 'community-list':
            header_row = r
            break
    data_start = header_row + 1
    for row in range(data_start, ws.max_row + 1):
        cl_val = ws.cell(row=row, column=1).value
        if not cl_val or not str(cl_val).strip():
            continue
        cl_id = str(cl_val).strip()
        if cl_id.lower() == 'community-list':
            continue  # skip header row if it appears in data range
        rule_val = ws.cell(row=row, column=2).value
        action_val = ws.cell(row=row, column=3).value
        community_val = ws.cell(row=row, column=4).value
        result.setdefault(cl_id, []).append({
            'id': str(rule_val).strip() if rule_val else '',
            'action': str(action_val).strip().lower() if action_val else '',
            'community': str(community_val).strip() if community_val else '',
        })
    return result


def _apply_named_overrides(base_list, overrides):
    """Replace/add entries in a list of ``{'id': ..., 'rule': [...]}`` dicts
    by id. Shared by the `route_map` and `community_list`
    override-merge in `generate_group_vars` — both are lists of
    ``{'id': ..., 'rule': [...]}`` entries in the source-inventory YAML.

    ``base_list``: the source-inventory list (e.g. `core_vars['route_map']`).
    ``overrides``: ``{entry_id: rule_list}`` as returned by
    `parse_route_policy_sheet` / `parse_community_lists_sheet`.

    Entries whose id exists in ``overrides`` are replaced in place (list
    order preserved); ids not already present are appended in the order
    they first appear in ``overrides``. Returns a new list — ``base_list``
    is not mutated. Absent/empty ``overrides`` returns ``base_list``
    unchanged (derive-by-default).
    """
    if not overrides:
        return base_list
    seen = set()
    result = []
    for entry in base_list or []:
        entry_id = entry.get('id')
        if entry_id in overrides:
            result.append({'id': entry_id, 'rule': overrides[entry_id]})
            seen.add(entry_id)
        else:
            result.append(entry)
    for entry_id, rules in overrides.items():
        if entry_id not in seen:
            result.append({'id': entry_id, 'rule': rules})
    return result


def _svi_gateway_ip(subnet, gateway):
    """Anycast (VRR) gateway address for an SVI.

    Returns the operator-declared Excel ``gateway`` when present, otherwise the
    first usable host of ``subnet``. Bare IP string, no prefix. Falls back to
    the legacy ``<base>.1`` only when the subnet can't be parsed, so callers
    never crash on malformed input.
    """
    gw = (gateway or '').split('/')[0].strip()
    if gw:
        return gw
    try:
        return str(next(ipaddress.ip_network(subnet, strict=False).hosts()))
    except (ValueError, StopIteration):
        return f"{subnet.split('/')[0].rsplit('.', 1)[0]}.1"


def _svi_switch_ip(subnet, gateway, core_num):
    """Per-switch SVI host IP that always lands inside ``subnet``.

    ``core_num`` is 1-based (core-01 -> 1). Host IPs are taken from
    ``ip_network().hosts()`` with the gateway excluded, so a non-``.0``-aligned
    subnet (e.g. ``100.82.254.128/27``) or a non-``.1`` gateway no longer
    produces addresses in the wrong network. For a ``.0``-aligned /24 with a
    ``.1`` gateway this reproduces the legacy ``<base>.<1+core_num>`` scheme
    exactly, keeping the shipping default Excels byte-identical.

    Falls back to the legacy scheme only when the subnet can't be parsed.
    """
    try:
        net = ipaddress.ip_network(subnet, strict=False)
    except ValueError:
        return f"{subnet.split('/')[0].rsplit('.', 1)[0]}.{1 + core_num}"
    gw = _svi_gateway_ip(subnet, gateway)
    hosts = [str(h) for h in net.hosts() if str(h) != gw]
    idx = core_num - 1
    if 0 <= idx < len(hosts):
        return hosts[idx]
    # More switches than usable hosts (not expected for real mgmt subnets) —
    # wrap rather than crash; the result is still inside the subnet.
    return hosts[idx % len(hosts)] if hosts else str(net.network_address)


# Prefix list classification constants
# OVERRIDABLE_PREFIX_LISTS: global-subnet lists that operators may override via the
# Prefix lists Excel sheet; DERIVED_ONLY_PREFIX_LISTS: per-switch /32 lists that
# are computed and must never be user-overridable.
OVERRIDABLE_PREFIX_LISTS = frozenset({
    'ERA_PREFIXES',
    'INBAND_PREFIXES',
    'OOB_PREFIXES',
    'VTEP_PREFIXES',
    'ALL_PREFIXES',
    # ERA-52. Deliberately overridable: EVPN_OOB_OUT rule 20 denies every macip
    # NOT matched here, so a site whose BMCs sit outside the derived OOB subnet
    # must be able to widen this from the Prefix lists sheet. Without that, the
    # filter's failure mode is loss of OOB reachability with no way out short of
    # editing the tool.
    'OOB_HOSTS',
})
DERIVED_ONLY_PREFIX_LISTS = frozenset({
    'EXIT_LOCAL_IF',
    'INBAND_LOCAL_IF',
    'OOB_LOCAL_IF',
    'LOCAL_OOB_LOOPBACK',
})


# ADR-0043 §3: every prefix-list entry carries a description stating its
# purpose. Keyed by list id, because every rule within one of these lists serves
# the same purpose — successive address ranges of one category — so describing
# the LIST is honest, whereas inventing a distinct purpose per range would not
# be. Charset is [a-z0-9_]: ERA-71 established the Cumulus docs say "Avoid using
# apostrophes or non-ASCII characters. Cumulus Linux does not parse these
# characters", and the 255-char limit is nowhere near binding (longest is 31).
# ERA-64/ERA-71: prefix-list and community-list rules have NO `description`
# attribute in NVUE. Confirmed on live switches at both supported releases:
#
#   prefix-list rule     Error: 'description' is not one of ['match', 'action']
#   community-list rule  Error: 'description' is not one of ['community', 'action']
#
# The table and helper that used to live here generated 34 prefix-list and 2
# community-list descriptions per arch. They were emitted into root-executed
# configs, `nv config apply` rejected the whole config, and every switch in the
# 2-8-9-800 e2e cell stayed unreachable — a flat 0/8 — because a switch
# configured via Air Node Instructions never received its management address.
#
# !215 removed the TEMPLATE emission under time pressure and left this
# machinery in place, so the descriptions kept being computed and written into
# the generated inventory where nothing could ever emit them. Inert, but shaped
# exactly like configuration — the next person to add a prefix-list emission
# block would have wired it straight back into the same failure.
#
# Route-map rule descriptions are valid and unaffected; they are attached where
# route-maps are built. Max length is 80 characters (ERA-71, bisected).


def generate_prefix_lists(vlans, core_num, loopback_base=None, prefix_list_directives=None,
                          vrf_loopback_ips=None, oob_subnets=None, oob_uplink_mode='l2'):
    """Generate prefix_list configurations based on VLANs and switch number.

    vrf_loopback_ips: optional {VRF: 'a.b.c.d/32'} dict (already merged
    with any Loopbacks-sheet overrides). When provided, the per-VRF
    match rules (EXIT_LOCAL_IF, INBAND_LOCAL_IF, INBAND_PREFIXES,
    LOCAL_OOB_LOOPBACK, OOB_LOCAL_IF, OOB_PREFIXES) use these IPs so
    that BGP advertisement policy tracks the actual loopback values.
    Supernet rules (ERA_PREFIXES, VTEP_PREFIXES) still derive from
    loopback_base.

    oob_subnets: optional list of OOB management subnets from Settings
    (e.g., ['192.168.200.0/24']). These are added to ERA_PREFIXES so
    the EXIT-VRF outbound route-map permits them to cust-net-edge —
    required for the L3 OOB return path. The VLAN sheet's OOB entry
    typically declares a different (legacy) subnet, so we plumb the
    real OOB subnets through here.

    oob_uplink_mode: 'l2' (default) or 'l3'. This affects OOB_LOCAL_IF: in
    l2 mode CSL owns a real, locally-connected SVI on the OOB VLAN itself
    (the OOB switches are dumb L2 bridges in that design), which needs its
    own anti-duplicate protection rule since nothing else in OOB_FILTER
    covers it. In l3 mode that SVI doesn't exist (the OOB gateway lives on
    the OOB/MG switches instead), so OOB_LOCAL_IF is just the OOB VRF
    loopback with nothing else to add — see the inline comment at the
    OOB_LOCAL_IF construction below for why INBAND SVIs are deliberately
    NOT substituted in as a stand-in.
    """
    prefix_lists = []

    # ALL_PREFIXES - always present
    prefix_lists.append({
        'id': 'ALL_PREFIXES',
        'rule': [{'id': '10', 'match': '0.0.0.0/0', 'max_len': '32'}]
    })

    # Group VLANs by VRF
    vrf_vlans = {}
    for vlan in vlans:
        vrf = vlan['vrf']
        if vrf not in vrf_vlans:
            vrf_vlans[vrf] = []
        vrf_vlans[vrf].append(vlan)

    lb = loopback_base or LOOPBACK_BASE
    vrf_loopback_ips = vrf_loopback_ips or {}

    def _vrf_ip(vrf, computed):
        """Return the per-VRF loopback /32 — override if present, else computed."""
        ip = vrf_loopback_ips.get(vrf)
        if not ip:
            return computed
        return ip if '/' in ip else f"{ip}/32"

    # EXIT_LOCAL_IF - for EXIT VRF loopback (per-switch)
    prefix_lists.append({
        'id': 'EXIT_LOCAL_IF',
        'rule': [{'id': '10', 'match': _vrf_ip('EXIT', f'{lb}.{4 + core_num}/32'), 'max_len': '32'}]
    })

    # INBAND VRF prefix lists (per-switch loopback IPs)
    if 'INBAND' in vrf_vlans:
        inband_ip = _vrf_ip('INBAND', f'{lb}.{2 + core_num}/32')
        inband_rules = []
        # INBAND loopback (per-switch)
        inband_rules.append({'id': '10', 'match': inband_ip, 'max_len': '32'})
        rule_id = 20
        for vlan in vrf_vlans['INBAND']:
            if vlan['subnet']:
                svi_ip = _svi_switch_ip(vlan['subnet'], vlan.get('gateway'), core_num)
                inband_rules.append({'id': str(rule_id), 'match': f'{svi_ip}/32', 'max_len': '32'})
                rule_id += 10
        prefix_lists.append({'id': 'INBAND_LOCAL_IF', 'rule': inband_rules})

        # INBAND_PREFIXES
        inband_prefix_rules = []
        rule_id = 10
        for vlan in vrf_vlans['INBAND']:
            if vlan['subnet']:
                inband_prefix_rules.append({'id': str(rule_id), 'match': vlan['subnet'], 'max_len': '32'})
                rule_id += 10
        inband_prefix_rules.append({'id': str(rule_id), 'match': inband_ip, 'max_len': '32'})
        prefix_lists.append({'id': 'INBAND_PREFIXES', 'rule': inband_prefix_rules})

    # ERA_PREFIXES - loopback supernet (still tied to loopback_base; see
    # docs/LOOPBACKS.md if you override loopbacks into a different range)
    # PR2: rule 10's `max-prefix-len 24` silently drops any VLAN subnet
    # narrower than /24 (e.g. /25 INBAND halves) from outbound EXIT
    # advertisement. Walk declared INBAND/OOB VLAN subnets and add
    # explicit rules for any with prefix > 24 so they ride along.
    era_rules = [
        {'id': '10', 'match': f'{lb}.0/21', 'max_len': '24'},
        {'id': '20', 'match': f'{lb}.0/24', 'max_len': '32'},
    ]
    narrow_rule_id = 30
    seen_narrow = set()  # dedup across VRFs
    try:
        loopback_supernet = ipaddress.ip_network(f'{lb}.0/21', strict=False)
    except ValueError:
        loopback_supernet = None
    for advertised_vrf in ('INBAND', 'OOB'):
        for vlan in vrf_vlans.get(advertised_vrf, []):
            subnet = vlan.get('subnet')
            if not subnet or '/' not in subnet:
                continue
            try:
                net = ipaddress.ip_network(subnet, strict=False)
            except ValueError:
                continue
            # rule 10 covers anything within the ERA loopback supernet at
            # prefix <= /24. Subnets outside that supernet (e.g. OOB
            # 192.168.200.0/24) still need explicit rules — otherwise the
            # EXIT-VRF outbound route-map filters them and cust-net-edge
            # never learns them, breaking L3 OOB return path.
            if (loopback_supernet
                    and net.prefixlen <= 24
                    and net.subnet_of(loopback_supernet)):
                continue
            if subnet in seen_narrow:
                continue
            seen_narrow.add(subnet)
            era_rules.append({'id': str(narrow_rule_id),
                              'match': subnet, 'max_len': '32'})
            narrow_rule_id += 10
    # oob_subnets — the real OOB subnets, resolved from the OOB VLAN(s)
    # (the VLAN sheet's OOB entry is typically the legacy 172.16.177.0/24
    # placeholder, but the actual operator-facing OOB SVI lives on the
    # resolved oob_subnets value).
    for subnet in (oob_subnets or []):
        if not subnet or subnet in seen_narrow:
            continue
        try:
            ipaddress.ip_network(subnet, strict=False)
        except ValueError:
            continue
        seen_narrow.add(subnet)
        era_rules.append({'id': str(narrow_rule_id),
                          'match': subnet, 'max_len': '32'})
        narrow_rule_id += 10
    prefix_lists.append({'id': 'ERA_PREFIXES', 'rule': era_rules})

    # ERA-52: OOB_HOSTS — the address ranges a genuine OOB host may occupy.
    # Used by EVPN_OOB_OUT to distinguish a real BMC address from a non-OOB
    # address that a misconfigured multihomed node has put on VLAN 200.
    #
    # Derived from the resolved OOB subnets rather than the sample ticket's
    # hardcoded 10.184.177.0/24, and emitted as one rule per subnet because
    # resolve_oob_vlans() supports per-switch OOB VLANs (200/201) — a single
    # rule would silently exclude every switch on the second VLAN.
    oob_host_rules = []
    _rid = 10
    for subnet in (oob_subnets or []):
        if not subnet:
            continue
        try:
            ipaddress.ip_network(subnet, strict=False)
        except ValueError:
            continue
        oob_host_rules.append({'id': str(_rid), 'match': subnet, 'max_len': '32'})
        _rid += 10
    if oob_host_rules:
        prefix_lists.append({'id': 'OOB_HOSTS', 'rule': oob_host_rules})

    # OOB VRF prefix lists (per-switch loopback IPs)
    if 'OOB' in vrf_vlans:
        oob_vlan = vrf_vlans['OOB'][0]
        if oob_vlan['subnet']:
            oob_svi_ip = _svi_switch_ip(oob_vlan['subnet'], oob_vlan.get('gateway'), core_num)
            oob_ip = _vrf_ip('OOB', f'{lb}.{core_num}/32')

            prefix_lists.append({
                'id': 'LOCAL_OOB_LOOPBACK',
                'rule': [{'id': '10', 'match': oob_ip, 'max_len': '32'}]
            })

            oob_local_rules = [{'id': '10', 'match': oob_ip, 'max_len': '32'}]
            # l2 mode: CSL genuinely owns an SVI on the OOB VLAN itself (the
            # OOB switches are dumb L2 bridges in that design), and nothing
            # else in OOB_FILTER blocks it -- INBAND_PREFIXES only covers
            # INBAND subnets, not the OOB VLAN's own subnet -- so it needs
            # explicit anti-duplicate protection here.
            #
            # l3 mode: no such SVI exists on this switch (the OOB gateway
            # lives on the OOB/MG switches instead), so there is nothing
            # real left to protect here. We deliberately do NOT substitute
            # this switch's INBAND SVI addresses instead: OOB_FILTER rule 10
            # (INBAND_PREFIXES) already unconditionally denies the entire
            # INBAND subnet range before OOB_LOCAL_IF (rule 15) is ever
            # evaluated, so an INBAND-SVI /32 entry here would be dead,
            # redundant config -- the same class of leftover mistake flagged
            # in production's own OOB_LOCAL_IF rule 20 (see
            # era-documentation/guides/csl-routing-policy-analysis.md).
            if str(oob_uplink_mode).strip().lower() != 'l3':
                oob_local_rules.append({'id': '20', 'match': f'{oob_svi_ip}/32', 'max_len': '32'})

            prefix_lists.append({'id': 'OOB_LOCAL_IF', 'rule': oob_local_rules})

            prefix_lists.append({
                'id': 'OOB_PREFIXES',
                'rule': [
                    {'id': '10', 'match': oob_vlan['subnet'], 'max_len': '32'},
                    {'id': '20', 'match': oob_ip, 'max_len': '32'},
                ]
            })

    # VTEP_PREFIXES — supernet, still tied to loopback_base
    prefix_lists.append({
        'id': 'VTEP_PREFIXES',
        'rule': [{'id': '5', 'match': f'{lb}.8/29', 'max_len': '32'}]
    })

    # Apply directives from the Excel 'Prefix lists' sheet.
    # prefix_list_directives is the full {'lists': {...}, 'suppress': set()}
    # dict returned by parse_prefix_lists_sheet — NOT pre-unwrapped. Directive
    # application is gated on the OVERRIDABLE_PREFIX_LISTS /
    # DERIVED_ONLY_PREFIX_LISTS classification so per-switch /32 lists
    # (EXIT_LOCAL_IF, INBAND_LOCAL_IF, OOB_LOCAL_IF, LOCAL_OOB_LOOPBACK) can
    # never be overridden or suppressed by an operator.
    directives = prefix_list_directives or {}
    directive_lists = directives.get('lists') or {}
    directive_suppress = directives.get('suppress') or set()

    existing_ids = {pl['id'] for pl in prefix_lists}
    for list_id, rules in directive_lists.items():
        if list_id in existing_ids and list_id in OVERRIDABLE_PREFIX_LISTS:
            # override: replace the existing global-subnet list's rules
            for pl in prefix_lists:
                if pl['id'] == list_id:
                    pl['rule'] = rules
        elif list_id not in existing_ids and list_id not in DERIVED_ONLY_PREFIX_LISTS:
            # add: a brand-new operator-defined list
            prefix_lists.append({'id': list_id, 'rule': rules})
            existing_ids.add(list_id)
        else:
            # derived-only target (existing or by name) — never overridable
            print(f"  ⚠️  Prefix list directive for '{list_id}' ignored "
                  f"— derived-only per-switch list, never user-overridable.")

    for list_id in directive_suppress:
        if list_id in OVERRIDABLE_PREFIX_LISTS:
            prefix_lists[:] = [pl for pl in prefix_lists if pl['id'] != list_id]
            existing_ids.discard(list_id)
        elif list_id in DERIVED_ONLY_PREFIX_LISTS:
            print(f"  ⚠️  Suppress directive for '{list_id}' ignored "
                  f"— derived-only per-switch list, never user-overridable.")

    return prefix_lists


# Control-plane ACL classification (ADR-0030). acl-default-dos / acl-default-whitelist
# are NVUE built-ins bound inbound on every switch; the whitelist additionally carries
# a spine-only service rule. Operators override/add/suppress via the 'ACLs' sheet.
OVERRIDABLE_ACLS = frozenset({'acl-default-dos', 'acl-default-whitelist'})
# Suppressing either baseline control-plane ACL is honored but flagged (security nudge).
SECURITY_DEFAULT_ACLS = frozenset({'acl-default-dos', 'acl-default-whitelist'})
# Categories that render the spine template (roles/spine/templates/spine_nvue_cli.j2)
# — the only roles that define the whitelist service rule. Matches the playbook
# group gate (cs / gs_plane1 / gs_plane2). Collapsed gsl/gl planes render the gl
# template (bindings only), so they are intentionally excluded.
_SPINE_ACL_CATEGORIES = frozenset({'cs', 'gs-plane1', 'gs-plane2'})


# Functions whose nodes actually receive a generated config file, per the `when:`
# conditions in playbooks/generate-cli-configs.yml. This is deliberately NOT
# utils.is_switch(): that returns True for `edge` and `air-oob`, which render no
# template. Targeting them would pass validation and then silently produce nothing —
# the exact failure Custom_Config validation exists to prevent (ADR-0055).
# A boundary test pins this set against the playbook.
CONFIG_GENERATING_FUNCTIONS = {
    "core", "csl", "cl",                      # core template
    "gsl-plane1", "gsl-plane2",               # gl template
    "gl-plane1", "gl-plane2",                 # gl template
    "cs", "gs-plane1", "gs-plane2",           # spine template
    "oob-switch",                             # oob-switch template
}


def _node_function(node):
    """The Function-column value of a node dict, whichever producer made it.

    This codebase has TWO node parsers that disagree on the key:
    `excel_parser.parse_nodes()` emits `role`, `validate_excel.validate_nodes()`
    emits `function`. Reading only one silently yields an empty switch inventory in
    the other's context — which made validate_excel reject every Custom_Config row
    with "Valid: ." Same shape as ADR-0054's trap 1, where the models spelled
    functions `oob-switch`/`gsl-plane1` and classify_node() spelled them `oob`/`gsl`.
    """
    return str(node.get("role") or node.get("function") or "").strip()


def switches_by_function_from_nodes(nodes):
    """{function: [switch names]} for functions that receive a generated config.

    Nodes with no template — servers, `edge`, `air-oob` — are excluded, so they can
    never be targeted by a Custom_Config row.
    """
    out = {}
    for node in nodes or []:
        func = _node_function(node)
        name = str(node.get("name") or "").strip()
        if not func or not name or func not in CONFIG_GENERATING_FUNCTIONS:
            continue
        out.setdefault(func, []).append(name)
    return {k: sorted(v) for k, v in sorted(out.items())}


def servers_by_function_from_nodes(nodes):
    """{function: [names]} for everything that is NOT config-generating.

    Used only to improve the error message when someone targets `gpu`.
    """
    out = {}
    for node in nodes or []:
        func = _node_function(node)
        name = str(node.get("name") or "").strip()
        if not func or not name or func in CONFIG_GENERATING_FUNCTIONS:
            continue
        out.setdefault(func, []).append(name)
    return {k: sorted(v) for k, v in sorted(out.items())}


class CustomConfigError(ValueError):
    """A `Custom_Config` row is unusable. Always fatal — never downgraded to a warning.

    ADR-0055: a target that matches nothing, while the operator believes their config
    shipped, is the same silent-no-op class as `make fix-ext-storage` exiting 0 having
    done nothing (#59). Refusing the import is the whole point of this parser.
    """


# A legitimate `nv` line never needs shell metacharacters; values with spaces are
# quoted by the templates. Rejecting these is best-effort containment, not a sandbox.
_SHELL_METACHARACTERS = ("`", "$(", ";", "&&", "||", "|", ">", "<")

# Apply-lifecycle control (`nv config`) belongs to the tool: a mid-script apply leaves
# the rest of the file staging against live state, and `nv config replace` can discard
# the entire generated configuration. `nv action` is imperative — it leaves no config
# state (so it is unverifiable), is non-idempotent in a file ZTP and push-switch-configs
# both re-run, and includes destructive verbs like `install system image`.
_BLOCKED_NV_VERBS = ("config", "action")


def _validate_custom_config_line(line, row_no, line_no):
    """Raise CustomConfigError unless `line` is an acceptable `nv` command."""
    where = f"Custom_Config row {row_no}, line {line_no}"
    if not line.startswith("nv "):
        raise CustomConfigError(
            f"{where}: {line!r} does not start with 'nv '. This sheet accepts NVUE "
            f"commands only — configuration, not shell."
        )
    if line.endswith("\\"):
        raise CustomConfigError(
            f"{where}: {line!r} ends with a line continuation, which is not allowed."
        )
    for meta in _SHELL_METACHARACTERS:
        if meta in line:
            raise CustomConfigError(
                f"{where}: {line!r} contains {meta!r}. Shell metacharacters are "
                f"rejected; a legitimate nv command never needs them."
            )
    verb = line.split()[1] if len(line.split()) > 1 else ""
    if verb in _BLOCKED_NV_VERBS:
        raise CustomConfigError(
            f"{where}: 'nv {verb}' is not allowed. "
            + ("Applying or replacing the configuration is the tool's job — a "
               "mid-script apply leaves the rest of the file staging against live "
               "state." if verb == "config" else
               "Actions are imperative and leave no configuration state, so they "
               "cannot be verified and are not idempotent when ZTP re-runs the file.")
        )


def _resolve_targets(location, row_no, switches_by_function, servers_by_function):
    """Resolve a Switch_Location expression to a sorted list of switch names."""
    text = str(location).strip()
    all_switches = sorted(n for names in switches_by_function.values() for n in names)
    if text.upper() == "ALL":
        return all_switches

    if ":" not in text:
        raise CustomConfigError(
            f"Custom_Config row {row_no}: {text!r} is not a valid Switch_Location. "
            f"Use 'ALL', 'Function: <list>' or 'Host: <list>'."
        )
    keyword, _, rest = text.partition(":")
    keyword = keyword.strip()
    tokens = [t.strip() for t in rest.split(",") if t.strip()]
    if not tokens:
        raise CustomConfigError(
            f"Custom_Config row {row_no}: {text!r} names no {keyword.lower()}s."
        )

    if keyword == "Function":
        valid = sorted(switches_by_function)
        resolved = []
        for tok in tokens:
            if tok in switches_by_function:
                resolved.extend(switches_by_function[tok])
                continue
            hint = ""
            if tok in (servers_by_function or {}):
                hint = (f" {tok!r} is a server function; this sheet configures "
                        f"switches only.")
            raise CustomConfigError(
                f"Custom_Config row {row_no}: {tok!r} is not a valid switch function "
                f"in this workbook.{hint} Valid: {', '.join(valid)}."
            )
        return sorted(set(resolved))

    if keyword == "Host":
        known = {n for names in switches_by_function.values() for n in names}
        for tok in tokens:
            if tok not in known:
                raise CustomConfigError(
                    f"Custom_Config row {row_no}: {tok!r} is not a switch in this "
                    f"workbook. Valid: {', '.join(sorted(known))}."
                )
        return sorted(set(tokens))

    raise CustomConfigError(
        f"Custom_Config row {row_no}: {keyword!r} is not a valid Switch_Location "
        f"keyword. Use 'ALL', 'Function: <list>' or 'Host: <list>'."
    )


def custom_config_keypaths(lines):
    """NVUE key paths a switch's custom config touches, for validate-config.

    Any key path custom config writes to produces textual differences on BOTH sides of
    the comparison, so both must be suppressed:

      generated:  nv set system message pre-login "Authorized use only"
      running:    nv set system message pre-login 'Authorized use only'

    NVUE re-quotes values, so the operator line reads MISSING; if the operator overrode
    a key the tool also sets, the tool line reads MISSING too; and the running form
    reads EXTRA. A live Air run on 2026-08-17 showed all three at once.

    Key path = the tokens after the verb, minus the value:
      nv unset <path>            -> <path>              (no value)
      nv set <path> "<value>"    -> <path>              (quoted value)
      nv set <path> <value>      -> <path>              (bare final token)

    `nv show` yields nothing: it changes no state, so it must not suppress anything.
    The split is a heuristic, and deliberately so — v1 policy is to WARN rather than
    fail, so an imprecise boundary costs a warning, never a false pass.
    """
    out = []
    for raw in lines or []:
        line = str(raw).strip()
        if line.startswith("nv unset "):
            path = line[len("nv unset "):].strip()
        elif line.startswith("nv set "):
            rest = line[len("nv set "):].strip()
            cut = min((i for i in (rest.find('"'), rest.find("'")) if i > -1), default=-1)
            if cut > -1:
                path = rest[:cut].strip()
            else:
                path = rest.rsplit(" ", 1)[0].strip() if " " in rest else rest
        else:
            continue
        if path and path not in out:
            out.append(path)
        # normalize_nvue.py:155 rewrites `state enabled` -> `enable on` before the
        # comparison, so the same setting reaches the comparator under a different
        # final token. An operator writing `nv unset system wjh state` must also
        # suppress the generated `nv set system wjh enable on`. Only the TRAILING
        # token aliases — `state` mid-path is not the enable/disable idiom.
        for a, b in (("state", "enable"), ("enable", "state")):
            if path.endswith(" " + a):
                alias = path[: -len(a)] + b
                if alias not in out:
                    out.append(alias)
    return sorted(out)


def parse_custom_config_sheet(ws, switches_by_function, servers_by_function=None):
    """Parse the optional 'Custom_Config' sheet (ADR-0055).

    Columns: Switch_Location, Config. Targeting is resolved HERE, in Python, so it is
    testable without rendering and so security-relevant filtering does not live in a
    Jinja2 template.

    Returns: {switch_name: [line, ...]} accumulated in sheet order. An absent or empty
    sheet yields {}, and the feature is inert.

    Raises CustomConfigError on any unresolvable target or unacceptable line — never a
    warning, never a silent skip.
    """
    result = {}
    if ws is None or ws.max_row < 2:
        return result

    for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not any(c not in (None, "") for c in row[:2]):
            continue
        location, config = (row[0], row[1] if len(row) > 1 else None)
        # A '#' row is guidance, not configuration. The shipped sheet carries usage
        # help, and without this the help text parsed as data and failed
        # `make generate`. Also lets an operator comment a row out.
        if str(location or "").lstrip().startswith("#"):
            continue
        if location in (None, ""):
            raise CustomConfigError(
                f"Custom_Config row {row_no}: has configuration but no Switch_Location."
            )
        if config in (None, "") or not str(config).strip():
            raise CustomConfigError(
                f"Custom_Config row {row_no}: targets {str(location).strip()!r} but "
                f"has no configuration. Remove the row or fill in the Config column."
            )

        lines = []
        for line_no, raw in enumerate(str(config).splitlines(), start=1):
            line = raw.strip()
            if not line:
                continue
            _validate_custom_config_line(line, row_no, line_no)
            lines.append(line)
        if not lines:
            raise CustomConfigError(
                f"Custom_Config row {row_no}: configuration is blank after stripping."
            )

        for switch in _resolve_targets(location, row_no, switches_by_function,
                                       servers_by_function or {}):
            result.setdefault(switch, []).extend(lines)

    return result


def parse_acls_sheet(ws):
    """
    Parse the optional 'ACLs' sheet (ADR-0030 control-plane ACL override).
    Columns: ACL name, Rule id, Protocol, Dest port, Action.
    Header is the 'ACL name' row; data below. Fully optional — an absent
    workbook, or an absent/empty sheet, produces byte-identical output.

    Action semantics (mirrors parse_prefix_lists_sheet):
      - blank ⇒ the row's rule is collected under its acl name in ``acls``.
      - 'suppress' (case-insensitive) ⇒ the acl name is added to ``suppress``
        (any rule on that row is ignored); suppress wins regardless of row order.

    Set-agnostic: override-vs-add is decided in generate_acls.

    Returns: {'acls': {acl_name: [ {id, protocol, dest_port}, ... ]},
              'suppress': set()}
    """
    result = {'acls': {}, 'suppress': set()}
    acls = result['acls']
    suppress = result['suppress']
    if ws.max_row < 2:
        return result
    # Find header row (row with "ACL name" in col 1); data starts next row
    header_row = 1
    for r in range(1, min(ws.max_row + 1, 5)):
        val = ws.cell(row=r, column=1).value
        if val and str(val).strip().lower() == 'acl name':
            header_row = r
            break
    data_start = header_row + 1
    for row in range(data_start, ws.max_row + 1):
        acl_name = ws.cell(row=row, column=1).value
        rule_id = ws.cell(row=row, column=2).value
        protocol = ws.cell(row=row, column=3).value
        dest_port = ws.cell(row=row, column=4).value
        if not acl_name or not str(acl_name).strip():
            continue
        name = str(acl_name).strip()
        if name.lower() == 'acl name':
            continue  # skip header row if it appears in data range
        action_val = ws.cell(row=row, column=5).value
        action = str(action_val).strip().lower() if action_val else ''
        if action == 'suppress':
            suppress.add(name)
            acls.pop(name, None)
            continue
        rule = {
            'id': str(rule_id).strip() if rule_id else str(row - 1),
            'protocol': str(protocol).strip() if protocol else '',
            'dest_port': (str(int(dest_port)) if isinstance(dest_port, (int, float))
                          else str(dest_port).strip()) if dest_port not in (None, '') else '',
        }
        if not rule['protocol'] or not rule['dest_port']:
            continue  # a whitelist rule needs both a protocol and a dest port
        if name in suppress:
            continue  # suppress wins even if the suppress row came first
        if name not in acls:
            acls[name] = []
        acls[name].append(rule)
    return result


def generate_acls(category, acl_directives=None):
    """Generate the per-switch control-plane ACL set (ADR-0030).

    Returns the host_vars['acls'] list: ordered ACL dicts
    ``{id, control_plane_inbound, type, rule: [{id, protocol, dest_port}, ...]}``.
    Bindings render before rule/type lines (see the role templates), so the
    default order is preserved byte-for-byte.

    Defaults (tool-owned, moved out of the Jinja templates):
      - every switch binds acl-default-dos + acl-default-whitelist inbound;
      - the spine role (cs / gs-plane1 / gs-plane2) additionally defines
        acl-default-whitelist rule 200 (tcp/8251) + type ipv4.

    acl_directives is the full {'acls': {...}, 'suppress': set()} dict from
    parse_acls_sheet (not pre-unwrapped); override/add/suppress decided here.
    An existing acl name ⇒ its rules are replaced (override); a new name ⇒ a new
    inbound-bound ACL is added; suppress removes the acl entirely.
    """
    is_spine = category in _SPINE_ACL_CATEGORIES
    acls = [
        {'id': 'acl-default-dos', 'control_plane_inbound': True,
         'type': None, 'rule': []},
        {'id': 'acl-default-whitelist', 'control_plane_inbound': True,
         'type': 'ipv4' if is_spine else None,
         'rule': ([{'id': '200', 'protocol': 'tcp', 'dest_port': '8251'}]
                  if is_spine else [])},
    ]

    directives = acl_directives or {}
    directive_acls = directives.get('acls') or {}
    directive_suppress = directives.get('suppress') or set()

    existing_ids = {a['id'] for a in acls}
    for name, rules in directive_acls.items():
        if name in existing_ids:
            # override: replace the default ACL's rules (a rule set implies ipv4)
            for a in acls:
                if a['id'] == name:
                    a['rule'] = rules
                    if rules and not a['type']:
                        a['type'] = 'ipv4'
        else:
            # add: a brand-new operator-defined ACL, bound inbound
            acls.append({'id': name, 'control_plane_inbound': True,
                         'type': 'ipv4' if rules else None, 'rule': rules})
            existing_ids.add(name)

    for name in directive_suppress:
        acls[:] = [a for a in acls if a['id'] != name]
        existing_ids.discard(name)

    return acls


def _ensure_mask(value, default_mask='/32'):
    """Append a /32 mask if the value is a bare IP."""
    if not value:
        return value
    return value if '/' in value else f"{value}{default_mask}"


def _strip_gpu_plane(core_vars, vlans):
    """Return a deep copy of core_vars with GPU-plane entries removed.

    Used when writing csl.yml in dedicated_gpu architectures: the CSL
    fabric only handles CPU/Storage/Support/OOB traffic, while the GPU
    plane (VLAN 900, GPU VRF, GPU VNI) lives on separate GSL switches.
    Without stripping, csl.yml inherits GPU VRF + vlan900 SVI from the
    source inventory and the rendered CSL config gains spurious
    `vlan900` SVIs and a GPU VRF that never gets used.
    """
    import copy
    stripped = copy.deepcopy(core_vars)
    gpu_vlan_ids = {v['id'] for v in vlans if v.get('vrf') == 'GPU'}

    if 'vlans' in stripped and isinstance(stripped['vlans'], list):
        stripped['vlans'] = [v for v in stripped['vlans'] if v not in gpu_vlan_ids]
    if 'vnis' in stripped and isinstance(stripped['vnis'], dict):
        stripped['vnis'] = {k: v for k, v in stripped['vnis'].items() if k not in gpu_vlan_ids}
    if 'vrf_vnis' in stripped and isinstance(stripped['vrf_vnis'], dict):
        stripped['vrf_vnis'].pop('GPU', None)
    if 'vrf_config' in stripped and isinstance(stripped['vrf_config'], list):
        stripped['vrf_config'] = [v for v in stripped['vrf_config']
                                  if v.get('id') != 'GPU']
    for key in ('gpu_interfaces', 'gpu_rail_interfaces', 'gpu_rail_planes'):
        stripped.pop(key, None)
    return stripped


def generate_vrf_loopbacks(vlans, core_num, loopback_base=None, switch_overrides=None,
                           skip_gpu=False):
    """Generate VRF loopback IP assignments — unique per switch.

    If switch_overrides is provided (from the Loopbacks Excel sheet), any
    VRF whose entry is non-empty in the override dict replaces the
    computed default. Empty/missing entries fall back to computed values.

    skip_gpu=True drops the GPU VRF entry entirely — used for CSL nodes in
    dedicated_gpu archs where the GPU plane lives on GSL, not CSL.
    """
    lb = loopback_base or LOOPBACK_BASE
    # One contiguous block per VRF (see VRF_LOOPBACK_BLOCKS) so the series can
    # never overlap each other or the switch loopbacks, at any N/S leaf count.
    if core_num > VRF_LOOPBACK_BLOCK_SIZE:
        raise ValueError(
            f"switch index {core_num} exceeds the per-VRF loopback block size "
            f"({VRF_LOOPBACK_BLOCK_SIZE}); widen VRF_LOOPBACK_BLOCK_SIZE and "
            f"re-space VRF_LOOPBACK_BLOCKS in excel_parser.py"
        )
    vrf_loopbacks = {
        'EXIT':   f'{lb}.{VRF_LOOPBACK_BLOCKS["EXIT"] + core_num}/32',
        'INBAND': f'{lb}.{VRF_LOOPBACK_BLOCKS["INBAND"] + core_num}/32',
        'OOB':    f'{lb}.{VRF_LOOPBACK_BLOCKS["OOB"] + core_num}/32',
    }

    # STORAGE (ERA-39): every arch declares the STORAGE VRF in its VLANs &
    # Profiles sheet, so give it a computed loopback wherever that VRF exists
    # rather than leaving it opt-in per workbook. Previously this was
    # override-only, which is why the collapsed-core archs ended up with a
    # STORAGE VRF that had EVPN + BGP but no loopback.
    if any(vlan.get('vrf') == 'STORAGE' for vlan in vlans):
        vrf_loopbacks['STORAGE'] = f'{lb}.{VRF_LOOPBACK_BLOCKS["STORAGE"] + core_num}/32'

    if not skip_gpu:
        for vlan in vlans:
            if vlan['vrf'] == 'GPU' and vlan['subnet']:
                gpu_subnet_base = vlan['subnet'].rsplit('.', 1)[0]
                vrf_loopbacks['GPU'] = f'{gpu_subnet_base}.{4 + core_num}/32'
                break

    if switch_overrides:
        loopback_vrfs = ('OOB', 'INBAND', 'EXIT', 'STORAGE') if skip_gpu else \
                        ('OOB', 'INBAND', 'EXIT', 'GPU', 'STORAGE')
        for vrf in loopback_vrfs:
            override = switch_overrides.get(vrf)
            if override:
                vrf_loopbacks[vrf] = _ensure_mask(override)

    return vrf_loopbacks


def _apply_oob_l3_uplink_mode(core_vars, settings):
    """Rewrite default-VRF BGP intent when OOB uplinks are direct L3 links."""
    if _normalize_oob_uplink_mode(settings) != 'l3':
        return
    if not isinstance(core_vars.get('oob_uplink_interfaces'), dict):
        return
    default_vrf_bgp = core_vars.get('default_vrf_bgp')
    if not isinstance(default_vrf_bgp, dict):
        return

    # Dedicated cl/cs (ns_tiers > 1) overlay is hub-and-spoke via the
    # cs spines — cl leaves peer the cs spine loopbacks (not the oob-switches).
    # Converged csl (ns_tiers = 1, no cs spine) keeps the cl↔oob overlay mesh.
    if int((settings or {}).get('ns_tiers') or 1) > 1 and settings.get('_derived_cs_spine_overlay_peers'):
        overlay_peers = list(settings.get('_derived_cs_spine_overlay_peers') or [])
    else:
        overlay_peers = list(settings.get('_derived_oob_overlay_peers') or [])
    underlay_remote_as = 'external'
    overlay_remote_as = 'external'
    overlay_ttl = 2

    neighbors = []
    if isinstance(core_vars.get('isl_interfaces'), dict):
        neighbors.append({
            'interfaces': 'isl',
            'peer_group': 'internal_isl',
            'type': 'unnumbered',
        })
    neighbors.append({
        'interfaces': 'oob_uplink',
        'peer_group': 'underlay',
        'type': 'unnumbered',
    })
    if overlay_peers:
        neighbors.append({
            'interfaces': overlay_peers,
            'peer_group': 'overlay',
            'type': 'numbered',
        })

    peer_groups = []
    if isinstance(core_vars.get('isl_interfaces'), dict):
        _ns_tiers = int((settings or {}).get('ns_tiers') or 1)
        _isl_ipv4_unicast = {'enable': True}
        if _ns_tiers > 1:
            # W-ECMP only works over eBGP (NVIDIA Cumulus Linux docs:
            # "W-ECMP is only supported in EBGP fabrics"). At ns_tiers > 1
            # this ISL is a genuine leaf (cl) uplink to the cs spine, not a
            # converged-csl horizontal peer-link, and it's eBGP -- so it
            # qualifies the same way the underlay-to-MG peer-group does.
            # Verified via a generated SU16 (ns_tiers=2) workbook: cl's
            # internal_isl is eBGP with ipv4-unicast enabled, mirroring the
            # cs spine's own 'underlay' downlink on the other end (see
            # WEIGHTED_ECMP_CUMULATIVE in spine_nvue_cli.j2). No live 2-tier
            # deployment exists to verify byte-for-byte, unlike the
            # underlay-to-MG fix.
            _isl_ipv4_unicast['policy_outbound_route_map'] = 'WEIGHTED_ECMP'
        peer_groups.append({
            'id': 'internal_isl',
            # A dedicated cl/cs spine-leaf (ns_tiers > 1) runs an eBGP
            # Clos underlay — the cl↔cs ISL must be eBGP ('external') to match the
            # cs side (peer-group 'underlay', remote-as external) or the OPEN is
            # rejected and every cl↔cs session sits IDLE (no loopback propagation
            # → no EVPN → OOB VLAN-200 stretch dead). Converged csl (ns_tiers=1)
            # keeps iBGP ('internal') — the golden csl-pair peer-link, unchanged.
            # Mirrors the proven gl/gs ew_tiers fix; design confirmed by the
            # Common Networking RA (BGP-based ECMP underlay).
            'remote_as': 'external' if _ns_tiers > 1 else 'internal',
            'bfd_enable': True,
            'description': 'internal_isl_interconnect',
            'address_family': {
                'ipv4_unicast': _isl_ipv4_unicast,
                'l2vpn_evpn': {'enable': True},
            },
        })
    peer_groups.append({
        'id': 'underlay',
        'remote_as': underlay_remote_as,
        'bfd_enable': True,
        'description': 'oob_underlay_interconnect',
        'address_family': {
            # W-ECMP (BGP link-bandwidth extended community) only works over
            # eBGP (NVIDIA Cumulus Linux docs: "W-ECMP is only supported in
            # EBGP fabrics") -- this peer-group's remote_as is always
            # 'external' here, so it always qualifies. Mirrors the identical,
            # already-correct WEIGHTED_ECMP outbound policy on the OOB/MG
            # switch's own 'underlay' peer-group (the other end of this same
            # link) in oob_nvue_cli.j2.
            'ipv4_unicast': {'enable': True, 'policy_outbound_route_map': 'WEIGHTED_ECMP'},
        },
    })
    if overlay_peers:
        peer_groups.append({
            'id': 'overlay',
            'remote_as': overlay_remote_as,
            'bfd_enable': True,
            'bfd_profile': 'overlay',
            'description': 'oob_overlay_interconnect',
            'multihop_ttl': overlay_ttl,
            'update_source': 'lo',
            'address_family': {
                'ipv4_unicast': {'enable': False},
                # ERA-52: filter what this fabric ORIGINATES into the OOB VNI.
                #
                # The leak vector is the support-server bonds on core/csl, which
                # trunk VLAN 200 alongside the data VLANs on one EVPN-MH bond
                # (`bridge domain br_default vlan 200,300,400`). A node that puts
                # a non-OOB address on VLAN 200 makes core/csl originate an EVPN
                # type-2 macip into VNI 4200, which lands in the OOB VRF.
                #
                # NOT already covered by OOB_FILTER: that is a
                # `route-import from-vrf` policy on ipv4-unicast, governing what
                # the OOB VRF IMPORTS from other VRFs. It never sees a type-2
                # macip this fabric originates itself.
                'l2vpn_evpn': {
                    'enable': True,
                    'policy_outbound_route_map': 'EVPN_OOB_OUT',
                },
            },
        })

    default_vrf_bgp['neighbors'] = neighbors
    default_vrf_bgp['peer_groups'] = peer_groups


def _derive_oob_overlay_peers(nodes=None, loopback_overrides=None):
    """Return default-loopback IPs for OOB switches, in node order."""
    peers = []
    loopback_overrides = loopback_overrides or {}
    for node in nodes or []:
        name = (node.get('name') or '').strip()
        if not name:
            continue
        if canonical_category(node.get('role'), name) != 'oob-switch':
            continue
        lo = (loopback_overrides.get(name) or {}).get('lo')
        if lo:
            peers.append(str(lo).split('/')[0].strip())
    return peers


def _derive_core_overlay_peers(nodes=None, loopback_overrides=None):
    """Return overlay peer loopback IPs that OOB switches should peer to.

    In standard eBGP EVPN (Approach B — separate overlay/underlay), leaves
    peer overlay to spines.  When dedicated spines exist (csl-spine), return
    their loopbacks.  In collapsed designs (no spines — core or bare csl),
    return the core/csl loopbacks since those switches act as both spine and
    leaf.
    """
    loopback_overrides = loopback_overrides or {}
    spine_peers = []
    leaf_peers = []
    for node in nodes or []:
        name = (node.get('name') or '').strip()
        if not name:
            continue
        cat = canonical_category(node.get('role'), name)
        lo = (loopback_overrides.get(name) or {}).get('lo')
        if not lo:
            continue
        ip = str(lo).split('/')[0].strip()
        if cat == 'cs':
            spine_peers.append(ip)
        elif cat in ('core', 'csl', 'cl'):
            leaf_peers.append(ip)
    return spine_peers if spine_peers else leaf_peers


def resolve_oob_vlans(vlans, oob_nodes):
    """Resolve each OOB switch's OOB VLAN (and subnet) from the VLANs &
    Profiles sheet.

    Args:
        vlans: Parsed 'VLANs & Profiles' rows (list of dicts with at least
            'id', 'vrf', 'subnet'). VRF-OOB rows (vrf == 'OOB', case-
            insensitive) are the candidate OOB VLANs.
        oob_nodes: Parsed OOB switch nodes (list of dicts with 'name' and
            optional 'oob_vlan' — the Nodes-tab 'OOB VLAN' column value).

    Sole-VLAN default rule: when exactly one VRF-OOB VLAN exists, a blank
    'oob_vlan' on a switch defaults to that VLAN's id. With zero or more than
    one VRF-OOB VLAN, a blank 'oob_vlan' does NOT default — it resolves to
    None (see validate_excel.py's cross-sheet guardrail, which errors on
    this case for >1 OOB VLAN). A non-blank 'oob_vlan' that doesn't match
    any VRF-OOB VLAN id also resolves to None.

    Returns:
        dict with:
          'oob_vlans': list of the VRF-OOB VLAN rows.
          'default_vlan_id': the sole OOB VLAN's int id if exactly one
              exists, else None.
          'vlan_by_switch': dict of switch name -> resolved VLAN row (or
              None if unresolved).
          'subnets': de-duplicated list of OOB VLAN subnets actually in use
              by the given oob_nodes (in first-seen order).
    """
    oob_vlans = [v for v in (vlans or []) if str(v.get('vrf', '')).strip().upper() == 'OOB']
    by_id = {}
    for v in oob_vlans:
        try:
            by_id[int(v['id'])] = v
        except (TypeError, ValueError, KeyError):
            continue
    default_vlan_id = int(oob_vlans[0]['id']) if len(oob_vlans) == 1 else None

    vlan_by_switch = {}
    subnets = []
    seen = set()
    for node in (oob_nodes or []):
        raw = str(node.get('oob_vlan', '') or '').strip()
        vid = None
        if raw:
            try:
                vid = int(float(raw))
            except (TypeError, ValueError):
                vid = None
        elif default_vlan_id is not None:
            vid = default_vlan_id
        vlan = by_id.get(vid) if vid is not None else None
        vlan_by_switch[node['name']] = vlan
        if vlan and vlan.get('subnet') and vlan['subnet'] not in seen:
            seen.add(vlan['subnet'])
            subnets.append(vlan['subnet'])
    return {
        'oob_vlans': oob_vlans,
        'default_vlan_id': default_vlan_id,
        'vlan_by_switch': vlan_by_switch,
        'subnets': subnets,
    }


def get_oob_nodes_for_inventory(nodes, settings, vlans):
    """
    Return OOB switch nodes for inventory, derived from the oob-switch rows
    on the Nodes tab and the OOB VLAN mapping (per-switch `oob_vlan` resolved
    against the VRF=OOB VLANs from the VLANs & Profiles sheet — see
    resolve_oob_vlans()). Every Active oob-switch Nodes row is a real switch;
    there is no synthetic padding (OOB switches must carry a real
    `OOB VLAN` mapping).

    SVI IPs and gateways are derived from each switch's mapped OOB VLAN subnet:
      - Shared subnet (multiple switches map to the same VLAN): anycast
        gateway=.1, SVI IPs=.2,.3,.4 assigned in switch order
      - Distinct subnet per switch: gateway=.1, SVI=.2 in each subnet
    """
    oob_from_sheet = sorted(
        [n for n in nodes if n['status'] == 'Active' and n.get('category') == 'oob-switch'],
        key=lambda x: x['name'],
    )
    result = list(oob_from_sheet)

    mapping = resolve_oob_vlans(vlans, result)
    # Group switches by their resolved OOB subnet so shared-subnet switches get
    # sequential .2/.3/.4 SVIs while distinct-subnet switches each get .2.
    per_subnet_index = {}
    for node in result:
        vlan = mapping['vlan_by_switch'].get(node['name'])
        if not vlan or not vlan.get('subnet'):
            continue
        parsed = _parse_cidr(vlan['subnet'], context=f"OOB VLAN {vlan.get('id')}")
        if not parsed:
            continue
        net_ip, prefix = parsed
        base = net_ip.rsplit('.', 1)[0]
        try:
            net_last = int(net_ip.rsplit('.', 1)[1])
        except (ValueError, IndexError):
            net_last = 0
        idx = per_subnet_index.get(vlan['subnet'], 0)
        per_subnet_index[vlan['subnet']] = idx + 1
        node['svi_ip'] = _oob_switch_svi_ip(base, net_last, idx, prefix)
        gw = (vlan.get('gateway') or '').split('/')[0].strip()
        node['gateway'] = gw or f"{base}.{net_last + 1}"
        node['prefix'] = prefix

    return result


def categorize_nodes(nodes, settings=None, vlans=None):
    """Categorize nodes by their role type. OOB count comes from the Active oob-switch rows on the Nodes tab.

    Recognised role prefixes:
      core-*, csl-*                  → categories['core'] (csl is the
                                       non-collapsed equivalent of core)
      gsl-plane1-*, gsl-plane2-*     → categories['gsl_plane1' / 'gsl_plane2']
      oob-switch-*                   → categories['oob'] (via get_oob_nodes_for_inventory)
      su-N-node-N, gpu-NN            → categories['gpu_nodes']
      support-*, k8s-*, bcm-*,
        slurm-*, bcme-*              → categories['support']
      storage-*                      → categories['storage']
    """
    settings = settings or {}
    categories = {
        'core': [],
        'oob': [],
        'gsl_plane1': [],
        'gsl_plane2': [],
        'cl': [],
        'cs': [],
        'gl_plane1': [],
        'gl_plane2': [],
        'gs_plane1': [],
        'gs_plane2': [],
        'gpu_nodes': [],
        'support': [],
        'storage': [],
        'k8s': [],
    }
    for node in nodes:
        # Skip Disabled (operator-excluded) AND Air (documentary for
        # auto-injected infra) rows from category buckets — those rows
        # are not provisioned cluster nodes.
        if node['status'] in ('Disabled', 'Air'):
            continue
        cat = node.get('category')
        if cat in ('core', 'csl'):
            categories['core'].append(node)
        elif cat == 'gsl-plane1':
            categories['gsl_plane1'].append(node)
        elif cat == 'gsl-plane2':
            categories['gsl_plane2'].append(node)
        elif cat == 'cl':
            categories['cl'].append(node)
        elif cat == 'cs':
            categories['cs'].append(node)
        elif cat == 'gl-plane1':
            categories['gl_plane1'].append(node)
        elif cat == 'gl-plane2':
            categories['gl_plane2'].append(node)
        elif cat == 'gs-plane1':
            categories['gs_plane1'].append(node)
        elif cat == 'gs-plane2':
            categories['gs_plane2'].append(node)
        elif cat == 'oob-switch':
            # Collected below via get_oob_nodes_for_inventory
            pass
        elif cat == 'gpu':
            categories['gpu_nodes'].append(node)
        elif cat == 'support':
            categories['support'].append(node)
        elif cat == 'storage':
            categories['storage'].append(node)
    categories['oob'] = get_oob_nodes_for_inventory(nodes, settings, vlans)
    return categories


def generate_hosts_file(settings, nodes, output_dir, air_virtual_nodes=None, vlans=None):
    """Generate the Ansible hosts inventory file."""
    arch = settings.get('architecture', '2-4-3-200')
    categories = categorize_nodes(nodes, settings, vlans)
    air_virtual_nodes = air_virtual_nodes or set()
    
    # Switch census for the header. Counting only categories['core'] + ['oob']
    # under-reported every dedicated-GPU arch — 2-8-9-800 read "2 Core switches,
    # 2 OOB switches" while carrying 2 CSL + 4 GSL, and 2-4-5-800 omitted 8 of
    # its 10 switches. categories['core'] also lumps core/csl/cl together, so
    # split on node['category'] to name each role as the inventory groups do.
    _core_by_cat = Counter(n.get('category') or 'core' for n in categories['core'])
    _census = [
        ('Core', _core_by_cat.get('core', 0)),
        ('CSL', _core_by_cat.get('csl', 0)),
        ('CL', _core_by_cat.get('cl', 0)),
        ('CS', len(categories.get('cs') or [])),
        ('GSL', sum(len(categories.get(k) or [])
                    for k in ('gsl_plane1', 'gsl_plane2'))),
        ('GL', sum(len(categories.get(k) or [])
                   for k in ('gl_plane1', 'gl_plane2'))),
        ('GS', sum(len(categories.get(k) or [])
                   for k in ('gs_plane1', 'gs_plane2'))),
        ('OOB', len(categories.get('oob') or [])),
    ]
    _present = [f"{n} {label}" for label, n in _census if n]
    _total = sum(n for _, n in _census)
    _census_line = (f"# {_total} switches: " + ", ".join(_present)
                    if _present else "# no switches")

    lines = [
        "# ============================================================================",
        f"# {arch} Deployment Inventory (Generated from Excel)",
        "# ============================================================================",
        _census_line,
        "# ============================================================================",
        "",
    ]
    
    # Core / CSL switches — both routed into categories['core'] but emitted
    # under separate group names so the playbook can pick the right template.
    # Use node['name'] for inventory entries (hostnames) and node['category']
    # for the role filter.
    core_nodes = [n for n in categories['core'] if n.get('category') == 'core']
    csl_nodes  = [n for n in categories['core'] if n.get('category') == 'csl']
    if core_nodes:
        lines.append("[core]")
        for node in sorted(core_nodes, key=lambda x: x['name']):
            lines.append(node['name'])
        lines.append("")
    if csl_nodes:
        lines.append("[csl]")
        for node in sorted(csl_nodes, key=lambda x: x['name']):
            lines.append(node['name'])
        lines.append("")

    # CL group — split-role (non-converged) CPU/Storage leaf.
    if categories.get('cl'):
        lines.append("[cl]")
        for node in sorted(categories['cl'], key=lambda x: x['name']):
            lines.append(node['name'])
        lines.append("")

    # CS group — split-role (non-converged) CPU/Storage spine.
    if categories.get('cs'):
        lines.append("[cs]")
        for node in sorted(categories['cs'], key=lambda x: x['name']):
            lines.append(node['name'])
        lines.append("")

    # GSL plane groups (dedicated_gpu only)
    if categories.get('gsl_plane1'):
        lines.append("[gsl_plane1]")
        for node in sorted(categories['gsl_plane1'], key=lambda x: x['name']):
            lines.append(node['name'])
        lines.append("")
    if categories.get('gsl_plane2'):
        lines.append("[gsl_plane2]")
        for node in sorted(categories['gsl_plane2'], key=lambda x: x['name']):
            lines.append(node['name'])
        lines.append("")
    if categories.get('gsl_plane1') or categories.get('gsl_plane2'):
        lines.extend(["[gsl:children]"])
        if categories.get('gsl_plane1'):
            lines.append("gsl_plane1")
        if categories.get('gsl_plane2'):
            lines.append("gsl_plane2")
        lines.append("")

    # GL plane groups — split-role (non-converged) GPU leaf.
    if categories.get('gl_plane1'):
        lines.append("[gl_plane1]")
        for node in sorted(categories['gl_plane1'], key=lambda x: x['name']):
            lines.append(node['name'])
        lines.append("")
    if categories.get('gl_plane2'):
        lines.append("[gl_plane2]")
        for node in sorted(categories['gl_plane2'], key=lambda x: x['name']):
            lines.append(node['name'])
        lines.append("")
    if categories.get('gl_plane1') or categories.get('gl_plane2'):
        lines.extend(["[gl:children]"])
        if categories.get('gl_plane1'):
            lines.append("gl_plane1")
        if categories.get('gl_plane2'):
            lines.append("gl_plane2")
        lines.append("")

    # GS Spine plane groups — split-role (non-converged) GPU spine.
    if categories.get('gs_plane1'):
        lines.append("[gs_plane1]")
        for node in sorted(categories['gs_plane1'], key=lambda x: x['name']):
            lines.append(node['name'])
        lines.append("")
    if categories.get('gs_plane2'):
        lines.append("[gs_plane2]")
        for node in sorted(categories['gs_plane2'], key=lambda x: x['name']):
            lines.append(node['name'])
        lines.append("")

    # OOB switches
    if categories['oob']:
        lines.append("[oob]")
        for node in sorted(categories['oob'], key=lambda x: x['name']):
            lines.append(node['name'])
        lines.append("")
    
    # GPU nodes — use node['name'] (OEM name) so inventory_hostname matches devices dict key
    if categories['gpu_nodes']:
        lines.append("[nodes]")
        for node in sorted(categories['gpu_nodes'], key=lambda x: x['name'] or x['role']):
            lines.append(node['name'] or node['role'])
        lines.append("")

    # Storage
    if categories['storage']:
        lines.append("[storage]")
        for node in sorted(categories['storage'], key=lambda x: x['name'] or x['role']):
            lines.append(node['name'] or node['role'])
        lines.append("")

    # Support
    if categories['support']:
        lines.append("[support]")
        for node in sorted(categories['support'], key=lambda x: x['name'] or x['role']):
            lines.append(node['name'] or node['role'])
        lines.append("")
    
    # Air virtual node groups — mode-aware.
    # Both modes use the same group names so playbooks can target groups
    # (e.g., `hosts: dhcp`) without variable indirection. The group → host
    # mapping differs per mode:
    #   L2: [dhcp] dhcp-oob, [oob_server] oob-server-01, [jump] dhcp-oob
    #   L3: [dhcp] external-dhcp, [oob_server] external-conn, [jump] utility
    # The [jump] group is the Ansible SSH jump host (validate-* playbooks
    # run from this host). In L2 it co-locates with the DHCP server box;
    # in L3 they're separate Air nodes.
    dhcp_nodes = sorted(n for n in air_virtual_nodes if n.startswith('dhcp-'))
    oob_server_nodes = sorted(n for n in air_virtual_nodes if n.startswith('oob-server'))
    # L3 trio mapping into the existing role groups.
    if 'external-dhcp' in air_virtual_nodes:
        dhcp_nodes = (dhcp_nodes or []) + ['external-dhcp']
    if 'external-conn' in air_virtual_nodes:
        oob_server_nodes = (oob_server_nodes or []) + ['external-conn']
    # Jump host group: L2 reuses dhcp-oob, L3 uses utility.
    jump_nodes = []
    if 'utility' in air_virtual_nodes:
        jump_nodes = ['utility']
    elif 'dhcp-oob' in air_virtual_nodes:
        jump_nodes = ['dhcp-oob']

    if dhcp_nodes:
        lines.append("[dhcp]")
        lines.extend(sorted(set(dhcp_nodes)))
        lines.append("")

    if oob_server_nodes:
        lines.append("[oob_server]")
        lines.extend(sorted(set(oob_server_nodes)))
        lines.append("")

    if jump_nodes:
        lines.append("[jump]")
        lines.extend(jump_nodes)
        lines.append("")

    # Groups — switches:children gets every switch group that exists in this arch
    switch_children = []
    if core_nodes:
        switch_children.append("core")
    if csl_nodes:
        switch_children.append("csl")
    if categories.get('cl'):
        switch_children.append("cl")
    if categories.get('cs'):
        switch_children.append("cs")
    if categories.get('gsl_plane1') or categories.get('gsl_plane2'):
        switch_children.append("gsl")
    if categories.get('gl_plane1') or categories.get('gl_plane2'):
        switch_children.append("gl")
    if categories.get('gs_plane1'):
        switch_children.append("gs_plane1")
    if categories.get('gs_plane2'):
        switch_children.append("gs_plane2")
    if categories['oob']:
        switch_children.append("oob")
    lines.extend([
        "[switches:children]",
        *switch_children,
        "",
        "[switches:vars]",
        "ansible_user=cumulus",
        "",
    ])
    servers_children = [g for g in ["nodes", "storage", "support"] if categories.get({"nodes": "gpu_nodes"}.get(g, g))]
    if dhcp_nodes:
        servers_children.append("dhcp")
    if oob_server_nodes:
        servers_children.append("oob_server")
    if jump_nodes:
        # Keep [jump] in the [servers] tree so jump hosts inherit
        # ansible_password / ansible_user from group_vars/servers.yml.
        # Without this, L3-mode utility is orphaned and Ansible can't
        # fall back to password auth — every play that targets `hosts:
        # jump` fails when the operator's local SSH key is locked.
        servers_children.append("jump")
    lines.extend([
        "[servers:children]",
        *servers_children,
        "",
        "[servers:vars]",
        "ansible_user=ubuntu",
        "",
    ])
    
    output_file = output_dir / "hosts"
    with open(output_file, 'w') as f:
        f.write('\n'.join(lines))
    
    return output_file


def build_per_switch_gpu_rail_interfaces(wiremap_rows, switch_name, vlans,
                                         gpu_profile=None):
    """Return the physical per-rail GPU ports wired to one converged switch.

    ``Display in Air`` and node status control topology rendering only. Hidden
    or disabled scale-out rows are still physical cables and must remain in the
    generated switch config.
    """
    rail_vlan_map = {}
    for vlan in vlans or []:
        match = re.match(
            r'^gpu_rail(\d+)$',
            (vlan.get('name') or '').lower(),
        )
        if match and vlan.get('id'):
            rail_vlan_map[int(match.group(1))] = vlan['id']
    if not rail_vlan_map:
        return {}

    per_switch_rails = defaultdict(lambda: defaultdict(set))
    for row in wiremap_rows or []:
        if row.get('switch_name') != switch_name:
            continue
        profile = (row.get('net_profile') or '').lower()
        match = re.match(r'^gpu[\s_-]*rail[\s_-]*(\d+)$', profile)
        if not match:
            continue
        rail_idx = int(match.group(1))
        if rail_idx not in rail_vlan_map:
            continue
        switch_port = row.get('switch_port', '') or ''
        port_match = re.match(r'^swp(\d+)s(\d+)$', switch_port)
        if port_match:
            per_switch_rails[rail_idx][int(port_match.group(1))].add(
                int(port_match.group(2))
            )

    # Geometry comes from the GPU Port Profile row when the caller supplies it,
    # falling back to the _ROLE_HW constants only when it cannot. This builder
    # used the constants unconditionally, which meant the sheet's Speed column
    # never reached per-rail ports and every rail rendered breakout with no
    # `link speed` (ERA-55). Breakout/lanes were already correct by coincidence
    # — the constants happened to match — so only speed was visibly missing,
    # but reading the sheet makes all three follow the workbook.
    gpu_hw_defaults = _ROLE_HW.get('gpu', {})
    gpu_profile = gpu_profile or {}
    rail_breakout = gpu_profile.get('breakout') or gpu_hw_defaults.get('breakout', 2)
    rail_lanes = gpu_profile.get('lanes') or gpu_hw_defaults.get('lanes', 4)
    rail_speed = gpu_profile.get('speed') or gpu_hw_defaults.get('speed')
    result = {}
    for rail_idx, port_data in sorted(per_switch_rails.items()):
        overrides = {}
        for base, subports in port_data.items():
            active = sorted(subports)
            if active != list(range(rail_breakout)):
                overrides[base] = {'subports': active}
        result[f'rail{rail_idx}'] = {
            'ports': sorted(port_data),
            'breakout': rail_breakout,
            'lanes': rail_lanes,
            'vlan': rail_vlan_map[rail_idx],
            'state': 'up',
            'port_overrides': overrides,
        }
        if rail_speed:
            result[f'rail{rail_idx}']['speed'] = rail_speed
    return result


def _record_loopback_claims(claims, hostname, host_vars):
    """Accumulate every loopback /32 this host claims, keyed by semantic slot.

    Slots exist so that `lo_ip` and `router_id` — the same address by design —
    are one claim rather than a self-collision. Two DIFFERENT (host, slot)
    pairs claiming one address is the real defect (ERA-97 / GitLab #65).
    """
    def add(slot, value):
        if not value:
            return
        ip = str(value).split('/')[0].strip()
        if ip:
            claims.setdefault(ip, set()).add((hostname, slot))

    # Both are recorded rather than coalesced with `or`. Every assignment site
    # writes the pair from one source, so today they are always equal and the
    # shared 'lo' slot collapses them into a single claim. If a future path ever
    # sets them apart, coalescing would silently stop checking `router_id` --
    # blinding this guard to exactly the collision it exists to catch. Recording
    # both is correct either way and cannot produce a false positive.
    add('lo', host_vars.get('lo_ip'))
    add('lo', host_vars.get('router_id'))
    add('vrf:OOB', host_vars.get('oob_vrf_loopback'))
    add('vrf:GPU', host_vars.get('vrf_gpu_loopback'))
    for vrf_name, vrf_ip in (host_vars.get('vrf_loopbacks') or {}).items():
        add(f'vrf:{vrf_name}', vrf_ip)


def _assert_no_duplicate_loopbacks(claims):
    """Fail generation if any loopback /32 is claimed by two switches.

    This is deliberately AUTHORITY-NEUTRAL. `loopback_allocation.py` declares a
    block map, the workbook generator emits a slightly different one, and the
    blank-Loopbacks-sheet fallback in this file emits a third; which of the
    three is correct is an open decision (ERA-97 / D1). This check does not
    take a side — it only catches the harm all three can produce, which is two
    switches sharing a router-id. A duplicate /32 in an EVPN fabric fails
    silently at runtime, so failing loudly at generate time is strictly better.

    Known margins at the time of writing: the fallback path collides at 11 N/S
    leaves (largest shipped fabric has 8) and the workbook path at 21 OOB
    switches (largest shipped has 16). Both latent, neither guarded until now.
    """
    dupes = {ip: sorted(owners) for ip, owners in claims.items() if len(owners) > 1}
    if not dupes:
        return
    lines = ["Duplicate switch loopback /32 detected — refusing to generate.", ""]
    for ip, owners in sorted(dupes.items()):
        lines.append(f"  {ip} claimed by:")
        for host, slot in owners:
            lines.append(f"      {host}  ({slot})")
    lines += [
        "",
        "Two switches sharing a loopback means two VTEPs sharing a BGP router-id;",
        "the fabric would come up and misroute silently.",
        "",
        "Fix by giving each switch a distinct address in the 'Loopbacks & ASNs'",
        "sheet. If those cells are blank the value came from the fallback in",
        "excel_parser.py, whose per-role offsets are declared in",
        "loopback_allocation.py:20-27 (see ERA-97).",
    ]
    raise ValueError("\n".join(lines))


def generate_host_vars(nodes, vlans, output_dir, arch, settings, prefix_list_directives=None, oob_switch_configs=None, vrfs=None, air_settings=None, gsl_port_configs=None, loopback_overrides=None, wiremap_rows=None, physical_wiremap_rows=None, storage_uplink_ports=None, bond_descriptions_per_switch=None, per_switch_network_roles=None, per_switch_direct_interfaces=None, acl_directives=None, storage_uplink_hw=None, group_port_blocks=None,
                       gpu_profile_hw=None, custom_config=None):
    """Generate host_vars YAML files for each node. OOB count comes from the Active oob-switch rows on the Nodes tab. prefix_list_directives is the full {'lists': {...}, 'suppress': set()} dict from the Excel 'Prefix lists' sheet; generate_prefix_lists applies it scoped to overridable lists only. oob_switch_configs derived from Wire Map. gsl_port_configs (per-host dict from parse_gsl_port_config) controls which GSL ports are bridged/broken-out."""
    host_vars_dir = output_dir / "host_vars"
    host_vars_dir.mkdir(exist_ok=True)
    generated_files = []
    # ip -> {(hostname, slot)} — checked once every host is built (ERA-97).
    loopback_claims = {}
    categories = categorize_nodes(nodes, settings, vlans)
    nodes_to_process = (
        categories['core']
        + categories['oob']
        + categories.get('cl', [])
        + categories.get('cs', [])
        + categories['gsl_plane1'] + categories['gsl_plane2']
        + categories.get('gl_plane1', []) + categories.get('gl_plane2', [])
        + categories.get('gs_plane1', []) + categories.get('gs_plane2', [])
        + categories['gpu_nodes']
        + categories['storage'] + categories['support']
    )
    loopback_base = str(settings.get('loopback_base') or LOOPBACK_BASE).strip()

    # Air management subnet for switch eth0 IPs
    air_settings = air_settings or {}
    air_mgmt_subnet = air_settings.get('air_mgmt_subnet', '172.20.0.0/24')
    air_mgmt_base = air_mgmt_subnet.split('/')[0].rsplit('.', 1)[0]
    settings['_air_mgmt_base'] = air_mgmt_base  # pass to switch host_vars generation
    # ERA-48: 'Deploy in Air' (Settings, default Yes). When set to No the
    # operator is generating configs to push to real hardware locally — the
    # inventory must then use each switch's real Nodes-tab mgmt IP, NOT an
    # auto-assigned air-mgmt (172.20.0.x) IP the local network can't reach.
    deploy_in_air = str(settings.get('deploy_in_air', 'Yes')).strip().lower() in ('yes', 'true', '1')
    # Switch eth0 mgmt IPs live in the air-mgmt /24. Historically assigned
    # `.{200+idx}` starting at .201, which silently OVERFLOWS past .255 once
    # there are >~53 switches (maxscale 2-4-5-800 = 71, 2-8-9-800 SU32 = 76):
    # NVUE rejects the invalid octet (e.g. `172.20.0.260/24`), the deferred
    # apply aborts, and the switch never configures → unreachable.
    # Fix: keep the .201 base for archs that fit (byte-identical small/default
    # output), but repack from .2 when .201+N would overflow; skip the IPs the
    # air-deploy L3 trio + cust-net-edge reserve on this /24.
    _AIR_MGMT_RESERVED = AIR_MGMT_RESERVED_OCTETS  # canonical set: scripts/oob_reserved.py

    # Operator-pinned switch eth0 addresses. A brownfield operator often knows
    # the real management addresses their switches already answer on, and wants
    # the sim to reproduce them rather than accept our .201+ walk. A switch
    # whose Nodes-tab mgmt IP sits INSIDE air_mgmt_subnet is therefore honoured
    # verbatim; the auto-assign walk then fills the remaining switches around
    # those pins so a pin can never be handed out twice. A mgmt IP OUTSIDE the
    # subnet is rejected by validate-excel (it would be unreachable) rather
    # than silently overridden here.
    _pinned_ips = {}
    _pinned_octets = set()
    try:
        _air_net = ipaddress.IPv4Network(str(air_mgmt_subnet).strip(), strict=False)
    except ValueError:
        _air_net = None
    if _air_net is not None:
        for _n in nodes_to_process:
            if not is_switch(_n['role']):
                continue
            _raw = str(_n.get('mgmt_ip') or '').strip().split('/')[0]
            if not _raw:
                continue
            try:
                _addr = ipaddress.IPv4Address(_raw)
            except ValueError:
                continue  # malformed IP already reported by validate-excel
            if _addr in _air_net:
                _pinned_ips[_n['name'] or _n['role']] = str(_addr)
                _pinned_octets.add(int(_addr) - int(_air_net.network_address))

    # Only the UNPINNED switches consume auto-assigned octets, so the .201-vs-.2
    # repack decision counts those — not every switch.
    _n_switches = sum(1 for _n in nodes_to_process
                      if is_switch(_n['role'])
                      and (_n['name'] or _n['role']) not in _pinned_ips)
    _switch_octet = 201 if (201 + _n_switches) <= 254 else 2

    for node in nodes_to_process:
        role = node['role']
        
        # Build host vars
        host_vars = {
            'ansible_host': node['mgmt_ip'],
            # Use the Excel Name column as hostname — this is the OEM/customer name.
            # Config files and DHCP hostname use this value so ZTP downloads match
            # what Air names the VM (from the topology JSON node name).
            'hostname': node['name'] or role,
        }
        
        # Switches ZTP via eth0 on the air-mgmt subnet.
        # Auto-assign air-mgmt IPs (.200+ for switches) and eth0 MACs.
        if is_switch(role):
            topo_name = node['name'] or role
            host_vars['mac_address'] = generate_mac(topo_name, 'eth0')
            if deploy_in_air:
                host_vars['ip_assignment_mode'] = 'dhcp'
                if topo_name in _pinned_ips:
                    # Operator pinned this switch on the air-mgmt plane.
                    host_vars['ansible_host'] = _pinned_ips[topo_name]
                else:
                    while (_switch_octet in _AIR_MGMT_RESERVED
                           or _switch_octet in _pinned_octets):
                        _switch_octet += 1
                    if _switch_octet > 254:
                        raise ValueError(
                            f"air_mgmt subnet {air_mgmt_subnet} too small for "
                            f"{_n_switches} switches (ran past .254). Widen the subnet."
                        )
                    host_vars['ansible_host'] = f"{air_mgmt_base}.{_switch_octet}"
                    _switch_octet += 1
            else:
                # ERA-48: Deploy in Air = No -> local push to real hardware.
                # Keep the real Nodes-tab mgmt IP (already set on host_vars) and
                # mark the switch statically addressed instead of Air-DHCP.
                host_vars['ip_assignment_mode'] = 'static'
            # Control-plane ACLs (ADR-0030): tool-owned defaults (universal
            # dos/whitelist bindings + spine-only whitelist rule) plus the
            # optional 'ACLs' sheet override/add/suppress. Set for every switch
            # role so the gl/spine/oob templates render them too (not just core).
            host_vars['acls'] = generate_acls(node.get('category'), acl_directives)
        # ADR-0055: operator lines land last in the generated .sh, so they win on any
        # key the tool also sets. Absent sheet -> absent key -> templates emit nothing.
        _custom = (custom_config or {}).get(str(node.get('name') or '').strip())
        if _custom:
            host_vars['custom_config'] = list(_custom)
            # Key paths custom config touches. validate-config suppresses these on BOTH
            # sides of the comparison — see custom_config_keypaths().
            host_vars['custom_config_keypaths'] = custom_config_keypaths(_custom)

        # Add VLAN interfaces for core/csl switches (csl == core in
        # dedicated_gpu designs — same template, same vars).
        # CSL switches host CPU/storage/support/OOB traffic only; the GPU
        # plane lives on separate GSL switches, so we filter GPU-VRF VLANs
        # (and the GPU VRF loopback) out of csl host_vars.
        if node.get('category') in ('core', 'csl', 'cl'):
            is_csl = node.get('category') in ('csl', 'cl')
            # Use sequential position within the core list for SVI IP assignment,
            # not hostname-derived index (which can collide for OEM names that
            # share a trailing digit, e.g. two leaves both resolving to index 6).
            # CSL nodes live in categories['core'] regardless of their category field.
            core_list = categories.get('core', [])
            core_num = next((i + 1 for i, n in enumerate(core_list) if n is node), node.get('index') or 1)
            # Excel Loopbacks sheet wins when present; fall back to computed.
            sw_loop = (loopback_overrides or {}).get(node['name'] or role, {})
            if sw_loop.get('lo'):
                host_vars['lo_ip'] = _ensure_mask(sw_loop['lo'])
                host_vars['router_id'] = sw_loop['lo'].split('/')[0]
            else:
                host_vars['router_id'] = f"{loopback_base}.{10 + core_num}"
                host_vars['lo_ip'] = f"{loopback_base}.{10 + core_num}/32"

            # On a dedicated cl/cs spine-leaf (ns_tiers > 1) the underlay
            # is eBGP Clos, so each cl leaf needs a UNIQUE ASN — otherwise routes
            # from one cl, via a cs spine, to another cl hit an AS-path loop and
            # never install (cl↔cs sessions also idle on the iBGP/eBGP mismatch).
            # Converged csl (ns_tiers=1) leaves the ASN unset → shared base (iBGP),
            # which is golden and unchanged. Spines already get CSL_SPINE_ASN_OFFSET.
            if int((settings or {}).get('ns_tiers') or 1) > 1 and node.get('category') in ('csl', 'cl'):
                # The core template renders `router bgp autonomous-system {{ bgp_asn }}`
                # (default VRF + all VRFs), so override bgp_asn per-host (host_vars win
                # over group_vars/csl.yml) to give each cl leaf a unique switch ASN.
                _asn_ovr = (sw_loop or {}).get('asn')
                host_vars['bgp_asn'] = (int(_asn_ovr) if _asn_ovr else
                    asn_alloc.csl_leaf_asn(int(settings.get('bgp_asn') or DEFAULT_BASE_ASN), core_num))

            vlan_interfaces = []
            for vlan in vlans:
                if not vlan['subnet']:
                    continue
                if is_csl and vlan.get('vrf') == 'GPU':
                    continue  # GPU plane lives on GSL; skip for CSL
                if is_csl and vlan.get('vrf') == 'OOB':
                    continue  # OOB VLANs live on OOB switches; skip for CSL
                # Per-switch SVI IP + anycast VRR derived from the declared
                # subnet's host range and the Excel gateway, so non-.0-aligned
                # subnets / non-.1 gateways stay inside the network.
                pfx = vlan['subnet'].split('/')[1]
                svi_ip = _svi_switch_ip(vlan['subnet'], vlan.get('gateway'), core_num)
                vrr_ip = _svi_gateway_ip(vlan['subnet'], vlan.get('gateway'))

                vlan_interfaces.append({
                    'id': f"vlan{vlan['id']}",
                    'ip': f"{svi_ip}/{pfx}",
                    'vrr': f"{vrr_ip}/{pfx}",
                    'vlan': str(vlan['id']),
                    'vrf': vlan['vrf'],  # Use VRF from VLAN definition
                })

            if _normalize_oob_uplink_mode(settings) == 'l3':
                for vi in vlan_interfaces:
                    if vi.get('vrf') == 'OOB':
                        vi.pop('ip', None)
                        vi.pop('vrr', None)
                        vi['no_svi_type'] = True

            if vlan_interfaces:
                host_vars['vlan_interfaces'] = vlan_interfaces

            # Per-spine per-rail GPU port assignment. The group-level
            # `gpu_rail_interfaces` aggregates all rail ports across both
            # spines; for asymmetric per-rail layouts (e.g. prod-285200
            # where rails 1+3 land on spine-1 only and rails 2+4 on
            # spine-2 only), each spine must emit just its own rails.
            # Walk the Wire Map per-host and build a per-spine override.
            #
            # CSL switches in dedicated_gpu designs (e.g. 2-8-9-800) never
            # carry GPU traffic — those rails live on GSL switches — so
            # we skip per-rail port emission entirely for csl nodes.
            if not is_csl and (physical_wiremap_rows or wiremap_rows) and vlans:
                gpu_rail_interfaces = build_per_switch_gpu_rail_interfaces(
                    physical_wiremap_rows or wiremap_rows,
                    node.get('name'),
                    vlans,
                    gpu_profile=gpu_profile_hw,
                )
                if gpu_rail_interfaces:
                    host_vars['gpu_rail_interfaces'] = gpu_rail_interfaces

            # Generate VRF loopbacks first (computed + Loopbacks-sheet overrides)
            # so prefix-list generation can reference the post-override IPs.
            host_vars['vrf_loopbacks'] = generate_vrf_loopbacks(
                vlans, core_num, loopback_base, sw_loop, skip_gpu=is_csl)
            # Generate prefix_lists (loopback_base + Option C overrides; per-VRF
            # match rules track the override IPs from the Loopbacks sheet).
            # OOB subnets come from the OOB VLAN(s), resolved per deployment.
            # Needed so ERA_PREFIXES permits 192.168.200.0/24 outbound to the
            # cust-net-edge eBGP session (return path for L3 OOB NAT).
            oob_mapping = resolve_oob_vlans(vlans, categories['oob'])
            host_vars['prefix_list'] = generate_prefix_lists(
                vlans, core_num, loopback_base, prefix_list_directives,
                vrf_loopback_ips=host_vars['vrf_loopbacks'],
                oob_subnets=oob_mapping['subnets'],
                oob_uplink_mode=_normalize_oob_uplink_mode(settings),
            )

            host_name = node['name'] or role
            switch_direct = (per_switch_direct_interfaces or {}).get(host_name)
            if switch_direct is not None:
                host_vars.update(switch_direct)
                # The group value is a Wire Map union.  Override it before
                # deriving this switch's actual 8x-consumed neighbors.
                host_vars['interfaces_disabled'] = []

            # Disabled interfaces - read from settings or use defaults
            disabled_ports = settings.get('disabled_ports', '')
            if disabled_ports and switch_direct is None:
                host_vars['interfaces_disabled'] = [int(p.strip()) for p in str(disabled_ports).split(',')]
            elif switch_direct is None and arch in DEFAULT_DISABLED_INTERFACES:
                host_vars['interfaces_disabled'] = DEFAULT_DISABLED_INTERFACES[arch]

            # An 8x breakout consumes the adjacent even port's lanes, so that
            # port being disabled is a CONSEQUENCE of the breakout, not an
            # independent choice. Deriving it is the difference between one rule
            # and twelve workbooks that each have to remember it — four of them
            # did not, and the template answered by writing a `# WARNING:`
            # comment into the config rather than failing.
            _merged_blocks = dict(group_port_blocks or {})
            _merged_blocks.update(host_vars)
            _consumed = sorted(
                p + 1 for p in collect_8x_ports(_merged_blocks)
                if p % 2 == 1 and p + 1 <= MAX_SWITCH_PORT
            )
            if _consumed:
                _existing = list(
                    host_vars['interfaces_disabled']
                    if 'interfaces_disabled' in host_vars
                    else (group_port_blocks or {}).get('interfaces_disabled') or [])
                host_vars['interfaces_disabled'] = sorted(
                    {int(p) for p in _existing} | set(_consumed))

            # STORAGE VRF (PR-b): per-switch enablement. If the Wire Map
            # declared L3 Storage Uplink ports for this switch, emit:
            #   - storage_interfaces (ports + breakout shape; template
            #     consumes in PR-c for the L3 port-VRF binding line)
            #   - vrf_config_extra: the STORAGE VRF block (peer-group,
            #     neighbors, BGP) appended on top of group_vars vrf_config
            # See docs/plans/2026-05-19-storage-vrf-design.md.
            # Per-bond descriptions naming the connected node hostname.
            sw_bond_descs = (bond_descriptions_per_switch or {}).get(host_name, {})
            if sw_bond_descs:
                host_vars['bond_descriptions'] = sw_bond_descs
            # On the dedicated cl/cs tier, override the shared group
            # network_roles with this switch's OWN server bonds (cpu/storage/
            # support) carrying per-node ESIs. host_vars win over group_vars, so
            # each cl-* emits only its cabled sub-ports (no ifreload rollback)
            # and a node's ESI matches across its leaves (MH converges).
            # per_switch_network_roles contains cl-* (+ csl-* on dedicated tiers,
            # ns_tiers>1); converged csl-*/core (ns_tiers==1)
            # are absent and keep the golden group config.
            if per_switch_network_roles and host_name in per_switch_network_roles:
                host_vars['network_roles'] = per_switch_network_roles[host_name]
            sw_storage_ports = (storage_uplink_ports or {}).get(host_name, [])
            # A switch needs the STORAGE VRF *defined* whenever it hosts a
            # STORAGE-VRF SVI (the L2-stretched storage VLAN) OR terminates
            # L3 STORAGE uplinks. The SVI alone emits `interface vlanN vrf
            # STORAGE`; if the VRF is never defined, NVUE rejects the whole
            # config ("map ... to a non-existent vrf.STORAGE") and rolls the
            # apply back, leaving a bare switch. On dedicated-tier archs
            # (ns_tiers=2, e.g. 2-8-9-800 maxscale) the storage SVI lives on
            # the csl leaves while uplinks (if any) live elsewhere, so the two
            # conditions diverge — converged archs happened to satisfy both on
            # the same switch, which is why this only surfaced at max scale.
            has_storage_svi = any(
                (vi.get('vrf') or '').strip().upper() == 'STORAGE'
                for vi in host_vars.get('vlan_interfaces', [])
            )
            if sw_storage_ports or has_storage_svi:
                # Sort + dedup the port list for deterministic output.
                ordered_ports = sorted(set(sw_storage_ports))
                if sw_storage_ports:
                    # Map swpNsM → parent N for breakout extraction. Storage
                    # uplinks are typically 100G break-out of a 400G port
                    # (4x lanes/port); record the parent in a parents set.
                    parents = set()
                    for p in ordered_ports:
                        m = re.match(r'^swp(\d+)(?:s\d+)?$', p)
                        if m:
                            parents.add(int(m.group(1)))
                    # Shape comes from the Port Profiles row, not a constant —
                    # see find_storage_uplink_hw for why that matters.
                    _su_hw = storage_uplink_hw or _ROLE_HW['storage_uplink']
                    host_vars['storage_interfaces'] = {
                        'ports': sorted(parents),
                        'breakout': _su_hw['breakout'],
                        'lanes': _su_hw['lanes'],
                        'vrf': 'STORAGE',
                        'subports': ordered_ports,
                    }
                    if _su_hw.get('speed'):
                        host_vars['storage_interfaces']['speed'] = _su_hw['speed']
                # Build STORAGE vrf_config_extra entry. L3 VNI + L2 VLAN
                # come from the VRFs section (parse_vrfs) — pluck them
                # from the `vrfs` parameter if available. parse_vrfs
                # returns a {name: {...}} dict, not a list.
                storage_vrf_def = None
                if isinstance(vrfs, dict):
                    for vrf_name, vrf_meta in vrfs.items():
                        if str(vrf_name).strip().upper() == 'STORAGE':
                            storage_vrf_def = vrf_meta
                            break
                elif isinstance(vrfs, (list, tuple)):
                    for v in vrfs:
                        if isinstance(v, dict) and (v.get('name') or '').strip().upper() == 'STORAGE':
                            storage_vrf_def = v
                            break
                storage_l3_vni = (storage_vrf_def or {}).get('l3_vni') or 5005
                # The STORAGE VRF's EVPN L3VNI binds to its dedicated L3VNI
                # VLAN (the 300x symmetric-routing VLAN, e.g. 3005 — matching
                # OOB=3001/EXIT=3004), NOT the L2-stretched storage data VLAN
                # (e.g. 500). parse_vrfs stores the L3VNI VLAN under 'vlan'
                # (VRFs sheet col 4) — the same field the group vrf_config uses
                # for `evpn vlan`. Using the data VLAN here emitted a second,
                # conflicting `nv set vrf STORAGE evpn vlan 500` that overrode
                # the golden `vlan 3005` (last-wins in NVUE).
                storage_l3_vlan = (storage_vrf_def or {}).get('vlan')
                storage_entry = {
                    'id': 'STORAGE',
                    'vni': str(storage_l3_vni),
                    'route_export': True,
                }
                # L3 STORAGE uplinks need a kernel routing table for the
                # external eBGP underlay. Pure SVI-only leaves are an overlay
                # tenant VRF (like INBAND) and omit table_auto — keying it off
                # uplink presence keeps uplink switches byte-identical.
                if sw_storage_ports:
                    storage_entry['table_auto'] = True
                storage_bgp = {
                    'address_family': {
                        'ipv4_unicast': {
                            'enable': True,
                            'redistribute_connected': True,
                            'route_export_to_evpn': True,
                        },
                        'l2vpn_evpn': {'enable': True},
                    },
                }
                # Underlay eBGP (peer-group + unnumbered neighbors) only exists
                # where physical STORAGE uplinks terminate. SVI-only leaves
                # reach external storage over the EVPN overlay, no local peers.
                if sw_storage_ports:
                    storage_bgp['neighbors'] = [{
                        'interfaces': ordered_ports,
                        'peer_group': 'storage',
                        'type': 'unnumbered',
                    }]
                    storage_bgp['peer_groups'] = [{
                        'id': 'storage',
                        'remote_as': 'external',
                        'bfd_enable': True,
                        # ADR-0043: STORAGE gets its own BFD profile rather than
                        # borrowing underlay's, so its timers can diverge later
                        # without a fabric-wide change. Starts as a copy of
                        # underlay's 300/300/3, so this is rendered-config-only.
                        'bfd_profile': 'storage',
                        'description': 'underlay_era_storage_interconnect',
                        'address_family': {
                            'ipv4_unicast': {'enable': True},
                            'l2vpn_evpn': {'enable': True},
                        },
                    }]
                storage_entry['bgp'] = storage_bgp
                if storage_l3_vlan:
                    storage_entry['vlan'] = (int(storage_l3_vlan)
                                             if str(storage_l3_vlan).isdigit()
                                             else storage_l3_vlan)
                host_vars['vrf_config_extra'] = [storage_entry]

        # GSL (GPU Spine/Leaf) host vars — plane-aware
        # Plane comes from node['category'] (canonical: gsl-plane1 / gsl-plane2);
        # index from node['index'] (trailing digits of name, or order-fallback).
        gsl_cat = node.get('category')
        if gsl_cat in ('gsl-plane1', 'gsl-plane2', 'gl-plane1', 'gl-plane2'):
            plane_num = 1 if gsl_cat in ('gsl-plane1', 'gl-plane1') else 2
            plane_idx = node.get('index') or 1
            lo_oct = plane_idx
            host_vars['plane'] = plane_num
            # Excel Loopbacks sheet wins when present; fall back to computed.
            sw_loop = (loopback_overrides or {}).get(node['name'] or role, {})
            if sw_loop.get('lo'):
                host_vars['lo_ip'] = _ensure_mask(sw_loop['lo'])
            else:
                host_vars['lo_ip'] = (
                    f"10.{plane_num}.1.{plane_loopback_octet('leaf', lo_oct)}/32")
            if sw_loop.get('GPU'):
                host_vars['vrf_gpu_loopback'] = _ensure_mask(sw_loop['GPU'])
            else:
                host_vars['vrf_gpu_loopback'] = (
                    f"10.{plane_num}.1.{plane_loopback_octet('leaf_gpu', lo_oct)}/32")
            # Plane leaf count decides BOTH the ASN scheme and overlay peering.
            # Merge legacy (gsl_plane*) ∪ new (gl_plane*) buckets so a plane
            # declared with either taxonomy resolves the same peer set.
            gsl_leaves_in_plane = (categories.get(f'gsl_plane{plane_num}', [])
                                   + categories.get(f'gl_plane{plane_num}', []))
            # BGP ASN. A COLLAPSED (<=2-leaf) plane peers plane-mate↔plane-mate
            # via iBGP (`remote-as internal` in gl_nvue_cli.j2 for 1-tier), so
            # BOTH mates MUST share ONE ASN — matches the OG/golden REFERENCES
            # config (2-8-9-800: both gsl-plane1 leaves = 4260397297). Only a
            # SPINED (>2-leaf) plane is an eBGP Clos where each leaf needs a
            # UNIQUE ASN (it peers the spine via `remote-as external`). The
            # per-leaf +(plane_idx-1) offset must therefore apply to spined
            # planes ONLY — applying it to collapsed planes (commit 2d8837e)
            # gave mates different ASNs and left their iBGP sessions Idle.
            _asn_ovr = (sw_loop or {}).get('asn')
            host_vars['asn'] = (int(_asn_ovr) if _asn_ovr else
                asn_alloc.gsl_leaf_asn(int(settings.get('bgp_asn') or DEFAULT_BASE_ASN),
                                       plane_num, plane_idx, len(gsl_leaves_in_plane)))
            # Overlay peers: for 2-leaf planes, peer with the mate leaf.
            # For 4+ leaf planes, peer with the spine loopbacks instead.
            if len(gsl_leaves_in_plane) <= 2:
                # 2-leaf plane: peer with plane-mate.
                # The mate's address MUST come from the Loopbacks sheet when it
                # carries one (ADR-0033: that tab is the single home for per-node
                # loopbacks). This used to be the computed formula only, so a
                # switch honoured its OWN override while its mate still peered
                # the formula address — the session then never established, with
                # no error anywhere. The 4+-leaf branch below already resolved
                # spines this way; this is the same lookup for the mate.
                mate_idx = 2 if plane_idx == 1 else 1
                mate_lo = None
                for _mate in gsl_leaves_in_plane:
                    if _mate is node:
                        continue
                    if (_mate.get('index') or 1) == mate_idx or len(gsl_leaves_in_plane) == 2:
                        _mate_loop = (loopback_overrides or {}).get(_mate.get('name') or '', {})
                        if _mate_loop.get('lo'):
                            mate_lo = _mate_loop['lo'].split('/')[0]
                        break
                host_vars['plane_mate_lo_ip'] = (
                    mate_lo or f"10.{plane_num}.1.{plane_loopback_octet('leaf', mate_idx)}")
            else:
                # 4+ leaf plane: peer with spine loopbacks (not self, not other leaves)
                spine_nodes = categories.get(f'gs_plane{plane_num}', [])
                spine_los = []
                for sp in spine_nodes:
                    sp_loop = (loopback_overrides or {}).get(sp.get('name', ''), {})
                    if sp_loop.get('lo'):
                        spine_los.append(sp_loop['lo'].split('/')[0])
                    else:
                        sp_idx = sp.get('index') or 1
                        spine_los.append(
                            f"10.{plane_num}.1.{plane_loopback_octet('spine', sp_idx)}")
                host_vars['overlay_peers'] = spine_los
                # Keep plane_mate_lo_ip as fallback (won't be used if overlay_peers exists)
                host_vars['plane_mate_lo_ip'] = (
                    spine_los[0] if spine_los
                    else f"10.{plane_num}.1.{plane_loopback_octet('spine', 1)}")
            # vlan900 SVI: plane1=192.168.0.0/20, plane2=192.168.16.0/20.
            # Each leaf gets a distinct host IP within the /20 anycast group.
            vlan900_subnet_base = 0 if plane_num == 1 else 16
            host_vars['vlan900_ip'] = f"192.168.{vlan900_subnet_base}.{1 + plane_idx}/20"
            host_vars['vlan900_vrr'] = f"192.168.{vlan900_subnet_base}.1/20"

            # Per-rail SVI/VRR host_vars. Template iterates this dict with
            # gpu_rail_subports and emits one SVI + bridge-access line per
            # rail that this GSL switch physically owns.
            gpu_mode = str(settings.get('gpu_vlan_mode', 'single')).strip().lower()
            if gpu_mode == 'per_rail':
                rails = {}
                for v in vlans:
                    vname = (v.get('name') or '').lower()
                    m_r = re.match(r'^gpu_rail(\d+)$', vname)
                    if not m_r:
                        continue
                    rail_idx = int(m_r.group(1))
                    subnet = v.get('subnet')
                    vlan_id = v.get('id')
                    if not subnet or not vlan_id:
                        continue
                    net = ipaddress.ip_network(subnet, strict=False)
                    sw_offset = 1 + plane_idx
                    if sw_offset >= net.num_addresses - 1:
                        continue
                    sw_ip = f"{net.network_address + sw_offset}/{net.prefixlen}"
                    vrr = _svi_gateway_ip(subnet, v.get('gateway'))
                    if '/' not in vrr:
                        vrr = f"{vrr}/{net.prefixlen}"
                    rails[f'rail{rail_idx}'] = {
                        'vlan_id': vlan_id,
                        'vni': v.get('vni') or (vlan_id + 4000),
                        'subnet': subnet,
                        'ip': sw_ip,
                        'vrr': vrr,
                        'vrf': v.get('vrf') or 'GPU',
                    }
                if rails:
                    host_vars['gpu_rails'] = rails

            # Per-rail-per-plane SVI/VRR host_vars. When gpu_vlan_mode is
            # per_rail_per_plane, build a gpu_rail_planes dict containing
            # only the rails that belong to THIS GSL switch's plane.
            # Template iterates this dict and emits one SVI + VRR + bridge
            # access line per rail. Legacy vlan900 block stays defined as a
            # fallback but the template skips it when gpu_rail_planes exists.
            elif gpu_mode == 'per_rail_per_plane':
                rail_planes = {}
                for v in vlans:
                    vname = (v.get('name') or '').lower()
                    m_rp = re.match(r'^gpu_rail(\d+)_plane(\d+)$', vname)
                    if not m_rp:
                        continue
                    rail_idx = int(m_rp.group(1))
                    p_idx = int(m_rp.group(2))
                    if p_idx != plane_num:
                        continue  # other plane's rail — not on this switch
                    subnet = v.get('subnet')
                    vlan_id = v.get('id')
                    if not subnet or not vlan_id:
                        continue
                    net = ipaddress.ip_network(subnet, strict=False)
                    # Switch SVI host octet: 1 + plane_idx (per-switch unique
                    # within plane, anycast VRR on .1).
                    sw_offset = 1 + plane_idx
                    if sw_offset >= net.num_addresses - 1:
                        continue
                    sw_ip = f"{net.network_address + sw_offset}/{net.prefixlen}"
                    # VRR: use Excel gateway if provided, else .1 of subnet
                    vrr = _svi_gateway_ip(subnet, v.get('gateway'))
                    if '/' not in vrr:
                        vrr = f"{vrr}/{net.prefixlen}"
                    rail_planes[f'rail{rail_idx}_plane{p_idx}'] = {
                        'vlan_id': vlan_id,
                        'vni': v.get('vni') or (vlan_id + 4000),
                        'subnet': subnet,
                        'ip': sw_ip,
                        'vrr': vrr,
                        'vrf': v.get('vrf') or 'GPU',
                    }
                if rail_planes:
                    host_vars['gpu_rail_planes'] = rail_planes

            # Wire Map–derived per-host port config (must match topology JSON
            # exactly — referencing a port that's not in topology causes
            # ifreload-nvue to roll back the apply transaction). Dict is keyed
            # by switch hostname (System Name col) so the lookup works after
            # canonical conversion of the System Role column.
            gsl_cfg = (gsl_port_configs or {}).get(node.get('name') or role, {})
            for k in ('gpu_subports', 'gpu_breakout_parents',
                      'isl_subports', 'isl_breakout_parents',
                      'gpu_rail_subports', 'gpu_rail_plane_subports'):
                if gsl_cfg.get(k):
                    host_vars[k] = gsl_cfg[k]

        # CS (CPU/Storage Spine) host vars — loopback from Loopbacks sheet
        if node.get('category') == 'cs':
            cs_nodes = categories.get('cs', [])
            spine_idx = next((i for i, n in enumerate(cs_nodes, start=1) if n is node), 1)
            sw_loop = (loopback_overrides or {}).get(node['name'] or role, {})
            if sw_loop.get('lo'):
                host_vars['lo_ip'] = _ensure_mask(sw_loop['lo'])
                host_vars['router_id'] = sw_loop['lo'].split('/')[0]
            else:
                host_vars['lo_ip'] = f"{loopback_base}.{20 + spine_idx}/32"
                host_vars['router_id'] = f"{loopback_base}.{20 + spine_idx}"
            _asn_ovr = (sw_loop or {}).get('asn')
            host_vars['asn'] = (int(_asn_ovr) if _asn_ovr else
                asn_alloc.csl_spine_asn(int(settings.get('bgp_asn') or DEFAULT_BASE_ASN), spine_idx))
            isl_cfg = _derive_host_isl_port_config(
                wiremap_rows, node.get('name') or role, profile_names=['isl'])
            for k in ('isl_subports', 'isl_breakout_parents'):
                if isl_cfg.get(k):
                    host_vars[k] = isl_cfg[k]
            if host_vars.get('isl_breakout_parents'):
                host_vars['isl_leaf_ports'] = host_vars['isl_breakout_parents']
            if host_vars.get('isl_subports'):
                host_vars['underlay_neighbors'] = [
                    p for p in host_vars['isl_subports'].split(',') if p
                ]

            csl_leaf_peers = []
            # N/S leaves live in different buckets by tier: converged (core/csl)
            # land in categories['core']; 2-tier split leaves (cl) land in
            # categories['cl']. A cs spine only exists at 2-tier, where the
            # leaves are cl — so the overlay peer set MUST include
            # categories['cl']. Reading categories['core'] alone left the spine
            # with zero overlay peers → it was an EVPN RR for nobody → the cl
            # leaves' overlay sessions to the spine loopbacks never came up and
            # VLAN-200 (OOB) never stretched → servers behind OOB unreachable.
            core_nodes = categories.get('core', []) + categories.get('cl', [])
            for leaf_idx, leaf in enumerate(core_nodes, start=1):
                if leaf.get('category') not in ('core', 'csl', 'cl'):
                    continue
                leaf_loop = (loopback_overrides or {}).get(leaf.get('name') or leaf.get('role'), {})
                if leaf_loop.get('lo'):
                    csl_leaf_peers.append(leaf_loop['lo'].split('/')[0])
                else:
                    csl_leaf_peers.append(f"{loopback_base}.{10 + leaf_idx}")
            # A dedicated cs spine is the EVPN overlay HUB for both the
            # cl leaves (INBAND/storage VTEPs) AND the oob-switches (VLAN-200
            # VTEPs). Add the oob-switch loopbacks so oob↔cs↔oob forms and
            # VLAN-200 stretches (the oob side already peers cs via
            # _derive_core_overlay_peers; the spine must peer back). Converged
            # csl (ns_tiers = 1) has no cs spine so this block never runs there.
            if csl_leaf_peers and int((settings or {}).get('ns_tiers') or 1) > 1:
                csl_leaf_peers = csl_leaf_peers + _derive_oob_overlay_peers(
                    nodes=nodes, loopback_overrides=loopback_overrides)
            if csl_leaf_peers:
                host_vars['overlay_peers'] = csl_leaf_peers

        # GS-Spine (GPU E/W Spine) host vars — plane-aware, loopback from Loopbacks sheet
        if node.get('category') in ('gs-plane1', 'gs-plane2'):
            plane_num = 1 if node.get('category') == 'gs-plane1' else 2
            gs_spine_nodes = categories.get(f'gs_plane{plane_num}', [])
            spine_idx = next((i for i, n in enumerate(gs_spine_nodes, start=1) if n is node), 1)
            host_vars['plane'] = plane_num
            sw_loop = (loopback_overrides or {}).get(node['name'] or role, {})
            if sw_loop.get('lo'):
                host_vars['lo_ip'] = _ensure_mask(sw_loop['lo'])
                host_vars['router_id'] = sw_loop['lo'].split('/')[0]
            else:
                _sp_oct = plane_loopback_octet('spine', spine_idx)
                host_vars['lo_ip'] = f"10.{plane_num}.1.{_sp_oct}/32"
                host_vars['router_id'] = f"10.{plane_num}.1.{_sp_oct}"
            # Base-relative per-plane ASN (was hardcoded 4200101100/4200102100,
            # which ignored Settings.bgp_asn → wrong AS family for any arch whose
            # base ASN isn't 4200100001). Plane offset keeps it clear of the
            # gsl-leaf +1000 block and of OOB/csl-spine.
            _asn_ovr = (sw_loop or {}).get('asn')
            host_vars['asn'] = (int(_asn_ovr) if _asn_ovr else
                asn_alloc.gsl_spine_asn(int(settings.get('bgp_asn') or DEFAULT_BASE_ASN), plane_num))
            gsl_cfg = (gsl_port_configs or {}).get(node.get('name') or role, {})
            for k in ('isl_subports', 'isl_breakout_parents'):
                if gsl_cfg.get(k):
                    host_vars[k] = gsl_cfg[k]
            if host_vars.get('isl_breakout_parents'):
                host_vars['isl_ports'] = host_vars['isl_breakout_parents']
            if host_vars.get('isl_subports'):
                host_vars['underlay_neighbors'] = [
                    p for p in host_vars['isl_subports'].split(',') if p
                ]

            gsl_leaf_peers = []
            _plane_leaves = (categories.get(f'gsl_plane{plane_num}', [])
                             + categories.get(f'gl_plane{plane_num}', []))
            for leaf_idx, leaf in enumerate(_plane_leaves, start=1):
                leaf_loop = (loopback_overrides or {}).get(leaf.get('name') or leaf.get('role'), {})
                if leaf_loop.get('lo'):
                    gsl_leaf_peers.append(leaf_loop['lo'].split('/')[0])
                else:
                    leaf_num = leaf.get('index') or leaf_idx
                    gsl_leaf_peers.append(
                        f"10.{plane_num}.1.{plane_loopback_octet('leaf', leaf_num)}")
            if gsl_leaf_peers:
                host_vars['overlay_peers'] = gsl_leaf_peers

        # Add OOB switch specific variables derived from Wire Map
        if node.get('category') == 'oob-switch':
            # SVI IP on VLAN 200 — pre-computed by get_oob_nodes_for_inventory()
            # (sequential .2/.3/.4 if single subnet, or per-switch if multi-subnet)
            host_vars['svi_ip'] = f"{node.get('svi_ip', node['mgmt_ip'])}/{node['prefix']}"
            host_vars['default_gateway'] = node['gateway']

            oob_switch_configs = oob_switch_configs or {}
            oob_key = node.get('name') or role
            if oob_key in oob_switch_configs:
                cfg = oob_switch_configs[oob_key]
                host_vars['access_ports'] = cfg['access_ports']
                host_vars['uplink_ports'] = cfg['uplink_ports']
                host_vars['spine_bond_members'] = cfg['spine_bond_members']
            else:
                # Fallback: no Wire Map data for this switch. These are the
                # reference-topology pins, NOT anything the operator asked for —
                # say so loudly rather than emitting them as if they were derived.
                host_vars['access_ports'] = 'swp1-48'
                host_vars['uplink_ports'] = 'swp1-49,swp51'
                host_vars['spine_bond_members'] = ['swp49', 'swp51']
                print(f"  ⚠️  No Wire Map rows for OOB switch '{oob_key}' — falling back to "
                      f"reference port pins (access swp1-48, uplinks swp49/swp51). "
                      f"Add its rows to the Wire Map if the real cabling differs.",
                      file=sys.stderr)

            if _normalize_oob_uplink_mode(settings) == 'l3':
                # Sequential index avoids ASN collision when hostnames
                # share trailing digits (e.g. two OOB switches whose names
                # both end in the same digit → same derived index). Dedicated spines live in
                # their own ASN block (CSL_SPINE_ASN_OFFSET), so OOB sits
                # directly above the base (base+1..) — no spine collision.
                oob_nodes = [n for n in nodes_to_process if n.get('category') == 'oob-switch']
                oob_idx = next((i for i, n in enumerate(oob_nodes, start=1) if n is node), 1)
                sw_loop = (loopback_overrides or {}).get(oob_key, {})

                # Excel Loopbacks sheet wins; otherwise compute from
                # loopback_base with a fixed offset reserved for OOB
                # switches (30+oob_idx; cores live at 10+core_num).
                if sw_loop.get('lo'):
                    host_vars['lo_ip'] = _ensure_mask(sw_loop['lo'])
                    host_vars['router_id'] = sw_loop['lo'].split('/')[0]
                else:
                    host_vars['router_id'] = f"{loopback_base}.{30 + oob_idx}"
                    host_vars['lo_ip'] = f"{loopback_base}.{30 + oob_idx}/32"

                _asn_ovr = (sw_loop or {}).get('asn')
                host_vars['bgp_asn'] = (int(_asn_ovr) if _asn_ovr else
                    asn_alloc.oob_asn(int(settings.get('bgp_asn', 65000)), oob_idx))

                host_vars['vrr_ip'] = f"{node['gateway']}/{node['prefix']}"

                # Per-switch OOB access VLAN. The Nodes sheet carries an
                # `OOB VLAN` column per switch (ADR-0028) and resolve_oob_vlans
                # already builds a vlan_by_switch map from it — but only the L3
                # side consumed it, to compute svi_ip/gateway. The access-port
                # bridge VLAN was resolved GLOBALLY here ("first VLAN named
                # OOB*") and then hardcoded to 200 in the template, so a switch
                # placed on VLAN 201 got an SVI on the 201 subnet while its host
                # ports stayed in VLAN 200 — BMCs in a different broadcast
                # domain from their own gateway, silently, because every piece
                # was individually valid.
                #
                # Honour the per-node value first, then the single declared OOB
                # VLAN, then 200. `oob_vni` follows the SAME VLAN rather than
                # being looked up independently, so the two cannot disagree.
                _node_vlan = str(node.get('oob_vlan', '') or '').strip()
                oob_vlan_id = None
                if _node_vlan:
                    try:
                        oob_vlan_id = int(float(_node_vlan))
                    except (TypeError, ValueError):
                        oob_vlan_id = None
                if oob_vlan_id is None:
                    oob_vlan_id = next((v['id'] for v in vlans
                                        if (v.get('name') or '').upper().startswith('OOB')), 200)
                host_vars['oob_access_vlan'] = int(oob_vlan_id)
                oob_vni = next((v.get('vni') for v in vlans
                                if v['id'] == oob_vlan_id and v.get('vni')), None)
                if oob_vni:
                    host_vars['oob_vni'] = int(oob_vni)

                oob_vrf = vrfs.get('OOB', {}) if vrfs else {}
                if oob_vrf.get('l3_vni'):
                    host_vars['oob_vrf_vni'] = int(oob_vrf['l3_vni'])
                if oob_vrf.get('vlan'):
                    host_vars['oob_vrf_vlan'] = int(oob_vrf['vlan'])
                elif oob_vrf.get('l3_vni'):
                    host_vars['oob_vrf_vlan'] = 3001

                # OOB-VRF loopback: Excel Loopbacks sheet wins; otherwise
                # compute from loopback_base (.40+oob_idx, distinct from the
                # switch's primary loopback at .30+oob_idx).
                if sw_loop.get('OOB'):
                    host_vars['oob_vrf_loopback'] = _ensure_mask(sw_loop['OOB'])
                else:
                    host_vars['oob_vrf_loopback'] = f"{loopback_base}.{40 + oob_idx}/32"

                host_vars['overlay_peers'] = _derive_core_overlay_peers(
                    nodes=nodes, loopback_overrides=loopback_overrides)

        
        # Write YAML file — filename must match the inventory hostname.
        # Use node['name'] (the hostname from the Name column) for every
        # node. Legacy Excels had role == name so this preserves prior
        # filenames; canonical Excels carry the hostname only in name.
        inv_hostname = assert_valid_inv_hostname(node['name'] or role)
        # Group port blocks belong to the CORE group only — applying them to an
        # OOB switch invents an edge uplink it does not have.
        _grp = group_port_blocks if node.get('category') in ('core', 'csl', 'cl') else None
        assert_valid_8x_breakout(inv_hostname, host_vars, _grp)
        _record_loopback_claims(loopback_claims, inv_hostname, host_vars)
        output_file = host_vars_dir / f"{inv_hostname}.yml"
        with open(output_file, 'w') as f:
            f.write("---\n")
            f.write(f"# Host variables for {inv_hostname} (Generated from Excel)\n")
            yaml.dump(host_vars, f, default_flow_style=False, sort_keys=False)
        
        generated_files.append(output_file)

    _assert_no_duplicate_loopbacks(loopback_claims)

    return generated_files


def _breakout_token(hw):
    """{'breakout': 2} -> '2x'. None when the sheet declares nothing.

    ERA-73. Returning None (rather than a default) keeps the template's own
    `| default('2x')` as the single place the fallback lives, so a workbook
    that declares nothing behaves exactly as before this change.
    """
    if not hw:
        return None
    raw = hw.get('breakout')
    if raw in (None, ''):
        return None
    try:
        n = int(str(raw).strip().rstrip('xX'))
    except (TypeError, ValueError):
        return None
    return f'{n}x' if n > 1 else None


def generate_group_vars(settings, vlans, vrfs, output_dir, arch, nodes=None, port_config=None,
                        node_oob_mapping=None, versions=None, wiremap_rows=None,
                        air_settings=None, dhcp_relay_table=None,
                        loopback_overrides=None, route_policy_directives=None,
                        community_list_directives=None, plane_breakout=None):
    """Generate group_vars YAML files.

    route_policy_directives / community_list_directives:
    optional {route_map_id/community_list_id: rule_list} dicts from the
    'Route policy' / 'Community lists' Excel sheets, as returned by
    parse_route_policy_sheet / parse_community_lists_sheet. Absent (None)
    ⇒ the source-inventory route_map/community_list merged below are used
    unchanged (derive-by-default). Present ⇒ same-named entries are
    replaced and new ones appended via `_apply_named_overrides`.
    """
    group_vars_dir = output_dir / "group_vars"
    group_vars_dir.mkdir(exist_ok=True)
    generated_files = []
    
    # all.yml (ntp_servers here so core and oob both get it from group_vars/all)
    # Accept comma- or newline-separated (Excel may show one per line with wrap_text)
    ntp_str = settings.get('ntp_servers', '')
    ntp_entries = [s.strip() for s in re.split(r'[,\n]', str(ntp_str)) if ntp_str and s.strip()] if ntp_str else []
    if not ntp_entries:
        ntp_entries = [
            '0.cumulusnetworks.pool.ntp.org',
            '1.cumulusnetworks.pool.ntp.org',
            '2.cumulusnetworks.pool.ntp.org',
            '3.cumulusnetworks.pool.ntp.org',
        ]
    # Each entry is `<address>` or `<address> [association-type] <type>`
    # (ERA-59). NVUE's default is `server`, and we now emit it explicitly on
    # every role so the config itself shows an operator where `pool` or `peer`
    # would go — previously only the spine template emitted the suffix, and
    # nothing could express anything but `server`.
    #
    # The address and the type are kept in SEPARATE structures on purpose.
    # `ntp_servers` stays a flat list of bare addresses so it renders through
    # `| quote` unchanged and any hand-written inventory that overrides it keeps
    # working; the type is looked up beside it and constrained to a closed set,
    # so neither half can carry a spaced or injected value into a root-executed
    # config script. `ntp_servers` is a SHELL_INJECTION_PRONE_KEY.
    ntp_list = []
    ntp_association_types = {}
    for entry in ntp_entries:
        parts = entry.split()
        address = parts[0]
        rest = [p for p in parts[1:] if p.lower() != 'association-type']
        assoc = rest[0].lower() if rest else 'server'
        if assoc not in _VALID_NTP_ASSOCIATION_TYPES:
            print(f"  ⚠️  Settings.ntp_servers: {entry!r} declares association type "
                  f"{assoc!r}; expected one of "
                  f"{', '.join(sorted(_VALID_NTP_ASSOCIATION_TYPES))}. Using 'server'.",
                  file=sys.stderr)
            assoc = 'server'
        ntp_list.append(address)
        ntp_association_types[address] = assoc
    # `tiers` was split into ns_tiers (compute) + ew_tiers (GPU). A bare legacy
    # `tiers` seeds both for back-compat; warn so the operator migrates.
    if settings.get('tiers') is not None and \
            settings.get('ns_tiers') is None and settings.get('ew_tiers') is None:
        print("  ⚠️  Settings.tiers is deprecated — use ns_tiers (compute) and "
              "ew_tiers (GPU); seeding both from tiers for now.",
              file=sys.stderr)

    # NOTE: scalable_units / nodes_per_su are deliberately NOT emitted. Neither
    # has ever appeared in a shipped workbook, so they always fell back to the
    # hardcoded 8/4 — which is simply wrong for most archs (2-8-9-800 has 2 SUs,
    # not 8). Nothing read them, so the only effect was a false fact sitting in
    # the generated inventory. If a consumer ever needs them, derive from the
    # Nodes sheet (`su-<N>-node-<M>`) rather than reintroducing a default.
    all_vars = {
        'architecture': settings.get('architecture', '2-4-3-200'),
        'ns_tiers': int(settings.get('ns_tiers', settings.get('tiers', 1)) or 1),
        'ew_tiers': int(settings.get('ew_tiers', settings.get('tiers', 1)) or 1),
        'convergence': settings.get('convergence', 'full'),
        'gpu_planes': int(settings.get('gpu_planes', 1)),
        'oob_uplink_mode': _normalize_oob_uplink_mode(settings),
        'ntp_servers': ntp_list,
        'ntp_association_types': ntp_association_types,
    }

    # Add VLAN/network info
    common = {}
    for vlan in vlans:
        # Normalize VLAN name to valid identifier (e.g. "cpu/in-band" → "cpu")
        raw_name = vlan['name'].lower() if vlan['name'] else f"vlan{vlan['id']}"
        name_key = raw_name.split('/')[0].replace('-', '_').replace(' ', '_')
        if vlan['subnet']:
            subnet_parts = vlan['subnet'].split('/')
            base_ip = subnet_parts[0].rsplit('.', 1)[0]
            common[f"{name_key}_network"] = vlan['subnet']
            # Prefer the Excel Gateway column; fall back to base.1 if empty.
            common[f"{name_key}_gateway"] = vlan.get('gateway') or f"{base_ip}.1"
            common[f"{name_key}_vlan"] = vlan['id']
            # Per-VLAN VRF assignment from the Excel. Consumed by
            # validate-ping-matrix to classify cross-VRF pings correctly:
            # on archs where Storage lives in its own STORAGE VRF (2-4-5-800,
            # 2-8-9-800), compute→storage pings are CROSS-VRF and expected
            # to fail. On legacy archs where Storage is merged into INBAND,
            # they're same-VRF and expected to pass.
            common[f"{name_key}_vrf"] = (vlan.get('vrf') or 'default').strip() or 'default'

    # Storage server VLAN tagging signal (consumed by air-deploy.py server
    # netplan). The switch storage port is untagged/access when the Storage
    # role carries only the storage VLAN (e.g. `access 500` on collapsed-core
    # archs) — the server must put the storage IP on the raw bond. It is a
    # tagged trunk member when the role trunks multiple VLANs and the storage
    # VLAN is not the native/untagged one (e.g. `vlan 400,500` untagged 300 on
    # 2-4-5-800) — the server must tag via bond.<vlan>. Mismatch = frames
    # dropped at the switch port. Derived from the Storage network role's
    # actual VLAN membership; defaults to untagged when unknown (the
    # historically-safe collapsed-core behavior).
    storage_role = ((port_config or {}).get('network_roles') or {}).get('storage') or {}
    if 'storage_vlan' in common:
        role_vlan = str(storage_role.get('vlan') or '').strip()
        common['storage_vlan_tagged'] = bool(role_vlan) and role_vlan != str(common['storage_vlan'])

    # Optional operator-configurable SSH login banners (pre + post-login).
    # Empty → templates emit no `nv set system message ...` line and leave
    # any existing banner on the switch alone. Multi-line content is
    # preserved verbatim. Templates substitute {hostname}/{site}/{arch}
    # per-switch at render time. See
    # docs/plans/2026-05-26-switch-login-messages-design.md.
    common['pre_login_message'] = str(settings.get('pre_login_message') or '')
    common['post_login_message'] = str(settings.get('post_login_message') or '')

    # Site + arch into common so templates can substitute them into the
    # banner placeholders. `architecture` is the Settings field name;
    # default to the `arch` function argument when missing.
    common['site'] = str(settings.get('site_name') or 'default').strip()
    common['arch'] = str(settings.get('architecture') or arch).strip()

    all_vars['common'] = common

    # LDAP (#3): read ldap_* settings and generate ldap block
    ldap_enabled = str(settings.get('ldap_enabled', 'No')).strip().lower() in ('yes', 'true', '1')
    ldap_servers_str = str(settings.get('ldap_servers', '')).strip()
    ldap_server_ips = [s.strip() for s in ldap_servers_str.split(',') if s.strip()] if ldap_servers_str else []
    all_vars['ldap'] = {
        'enabled': ldap_enabled,
        'domain': settings.get('ldap_domain', 'example.com'),
        'organization': settings.get('ldap_organization', 'Example Org'),
        'admin_password': '{{ ldap_admin_password }}',
        'base_dn': settings.get('ldap_base_dn', ''),
        'root_dn': settings.get('ldap_root_dn', ''),
        'servers': [{'ip': ip, 'priority': i + 1} for i, ip in enumerate(ldap_server_ips)],
    }

    # LDAP users (default placeholder accounts)
    all_vars['ldap']['users'] = [
        {'firstname': 'John', 'lastname': 'Doe', 'username': 'jdoe', 'password': '{{ ldap_user_default_password }}'},
        {'firstname': 'Alice', 'lastname': 'Smith', 'username': 'asmith', 'password': '{{ ldap_user_default_password }}'},
    ]

    # Status page toggle (for Air HTTP service + nginx basic auth)
    status_page_enabled = str(settings.get('status_page_enabled', 'No')).strip().lower() in ('yes', 'true', '1')
    all_vars['status_page_enabled'] = status_page_enabled

    # Build devices dict from Nodes tab (for DHCP reservations + server netplan config)
    # GPU VLAN topology mode. Three values:
    #   single             — one GPU VLAN (default; legacy behavior)
    #   per_rail           — one VLAN per rail; gpu_rail<N> VLAN rows
    #   per_rail_per_plane — one VLAN per (rail, plane); gpu_rail<R>_plane<P> rows
    # See docs/plans/2026-05-18-gpu-plane-per-rail.md.
    gpu_vlan_mode = str(settings.get('gpu_vlan_mode', 'single')).strip().lower()
    if gpu_vlan_mode not in ('single', 'per_rail', 'per_rail_per_plane'):
        gpu_vlan_mode = 'single'
    devices = build_devices(nodes or [], vlans, node_oob_mapping, wiremap_rows,
                            gpu_vlan_mode=gpu_vlan_mode)
    if devices:
        all_vars['host_dhcp'] = True
        all_vars['devices'] = devices
        auto_mac_count = sum(1 for d in devices.values() if d.get('mac', '').startswith('48:b0:2d:'))
        # Count hosts missing data-plane IPs due to overflow or unknown role
        no_dataplane = sum(1 for d in devices.values()
                          if not any(k in d for k in ('bond_ip', 'bond_ip1', 'gpu_ip1')))
        print(f"  Devices: {len(devices)} hosts ({auto_mac_count} with auto-generated MACs)")
        if no_dataplane:
            roles_without = set()
            for name, d in devices.items():
                if not any(k in d for k in ('bond_ip', 'bond_ip1', 'gpu_ip1')):
                    roles_without.add(classify_host_role(name))
            print(f"    {no_dataplane} hosts without data-plane IPs (roles: {', '.join(sorted(roles_without))})")

    all_vars['switch_user'] = 'cumulus'

    # Mode-aware Ansible target hosts. L2 default uses the dhcp-oob trio;
    # L3 uses the external-conn / external-dhcp / utility Ubuntu trio per
    # docs/plans/2026-05-20-l3-oob-air-topology.md.
    _mode = _normalize_oob_uplink_mode(settings)
    if _mode == 'l3':
        all_vars['ztp_server_host'] = 'external-dhcp'
        all_vars['dhcp_server_host'] = 'external-dhcp'
        all_vars['jump_host'] = 'utility'
        all_vars['ansible_target'] = 'utility'
        all_vars['nat_host'] = 'external-conn'
    else:
        all_vars['ztp_server_host'] = 'dhcp-oob'
        all_vars['dhcp_server_host'] = 'dhcp-oob'
        all_vars['jump_host'] = 'dhcp-oob'
        all_vars['ansible_target'] = 'oob-server-01'

    # Generate ztp_interfaces from Air Management Subnet (for switch ZTP on air-oob-switch)
    air_settings = air_settings or {}
    air_mgmt_subnet = air_settings.get('air_mgmt_subnet', '172.20.0.0/24')
    _parsed_air = _parse_cidr(air_mgmt_subnet, context="Air_Only.air_mgmt_subnet")
    if not _parsed_air:
        # Fall back to default if the operator entered garbage.
        air_mgmt_subnet = '172.20.0.0/24'
        _parsed_air = _parse_cidr(air_mgmt_subnet)
    _air_ip, air_mgmt_prefix = _parsed_air
    air_mgmt_base = _air_ip.rsplit('.', 1)[0]  # e.g., "172.20.0"
    all_vars['air_mgmt_subnet'] = air_mgmt_subnet

    # Build ztp_interfaces: eth1 (air-mgmt for switch ZTP) + ethN per OOB VLAN subnet.
    # In L3 mode external-dhcp only needs the air-mgmt interface (cust-net-edge
    # bridge) — server-side DHCP on VLAN 200 is hosted by utility, separately.
    # Its DHCP gateway is the cust-net-edge-01 bridge SVI (.254); external-conn
    # sits on separate routed EXIT egress legs, not on this management subnet.
    air_mgmt_gateway = (
        f"{air_mgmt_base}.254" if _mode == 'l3' else f"{air_mgmt_base}.1"
    )
    ztp_ifaces = [{
        'name': 'eth1',
        'ip': f"{air_mgmt_base}.77",
        'network': air_mgmt_subnet,
        'gateway': air_mgmt_gateway,
        'purpose': 'air-mgmt',
        'dnsmasq_listen': True,
    }]
    # OOB VLAN subnets (used for per-VLAN ZTP interfaces below, L2 only, and
    # for the nginx allow-list further down regardless of mode).
    oob_subnets_list = resolve_oob_vlans(vlans, get_oob_nodes_for_inventory(nodes or [], settings, vlans))['subnets']
    if _mode != 'l3':
        # Per-VLAN interfaces from OOB VLAN subnets (L2 only — see comment above).
        for i, subnet_str in enumerate(oob_subnets_list):
            parsed = _parse_cidr(subnet_str, context=f"oob_vlan_subnets[{i}]")
            if not parsed:
                continue
            net_ip, prefix = parsed
            base = net_ip.rsplit('.', 1)[0]
            try:
                net_last = int(net_ip.rsplit('.', 1)[1])
            except (ValueError, IndexError):
                continue
            ztp_ifaces.append({
                'name': f'eth{2 + i}',
                'ip': f"{base}.{net_last + 78}",
                'network': subnet_str,
                'gateway': f"{base}.{net_last + 1}",
                'purpose': f'mgmt-subnet-{i + 1}',
                'dnsmasq_listen': True,
            })
    all_vars['ztp_interfaces'] = ztp_ifaces

    # SEC (scan finding #0): the nginx ZTP vhost restricts /scripts/, /configs/,
    # and /authorized_keys to the OOB management subnet(s) switches actually
    # source ZTP from. In L3-OOB mode switch eth0s live on air_mgmt_subnet
    # (e.g. 172.20.0.0/24); in L2/production they live on the OOB VLAN
    # subnet(s) (e.g. 192.168.200.0/24). Emit the union so the allow-list
    # always matches the real ZTP source subnet — a mismatch makes nginx 403
    # every ztp.sh fetch.
    ztp_allow_subnets = [air_mgmt_subnet]
    for subnet_str in oob_subnets_list:
        subnet_str = subnet_str.strip()
        if subnet_str and _parse_cidr(subnet_str) and subnet_str not in ztp_allow_subnets:
            ztp_allow_subnets.append(subnet_str)
    all_vars['ztp_allow_subnets'] = ztp_allow_subnets

    # EXIT-VRF inter-VRF DHCP relay scopes (external-dhcp dnsmasq).
    # If the DHCP Relay table declares an EXIT row AND any VLAN opts into
    # EXIT relay via `DHCP Relay Client = EXIT`, external-dhcp's dnsmasq
    # needs (a) to listen on eth2 (the customer-DC-side leg) and (b) a
    # dhcp-range per client VLAN subnet so dnsmasq has a pool to offer
    # from when relayed DISCOVERs arrive with giaddr in those subnets.
    # See docs/plans/2026-05-27-l3-oob-exit-dhcp-relay.md.
    if _mode == 'l3' and dhcp_relay_table:
        has_exit_row = any(
            r['vrf'].upper() == 'EXIT' for r in dhcp_relay_table
        )
        exit_target_vlans = []
        if has_exit_row:
            for v in vlans:
                targets = {c.strip().upper() for c in
                           str(v.get('dhcp_relay_client', '')).split(',')
                           if c.strip()}
                if 'EXIT' in targets:
                    exit_target_vlans.append(v)
        if exit_target_vlans:
            all_vars['inter_vrf_dhcp_listen_iface'] = 'eth2'
            all_vars['inter_vrf_dhcp_scopes'] = [
                {
                    'vlan_id': v['id'],
                    'subnet': v['subnet'],
                    'gateway': v.get('gateway', ''),
                    'name': v.get('name', f"vlan{v['id']}"),
                }
                for v in exit_target_vlans
            ]

    # Merge invariant all.yml defaults for variables the parser doesn't generate
    # (ztp_*, ssh_*, cumulus_target_version, nvue_syntax, devices, ldap, …).
    # Sourced from the single-home consolidated inventory_defaults.yml
    # (section 'all', keyed by arch) instead of the per-arch seed all.yml.
    source_vars = arch_group_vars("all", arch)
    merged_count = 0
    for key, value in source_vars.items():
        if key not in all_vars:
            all_vars[key] = value
            merged_count += 1
    if merged_count:
        print(f"    Merged {merged_count} variables from inventory_defaults (all/{arch})")

    # Write to all/main.yml (directory form — all/secrets.yml also lives here)
    all_dir = group_vars_dir / "all"
    all_dir.mkdir(exist_ok=True)
    all_file = all_dir / "main.yml"
    with open(all_file, 'w') as f:
        f.write("---\n")
        f.write("# ============================================================================\n")
        f.write("# Global Variables - Applied to All Switches\n")
        f.write("# ============================================================================\n")
        yaml.dump(all_vars, f, default_flow_style=False, sort_keys=False)
    generated_files.append(all_file)

    # Remove stale flat all.yml if it exists (avoid Ansible double-loading)
    stale_all = group_vars_dir / "all.yml"
    if stale_all.exists():
        stale_all.unlink()
    
    # core.yml - Core switch configuration from Excel
    bgp_asn = settings.get('bgp_asn', 4260394788)
    # NOTE: no `pre_login_message` here. The switch login banners come from
    # Settings.pre_login_message / post_login_message via `common` (see above) —
    # every template reads `common.pre_login_message`. A bare `pre_login_message`
    # key was emitted into core.yml/oob.yml for years carrying a stock Cumulus VX
    # banner that nothing rendered; it only made the group_vars look like the
    # banner source.
    core_vars = {
        'timezone': settings.get('timezone', 'Etc/Zulu'),
        'mh_mac': settings.get('mh_mac', '44:38:39:ff:00:aa'),
        'anycast_mac': settings.get('anycast_mac', '44:38:39:ff:00:ff'),
        'bgp_asn': bgp_asn,
    }
    # Per-function cumulus version from VERSIONS table (new format only)
    _core_version = resolve_role_version(versions, 'core')
    if _core_version:
        core_vars['cumulus_target_version'] = _core_version
    # NTP servers — reuse ntp_list parsed earlier for all_vars
    core_vars['ntp_servers'] = ntp_list

    # Add VLANs list
    core_vars['vlans'] = [v['id'] for v in vlans]
    
    # Build VNIs dict from VLANs (VNI always present — fallback is VLAN_ID + 4000)
    core_vars['vnis'] = {v['id']: v['vni'] for v in vlans}
    
    # VRF VNIs from VRFs section
    if vrfs:
        vrf_vnis = {}
        for vrf_name, vrf_data in vrfs.items():
            if vrf_data.get('l3_vni'):
                vrf_vnis[vrf_name] = int(vrf_data['l3_vni'])
        if vrf_vnis:
            core_vars['vrf_vnis'] = vrf_vnis
            # Excel-driven L3 VNIs get overlaid onto `vrf_config` AFTER the
            # source-inventory merge below — see `_apply_excel_vrf_vnis`.

    # Add disabled interfaces — Wire Map takes precedence over settings/defaults
    if port_config and port_config.get('interfaces_disabled'):
        core_vars['interfaces_disabled'] = port_config['interfaces_disabled']
    else:
        disabled_ports_str = settings.get('disabled_ports', '')
        if disabled_ports_str:
            core_vars['interfaces_disabled'] = [
                int(p.strip()) for p in str(disabled_ports_str).split(',') if p.strip()
            ]
        else:
            core_vars['interfaces_disabled'] = DEFAULT_DISABLED_INTERFACES.get(arch, [])

    # num_physical_ports (#4): SN5610 default is 64; override via Settings
    core_vars['num_physical_ports'] = int(settings.get('num_physical_ports', 64))

    # Inject Wire Map-derived port topology (network_roles, gpu/isl/edge_interfaces)
    if port_config:
        if port_config.get('network_roles'):
            core_vars['network_roles'] = port_config['network_roles']
        for key in ('gpu_interfaces', 'isl_interfaces', 'edge_interfaces',
                    'oob_uplink_interfaces',
                    'gpu_rail_interfaces'):
            if port_config.get(key):
                core_vars[key] = port_config[key]

    # dhcp_relay: VRF-aware DHCP relay (Excel-driven).
    # Built from the DHCP Relay table on the VLANs & Profiles sheet +
    # per-VLAN 'DHCP Relay Client' column. One entry per server-group
    # (= per VRF where the relay daemon runs).
    if dhcp_relay_table:
        dhcp_relay = []
        for server_row in dhcp_relay_table:
            vrf = server_row['vrf'].upper()
            server_group = f"{vrf.lower()}-dhcp-servers"
            downstream_ifaces = []
            for v in vlans:
                client = str(v.get('dhcp_relay_client', '')).strip()
                if not client or client.lower() == 'no':
                    continue
                targets = [c.strip().upper() for c in client.split(',') if c.strip()]
                if vrf in targets:
                    downstream_ifaces.append(f"vlan{v['id']}")
            if not downstream_ifaces:
                # No client VLANs reference this server-group — skip emission.
                continue
            dhcp_relay.append({
                'vrf': vrf,
                'server_group': server_group,
                'servers': server_row['servers'],
                'upstream_interfaces': server_row['upstream_interfaces'],
                'downstream_interfaces': downstream_ifaces,
            })
        if dhcp_relay:
            core_vars['dhcp_relay'] = dhcp_relay

    # Merge in source inventory variables that the Excel parser doesn't generate.
    # Excel-derived values take precedence; source inventory fills in complex routing
    # config (vrf_config, default_vrf_bgp, nve_vxlan, route_map, community_list, etc.)
    project_root = Path(__file__).resolve().parent.parent
    # Merge the converged-leaf routing policy from the single-home
    # consolidated defaults (section 'core', keyed by arch) instead of the
    # per-arch seed core.yml/csl.yml. Link-state defaults (interfaces_up/down)
    # are never inherited — they are tied to the reference port layout and
    # destructive on site-specific wire maps.
    source_vars = arch_group_vars("core", arch)
    merged_count = 0
    for key, value in source_vars.items():
        if key in ("interfaces_up", "interfaces_down"):
            continue
        if key not in core_vars:
            core_vars[key] = value
            merged_count += 1
    if merged_count:
        print(f"    Merged {merged_count} variables from inventory_defaults (core/{arch})")

    # Apply optional 'Route policy' / 'Community lists'
    # sheet overrides on top of the source-inventory route_map/community_list
    # just merged above. Same-named entries are replaced, new ones appended;
    # absent directives leave the merged defaults untouched.
    if route_policy_directives:
        core_vars['route_map'] = _apply_named_overrides(
            core_vars.get('route_map', []), route_policy_directives)
    if community_list_directives:
        core_vars['community_list'] = _apply_named_overrides(
            core_vars.get('community_list', []), community_list_directives)

    if _normalize_oob_uplink_mode(settings) == 'l3':
        settings = dict(settings)
        settings['_derived_oob_overlay_peers'] = _derive_oob_overlay_peers(
            nodes=nodes, loopback_overrides=loopback_overrides)
        # On a dedicated cl/cs spine-leaf tier (ns_tiers > 1) the EVPN
        # overlay is hub-and-spoke through the cs spines, NOT a cl↔oob mesh. The
        # cl leaves must peer the cs spine loopbacks for overlay (so cl↔cs↔cl
        # stretches INBAND/storage), while the oob-switches peer cs for VLAN-200.
        # _derive_core_overlay_peers returns the cs spine loopbacks when spines
        # exist (same list the oob-switches use), so reuse it here.
        settings['_derived_cs_spine_overlay_peers'] = _derive_core_overlay_peers(
            nodes=nodes, loopback_overrides=loopback_overrides)
    _apply_oob_l3_uplink_mode(core_vars, settings)
    # Order matters: supplement first (may add new vrf_config entries with
    # placeholder `interfaces: []`), then sync overwrites those placeholders
    # with the Excel-driven edge_interfaces port list.
    _supplement_vrf_config_from_excel(core_vars, vrfs, dhcp_relay_table, vlans)
    _sync_edge_vrf_neighbors(core_vars)

    # Overlay Excel-driven L3 VNIs onto the source-inventory `vrf_config`
    # list. Without this, the Excel "L3 VNI" column on the VLANs & Profiles
    # sheet is dead data — the core template iterates `vrf_config` and
    # emits the source-inventory default no matter what the sheet says.
    if isinstance(core_vars.get('vrf_vnis'), dict) and \
            isinstance(core_vars.get('vrf_config'), list):
        for entry in core_vars['vrf_config']:
            vid = entry.get('id')
            if vid in core_vars['vrf_vnis']:
                entry['vni'] = str(core_vars['vrf_vnis'][vid])

    core_file = group_vars_dir / "core.yml"
    with open(core_file, 'w') as f:
        f.write("---\n")
        f.write(f"# Core Switch Configuration (Generated from Excel + Source Inventory - {arch})\n\n")
        yaml.dump(core_vars, f, default_flow_style=False, sort_keys=False)
    generated_files.append(core_file)

    active_nodes_for_group_vars = [
        node for node in (nodes or [])
        if node.get('status', 'Active') not in ('Disabled', 'Air')
        and node.get('enabled', True)
    ]
    group_var_categories = {
        canonical_category(
            node.get('category') or node.get('role') or node.get('function'),
            node.get('name'),
        )
        for node in active_nodes_for_group_vars
    }
    has_csl_nodes = 'csl' in group_var_categories
    has_cl_nodes = 'cl' in group_var_categories
    gsl_planes_present = {
        plane_num
        for plane_num, categories_for_plane in (
            (1, ('gsl-plane1', 'gl-plane1')),
            (2, ('gsl-plane2', 'gl-plane2')),
        )
        if any(c in group_var_categories for c in categories_for_plane)
    }

    # csl.yml — for dedicated_gpu designs where the converged switches are
    # CSL (CPU/Storage Leaf) rather than Core. Same vars as core.yml minus
    # the GPU plane (VLAN 900, GPU VRF) which belongs on GSL switches only.
    import shutil  # used below
    # Gate purely on node categories (Excel-derived); the seed csl.yml
    # is gone, and has_csl_nodes covers every arch that ships CSL leaves.
    if has_csl_nodes:
        csl_vars = _strip_gpu_plane(core_vars, vlans)
        csl_file = group_vars_dir / "csl.yml"
        with open(csl_file, 'w') as f:
            f.write("---\n")
            f.write(f"# CSL Switch Configuration (Generated from Excel + Source Inventory - {arch})\n\n")
            yaml.dump(csl_vars, f, default_flow_style=False, sort_keys=False)
        generated_files.append(csl_file)

    # cl.yml — split-role (non-converged) CPU/Storage leaf. Same vars as
    # csl.yml (core_vars minus the GPU plane), emitted under the cl group
    # name so the playbook can select the leaf template. Only generated when
    # cl nodes exist or a source cl.yml is present; converged archs are
    # unaffected.
    if has_cl_nodes:  # seed cl.yml never existed; gate on Excel nodes only
        cl_vars = _strip_gpu_plane(core_vars, vlans)
        cl_file = group_vars_dir / "cl.yml"
        with open(cl_file, 'w') as f:
            f.write("---\n")
            f.write(f"# CL Switch Configuration (Generated from Excel + Source Inventory - {arch})\n\n")
            yaml.dump(cl_vars, f, default_flow_style=False, sort_keys=False)
        generated_files.append(cl_file)

    # GPU leaf plane group_vars — GENERATED from Excel (replaces the former
    # verbatim shutil.copy2 of the source files). Split-role
    # archs (gl-plane leaves paired with gs-plane spines, e.g. 2-4-5-800) carry
    # the GPU EVPN VNI numbering (gpu_vni / gpu_evpn_vni) on the leaf group file,
    # sourced from the Excel `VLANs & Profiles` VNI columns so any deployment
    # can set its own value. Collapsed gsl-plane archs emit only the shadowed
    # asn/vlan900_vrr/gpu_subnet — their GPU VNI comes from the all-group /
    # template default, unchanged. Every other key in the legacy source files
    # was SHADOWED by per-host Wire-Map data or dead — dropped.
    gpu_vlans = [v for v in vlans if str(v.get('vrf') or '').strip().upper() == 'GPU']
    gpu_vrf = vrfs.get('GPU') or {}
    for plane_num in sorted(gsl_planes_present):
        # OUTPUT filename must match the inventory group the leaves land in
        # ([gl_plane{N}] split-role vs legacy [gsl_plane{N}] collapsed).
        split_role = f'gl-plane{plane_num}' in group_var_categories
        plane_file = f"gl_plane{plane_num}.yml" if split_role else f"gsl_plane{plane_num}.yml"
        preferred = next(
            (
                v for v in gpu_vlans
                if str(v.get('name') or '').strip().lower() == f'gpu_plane{plane_num}'
            ),
            gpu_vlans[plane_num - 1] if len(gpu_vlans) >= plane_num else (gpu_vlans[0] if gpu_vlans else {}),
        )
        subnet = preferred.get('subnet') or ('192.168.0.0/20' if plane_num == 1 else '192.168.16.0/20')
        try:
            net = ipaddress.ip_network(subnet, strict=False)
            default_vrr = f"{net.network_address + 1}/{net.prefixlen}"
        except ValueError:
            default_vrr = '192.168.0.1/20' if plane_num == 1 else '192.168.16.1/20'
        vlan900_vrr = preferred.get('gateway') or default_vrr
        if '/' not in str(vlan900_vrr):
            vlan900_vrr = f"{vlan900_vrr}/{str(default_vrr).split('/')[-1]}"
        try:
            plane_asn = int(core_vars.get('bgp_asn') or settings.get('bgp_asn') or 4200100001) + 1000 + (plane_num - 1)
        except (TypeError, ValueError):
            plane_asn = 4200101000 + (plane_num - 1)

        # GPU EVPN VNI numbering — Excel `VLANs & Profiles` is the source of
        # truth (VLAN-900 VNI column -> gpu_vni; GPU VRF L3 VNI -> gpu_evpn_vni),
        # replacing the former hardcoded source group_vars. Emitting it on the
        # leaf plane group makes it per-deployment-overridable via the sheet.
        # asn/vlan900_vrr/gpu_subnet are SHADOWED by per-host values / dead, but
        # emitted for continuity with the prior file shape.
        plane_vars = {}
        if preferred.get('id') is not None:
            plane_vars['gpu_vlan'] = int(preferred['id'])
        if preferred.get('vni') is not None:
            plane_vars['gpu_vni'] = int(preferred['vni'])
        if gpu_vrf.get('name'):
            plane_vars['gpu_vrf_name'] = gpu_vrf['name']
        if gpu_vrf.get('vlan') is not None:
            plane_vars['gpu_evpn_vlan'] = int(gpu_vrf['vlan'])
        if gpu_vrf.get('l3_vni') is not None:
            plane_vars['gpu_evpn_vni'] = int(gpu_vrf['l3_vni'])
        plane_vars['asn'] = plane_asn
        plane_vars['vlan900_vrr'] = vlan900_vrr
        plane_vars['gpu_subnet'] = subnet
        # ERA-73: breakout LEVEL from the sheet. The plane group_vars already
        # carried which ports break out (`*_breakout_parents` in host_vars) but
        # never how far, so gl/gs silently used the template's '2x' regardless
        # of what the GPU Network / ISL rows declared. Both currently declare
        # 2 on every shipped arch, so emitting them changes no output — it makes
        # the coupling real, so that editing the sheet now works.
        _gpu_bo = _breakout_token((plane_breakout or {}).get('gpu'))
        _isl_bo = _breakout_token((plane_breakout or {}).get('isl'))
        if _gpu_bo:
            plane_vars['gpu_breakout'] = _gpu_bo
        if _isl_bo:
            plane_vars['isl_breakout'] = _isl_bo

        dst = group_vars_dir / plane_file
        with open(dst, 'w') as f:
            f.write("---\n")
            f.write(f"# GPU Plane {plane_num} Configuration (Generated from Excel - {arch})\n\n")
            yaml.dump(plane_vars, f, default_flow_style=False, sort_keys=False)
        generated_files.append(dst)

    # Spine group_vars — GENERATED from Excel (replaces the former verbatim
    # shutil.copy2). The only spine-template key not shadowed by
    # per-host Wire-Map data is roce_traffic_pool — a structural constant, true
    # for GPU spines. The cs (CPU/Storage spine) role needs no such flag; all its
    # source keys are shadowed by host_vars. Converged archs ship no spine role
    # and generate nothing here.
    for plane_num in (1, 2):
        if f'gs-plane{plane_num}' in group_var_categories:
            dst = group_vars_dir / f"gs_plane{plane_num}.yml"
            with open(dst, 'w') as f:
                f.write("---\n")
                f.write(f"# GPU Spine Plane {plane_num} Configuration (Generated from Excel - {arch})\n\n")
                _spine_vars = {'roce_traffic_pool': True}
                # ERA-73 — same reasoning as the leaf planes above.
                _isl_bo = _breakout_token((plane_breakout or {}).get('isl'))
                if _isl_bo:
                    _spine_vars['isl_breakout'] = _isl_bo
                yaml.dump(_spine_vars, f, default_flow_style=False, sort_keys=False)
            generated_files.append(dst)
    if 'cs' in group_var_categories:
        dst = group_vars_dir / "cs.yml"
        with open(dst, 'w') as f:
            f.write("---\n")
            f.write(f"# CS (CPU/Storage Spine) Configuration (Generated from Excel - {arch})\n")
        generated_files.append(dst)

    # oob.yml — generated from Excel when versions table is present (new format),
    # otherwise copied from source inventory (old format).
    oob_file = group_vars_dir / "oob.yml"
    if versions is not None:
        # New format: generate oob.yml from Excel data + source inventory merge.
        # Key off the VRF column, not the free-text Name: VRF is a constrained,
        # normalized vocabulary (see validate_excel._LOOPBACK_VRFS) and the
        # validator already identifies the OOB VLAN this exact way. The old
        # `name.startswith('OOB')` was case-sensitive and silently fell through
        # to 200 for a workbook that named the VLAN 'oob' or 'Mgmt-OOB'.
        oob_vlan_id = next(
            (v['id'] for v in vlans if str(v.get('vrf') or '').strip().upper() == 'OOB'),
            None)
        if oob_vlan_id is None:
            oob_vlan_id = 200
            print(f"  ⚠️  No VLAN with VRF 'OOB' on the 'VLANs & Profiles' sheet — "
                  f"defaulting oob_vlan to {oob_vlan_id}. Set the VRF column to OOB "
                  f"on the management VLAN so this is driven by the workbook.",
                  file=sys.stderr)
        oob_vars = {
            'timezone': settings.get('timezone', 'Etc/Zulu'),
            'oob_vlan': str(oob_vlan_id),
            'oob_uplink_mode': _normalize_oob_uplink_mode(settings),
        }
        _oob_version = resolve_role_version(versions, 'oob')
        if _oob_version:
            oob_vars['cumulus_target_version'] = _oob_version
        # Merge non-derived oob vars from the single-home consolidated
        # defaults (section 'oob', keyed by arch) instead of the per-arch seed.
        source_oob_vars = arch_group_vars("oob", arch)
        merged_oob = 0
        for key, value in source_oob_vars.items():
            if key not in oob_vars:
                oob_vars[key] = value
                merged_oob += 1
        if merged_oob:
            print(f"    Merged {merged_oob} variables from inventory_defaults (oob/{arch})")
        with open(oob_file, 'w') as f:
            f.write("---\n")
            f.write("# SPDX-FileCopyrightText: Copyright (c) 2026 NVIDIA CORPORATION & AFFILIATES. All rights reserved.\n")
            f.write("# SPDX-License-Identifier: MIT\n")
            f.write(f"# OOB Switch Configuration (Generated from Excel + Source Inventory - {arch})\n\n")
            yaml.dump(oob_vars, f, default_flow_style=False, sort_keys=False)
        generated_files.append(oob_file)
    else:
        # Old format (versions is None): write oob defaults from the single-home
        # consolidated defaults instead of copying the per-arch seed.
        legacy_oob = arch_group_vars("oob", arch)
        if legacy_oob and not oob_file.exists():
            with open(oob_file, 'w') as f:
                f.write("---\n")
                f.write("# SPDX-FileCopyrightText: Copyright (c) 2026 NVIDIA CORPORATION & AFFILIATES. All rights reserved.\n")
                f.write("# SPDX-License-Identifier: MIT\n")
                f.write(f"# OOB Switch Configuration (inventory_defaults - {arch})\n\n")
                yaml.dump(legacy_oob, f, default_flow_style=False, sort_keys=False)
            generated_files.append(oob_file)

    # Emit switches.yml / servers.yml from the single-home constants
    # module (scripts/inventory_defaults.py) instead of shutil.copy2-ing the
    # per-arch seed. Content is uniform across archs and byte-identical to the
    # prior golden — the seed copies were a drift hazard with no unique data.
    for extra_file, content in (("switches.yml", SWITCHES_GROUP_VARS),
                                ("servers.yml", SERVERS_GROUP_VARS)):
        dst = group_vars_dir / extra_file
        if not dst.exists():
            dst.write_text(content)
            generated_files.append(dst)

    # Also copy the secrets template if not present. One shared
    # inventories/secrets.yml for every arch (the six per-arch copies were
    # functionally identical — a drift hazard); fall back to secrets.yml.example.
    secrets_dst = all_dir / "secrets.yml"
    if not secrets_dst.exists():
        secrets_src = project_root / "inventories" / "secrets.yml"
        if not secrets_src.exists():
            secrets_src = project_root / "inventories" / "secrets.yml.example"
        if secrets_src.exists():
            shutil.copy2(secrets_src, secrets_dst)

    return generated_files


def process_excel_template(excel_path, output_dir):
    """Process an Excel template and generate inventory files.

    Args:
        excel_path: Path to the Excel workbook.
        output_dir: Final output directory (e.g. output/<arch>/<site>/inventory/).
                    Files are written directly here — no subdirectories are created.
    """
    # Start each generation run with a clean MAC collision registry so repeated
    # in-process runs (tests, batch invocations) can't accumulate stale state
    # and mis-detect collisions. No effect on a single `make generate` process.
    reset_mac_registry()
    print(f"\nProcessing: {excel_path}")

    wb = load_workbook_safe(excel_path, data_only=True)

    # R4-05: pre-check required sheets so direct callers (not via
    # `make generate`'s validate gate) get a friendly error instead of
    # an uncaught `KeyError: 'Worksheet X does not exist.'` traceback.
    required_sheets = ('Settings', 'Nodes', 'VLANs & Profiles')
    missing = [s for s in required_sheets if s not in wb.sheetnames]
    if missing:
        raise SystemExit(
            f"❌ {excel_path}: required sheets missing: {', '.join(missing)}. "
            f"Run `python3 scripts/validate_excel.py {excel_path}` for details, "
            f"or use `make generate` which gates on validation.")

    # Detect format: new format has Air_Only sheet
    new_format = 'Air_Only' in wb.sheetnames

    # Parse sheets
    settings = parse_settings(wb['Settings'])
    versions = parse_versions(wb['Settings']) if new_format else {}
    nodes = parse_nodes(wb['Nodes'])
    # Single source of truth for "what is this host" — used by Wire Map / Air_Only
    # row parsing to resolve a row's Function when its own cell is blank. See
    # _build_wiremap_row_list for the cascading lookup.
    nodes_function_map = build_nodes_function_map(nodes)
    vlans = parse_vlans(wb['VLANs & Profiles'])
    vrfs = parse_vrfs(wb['VLANs & Profiles'])
    dhcp_relay_table = parse_dhcp_relay_table(wb['VLANs & Profiles'])
    # parse_prefix_lists_sheet returns the full
    # {'lists': {...}, 'suppress': set()} directive dict, threaded unchanged
    # through generate_host_vars -> generate_prefix_lists, which applies it
    # scoped to OVERRIDABLE_PREFIX_LISTS (derived-only per-switch lists are
    # never overridable/suppressible — see the frozensets above).
    prefix_list_directives = parse_prefix_lists_sheet(wb['Prefix lists']) if 'Prefix lists' in wb.sheetnames else None
    # ACLs sheet (ADR-0030): optional control-plane ACL override, same
    # derive-by-default + directive model as the prefix-list sheet.
    acl_directives = parse_acls_sheet(wb['ACLs']) if 'ACLs' in wb.sheetnames else None
    # ADR-0055. Targeting resolves against the functions that actually receive a
    # generated config, so a target with no template fails loudly here rather than
    # silently producing nothing.
    custom_config = (
        parse_custom_config_sheet(
            wb['Custom_Config'],
            switches_by_function=switches_by_function_from_nodes(nodes),
            servers_by_function=servers_by_function_from_nodes(nodes),
        )
        if 'Custom_Config' in wb.sheetnames else {}
    )
    _lb_sheet = loopbacks_sheet_name(wb)
    loopback_overrides = parse_loopbacks_sheet(wb[_lb_sheet]) if _lb_sheet else None
    # The Loopbacks ASN column is the home for per-node BGP ASNs;
    # Settings.bgp_asn is removed from the shipped workbooks. When it is absent,
    # recover the fabric base ASN from the tab (a converged core/csl node's ASN =
    # the base, offset 0 — only exists for ns_tiers==1) and inject it into
    # settings, so the derived fallbacks + converged-core group ASN keep working
    # unchanged. A workbook that still carries Settings.bgp_asn (older customer
    # input) is untouched here and keeps deriving as before.
    if (not settings.get('bgp_asn')) and loopback_overrides:
        _ns_tiers = int((settings or {}).get('ns_tiers',
                                             (settings or {}).get('tiers', 1)) or 1)
        _base = None
        # ns_tiers==1: a converged core/csl node's ASN IS the base (offset 0).
        if _ns_tiers == 1:
            for _n in nodes:
                if _n.get('category') in ('core', 'csl'):
                    _a = (loopback_overrides.get(_n.get('name') or '') or {}).get('asn')
                    if _a:
                        _base = int(_a)
                        break
        # Fallback (any arch, incl. dedicated ns_tiers>1 with no converged core):
        # OOB switches occupy the lowest offset block, oob asn = base + oob_idx
        # (oob_idx starts at 1), so base = min(OOB ASN) - 1.
        if _base is None:
            _oob = [int((loopback_overrides.get(_n.get('name') or '') or {}).get('asn'))
                    for _n in nodes if _n.get('category') == 'oob-switch'
                    and (loopback_overrides.get(_n.get('name') or '') or {}).get('asn')]
            if _oob:
                _base = min(_oob) - OOB_ASN_OFFSET - 1
        if _base is not None:
            settings['bgp_asn'] = _base
    # Optional override sheets for route_map / community_list.
    # Absent sheet -> None -> generate_group_vars leaves the source-inventory
    # defaults merged in untouched (derive-by-default).
    route_policy_directives = parse_route_policy_sheet(wb['Route policy']) if 'Route policy' in wb.sheetnames else None
    community_list_directives = parse_community_lists_sheet(wb['Community lists']) if 'Community lists' in wb.sheetnames else None
    air_virtual_nodes = set()
    node_oob_mapping = {}
    wiremap_rows = None
    physical_wiremap_rows = None
    # Build the disabled-name set once (used by row-list and topology callers).
    disabled_names = {(n.get('name') or '').strip() for n in nodes
                       if n.get('status') == 'Disabled' and n.get('name')}
    if 'Wire Map' in wb.sheetnames:
        ws_wm = wb['Wire Map']
        ws_air_only = wb['Air_Only'] if 'Air_Only' in wb.sheetnames else None
        oob_switch_configs = parse_oob_switch_configs(ws_wm, ws_air_only,
                                                        nodes_function_map=nodes_function_map)
        port_config = parse_core_port_config(
            ws_wm, wb['VLANs & Profiles'],
            nodes_function_map=nodes_function_map,
            vlans=vlans,
            oob_uplink_mode=_normalize_oob_uplink_mode(settings),
        )
        gsl_port_configs = parse_gsl_port_config(ws_wm, nodes_function_map=nodes_function_map)
        # Per-switch server bonds + per-node ESI for the dedicated
        # N/S compute leaf tier (cl-* always; csl-* when ns_tiers>1, where the
        # csl name is legacy for a cl role — e.g. 2-8-9-800 maxscale). Empty for
        # converged csl/core (ns_tiers==1), which keep the golden shared config.
        per_switch_network_roles = build_per_switch_server_roles(
            ws_wm, (port_config or {}).get('network_roles'),
            nodes_function_map=nodes_function_map,
            dedicated_ns_tier=int((settings or {}).get('ns_tiers') or 1) > 1)
        per_switch_direct_interfaces = build_per_switch_direct_interfaces(
            ws_wm,
            port_config,
            nodes_function_map=nodes_function_map,
            oob_uplink_mode=_normalize_oob_uplink_mode(settings),
        )
        node_oob_mapping = parse_node_mgmt_mapping(ws_wm, new_format=new_format)
        # Build two Wire Map views. Topology and server configuration use only
        # active nodes; switch port configuration must retain every physical
        # row, including disabled scale-out nodes whose ports stay provisioned.
        physical_wiremap_rows = _build_wiremap_row_list(
            ws_wm,
            ws_air_only,
            nodes_function_map=nodes_function_map,
        )
        wiremap_rows = _build_wiremap_row_list(ws_wm, ws_air_only,
                                                nodes_function_map=nodes_function_map,
                                                disabled_names=disabled_names)
        # STORAGE VRF (PR-b): find any L3 STORAGE Port Profiles, then scan
        # Wire Map for rows using them. Empty result = no STORAGE on this
        # site, defaults stay unchanged. See
        # docs/plans/2026-05-19-storage-vrf-design.md.
        l3_storage_profiles = find_l3_storage_profiles(wb['VLANs & Profiles'])
        storage_uplink_hw = find_storage_uplink_hw(wb['VLANs & Profiles'])
        # Same idea for the GPU profile: the per-switch rail builder used
        # hardcoded _ROLE_HW constants, so the sheet's Speed never reached
        # per-rail ports (ERA-55).
        gpu_profile_hw = find_port_profile_hw(wb['VLANs & Profiles'], 'GPU Network')
        # ERA-73: the ISL profile's declared breakout governed core/csl but was
        # ignored by gl/gs/gsl, which fell back to the template's hardcoded
        # '2x'. Read it here so the GPU fabric is driven by the same sheet row
        # that already drives the N/S fabric (ADR-0035).
        isl_profile_hw = find_port_profile_hw(wb['VLANs & Profiles'], 'ISL')
        storage_uplink_ports = get_storage_uplink_ports_per_switch(
            ws_wm, l3_storage_profiles,
            nodes_function_map=nodes_function_map,
            disabled_names=disabled_names)
        # Per-bond descriptions naming the connected node hostname.
        # Matches the customer golden config:
        #   nv set interface bond1s0 description <host>-gpu-01
        known_node_names_set = {(n.get('name') or '').strip()
                                  for n in nodes
                                  if (n.get('name') or '').strip()}
        bond_descriptions_per_switch = get_bond_descriptions_per_switch(
            ws_wm,
            nodes_function_map=nodes_function_map,
            disabled_names=disabled_names,
            known_node_names=known_node_names_set)
        if new_format:
            # New format: Air virtual nodes come from dedicated Air_Only sheet
            air_virtual_nodes = parse_air_virtual_nodes(wb['Air_Only'], new_format=True)
            # Mode-aware default Air infra nodes (always injected by the
            # topology generator). L2: flat-bridge trio. L3: Ubuntu trio
            # per docs/plans/2026-05-20-l3-oob-air-topology.md.
            if _normalize_oob_uplink_mode(settings) == 'l3':
                air_virtual_nodes |= {'external-conn', 'external-dhcp', 'utility'}
            else:
                air_virtual_nodes |= {'dhcp-oob', 'oob-server-01'}
            # Air settings (Air Management Subnet, etc.)
            air_settings = parse_air_settings(wb['Air_Only'])
        else:
            # Old format: Air virtual nodes are tagged rows in Wire Map
            air_virtual_nodes = parse_air_virtual_nodes(ws_wm, new_format=False)
            air_settings = {}
    else:
        oob_switch_configs = {}
        port_config = None
        gsl_port_configs = {}
        per_switch_network_roles = {}
        per_switch_direct_interfaces = {}

    print(f"  Format: {'new (Air_Only sheet)' if new_format else 'legacy'}")
    print(f"  Settings: {len(settings)} items")
    if versions:
        print(f"  Versions: {versions}")
    print(f"  Nodes: {len(nodes)} total ({len([n for n in nodes if n['status'] == 'Active'])} active)")
    print(f"  VLANs: {len(vlans)}, VRFs: {len(vrfs)} defined")
    if prefix_list_directives and (prefix_list_directives.get('lists') or prefix_list_directives.get('suppress')):
        print(f"  Prefix list directives: lists={list(prefix_list_directives.get('lists', {}).keys())} "
              f"suppress={sorted(prefix_list_directives.get('suppress', set()))}")
    if acl_directives and (acl_directives.get('acls') or acl_directives.get('suppress')):
        print(f"  ACL directives: acls={list(acl_directives.get('acls', {}).keys())} "
              f"suppress={sorted(acl_directives.get('suppress', set()))}")
    if route_policy_directives:
        print(f"  Route policy overrides: {sorted(route_policy_directives.keys())}")
    if community_list_directives:
        print(f"  Community list overrides: {sorted(community_list_directives.keys())}")
    print(f"  OOB switches from Wire Map: {sorted(oob_switch_configs.keys())}")
    if gsl_port_configs:
        print(f"  GSL switches from Wire Map: {sorted(gsl_port_configs.keys())}")
    if air_virtual_nodes:
        print(f"  Air virtual nodes: {sorted(air_virtual_nodes)}")
    if node_oob_mapping:
        print(f"  Node mgmt mapping: {len(node_oob_mapping)} nodes with eth0 → OOB switch from Wire Map")
    if port_config:
        roles = list(port_config.get('network_roles', {}).keys())
        extras = [k for k in ('gpu_interfaces', 'isl_interfaces', 'edge_interfaces') if port_config.get(k)]
        print(f"  Core port config from Wire Map: roles={roles} direct={extras} disabled={port_config['interfaces_disabled']}")

    # Get architecture name
    arch = settings.get('architecture', Path(excel_path).stem)
    output_dir.mkdir(parents=True, exist_ok=True)

    # Generate files
    hosts_file = generate_hosts_file(settings, nodes, output_dir, air_virtual_nodes, vlans)
    print(f"  Generated: {hosts_file}")

    host_vars_files = generate_host_vars(nodes, vlans, output_dir, arch, settings,
                                          prefix_list_directives, oob_switch_configs,
                                          vrfs, air_settings, gsl_port_configs,
                                          loopback_overrides, wiremap_rows=wiremap_rows,
                                          physical_wiremap_rows=physical_wiremap_rows,
                                          storage_uplink_ports=storage_uplink_ports if 'Wire Map' in wb.sheetnames else None,
                                          bond_descriptions_per_switch=bond_descriptions_per_switch if 'Wire Map' in wb.sheetnames else None,
                                          per_switch_network_roles=per_switch_network_roles,
                                          per_switch_direct_interfaces=per_switch_direct_interfaces,
                                          storage_uplink_hw=storage_uplink_hw,
                                          gpu_profile_hw=gpu_profile_hw,
                                          group_port_blocks=port_config,
                                          acl_directives=acl_directives,
                                          custom_config=custom_config)

    # Merge host_vars for Air virtual nodes from source inventory.
    # These nodes don't appear in the Excel Nodes sheet but need host_vars.
    # Merge preserves air-deploy.py connection details (ansible_host/port)
    # while updating config vars (oob_server_interfaces, etc.)
    host_vars_dir = output_dir / "host_vars"
    # Air Management Subnet for virtual node IPs
    _air_mgmt = air_settings.get('air_mgmt_subnet', '172.20.0.0/24')
    _parsed_air2 = _parse_cidr(_air_mgmt, context="Air_Only.air_mgmt_subnet")
    if not _parsed_air2:
        _air_mgmt = '172.20.0.0/24'
        _parsed_air2 = _parse_cidr(_air_mgmt)
    _air_base = _parsed_air2[0].rsplit('.', 1)[0]
    _air_prefix = _parsed_air2[1]

    _connection_keys = {'ansible_host', 'ansible_port', 'ansible_user'}
    for vnode in sorted(air_virtual_nodes):
        dst = host_vars_dir / f"{vnode}.yml"
        # Source Air vnode host_vars from the single-home consolidated
        # defaults instead of the per-arch seed inventories/<arch>/host_vars/.
        src_vars = air_vnode_host_vars(arch, vnode)
        if src_vars is None:
            continue
        # Load existing output (may have real SSH details from air-deploy)
        existing = {}
        if dst.exists():
            with open(dst) as f:
                existing = yaml.safe_load(f) or {}
        # Merge: source vars as base, preserve existing connection keys
        merged = {**src_vars}
        for key in _connection_keys:
            if key in existing and existing[key] != 'CHANGE_ME':
                merged[key] = existing[key]

        # oob-server-01: eth1 (air-mgmt gateway) + ethN per OOB VLAN subnet
        if 'oob-server' in vnode:
            _mgmt_list = resolve_oob_vlans(
                vlans, get_oob_nodes_for_inventory(nodes, settings, vlans)
            )['subnets']
            oob_ifaces = [{
                'name': 'eth1',
                'ip': f"{_air_base}.1",
                'netmask': _air_prefix,
                'network': _air_mgmt,
                'purpose': 'Air Management Gateway',
            }]
            for i, subnet_str in enumerate(_mgmt_list):
                net_ip, prefix = subnet_str.split('/')
                prefix = int(prefix)
                base = net_ip.rsplit('.', 1)[0]
                net_last = int(net_ip.rsplit('.', 1)[1])
                oob_ifaces.append({
                    'name': f'eth{2 + i}',
                    'ip': f"{base}.{net_last + 1}",
                    'netmask': prefix,
                    'network': subnet_str,
                    'purpose': f'OOB Mgmt Subnet {i + 1} Gateway',
                })
            merged['oob_server_interfaces'] = oob_ifaces

        with open(dst, 'w') as f:
            f.write("---\n")
            yaml.dump(merged, f, default_flow_style=False, sort_keys=False)
        host_vars_files.append(dst)

    # Sweep stale Air-virtual-node host_vars from the OTHER mode. If an
    # operator switches a site between L2 and L3 (or vice versa), the old
    # mode's trio (dhcp-oob/oob-server-01 OR external-conn/external-dhcp/
    # utility) would otherwise persist with stale IPs and confuse audits.
    L2_INFRA_NAMES = {'dhcp-oob', 'oob-server-01'}
    L3_INFRA_NAMES = {'external-conn', 'external-dhcp', 'utility'}
    other_mode_files = (L3_INFRA_NAMES if 'dhcp-oob' in air_virtual_nodes
                        else L2_INFRA_NAMES if 'utility' in air_virtual_nodes
                        else set())
    swept = 0
    for stale in other_mode_files:
        stale_path = host_vars_dir / f"{stale}.yml"
        if stale_path.exists():
            stale_path.unlink()
            swept += 1
    if swept:
        print(f"  Swept {swept} stale host_vars from previous mode")

    print(f"  Generated: {len(host_vars_files)} host_vars files")
    
    group_vars_files = generate_group_vars(
        settings, vlans, vrfs, output_dir, arch, nodes, port_config,
        node_oob_mapping, versions=versions or None, wiremap_rows=wiremap_rows,
        air_settings=air_settings, dhcp_relay_table=dhcp_relay_table,
        loopback_overrides=loopback_overrides,
        route_policy_directives=route_policy_directives,
        community_list_directives=community_list_directives,
        plane_breakout={'gpu': gpu_profile_hw, 'isl': isl_profile_hw})
    print(f"  Generated: {len(group_vars_files)} group_vars files")

    # Loopback visibility (#5): print assignments so users know what IPs get configured
    lb = str(settings.get('loopback_base') or LOOPBACK_BASE).strip()
    core_nodes = sorted(
        [n for n in nodes if n.get('category') == 'core'],
        key=lambda x: x['name'],
    )
    if core_nodes:
        print(f"  Loopback assignments (base={lb}):")
        for node in core_nodes:
            core_num = node.get('index') or 1
            # Derive from the allocator itself rather than restating its
            # formula — a hand-copied copy silently went stale when the
            # per-VRF layout moved to VRF_LOOPBACK_BLOCKS.
            vrf_los = generate_vrf_loopbacks(
                vlans, core_num, loopback_base=lb,
                switch_overrides=(loopback_overrides or {}).get(node['name'], {}),
            )
            shown = '  '.join(
                f"{vrf}={ip.split('/')[0]}"
                for vrf, ip in sorted(vrf_los.items())
                if ip.split('/')[0].startswith(lb)
            )
            print(f"    {node['name']}: lo={lb}.{10 + core_num}  {shown}")
    
    return output_dir


def _run_validator(excel_path):
    """R4-10/R4-11: run validate_excel as a subprocess and bail on errors.

    Direct invocations of excel_parser.py (outside `make generate`) used
    to skip validation entirely and silently emit broken inventory when
    the validator would have caught the bad input. Invoke it here.
    """
    import subprocess
    here = Path(__file__).resolve().parent
    cmd = ['python3', str(here / 'validate_excel.py'), str(excel_path)]
    rc = subprocess.call(cmd)
    if rc != 0:
        raise SystemExit(
            f"\n❌ Validation failed for {excel_path}. Fix the errors above, "
            f"or pass --skip-validate to bypass (not recommended).")


def main():
    import argparse
    parser = argparse.ArgumentParser(
        description="Parse ERA Excel templates and generate Ansible inventory.",
    )
    parser.add_argument('--arch', metavar='ARCH',
                        help="Process only this architecture (e.g. 2-8-5-200)")
    parser.add_argument('--site', metavar='SITE', default='default',
                        help="Site name (default: 'default')")
    parser.add_argument('--skip-validate', action='store_true',
                        help="Skip the validate-excel pre-check (not recommended).")
    args = parser.parse_args()

    base_dir = Path(__file__).resolve().parent.parent
    input_dir = base_dir / "input"
    output_base = base_dir / "output"

    print("=" * 60)
    print("ERA Excel → Inventory Generator")
    print("=" * 60)

    if args.arch:
        # Process a single architecture/site
        excel_path = input_dir / args.arch / args.site / f"{args.arch}.xlsx"
        if not excel_path.exists():
            print(f"❌  Excel not found: {excel_path}")
            raise SystemExit(1)
        if not args.skip_validate:
            _run_validator(excel_path)
        output_dir = output_base / args.arch / args.site / "inventory"
        process_excel_template(excel_path, output_dir)
        print(f"\n✅ Inventory written to: {output_dir.relative_to(base_dir)}/")
    else:
        # Discover and process all default templates
        templates = sorted(input_dir.glob("*/default/*.xlsx"))
        if not templates:
            print(f"No templates found under {input_dir}/<arch>/default/")
            return
        for template in templates:
            arch = template.stem
            if not args.skip_validate:
                _run_validator(template)
            output_dir = output_base / arch / "default" / "inventory"
            process_excel_template(template, output_dir)
        print("\n✅ Inventories written to: output/<arch>/default/inventory/")

    print("=" * 60)


if __name__ == "__main__":
    main()
