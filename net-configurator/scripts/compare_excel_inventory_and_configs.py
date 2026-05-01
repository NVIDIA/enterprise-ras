#!/usr/bin/env python3
# SPDX-FileCopyrightText: Copyright (c) 2026 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: MIT
"""
Compare Excel defaults to inventory and generated configs to enterprise-ras source of truth.

Reads input/defaults/*.xlsx, builds inventory in NEW_TESTS/<arch>/, generates CLI configs,
and compares to enterprise-ras/spectrum-era/<arch>/exported-configurations/CLI/.
DO NOT MODIFY any existing files; report-only.

Note: input/defaults Excel "Nodes" sheet uses columns: Function, Name, MAC Address for ZTP,
Mgmt IP Address, Prefix, Gateway, ZTP, Enabled. This script parses by header. excel_parser
uses position-based columns (mgmt_ip=5) which does not match that layout.

Usage:
    python3 scripts/compare_excel_inventory_and_configs.py [--no-generate] [--enterprise-ras PATH]
"""

import argparse
import os
import re
import shutil
import subprocess
import sys
from pathlib import Path

import openpyxl
import yaml


# ---------------------------------------------------------------------------
# NVUE interface range normalization helpers
# ---------------------------------------------------------------------------

def _natural_key(s):
    """Sort key: splits on digit boundaries for natural ordering."""
    return [int(c) if c.isdigit() else c for c in re.split(r'(\d+)', s)]


def _expand_iface_token(token):
    """
    Expand a NVUE comma+range interface spec token to sorted, expanded form.

    Examples:
      'bond1s0-3'                        -> 'bond1s0,bond1s1,bond1s2,bond1s3'
      'swp49-52'                         -> 'swp49,swp50,swp51,swp52'
      'swp49,swp50,swp51,swp52'         -> 'swp49,swp50,swp51,swp52'  (unchanged, sorted)
      'spine_bond,swp1-6,10-17,28-40'   -> canonical expanded+sorted form
      'bond1s0-3,bond2s0-3'             -> 'bond1s0,...,bond1s3,bond2s0,...,bond2s3'

    The prefix is everything up to and including the last letter before the trailing
    digit sequence, so 'swp49' has prefix 'swp' (not 'swp4').
    """
    parts = token.split(',')
    expanded = []
    last_alpha_prefix = None
    for part in parts:
        # Prefix must end with a letter: (.*[a-zA-Z]) ensures correct split for 'swp49-52'
        m = re.match(r'^(.*[a-zA-Z])(\d+)(?:-(\d+))?$', part)
        if m:
            raw_prefix = m.group(1)
            start = int(m.group(2))
            end = int(m.group(3)) if m.group(3) else start
            last_alpha_prefix = raw_prefix
            for i in range(start, end + 1):
                expanded.append(f'{raw_prefix}{i}')
        else:
            # Digits-only part: carry forward the prefix from the previous item
            # e.g. '10-17' after 'swp1-6' expands to swp10..swp17
            m2 = re.match(r'^(\d+)(?:-(\d+))?$', part)
            if m2 and last_alpha_prefix is not None:
                start = int(m2.group(1))
                end = int(m2.group(2)) if m2.group(2) else start
                for i in range(start, end + 1):
                    expanded.append(f'{last_alpha_prefix}{i}')
            else:
                expanded.append(part)
                last_alpha_prefix = None
    expanded.sort(key=_natural_key)
    return ','.join(expanded)


def normalize_nvue_line(line):
    """
    Normalize NVUE interface range notation in a nv set/unset line so that
    compact ranges ('swp49-52') and expanded lists ('swp49,swp50,...') compare equal.
    Applies to any whitespace-delimited token that looks like an interface spec.
    """
    tokens = line.split()
    result = []
    for tok in tokens:
        # Expand token if it contains a letter-digit range or a comma-separated iface list
        if re.search(r'[a-zA-Z]\d+-\d', tok) or (
            ',' in tok
            and re.search(r'[a-zA-Z]', tok)
            and re.search(r'\d', tok)
        ):
            result.append(_expand_iface_token(tok))
        else:
            result.append(tok)
    return ' '.join(result)

def functional_lines(path):
    """Extract nv set/unset command lines from a config file; normalize whitespace and interface ranges."""
    out = []
    with open(path) as f:
        for line in f:
            s = line.strip()
            if not s or s.startswith("#"):
                continue
            if s.startswith("nv set ") or s.startswith("nv unset "):
                out.append(normalize_nvue_line(" ".join(s.split())))
    return out


# Import parser helpers from excel_parser (same repo)
SCRIPT_DIR = Path(__file__).resolve().parent
BASE_DIR = SCRIPT_DIR.parent
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

from excel_parser import (
    DEFAULT_DISABLED_INTERFACES,
    LOOPBACK_BASE,
    generate_group_vars,
    generate_hosts_file,
    generate_host_vars,
    parse_prefix_lists_sheet,
    parse_settings,
    parse_vlans,
    parse_vrfs,
)


def parse_nodes_by_headers(ws):
    """
    Parse Nodes sheet using header row to find columns.
    input/defaults Excel has: Function, Name, MAC Address for ZTP, Mgmt IP Address, Prefix, Gateway, ZTP, Enabled
    """
    nodes = []
    headers_row = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    headers = {}
    for col, val in enumerate(headers_row, 1):
        if val:
            key = str(val).strip().lower()
            headers[key] = col
    # Map possible header names to our node keys
    role_col = headers.get("function") or 1
    name_col = headers.get("name") or 2
    mac_col = headers.get("mac address for ztp") or headers.get("mac address") or 0
    mgmt_col = headers.get("mgmt ip address") or headers.get("mgmt ip") or headers.get("ip address") or 0
    prefix_col = headers.get("prefix") or 0
    gateway_col = headers.get("gateway") or 0
    enabled_col = headers.get("enabled") or 0

    for row in range(2, ws.max_row + 1):
        role = ws.cell(row=row, column=role_col).value
        if not role:
            continue
        role = str(role).strip()
        name = (ws.cell(row=row, column=name_col).value or role) if name_col else role
        if name is not None:
            name = str(name).strip()
        mac = (ws.cell(row=row, column=mac_col).value or "") if mac_col else ""
        if mac is not None:
            mac = str(mac).strip()
        mgmt_ip = (ws.cell(row=row, column=mgmt_col).value or "") if mgmt_col else ""
        if mgmt_ip is not None:
            mgmt_ip = str(mgmt_ip).strip()
        prefix = (ws.cell(row=row, column=prefix_col).value or 24) if prefix_col else 24
        if prefix is not None:
            prefix = int(prefix) if isinstance(prefix, (int, float)) else 24
        gateway = (ws.cell(row=row, column=gateway_col).value or "") if gateway_col else ""
        if gateway is not None:
            gateway = str(gateway).strip()
        enabled = (ws.cell(row=row, column=enabled_col).value if enabled_col else "Yes")
        status = "Active" if (enabled and str(enabled).strip().lower() in ("yes", "true", "1", "active")) else "Disabled"

        nodes.append({
            "role": role,
            "name": name or role,
            "status": status,
            "mac_address": mac or "",
            "mgmt_ip": mgmt_ip or "",
            "prefix": prefix,
            "gateway": gateway or "",
        })
    return nodes


def compare_excel_to_inventory(base_dir, issues):
    """Compare Excel default values to current inventory; append to issues list."""
    input_dir = base_dir / "input" / "defaults"
    inventories_dir = base_dir / "inventories"
    if not input_dir.exists():
        issues.append("input/defaults/ not found - cannot compare Excel to inventory")
        return

    for xlsx in sorted(input_dir.glob("*.xlsx")):
        arch = xlsx.stem
        inv_dir = inventories_dir / arch
        if not inv_dir.exists():
            issues.append(f"[{arch}] No inventory at inventories/{arch}/ - skip Excel vs inventory")
            continue

        wb = openpyxl.load_workbook(xlsx, data_only=True)
        settings = parse_settings(wb["Settings"])
        nodes = parse_nodes_by_headers(wb["Nodes"])
        wb.close()

        # Settings vs inventory
        core_yml = inv_dir / "group_vars" / "core.yml"
        inv_core = {}
        if core_yml.exists():
            with open(core_yml) as f:
                inv_core = yaml.safe_load(f) or {}

        if inv_core:
            ex_tz = settings.get("timezone")
            if ex_tz is not None and str(ex_tz) != str(inv_core.get("timezone", "")):
                issues.append(f"[{arch}] timezone: Excel={ex_tz!r} vs inventory={inv_core.get('timezone')!r}")
            ex_bgp = settings.get("bgp_asn")
            if ex_bgp is not None:
                try:
                    ex_bgp = int(ex_bgp)
                    if ex_bgp != inv_core.get("bgp_asn"):
                        issues.append(f"[{arch}] bgp_asn: Excel={ex_bgp} vs inventory={inv_core.get('bgp_asn')}")
                except (TypeError, ValueError):
                    pass

        # Nodes (host_vars) comparison
        for node in nodes:
            if node["status"] == "Disabled":
                continue
            role = node["role"]
            host_var_file = inv_dir / "host_vars" / f"{role}.yml"
            if not host_var_file.exists():
                continue
            with open(host_var_file) as f:
                hv = yaml.safe_load(f) or {}
            if node.get("mgmt_ip") and str(hv.get("ansible_host")) != str(node["mgmt_ip"]):
                issues.append(f"[{arch}] {role} ansible_host: Excel={node['mgmt_ip']!r} vs inventory={hv.get('ansible_host')!r}")
            if node.get("name") and str(hv.get("hostname")) != str(node["name"]):
                issues.append(f"[{arch}] {role} hostname: Excel={node['name']!r} vs inventory={hv.get('hostname')!r}")


def process_excel_to_new_tests(base_dir, new_tests_root, issues, input_dir=None):
    """Parse Excel files into <output>/<arch>/ using header-based Nodes parsing."""
    input_dir = input_dir if input_dir is not None else base_dir / "input" / "defaults"
    inventories_dir = base_dir / "inventories"
    if not input_dir.exists():
        issues.append("input/defaults/ not found")
        return []

    processed = []
    for xlsx in sorted(input_dir.glob("*.xlsx")):
        arch = xlsx.stem
        wb = openpyxl.load_workbook(xlsx, data_only=True)
        settings = parse_settings(wb["Settings"])
        nodes = parse_nodes_by_headers(wb["Nodes"])
        vlans = parse_vlans(wb["VLANs & Profiles"])
        vrfs = parse_vrfs(wb["VLANs & Profiles"])
        prefix_list_overrides = parse_prefix_lists_sheet(wb["Prefix lists"]) if "Prefix lists" in wb.sheetnames else None
        wb.close()

        arch_from_settings = settings.get("architecture") or arch
        if isinstance(arch_from_settings, float):
            arch_from_settings = int(arch_from_settings)
        arch_from_settings = str(arch_from_settings)

        output_dir = new_tests_root / arch
        output_dir.mkdir(parents=True, exist_ok=True)

        generate_hosts_file(settings, nodes, output_dir)
        generate_host_vars(nodes, vlans, output_dir, arch_from_settings, settings, prefix_list_overrides)
        generate_group_vars(settings, vlans, vrfs, output_dir, arch_from_settings)

        # Overwrite with full group_vars from inventory so config generation has network_roles, telemetry, etc.
        # Only copy files that don't require vault secrets (switches.yml has ansible_password vault ref)
        for gvar_file in ["core.yml", "oob.yml"]:
            inv_gvar = inventories_dir / arch / "group_vars" / gvar_file
            if inv_gvar.exists():
                shutil.copy2(inv_gvar, output_dir / "group_vars" / gvar_file)
        if not (inventories_dir / arch / "group_vars" / "core.yml").exists():
            issues.append(f"[{arch}] No inventories/{arch}/group_vars/core.yml - config generation may fail")

        # Copy real core host_vars so VLAN SVI IPs / prefix lists match truth (Excel has mgmt IPs, not ERA fabric)
        inv_host_vars_dir = inventories_dir / arch / "host_vars"
        out_host_vars_dir = output_dir / "host_vars"
        if inv_host_vars_dir.exists():
            for hv_file in inv_host_vars_dir.glob("core-*.yml"):
                shutil.copy2(hv_file, out_host_vars_dir / hv_file.name)

        processed.append(arch)
    return processed


def run_generate(base_dir, arch, new_tests_root, issues):
    """Run ansible-playbook generate-cli-configs.yml with NEW_TESTS inventory."""
    inventory_path = new_tests_root / arch / "hosts"
    generated_dir = new_tests_root / arch / "generated-cli"
    generated_dir.mkdir(parents=True, exist_ok=True)
    # config_output_dir relative to playbooks/ so playbooks/../NEW_TESTS/arch/generated-cli
    config_output_dir = str((new_tests_root / arch / "generated-cli").resolve())
    playbook_dir = base_dir / "playbooks"
    # Ansible file module resolves paths relative to playbook dir; use absolute to be clear
    cmd = [
        "ansible-playbook",
        str(base_dir / "playbooks" / "generate-cli-configs.yml"),
        "-i", str(inventory_path),
        "-e", f"config_output_dir={config_output_dir}",
    ]
    try:
        result = subprocess.run(
            cmd,
            cwd=str(base_dir),
            capture_output=True,
            text=True,
            timeout=120,
        )
        if result.returncode != 0:
            issues.append(f"[{arch}] Config generation failed: {result.stderr or result.stdout or 'non-zero exit'}")
            return False
    except FileNotFoundError:
        issues.append(f"[{arch}] ansible-playbook not found - skip config generation")
        return False
    except subprocess.TimeoutExpired:
        issues.append(f"[{arch}] Config generation timed out")
        return False
    return True


def compare_configs_to_enterprise_ras(arch, new_tests_root, enterprise_ras_root, issues):
    """
    Compare NEW_TESTS/<arch>/generated-cli/*-config.sh to
    enterprise-ras/spectrum-era/<arch>/exported-configurations/CLI/*.sh
    """
    generated_dir = new_tests_root / arch / "generated-cli"
    truth_dir = Path(enterprise_ras_root) / "spectrum-era" / arch / "exported-configurations" / "CLI"
    if not generated_dir.exists():
        issues.append(f"[{arch}] Generated config dir missing: {generated_dir}")
        return
    generated_files = {f.stem.replace("-config", ""): f for f in generated_dir.glob("*-config.sh")}
    if not generated_files:
        issues.append(f"[{arch}] No *-config.sh in generated dir (config generation may have failed)")
        return
    if not truth_dir.exists():
        issues.append(f"[{arch}] Enterprise-ras source of truth missing: {truth_dir}")
        return

    truth_files = {f.stem: f for f in truth_dir.glob("*.sh")}

    for host_key in set(generated_files) | set(truth_files):
        gen_file = generated_files.get(host_key)
        truth_file = truth_files.get(host_key)
        if not gen_file:
            issues.append(f"[{arch}] Missing in generated: {host_key}-config.sh (truth has {host_key}.sh)")
            continue
        if not truth_file:
            issues.append(f"[{arch}] Extra in generated: {gen_file.name} (no {host_key}.sh in truth)")
            continue

        gen_cmds = functional_lines(gen_file)
        truth_cmds = functional_lines(truth_file)
        gen_set = set(gen_cmds)
        truth_set = set(truth_cmds)
        only_gen = gen_set - truth_set
        only_truth = truth_set - gen_set

        if only_truth:
            for cmd in sorted(only_truth)[:10]:
                issues.append(f"[{arch}] {host_key}.sh only in TRUTH: {cmd[:90]}")
            if len(only_truth) > 10:
                issues.append(f"[{arch}] {host_key}.sh ... and {len(only_truth) - 10} more only in truth")
        if only_gen:
            for cmd in sorted(only_gen)[:10]:
                issues.append(f"[{arch}] {host_key}.sh only in GENERATED: {cmd[:90]}")
            if len(only_gen) > 10:
                issues.append(f"[{arch}] {host_key}.sh ... and {len(only_gen) - 10} more only in generated")


def compare_configs_to_template(arch, new_tests_root, template_configs_by_host, issues):
    """
    Compare generated-cli/*-config.sh to configs extracted from template Excel Switch Configs tab.
    template_configs_by_host: dict host_key -> list of 'nv set ...' strings (e.g. from extract_template_switch_configs).
    """
    generated_dir = new_tests_root / arch / "generated-cli"
    if not generated_dir.exists():
        issues.append(f"[{arch}] Generated config dir missing: {generated_dir}")
        return
    generated_files = {f.stem.replace("-config", ""): f for f in generated_dir.glob("*-config.sh")}
    if not generated_files:
        issues.append(f"[{arch}] No *-config.sh in generated dir (config generation may have failed)")
        return

    for host_key in set(generated_files) | set(template_configs_by_host):
        gen_file = generated_files.get(host_key)
        truth_cmds = template_configs_by_host.get(host_key) or []
        truth_set = set(normalize_nvue_line(" ".join(str(c).strip().split())) for c in truth_cmds if c and str(c).strip().startswith("nv "))
        if not gen_file:
            issues.append(f"[{arch}] Missing in generated: {host_key} (template has {host_key})")
            continue
        gen_cmds = functional_lines(gen_file)
        gen_set = set(gen_cmds)
        only_gen = gen_set - truth_set
        only_truth = truth_set - gen_set
        if only_truth:
            for cmd in sorted(only_truth)[:10]:
                issues.append(f"[{arch}] {host_key} only in TEMPLATE: {cmd[:90]}")
            if len(only_truth) > 10:
                issues.append(f"[{arch}] {host_key} ... and {len(only_truth) - 10} more only in template")
        if only_gen:
            for cmd in sorted(only_gen)[:10]:
                issues.append(f"[{arch}] {host_key} only in GENERATED: {cmd[:90]}")
            if len(only_gen) > 10:
                issues.append(f"[{arch}] {host_key} ... and {len(only_gen) - 10} more only in generated")


def main():
    parser = argparse.ArgumentParser(description="Compare Excel defaults to inventory and configs to enterprise-ras (report only)")
    parser.add_argument("--no-generate", action="store_true", help="Skip config generation and config comparison")
    parser.add_argument("--skip-generate", action="store_true", help="Skip config generation step but still compare (use with pre-built configs)")
    parser.add_argument("--generated-dir", type=str, default=None,
                        help="Directory containing pre-built *-config.sh files (overrides NEW_TESTS/<arch>/generated-cli lookup)")
    parser.add_argument("--arch", type=str, default=None,
                        help="Comma-separated list of architectures to compare (e.g. 2-4-3-200,2-8-5-200). Auto-detected from output/ if omitted.")
    parser.add_argument("--enterprise-ras", type=str, default=None,
                        help="Path to enterprise-ras repo (default: $ENTERPRISE_RAS or ../enterprise-ras)")
    parser.add_argument("--output-dir", type=str, default="NEW_TESTS",
                        help="Output directory for generated inventory and configs (default: NEW_TESTS)")
    parser.add_argument("--input-dir", type=str, default=None,
                        help="Input directory for Excel files (default: input/defaults)")
    parser.add_argument("--compare-to-template", action="store_true",
                        help="Compare to template Excel Switch Configs tab (TEST1b: 2-8-5, 2-8-9 only)")
    args = parser.parse_args()

    base_dir = BASE_DIR
    new_tests_root = base_dir / args.output_dir
    input_dir_explicit = Path(args.input_dir) if args.input_dir else base_dir / "input" / "defaults"
    enterprise_ras = args.enterprise_ras or os.environ.get("ENTERPRISE_RAS") or str(base_dir.parent / "enterprise-ras")
    if not Path(enterprise_ras).exists() and not args.no_generate:
        enterprise_ras = "/home/stovar/GOLDENTURTLE/enterprise-ras"
    if not Path(enterprise_ras).exists():
        enterprise_ras = ""

    issues = []

    # 1) Excel vs inventory comparison (only when using default input)
    if input_dir_explicit == base_dir / "input" / "defaults":
        print("Step 1: Comparing input/defaults Excel to inventories/...")
        compare_excel_to_inventory(base_dir, issues)
    else:
        print("Step 1: Using custom input dir (skip Excel vs inventory comparison)")

    # 2) Parse Excel -> output dir (when --compare-to-template, only 2-8-5-200 and 2-8-9-400 from input/defaults)
    processed = []

    # When --arch is given (or --skip-generate with no input/defaults), populate processed directly
    if args.arch:
        processed = [a.strip() for a in args.arch.split(",") if a.strip()]
        print(f"Step 2: Using explicit arch list: {', '.join(processed)}")
    elif args.skip_generate and not input_dir_explicit.exists():
        # Auto-detect from output/
        output_root = base_dir / "output"
        if output_root.exists():
            processed = sorted(d.name for d in output_root.iterdir() if d.is_dir())
        print(f"Step 2: Auto-detected archs from output/: {', '.join(processed) or 'none'}")
    elif args.compare_to_template:
        input_dir_explicit = base_dir / "input" / "defaults"
        excel_files = [input_dir_explicit / "2-8-5-200.xlsx", input_dir_explicit / "2-8-9-400.xlsx"]
        excel_files = [p for p in excel_files if p.exists()]
        if not excel_files:
            issues.append("--compare-to-template: no input/defaults 2-8-5-200.xlsx or 2-8-9-400.xlsx")
        else:
            for xlsx in excel_files:
                arch = xlsx.stem
                wb = openpyxl.load_workbook(xlsx, data_only=True)
                settings = parse_settings(wb["Settings"])
                nodes = parse_nodes_by_headers(wb["Nodes"])
                vlans = parse_vlans(wb["VLANs & Profiles"])
                vrfs = parse_vrfs(wb["VLANs & Profiles"])
                prefix_list_overrides = parse_prefix_lists_sheet(wb["Prefix lists"]) if "Prefix lists" in wb.sheetnames else None
                wb.close()
                arch_from_settings = settings.get("architecture") or arch
                if isinstance(arch_from_settings, float):
                    arch_from_settings = int(arch_from_settings)
                arch_from_settings = str(arch_from_settings)
                output_dir = new_tests_root / arch
                output_dir.mkdir(parents=True, exist_ok=True)
                generate_hosts_file(settings, nodes, output_dir)
                generate_host_vars(nodes, vlans, output_dir, arch_from_settings, settings, prefix_list_overrides)
                generate_group_vars(settings, vlans, vrfs, output_dir, arch_from_settings)
                inv_core = base_dir / "inventories" / arch / "group_vars" / "core.yml"
                if inv_core.exists():
                    (output_dir / "group_vars").mkdir(parents=True, exist_ok=True)
                    shutil.copy2(inv_core, output_dir / "group_vars" / "core.yml")
                processed.append(arch)
    else:
        processed = process_excel_to_new_tests(base_dir, new_tests_root, issues, input_dir=input_dir_explicit)
        print(f"  Created output for: {', '.join(processed) or 'none'}")

    if not args.no_generate:
        if not args.skip_generate:
            for arch in processed:
                print(f"Step 3: Generating configs for {arch}...")
                run_generate(base_dir, arch, new_tests_root, issues)
        else:
            # --skip-generate: symlink or copy pre-built configs into NEW_TESTS structure
            if args.generated_dir:
                gen_src = Path(args.generated_dir)
                for arch in processed:
                    dest = new_tests_root / arch / "generated-cli"
                    dest.mkdir(parents=True, exist_ok=True)
                    arch_src = gen_src / arch / "default" / "configs"
                    if not arch_src.exists():
                        arch_src = gen_src / arch
                    for sh in arch_src.glob("*-config.sh"):
                        target = dest / sh.name
                        if not target.exists():
                            target.symlink_to(sh.resolve())
                print(f"Step 3: Skipped generation; linked configs from {args.generated_dir}")
            else:
                # Default: look in output/<arch>/default/configs/
                for arch in processed:
                    src = base_dir / "output" / arch / "default" / "configs"
                    dest = new_tests_root / arch / "generated-cli"
                    dest.mkdir(parents=True, exist_ok=True)
                    if src.exists():
                        for sh in src.glob("*-config.sh"):
                            target = dest / sh.name
                            if not target.exists():
                                target.symlink_to(sh.resolve())
                    else:
                        issues.append(f"[{arch}] --skip-generate: no configs at output/{arch}/default/configs/")
                print("Step 3: Skipped generation; linked configs from output/<arch>/default/configs/")
        if args.compare_to_template:
            print("Step 4: Comparing generated configs to template Switch Configs tab...")
            try:
                from extract_template_switch_configs import extract_configs_for_arch
            except ImportError:
                extract_configs_for_arch = None
            if extract_configs_for_arch:
                for arch in processed:
                    template_configs = extract_configs_for_arch(arch)
                    if template_configs:
                        compare_configs_to_template(arch, new_tests_root, template_configs, issues)
                    else:
                        issues.append(f"[{arch}] No configs extracted from template Switch Configs tab")
            else:
                issues.append("Could not import extract_template_switch_configs")
        elif enterprise_ras:
            print("Step 4: Comparing generated configs to enterprise-ras source of truth...")
            for arch in processed:
                compare_configs_to_enterprise_ras(arch, new_tests_root, enterprise_ras, issues)
        else:
            issues.append("Enterprise-ras path not set/found - skip config comparison (use --enterprise-ras PATH)")

    # Report
    print("\n" + "=" * 60)
    print("REPORT (do not modify any files)")
    print("=" * 60)
    if not issues:
        print("No issues found.")
    else:
        for i, msg in enumerate(issues, 1):
            print(f"  {i}. {msg}")
    print("=" * 60)
    print(f"Output written to: {new_tests_root}")
    # Write report to output dir
    report_path = new_tests_root / "report.txt"
    try:
        new_tests_root.mkdir(parents=True, exist_ok=True)
        with open(report_path, "w") as f:
            f.write("REPORT (do not modify any files)\n")
            f.write("=" * 60 + "\n")
            if not issues:
                f.write("No issues found.\n")
            else:
                for i, msg in enumerate(issues, 1):
                    f.write(f"  {i}. {msg}\n")
            f.write("=" * 60 + "\n")
        print(f"Report saved to: {report_path}")
    except Exception as e:
        print(f"Could not write report file: {e}")
    if issues:
        sys.exit(1)
    sys.exit(0)


if __name__ == "__main__":
    main()
