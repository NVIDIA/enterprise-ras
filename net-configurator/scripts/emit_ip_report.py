#!/usr/bin/env python3
# SPDX-FileCopyrightText: Copyright (c) 2026 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: MIT
"""
Emit a read-only "IP assignments" report — every resolved IP for a deployment.

Sources (all already-generated artifacts, never re-read as input; see ADR-0034):
  - servers:  the resolved `devices` map in
              output/<arch>/<site>/inventory/group_vars/all/main.yml
  - switches: output/<arch>/<site>/inventory/host_vars/<switch>.yml
              (loopbacks, per-SVI IP + anycast VRR gateway, VRF loopbacks, mgmt)
  - Air infra: infra host_vars (mgmt IP; shown as "assigned by Air" when the
              static inventory carries a CHANGE_ME placeholder)

Flattens to one row per (device, interface, IP) and writes
output/<arch>/<site>/reports/ip-assignments.xlsx plus a text preview. The
canonical source of truth remains the input Excel — this is output only.

Usage:
    python3 scripts/emit_ip_report.py --arch 2-8-5-200 [--site default]
"""
import argparse
import ipaddress
from pathlib import Path

import yaml

# Columns mirror the Wire Map row shape: the device is the A side ("Port (A)" is the
# Wire Map's own physical-NIC name verbatim, e.g. "B3220 PCIe Slot 1 Port 1"), and the
# far end is split into "System Name (B)" + "Port (B)" exactly like the Excel. "Logical
# Port (A)" is the tool's logical name for the same A-side port (eth1, bond0, vlan300, lo).
IP_COL = 7  # index of "IP Address" in COLUMNS below
COLUMNS = ["System Name (A)", "Type", "Network Profile", "Port (A)",
           "Logical Port (A)", "System Name (B)", "Port (B)",
           "IP Address", "Subnet", "Gateway"]

# Second, human-facing sheet ("By Node") — LaunchPad P2P style: one row per physical
# port, the host name merged vertically over its block, rows tinted by Network Profile.
# No bond abstraction — each port stands on its own line as in the LaunchPad wiremap.
GROUPED_COLUMNS = ["System Name (A)", "Port (A)", "Network Profile",
                   "IP Address", "Gateway", "System Name (B)", "Port (B)"]


# Per-node fill palette (LaunchPad's own cycle) — colors are NODE-centric: each device
# block gets the next colour so adjacent blocks are visually distinct. Not network-keyed.
NODE_PALETTE = ["F4CCCC", "FCE5CD", "FFF2CC", "D9EAD3", "C9DAF8", "D9D2E9",
                "EA9999", "F9CB9C", "FFE599", "B6D7A8", "A4C2F4", "B4A7D6",
                "E06666", "F6B26B", "FFD966", "93C47D", "6FA8DC", "8E7CC3"]

# VLAN id -> human network name (ERA convention). Falls back to "vlan<id>".
VLAN_NAME = {"200": "oob", "300": "in-band", "400": "support", "500": "storage",
             "900": "gpu"}
# Server interface-group key -> the same network vocabulary as the switch SVIs.
NET_DISPLAY = {"oob": "oob", "cpu": "in-band", "support": "support",
               "storage": "storage", "gpu": "gpu"}
# Internal net token -> Wire Map "Network Profile" name. Used ONLY for rows with no
# Wire Map link (switch SVIs, bond header rows, loopback/mgmt) — linked rows show the
# profile string read straight from the Wire Map. Tokens absent here (loopback, mgmt,
# air-mgmt, ztp, vlan<id>) have no profile and pass through unchanged.
NET_PROFILE = {"oob": "OOB / IPMI", "in-band": "CPU/In-Band Network",
               "support": "Support", "storage": "Storage Uplink",
               "gpu": "GPU Network"}
# Sort order for a tidy report.
NET_ORDER = {"mgmt": 0, "oob": 1, "loopback": 2, "in-band": 3, "support": 4,
             "storage": 5, "gpu": 6}
TYPE_ORDER = {"switch": 0, "oob-switch": 1, "server": 2, "infra": 3}


def _norm(ip):
    """Ensure an 'a.b.c.d/pfx' string; bare addresses get a /32."""
    ip = str(ip).strip()
    return ip if "/" in ip else f"{ip}/32"


def _net(cidr):
    try:
        return str(ipaddress.ip_interface(cidr).network)
    except ValueError:
        return ""


def _derived_gateway(cidr):
    try:
        net = ipaddress.ip_interface(cidr).network
        return str(next(net.hosts())) if net.num_addresses > 2 else ""
    except (ValueError, StopIteration):
        return ""


def load_switches(host_vars_dir):
    """Return {name: host_vars} for switch host_vars (have lo_ip or vlan_interfaces)."""
    switches = {}
    for f in sorted(host_vars_dir.glob("*.yml")):
        hv = yaml.safe_load(f.read_text()) or {}
        if "lo_ip" in hv or "vlan_interfaces" in hv:
            switches[f.stem] = hv
    return switches


def svi_gateway_map(switches):
    """network-CIDR -> anycast VRR gateway address, harvested from switch SVIs."""
    gw = {}
    for hv in switches.values():
        for svi in hv.get("vlan_interfaces", []) or []:
            if svi.get("ip") and svi.get("vrr"):
                gw[_net(_norm(svi["ip"]))] = svi["vrr"].split("/")[0]
    return gw


# Server data plane preference, mirroring the role interfaces.j2 templates
# (support/storage use their own key, compute uses cpu). eth0 is always OOB.
DATA_PLANES = ("support", "storage", "cpu")


def _bond_plane(ifaces):
    """Return (plane, [member-eths]) for the single bonded data plane, eth0 excluded."""
    for plane in DATA_PLANES:
        eths = [e for e in (ifaces.get(plane) or []) if e != "eth0"]
        if eths:
            return plane, eths
    return None, []


def server_rows(devices):
    """Yield interface records (device, type, network, iface, ip, members, is_bond).

    ``members`` is the list of physical eths behind the interface; ``is_bond`` marks
    a bond that build_rows expands into a header row (the IP) plus one row per member
    (its Excel physical-port label). Bond membership mirrors
    roles/{nodes,storage,support}/templates/interfaces.j2 exactly: compute (single
    ``bond_ip``) bonds ALL cpu members into bond0; storage/support
    (``bond_ip1``/``bond_ip2``) put the first two data eths on bond0 and — only when
    there are >=4 data eths — the next two on bond1.
    """
    for name, dev in devices.items():
        ifaces = dev.get("interfaces", {}) or {}
        if dev.get("eth0_ip"):
            ip = f"{dev['eth0_ip']}/24" if "/" not in str(dev["eth0_ip"]) else dev["eth0_ip"]
            yield (name, "server", "oob", "eth0", _norm(ip), ["eth0"], False)
        plane, data = _bond_plane(ifaces)
        net = NET_DISPLAY.get(plane, plane) if plane else "bond"
        if dev.get("bond_ip"):
            # Compute node: one bond0 over every cpu member.
            yield (name, "server", net, "bond0", _norm(dev["bond_ip"]), data, True)
        else:
            # storage/support: bond0 = data[0:2], bond1 = data[2:4] (>=4 eths only).
            if len(data) >= 2 and dev.get("bond_ip1"):
                yield (name, "server", net, "bond0", _norm(dev["bond_ip1"]), data[0:2], True)
            if len(data) >= 4 and dev.get("bond_ip2"):
                yield (name, "server", net, "bond1", _norm(dev["bond_ip2"]), data[2:4], True)
        # GPU (single-homed): dual-plane per-NIC objects, else flat gpu_ips list.
        gpu_ifaces = [e for e in (ifaces.get("gpu") or []) if e != "eth0"]
        if dev.get("gpu_interfaces"):
            for gnic in dev["gpu_interfaces"]:
                iface = gnic.get("iface", "gpu")
                yield (name, "server", "gpu", iface, _norm(gnic["ip"]), [iface], False)
        else:
            for i, ip in enumerate(dev.get("gpu_ips") or []):
                eth = gpu_ifaces[i] if i < len(gpu_ifaces) else f"gpu{i + 1}"
                yield (name, "server", "gpu", eth, _norm(ip), [eth], False)


def switch_rows(switches):
    for name, hv in switches.items():
        dtype = "oob-switch" if name.startswith("oob-switch") else "switch"
        mgmt = hv.get("ansible_host")
        if mgmt and str(mgmt) != "CHANGE_ME":
            yield (name, dtype, "mgmt", "eth0", _norm(f"{mgmt}/24"), [], False)
        if hv.get("lo_ip"):
            yield (name, dtype, "loopback", "lo", _norm(hv["lo_ip"]), [], False)
        for svi in hv.get("vlan_interfaces", []) or []:
            if not svi.get("ip"):
                continue
            vid = str(svi.get("vlan", "")).lstrip("vlan") or "?"
            net = VLAN_NAME.get(vid, f"vlan{vid}")
            yield (name, dtype, net, f"vlan{vid}", _norm(svi["ip"]), [], False)
        # OOB switches carry their OOB-VLAN SVI as a flat `svi_ip` rather than
        # in `vlan_interfaces`, so the loop above never saw it and the report
        # contained zero VLAN-200 rows. That omission is why this report could
        # not surface the switch-SVI/node-eth0 collisions that shipped on four
        # largescale sites: it printed one side of every duplicate and not the
        # other.
        if hv.get("svi_ip"):
            vid = str(hv.get("oob_access_vlan") or "200")
            yield (name, dtype, VLAN_NAME.get(vid, f"vlan{vid}"),
                   f"vlan{vid}", _norm(hv["svi_ip"]), [], False)
        for vrf, ip in (hv.get("vrf_loopbacks") or {}).items():
            yield (name, dtype, "loopback", f"lo ({vrf})", _norm(ip), [], False)


def infra_rows(host_vars_dir, switches, devices, main):
    """Infra nodes. Tool-assigned IPs (ztp/dnsmasq) come from main.yml, sourced by
    role binding; nodes with only an Air-assigned mgmt IP show a placeholder."""
    ztp_host = main.get("ztp_server_host")
    # The OOB/NAT server renders roles/oob-server/netplan-config.yaml.j2.
    oob_server_hosts = {h for h in (main.get("nat_host"),) if h}
    oob_server_ifaces = main.get("oob_server_interfaces") or []
    oob_mode = str(main.get("oob_uplink_mode", "l2")).lower()
    jump_host = main.get("jump_host")
    known = set(switches) | set(devices)
    for f in sorted(host_vars_dir.glob("*.yml")):
        if f.stem in known:
            continue
        hv = yaml.safe_load(f.read_text()) or {}
        name = f.stem
        emitted = False
        # Tool-assigned addressing for the ZTP/DHCP host (external-dhcp).
        if name == ztp_host:
            for itf in main.get("ztp_interfaces", []) or []:
                if itf.get("ip"):
                    net = itf.get("purpose", "ztp")
                    yield (name, "infra", net, itf.get("name", "eth?"),
                           _norm(f"{itf['ip']}/{itf['network'].split('/')[1]}"
                                 if itf.get("network") else f"{itf['ip']}/24"), [], False)
                    emitted = True
            if main.get("ztp_server_ip"):
                yield (name, "infra", "ztp", "ztp", _norm(main["ztp_server_ip"]), [], False)
                emitted = True
        # OOB server (external-conn): explicit oob_server_interfaces win; otherwise
        # its data-plane IPs are NOT in the generated artifacts — L2 gateways are
        # hard-coded in netplan-config.yaml.j2 and L3 addresses are written by
        # air-deploy.py at first boot (they live on the OOB-switch anycast VRR,
        # already reported via the oob-switch SVI rows). We label honestly rather
        # than fabricate. (ADR-0034 tweak #1.)
        if name in oob_server_hosts or (not oob_server_hosts and oob_server_ifaces):
            for itf in oob_server_ifaces:
                if itf.get("ip"):
                    nm = itf.get("netmask", 24)
                    yield (name, "infra", itf.get("purpose", "oob"),
                           itf.get("name", "eth?"), _norm(f"{itf['ip']}/{nm}"), [], False)
                    emitted = True
            if not emitted:
                note = ("(L3: gateways on OOB anycast VRR — first boot)"
                        if oob_mode == "l3" else "(L2 OOB gateways — see netplan template)")
                yield (name, "infra", "oob", "eth1+", note, [], False)
                emitted = True
        mgmt = hv.get("ansible_host")
        if mgmt is None:
            continue
        if str(mgmt) == "CHANGE_ME":
            if not emitted:   # only note the Air placeholder if we found no tool IP
                label = ("(Air-assigned mgmt)" if name == jump_host
                         else "(assigned by Air)")
                yield (name, "infra", "mgmt", "eth0", label, [], False)
        else:
            yield (name, "infra", "mgmt", "eth0", _norm(f"{mgmt}/24"), [], False)


def load_links(topology_path):
    """(node, iface) -> (peer_node, peer_iface) from topology content.links (paired)."""
    link_map = {}
    if not topology_path.exists():
        return link_map
    import json
    links = (json.loads(topology_path.read_text()).get("content", {}).get("links") or [])
    for pair in links:
        if not isinstance(pair, (list, tuple)) or len(pair) < 2:
            continue
        a, b = pair[0], pair[1]
        if not (isinstance(a, dict) and isinstance(b, dict)):
            continue
        ka, kb = (a.get("node"), a.get("interface")), (b.get("node"), b.get("interface"))
        link_map[ka] = kb
        link_map[kb] = ka
    return link_map


def load_wiremap(xlsx_path):
    """(switch, swp-port) -> (server-side physical NIC label, Network Profile) from the
    input Excel Wire Map, e.g. ('B3220 PCIe Slot 1 Port 1', 'CPU/In-Band Network').
    Read-only — the input Excel stays canonical (ADR-0034). Returns {} if the workbook
    / sheet is absent (e.g. a third-party sim imported without its source Excel).

    The Wire Map is orientation-agnostic (some rows list the switch on side A, some on
    side B), so we key off whichever side is a ``swp*`` port and take the other side's
    label as the physical NIC name.
    """
    wm = {}
    if not xlsx_path.exists():
        return wm
    try:
        from openpyxl import load_workbook
    except ImportError:
        return wm
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    sheet = next((s for s in wb.sheetnames if "wire" in s.lower()), None)
    if not sheet:
        return wm
    rows = wb[sheet].iter_rows(values_only=True)
    header = [str(h) if h is not None else "" for h in (next(rows, []) or [])]

    def col(label):
        return next((i for i, h in enumerate(header) if label in h), None)

    ia, ipa = col("System Name (A)"), col("Port (A)")
    ib, ipb = col("System Name (B)"), col("Port (B)")
    iprof = col("Network Profile")
    if None in (ia, ipa, ib, ipb):
        return wm
    for row in rows:
        sysA, portA, sysB, portB = row[ia], row[ipa], row[ib], row[ipb]
        prof = str(row[iprof]) if iprof is not None and row[iprof] is not None else ""
        a_sw = str(portA or "").startswith("swp")
        b_sw = str(portB or "").startswith("swp")
        if a_sw and not b_sw and sysA and portB:
            wm[(str(sysA), str(portA))] = (str(portB), prof)
        elif b_sw and not a_sw and sysB and portA:
            wm[(str(sysB), str(portB))] = (str(portA), prof)
    return wm


def build_rows(devices, host_vars_dir, topology_path, wm_map):
    switches = load_switches(host_vars_dir)
    gw_map = svi_gateway_map(switches)
    link_map = load_links(topology_path)
    main = yaml.safe_load((host_vars_dir.parent / "group_vars" / "all" / "main.yml").read_text()) or {}
    raw = (list(switch_rows(switches))
           + list(server_rows(devices))
           + list(infra_rows(host_vars_dir, switches, devices, main)))
    raw.sort(key=lambda r: (TYPE_ORDER.get(r[1], 9), r[0],
                            NET_ORDER.get(r[2], 9), r[3]))

    def _with_ip(cells, ip):
        """Append IP/Subnet/Gateway; note-strings ('(...)') carry no subnet/gw."""
        if not ip or str(ip).startswith("("):
            return cells + [ip, "", ""]
        subnet = _net(ip)
        return cells + [ip, subnet, gw_map.get(subnet) or _derived_gateway(ip)]

    def _far(peer):
        """(System Name (B), Port (B)) for a topology peer, or ('', '')."""
        return (peer[0], peer[1]) if peer else ("", "")

    # Learn each net token's Network Profile string from THIS workbook's own linked
    # ports, so bond-header / SVI rows (which have no Wire Map link) echo the same
    # wording as their member rows — one workbook may say 'OOB', another 'OOB / IPMI'.
    learned = {}
    for dev, _dt, net, _if, _ip, members, _ib in raw:
        if net in learned or not members:
            continue
        peer = link_map.get((dev, members[0]))
        if peer and wm_map.get(peer, ("", ""))[1]:
            learned[net] = wm_map[peer][1]

    def _profile(net):
        return learned.get(net) or NET_PROFILE.get(net, net)

    out = []
    for dev, dtype, net, iface, ip, members, is_bond in raw:
        if is_bond:
            # Header row (logical bond name + IP); each member is its own physical row.
            out.append(_with_ip([dev, dtype, _profile(net), "", iface, "", ""], ip))
            for eth in members:
                peer = link_map.get((dev, eth))
                sys_b, port_b = _far(peer)
                phys, prof = wm_map.get(peer, ("", "")) if peer else ("", "")
                # Qualify the member with its bond so the row is self-describing even
                # once the sheet is sorted/filtered away from its header row.
                out.append([dev, dtype, prof or _profile(net), phys,
                            f"{iface}:{eth}", sys_b, port_b, "", "", ""])
        else:
            eth = members[0] if members else None
            peer = link_map.get((dev, eth)) if eth else None
            sys_b, port_b = _far(peer)
            phys, prof = wm_map.get(peer, ("", "")) if peer else ("", "")
            out.append(_with_ip([dev, dtype, prof or _profile(net), phys,
                                 iface, sys_b, port_b], ip))
    return out


def build_grouped(devices, host_vars_dir, topology_path, wm_map):
    """Rows for the 'By Node' sheet: one row per physical port, grouped by host.

    Returns (rows, host_spans) where host_spans is a list of 0-based [start, end]
    index ranges used to merge the System Name (A) cell over each device's block.
    No bonds — every physical port is its own line; a shared bond IP simply lands on
    its first member's row (the other member's IP cell is left blank), as the
    LaunchPad wiremap does.
    """
    switches = load_switches(host_vars_dir)
    gw_map = svi_gateway_map(switches)
    link_map = load_links(topology_path)
    main = yaml.safe_load((host_vars_dir.parent / "group_vars" / "all" / "main.yml").read_text()) or {}
    raw = (list(switch_rows(switches))
           + list(server_rows(devices))
           + list(infra_rows(host_vars_dir, switches, devices, main)))
    raw.sort(key=lambda r: (TYPE_ORDER.get(r[1], 9), r[0],
                            NET_ORDER.get(r[2], 9), r[3]))

    learned = {}
    for dev, _dt, net, _if, _ip, members, _ib in raw:
        if net in learned or not members:
            continue
        peer = link_map.get((dev, members[0]))
        if peer and wm_map.get(peer, ("", ""))[1]:
            learned[net] = wm_map[peer][1]

    def _profile(net):
        return learned.get(net) or NET_PROFILE.get(net, net)

    def _gw(ip):
        if not ip or str(ip).startswith("("):
            return ""
        return gw_map.get(_net(ip)) or _derived_gateway(ip)

    def _far(peer):
        return (peer[0], peer[1]) if peer else ("", "")

    rows, host_spans = [], []
    cur_host, host_start = None, 0
    for dev, dtype, net, iface, ip, members, is_bond in raw:
        if dev != cur_host:
            if cur_host is not None:
                host_spans.append((host_start, len(rows) - 1))
            cur_host, host_start = dev, len(rows)

        def emit(port_a, prof, ip_cell, sys_b, port_b):
            name = dev if len(rows) == host_start else ""   # value on top row only
            rows.append([name, port_a, prof, ip_cell, _gw(ip_cell), sys_b, port_b])

        if is_bond:
            for idx, eth in enumerate(members):
                peer = link_map.get((dev, eth))
                sys_b, port_b = _far(peer)
                phys, pf = wm_map.get(peer, ("", "")) if peer else ("", "")
                emit(phys or eth, pf or _profile(net), ip if idx == 0 else "", sys_b, port_b)
        else:
            eth = members[0] if members else None
            peer = link_map.get((dev, eth)) if eth else None
            sys_b, port_b = _far(peer)
            phys, pf = wm_map.get(peer, ("", "")) if peer else ("", "")
            emit(phys or iface, pf or _profile(net), ip, sys_b, port_b)
    if cur_host is not None:
        host_spans.append((host_start, len(rows) - 1))
    return rows, host_spans


def add_grouped_sheet(wb, grouped):
    """LaunchPad-style 'By Node' sheet: host name merged per block, rows tinted by
    Network Profile. One row per physical port; no bond merges."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    rows, host_spans = grouped
    ws = wb.create_sheet("By Node")
    ws.append(GROUPED_COLUMNS)
    hdr_fill = PatternFill("solid", fgColor="EFEFEF")
    for cell in ws[1]:
        cell.font = Font(bold=True, size=10)
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in rows:
        ws.append(row)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ncol = len(GROUPED_COLUMNS)
    # Colour each device block with the next palette entry (node-centric, like LaunchPad).
    for bi, (s, e) in enumerate(host_spans):
        fill = PatternFill("solid", fgColor=NODE_PALETTE[bi % len(NODE_PALETTE)])
        for i in range(s, e + 1):
            for c in range(1, ncol + 1):
                cell = ws.cell(i + 2, c)
                cell.fill = fill
                cell.border = border
                cell.alignment = Alignment(vertical="center", horizontal="left")
        ws.cell(s + 2, 1).font = Font(bold=True)   # host name
        if e > s:   # merge the host name over its block (value only on the top row)
            ws.merge_cells(start_row=2 + s, start_column=1, end_row=2 + e, end_column=1)
    for col, header in enumerate(GROUPED_COLUMNS, 1):
        width = max([len(header)] + [len(str(r[col - 1])) for r in rows] + [8]) + 2
        ws.column_dimensions[ws.cell(1, col).column_letter].width = min(width, 44)
    ws.freeze_panes = "A2"
    return ws


def write_xlsx(rows, grouped, dest):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "IP Assignments"
    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append(row)
    for col, header in enumerate(COLUMNS, 1):
        width = max([len(header)] + [len(str(r[col - 1])) for r in rows] + [8]) + 2
        ws.column_dimensions[ws.cell(1, col).column_letter].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    add_grouped_sheet(wb, grouped)
    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)


def find_duplicate_ips(rows):
    """Addresses claimed by more than one owner, ignoring prefix length.

    The report used to be a listing only: it collected every address on every
    plane and never compared them. That is why it could sit on a real defect
    and read as clean — four largescale sites shipped a switch SVI and a node
    eth0 on the same OOB address, and both lines were in this table.

    A /32 loopback and a /24 SVI at the same address are still a duplicate, so
    compare on the bare address. Returns ``{address: [owner, ...]}`` sorted for
    stable output.
    """
    seen = {}
    for r in rows:
        addr = str(r[IP_COL]).split("/")[0].strip()
        if not addr or addr in ("", "-", "CHANGE_ME"):
            continue
        owner = f"{r[0]} ({r[4] or r[3] or '?'})"
        seen.setdefault(addr, [])
        if owner not in seen[addr]:
            seen[addr].append(owner)
    return {a: o for a, o in sorted(seen.items()) if len(o) > 1}


def print_duplicate_summary(rows):
    """Print the duplicate section. Returns the number of duplicated addresses."""
    dupes = find_duplicate_ips(rows)
    print()
    if not dupes:
        print("No duplicate IP addresses found.")
        return 0
    print(f"DUPLICATE IP ADDRESSES ({len(dupes)}) "
          f"- two owners on one address is an ARP/DAD war on that segment:")
    for addr, owners in dupes.items():
        print(f"  {addr}  <-  {'; '.join(owners)}")
    return len(dupes)


def print_preview(rows):
    cols = list(zip(COLUMNS, *rows)) if rows else [(c,) for c in COLUMNS]
    widths = [max(len(str(x)) for x in col) for col in cols]
    fmt = "  ".join("{:<%d}" % w for w in widths)
    print(fmt.format(*COLUMNS))
    print(fmt.format(*("-" * w for w in widths)))
    for r in rows:
        print(fmt.format(*(str(c) for c in r)))


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--arch", required=True)
    ap.add_argument("--site", default="default")
    args = ap.parse_args()

    base = Path("output") / args.arch / args.site / "inventory"
    main_yml = base / "group_vars" / "all" / "main.yml"
    host_vars = base / "host_vars"
    topology = Path("output") / args.arch / args.site / "topology" / f"{args.arch}-topology.json"
    if not main_yml.exists():
        raise SystemExit(f"⚠️  {main_yml} not found — run `make generate ARCH={args.arch} SITE={args.site}` first")

    devices = (yaml.safe_load(main_yml.read_text()) or {}).get("devices", {})
    wm_path = Path("input") / args.arch / args.site / f"{args.arch}.xlsx"
    wm_map = load_wiremap(wm_path)
    if not wm_map:
        print(f"⚠️  no Wire Map found at {wm_path} — 'Physical Port' column left blank")
    rows = build_rows(devices, host_vars, topology, wm_map)
    grouped = build_grouped(devices, host_vars, topology, wm_map)

    dest = Path("output") / args.arch / args.site / "reports" / "ip-assignments.xlsx"
    write_xlsx(rows, grouped, dest)
    print_preview(rows)
    # Read-only by contract, so this reports rather than exiting non-zero;
    # validate_excel is the gate that fails a build on duplicates.
    print_duplicate_summary(rows)
    n_dev = len({r[0] for r in rows})
    print(f"\n✓  {len(rows)} IP rows across {n_dev} devices → {dest}")


if __name__ == "__main__":
    main()
