#!/usr/bin/env python3
# SPDX-FileCopyrightText: Copyright (c) 2026 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: MIT
"""
Topology Generator - Generate and validate NVIDIA Air topology from Excel wiremap.

Reads the Wire Map sheet from the customer-filled Excel workbook
(input/<arch>/<site>/<arch>.xlsx) and produces a complete Air 2.0 topology
JSON with all switch ports represented — including unconnected stubs for
ports without connections.

Old format: Air-specific connections (dhcp-oob, oob-server-01, dhcp-edge, eth0
management) are included at the top of the Wire Map sheet with "Air - " prefixed
profiles.

New format: Air-specific connections are in a dedicated Air_Only sheet.
The Air_Only sheet also contains a version→image name mapping table used to
resolve the correct qcow2 image for each switch function.

Usage:
    # Generate topology from Excel wiremap
    python scripts/topology_generator.py generate --arch 2-8-5-200

    # Validate existing topology against wiremap
    python scripts/topology_generator.py validate --arch 2-8-5-200
    python scripts/topology_generator.py validate --arch 2-8-5-200 --topology path/to/file.json
"""

import argparse
import json
import re
import sys
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Set, Tuple

import openpyxl
import yaml

from utils import (generate_mac, classify_node, is_switch, is_valid_hostname,
                   parse_swp_port)
from excel_parser import (build_wiremap_column_map, _wm_cell_ws,
                          parse_nodes, build_nodes_function_map,
                          parse_vlans, resolve_oob_vlans)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# Minimum physical port counts per switch model.
# The generator also scans the wiremap for higher-numbered ports and extends
# the range automatically (e.g., SN2201 swp49-52 uplinks).
SWITCH_MODELS = {
    "core":    {"model": "SN5610", "ports": 64},
    "csl":     {"model": "SN5610", "ports": 64},
    "gsl":     {"model": "SN5610", "ports": 64},
    "oob":     {"model": "SN2201", "ports": 48},
    "edge":    {"model": "SN2201", "ports": 48},
    "air-oob": {"model": "SN2201", "ports": 48},
}

# Set of switch role keys, used by topology walks to distinguish switches
# from servers / infra / unknown nodes.
SWITCH_ROLES = frozenset(SWITCH_MODELS.keys())

# Last-resort OS image. Only reachable for a switch whose role never appeared in
# the workbook's VERSIONS table at all; a role that IS pinned but whose version
# has no image-map row now raises UnresolvedSwitchVersionError instead of
# silently landing here. See that class for why.
SWITCH_OS_FALLBACK = "cumulus-linux-vx-amd64-5.16.0.qcow2"
SERVER_OS = "generic/ubuntu2204"


# SN5600-class platform port limit for the cust-net-edge sim switches. NVUE
# will happily accept and apply config on swp65+, so nothing errors -- the
# ports simply are not in the datapath. The margin absorbs late allocations
# (a trunk for an edge added after sizing, an extra ext-storage node).
EDGE_PORT_LIMIT = 64
EDGE_PORT_MARGIN = 6
MAX_EDGES = 32


class UnresolvedSwitchVersionError(RuntimeError):
    """A pinned Cumulus version has no row in the workbook's image map.

    Raised instead of quietly substituting SWITCH_OS_FALLBACK. During the
    5.18.0 upgrade the workbook pinned `core -> 5.18.0` with no matching image
    map row, and the generator emitted cumulus-linux-vx-amd64-5.16.0.qcow2 —
    not 5.18.0, and not even the previous 5.16.1 — with no warning. The sim
    would have come up on 5.16.0 and been signed off as "5.18.0 validated".

    A wrong-but-plausible image is worse than a hard stop: it invalidates the
    test silently. Fail loudly and make the operator add the image-map row.
    """


# Node resource defaults by role.
#
# Storage is uniformly 20 GB across all roles: public NGC Air requires
# a 20 GB minimum per node and we keep one value so Air-Inside runs use
# the same topology without per-env conditionals.
#
# AIR LIMITATION — these are not tuning knobs, they are a hard floor.
# Every Air image publishes `minimum_resources` (GET /api/v3/images/), and Air
# ENFORCES it on all three axes (cpu, memory, storage) at exactly the declared
# value. A node below the floor on ANY single axis does not fail on its own —
# the ENTIRE simulation is rejected into state INVALID with ZERO nodes.
#
# The rejection is asynchronous and silent: POST /simulations/import/ still
# returns HTTP 200 with a simulation id, and nothing on the simulation object
# says why. Air rejects, it never clamps.
#
# So any role's values here must be >= the declared minimum of the image that
# role actually runs. Measured 2026-08-11 (2-4-3-200, import-only):
#   cumulus-vx-5.18.0 declares {cpu 2, memory 4096, storage 20}
#   REJECTED: 1/2048, 2/2048, 1/4096, 2/2560, 2/3072, 2/4096/10GB, 2/4096/15GB
#   ACCEPTED: 2/4096/20GB, and anything above it
# oob/air-oob were 1 cpu / 2048 MB, which exactly matched cumulus-vx-5.15.1's
# {1, 2048, 10} — 5.18.0 is the first pin that put a role BELOW its image floor.
# Full evidence + reproduction:
#   internal-docs/validation-evidence/2026-08-11-cumulus-518-air-minimum-resources.md
#
# NOTE: the declared values are arbitrary per-upload metadata and do NOT track
# any measured requirement (cumulus-vx-5.16.0 was uploaded later than 5.16.1 and
# declares a LOWER minimum). That makes them no less binding — but it also means
# these numbers say nothing about what the OS needs on physical hardware.
NODE_DEFAULTS = {
    "core":    {"cpu": 4, "memory": 4096, "storage": 20},
    "csl":     {"cpu": 4, "memory": 4096, "storage": 20},
    "gsl":     {"cpu": 4, "memory": 4096, "storage": 20},
    # 2/4096 is the cumulus-vx-5.18.0 floor — do NOT lower (see above).
    "oob":     {"cpu": 2, "memory": 4096, "storage": 20},
    "air-oob": {"cpu": 2, "memory": 4096, "storage": 20},
    "edge":    {"cpu": 4, "memory": 4096, "storage": 20},  # cust-net-edge: SN5600-class L2 bridge + eBGP underlay; 2GB/1CPU left it unable to boot the bridge config
    "compute": {"cpu": 1, "memory": 1024, "storage": 20},
    "storage": {"cpu": 1, "memory": 1024, "storage": 20},
    "support": {"cpu": 1, "memory": 1024, "storage": 20},
    "k8s":     {"cpu": 1, "memory": 1024, "storage": 20},
    "bcme":    {"cpu": 1, "memory": 1024, "storage": 20},
    "infra":   {"cpu": 1, "memory": 1024, "storage": 20},
    "unknown": {"cpu": 1, "memory": 1024, "storage": 20},
}

# Wire Map sheet column indices (1-based, openpyxl convention)
# These are the same in both old and new format.
COL_DISPLAY_IN_AIR = 1
COL_SYSTEM_ROLE = 2
COL_SYSTEM_NAME = 3   # Actual node name (may differ from role for OEM naming)
COL_NIC_PORT = 4
COL_NET_PROFILE = 7
COL_SWITCH_ROLE = 11
COL_SWITCH_NAME = 12  # Actual switch name (may differ from role for OEM naming)
COL_SWITCH_PORT = 13

# Air_Only sheet column indices (1-based, new format only)
AIR_ONLY_COL_DISPLAY_IN_AIR = 1
AIR_ONLY_COL_SYSTEM_ROLE = 2
AIR_ONLY_COL_SYSTEM_NAME = 3  # Actual node name (formula-computed)
AIR_ONLY_COL_NIC_PORT = 4
AIR_ONLY_COL_NET_PROFILE = 5
AIR_ONLY_COL_SWITCH_ROLE = 6
AIR_ONLY_COL_SWITCH_NAME = 7  # Actual switch name (formula-computed)
AIR_ONLY_COL_SWITCH_PORT = 8


# ---------------------------------------------------------------------------
# Data class for parsed wiremap rows
# ---------------------------------------------------------------------------

@dataclass
class WiremapRow:
    """One row from the Wire Map Excel sheet."""
    display_in_air: bool
    system_role: str       # Left-side function/role (su-01-node-01, support-07, …) — used for classification
    system_name: str       # Left-side actual node name (may differ from role for OEM naming)
    nic_port: str          # Left-side interface (swpNsM for switches, HW NIC for servers)
    net_profile: str       # Network profile / description
    switch_role: str       # Right-side device role (core-01, oob-switch-01, …) — used for classification
    switch_name: str       # Right-side actual device name (may differ from role for OEM naming)
    switch_port: str       # Right-side interface (swpN, ethNN, …)
    enabled: bool = True   # False if either side's node is Enabled=No on the Nodes tab.
                           # Disabled rows are KEPT (so breakout-analysis sees the
                           # switch_port → subport mapping and emits unconnected stubs)
                           # but skipped by node enumeration and link creation, so the
                           # disabled node doesn't spawn a phantom Air VM. See feedback
                           # memory: "Enabled=No keeps switch port wired".



# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

# classify_node(), is_switch(), generate_mac(), is_valid_hostname()
# imported from utils.py — shared with excel_parser.py


def _cell_str(ws, row: int, col: int) -> str:
    """Read a cell value as a stripped string, treating None as ''."""
    val = ws.cell(row, col).value
    if val is None:
        return ""
    return str(val).strip()


# ---------------------------------------------------------------------------
# Excel parser
# ---------------------------------------------------------------------------

def parse_wiremap_excel(excel_path: Path) -> List[WiremapRow]:
    """Read the Wire Map sheet from the Excel workbook.

    Returns a list of WiremapRow for every data row (skips the header).
    Applies the Nodes-tab Function lookup so rows with blank Function
    cells get their roles resolved from the single source of truth
    (matches `excel_parser._build_wiremap_row_list` behavior).
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    if "Wire Map" not in wb.sheetnames:
        raise ValueError(f"Sheet 'Wire Map' not found in {excel_path}")

    # Build the {hostname: canonical_function} lookup from the Nodes tab.
    # Used to resolve role values for Wire Map rows where the Function cell
    # is blank — same cascading lookup the parser uses.
    nodes_map = {}
    # Build name→Function map AND collect disabled hostnames. R3-7:
    # topology generator previously emitted disabled nodes into the
    # topology JSON, creating phantom Air VMs.
    disabled_names: Set[str] = set()
    if "Nodes" in wb.sheetnames:
        try:
            parsed = parse_nodes(wb["Nodes"])
            nodes_map = build_nodes_function_map(parsed)
            for n in parsed:
                if n.get('status') == 'Disabled':
                    name = (n.get('name') or '').strip()
                    if name:
                        disabled_names.add(name)
        except Exception:
            nodes_map = {}  # missing/garbled Nodes — fall through

    def _resolve(role, name):
        return (nodes_map.get(name) if name else None) or role

    ws = wb["Wire Map"]
    rows: List[WiremapRow] = []
    col_map = build_wiremap_column_map(ws, sheet_kind='wiremap')

    for row_idx in range(2, ws.max_row + 1):
        display_raw = _wm_cell_ws(ws, row_idx, col_map, 'display_in_air').lower()
        system_role_raw = _wm_cell_ws(ws, row_idx, col_map, 'system_role')
        system_name = _wm_cell_ws(ws, row_idx, col_map, 'system_name') or system_role_raw
        nic_port = _wm_cell_ws(ws, row_idx, col_map, 'nic_port')
        net_profile = _wm_cell_ws(ws, row_idx, col_map, 'network_profile')
        switch_role_raw = _wm_cell_ws(ws, row_idx, col_map, 'switch_role')
        switch_name = _wm_cell_ws(ws, row_idx, col_map, 'switch_name') or switch_role_raw
        switch_port = _wm_cell_ws(ws, row_idx, col_map, 'switch_port')

        # Skip empty/spacer rows: both Role AND Name cells blank.
        if not system_role_raw and not system_name:
            continue

        # R3-7 revised: KEEP rows for disabled nodes but tag them disabled.
        # We need the switch_port info for breakout analysis (per feedback
        # memory rule "Enabled=No keeps switch port wired") — otherwise the
        # config emits bondNsM/swpNsM references but the topology has only
        # parent swpN, and ifreload-nvue rolls the apply back. Downstream
        # consumers (_build_nodes, _build_connected_links) skip disabled
        # rows so no phantom Air VM is created.
        row_enabled = (system_name not in disabled_names
                       and switch_name not in disabled_names)

        # Resolve roles via Nodes-tab lookup (cascading fallback).
        system_role = _resolve(system_role_raw, system_name)
        switch_role = _resolve(switch_role_raw, switch_name)

        rows.append(WiremapRow(
            display_in_air=(display_raw == "yes"),
            system_role=system_role,
            system_name=system_name,
            nic_port=nic_port,
            net_profile=net_profile,
            switch_role=switch_role,
            switch_name=switch_name,
            switch_port=switch_port,
            enabled=row_enabled,
        ))

    wb.close()
    return rows


def parse_air_only_sheet(wb) -> List[WiremapRow]:
    """Read the Air_Only sheet from a workbook (new format).

    Air_Only columns (1-based):
      1 = Display in Air
      2 = System Role
      3 = System Name (formula — computed value via data_only)
      4 = NIC/Port
      5 = Network Profile
      6 = Switch Role
      7 = Switch Name (formula — computed value via data_only)
      8 = Switch Port

    Returns WiremapRow list compatible with Wire Map rows.
    """
    if "Air_Only" not in wb.sheetnames:
        return []

    ws = wb["Air_Only"]
    rows: List[WiremapRow] = []
    try:
        col_map = build_wiremap_column_map(ws, sheet_kind='air_only')
    except ValueError:
        # Air_Only sheet is sometimes used purely for non-row data (e.g.
        # version-image mapping). If required headers are absent, treat as
        # "no connection rows" rather than failing the whole generate.
        return []

    for row_idx in range(2, ws.max_row + 1):
        display_raw = _wm_cell_ws(ws, row_idx, col_map, 'display_in_air').lower()
        system_role = _wm_cell_ws(ws, row_idx, col_map, 'system_role')
        system_name = _wm_cell_ws(ws, row_idx, col_map, 'system_name') or system_role
        nic_port    = _wm_cell_ws(ws, row_idx, col_map, 'nic_port')
        net_profile = _wm_cell_ws(ws, row_idx, col_map, 'network_profile')
        switch_role = _wm_cell_ws(ws, row_idx, col_map, 'switch_role')
        switch_name = _wm_cell_ws(ws, row_idx, col_map, 'switch_name') or switch_role
        switch_port = _wm_cell_ws(ws, row_idx, col_map, 'switch_port')

        if not system_role and not system_name:
            continue

        rows.append(WiremapRow(
            display_in_air=(display_raw == "yes"),
            system_role=system_role,
            system_name=system_name,
            nic_port=nic_port,
            net_profile=net_profile,
            switch_role=switch_role,
            switch_name=switch_name,
            switch_port=switch_port,
        ))

    return rows


def parse_version_image_map(wb) -> dict:
    """Read the version→Air image mapping table from the Air_Only sheet.

    The table starts after the connection rows, identified by a row where
    col 1 = 'Friendly Version' (or similar) followed by data rows.

    Returns dict: {'5.16.1': 'cumulus-linux-vx-amd64-5.16.1.0008.qcow2', ...}
    """
    if "Air_Only" not in wb.sheetnames:
        return {}

    ws = wb["Air_Only"]
    image_map = {}
    in_table = False

    for row_idx in range(1, ws.max_row + 1):
        col1 = _cell_str(ws, row_idx, 1)
        col2 = _cell_str(ws, row_idx, 2)
        if col1.lower() in ('friendly version', 'version'):
            in_table = True
            continue
        if in_table:
            if not col1:
                break
            image_map[col1] = col2
    return image_map


def parse_air_settings(wb) -> dict:
    """Read key-value settings from the Air_Only sheet.

    Scans for rows where col1 is a known setting name (e.g., 'Air Management Subnet').
    Returns dict: {'air_mgmt_subnet': '172.20.0.0/24', ...}
    """
    if "Air_Only" not in wb.sheetnames:
        return {}

    ws = wb["Air_Only"]
    settings = {}
    _KNOWN_KEYS = {
        'air management subnet': 'air_mgmt_subnet',
    }

    for row_idx in range(1, ws.max_row + 1):
        col1 = _cell_str(ws, row_idx, 1).lower()
        col2 = _cell_str(ws, row_idx, 2)
        if col1 in _KNOWN_KEYS and col2:
            settings[_KNOWN_KEYS[col1]] = col2
    return settings


def parse_switch_versions(wb) -> dict:
    """Read the VERSIONS table from the Settings sheet (new format).

    Returns dict: {'core': '5.16.1', 'oob': '5.15.1'} or {} if not present.
    """
    if "Settings" not in wb.sheetnames:
        return {}

    ws = wb["Settings"]
    versions = {}
    in_versions = False

    for row_idx in range(1, ws.max_row + 1):
        key = ws.cell(row_idx, 1).value
        if key is None:
            continue
        key_str = str(key).strip()
        if key_str.lower() == 'switch function':
            in_versions = True
            continue
        if in_versions:
            value = ws.cell(row_idx, 2).value
            if not key_str or value is None:
                break
            versions[key_str.lower()] = str(value).strip()
    return versions


# ---------------------------------------------------------------------------
# Generator
# ---------------------------------------------------------------------------

class TopologyGenerator:
    """Generate an Air 2.0 topology JSON from the Excel wiremap."""

    # Infra/jump nodes that must survive a switches-only strip — the Ansible
    # jump host + L3-OOB plumbing that validation needs to reach the switches.
    # (Cumulus nodes are kept by OS; these are the Ubuntu nodes we keep.)
    SWITCHES_ONLY_INFRA_KEEP = (
        "oob-server", "dhcp-oob", "dhcp-edge", "utility",
        "external-conn", "external-dhcp", "cust-net-edge", "air-oob",
    )

    def __init__(self, excel_path: Path, arch: str, site: str = "default",
                 switches_only: bool = False):
        self.excel_path = excel_path
        self.arch = arch
        self.site = site
        self.switches_only = switches_only

        wb = openpyxl.load_workbook(excel_path, data_only=True)
        new_format = "Air_Only" in wb.sheetnames

        # Wire Map rows — always present
        self.rows = parse_wiremap_excel(excel_path)

        if new_format:
            # New format: Air topology rows (infrastructure + outbound) are in Air_Only
            # Air_Only rows go FIRST so they take priority in dedup over Wire Map rows
            self.rows = parse_air_only_sheet(wb) + self.rows
            # Build version → OS image mapping from Air_Only and per-function versions
            image_map = parse_version_image_map(wb)
            switch_versions = parse_switch_versions(wb)
            self._switch_os = self._build_switch_os(image_map, switch_versions)
        else:
            self._switch_os = {}

        # OOB uplink mode — drives whether we inject the L2 air-oob-switch
        # flat bridge (existing default) or the L3 three-Ubuntu-node trio
        # (external-conn + external-dhcp + utility). Read from Settings sheet
        # with same snake_case normalization as scripts/excel_parser.py.
        self._oob_uplink_mode = "l2"
        if "Settings" in wb.sheetnames:
            ws_settings = wb["Settings"]
            for row_idx in range(1, ws_settings.max_row + 1):
                key = ws_settings.cell(row_idx, 1).value
                value = ws_settings.cell(row_idx, 2).value
                if key is None or value is None:
                    continue
                key_clean = str(key).strip().lower().replace(" ", "_").replace("-", "_")
                if key_clean == "oob_uplink_mode":
                    self._oob_uplink_mode = str(value).strip().lower()
                    break

        wb.close()

        # Build actual-name → function/role mapping for OEM-named devices.
        # Classification helpers (is_switch, classify_node) need the function name.
        # Only map when we have a non-empty role; otherwise classify_node falls
        # back to the name itself (correctly recognising cust-net-edge-* etc.).
        self._name_to_role: Dict[str, str] = {}
        for r in self.rows:
            if r.system_name and r.system_role and r.system_name != r.system_role:
                self._name_to_role[r.system_name] = r.system_role
            if r.switch_name and r.switch_role and r.switch_name != r.switch_role:
                self._name_to_role[r.switch_name] = r.switch_role

        # Supplement from Nodes tab (Function → Name) when Wire Map lacks role columns.
        # This handles OEM-named deployments where the Wire Map only has System Name.
        if "Nodes" in wb.sheetnames:
            wb_nodes = openpyxl.load_workbook(str(self.excel_path), data_only=True, read_only=True)
            ws_n = wb_nodes["Nodes"]
            for row in ws_n.iter_rows(min_row=2, max_row=ws_n.max_row, values_only=True):
                func = str(row[0]).strip() if row[0] else ''
                name = str(row[1]).strip() if row[1] else ''
                if func and name and name not in self._name_to_role:
                    self._name_to_role[name] = func
            wb_nodes.close()

        # Per-server counter for sequential ethN assignment
        self._server_eth_counter: Dict[str, int] = defaultdict(int)

        # Pre-scan for explicit ethN assignments (e.g., eth0 rows from Air_Only or Wire Map)
        # so _next_eth() can skip them and avoid duplicates
        self._explicit_eth: Dict[str, Set[int]] = defaultdict(set)
        for r in self.rows:
            if r.display_in_air and not is_switch(r.system_role):
                m = re.search(r"eth(\d+)", r.nic_port or "")
                if m:
                    self._explicit_eth[r.system_name].add(int(m.group(1)))

        # Pre-scan: reserve eth0 for the first oob-switch connection per server.
        # This ensures the OOB management interface is always eth0, regardless of
        # row order in the Wire Map.  Stores (system_name, switch_name, switch_port)
        # tuples to identify the chosen row in _build_connected_links().
        self._oob_eth0: Dict[str, Tuple[str, str]] = {}  # node → (switch_name, switch_port)
        for r in self.rows:
            if not r.display_in_air or is_switch(r.system_role):
                continue
            if r.system_name in self._oob_eth0:
                continue  # already found one for this node
            peer_role = classify_node(self._name_to_role.get(r.switch_name, r.switch_role))
            if peer_role == "oob" and r.switch_port:
                self._oob_eth0[r.system_name] = (r.switch_name, r.switch_port)
                self._explicit_eth[r.system_name].add(0)  # reserve eth0

        # Same reservation for ext-storage-*, which reaches the management
        # network by a different route and so never matches the loop above.
        #
        # It has no oob-switch row in the Wire Map — its eth0 is synthesised
        # later against cust-net-edge (the Air-mgmt bridge) so it can hold a
        # 172.20.0.x address and reach NAT/DNS. Because that link is created
        # after data-link allocation and was never registered here, _next_eth()
        # started at 0 and handed eth0 to the FIRST STORAGE UPLINK. The mgmt
        # link then claimed eth0 as well, and Air rejected the whole topology:
        #
        #   Interface eth0 is already defined for node ext-storage-01
        #
        # That failed the import in ~479ms and left an empty INVALID simulation,
        # which is not visible in the Air UI — so `make air-deploy` was broken
        # on every architecture with no obvious cause.
        #
        # eth0 is the management interface on every node, always. Reserving it
        # here means the Wire Map's `swp1`/`swp2` for ext-storage land on
        # eth1/eth2, matching both the compute-node convention (18 data links,
        # none on eth0) and this file's own long-standing comment that
        # ext-storage is "wired to CSL swp63 ports via Wire Map (eth1, eth2)".
        for name in {r.system_name for r in self.rows
                     if r.display_in_air and not is_switch(r.system_role)
                     and (r.system_name or "").startswith("ext-storage-")}:
            self._explicit_eth[name].add(0)

    # ---- public -----------------------------------------------------------

    def generate(self) -> dict:
        """Build the complete topology dict."""
        nodes = self._build_nodes()
        links, switch_connected = self._build_connected_links()
        if self._oob_uplink_mode == "l3":
            nodes, links, switch_connected = self._inject_l3_oob_nodes(
                nodes, links, switch_connected,
            )
        else:
            nodes, links, switch_connected = self._inject_air_oob_switch(
                nodes, links, switch_connected,
            )
        links += self._build_unconnected_stubs(switch_connected)
        if self.switches_only:
            nodes, links = self._strip_to_switches(nodes, links)
        # Re-run layout now that all infra/injected nodes exist.
        self._apply_layout(nodes)
        self._assert_no_duplicate_interfaces(links)

        return {
            "format": "JSON",
            "title": f"ERA-{self.site}-{self.arch}",
            "ztp": None,
            "content": {
                "nodes": nodes,
                "links": links,
                "oob": False,
            },
            # Metadata for air-deploy.py Node Instructions
            "_air_oob": getattr(self, '_air_oob_metadata', {}),
            "_l3_oob": getattr(self, '_l3_oob_metadata', {}),
            "_oob_uplink_mode": self._oob_uplink_mode,
        }

    def _strip_to_switches(self, nodes: dict, links: list) -> tuple:
        """Drop server VMs for a switches-only sim, keeping every switch + the
        infra/jump nodes validation needs.

        Server-facing switch ports are preserved as ``unconnected`` stubs so the
        generated switch configs still reference existing interfaces (Air rolls
        back the whole apply if a referenced port is missing). The bulk server
        VMs (compute/gpu/storage/support/…) are what we shed — at 2-tier scale
        they're most of the node count, so dropping them frees the budget to run
        much larger switch fabrics. Breakout parents are turned into sub-port
        stubs afterwards by patch-air-breakout-stubs.py.
        """
        def _keep(name: str, node: dict) -> bool:
            if "cumulus" in (node.get("os") or "").lower():
                return True  # any switch
            return name.startswith(self.SWITCHES_ONLY_INFRA_KEEP)

        kept = {n: v for n, v in nodes.items() if _keep(n, v)}
        dropped = set(nodes) - set(kept)

        new_links = []
        for link in links:
            if not isinstance(link, list):
                new_links.append(link)
                continue
            eps = [e for e in link if isinstance(e, dict)]
            if not any(e.get("node") in dropped for e in eps):
                new_links.append(link)  # untouched (switch↔switch, infra, etc.)
                continue
            # A dropped server sat on one end; keep each surviving endpoint as an
            # unconnected stub so its switch interface still exists in the sim.
            for e in eps:
                if e.get("node") in kept:
                    new_links.append([e, "unconnected"])

        return kept, new_links

    def write(self, output_path: Path) -> None:
        """Generate and write topology JSON to *output_path*."""
        topology = self.generate()
        output_path.parent.mkdir(parents=True, exist_ok=True)
        with open(output_path, "w") as f:
            json.dump(topology, f, indent=4)

        node_count = len(topology["content"]["nodes"])
        link_count = len(topology["content"]["links"])
        connected = sum(
            1 for lnk in topology["content"]["links"]
            if isinstance(lnk[1], dict)
        )
        unconnected = sum(
            1 for lnk in topology["content"]["links"]
            if isinstance(lnk[1], str) and lnk[1] == "unconnected"
        )
        outbound = sum(
            1 for lnk in topology["content"]["links"]
            if isinstance(lnk[1], str) and lnk[1] == "outbound"
        )

        print(f"  Topology written to {output_path}")
        print(f"  Nodes:       {node_count}")
        print(f"  Links:       {link_count}")
        print(f"    connected:   {connected}")
        print(f"    unconnected: {unconnected}")
        print(f"    outbound:    {outbound}")

    # ---- nodes ------------------------------------------------------------

    def _build_nodes(self) -> dict:
        """Collect every device that should appear as a node."""
        device_names: Set[str] = set()

        for r in self.rows:
            if not r.display_in_air:
                continue
            # Skip rows whose A-side or B-side node is disabled — those rows
            # exist only to retain switch-port info for breakout analysis
            # (see WiremapRow.enabled). Don't spawn a phantom Air VM for them.
            if not r.enabled:
                continue
            # A-side: collect if we have a valid hostname (regardless of
            # whether the Function column was filled — it's optional now).
            if is_valid_hostname(r.system_name):
                device_names.add(r.system_name)
            # B-side: same. Skip when role explicitly says NA/OUTBOUND
            # (those aren't real peers) — note the role check is _OR_
            # against blank, since post-Nodes-lookup an empty Function (B)
            # cell is normal.
            sr_upper = r.switch_role.upper() if r.switch_role else ''
            if sr_upper not in ("NA", "OUTBOUND") and is_valid_hostname(r.switch_name):
                device_names.add(r.switch_name)

        nodes = {}
        for name in sorted(device_names):
            role_name = self._name_to_role.get(name, name)
            role = classify_node(role_name)
            defaults = NODE_DEFAULTS.get(role, NODE_DEFAULTS["support"])
            nodes[name] = {
                "cpu": defaults["cpu"],
                "memory": defaults["memory"],
                "storage": defaults["storage"],
                "positioning": {"x": 0, "y": 0},
                "os": self._resolve_os(name),
                "features": {"uefi": False, "tpm": False},
                "pxehost": False,
                "secureboot": False,
                "oob": False,
                "emulation_type": None,
                "network_pci": {},
            }

        # Layout is applied later (in generate()) once air-oob-switch
        # has been injected — see comment there.
        return nodes

    @staticmethod
    def _build_switch_os(image_map: dict, switch_versions: dict) -> Dict[str, str]:
        """Resolve the VERSIONS table into a role → Air image mapping.

        The workbook's VERSIONS table is keyed by the arch models' function
        vocabulary (`oob-switch`, `gsl-plane1`, `cs`, …). `_resolve_os()` looks
        roles up by the classify_node() vocabulary (`oob`, `gsl`, `csl`, …).
        Those two only coincide for `core`, so keying this dict by the raw
        function name meant every other role missed the lookup — a 7-model
        version bump moved exactly one row. Index under BOTH spellings so
        either vocabulary resolves.

        Raises UnresolvedSwitchVersionError rather than substituting
        SWITCH_OS_FALLBACK: a wrong-but-plausible image silently invalidates
        whatever was validated on it.
        """
        by_key: Dict[str, str] = {}
        conflicts: Dict[str, set] = {}

        for func, version in switch_versions.items():
            role = classify_node(func)
            # Index under the model spelling AND the classified role.
            for key in {str(func).strip().lower(), role}:
                if not key:
                    continue
                if key in by_key and by_key[key] != version:
                    conflicts.setdefault(key, {by_key[key]}).add(version)
                by_key[key] = version

        if conflicts:
            detail = "; ".join(
                f"{k} <- {sorted(v)}" for k, v in sorted(conflicts.items())
            )
            raise UnresolvedSwitchVersionError(
                "Conflicting Cumulus versions collapse onto the same switch "
                f"role: {detail}. Model functions that map to one role (e.g. "
                "`csl` and `cs`, or `gsl-plane1` and `gs-plane1`) must pin the "
                "same version."
            )

        resolved: Dict[str, str] = {}
        unresolved: Dict[str, str] = {}
        for key, version in by_key.items():
            image = image_map.get(version)
            if image:
                resolved[key] = image
            else:
                unresolved[key] = version

        if unresolved:
            wanted = sorted(set(unresolved.values()))
            known = sorted(image_map) or ["<empty>"]
            raise UnresolvedSwitchVersionError(
                f"No Air image mapping for Cumulus version(s) {wanted} "
                f"pinned by switch function(s) {sorted(unresolved)}. "
                f"The workbook's Air_Only image map knows: {known}. "
                "Add the missing 'Friendly Version' → 'Air Image' row "
                "(data-models/generate_arch_excel.py, append_air_only) and "
                "regenerate the workbook. Refusing to substitute a different "
                "image, which would silently validate the wrong version."
            )
        return resolved

    def _resolve_os(self, name: str) -> str:
        """Return the Air OS image string for a node.

        For servers always returns SERVER_OS.
        For switches, looks up per-function image from _switch_os (populated from
        the Air_Only version mapping table when using new format).
        Falls back to SWITCH_OS_FALLBACK if no mapping is available.
        """
        role_name = self._name_to_role.get(name, name)
        if not is_switch(role_name):
            return SERVER_OS
        role = classify_node(role_name)  # 'core', 'oob', 'edge', …
        if role in self._switch_os:
            return self._switch_os[role]
        # air-oob shares the oob image (SN2201 Air OOB switch)
        if role == "air-oob" and "oob" in self._switch_os:
            return self._switch_os["oob"]
        # cust-net-edge (edge) is an SN5600-class L2-bridge + eBGP-underlay node
        # that runs the same heavy config as core/csl, so it tracks core's image
        # (5.16.1) — NOT oob's 5.15.1. (It previously inherited oob and shipped
        # 5.15.1, which under the 2GB 'edge' default left edge01 unable to boot
        # the bridge → the whole air-mgmt plane stayed isolated.)
        if role in ("edge", "csl", "gsl") and "core" in self._switch_os:
            return self._switch_os["core"]
        return SWITCH_OS_FALLBACK

    # ---- layout -------------------------------------------------------------

    def _apply_layout(self, nodes: dict) -> None:
        """Assign grid positions to nodes based on role.

        Uses a 275-unit grid matching NVIDIA Air visual layout conventions.
        Layout regions (left-to-right):
          - Left:   infrastructure (edge switches, OOB switches, dhcp/oob-server)
          - Center: core switches, k8s, bcme, storage columns
          - Right:  compute (su-*) nodes grouped by SU

        After role-based placement, a final sweep scatters any nodes that
        didn't get explicit positions and resolves overlaps — keeps the
        Air UI readable when an arch ships nodes the layout heuristics
        don't know about (e.g. ext-storage in L2 mode, future virtual
        nodes added to the Wire Map).
        """
        G = 275  # grid unit (pixels)

        def _role(n: str) -> str:
            return classify_node(self._name_to_role.get(n, n))

        # Categorise nodes by role
        edges = sorted(n for n in nodes if _role(n) == "edge")
        cores = sorted(n for n in nodes if _role(n) == "core")
        # csl (CPU/Storage Leaf) is positionally equivalent to core in
        # non-collapsed designs (2-8-9-800 dedicated_gpu).
        csls = sorted(n for n in nodes if _role(n) == "csl")
        gsls_p1 = sorted(n for n in nodes if _role(n) == "gsl" and "plane1" in n.lower())
        gsls_p2 = sorted(n for n in nodes if _role(n) == "gsl" and "plane2" in n.lower())
        oobs = sorted(n for n in nodes if _role(n) == "oob")
        infras = sorted(n for n in nodes if _role(n) == "infra")
        k8s_nodes = sorted(n for n in nodes if n.lower().startswith("k8s"))
        bcme_nodes = sorted(n for n in nodes if n.lower().startswith("bcme"))
        bcm_nodes = sorted(n for n in nodes
                           if n.lower().startswith("bcm-") and not n.lower().startswith("bcme"))
        slurm_nodes = sorted(n for n in nodes if n.lower().startswith("slurm"))
        storage_nodes = sorted(n for n in nodes if _role(n) == "storage")
        # Support nodes that aren't already in a more specific bucket
        _specific = set(k8s_nodes) | set(bcme_nodes) | set(bcm_nodes) | set(slurm_nodes)
        support_other = sorted(
            n for n in nodes
            if _role(n) == "support" and n not in _specific
        )
        compute = sorted(n for n in nodes if _role(n) == "compute")

        # Dedicated-GPU layout (csl/gsl arch, e.g. 2-8-9-800) — distinct
        # placement rules:
        #   row 0 (y = -G):  GSLs on a single row, plane1 leaves grouped left,
        #                    plane2 leaves grouped right (one-column gap between).
        #   row 1 (y =  0):  CSLs adjacent on the same row, centered under GSLs.
        #   row 2 (y =  G):  OOB row — air-oob-switch, oob-switches, dhcp-oob,
        #                    oob-server, dhcp-edge all on one row, in that order.
        #   row 3+ (y = 2G+): server columns (gpu | bcm | slurm | k8s | …).
        if csls or gsls_p1 or gsls_p2:
            self._apply_dedicated_gpu_layout(
                nodes, G, csls, gsls_p1, gsls_p2, oobs, infras,
                compute, bcm_nodes, slurm_nodes, k8s_nodes, bcme_nodes,
                storage_nodes, support_other,
            )
            return

        # Parse compute nodes into SU groups: { su_id: [node_names sorted] }
        su_groups: Dict[str, list] = {}
        for name in compute:
            # Extract SU identifier (e.g., "su-01" from "su-01-node-03")
            m = re.match(r"(su-\d+)", name.lower())
            su_id = m.group(1) if m else "su-00"
            su_groups.setdefault(su_id, []).append(name)

        su_ids = sorted(su_groups.keys())

        # Determine how many "server columns" we need to the right of core-01
        # Columns: k8s | bcme | storage | su-01 | su-02 | su-03 | ...
        server_col_names = []
        if k8s_nodes:
            server_col_names.append("k8s")
        if bcme_nodes:
            server_col_names.append("bcme")
        if storage_nodes:
            server_col_names.append("storage")
        if support_other:
            server_col_names.append("support")
        for su_id in su_ids:
            server_col_names.append(su_id)

        # X assignments
        # Infrastructure on the left: cols -1, 0, 1 (x = -G, 0, G)
        infra_x_base = -1  # grid col for leftmost infra

        # Core-01 starts at col 2 (x = 2*G = 550)
        core_start_col = 2
        # Server columns start at core_start_col + 1
        server_start_col = core_start_col + 1

        # Core-02 goes to the right of all server columns
        core2_col = server_start_col + len(server_col_names)

        # Map server column names to x grid columns
        server_col_x = {}
        for i, col_name in enumerate(server_col_names):
            server_col_x[col_name] = server_start_col + i

        # --- Row 0 (Y=0): edge switch #2 (top) ---
        # --- Row 1 (Y=G): edge switch #1, core-01, core-02 ---
        # --- Row 2 (Y=2G): OOB switches + first server row ---
        # --- Row 3+ : more server rows ---
        # --- Below servers: infra nodes (dhcp-edge, oob-server, dhcp-oob) ---

        # Place edge switches
        for i, name in enumerate(edges):
            if i == 0:
                nodes[name]["positioning"] = {"x": 0, "y": G}
            elif i == 1:
                nodes[name]["positioning"] = {"x": G, "y": 0}
            else:
                nodes[name]["positioning"] = {"x": i * G, "y": 0}

        # Place core switches (collapsed/converged designs only — dedicated_gpu
        # arches are handled by _apply_dedicated_gpu_layout above).
        if len(cores) >= 1:
            nodes[cores[0]]["positioning"] = {
                "x": core_start_col * G, "y": G,
            }
        if len(cores) >= 2:
            nodes[cores[1]]["positioning"] = {
                "x": core2_col * G, "y": G,
            }
        for i, name in enumerate(cores[2:], 2):
            nodes[name]["positioning"] = {"x": (core2_col + i - 1) * G, "y": G}

        # Place OOB switches — stacked in a single column (left infra zone).
        # In L2 mode the air-oob-switch sits next to them; in L3 mode utility
        # sits to their LEFT (one column left). The column placement keeps
        # them visually grouped regardless of how many OOB switches the arch
        # has (2, 3, or more).
        for i, name in enumerate(oobs):
            nodes[name]["positioning"] = {
                "x": (infra_x_base + 1) * G,
                "y": (2 + i) * G,
            }

        # Place air-oob-switch (row 3, left side — between OOB switches and infra)
        if "air-oob-switch" in nodes:
            nodes["air-oob-switch"]["positioning"] = {
                "x": infra_x_base * G, "y": 3 * G,
            }

        # Determine max server rows (tallest column)
        col_heights = {
            "k8s": len(k8s_nodes),
            "bcme": len(bcme_nodes),
            "storage": len(storage_nodes),
            "support": len(support_other),
        }
        for su_id in su_ids:
            col_heights[su_id] = len(su_groups[su_id])
        max_server_rows = max(col_heights.values()) if col_heights else 0

        # Place server columns starting at row 2
        server_row_base = 2

        def _place_column(node_list: list, col_name: str) -> None:
            if col_name not in server_col_x:
                return
            x = server_col_x[col_name] * G
            for row_i, name in enumerate(node_list):
                nodes[name]["positioning"] = {
                    "x": x, "y": (server_row_base + row_i) * G,
                }

        _place_column(k8s_nodes, "k8s")
        _place_column(bcme_nodes, "bcme")
        _place_column(storage_nodes, "storage")
        _place_column(support_other, "support")
        for su_id in su_ids:
            _place_column(su_groups[su_id], su_id)

        # Place infra nodes (dhcp-edge, oob-server, dhcp-oob) below OOB switches
        # Sort: edge-related first, then oob-server in the middle, then dhcp-oob
        def _infra_sort_key(n: str) -> tuple:
            nl = n.lower()
            if "edge" in nl:
                return (0, nl)
            if "server" in nl:
                return (1, nl)
            return (2, nl)
        infras = sorted(infras, key=_infra_sort_key)

        infra_row = 4  # default position matching reference layout
        if max_server_rows > 2:
            infra_row = server_row_base + 2  # two rows below OOB switches
        for i, name in enumerate(infras):
            nodes[name]["positioning"] = {
                "x": (infra_x_base + i) * G, "y": infra_row * G,
            }

        # L3 OOB layout overrides — see docs/plans/2026-05-20-l3-oob-air-topology.md.
        # Visual hierarchy from top to bottom:
        #   y = -2G : external-conn, external-dhcp   (Ubuntu, simulated ISP/DC)
        #   y = -G  : cust-net-edge-01, cust-net-edge-02   (Cumulus, EXIT VRF)
        #   y =  G  : cores                                 (existing)
        #   y = 2G  : oob-switches + first server row       (existing)
        #   ...     : server columns + utility in infra row
        if self._oob_uplink_mode == "l3":
            # Place cust-net-edge above cores (one row up).
            if len(edges) >= 1 and edges[0] in nodes:
                nodes[edges[0]]["positioning"] = {
                    "x": core_start_col * G, "y": -G,
                }
            if len(edges) >= 2 and edges[1] in nodes:
                nodes[edges[1]]["positioning"] = {
                    "x": core2_col * G, "y": -G,
                }
            # external-conn above cust-net-edge-01 (top row), external-dhcp
            # LEFT of cust-net-edge-01 (same row as the edges).
            if "external-conn" in nodes:
                nodes["external-conn"]["positioning"] = {
                    "x": core_start_col * G, "y": -2 * G,
                }
            if "external-dhcp" in nodes:
                nodes["external-dhcp"]["positioning"] = {
                    "x": (core_start_col - 1) * G, "y": -G,
                }
            # utility sits to the LEFT of oob-switch-01 (same row).
            if "utility" in nodes:
                nodes["utility"]["positioning"] = {
                    "x": infra_x_base * G,
                    "y": 2 * G,
                }

        self._resolve_position_collisions(nodes, G)

    def _resolve_position_collisions(self, nodes: dict, G: int) -> None:
        """Resolve overlapping (x, y) positions in the Air topology.

        Two failure modes this catches:
          - Nodes that never received an explicit positioning call sit at
            the default (0, 0) — happens when a Wire Map references a
            virtual node (e.g. ext-storage-*) the layout heuristic doesn't
            classify.
          - Two role-based placements that happen to land on the same
            grid cell (e.g. oob-switch-03 + dhcp-oob both at (0, 4G)
            when the arch templates 3 OOB switches).

        Strategy: any group of >1 node sharing a coordinate gets fanned
        out horizontally on the grid starting at the original x. Stable
        order: alphabetical by name, so regenerations are deterministic.
        """
        from collections import defaultdict
        occupied = defaultdict(list)
        for name, nd in nodes.items():
            pos = nd.get("positioning")
            if not isinstance(pos, dict):
                continue
            occupied[(pos.get("x", 0), pos.get("y", 0))].append(name)
        for (x, y), names in occupied.items():
            if len(names) <= 1:
                continue
            # Stable order so output is deterministic.
            names.sort()
            # Leave the first node where it is; shift the rest right by G.
            # If a shifted cell is itself occupied, keep stepping until
            # we find empty space.
            all_taken = {k for k, v in occupied.items()}
            for i, n in enumerate(names[1:], 1):
                step = i
                while True:
                    candidate = (x + step * G, y)
                    if candidate not in all_taken:
                        break
                    step += 1
                nodes[n]["positioning"] = {"x": candidate[0], "y": candidate[1]}
                all_taken.add(candidate)

    def _apply_dedicated_gpu_layout(
        self, nodes, G, csls, gsls_p1, gsls_p2, oobs, infras,
        compute, bcm_nodes, slurm_nodes, k8s_nodes, bcme_nodes,
        storage_nodes, support_other,
    ) -> None:
        """Placement for non-collapsed designs (e.g. 2-8-9-800 dual-plane).

        Visual structure (validated against the Air UI by hand):

            y = -G    Air infra row:    dhcp-edge | oob-server | air-oob | dhcp-oob
            y =  0    Switch row:       CSL-01 |·| CSL-02 |····| GSLs plane1 |·| GSLs plane2
                                                ▲           ▲                  ▲
            y = G+    Server columns:        ctrl col    oob col            gpu col
                                            (k8s/bcm/  (oob-switch       (gpu compute,
                                             slurm)     vertical stack)   between planes)

        Three vertical channels interleave between switch positions on the
        switch row; servers stack downward in those channels. Air-management
        nodes sit on a row above the switches to keep the data plane (CSL/GSL
        + servers) visually grouped.
        """
        # Split spine vs leaf per fabric — in 2-tier archs (e.g. 2-4-5-800)
        # the `gs-*` nodes are dedicated GPU spines and `cs-*` are dedicated
        # compute spines, which should sit visually ABOVE their respective
        # leaves. In converged 1-tier archs (e.g. 2-8-9-800) the gsl/csl
        # nodes contain everything and these "spine" lists are empty, so
        # layout stays as it was.
        gpu_spines_p1 = [n for n in gsls_p1 if n.lower().startswith('gs-')]
        gpu_leaves_p1 = [n for n in gsls_p1 if not n.lower().startswith('gs-')]
        gpu_spines_p2 = [n for n in gsls_p2 if n.lower().startswith('gs-')]
        gpu_leaves_p2 = [n for n in gsls_p2 if not n.lower().startswith('gs-')]
        cpu_spines    = [n for n in csls if n.lower().startswith('cs-')]
        cpu_leaves    = [n for n in csls if not n.lower().startswith('cs-')]
        has_two_tier = bool(gpu_spines_p1 or gpu_spines_p2 or cpu_spines)
        leaf_y  = 0
        spine_y = -G            # only populated in 2-tier archs
        air_y   = -2 * G if has_two_tier else -G

        # ---- Leaf row (y = 0): CSLs + GSL leaves ----
        # CSLs use even-spaced columns so the control-plane column can slot
        # between them.
        csl_start = 6
        csl_step = 2
        for i, name in enumerate(cpu_leaves):
            nodes[name]["positioning"] = {"x": (csl_start + i * csl_step) * G, "y": leaf_y}

        # OOB switch column sits 2 cols to the right of the last CSL leaf.
        csl_end = csl_start + (len(cpu_leaves) - 1) * csl_step if cpu_leaves else csl_start
        oob_col = csl_end + 2

        # GPU leaf plane1 — adjacent leaves, starting 2 cols after the OOB column.
        gsl_p1_start = oob_col + 2
        for i, name in enumerate(gpu_leaves_p1):
            nodes[name]["positioning"] = {"x": (gsl_p1_start + i) * G, "y": leaf_y}

        # GPU column sits between plane1 and plane2.
        gpu_col = gsl_p1_start + len(gpu_leaves_p1)

        # GPU leaf plane2 — adjacent leaves, one column past the GPU column.
        gsl_p2_start = gpu_col + 1
        for i, name in enumerate(gpu_leaves_p2):
            nodes[name]["positioning"] = {"x": (gsl_p2_start + i) * G, "y": leaf_y}

        # ---- Spine row (y = -G) — 2-tier archs only ----
        # Place GPU spines centered above their plane's leaves.
        for i, name in enumerate(gpu_spines_p1):
            # Spine columns centered in the leaf block (4 leaves -> spines span
            # cols 0..1 of that block); for typical 2 spines they sit cleanly.
            nodes[name]["positioning"] = {"x": (gsl_p1_start + i) * G, "y": spine_y}
        for i, name in enumerate(gpu_spines_p2):
            nodes[name]["positioning"] = {"x": (gsl_p2_start + i) * G, "y": spine_y}
        # CPU spines above their CSL leaves (cs-* sits over cl-*).
        for i, name in enumerate(cpu_spines):
            nodes[name]["positioning"] = {"x": (csl_start + i * csl_step) * G, "y": spine_y}

        # ---- Air infra row (y = -G in 1-tier, y = -2G in 2-tier) ----
        # L2 mode: dhcp-edge | oob-server | air-oob | dhcp-oob
        # L3 mode: external-conn | external-dhcp | (cust-net-edge already
        # placed in edges row) | utility (sits above OOB column on VLAN 200)
        ctrl_col = csl_start + 1   # between CSL-01 and CSL-02
        # air_y was computed above based on has_two_tier.
        placements = {
            # L2 mode infra
            'dhcp-edge':       ctrl_col,
            'oob-server-01':   ctrl_col + 2,
            'air-oob-switch':  oob_col,
            'dhcp-oob':        oob_col + 1,
            # L3 mode infra (none of these clash with L2 names since they're
            # mode-exclusive — only the relevant set will exist in nodes)
            'external-conn':   ctrl_col,
            'external-dhcp':   ctrl_col + 2,
            'utility':         oob_col + 1,
        }
        placed_infra = set()
        for name, col in placements.items():
            if name in nodes:
                nodes[name]["positioning"] = {"x": col * G, "y": air_y}
                placed_infra.add(name)
        # Any extra infra nodes get sequential columns past dhcp-oob.
        next_extra = oob_col + 2
        for name in infras:
            if name not in placed_infra:
                nodes[name]["positioning"] = {"x": next_extra * G, "y": air_y}
                next_extra += 1

        # ---- Server columns (y = G+) ----
        server_row_start = 1
        # Control-plane column: k8s/bcm/slurm/bcme/storage/support stacked
        ctrl_servers = (
            list(k8s_nodes) + list(bcm_nodes) + list(slurm_nodes)
            + list(support_other) + list(bcme_nodes) + list(storage_nodes)
        )
        for i, name in enumerate(ctrl_servers):
            nodes[name]["positioning"] = {
                "x": ctrl_col * G, "y": (server_row_start + i) * G,
            }
        # OOB switch column
        for i, name in enumerate(oobs):
            nodes[name]["positioning"] = {
                "x": oob_col * G, "y": (server_row_start + i) * G,
            }
        # GPU servers — grouped by SU, each SU a vertical column (its nodes stack
        # downward, matching the collapsed-arch convention). SU columns are laid
        # out in a grid `num_leaves` wide so the row of SUs lines up with the
        # rail-leaf block above; once the row fills it wraps to a new band below
        # (SU 9 sits under SU 1 for 8 leaves), with a blank spacer row between
        # bands. Flat gpu-NN naming (no SU structure) falls back to one column.
        gpu_leaf_cols = [gsl_p1_start + i for i in range(len(gpu_leaves_p1))]
        block_x = gpu_leaf_cols[0] if gpu_leaf_cols else gpu_col
        num_leaves = len(gpu_leaves_p1) or 1
        gpu_su_groups: Dict[str, list] = {}
        for name in compute:
            m = re.match(r"(su-\d+)", name.lower())
            gpu_su_groups.setdefault(m.group(1) if m else "", []).append(name)
        if len(gpu_su_groups) == 1 and "" in gpu_su_groups:
            for i, name in enumerate(compute):
                nodes[name]["positioning"] = {
                    "x": gpu_col * G, "y": (server_row_start + i) * G,
                }
        else:
            su_ids = sorted(gpu_su_groups)
            su_height = max((len(v) for v in gpu_su_groups.values()), default=1)
            band_stride = su_height + 1  # +1 = blank spacer row between bands
            for idx, su_id in enumerate(su_ids):
                band, col = divmod(idx, num_leaves)
                x = (block_x + col) * G
                y0 = server_row_start + band * band_stride
                for j, name in enumerate(sorted(gpu_su_groups[su_id])):
                    nodes[name]["positioning"] = {"x": x, "y": (y0 + j) * G}

        # L3 OOB layout overrides for dedicated-GPU arches.
        # The cust-net-edge row must clear the spine row. In 2-tier archs
        # (cs-*/gs-* present) the spine row occupies y = -G, so the edges sit
        # one row higher; in 1-tier (converged csl/gsl) there is no spine row,
        # so edges keep the legacy y = -G. external-* always sits one row above
        # the edges.
        #   2-tier (e.g. 2-4-5-800)       1-tier (e.g. 2-8-9-800)
        #     y = -3G : external-conn/dhcp   y = -2G : external-conn/dhcp
        #     y = -2G : cust-net-edge-*      y =  -G : cust-net-edge-*
        #     y =  -G : cs/gs spines         y =   0 : CSLs and GSLs
        #     y =   0 : cl/gl leaves
        if self._oob_uplink_mode == "l3":
            edges = sorted(
                n for n in nodes
                if classify_node(self._name_to_role.get(n, n)) == "edge"
            )
            edge_y = (spine_y - G) if has_two_tier else -G
            ext_y = edge_y - G
            # Place every cust-net-edge on its own row (handles the N-edge star
            # at max scale, not just the first two), spread across columns.
            for i, name in enumerate(edges):
                nodes[name]["positioning"] = {
                    "x": (csl_start + i * csl_step) * G, "y": edge_y,
                }
            # external-conn / external-dhcp one row above the edges. In 1-tier
            # the legacy layout kept external-dhcp on the edge row; preserve that
            # so existing 1-tier topologies don't shift.
            if "external-conn" in nodes:
                nodes["external-conn"]["positioning"] = {"x": csl_start * G, "y": ext_y}
            if "external-dhcp" in nodes:
                dhcp_y = ext_y if has_two_tier else -G
                nodes["external-dhcp"]["positioning"] = {"x": (csl_start - 1) * G, "y": dhcp_y}
            # utility sits to the LEFT of the OOB column (one col left of oob_col).
            if "utility" in nodes:
                nodes["utility"]["positioning"] = {
                    "x": (oob_col - 1) * G,
                    "y": server_row_start * G,
                }

        self._resolve_position_collisions(nodes, G)

    # ---- connected links --------------------------------------------------

    def _build_connected_links(
        self,
    ) -> Tuple[list, Dict[str, Set[str]]]:
        """Build links from Display-in-Air=Yes rows.

        Air requires each (node, interface) pair to be unique across all links.
        When the wiremap maps multiple devices to the same switch port (e.g.
        BMC + LOM + iDRAC all sharing oob-switch-01:swp26), only the first
        connection is kept — duplicates are skipped with a warning.

        Handles special switch_role values:
        - "outbound": creates an outbound link (internet access for virtual nodes)
        - "NA": skipped entirely

        Returns (links_list, switch_connected_ports) where
        switch_connected_ports maps switch name -> set of port names already
        wired to something.
        """
        links: list = []
        switch_connected: Dict[str, Set[str]] = defaultdict(set)
        used_endpoints: Set[Tuple[str, str]] = set()
        skipped = 0

        for r in self.rows:
            if not r.display_in_air:
                continue
            # Disabled rows are retained for breakout analysis only — don't
            # create a link to a node that won't be in the topology.
            if not r.enabled:
                continue
            if not r.switch_role or r.switch_role.upper() == "NA":
                continue
            # Skip rows with invalid hostnames (e.g. "SPARE ISL")
            if not is_valid_hostname(r.system_name) or not is_valid_hostname(r.switch_name):
                continue

            # Determine left-side (system) interface name
            # Switches: use wiremap port name as-is
            # OOB-switch connection reserved as eth0: use eth0
            # Servers with ethN in wiremap (Air rows): use as-is
            # Servers with HW NIC names: assign sequential ethN
            if is_switch(r.system_role):
                system_port = r.nic_port
            elif (r.system_name in self._oob_eth0
                  and self._oob_eth0[r.system_name] == (r.switch_name, r.switch_port)):
                system_port = "eth0"
            elif r.nic_port and re.search(r"eth(\d+)", r.nic_port):
                system_port = f"eth{re.search(r'eth(\d+)', r.nic_port).group(1)}"
            else:
                system_port = self._next_eth(r.system_name)

            # Handle outbound links (e.g., dhcp-oob:eth0 → outbound)
            if r.switch_role.lower() == "outbound":
                ep = (r.system_name, system_port)
                if ep in used_endpoints:
                    skipped += 1
                    continue
                used_endpoints.add(ep)

                links.append([
                    {
                        "interface": system_port,
                        "node": r.system_name,
                        "mac": generate_mac(r.system_name, system_port),
                        "network_pci": None,
                    },
                    "outbound",
                ])
                # Track the port as connected so the unconnected-stub pass
                # doesn't emit a duplicate for it.
                if is_switch(r.system_role):
                    switch_connected[r.system_name].add(system_port)
                continue

            if not r.switch_port:
                continue

            # Right-side interface comes directly from the wiremap
            other_port = r.switch_port

            # Check for duplicate endpoints — Air only allows one link per port
            left_ep = (r.system_name, system_port)
            right_ep = (r.switch_name, other_port)

            if left_ep in used_endpoints or right_ep in used_endpoints:
                skipped += 1
                continue

            used_endpoints.add(left_ep)
            used_endpoints.add(right_ep)

            link = self._make_link(
                r.system_name, system_port,
                r.switch_name, other_port,
            )
            links.append(link)

            # Track connected ports on switches
            if is_switch(r.system_role):
                switch_connected[r.system_name].add(system_port)
            if is_switch(r.switch_role):
                switch_connected[r.switch_name].add(other_port)

        if skipped:
            print(f"    (skipped {skipped} duplicate port assignments)")

        return links, switch_connected

    def _oob_subnets(self) -> List[str]:
        """Derive OOB subnets from the OOB VLAN rows + the Nodes `OOB VLAN`
        mapping — replaces the old Settings-tab CSV field.

        Uses the same `resolve_oob_vlans()` resolver as scripts/excel_parser.py
        so topology and inventory generation always agree on which VLANs/
        subnets constitute the OOB plane.

        Returns list of subnet strings, e.g., ['192.168.200.0/24', '192.168.210.0/24'].
        """
        wb = openpyxl.load_workbook(self.excel_path, data_only=True)
        vlans = parse_vlans(wb["VLANs & Profiles"]) if "VLANs & Profiles" in wb.sheetnames else []
        # NOTE: passes UNFILTERED parse_nodes() (all OOB-category nodes on the
        # sheet, regardless of Active/Inactive status), unlike the deploy
        # path's get_oob_nodes_for_inventory() which filters to Active only.
        # There is no more Settings-driven trim/pad (management_switches is
        # retired) — the only remaining divergence is the status filter,
        # which can at most resolve a harmless superset of subnets here.
        oob_nodes = parse_nodes(wb["Nodes"]) if "Nodes" in wb.sheetnames else []
        wb.close()
        return resolve_oob_vlans(vlans, oob_nodes)['subnets']

    # ---- air-oob-switch injection -------------------------------------------

    def _inject_air_oob_switch(
        self,
        nodes: dict,
        links: list,
        switch_connected: Dict[str, Set[str]],
    ) -> Tuple[dict, list, Dict[str, Set[str]]]:
        """Inject a VLAN-aware air-oob-switch for OOB management.

        air-oob-switch provides:
          - air-mgmt (untagged): switch eth0s for ZTP
          - Per OOB subnet VLAN: uplink from each OOB switch, plus interfaces
            for oob-server-01 (gateway .1) and dhcp-oob (DHCP server)

        oob-server-01 and dhcp-oob each get:
          - eth0 → outbound (internet/SSH)
          - eth1 → air-oob-switch (air-mgmt, untagged)
          - eth2+ → air-oob-switch (one per OOB subnet VLAN)
        """
        AIR_OOB = "air-oob-switch"
        next_swp = 1
        air_connected: Set[str] = set()

        def _alloc_swp() -> str:
            nonlocal next_swp
            port = f"swp{next_swp}"
            next_swp += 1
            air_connected.add(port)
            return port

        # Derive OOB subnets from the OOB VLANs
        oob_subnets = self._oob_subnets()
        n_subnets = len(oob_subnets)

        # Collect OOB switch names present in the topology
        oob_switch_names = sorted(
            n for n in nodes
            if classify_node(self._name_to_role.get(n, n)) == "oob"
        )

        # --- Pass 1: Process existing links ---
        # Rewire switch eth0s to air-oob-switch, drop old infra→oob-switch links
        new_links: list = []
        oob_set = set(oob_switch_names)

        for link in links:
            if not isinstance(link[0], dict) or not isinstance(link[1], dict):
                new_links.append(link)
                continue

            ep0, ep1 = link[0], link[1]

            # Identify OOB switch endpoint
            oob_ep, other_ep = None, None
            if ep1["node"] in oob_set:
                oob_ep, other_ep = ep1, ep0
            elif ep0["node"] in oob_set:
                oob_ep, other_ep = ep0, ep1
            else:
                new_links.append(link)
                continue

            other_role = classify_node(
                self._name_to_role.get(other_ep["node"], other_ep["node"])
            )
            other_name = other_ep["node"].lower()

            # Switch eth0 → rewire to air-oob-switch
            if other_ep["interface"] == "eth0" and other_role in SWITCH_ROLES \
                    and other_role != "air-oob":
                swp = _alloc_swp()
                new_links.append(self._make_link(other_ep["node"], "eth0", AIR_OOB, swp))
                switch_connected[oob_ep["node"]].discard(oob_ep["interface"])
                continue

            # Drop dhcp-oob/oob-server links to oob-switches (rebuilt in Pass 4)
            if ("dhcp-oob" in other_name or "oob-server" in other_name) \
                    and other_ep["interface"].startswith("eth"):
                continue

            # Keep everything else (server connections, etc.)
            new_links.append(link)

        # --- Pass 2: Connect ALL switch eth0s to air-oob-switch ---
        already_on_air_oob = {
            ep["node"]
            for link in new_links
            if isinstance(link[0], dict) and isinstance(link[1], dict)
            for ep in link
            if ep.get("interface") == "eth0"
            and any(e.get("node") == AIR_OOB for e in link if isinstance(e, dict))
        }
        for sw_name in sorted(nodes):
            role = classify_node(self._name_to_role.get(sw_name, sw_name))
            if role not in SWITCH_ROLES or role == "air-oob":
                continue
            if sw_name in already_on_air_oob:
                continue
            swp = _alloc_swp()
            new_links.append(self._make_link(sw_name, "eth0", AIR_OOB, swp))
            switch_connected[sw_name].add("eth0")

        # --- Pass 3: OOB switch uplinks to air-oob-switch ---
        # Each OOB switch gets an uplink port to air-oob-switch (for its mgmt VLAN)
        for oob_name in oob_switch_names:
            used = switch_connected.get(oob_name, set())
            uplink_port = None
            for p in range(1, 53):
                candidate = f"swp{p}"
                if candidate not in used:
                    uplink_port = candidate
                    break
            if uplink_port:
                swp = _alloc_swp()
                new_links.append(self._make_link(oob_name, uplink_port, AIR_OOB, swp))
                switch_connected[oob_name].add(uplink_port)

        # --- Pass 4: Create dhcp-oob and oob-server-01 ---
        # Each gets: eth0 → outbound, eth1 → air-oob-switch (air-mgmt),
        #            eth2+ → air-oob-switch (one per OOB subnet)
        for infra_name in ["dhcp-oob", "oob-server-01"]:
            if infra_name not in nodes:
                defaults = NODE_DEFAULTS.get("infra", NODE_DEFAULTS["unknown"])
                nodes[infra_name] = {
                    "cpu": defaults["cpu"],
                    "memory": defaults["memory"],
                    "storage": defaults["storage"],
                    "positioning": {"x": 0, "y": 0},
                    "os": SERVER_OS,
                    "features": {"uefi": False, "tpm": False},
                    "pxehost": False,
                    "secureboot": False,
                    "oob": False,
                    "emulation_type": None,
                    "network_pci": {},
                }
            # eth0 → outbound
            outbound_exists = any(
                isinstance(link[0], dict) and link[0].get("node") == infra_name
                and link[1] == "outbound"
                for link in new_links
            )
            if not outbound_exists:
                new_links.append([
                    {
                        "interface": "eth0",
                        "node": infra_name,
                        "mac": generate_mac(infra_name, "eth0"),
                        "network_pci": None,
                    },
                    "outbound",
                ])
            # eth1 → air-oob-switch (air-mgmt, untagged)
            swp = _alloc_swp()
            new_links.append(self._make_link(infra_name, "eth1", AIR_OOB, swp))
            # eth2+ → air-oob-switch (one per OOB subnet)
            for i in range(n_subnets):
                eth_name = f"eth{2 + i}"
                swp = _alloc_swp()
                new_links.append(self._make_link(infra_name, eth_name, AIR_OOB, swp))

        # --- Create the air-oob-switch node ---
        role_defaults = NODE_DEFAULTS.get("air-oob", NODE_DEFAULTS["oob"])
        nodes[AIR_OOB] = {
            "cpu": role_defaults["cpu"],
            "memory": role_defaults["memory"],
            "storage": role_defaults["storage"],
            "positioning": {"x": 0, "y": 0},  # updated by _apply_layout
            "os": self._resolve_os(AIR_OOB),
            "features": {"uefi": False, "tpm": False},
            "pxehost": False,
            "secureboot": False,
            "oob": False,
            "emulation_type": None,
            "network_pci": {},
        }

        # Update switch_connected tracking
        switch_connected[AIR_OOB] = air_connected

        print(f"    air-oob-switch: {len(air_connected)} ports connected "
              f"(swp1-swp{next_swp - 1}), "
              f"{n_subnets} OOB subnet{'s' if n_subnets != 1 else ''}")

        # Store metadata for Node Instructions (used by air-deploy.py)
        self._air_oob_metadata = {
            'oob_subnets': oob_subnets,
            'oob_switch_names': oob_switch_names,
            'connected_ports': sorted(air_connected, key=lambda p: int(p.replace("swp", ""))),
        }

        return nodes, new_links, switch_connected

    # ---- l3-oob injection -------------------------------------------------

    def _inject_l3_oob_nodes(
        self,
        nodes: dict,
        links: list,
        switch_connected: Dict[str, Set[str]],
    ) -> Tuple[dict, list, Dict[str, Set[str]]]:
        """Inject L3-OOB Ubuntu trio (external-conn, external-dhcp, utility).

        In L3 mode every cluster switch's eth0 is rerouted to a cust-net-edge
        L2 Air-mgmt bridge (so ZTP works without the OOB switches being
        pre-configured). external-conn provides routed outbound NAT,
        external-dhcp answers switch ZTP DHCP, utility sits on the OOB VLAN
        200 plane as jump host + server-ZTP DHCP server.

        Design doc: docs/plans/2026-05-20-l3-oob-air-topology.md
        """
        EDGE = "cust-net-edge-01"
        OOB = "oob-switch-01"
        if OOB not in nodes:
            for n in nodes:
                role = self._name_to_role.get(n, '')
                if role == 'oob-switch' or classify_node(n) == 'oob':
                    OOB = n
                    break

        if OOB not in nodes:
            print("    [WARN] L3 OOB mode requested but OOB switch anchor "
                  "missing (need at least one oob-switch); skipping L3 injection.")
            return nodes, links, switch_connected
        sw_defaults = NODE_DEFAULTS.get("edge", NODE_DEFAULTS["core"])
        edge_os = nodes.get(OOB, {}).get("os", SWITCH_OS_FALLBACK)

        def _make_edge_node() -> dict:
            return {
                "cpu": sw_defaults["cpu"],
                "memory": sw_defaults["memory"],
                "storage": sw_defaults["storage"],
                "positioning": {"x": 0, "y": 0},
                "os": edge_os,
                "features": {"uefi": False, "tpm": False},
                "pxehost": False,
                "secureboot": False,
                "oob": False,
                "emulation_type": None,
                "network_pci": {},
            }

        # --- Multi-edge mgmt bridge -------------------------------
        # A single cust-net-edge bridging every switch eth0 overflows the 64-port
        # switch limit past ~60 switches (maxscale 2-4-5-800=93, 2-8-9-800=112).
        # Spread switch eth0s across N edges and span the air-mgmt L2 as a
        # hub-and-spoke star centred on cust-net-edge-01: each spoke trunks to
        # the hub (loop-free, no RSTP). The hub owns the bridge SVI and the
        # external-dhcp/utility management legs; external-conn is separate
        # routed EXIT egress. N grows sub-linearly with the switch count.
        def _is_cluster_sw(n: str) -> bool:
            r = classify_node(self._name_to_role.get(n, n))
            return r in SWITCH_ROLES and r not in ("edge", "air-oob")

        _cluster_switches = sorted(n for n in nodes if _is_cluster_sw(n))

        # The final edge count is sized below from the REAL load (EXIT uplinks
        # land in Pass 1b, so we must size after that — see _ensure_edges()).
        # Seed from every cust-net-edge the Nodes tab declares, not a
        # hardcoded pair: generate_arch_excel spreads the fabric EXIT uplinks
        # across CUST_EDGE_HA edges, and an edge missing from this list is
        # invisible to BOTH the sizing pass below and _least_loaded_edge, so
        # its Wire Map uplinks do not count toward the 64-port budget while
        # management eth0s are still stacked on top of them.
        # Names are zero-padded (cust-net-edge-01..NN) so a plain sort is
        # already ordinal; -01 must stay first because it is the star hub.
        edge_names = sorted(
            {n for n in nodes if str(n).startswith("cust-net-edge-")}
            | {"cust-net-edge-01", "cust-net-edge-02"}
        )
        for _en in edge_names:
            if _en not in nodes:
                nodes[_en] = _make_edge_node()
        # Per-edge used-port sets, shared with switch_connected so Pass 1b's
        # EXIT-uplink allocations are reflected when we balance eth0s.
        edge_used_map = {en: switch_connected.setdefault(en, set()) for en in edge_names}

        def _alloc_port_on(en: str) -> str:
            used = edge_used_map[en]
            k = 1
            while f"swp{k}" in used:
                k += 1
            port = f"swp{k}"
            used.add(port)
            return port

        def _least_loaded_edge() -> str:
            # The hub is charged its infra reserve so eth0s prefer the spokes;
            # without this the hub wins ties, then the star lands on top of it.
            return min(
                edge_names,
                key=lambda e: (
                    len(edge_used_map[e])
                    + (_hub_infra_reserve(len(edge_names)) if e == EDGE else 0)
                ),
            )

        def _alloc_edge_port() -> str:
            # Hub (cust-net-edge-01) allocator for infra legs that must land
            # on the management hub or edge-01 EXIT egress.
            return _alloc_port_on(EDGE)

        # Ports cust-net-edge-01 alone must still have free AFTER the eth0
        # balancing below: one trunk per spoke, the external-conn leg, both
        # external-dhcp legs, utility, one per ext-storage node, and its own
        # eth0 management stub. All of these are allocated later, so if they
        # are not reserved here the hub silently finishes as the most loaded
        # edge -- which is how it ended up on swp65+ (see ADR-0002).
        _HUB_FIXED_INFRA = 4  # external-conn(1) + external-dhcp(2) + utility(1)

        def _hub_infra_reserve(n_edges: int) -> int:
            n_ext_storage = sum(1 for n in nodes if str(n).startswith("ext-storage-"))
            return (n_edges - 1) + _HUB_FIXED_INFRA + n_ext_storage + 1

        def _project_loads(n_edges: int, pending_eth0: int) -> list:
            """Per-edge port count after eth0 placement, for `n_edges` edges.

            Mirrors _least_loaded_edge exactly -- including charging the hub
            its infra reserve while choosing -- so the projection the sizing
            loop trusts is the placement that actually happens. Modelling
            plain least-loaded here instead would under-count the spokes,
            since biasing eth0s off the hub is what pushes them onto spokes.
            """
            reserve = _hub_infra_reserve(n_edges)
            loads = [len(edge_used_map[e]) for e in edge_names[:n_edges]]
            loads += [0] * (n_edges - len(loads))
            # Each spoke costs 2 beyond its eth0 share: the trunk to the hub,
            # and -- if it carries EXIT uplinks -- its own external-conn egress
            # leg. Charged to every spoke rather than only the EXIT ones so the
            # projection stays an upper bound; over-counting a mgmt-only spoke
            # by one port is far cheaper than under-counting an EXIT one, which
            # is what silently pushes ports past the platform limit.
            # The hub's trunk ports are already in `reserve`.
            for i in range(1, n_edges):
                loads[i] += 2
            for _ in range(pending_eth0):
                biased = [n + (reserve if i == 0 else 0) for i, n in enumerate(loads)]
                loads[biased.index(min(biased))] += 1
            loads[0] += reserve
            return loads

        def _ensure_edges_for_load(pending_eth0: int) -> None:
            # Grow the edge count until the projected worst-case edge fits
            # under the platform port limit with margin. NVUE accepts config
            # on ports past the limit and `nv config apply` SUCCEEDS, so an
            # over-subscribed edge fails silently at the datapath -- the
            # air-mgmt bridge comes up with no working members and every
            # switch goes unreachable. Sizing from the REAL per-edge load
            # (Wire Map EXIT uplinks included) is what keeps that honest.
            n = max(2, len(edge_names))
            while n < MAX_EDGES and max(
                _project_loads(n, pending_eth0)
            ) > EDGE_PORT_LIMIT - EDGE_PORT_MARGIN:
                n += 1
            # Adding spokes only relieves eth0 pressure -- it cannot move a
            # Wire Map EXIT uplink off the edge it is cabled to. If an edge is
            # STILL over budget here, the workbook itself handed one edge more
            # uplinks than the platform has ports, and no amount of spokes will
            # fix it. Fail loudly: the alternative is a topology that deploys
            # cleanly, applies cleanly, and silently does not forward.
            projected = _project_loads(n, pending_eth0)
            if max(projected) > EDGE_PORT_LIMIT:
                worst = max(range(len(projected)), key=lambda i: projected[i])
                raise ValueError(
                    f"cust-net-edge-{worst + 1:02d} needs {projected[worst]} "
                    f"ports but the platform has {EDGE_PORT_LIMIT}. The Wire "
                    f"Map has concentrated too many EXIT uplinks on one edge "
                    f"-- spread them across more cust-net-edge switches "
                    f"(see cust_edge_count() in generate_arch_excel.py). "
                    f"Ports past swp{EDGE_PORT_LIMIT} accept config and apply "
                    f"successfully but are not in the datapath."
                )
            for i in range(len(edge_names) + 1, n + 1):
                en = f"cust-net-edge-{i:02d}"
                if en not in nodes:
                    nodes[en] = _make_edge_node()
                edge_names.append(en)
                edge_used_map[en] = switch_connected.setdefault(en, set())
                # Declared edges get their unconnected eth0 management stub
                # from _build_unconnected_stubs (they are Wire Map switches).
                # Edges spun up here are not in the Wire Map, so emit the same
                # stub explicitly rather than leaving them structurally
                # different from cust-net-edge-01..CUST_EDGE_HA.
                new_links.append(self._make_unconnected(en, "eth0"))

        oob_used = switch_connected.setdefault(OOB, set())

        def _alloc_oob_port() -> str:
            # Server eth0s already occupy swp1..N from Wire Map; utility
            # goes on the next free port.
            for n in range(1, 49):
                cand = f"swp{n}"
                if cand not in oob_used:
                    oob_used.add(cand)
                    return cand
            raise RuntimeError(f"No free port on {OOB} for utility:eth0")

        # --- Pass 1: rewrite existing links ------------------------------
        # Drop dead Wire-Map links from dhcp-oob/oob-server-01 (L2-mode
        # infra nodes) — those nodes don't exist in L3 mode. Rewire any
        # switch eth0 currently landing on an OOB switch onto cust-net-edge-01.
        rewired_switches: Set[str] = set()
        new_links: list = []
        for link in links:
            if not (isinstance(link[0], dict) and isinstance(link[1], dict)):
                new_links.append(link)
                continue
            ep0, ep1 = link[0], link[1]
            name0, name1 = ep0["node"].lower(), ep1["node"].lower()

            # Drop legacy dhcp-oob / oob-server-01 connections — they
            # appear in Wire Map for L2 mode but have no role in L3.
            if any("dhcp-oob" in n or "oob-server" in n for n in (name0, name1)):
                continue

            # Switch eth0 → re-home to cust-net-edge-01. Identify the switch
            # endpoint: must be on a switch role AND using interface 'eth0'.
            # This also handles the case where Wire Map already pinned eth0
            # to a cust-net-edge node (peer role is 'edge'): without dropping
            # the original link here, Pass 2 would add a second eth0 link
            # and Air rejects duplicate (node, interface) endpoints.
            sw_ep, other_ep = None, None
            for cand_sw, cand_other in ((ep0, ep1), (ep1, ep0)):
                role = classify_node(
                    self._name_to_role.get(cand_sw["node"], cand_sw["node"])
                )
                if (cand_sw["interface"] == "eth0"
                        and role in SWITCH_ROLES
                        and role != "edge"
                        and role != "air-oob"):
                    sw_ep, other_ep = cand_sw, cand_other
                    break
            if sw_ep is not None:
                # Drop the original eth0 link; we'll re-create it pointing at EDGE.
                # Free up the port on the OOB/edge switch (other_ep) so it's
                # still available for the unconnected-stubs pass.
                if other_ep["node"] in switch_connected:
                    switch_connected[other_ep["node"]].discard(other_ep["interface"])
                rewired_switches.add(sw_ep["node"])
                continue

            new_links.append(link)

        # --- Pass 1b: wire cust-net-edge nodes to cores (BGP EXIT VRF) ---
        # Wire Map "Edge Uplink" rows have blank switch_role (the customer
        # network doesn't fit a standard function), so _build_connected_links
        # skipped them. In L3 mode we need these uplinks live for the EXIT
        # VRF path, so re-scan self.rows and add them here.
        used_endpoints = set()
        for link in new_links:
            if isinstance(link[0], dict) and isinstance(link[1], dict):
                used_endpoints.add((link[0]["node"], link[0]["interface"]))
                used_endpoints.add((link[1]["node"], link[1]["interface"]))
        # Edges that carry real fabric EXIT uplinks; drives which ones get an
        # external-conn egress leg in Pass 3. Read from the Wire Map rather
        # than from the links Pass 1b creates: once an edge is DECLARED in the
        # Nodes tab, Pass 1 already wired it and Pass 1b skips the row as a
        # duplicate endpoint, so counting Pass 1b's work would see nothing.
        _edges_with_exit_uplinks = {
            str(name)
            for r in self.rows if r.display_in_air and r.enabled
            for name in (r.system_name, r.switch_name)
            if name and str(name).startswith("cust-net-edge")
        }
        for r in self.rows:
            if not r.display_in_air:
                continue
            # Skip disabled-node rows here too — Pass 1b creates LINKS, and
            # _build_nodes won't emit the disabled node, so a link to it
            # would dangle and Air would reject the topology.
            if not r.enabled:
                continue
            if not (r.switch_name and r.switch_name.startswith("cust-net-edge")):
                continue
            if not (r.system_name and r.switch_port and r.nic_port):
                continue
            if not is_valid_hostname(r.system_name) or not is_valid_hostname(r.switch_name):
                continue
            left_ep = (r.system_name, r.nic_port)
            right_ep = (r.switch_name, r.switch_port)
            if left_ep in used_endpoints or right_ep in used_endpoints:
                continue
            used_endpoints.add(left_ep)
            used_endpoints.add(right_ep)
            new_links.append(self._make_link(
                r.system_name, r.nic_port, r.switch_name, r.switch_port,
            ))
            switch_connected.setdefault(r.system_name, set()).add(r.nic_port)
            switch_connected.setdefault(r.switch_name, set()).add(r.switch_port)

        # --- Pass 2: connect every cluster-switch eth0 to cust-net-edge-01 ----
        # Includes both switches we rewired above and any switch that didn't
        # have an explicit Wire Map eth0 link. Match "eth0 already on ANY
        # cust-net-edge" (not just -01) so we don't double-link a switch
        # whose Wire Map eth0 row pointed at cust-net-edge-02 and somehow
        # survived Pass 1's rewrite.
        already_on_edge = set()
        for link in new_links:
            if not (isinstance(link[0], dict) and isinstance(link[1], dict)):
                continue
            n0, n1 = link[0].get("node", ""), link[1].get("node", "")
            if not (n0.startswith("cust-net-edge") or n1.startswith("cust-net-edge")):
                continue
            for ep in link:
                if ep.get("interface") == "eth0":
                    already_on_edge.add(ep["node"])
        _pending = [s for s in _cluster_switches if s not in already_on_edge]
        _ensure_edges_for_load(len(_pending))
        for sw_name in _pending:
            en = _least_loaded_edge()
            port = _alloc_port_on(en)
            new_links.append(self._make_link(sw_name, "eth0", en, port))
            switch_connected.setdefault(sw_name, set()).add("eth0")

        # Star trunks: each spoke edge gets one trunk to the hub
        # (cust-net-edge-01) so the air-mgmt L2 bridge spans all edges. Loop-free
        # star → RSTP not relied upon. air-deploy detects these edge↔edge links
        # and puts both ends in br_default.
        for en in edge_names[1:]:
            spoke_port = _alloc_port_on(en)
            hub_port = _alloc_port_on(EDGE)
            new_links.append(self._make_link(en, spoke_port, EDGE, hub_port))

        # Infra ports (external-conn, external-dhcp, utility) allocated
        # dynamically in Pass 3 after Pass 1b populates edge_used from Wire Map.

        # --- Pass 3: create the three Ubuntu nodes ---------------------------
        infra_defaults = NODE_DEFAULTS.get("infra", NODE_DEFAULTS["unknown"])

        def _ubuntu_node() -> dict:
            return {
                "cpu": infra_defaults["cpu"],
                "memory": infra_defaults["memory"],
                "storage": infra_defaults["storage"],
                "positioning": {"x": 0, "y": 0},
                "os": SERVER_OS,
                "features": {"uefi": False, "tpm": False},
                "pxehost": False,
                "secureboot": False,
                "oob": False,
                "emulation_type": None,
                "network_pci": {},
            }

        # external-conn: NAT egress. eth1 -> cust-net-edge-01 and eth2 ->
        # cust-net-edge-02 are dedicated routed EXIT legs, not bridge members.
        # The first two edges advertise default-route-origination, cores ECMP
        # between them, and each edge has a static 0/0 toward external-conn on
        # its own egress subnet (172.20.1.0/24 via -01, 172.20.2.0/24 via -02).
        # Every edge that carries EXIT uplinks needs its own routed egress leg
        # or its customer traffic black-holes, so this is derived rather than
        # hardcoded to -02. ADR-0002 left edges 03+ as mgmt-bridge spokes
        # "until the topology generator deliberately adds more EXIT uplinks
        # plus matching external-conn interfaces/subnets" — which this and the
        # companion air-deploy.py change now do, numbered on the existing
        # pattern (edge-0N <-> external-conn:ethN on 172.20.N.0/24). Spokes
        # that _ensure_edges_for_load spun up purely to absorb management eth0s
        # carry no customer traffic and are deliberately left without a leg.
        EXIT_EDGES = [en for en in edge_names[1:] if en in _edges_with_exit_uplinks]
        edge_ext_conn_port = _alloc_edge_port()
        nodes["external-conn"] = _ubuntu_node()
        new_links.append([
            {"interface": "eth0", "node": "external-conn",
             "mac": generate_mac("external-conn", "eth0"), "network_pci": None},
            "outbound",
        ])
        new_links.append(self._make_link("external-conn", "eth1", EDGE, edge_ext_conn_port))
        for _n, EDGE2 in enumerate(EXIT_EDGES, start=2):
            if EDGE2 not in nodes:
                continue
            edge2_used = switch_connected.setdefault(EDGE2, set())
            # Pick the next FREE port — don't hardcode swp1. The Wire Map may
            # already use swp1.. on cust-net-edge-02 for fabric EXIT uplinks
            # (e.g. the dense 2-4-5-800 edge uses swp1-4), and a duplicate
            # (node, interface) makes Air reject the whole topology (INVALID).
            # On arches where swp1 is free (e.g. 2-8-9-800) this still picks
            # swp1, so their topology is unchanged.
            edge2_port_n = 1
            while f"swp{edge2_port_n}" in edge2_used:
                edge2_port_n += 1
            edge2_ext_conn_port = f"swp{edge2_port_n}"
            edge2_used.add(edge2_ext_conn_port)
            new_links.append(self._make_link(
                "external-conn", f"eth{_n}", EDGE2, edge2_ext_conn_port,
            ))

        # external-dhcp: switch ZTP DHCP + EXIT-VRF relay scopes
        edge_ext_dhcp_port = _alloc_edge_port()
        nodes["external-dhcp"] = _ubuntu_node()
        new_links.append([
            {"interface": "eth0", "node": "external-dhcp",
             "mac": generate_mac("external-dhcp", "eth0"), "network_pci": None},
            "outbound",
        ])
        new_links.append(self._make_link("external-dhcp", "eth1", EDGE, edge_ext_dhcp_port))

        # eth2: customer-DC-side leg for EXIT-VRF inter-VRF DHCP relay
        # testing. cust-net-edge-01 brings this swp up as an L3 interface
        # (10.88.88.1/24) and `redistribute connected` (already on the
        # cust-net-edge BGP block) advertises 10.88.88.0/24 to the cores'
        # EXIT VRF via the existing ESL underlay eBGP session. Pairs with
        # external-dhcp:eth2 = 10.88.88.88/24 in L3_TRIO_NETPLAN. See
        # docs/plans/2026-05-27-l3-oob-exit-dhcp-relay.md.
        edge_ext_dhcp_eth2_port = _alloc_edge_port()
        new_links.append(self._make_link(
            "external-dhcp", "eth2", EDGE, edge_ext_dhcp_eth2_port,
        ))

        # utility: server ZTP DHCP + jumpbox + status page.
        # eth0 → outbound matches the dhcp-oob convention and is what Air
        # uses to expose SSH/HTTP services to the operator (Air services
        # bind to the node's outbound interface).
        # eth1 → oob-switch-01 puts utility on VLAN 200 (OOB VRF) so it
        # can L2-broadcast DHCP to ZTP'ing servers.
        # eth2 → cust-net-edge-01 puts utility on the same Air-mgmt L2
        # bridge as every cluster switch's eth0 (172.20.0.0/24). Without
        # this link, validate-config and any Ansible play that connects
        # from the jump host to switches gets "no route to host" — utility
        # has no path to the Air-mgmt subnet otherwise. Its on-link gateway
        # is the cust-net-edge-01 bridge SVI (.254).
        nodes["utility"] = _ubuntu_node()
        new_links.append([
            {"interface": "eth0", "node": "utility",
             "mac": generate_mac("utility", "eth0"), "network_pci": None},
            "outbound",
        ])
        utility_port = _alloc_oob_port()
        new_links.append(self._make_link("utility", "eth1", OOB, utility_port))
        utility_edge_port = _alloc_edge_port()
        new_links.append(self._make_link("utility", "eth2", EDGE, utility_edge_port))

        # ext-storage-*: customer-side simulated storage aggregate, wired to
        # CSL swp63 ports via Wire Map (eth1, eth2 for BGP unnumbered). They
        # also need an eth0 on the Air-mgmt bridge (cust-net-edge-01) so
        # they can use a 172.20.0.x management address, route through the
        # cust-net-edge-01 SVI to external-conn NAT, and resolve
        # archive.ubuntu.com to apt-install FRR at first boot. Without this,
        # the NI's apt-get hangs on DNS.
        # Pattern mirrors how every switch eth0 lands on cust-net-edge-01.
        ext_storage_names = sorted(
            n for n in nodes if n.startswith("ext-storage-")
        )
        for ext_name in ext_storage_names:
            ext_edge_port = _alloc_edge_port()
            new_links.append(self._make_link(
                ext_name, "eth0", EDGE, ext_edge_port,
            ))

        # Store L3 metadata so air-deploy.py can resolve roles by name.
        self._l3_oob_metadata = {
            "jump_host": "utility",
            "ansible_target": "utility",
            "status_page_host": "utility",
            "nat_host": "external-conn",
            "dhcp_relay_server": "external-dhcp",
            "edge_switch": EDGE,
            "oob_switch": OOB,
            "rewired_switches": sorted(rewired_switches),
            "utility_air_mgmt_iface": "eth2",
        }

        edge_loads = ", ".join(
            f"{en}={len(edge_used_map[en])}" for en in edge_names
        )
        print(f"    L3 OOB topology: {len(_cluster_switches)} switch eth0s "
              f"across {len(edge_names)} cust-net-edge nodes "
              f"({edge_loads}), utility on {OOB}:{utility_port}")

        return nodes, new_links, switch_connected

    # ---- unconnected stubs ------------------------------------------------

    def _build_unconnected_stubs(
        self, switch_connected: Dict[str, Set[str]]
    ) -> list:
        """Add unconnected stubs for every switch port not already linked."""
        links: list = []

        # Gather all switch names from wiremap (use actual names, not function roles)
        switch_names: Set[str] = set()
        for r in self.rows:
            if is_switch(r.system_role):
                switch_names.add(r.system_name)
            if r.switch_role and is_switch(r.switch_role):
                switch_names.add(r.switch_name)

        for switch_name in sorted(switch_names):
            breakout_map, disabled_ports = self._analyze_breakouts(switch_name)
            port_count = self._effective_port_count(switch_name, breakout_map)
            connected = switch_connected.get(switch_name, set())

            for port_num in range(1, port_count + 1):
                if port_num in disabled_ports:
                    continue

                if port_num in breakout_map:
                    max_sub = breakout_map[port_num]
                    for sub in range(max_sub + 1):
                        sub_port = f"swp{port_num}s{sub}"
                        if sub_port not in connected:
                            links.append(
                                self._make_unconnected(switch_name, sub_port))
                else:
                    port_name = f"swp{port_num}"
                    if port_name not in connected:
                        links.append(
                            self._make_unconnected(switch_name, port_name))

            # eth0 management port
            if "eth0" not in connected:
                links.append(self._make_unconnected(switch_name, "eth0"))

        return links

    def _analyze_breakouts(
        self, switch_name: str
    ) -> Tuple[Dict[int, int], Set[int]]:
        """Scan ALL rows (including Display=No) for breakout/disabled info.

        Also supplements from generated inventory group_vars when the Wire Map
        lacks server rows (e.g. switches-only Excel). The generated
        gpu_breakout_parents / isl_breakout_parents in group_vars tell us which
        ports the rendered config will break out, so the topology must create
        sub-port stubs rather than parent-port stubs for those ports.

        Returns:
            breakout_map: {base_port_num: max_sub_port_index}
            disabled_ports: set of port numbers marked 'Port Disabled by Neighbor'
        """
        breakout_map: Dict[int, int] = {}
        disabled_ports: Set[int] = set()

        for r in self.rows:
            if r.system_name == switch_name:
                self._scan_port_info(
                    r.nic_port, r.net_profile,
                    breakout_map, disabled_ports)
            if r.switch_name == switch_name:
                self._scan_port_info(
                    r.switch_port, r.net_profile,
                    breakout_map, disabled_ports)

        self._supplement_breakouts_from_inventory(switch_name, breakout_map)
        return breakout_map, disabled_ports

    def _supplement_breakouts_from_inventory(
        self, switch_name: str, breakout_map: Dict[int, int]
    ) -> None:
        """Add breakout parents from generated host_vars/group_vars.

        When the Wire Map omits server rows (switches-only deploy), the
        breakout_map won't include GPU access ports. But the Jinja2 config
        template uses gpu_breakout_parents from group_vars and will emit
        'link breakout 2x' for those ports. ifreload-nvue needs the sub-port
        interfaces to exist in the Air topology, so we must create sub-port
        stubs — not parent-port stubs.
        """
        inv_dir = Path("output") / self.arch / self.site / "inventory"
        if not inv_dir.exists():
            return

        if not hasattr(self, "_inv_breakout_cache"):
            self._inv_breakout_cache = self._load_inventory_breakouts(inv_dir)

        for base_num in self._inv_breakout_cache.get(switch_name, []):
            breakout_map.setdefault(base_num, 1)

    def _load_inventory_breakouts(
        self, inv_dir: Path
    ) -> Dict[str, Set[int]]:
        """Build {switch_name: {base_port_nums}} from generated inventory.

        Reads the hosts file for group membership, then reads each group's
        group_vars (and per-host host_vars) for *_breakout_parents keys.
        """
        result: Dict[str, Set[int]] = {}
        hosts_file = inv_dir / "hosts"
        if not hosts_file.exists():
            return result

        breakout_keys = (
            "gpu_breakout_parents", "isl_breakout_parents",
            "cpu_breakout_parents",
        )

        def _extract_ports(val: str) -> Set[int]:
            ports: Set[int] = set()
            for token in val.split(","):
                parsed = parse_swp_port(token.strip())
                if parsed and parsed[1] is None:
                    ports.add(parsed[0])
            return ports

        group_members: Dict[str, list] = {}
        current_group = ""
        for line in hosts_file.read_text().splitlines():
            s = line.strip()
            if not s or s.startswith("#"):
                continue
            if s.startswith("[") and s.endswith("]"):
                current_group = s[1:-1].split(":")[0]
                continue
            if ":" not in s.split("]", 1)[0] if "]" in s else True:
                group_members.setdefault(current_group, []).append(s)

        for group_name, members in group_members.items():
            gv_path = inv_dir / "group_vars" / f"{group_name}.yml"
            if not gv_path.exists():
                continue
            try:
                with open(gv_path) as f:
                    gv = yaml.safe_load(f) or {}
            except Exception:
                continue
            ports: Set[int] = set()
            for key in breakout_keys:
                if key in gv and gv[key]:
                    ports |= _extract_ports(str(gv[key]))
            if ports:
                for member in members:
                    result.setdefault(member, set()).update(ports)

        for hv_file in sorted((inv_dir / "host_vars").iterdir()):
            if not hv_file.name.endswith(".yml"):
                continue
            hostname = hv_file.stem
            try:
                with open(hv_file) as f:
                    hv = yaml.safe_load(f) or {}
            except Exception:
                continue
            for key in breakout_keys:
                if key in hv and hv[key]:
                    result.setdefault(hostname, set()).update(
                        _extract_ports(str(hv[key])))

        return result

    @staticmethod
    def _scan_port_info(
        port_str: str,
        description: str,
        breakout_map: Dict[int, int],
        disabled_ports: Set[int],
    ) -> None:
        """Update breakout_map and disabled_ports from one port entry."""
        if not port_str:
            return

        desc_lower = (description or "").lower()

        if "disabled by neighbor" in desc_lower or "port disabled" in desc_lower:
            parsed = parse_swp_port(port_str)
            if parsed:
                disabled_ports.add(parsed[0])
            return

        parsed = parse_swp_port(port_str)
        if parsed and parsed[1] is not None:
            base_num, sub_idx = parsed
            if base_num not in breakout_map or sub_idx > breakout_map[base_num]:
                breakout_map[base_num] = sub_idx

    def _effective_port_count(
        self, switch_name: str, breakout_map: Dict[int, int]
    ) -> int:
        """Base port count from the switch model.

        Ports above this count that appear in the wiremap are handled as
        connected links — they don't need unconnected stubs generated.
        We do NOT extend the range to fill gaps between the model max and
        the highest wiremap port, since those intermediate ports may not
        physically exist on the switch (e.g., SN2201 has swp1-48 + swp49-52
        uplinks, but swp53+ don't exist even though virtual connections
        may reference swp60/61).
        """
        role_name = self._name_to_role.get(switch_name, switch_name)
        role = classify_node(role_name)
        return SWITCH_MODELS.get(role, {}).get("ports", 48)

    # ---- helpers ----------------------------------------------------------

    def _next_eth(self, server_name: str) -> str:
        """Return next sequential ethN for a server node.

        Skips ethN values already used by explicit rows (e.g., eth0 from
        'Air - Management' entries) to avoid duplicate endpoints.
        """
        idx = self._server_eth_counter[server_name]
        explicit = self._explicit_eth.get(server_name, set())
        while idx in explicit:
            idx += 1
        self._server_eth_counter[server_name] = idx + 1
        return f"eth{idx}"

    def _assert_no_duplicate_interfaces(self, links: list) -> None:
        """Hard-fail if one (node, interface) is wired by more than one link.

        Air validates this server-side and rejects the ENTIRE topology:

            Topology failed to pass validation: {"content": {"links":
              {"271": {"0": {"interface": "Interface eth0 is already defined
               for node ext-storage-01: index 0 of link index 247"}}}}}

        The failure mode is what makes this worth a hard gate rather than a
        warning. The import returns 200 with a simulation id, then flips to
        INVALID about half a second later with no nodes, no links and no error
        field. An INVALID simulation never appears in the Air UI, so the
        operator sees `make air-deploy` emit "Node X not found in simulation"
        for every node in turn and no reason for any of it.

        It shipped in all six architectures at once and cost two failed
        deploys to find, because nothing between `make generate` and Air had an
        opinion about it. Checking here is trivially cheap and catches it at
        generation, where the fix is.
        """
        seen: Dict[Tuple[str, str], int] = {}
        clashes = []
        for idx, link in enumerate(links):
            if not isinstance(link, list):
                continue
            for ep in link:
                if not isinstance(ep, dict):
                    continue
                node, iface = ep.get("node"), ep.get("interface")
                if not node or not iface:
                    continue
                key = (node, iface)
                if key in seen:
                    clashes.append(
                        f"{node} {iface}: link {seen[key]} and link {idx}"
                    )
                else:
                    seen[key] = idx
        if clashes:
            raise ValueError(
                "topology has duplicate (node, interface) endpoints — Air will "
                "reject the whole import and leave an INVALID simulation:\n  "
                + "\n  ".join(clashes)
            )

    def _make_link(self, node_a: str, port_a: str,
                   node_b: str, port_b: str) -> list:
        # Use the actual node name (not function/role) for MAC generation so topology
        # MACs match the dnsmasq DHCP reservations built by the Excel parser (which
        # also uses the node name).
        return [
            {
                "interface": port_a,
                "node": node_a,
                "mac": generate_mac(node_a, port_a),
                "network_pci": None,
            },
            {
                "interface": port_b,
                "node": node_b,
                "mac": generate_mac(node_b, port_b),
                "network_pci": None,
            },
        ]

    def _make_unconnected(self, node: str, port: str) -> list:
        return [
            {
                "interface": port,
                "node": node,
                "mac": generate_mac(node, port),
                "network_pci": None,
            },
            "unconnected",
        ]


# ---------------------------------------------------------------------------
# Validator
# ---------------------------------------------------------------------------

class TopologyValidator:
    """Compare an existing topology JSON against the Excel wiremap."""

    def __init__(self, excel_path: Path, topology_json: Path, arch: str,
                 switches_only: bool = False):
        self.excel_path = excel_path
        self.topology_json = topology_json
        self.arch = arch
        # When the topology was generated switches-only, the server VMs were
        # intentionally dropped — don't flag them as "missing nodes".
        self.switches_only = switches_only
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        self.rows = parse_wiremap_excel(excel_path)
        if "Air_Only" in wb.sheetnames:
            self.rows = parse_air_only_sheet(wb) + self.rows
        wb.close()
        # Build actual-name → function/role mapping. Skip empty roles so
        # classify_node falls back to the name (e.g., cust-net-edge-01 →
        # 'edge'); storing '' would make classify_node return 'unknown'
        # and break the is_switch check below.
        self._name_to_role: Dict[str, str] = {}
        for r in self.rows:
            if r.system_name and r.system_role and r.system_name != r.system_role:
                self._name_to_role[r.system_name] = r.system_role
            if r.switch_name and r.switch_role and r.switch_name != r.switch_role:
                self._name_to_role[r.switch_name] = r.switch_role
        self.errors: List[str] = []
        self.warnings: List[str] = []

    def validate(self) -> bool:
        """Run all checks.  Returns True when no errors are found."""
        with open(self.topology_json, "r") as f:
            topo = json.load(f)

        nodes = topo.get("content", {}).get("nodes", {})
        links = topo.get("content", {}).get("links", [])

        self._check_nodes(nodes)
        self._check_links(links)
        self._check_os_versions(nodes)
        self._check_unconnected_stubs(links)

        self._print_results()
        return len(self.errors) == 0

    # ---- checks -----------------------------------------------------------

    def _check_nodes(self, nodes: dict) -> None:
        wiremap_devices: Set[str] = set()
        air_only_devices: Set[str] = set()
        for r in self.rows:
            if not r.display_in_air:
                continue
            # Mirror the generator's enabled gate — disabled-node rows are
            # retained in self.rows only so _check_unconnected_stubs sees
            # the breakout info; they should NOT count as "expected nodes"
            # in the validator's missing-node check.
            if not r.enabled:
                continue
            # Track Air-only nodes separately (Air - prefixed network profiles)
            is_air_row = r.net_profile and r.net_profile.startswith('Air')
            if r.system_role:
                # Skip invalid hostnames — generator skips them too
                if not is_valid_hostname(r.system_name):
                    continue
                if is_air_row:
                    air_only_devices.add(r.system_name)
                else:
                    wiremap_devices.add(r.system_name)
            # Track the B-side device unless the role is explicitly OUTBOUND
            # (which is a sink, not a device). Blank role and "NA" are both
            # valid — cust-net-edge-* and ext-storage-* rows often carry NA
            # or blank function because they're "Edge Uplink" / external
            # device connections that don't map to a standard function.
            if r.switch_name and is_valid_hostname(r.switch_name):
                if not (r.switch_role and r.switch_role.upper() == "OUTBOUND"):
                    if is_air_row:
                        air_only_devices.add(r.switch_name)
                    else:
                        wiremap_devices.add(r.switch_name)

        # Nodes always injected by the generator (Air infrastructure).
        # L2 mode injects the air-oob-switch flat-bridge trio; L3 mode injects
        # the external-conn / external-dhcp / utility Ubuntu trio. Whitelist
        # both sets so a topology in either mode validates cleanly.
        # dhcp-edge is the legacy 2-8-9-400 inter-VRF DHCP-relay node.
        auto_injected = {
            # L2 mode
            "air-oob-switch", "dhcp-oob", "oob-server-01", "dhcp-edge",
            # L3 mode
            "external-conn", "external-dhcp", "utility",
        }

        # Switches-only: the server VMs were intentionally dropped, so only
        # switches + kept infra are expected in the topology. Mirror the
        # generator's _strip_to_switches keep rule (switch by role, or infra by
        # name prefix) so dropped servers are not flagged as missing/extra.
        if self.switches_only:
            def _kept(n: str) -> bool:
                return (is_switch(self._name_to_role.get(n, n))
                        or n.startswith(TopologyGenerator.SWITCHES_ONLY_INFRA_KEEP))
            wiremap_devices = {n for n in wiremap_devices if _kept(n)}
            air_only_devices = {n for n in air_only_devices if _kept(n)}

        # All devices the topology should contain
        all_expected = wiremap_devices | air_only_devices | auto_injected
        topo_nodes = set(nodes.keys())

        for name in sorted(wiremap_devices - topo_nodes):
            self.errors.append(
                f"Missing node: {name} (in wiremap, not in topology)")
        for name in sorted(topo_nodes - all_expected):
            self.warnings.append(
                f"Extra node: {name} (in topology, not in wiremap)")

    def _check_links(self, links: list) -> None:
        # Build lookup: switch -> set of ports present in topology
        # Node names in topology JSON are actual names; use _name_to_role for classification
        topo_switch_ports: Dict[str, Set[str]] = defaultdict(set)
        for link in links:
            node = link[0]["node"]
            if is_switch(self._name_to_role.get(node, node)):
                topo_switch_ports[node].add(link[0]["interface"])
            if isinstance(link[1], dict):
                node2 = link[1]["node"]
                if is_switch(self._name_to_role.get(node2, node2)):
                    topo_switch_ports[node2].add(link[1]["interface"])
            elif isinstance(link[1], str) and link[1] not in ("outbound", "unconnected"):
                pass  # unknown string endpoint

        for r in self.rows:
            if not r.display_in_air:
                continue
            # Disabled-node rows are retained for breakout analysis only —
            # we shouldn't expect the corresponding switch link to be in
            # the topology, so skip them in the missing-link check.
            if not r.enabled:
                continue
            if not r.switch_role or r.switch_role.upper() in ("NA", "OUTBOUND"):
                continue
            if not r.switch_port:
                continue
            # "Air - *" network profiles are physical wiremap entries that the
            # topology generator rewrites through air-oob-switch (eg.
            # dhcp-oob:eth1 → oob-switch-01:swpN becomes
            # dhcp-oob:eth1 → air-oob-switch:swpX). Skip the original-port
            # check for these — the post-injection topology is correct.
            if r.net_profile and r.net_profile.startswith("Air"):
                continue

            # Check right-side switch port
            if is_switch(r.switch_role):
                if r.switch_port not in topo_switch_ports.get(r.switch_name, set()):
                    self.errors.append(
                        f"Missing link: {r.switch_name}:{r.switch_port} "
                        f"(connected to {r.system_name} in wiremap)")

            # Check left-side switch port (switch-to-switch)
            if is_switch(r.system_role) and r.nic_port:
                if r.nic_port not in topo_switch_ports.get(r.system_name, set()):
                    self.errors.append(
                        f"Missing link: {r.system_name}:{r.nic_port} "
                        f"(connected to {r.switch_name} in wiremap)")

    def _check_os_versions(self, nodes: dict) -> None:
        for name, props in sorted(nodes.items()):
            role_name = self._name_to_role.get(name, name)
            actual = props.get("os", "")
            # Switch OS may be version-mapped, so only flag if it's
            # clearly wrong (not a cumulus image for switches, not ubuntu for servers)
            if is_switch(role_name):
                if "cumulus" not in actual.lower():
                    self.warnings.append(
                        f"OS mismatch: {name} has '{actual}', expected a Cumulus image")
            elif actual != SERVER_OS:
                self.warnings.append(
                    f"OS mismatch: {name} has '{actual}', expected '{SERVER_OS}'")

    def _check_unconnected_stubs(self, links: list) -> None:
        topo_switch_ports: Dict[str, Set[str]] = defaultdict(set)
        for link in links:
            node = link[0]["node"]
            if is_switch(self._name_to_role.get(node, node)):
                topo_switch_ports[node].add(link[0]["interface"])
            if isinstance(link[1], dict):
                node2 = link[1]["node"]
                if is_switch(self._name_to_role.get(node2, node2)):
                    topo_switch_ports[node2].add(link[1]["interface"])

        switch_names: Set[str] = set()
        for r in self.rows:
            if is_switch(r.system_role):
                switch_names.add(r.system_name)
            if r.switch_role and is_switch(r.switch_role):
                switch_names.add(r.switch_name)

        for switch_name in sorted(switch_names):
            role_name = self._name_to_role.get(switch_name, switch_name)
            role = classify_node(role_name)
            model_ports = SWITCH_MODELS.get(role, {}).get("ports", 48)

            # Find disabled ports and the set of ports referenced by the
            # wiremap for THIS switch. We only validate ports that fall in
            # one of two sets:
            #   1. the model's standard range (e.g., swp1..48 for SN2201)
            #   2. explicit ports referenced in the wiremap (e.g., swp49,
            #      swp51, swp52 uplinks on SN2201)
            # This avoids false "missing port: swp50" warnings on SN2201
            # where swp50 doesn't exist physically.
            wiremap_ports: Set[int] = set()
            disabled_ports: Set[int] = set()
            for r in self.rows:
                port_str = None
                desc = r.net_profile
                if r.system_name == switch_name:
                    port_str = r.nic_port
                elif r.switch_name == switch_name:
                    port_str = r.switch_port
                if port_str:
                    parsed = parse_swp_port(port_str)
                    if parsed:
                        wiremap_ports.add(parsed[0])
                        desc_lower = (desc or "").lower()
                        if ("disabled by neighbor" in desc_lower
                                or "port disabled" in desc_lower):
                            disabled_ports.add(parsed[0])

            ports_in_topo = topo_switch_ports.get(switch_name, set())

            # Validate: every port in the model's standard range OR
            # explicitly named in the wiremap should appear in the topology.
            check_ports = set(range(1, model_ports + 1)) | wiremap_ports
            for port_num in sorted(check_ports):
                if port_num in disabled_ports:
                    continue
                port_name = f"swp{port_num}"
                has_plain = port_name in ports_in_topo
                has_sub = any(
                    p.startswith(f"swp{port_num}s") for p in ports_in_topo
                )
                if not has_plain and not has_sub:
                    self.warnings.append(
                        f"Missing port: {switch_name}:{port_name} "
                        f"(no entry in topology)")

    # ---- output -----------------------------------------------------------

    def _print_results(self) -> None:
        if not self.errors and not self.warnings:
            print("  Topology validation passed - no issues found")
            return

        if self.errors:
            print(f"\n  {len(self.errors)} error(s):")
            for err in self.errors:
                print(f"    x {err}")

        if self.warnings:
            print(f"\n  {len(self.warnings)} warning(s):")
            for warn in self.warnings:
                print(f"    ! {warn}")

        print()


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> int:
    ap = argparse.ArgumentParser(
        description="ERA Topology Generator & Validator")
    sub = ap.add_subparsers(dest="command", required=True)

    gen = sub.add_parser("generate", help="Generate topology from wiremap")
    gen.add_argument("--arch", required=True,
                     help="Architecture (e.g., 2-8-5-200)")
    gen.add_argument("--site", default="default",
                     help="Site name (default: default)")
    gen.add_argument("--switches-only", action="store_true",
                     help="Omit server VMs from the topology (switches + jump "
                          "infra only) to test switch configs at larger scale "
                          "within a tighter Air budget")

    val = sub.add_parser("validate",
                         help="Validate topology against wiremap")
    val.add_argument("--arch", required=True, help="Architecture")
    val.add_argument("--site", default="default", help="Site name")
    val.add_argument("--topology",
                     help="Path to topology JSON (default: auto-detect)")
    val.add_argument("--switches-only", action="store_true",
                     help="Topology was generated switches-only; do not flag "
                          "intentionally-dropped server VMs as missing nodes")

    args = ap.parse_args()

    project_root = Path(__file__).resolve().parent.parent

    # Locate Excel wiremap
    excel_path = (
        project_root / "input" / args.arch / args.site / f"{args.arch}.xlsx"
    )
    if not excel_path.exists():
        print(f"  Error: Excel workbook not found: {excel_path}")
        return 1

    if args.command == "generate":
        output_path = (
            project_root / "output" / args.arch / args.site
            / "topology" / f"{args.arch}-topology.json"
        )
        print(f"  Generating topology for {args.arch} (site: {args.site})")
        print(f"  Source: {excel_path}")
        generator = TopologyGenerator(excel_path, args.arch, args.site,
                                      switches_only=getattr(args, "switches_only", False))
        generator.write(output_path)
        return 0

    elif args.command == "validate":
        if args.topology:
            topo_path = Path(args.topology)
        else:
            topo_path = (
                project_root / "output" / args.arch / args.site
                / "topology" / f"{args.arch}-topology.json"
            )

        if not topo_path.exists():
            print(f"  Error: topology not found: {topo_path}")
            return 1

        print(f"  Validating {topo_path}")
        print(f"  Against wiremap {excel_path}")
        validator = TopologyValidator(excel_path, topo_path, args.arch,
                                      switches_only=getattr(args, "switches_only", False))
        return 0 if validator.validate() else 1

    return 1


if __name__ == "__main__":
    sys.exit(main())
