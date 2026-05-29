#!/usr/bin/env python3
# SPDX-FileCopyrightText: Copyright (c) 2026 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: MIT
"""Generate a sample Excel sized to a specific SU count, with CRA-conformant OOB.

Reads the default Excel for an architecture, then:
  1. Sets `Enabled` on the Nodes tab so the first N SUs are active.
  2. For each activated node, applies the new CRA OOB strategy on its
     Wire Map OOB rows:
       - BMC row    → Display=Yes, port allocated (the Air-active cable)
       - LOM Port 1 → Display=No, kept (real-HW only)
       - iLO/iDRAC  → Display=No, kept (real-HW only)
       - LOM Port 2 → DELETED (HA-LOM dropped from new ERA design)
  3. Port allocation picks the next free port on the row's already-
     templated `oob-switch-NN` (sequentially from swp1).

Air constraint: exactly ONE OOB row Display=Yes per node (plain Ubuntu
in Air can't bond / would create loop weirdness). The validator flags
violations.

Usage:
    python3 scripts/scale_sample_excel.py --arch 2-4-3-200 --sus 4 \
        --output input/2-4-3-200/sample-su-4/2-4-3-200.xlsx
"""

import argparse
import re
import sys
from pathlib import Path
from typing import Optional

import openpyxl

SCRIPTS = Path(__file__).resolve().parent
if str(SCRIPTS) not in sys.path:
    sys.path.insert(0, str(SCRIPTS))

from arch_scaling import max_single_tier_su, get_tier, node_name_to_su
from excel_parser import canonical_category

# Canonical categories that are switches (not eligible for CRA / data-plane
# Display rewrites). Mirrors validate_excel._SWITCH_CATEGORIES.
_SWITCH_CATEGORIES = frozenset({
    "core", "csl", "gsl", "gsl-plane1", "gsl-plane2",
    "oob-switch", "edge", "air-oob",
})


# Port-description heuristics for identifying which OOB row is which.
# Matched against the NIC/Port column (column 3) of the Wire Map.
BMC_KEYWORDS  = ("BMC", "IPMI")                     # primary — Display=Yes
LOM2_KEYWORDS = ("LOM Port 2", "OCP 3.0 NIC Port 2",
                 "OCP 3.0 SL1 P2")                  # delete entirely
# 1st LOM, iLO/iDRAC, XCC are kept as Display=No (real-HW only).
LOM1_ILO_KEYWORDS = ("LOM Port 1", "OCP 3.0 NIC Port 1", "OCP 3.0 SL1 P1",
                     "iLO", "iDRAC", "XCC")


def find_default_excel(arch: str) -> Path:
    project_root = Path(__file__).resolve().parent.parent
    default = project_root / "input" / arch / "default" / f"{arch}.xlsx"
    if not default.exists():
        raise FileNotFoundError(f"Default Excel not found: {default}")
    return default


_node_pattern_su = node_name_to_su  # backward-compat shim for inline calls


def _is_bmc_row(port_desc: str) -> bool:
    return any(k in (port_desc or "") for k in BMC_KEYWORDS)


def _is_lom2_row(port_desc: str) -> bool:
    return any(k in (port_desc or "") for k in LOM2_KEYWORDS)


def _is_lom1_or_ilo_row(port_desc: str) -> bool:
    return any(k in (port_desc or "") for k in LOM1_ILO_KEYWORDS)


def set_enabled_for_sus(wb: openpyxl.Workbook, target_sus: int) -> dict:
    """Toggle the Nodes tab Enabled column so SUs 1..N are active."""
    if "Nodes" not in wb.sheetnames:
        raise ValueError("Excel has no Nodes sheet")
    ws = wb["Nodes"]
    enabled_col = name_col = None
    for c in range(1, ws.max_column + 1):
        header = (ws.cell(1, c).value or "").strip().lower()
        if header == "enabled":
            enabled_col = c
        elif header == "name":
            name_col = c
    if not enabled_col or not name_col:
        raise ValueError("Nodes tab must have 'Name' and 'Enabled' columns")

    stats = {"su_enabled": 0, "su_disabled": 0}
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, name_col).value
        if not name:
            continue
        su = _node_pattern_su(name)
        if su is None:
            continue
        if su <= target_sus:
            ws.cell(r, enabled_col).value = "Yes"
            stats["su_enabled"] += 1
        else:
            ws.cell(r, enabled_col).value = "No"
            stats["su_disabled"] += 1
    return stats


def collect_server_states(wb: openpyxl.Workbook) -> tuple[set, set]:
    """Read the Nodes tab and return (active_server_names, all_server_names).

    A server is any node whose canonical_category is NOT in _SWITCH_CATEGORIES.
    "Active" means the Enabled column is Yes/blank/true. Used by both the
    data-plane Display sync and the CRA OOB rewrite so they share the same
    notion of "which nodes are real, deployed servers".
    """
    ws = wb["Nodes"]
    func_col = name_col = enabled_col = None
    for c in range(1, ws.max_column + 1):
        h = (ws.cell(1, c).value or "").strip().lower()
        if h == "function":
            func_col = c
        elif h == "name":
            name_col = c
        elif h == "enabled":
            enabled_col = c
    if not func_col or not name_col:
        raise ValueError("Nodes tab needs Function and Name columns")
    active, all_servers = set(), set()
    for r in range(2, ws.max_row + 1):
        func = ws.cell(r, func_col).value
        name = ws.cell(r, name_col).value
        if not func or not name:
            continue
        cat = canonical_category(str(func), str(name))
        if cat in _SWITCH_CATEGORIES:
            continue
        all_servers.add(str(name).strip())
        en = (ws.cell(r, enabled_col).value if enabled_col else "Yes") or "Yes"
        if str(en).strip().lower() in ("yes", "true", "1", ""):
            active.add(str(name).strip())
    return active, all_servers


def sync_data_plane_display(wb: openpyxl.Workbook,
                            active: set, all_servers: set) -> dict:
    """Flip Wire Map Display=Yes on non-OOB rows for every active server,
    and Display=No on rows for every disabled server.

    Skips OOB rows entirely — those are governed by apply_cra_oob_layout
    which enforces the single-Display=Yes-per-node rule. Switch nodes
    (peer = core/csl/gsl) get one row per (compute, data-plane) cable
    and need to be Display=Yes so the topology generator wires them in
    Air. The default Excel ships these rows Display=No for disabled SU
    nodes; this pass re-enables them in lockstep with the Nodes-tab
    enable flip.
    """
    ws = wb["Wire Map"]
    disabled = all_servers - active
    yes_flipped = no_flipped = 0
    for r in range(2, ws.max_row + 1):
        name = str(ws.cell(r, 2).value or "").strip()
        sw = str(ws.cell(r, 5).value or "")
        if not name or not sw or "oob-switch" in sw:
            continue
        cur = str(ws.cell(r, 1).value or "").strip().lower()
        if name in active and cur != "yes":
            ws.cell(r, 1).value = "Yes"
            yes_flipped += 1
        elif name in disabled and cur != "no":
            ws.cell(r, 1).value = "No"
            no_flipped += 1
    return {"data_plane_yes_flipped": yes_flipped,
            "data_plane_no_flipped": no_flipped}


def apply_cra_oob_layout(wb: openpyxl.Workbook, target_sus: int,
                         active_servers: set | None = None) -> dict:
    """Apply the new CRA OOB strategy to every active server node.

    For each active server node (not just SU/gpu compute — also support,
    storage, bcm, k8s, slurm, etc.):
      - Find OOB rows on the Wire Map sheet
      - Mark the 2nd-LOM rows for deletion
      - Allocate a port on the templated OOB switch for the BMC row (if blank)
      - Set Display=Yes on the primary BMC row
      - Demote 1st LOM + iLO/iDRAC + secondary BMCs to Display=No

    If `active_servers` is None, falls back to legacy behavior matching
    only su-N-node-M and gpu-N names — kept for backwards compat with
    older callers; new code paths should pass the set explicitly.

    Returns a stats dict. Performs deletions in a second pass bottom-up
    so row indices stay valid during port assignment.
    """
    ws = wb["Wire Map"]
    # Resolve column indices from header
    headers = {}
    for c in range(1, ws.max_column + 1):
        h = (ws.cell(1, c).value or "").strip().lower()
        if h:
            headers[h] = c
    # Resolve columns from header row (preferred) with positional fallback.
    # The user's 2026-05-28 Excel revision added `Cable Split (A)` between
    # `Port Side (A)` and `System Name (B)`, shifting B-side columns right
    # by 1. Header-name lookup is the only safe path.
    DISPLAY_COL  = headers.get("display in air", 1)
    SYSNAME_COL  = headers.get("system name (a)",
                   headers.get("system name", 2))
    PORTDESC_COL = headers.get("port (a)",
                   headers.get("nic/port", 3))
    SWNAME_COL   = headers.get("system name (b)",
                   headers.get("switch name", 5))
    SWPORT_COL   = headers.get("port (b)",
                   headers.get("switch port", 6))
    # Some legacy Excels use 12/13 for switch columns. Header lookup
    # above already handles this if the header matches; this extra check
    # preserves backward compat for headers absent entirely.
    if (ws.cell(1, 12).value or "").strip().lower() in ("switch name",):
        SWNAME_COL = 12
        SWPORT_COL = 13

    # Pass 1: scan OOB rows. For each compute node in an active SU,
    # identify LOM2 rows (delete) and BMC rows (port-assign). Other OOB
    # rows are kept as-is.
    rows_to_delete: list[int] = []
    lom_ilo_to_demote: list[int] = []  # rows to force Display=No
    # Per-node BMC rows. Each node gets ONE primary BMC (Display=Yes,
    # port-allocated) and any additional BMCs (e.g. chassis BMC vs. DPU
    # BMC on a node with a DPU per the CRA strategy) are demoted to
    # Display=No so Air sees a single OOB cable per node.
    bmc_rows_by_node: dict[str, list[tuple[int, str]]] = {}
    # Track nodes that already have a non-BMC Display=Yes OOB row (e.g.
    # the 2-8-9-800 default's pre-injected `eth0` cable). For those we
    # preserve the existing primary and demote BMC rows instead of
    # promoting them — avoids creating two Display=Yes rows per node.
    pre_existing_yes: set[str] = set()

    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, SYSNAME_COL).value
        sw = ws.cell(r, SWNAME_COL).value
        if not name or not sw or "oob-switch" not in str(sw):
            continue
        name_s = str(name).strip()
        # Determine whether this row's A-side is an active server.
        if active_servers is not None:
            if name_s not in active_servers:
                continue
        else:
            # Legacy fallback — name must look like su-N-node-M or gpu-N
            # AND its SU index must be within target_sus.
            su = _node_pattern_su(name_s)
            if su is None or su > target_sus:
                continue
        port_desc = ws.cell(r, PORTDESC_COL).value or ""
        display = (ws.cell(r, DISPLAY_COL).value or "").strip().lower()
        if _is_lom2_row(port_desc):
            rows_to_delete.append(r)
            continue
        if _is_lom1_or_ilo_row(port_desc):
            # Real-HW only: kept in Wire Map for the operator's reference,
            # but Display=No so Air doesn't try to cable a second OOB link.
            lom_ilo_to_demote.append(r)
            continue
        if _is_bmc_row(port_desc):
            bmc_rows_by_node.setdefault(name_s, []).append((r, str(sw)))
            continue
        # Any other OOB row that's already Display=Yes — note it so the
        # BMC-promotion pass leaves things alone.
        if display == "yes":
            pre_existing_yes.add(name_s)

    # Pass 2: compute the set of OOB ports that will remain claimed after
    # deletion. Any row with a port set AND not in rows_to_delete keeps
    # its port. LOM2 ports are freed.
    used_ports: dict[str, set] = {}
    delete_set = set(rows_to_delete)
    for r in range(2, ws.max_row + 1):
        if r in delete_set:
            continue
        sw = ws.cell(r, SWNAME_COL).value
        port = ws.cell(r, SWPORT_COL).value
        if not sw or "oob-switch" not in str(sw):
            continue
        m = re.match(r"swp(\d+)$", str(port or "").strip())
        if m:
            used_ports.setdefault(sw, set()).add(int(m.group(1)))

    # Pass 3: allocate ports for BMC rows that need one.
    stats = {
        "bmc_assigned": 0,
        "bmc_demoted": 0,  # additional BMC rows (e.g. DPU BMC) → Display=No
        "lom2_deleted": len(rows_to_delete),
        "lom_ilo_demoted": 0,
    }

    # Demote LOM1 / iLO / iDRAC / XCC rows to Display=No.
    for r in lom_ilo_to_demote:
        cur = (ws.cell(r, DISPLAY_COL).value or "").strip().lower()
        if cur != "no":
            ws.cell(r, DISPLAY_COL).value = "No"
            stats["lom_ilo_demoted"] += 1

    def _alloc_port(sw_name: str) -> Optional[int]:
        used = used_ports.setdefault(sw_name, set())
        for p in range(1, 49):  # SN2201 has 48 access ports
            if p not in used:
                used.add(p)
                return p
        return None

    # Per node: pick ONE BMC row as the primary Display=Yes cable. We
    # iterate BMC rows in row order and pick the first whose templated
    # oob-switch peer has a free port. The default templates often pin
    # multiple BMCs at different oob-switches (e.g. PCIe Slot 1 BMC →
    # oob-switch-01 and PCIe Slot 2 BMC → oob-switch-02). When the first
    # switch is full, falling through to the next BMC row keeps the
    # operator's templated peering intact instead of moving cables.
    #
    # If the node already has a non-BMC Display=Yes OOB row (e.g. the
    # pre-injected `eth0` row for 2-8-9-800-style topologies), preserve
    # that as the primary and demote ALL BMC rows. Avoids the
    # multi-Display=Yes per-node Air loop hazard.
    for node_name, bmcs in bmc_rows_by_node.items():
        bmcs.sort()  # ascending by row index
        if node_name in pre_existing_yes:
            # Pre-injected eth0/management row is the primary. Demote
            # every BMC row to Display=No.
            for r, _sw in bmcs:
                cur = (ws.cell(r, DISPLAY_COL).value or "").strip().lower()
                if cur != "no":
                    ws.cell(r, DISPLAY_COL).value = "No"
                    stats["bmc_demoted"] += 1
            continue

        # Walk BMC rows: first one with an already-allocated port wins,
        # else first one whose switch has a free port we can allocate.
        primary_idx = None
        for i, (r, sw) in enumerate(bmcs):
            if ws.cell(r, SWPORT_COL).value:
                primary_idx = i
                break
        if primary_idx is None:
            for i, (r, sw) in enumerate(bmcs):
                used = used_ports.setdefault(sw, set())
                free = [p for p in range(1, 49) if p not in used]
                if free:
                    p = free[0]
                    used.add(p)
                    ws.cell(r, SWPORT_COL).value = f"swp{p}"
                    primary_idx = i
                    break
        if primary_idx is None:
            print(f"  ⚠️  {node_name}: no free OOB port on any templated "
                  f"BMC switch (tried {[sw for _, sw in bmcs]})",
                  file=sys.stderr)
            continue

        primary_row, _ = bmcs[primary_idx]
        if (ws.cell(primary_row, DISPLAY_COL).value or "").strip().lower() != "yes":
            ws.cell(primary_row, DISPLAY_COL).value = "Yes"
        stats["bmc_assigned"] += 1
        # Demote every other BMC row to Display=No (keep their port if
        # already set so real-HW reference stays accurate).
        for i, (r, _sw) in enumerate(bmcs):
            if i == primary_idx:
                continue
            cur = (ws.cell(r, DISPLAY_COL).value or "").strip().lower()
            if cur != "no":
                ws.cell(r, DISPLAY_COL).value = "No"
                stats["bmc_demoted"] += 1

    # Pass 4: delete LOM2 rows bottom-up so indices stay valid
    for r in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(r, 1)

    return stats


def main() -> int:
    ap = argparse.ArgumentParser(
        description="Generate a CRA-conformant sample Excel sized to N SUs.")
    ap.add_argument("--arch", required=True)
    ap.add_argument("--sus", type=int, required=True)
    ap.add_argument("--output", type=Path,
                    help="Output path (default: input/<arch>/sample-su-<N>/<arch>.xlsx)")
    ap.add_argument("--force-multi-tier", action="store_true",
                    help="Allow SU count beyond single-tier max")
    args = ap.parse_args()

    cap = max_single_tier_su(args.arch)
    if cap is None:
        print(f"ERROR: unknown arch '{args.arch}'", file=sys.stderr)
        return 1
    if args.sus < 1:
        print("ERROR: --sus must be >= 1", file=sys.stderr)
        return 1
    if args.sus > cap and not args.force_multi_tier:
        print(
            f"ERROR: --sus={args.sus} exceeds single-tier max ({cap}) for "
            f"{args.arch}. Pass --force-multi-tier to bypass (not advised).",
            file=sys.stderr,
        )
        return 2

    tier = get_tier(args.arch, args.sus)
    src = find_default_excel(args.arch)
    if args.output:
        dst = args.output
    else:
        project_root = Path(__file__).resolve().parent.parent
        dst = (project_root / "input" / args.arch
               / f"sample-su-{args.sus}" / f"{args.arch}.xlsx")
    dst.parent.mkdir(parents=True, exist_ok=True)

    print(f"Reading default Excel: {src}")
    wb = openpyxl.load_workbook(src)

    n_stats = set_enabled_for_sus(wb, args.sus)
    active, all_servers = collect_server_states(wb)
    d_stats = sync_data_plane_display(wb, active, all_servers)
    o_stats = apply_cra_oob_layout(wb, args.sus, active_servers=active)
    wb.save(dst)
    print(f"Wrote sample: {dst}")
    print(f"  SUs active:        {args.sus} / max single-tier {cap}")
    print(f"  SU nodes enabled:  {n_stats['su_enabled']}")
    print(f"  SU nodes disabled: {n_stats['su_disabled']}")
    print(f"  Data-plane rows Display=Yes flipped: {d_stats['data_plane_yes_flipped']}")
    print(f"  Data-plane rows Display=No flipped:  {d_stats['data_plane_no_flipped']}")
    print(f"  Primary BMC rows set: {o_stats['bmc_assigned']}")
    print(f"  Secondary BMC rows demoted: {o_stats['bmc_demoted']}")
    print(f"  LOM2 rows deleted: {o_stats['lom2_deleted']}")
    print(f"  LOM1/iLO rows demoted to Display=No: {o_stats['lom_ilo_demoted']}")
    if tier:
        print(f"  Tier: {tier.notes}")
        print(f"  Expected OOB switches at this scale: {tier.oob_switches}")
    else:
        print(f"  Tier: (multi-tier — outside arch_scaling table)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
